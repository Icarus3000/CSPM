import sys
import os
import openpyxl

def verify_metrics():
    wb_path = os.path.join(os.path.dirname(__file__), "data", "CSPM.xlsm")
    if not os.path.exists(wb_path):
        print(f"Error: {wb_path} does not exist.")
        return
        
    print(f"Loading {wb_path} ...")
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    
    txn_sheet = wb["Transactions"]
    
    # 1. Transactions check
    transactions = list(txn_sheet.iter_rows(values_only=True))
    if not transactions:
        print("No transactions found!")
        return
        
    headers = transactions[0]
    rows = [dict(zip(headers, r)) for r in transactions[1:] if any(c is not None for c in r)]
    
    print(f"Total Transactions in DB: {len(rows)}")
    
    # Check income/revenue transactions
    income_txns = [t for t in rows if t.get("Type") == "Income"]
    print(f"Total Income Transactions: {len(income_txns)}")
    
    business_income_txns = [t for t in income_txns if t.get("Class") == "Business"]
    revenue_income_txns = [t for t in income_txns if t.get("Class") == "Revenue"]
    print(f"  - Business Class Income Txns: {len(business_income_txns)}")
    print(f"  - Revenue Class Income Txns: {len(revenue_income_txns)}")
    
    def amt(val):
        if val is None: return 0.0
        try: return float(val)
        except: return 0.0
        
    sum_business_income = sum(amt(t.get("Amount")) for t in business_income_txns)
    print(f"  - Sum of Business Income Amount: ${sum_business_income:,.2f}")
    
    # Check dockets to compute Adjusted Production and Revenue + WIP manually
    dockets_sheet = wb["Dockets"]
    dockets = list(dockets_sheet.iter_rows(values_only=True))
    d_headers = dockets[0]
    d_rows = [dict(zip(d_headers, r)) for r in dockets[1:] if any(c is not None for c in r)]
    
    print(f"\nTotal Raw Dockets in DB: {len(d_rows)}")
    d_rows_2026 = [d for d in d_rows if "2026" in str(d.get("Date", ""))]
    print(f"Total 2026 Raw Dockets: {len(d_rows_2026)}")
    
    print("\n--- Production Breakdown ---")
    import re
    actual_invoice_pattern = re.compile(r"^[A-Za-z0-9]+-\d+$", re.IGNORECASE)
    
    # Calculate ledger billings by invoice
    ledger_billings_by_invoice = {}
    for row in rows:
        ref = (str(row.get("Invoice_Ref", "")) or "").strip().upper()
        if actual_invoice_pattern.match(ref):
            cls = row.get("Class")
            typ = row.get("Type")
            
            # Use 'Business' instead of 'Revenue' for this diagnostic
            if cls == "Business" and typ == "Income":
                ledger_billings_by_invoice[ref] = ledger_billings_by_invoice.get(ref, 0.0) + amt(row.get("Amount"))
                
    # Calculate docket fees by invoice
    docket_fees_by_invoice = {}
    for row in d_rows_2026:
        ref = (str(row.get("Invoice #", "")) or "").strip().upper()
        if actual_invoice_pattern.match(ref):
            docket_fees_by_invoice[ref] = docket_fees_by_invoice.get(ref, 0.0) + amt(row.get("Amount to CS"))
            
    # Calculate production and wip
    adjusted_production = 0.0
    wip_amount = 0.0
    
    for row in d_rows_2026:
        ref = (str(row.get("Invoice #", "")) or "").strip().upper()
        raw_production = amt(row.get("Amount to CS"))
        bucket = "blank"
        if ref:
            if ref.upper() == "NO BILL": bucket = "no_bill"
            elif ref.upper() == "HOLD": bucket = "hold"
            elif actual_invoice_pattern.match(ref): bucket = "invoice"
            else: bucket = "legacy_billed"
            
        if bucket == "no_bill":
            production = 0.0
        elif bucket == "invoice" and actual_invoice_pattern.match(ref):
            invoice_billings = ledger_billings_by_invoice.get(ref, 0.0)
            invoice_dockets = docket_fees_by_invoice.get(ref, 0.0)
            if invoice_dockets > 0.0001:
                realization_rate = invoice_billings / invoice_dockets
                production = raw_production * realization_rate
            else:
                production = 0.0
        else:
            production = raw_production
            
        adjusted_production += production
        if bucket in ("blank", "hold"):
            wip_amount += production

    print(f"\nResulting Target Adjusted Production: ${adjusted_production:,.2f}")
    
    ledger_billings_total = sum(ledger_billings_by_invoice.values())
    rev_plus_wip = wip_amount + ledger_billings_total
    print(f"Resulting Target Revenue + WIP: ${rev_plus_wip:,.2f}")
    
if __name__ == '__main__':
    verify_metrics()
