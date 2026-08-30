#!/usr/bin/env python3
"""Governed repository-only repair for three confirmed orphan records.

The command defaults to a read-only preflight.  ``--apply`` creates and
verifies a repository-workbook backup, repairs an isolated candidate, runs
semantic/package/integrity gates, and atomically promotes only that candidate
to ``data/CSPM.xlsm``.  Configured live and cloud workbooks are read/hash-only
refusal boundaries.
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from decimal import Decimal, InvalidOperation
import hashlib
import json
import os
from pathlib import Path
import shutil
import sys
from typing import Any, Iterable, Mapping, Sequence
from uuid import uuid4
from zipfile import ZipFile, is_zipfile


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PYTHON_ROOT = PROJECT_ROOT / "src" / "python"
if str(PYTHON_ROOT) not in sys.path:
    sys.path.insert(0, str(PYTHON_ROOT))

from openpyxl import load_workbook  # noqa: E402
from openpyxl.utils.cell import get_column_letter, range_boundaries  # noqa: E402

from domain import schema_constants as sc  # noqa: E402
from services.paths import AppPaths  # noqa: E402
from services.workbook_integrity_service import WorkbookIntegrityService  # noqa: E402


REPOSITORY_WORKBOOK = (PROJECT_ROOT / "data" / "CSPM.xlsm").resolve()
AUTHORIZED_MATTER_ID = "17bcaa4f-658d-48b0-84e7-01c4411ce135"
AUTHORIZED_ENTRY_IDS = ("T_9a77764a3b", "T_38a43a4197")
AUTHORIZED_IDS = (AUTHORIZED_MATTER_ID, *AUTHORIZED_ENTRY_IDS)
PROTECTED_INVOICE = "26-0080"

EXPECTED_MATTER_NONBLANK_COLUMNS = {
    sc.COL_MATTER_ID,
    sc.COL_MATTER_NAME,
    sc.COL_MATTER_CLIENT_ID,
    sc.COL_MATTER_STATUS,
}
EXPECTED_TIME_NONBLANK_COLUMNS = {
    sc.COL_TIME_ENTRY_ID,
    sc.COL_TIME_DATE,
    sc.COL_TIME_CLIENT_ID,
    sc.COL_TIME_MATTER_ID,
    sc.COL_TIME_DESC,
    sc.COL_TIME_HOURS,
    sc.COL_TIME_RATE,
    sc.COL_TIME_GROSS,
    sc.COL_TIME_NET,
    sc.COL_TIME_HST,
    sc.COL_TIME_INVOICE_STATUS,
}
TARGET_TABLE_REMOVALS = {sc.TBL_MATTERS: 1, sc.TBL_TIME: 2}
EXPECTED_WARNING_COUNTS = {
    "missing_account_reference": 4,
    "missing_business_unit_reference": 3,
}


class RepairRefused(RuntimeError):
    """A sanitized precondition or verification failure."""


@dataclass(frozen=True)
class FileFingerprint:
    classification: str
    bytes: int
    last_modified_utc: str
    sha256: str

    def as_dict(self) -> dict[str, Any]:
        return {
            "classification": self.classification,
            "bytes": self.bytes,
            "lastModifiedUtc": self.last_modified_utc,
            "sha256": self.sha256,
        }


@dataclass
class TableRow:
    row_number: int
    values: dict[str, Any]
    styles: tuple[tuple[Any, ...], ...]


@dataclass
class TableSnapshot:
    name: str
    sheet: str
    ref: str
    headers: tuple[str, ...]
    rows: list[TableRow]
    header_styles: tuple[tuple[Any, ...], ...]
    table_style: dict[str, Any]
    auto_filter_ref: str


@dataclass
class WorkbookSnapshot:
    path: Path
    sha256: str
    tables: dict[str, TableSnapshot]
    structure: dict[str, Any]
    package: dict[str, Any]
    identifier_occurrences: dict[str, list[dict[str, str]]]


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _money(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0.00")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, TypeError) as exc:
        raise RepairRefused("A reviewed financial precondition is not numeric.") from exc


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float):
        return format(Decimal(str(value)), "f")
    if value is None:
        return None
    if isinstance(value, (str, int, bool)):
        return value
    return str(value)


def _stable_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=True, sort_keys=True, separators=(",", ":"), default=_json_value)


def _digest_json(value: Any) -> str:
    return hashlib.sha256(_stable_json(value).encode("utf-8")).hexdigest().upper()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def fingerprint(path: Path, classification: str) -> FileFingerprint:
    item = path.resolve()
    if not item.is_file():
        raise RepairRefused(f"Required {classification} file is missing.")
    stat = item.stat()
    modified = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat().replace("+00:00", "Z")
    return FileFingerprint(classification, stat.st_size, modified, _sha256(item))


def _style_signature(cell: Any) -> tuple[Any, ...]:
    style = getattr(cell, "_style", None)
    return (
        tuple(style) if style is not None else (),
        _clean(getattr(cell, "number_format", "")),
        _clean(getattr(cell, "style", "")),
    )


def _table_style_signature(table: Any) -> dict[str, Any]:
    style = getattr(table, "tableStyleInfo", None)
    if style is None:
        return {}
    return {
        "name": _clean(getattr(style, "name", "")),
        "showFirstColumn": bool(getattr(style, "showFirstColumn", False)),
        "showLastColumn": bool(getattr(style, "showLastColumn", False)),
        "showRowStripes": bool(getattr(style, "showRowStripes", False)),
        "showColumnStripes": bool(getattr(style, "showColumnStripes", False)),
    }


def _dimension_manifest(items: Mapping[str, Any]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for key in sorted(items):
        item = items[key]
        result.append(
            {
                "key": str(key),
                "hidden": bool(getattr(item, "hidden", False)),
                "outlineLevel": int(getattr(item, "outlineLevel", 0) or 0),
                "width": _json_value(getattr(item, "width", None)),
                "height": _json_value(getattr(item, "height", None)),
                "style": int(getattr(item, "style_id", 0) or 0),
            }
        )
    return result


def _validation_manifest(worksheet: Any) -> list[dict[str, Any]]:
    container = getattr(worksheet, "data_validations", None)
    validations = getattr(container, "dataValidation", []) if container is not None else []
    return sorted(
        [
            {
                "sqref": _clean(getattr(item, "sqref", "")),
                "type": _clean(getattr(item, "type", "")),
                "operator": _clean(getattr(item, "operator", "")),
                "formula1": _clean(getattr(item, "formula1", "")),
                "formula2": _clean(getattr(item, "formula2", "")),
                "allowBlank": bool(getattr(item, "allow_blank", False)),
            }
            for item in validations
        ],
        key=_stable_json,
    )


def _defined_name_manifest(workbook: Any) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for item in workbook.defined_names.values():
        result.append(
            {
                "name": _clean(getattr(item, "name", "")),
                "attrText": _clean(getattr(item, "attr_text", "")),
                "localSheetId": getattr(item, "localSheetId", None),
                "hidden": bool(getattr(item, "hidden", False)),
            }
        )
    return sorted(result, key=_stable_json)


def _package_manifest(path: Path) -> dict[str, Any]:
    if not is_zipfile(path):
        raise RepairRefused("The repository workbook is not a valid Excel package.")
    with ZipFile(path) as package:
        names = sorted(package.namelist())
        macro_name = "xl/vbaProject.bin"
        macro_hash = ""
        if macro_name in names:
            macro_hash = hashlib.sha256(package.read(macro_name)).hexdigest().upper()
        content_types = package.read("[Content_Types].xml") if "[Content_Types].xml" in names else b""
        return {
            "members": names,
            "vbaProjectSha256": macro_hash,
            "macroEnabledContentType": b"application/vnd.ms-excel.sheet.macroEnabled.main+xml" in content_types,
        }


def _close_workbook(workbook: Any) -> None:
    for attribute in ("vba_archive", "_archive"):
        archive = getattr(workbook, attribute, None)
        if archive is not None:
            try:
                archive.close()
            except Exception:
                pass
            try:
                setattr(workbook, attribute, None)
            except Exception:
                pass
    try:
        workbook.close()
    except Exception:
        pass


def snapshot_workbook(path: Path) -> WorkbookSnapshot:
    workbook_path = path.resolve()
    if not workbook_path.is_file():
        raise RepairRefused("Workbook snapshot source is missing.")
    package = _package_manifest(workbook_path)
    workbook = load_workbook(workbook_path, keep_vba=True, data_only=False, keep_links=True)
    tables: dict[str, TableSnapshot] = {}
    occurrences = {identifier: [] for identifier in AUTHORIZED_IDS}
    structure: dict[str, Any] = {
        "sheetOrder": list(workbook.sheetnames),
        "sheetStates": {},
        "formulas": [],
        "definedNames": _defined_name_manifest(workbook),
        "externalLinkCount": len(getattr(workbook, "_external_links", [])),
        "namedStyles": [_clean(getattr(item, "name", "")) for item in getattr(workbook, "_named_styles", [])],
        "styleCatalog": {
            "cellStyles": [tuple(item) for item in getattr(workbook, "_cell_styles", [])],
            "numberFormats": list(getattr(workbook, "_number_formats", [])),
        },
        "worksheets": {},
    }
    try:
        for worksheet in workbook.worksheets:
            structure["sheetStates"][worksheet.title] = _clean(worksheet.sheet_state)
            structure["worksheets"][worksheet.title] = {
                "freezePanes": _clean(worksheet.freeze_panes),
                "mergedRanges": sorted(str(item) for item in worksheet.merged_cells.ranges),
                "validations": _validation_manifest(worksheet),
                "conditionalFormattingCount": len(worksheet.conditional_formatting),
                "columnDimensions": _dimension_manifest(worksheet.column_dimensions),
                "rowDimensions": _dimension_manifest(worksheet.row_dimensions),
                "showGridLines": bool(getattr(worksheet.sheet_view, "showGridLines", True)),
                "defaultRowHeight": _json_value(getattr(worksheet.sheet_format, "defaultRowHeight", None)),
                "defaultColWidth": _json_value(getattr(worksheet.sheet_format, "defaultColWidth", None)),
            }
            table_cell_map: dict[tuple[int, int], tuple[str, str]] = {}
            for table_name in worksheet.tables:
                if table_name in tables:
                    raise RepairRefused("A workbook table name is duplicated.")
                table = worksheet.tables[table_name]
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                headers = tuple(_clean(worksheet.cell(min_row, col).value) for col in range(min_col, max_col + 1))
                if not all(headers) or len(set(headers)) != len(headers):
                    raise RepairRefused(f"Table {table_name} has invalid headers.")
                rows: list[TableRow] = []
                for row_number in range(min_row + 1, max_row + 1):
                    values = {
                        header: worksheet.cell(row_number, min_col + offset).value
                        for offset, header in enumerate(headers)
                    }
                    styles = tuple(
                        _style_signature(worksheet.cell(row_number, col))
                        for col in range(min_col, max_col + 1)
                    )
                    if any(value not in (None, "") for value in values.values()):
                        rows.append(TableRow(row_number, values, styles))
                    for offset, header in enumerate(headers):
                        table_cell_map[(row_number, min_col + offset)] = (table_name, header)
                header_styles = tuple(
                    _style_signature(worksheet.cell(min_row, col))
                    for col in range(min_col, max_col + 1)
                )
                tables[table_name] = TableSnapshot(
                    name=table_name,
                    sheet=worksheet.title,
                    ref=table.ref,
                    headers=headers,
                    rows=rows,
                    header_styles=header_styles,
                    table_style=_table_style_signature(table),
                    auto_filter_ref=_clean(getattr(getattr(table, "autoFilter", None), "ref", "")),
                )

            for row in worksheet.iter_rows():
                for cell in row:
                    value_text = _clean(cell.value)
                    if getattr(cell, "data_type", "") == "f":
                        structure["formulas"].append(
                            {"sheet": worksheet.title, "cell": cell.coordinate, "formula": value_text}
                        )
                    if not value_text:
                        continue
                    folded = value_text.casefold()
                    for identifier in AUTHORIZED_IDS:
                        if identifier.casefold() not in folded:
                            continue
                        table_info = table_cell_map.get((cell.row, cell.column))
                        occurrences[identifier].append(
                            {
                                "sheet": worksheet.title,
                                "cell": cell.coordinate,
                                "table": table_info[0] if table_info else "",
                                "column": table_info[1] if table_info else "",
                            }
                        )
        structure["formulas"] = sorted(structure["formulas"], key=_stable_json)
    finally:
        _close_workbook(workbook)
    return WorkbookSnapshot(workbook_path, _sha256(workbook_path), tables, structure, package, occurrences)


def _table(snapshot: WorkbookSnapshot, table_name: str) -> TableSnapshot:
    table = snapshot.tables.get(table_name)
    if table is None:
        raise RepairRefused(f"Required table {table_name} is missing.")
    return table


def _matching_rows(table: TableSnapshot, column: str, values: Iterable[str]) -> list[TableRow]:
    accepted = {value.casefold() for value in values}
    return [row for row in table.rows if _clean(row.values.get(column)).casefold() in accepted]


def _row_values_for_digest(row: TableRow, headers: Sequence[str]) -> dict[str, Any]:
    return {header: _json_value(row.values.get(header)) for header in headers}


def _row_payload(row: TableRow, headers: Sequence[str]) -> dict[str, Any]:
    return {"values": _row_values_for_digest(row, headers), "styles": row.styles}


def _filtered_table_payload(snapshot: WorkbookSnapshot, table_name: str) -> list[dict[str, Any]]:
    table = _table(snapshot, table_name)
    rows = table.rows
    if table_name == sc.TBL_MATTERS:
        rows = [row for row in rows if _clean(row.values.get(sc.COL_MATTER_ID)) != AUTHORIZED_MATTER_ID]
    elif table_name == sc.TBL_TIME:
        rows = [row for row in rows if _clean(row.values.get(sc.COL_TIME_ENTRY_ID)) not in AUTHORIZED_ENTRY_IDS]
    return [_row_payload(row, table.headers) for row in rows]


def _protected_invoice_payload(snapshot: WorkbookSnapshot) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    invoice_key = PROTECTED_INVOICE.casefold()
    for table_name in sorted(snapshot.tables):
        table = snapshot.tables[table_name]
        for row in table.rows:
            if not any(_clean(value).casefold() == invoice_key for value in row.values.values()):
                continue
            result.append(
                {
                    "table": table_name,
                    "row": _row_values_for_digest(row, table.headers),
                    "styles": row.styles,
                }
            )
    return result


def _unrelated_data_digest(snapshot: WorkbookSnapshot) -> str:
    return _digest_json(
        {
            table_name: _filtered_table_payload(snapshot, table_name)
            for table_name in sorted(snapshot.tables)
        }
    )


def _financial_control_digest(snapshot: WorkbookSnapshot) -> str:
    financial_tokens = (
        "amount",
        "balance",
        "bill",
        "credit",
        "fee",
        "gross",
        "hst",
        "paid",
        "rate",
        "receivable",
        "tax",
        "total",
        "writeoff",
    )
    controls: dict[str, dict[str, str]] = {}
    for table_name in sorted(snapshot.tables):
        table = snapshot.tables[table_name]
        payload = _filtered_table_payload(snapshot, table_name)
        for header in table.headers:
            if not any(token in header.casefold() for token in financial_tokens):
                continue
            total = Decimal("0")
            numeric = False
            for item in payload:
                value = item["values"].get(header)
                if value in (None, ""):
                    continue
                try:
                    total += Decimal(str(value))
                    numeric = True
                except (InvalidOperation, ValueError, TypeError):
                    continue
            if numeric:
                controls.setdefault(table_name, {})[header] = format(total, "f")
    return _digest_json(controls)


def _validated_protected_invoice_payload(snapshot: WorkbookSnapshot) -> list[dict[str, Any]]:
    protected_rows = _protected_invoice_payload(snapshot)
    protected_tables = [item["table"] for item in protected_rows]
    for required_table in (sc.TBL_INVOICE_LOG, sc.TBL_RECEIVABLES, sc.TBL_TIME, sc.TBL_LEDGER):
        if protected_tables.count(required_table) != 1:
            raise RepairRefused("The protected invoice chain no longer has the reviewed row cardinality.")
    protected_text = _stable_json(protected_rows).casefold()
    if any(identifier.casefold() in protected_text for identifier in AUTHORIZED_IDS):
        raise RepairRefused("The protected invoice unexpectedly depends on an authorized orphan record.")
    invoice_log_rows = _matching_rows(
        _table(snapshot, sc.TBL_INVOICE_LOG),
        sc.COL_INV_INVOICE_NUM,
        [PROTECTED_INVOICE],
    )
    if len(invoice_log_rows) != 1:
        raise RepairRefused("The protected invoice log cardinality changed.")
    invoice_log = invoice_log_rows[0].values
    protected_amounts = {
        sc.COL_INV_TOTAL_FEES: Decimal("5000.00"),
        sc.COL_INV_TOTAL_DISBURSEMENTS: Decimal("0.00"),
        sc.COL_INV_TOTAL_TAX: Decimal("650.00"),
        sc.COL_INV_AGGREGATE_BILLED: Decimal("5650.00"),
    }
    if any(_money(invoice_log.get(column)) != expected for column, expected in protected_amounts.items()):
        raise RepairRefused("The protected invoice totals no longer match the reviewed repository state.")
    return protected_rows


def evaluate_preconditions(snapshot: WorkbookSnapshot) -> dict[str, Any]:
    matter_table = _table(snapshot, sc.TBL_MATTERS)
    time_table = _table(snapshot, sc.TBL_TIME)
    matter_rows = _matching_rows(matter_table, sc.COL_MATTER_ID, [AUTHORIZED_MATTER_ID])
    time_rows = _matching_rows(time_table, sc.COL_TIME_ENTRY_ID, AUTHORIZED_ENTRY_IDS)
    present = {
        AUTHORIZED_MATTER_ID: len(matter_rows),
        **{
            entry_id: len(_matching_rows(time_table, sc.COL_TIME_ENTRY_ID, [entry_id]))
            for entry_id in AUTHORIZED_ENTRY_IDS
        },
    }
    protected_rows = _validated_protected_invoice_payload(snapshot)
    if all(count == 0 for count in present.values()):
        if any(snapshot.identifier_occurrences[identifier] for identifier in AUTHORIZED_IDS):
            raise RepairRefused("An authorized ID remains as a dangling reference after record removal.")
        return {
            "status": "already_repaired",
            "authorizedIdsPresent": 0,
            "protectedInvoiceDigest": _digest_json(protected_rows),
            "unrelatedDataDigest": _unrelated_data_digest(snapshot),
            "financialControlDigest": _financial_control_digest(snapshot),
        }
    if present != {identifier: 1 for identifier in AUTHORIZED_IDS}:
        raise RepairRefused("Authorized-ID cardinality does not match the reviewed three-record state.")
    if len(matter_rows) != 1 or len(time_rows) != 2:
        raise RepairRefused("The reviewed matter/time record cardinality changed.")

    matter = matter_rows[0].values
    matter_nonblank = {key for key, value in matter.items() if _clean(value)}
    if matter_nonblank != EXPECTED_MATTER_NONBLANK_COLUMNS:
        raise RepairRefused("The sparse matter no longer has the reviewed field-presence shape.")
    if _clean(matter.get(sc.COL_MATTER_NUMBER)):
        raise RepairRefused("The authorized sparse matter unexpectedly has a matter number.")
    if _clean(matter.get(sc.COL_MATTER_STATUS)).casefold() != "active":
        raise RepairRefused("The sparse matter status no longer matches the reviewed state.")

    time_rows = sorted(time_rows, key=lambda row: _clean(row.values.get(sc.COL_TIME_ENTRY_ID)))
    first_without_id = {
        key: _json_value(value)
        for key, value in time_rows[0].values.items()
        if key != sc.COL_TIME_ENTRY_ID
    }
    second_without_id = {
        key: _json_value(value)
        for key, value in time_rows[1].values.items()
        if key != sc.COL_TIME_ENTRY_ID
    }
    if first_without_id != second_without_id:
        raise RepairRefused("The two authorized time records are no longer identical apart from their IDs.")
    for row in time_rows:
        values = row.values
        nonblank = {key for key, value in values.items() if _clean(value)}
        if nonblank != EXPECTED_TIME_NONBLANK_COLUMNS:
            raise RepairRefused("An authorized time record no longer has the reviewed field-presence shape.")
        if _clean(values.get(sc.COL_TIME_MATTER_ID)) != AUTHORIZED_MATTER_ID:
            raise RepairRefused("An authorized time record no longer points to the authorized sparse matter.")
        if _clean(values.get(sc.COL_TIME_CLIENT_ID)) != _clean(matter.get(sc.COL_MATTER_CLIENT_ID)):
            raise RepairRefused("The reviewed client linkage between the authorized records changed.")
        if _clean(values.get(sc.COL_TIME_DATE)) != "2026-08-05":
            raise RepairRefused("An authorized time record date changed from the reviewed state.")
        expected_money = {
            sc.COL_TIME_HOURS: Decimal("0.00"),
            sc.COL_TIME_RATE: Decimal("0.00"),
            sc.COL_TIME_SHARE_PCT: Decimal("0.00"),
            sc.COL_TIME_GROSS: Decimal("5000.00"),
            sc.COL_TIME_NET: Decimal("5000.00"),
            sc.COL_TIME_HST: Decimal("650.00"),
            sc.COL_TIME_TOTAL: Decimal("0.00"),
            sc.COL_TIME_SECONDS: Decimal("0.00"),
            sc.COL_TIME_INVOICE_TOTAL: Decimal("0.00"),
            sc.COL_TIME_INVOICE_AMOUNT_PAID: Decimal("0.00"),
            sc.COL_TIME_INVOICE_BALANCE_DUE: Decimal("0.00"),
        }
        if any(_money(values.get(column)) != expected for column, expected in expected_money.items()):
            raise RepairRefused("An authorized time record financial precondition changed.")
        if _clean(values.get(sc.COL_TIME_INVOICE_STATUS)).casefold() != "unbilled":
            raise RepairRefused("An authorized time record is no longer unbilled.")
        for column in (
            sc.COL_TIME_INVOICE_REF,
            sc.COL_TIME_PAYMENT_STATUS,
            sc.COL_TIME_INVOICE_DATE,
            sc.COL_TIME_REISSUE_INVOICE_NUM,
            sc.COL_TIME_LOCK_AUDIT,
            sc.COL_TIME_STATUS,
        ):
            if _clean(values.get(column)):
                raise RepairRefused("An invoice, payment, reissue, lock, or status dependency now blocks repair.")

    expected_occurrences = {
        AUTHORIZED_MATTER_ID: {
            (sc.SHEET_MATTERS, sc.TBL_MATTERS, sc.COL_MATTER_ID),
            (sc.SHEET_TIME, sc.TBL_TIME, sc.COL_TIME_MATTER_ID),
        },
        AUTHORIZED_ENTRY_IDS[0]: {(sc.SHEET_TIME, sc.TBL_TIME, sc.COL_TIME_ENTRY_ID)},
        AUTHORIZED_ENTRY_IDS[1]: {(sc.SHEET_TIME, sc.TBL_TIME, sc.COL_TIME_ENTRY_ID)},
    }
    expected_counts = {AUTHORIZED_MATTER_ID: 3, AUTHORIZED_ENTRY_IDS[0]: 1, AUTHORIZED_ENTRY_IDS[1]: 1}
    for identifier in AUTHORIZED_IDS:
        occurrences = snapshot.identifier_occurrences.get(identifier, [])
        if len(occurrences) != expected_counts[identifier]:
            raise RepairRefused("An unexpected dependent or non-table reference now blocks repair.")
        locations = {(item["sheet"], item["table"], item["column"]) for item in occurrences}
        if locations != expected_occurrences[identifier]:
            raise RepairRefused("An authorized ID appears outside its reviewed operational columns.")

    return {
        "status": "ready",
        "authorizedIdsPresent": 3,
        "matterDependents": 2,
        "protectedInvoiceDigest": _digest_json(protected_rows),
        "unrelatedDataDigest": _unrelated_data_digest(snapshot),
        "financialControlDigest": _financial_control_digest(snapshot),
    }


def _find_table_object(workbook: Any, table_name: str) -> tuple[Any, Any]:
    matches: list[tuple[Any, Any]] = []
    for worksheet in workbook.worksheets:
        if table_name in worksheet.tables:
            matches.append((worksheet, worksheet.tables[table_name]))
    if len(matches) != 1:
        raise RepairRefused(f"Expected exactly one {table_name} table object.")
    return matches[0]


def _delete_exact_table_rows(workbook: Any, table_name: str, key_column: str, identifiers: Sequence[str]) -> None:
    worksheet, table = _find_table_object(workbook, table_name)
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [_clean(worksheet.cell(min_row, col).value) for col in range(min_col, max_col + 1)]
    if key_column not in headers:
        raise RepairRefused(f"Required key column is missing from {table_name}.")
    key_col = min_col + headers.index(key_column)
    rows_for_id: dict[str, list[int]] = {identifier: [] for identifier in identifiers}
    for row_number in range(min_row + 1, max_row + 1):
        value = _clean(worksheet.cell(row_number, key_col).value)
        if value in rows_for_id:
            rows_for_id[value].append(row_number)
    if any(len(rows) != 1 for rows in rows_for_id.values()):
        raise RepairRefused(f"Exact-key deletion cardinality changed in {table_name}.")
    delete_rows = sorted((rows[0] for rows in rows_for_id.values()), reverse=True)
    for row_number in delete_rows:
        worksheet.delete_rows(row_number, 1)
    new_max_row = max_row - len(delete_rows)
    if new_max_row <= min_row:
        raise RepairRefused(f"Exact-key deletion would invalidate {table_name}.")
    table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"


def _assert_matter_removal_safe(workbook: Any) -> None:
    for worksheet in workbook.worksheets:
        for table_name in worksheet.tables:
            table = worksheet.tables[table_name]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            headers = [_clean(worksheet.cell(min_row, col).value) for col in range(min_col, max_col + 1)]
            for row_number in range(min_row + 1, max_row + 1):
                for offset, header in enumerate(headers):
                    value = _clean(worksheet.cell(row_number, min_col + offset).value)
                    if AUTHORIZED_MATTER_ID.casefold() not in value.casefold():
                        continue
                    is_own_key = table_name == sc.TBL_MATTERS and header == sc.COL_MATTER_ID
                    if not is_own_key:
                        raise RepairRefused("The sparse matter still has a dependent record.")


def apply_candidate_record_removal(candidate_path: Path) -> dict[str, Any]:
    candidate = candidate_path.resolve()
    before = snapshot_workbook(candidate)
    preconditions = evaluate_preconditions(before)
    if preconditions["status"] == "already_repaired":
        return {"status": "already_repaired", "changed": False, "sha256": before.sha256}
    workbook = load_workbook(candidate, keep_vba=True, data_only=False, keep_links=True)
    staged = candidate.with_name(f".{candidate.stem}.authorized-repair-{uuid4().hex}{candidate.suffix}")
    try:
        _delete_exact_table_rows(workbook, sc.TBL_TIME, sc.COL_TIME_ENTRY_ID, AUTHORIZED_ENTRY_IDS)
        _assert_matter_removal_safe(workbook)
        _delete_exact_table_rows(workbook, sc.TBL_MATTERS, sc.COL_MATTER_ID, [AUTHORIZED_MATTER_ID])
        workbook.save(staged)
        _close_workbook(workbook)
        workbook = None
        if not is_zipfile(staged):
            raise RepairRefused("The repaired candidate did not save as a valid Excel package.")
        os.replace(staged, candidate)
    finally:
        if workbook is not None:
            _close_workbook(workbook)
        staged.unlink(missing_ok=True)
    return {
        "status": "candidate_repaired",
        "changed": True,
        "inputSha256": before.sha256,
        "outputSha256": _sha256(candidate),
        "removed": {sc.TBL_TIME: 2, sc.TBL_MATTERS: 1},
    }


def _expected_table_ref(before_ref: str, removed: int) -> str:
    min_col, min_row, max_col, max_row = range_boundaries(before_ref)
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row - removed}"


def _integrity_summary(workbook_path: Path) -> dict[str, Any]:
    service = WorkbookIntegrityService(AppPaths(PROJECT_ROOT))
    report = service.check(
        workbook_path=workbook_path,
        schema_path=PROJECT_ROOT / "schema" / "workbook_schema.yml",
    )
    warning_counts: dict[str, int] = {}
    for issue in report.issues:
        if issue.severity == "warning":
            warning_counts[issue.code] = warning_counts.get(issue.code, 0) + 1
    return {
        "ok": report.ok,
        "tablesChecked": report.tables_checked,
        "rowsChecked": report.rows_checked,
        "errorCount": report.error_count,
        "warningCount": report.warning_count,
        "warningCounts": warning_counts,
        "hasTargetMatterNumberError": any(
            issue.code == "missing_required_value"
            and issue.table == sc.TBL_MATTERS
            and issue.column == sc.COL_MATTER_NUMBER
            for issue in report.issues
        ),
    }


def verify_repaired_candidate(before: WorkbookSnapshot, candidate_path: Path) -> dict[str, Any]:
    after = snapshot_workbook(candidate_path)
    postconditions = evaluate_preconditions(after)
    if postconditions["status"] != "already_repaired":
        raise RepairRefused("The candidate still contains an authorized orphan record.")
    if any(after.identifier_occurrences[identifier] for identifier in AUTHORIZED_IDS):
        raise RepairRefused("The candidate still contains an authorized ID reference.")
    if set(before.tables) != set(after.tables):
        raise RepairRefused("Workbook table membership changed during candidate repair.")
    if before.package["members"] != after.package["members"]:
        raise RepairRefused("Excel package membership changed during candidate repair.")
    if before.package["vbaProjectSha256"] != after.package["vbaProjectSha256"]:
        raise RepairRefused("The VBA project state was not preserved byte-for-byte.")
    if not after.package["macroEnabledContentType"]:
        raise RepairRefused("The candidate lost its macro-enabled workbook content type.")
    if before.structure != after.structure:
        raise RepairRefused("Formulas, names, validations, formatting catalog, or workbook structure changed.")

    for table_name in sorted(before.tables):
        before_table = before.tables[table_name]
        after_table = after.tables[table_name]
        if before_table.sheet != after_table.sheet or before_table.headers != after_table.headers:
            raise RepairRefused(f"Table structure changed for {table_name}.")
        if before_table.header_styles != after_table.header_styles:
            raise RepairRefused(f"Table header formatting changed for {table_name}.")
        if before_table.table_style != after_table.table_style:
            raise RepairRefused(f"Table style changed for {table_name}.")
        if before_table.auto_filter_ref != after_table.auto_filter_ref:
            raise RepairRefused(f"Table filter metadata changed for {table_name}.")
        removed = TARGET_TABLE_REMOVALS.get(table_name, 0)
        expected_ref = _expected_table_ref(before_table.ref, removed) if removed else before_table.ref
        if after_table.ref != expected_ref:
            raise RepairRefused(f"Table dimensions changed unexpectedly for {table_name}.")
        if _filtered_table_payload(before, table_name) != _filtered_table_payload(after, table_name):
            raise RepairRefused(f"An unrelated row or row format changed in {table_name}.")

    protected_before = _digest_json(_protected_invoice_payload(before))
    protected_after = _digest_json(_protected_invoice_payload(after))
    if protected_before != protected_after:
        raise RepairRefused("The protected invoice chain changed during candidate repair.")
    unrelated_before = _unrelated_data_digest(before)
    unrelated_after = _unrelated_data_digest(after)
    if unrelated_before != unrelated_after:
        raise RepairRefused("Unrelated workbook data changed during candidate repair.")
    financial_before = _financial_control_digest(before)
    financial_after = _financial_control_digest(after)
    if financial_before != financial_after:
        raise RepairRefused("Unrelated financial controls changed during candidate repair.")

    integrity = _integrity_summary(candidate_path)
    if not integrity["ok"] or integrity["errorCount"] != 0:
        raise RepairRefused("The candidate did not pass canonical workbook integrity.")
    if integrity["warningCount"] != 7 or integrity["warningCounts"] != EXPECTED_WARNING_COUNTS:
        raise RepairRefused("The candidate did not retain the seven independent warnings exactly.")
    if integrity["hasTargetMatterNumberError"]:
        raise RepairRefused("The target missing MatterNumber error remains in the candidate.")
    return {
        "ok": True,
        "candidateSha256": after.sha256,
        "removed": {sc.TBL_TIME: 2, sc.TBL_MATTERS: 1},
        "macrosPreserved": True,
        "vbaProjectPresent": bool(after.package["vbaProjectSha256"]),
        "packageMembersPreserved": True,
        "formulasPreserved": True,
        "definedNamesPreserved": True,
        "validationsPreserved": True,
        "formattingPreserved": True,
        "tableStructurePreserved": True,
        "protectedInvoiceUnchanged": True,
        "unrelatedDataUnchanged": True,
        "financialControlsUnchanged": True,
        "integrity": integrity,
    }


def _contains_identifier_bytes(path: Path, identifiers: Sequence[str]) -> dict[str, bool]:
    found = {identifier: False for identifier in identifiers}
    if not is_zipfile(path):
        raise RepairRefused("A boundary workbook is not a valid Excel package.")
    needles = {
        identifier: (
            identifier.encode("utf-8"),
            identifier.encode("utf-16-le"),
            identifier.encode("utf-16-be"),
        )
        for identifier in identifiers
    }
    with ZipFile(path) as package:
        for name in package.namelist():
            payload = package.read(name)
            for identifier, encodings in needles.items():
                if not found[identifier] and any(encoding in payload for encoding in encodings):
                    found[identifier] = True
    return found


def _load_boundary_paths() -> dict[str, Path]:
    settings_path = AppPaths(PROJECT_ROOT).user_settings_path()
    try:
        settings = json.loads(settings_path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise RepairRefused("Configured workbook boundaries could not be read.") from exc
    local_dir_text = _clean(settings.get("localDataDir"))
    master_dir_text = _clean(settings.get("masterDataDir"))
    if not local_dir_text or not master_dir_text:
        raise RepairRefused("Configured live or cloud workbook directory is missing.")
    local_dir = Path(local_dir_text).expanduser()
    master_dir = Path(master_dir_text).expanduser()
    return {
        "live_local_cspm": (local_dir / "CSPM.xlsm").resolve(),
        "live_local_dockets": (local_dir / "Dockets.xlsm").resolve(),
        "cloud_master_cspm": (master_dir / "CSPM.xlsm").resolve(),
        "cloud_master_dockets": (master_dir / "Dockets.xlsm").resolve(),
    }


def assert_repository_target(
    target_path: Path,
    *,
    repository_path: Path = REPOSITORY_WORKBOOK,
    forbidden_paths: Sequence[Path] = (),
) -> Path:
    target = target_path.expanduser().resolve()
    expected = repository_path.expanduser().resolve()
    if target != expected:
        raise RepairRefused("Unknown workbook target refused; only the repository snapshot is permitted.")
    if target.is_symlink():
        raise RepairRefused("A symlinked workbook target is not permitted.")
    for forbidden in forbidden_paths:
        if target == forbidden.expanduser().resolve():
            raise RepairRefused("A configured live or cloud workbook target was refused.")
    if not target.is_file():
        raise RepairRefused("The repository workbook target is missing.")
    return target


def _boundary_fingerprints(boundaries: Mapping[str, Path], historical_evidence: Path) -> dict[str, FileFingerprint]:
    result = {label: fingerprint(path, label) for label, path in boundaries.items()}
    result["historical_evidence"] = fingerprint(historical_evidence, "historical_evidence")
    return result


def _verify_boundary_content(boundaries: Mapping[str, Path], historical_evidence: Path) -> None:
    for label, path in boundaries.items():
        found = _contains_identifier_bytes(path, AUTHORIZED_IDS)
        if any(found.values()):
            raise RepairRefused(f"Configured boundary {label} unexpectedly contains an authorized ID.")
    historical_found = _contains_identifier_bytes(historical_evidence, AUTHORIZED_ENTRY_IDS)
    if not all(historical_found.values()):
        raise RepairRefused("Historical evidence no longer contains both reviewed orphan-entry IDs.")


def _assert_fingerprints_unchanged(
    before: Mapping[str, FileFingerprint],
    after: Mapping[str, FileFingerprint],
) -> None:
    if set(before) != set(after):
        raise RepairRefused("Boundary fingerprint membership changed.")
    for label in before:
        if before[label].sha256 != after[label].sha256 or before[label].bytes != after[label].bytes:
            raise RepairRefused(f"Read-only boundary {label} changed during repository repair.")


def _atomic_write_json(path: Path, payload: Mapping[str, Any]) -> None:
    temporary = path.with_name(f".{path.name}.{uuid4().hex}.tmp")
    try:
        temporary.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _relative(path: Path) -> str:
    return str(path.resolve().relative_to(PROJECT_ROOT)).replace("\\", "/")


def _public_preflight(
    target_fingerprint: FileFingerprint,
    preconditions: Mapping[str, Any],
    boundary_fingerprints: Mapping[str, FileFingerprint],
    historical_fingerprint: FileFingerprint,
) -> dict[str, Any]:
    return {
        "ok": True,
        "status": preconditions["status"],
        "targetClassification": "divergent_repository_snapshot",
        "targetFingerprint": target_fingerprint.as_dict(),
        "authorizedIds": list(AUTHORIZED_IDS),
        "authorizedIdsPresent": int(preconditions.get("authorizedIdsPresent", 0)),
        "matterDependentCount": int(preconditions.get("matterDependents", 0)),
        "protectedInvoiceSnapshotDigest": preconditions["protectedInvoiceDigest"],
        "unrelatedDataDigest": preconditions["unrelatedDataDigest"],
        "financialControlDigest": preconditions["financialControlDigest"],
        "readOnlyBoundaries": {
            label: item.as_dict() for label, item in sorted(boundary_fingerprints.items())
        },
        "historicalEvidence": historical_fingerprint.as_dict(),
        "mutation": "none",
    }


def execute_repair(target_path: Path, historical_evidence: Path, *, apply: bool) -> dict[str, Any]:
    boundaries = _load_boundary_paths()
    target = assert_repository_target(
        target_path,
        forbidden_paths=list(boundaries.values()),
    )
    evidence = historical_evidence.expanduser().resolve()
    before_boundary_fingerprints = _boundary_fingerprints(boundaries, evidence)
    _verify_boundary_content(boundaries, evidence)
    before_target_fingerprint = fingerprint(target, "repository_cspm")
    authoritative_hashes = {
        before_boundary_fingerprints["live_local_cspm"].sha256,
        before_boundary_fingerprints["cloud_master_cspm"].sha256,
    }
    if before_target_fingerprint.sha256 in authoritative_hashes:
        raise RepairRefused("The repository workbook is no longer a divergent snapshot; repair was refused.")
    before_snapshot = snapshot_workbook(target)
    preconditions = evaluate_preconditions(before_snapshot)
    boundary_only = {
        label: before_boundary_fingerprints[label]
        for label in boundaries
    }
    public = _public_preflight(
        before_target_fingerprint,
        preconditions,
        boundary_only,
        before_boundary_fingerprints["historical_evidence"],
    )
    if not apply or preconditions["status"] == "already_repaired":
        return public

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%SZ")
    repair_root = PROJECT_ROOT / "backups" / "CSPM" / f"repository_record_repair_{stamp}"
    repair_root.mkdir(parents=True, exist_ok=False)
    backup_path = repair_root / "CSPM.before-repair.xlsm"
    candidate_dir = repair_root / "candidate"
    candidate_dir.mkdir()
    candidate_path = candidate_dir / "CSPM.xlsm"
    audit_path = repair_root / "repair_audit.json"

    shutil.copy2(target, backup_path)
    backup_fingerprint = fingerprint(backup_path, "repository_pre_repair_backup")
    if backup_fingerprint.sha256 != before_target_fingerprint.sha256:
        raise RepairRefused("The governed pre-repair backup failed hash verification.")
    shutil.copy2(target, candidate_path)
    if _sha256(candidate_path) != before_target_fingerprint.sha256:
        raise RepairRefused("The isolated candidate failed its initial hash verification.")

    removal = apply_candidate_record_removal(candidate_path)
    if not removal.get("changed"):
        raise RepairRefused("The isolated candidate was not changed as expected.")
    candidate_verification = verify_repaired_candidate(before_snapshot, candidate_path)

    audit: dict[str, Any] = {
        "operation": "repository-only confirmed orphan-record repair",
        "status": "candidate-verified-awaiting-promotion",
        "authorizedIds": list(AUTHORIZED_IDS),
        "targetClassification": "divergent_repository_snapshot",
        "startedAtUtc": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "beforeTarget": before_target_fingerprint.as_dict(),
        "backup": {
            **backup_fingerprint.as_dict(),
            "relativePath": _relative(backup_path),
            "verifiedAgainstBefore": True,
        },
        "candidate": {
            "relativePath": _relative(candidate_path),
            **candidate_verification,
        },
        "readOnlyBoundariesBefore": {
            label: item.as_dict() for label, item in sorted(before_boundary_fingerprints.items())
        },
        "rollback": {
            "sourceRelativePath": _relative(backup_path),
            "destinationRelativePath": "data/CSPM.xlsm",
            "expectedBackupSha256": backup_fingerprint.sha256,
            "instruction": "Verify the backup hash, copy it to the repository workbook path, then rerun canonical integrity.",
        },
    }
    _atomic_write_json(audit_path, audit)

    if _sha256(target) != before_target_fingerprint.sha256:
        raise RepairRefused("The repository workbook changed after candidate creation; promotion was refused.")
    before_promotion_boundaries = _boundary_fingerprints(boundaries, evidence)
    _assert_fingerprints_unchanged(before_boundary_fingerprints, before_promotion_boundaries)

    staged_target = target.with_name(f".{target.stem}.repository-repair-{uuid4().hex}.staged{target.suffix}")
    try:
        shutil.copy2(candidate_path, staged_target)
        if _sha256(staged_target) != candidate_verification["candidateSha256"]:
            raise RepairRefused("The same-volume promotion stage failed hash verification.")
        os.replace(staged_target, target)
    except PermissionError as exc:
        raise RepairRefused("Repository promotion was refused because the target is in use.") from exc
    finally:
        staged_target.unlink(missing_ok=True)

    promoted_verification = verify_repaired_candidate(before_snapshot, target)
    if promoted_verification["candidateSha256"] != candidate_verification["candidateSha256"]:
        raise RepairRefused("The promoted repository workbook does not match the verified candidate.")
    after_target_fingerprint = fingerprint(target, "repository_cspm")
    after_boundary_fingerprints = _boundary_fingerprints(boundaries, evidence)
    _assert_fingerprints_unchanged(before_boundary_fingerprints, after_boundary_fingerprints)
    _verify_boundary_content(boundaries, evidence)

    audit.update(
        {
            "status": "promoted-and-reverified",
            "completedAtUtc": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "afterTarget": after_target_fingerprint.as_dict(),
            "promoted": promoted_verification,
            "readOnlyBoundariesAfter": {
                label: item.as_dict() for label, item in sorted(after_boundary_fingerprints.items())
            },
            "liveAndCloudUnchanged": True,
            "historicalEvidenceUnchanged": True,
            "conflictCheckingChanged": False,
        }
    )
    _atomic_write_json(audit_path, audit)
    return {
        "ok": True,
        "status": audit["status"],
        "targetClassification": audit["targetClassification"],
        "authorizedIds": audit["authorizedIds"],
        "removed": candidate_verification["removed"],
        "beforeSha256": before_target_fingerprint.sha256,
        "afterSha256": after_target_fingerprint.sha256,
        "backupRelativePath": _relative(backup_path),
        "backupSha256": backup_fingerprint.sha256,
        "candidateRelativePath": _relative(candidate_path),
        "candidateSha256": candidate_verification["candidateSha256"],
        "auditRelativePath": _relative(audit_path),
        "candidateIntegrity": candidate_verification["integrity"],
        "promotedIntegrity": promoted_verification["integrity"],
        "protectedInvoiceUnchanged": promoted_verification["protectedInvoiceUnchanged"],
        "liveAndCloudUnchanged": True,
        "historicalEvidenceUnchanged": True,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=REPOSITORY_WORKBOOK,
        help="Repository CSPM.xlsm target. Any other workbook is refused.",
    )
    parser.add_argument(
        "--historical-evidence",
        type=Path,
        required=True,
        help="Read-only historical workbook containing both reviewed orphan entry IDs.",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Create backup/candidate, verify, and atomically promote the repository-only repair.",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(list(argv) if argv is not None else None)
    try:
        result = execute_repair(args.workbook, args.historical_evidence, apply=bool(args.apply))
    except RepairRefused as exc:
        result = {
            "ok": False,
            "status": "refused",
            "message": str(exc),
            "verifiedBackupMayExist": True,
        }
        print(json.dumps(result, indent=2, sort_keys=True))
        return 2
    print(json.dumps(result, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
