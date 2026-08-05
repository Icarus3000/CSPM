import openpyxl
import os
import json

def analyze_workbook(filepath):
    print(f"Analyzing {filepath}...")
    wb = openpyxl.load_workbook(filepath, data_only=False, keep_vba=True)
    report = {
        "workbook": os.path.basename(filepath),
        "worksheets": wb.sheetnames,
        "tables": {},
        "defined_names": list(wb.defined_names.keys()),
        "vba_present": wb.vba_archive is not None
    }
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for table in sheet.tables.values():
            table_name = table.name
            ref = table.ref
            
            # Get columns
            min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(ref)
            columns = []
            has_formulas = False
            for col in range(min_col, max_col + 1):
                cell_value = sheet.cell(row=min_row, column=col).value
                columns.append(str(cell_value) if cell_value else f"Col{col}")
            
            # Check for formulas in the first data row
            if max_row > min_row:
                for col in range(min_col, max_col + 1):
                    val = sheet.cell(row=min_row + 1, column=col).value
                    if isinstance(val, str) and val.startswith('='):
                        has_formulas = True
                        break
            
            row_count = max_row - min_row
            
            report["tables"][table_name] = {
                "sheet": sheet_name,
                "columns": columns,
                "row_count": row_count,
                "has_formulas": has_formulas
            }
            
    return report

if __name__ == "__main__":
    cspm_path = r"C:\Projects\__CSPM\data\CSPM.xlsm"
    dockets_path = r"C:\Projects\__CSPM\data\Dockets.xlsm"
    
    final_report = {}
    
    if os.path.exists(cspm_path):
        final_report["CSPM"] = analyze_workbook(cspm_path)
    else:
        print(f"File not found: {cspm_path}")
        
    if os.path.exists(dockets_path):
        final_report["Dockets"] = analyze_workbook(dockets_path)
    else:
        print(f"File not found: {dockets_path}")
        
    out_path = r"C:\Projects\__CSPM\outputs\full_inventory.json"
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w") as f:
        json.dump(final_report, f, indent=2)
    print(f"Full inventory generated at {out_path}")
