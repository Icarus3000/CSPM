from __future__ import annotations

import json
import re
import shutil
import sys
from collections import Counter, defaultdict, deque
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
PYTHON_SRC = REPO_ROOT / "src" / "python"
if str(PYTHON_SRC) not in sys.path:
    sys.path.insert(0, str(PYTHON_SRC))

from domain import schema_constants as sc  # noqa: E402
from repositories.excel_repo import ExcelRepo  # noqa: E402
from services.paths import AppPaths  # noqa: E402


SOURCE = REPO_ROOT / "data" / "Dockets.xlsm"
TARGET = REPO_ROOT / "data" / "CSPM.xlsm"
REPORT_DIR = REPO_ROOT / "outputs" / "reports"


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value.strip())
    return str(value).strip()


def date_key(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean(value)
    return text[:10] if text else ""


def dec(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0.00")
    if isinstance(value, str):
        value = value.replace("$", "").replace(",", "").strip()
        if not value:
            return Decimal("0.00")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def money_float(value: Any) -> float:
    return float(dec(value))


def compact_description(value: Any) -> str:
    return re.sub(r"\s*\[[^\]]+\]\s*$", "", clean(value))


def is_actual_invoice(value: Any) -> bool:
    return bool(re.match(r"^\d{2}-\d{4}(?:-[A-Z])?$", clean(value)))


def invoice_bucket(value: Any) -> str:
    text = clean(value)
    upper = text.upper()
    if not text:
        return "blank"
    if is_actual_invoice(text):
        return "invoice"
    if upper == "BILLED":
        return "legacy_billed"
    if "HOLD" in upper or "FORGOT" in upper:
        return "hold"
    if "FREE" in upper or "DO NOT BILL" in upper or "WRITE OFF" in upper:
        return "no_bill"
    return "other"


def invoice_status(value: Any) -> str:
    bucket = invoice_bucket(value)
    if bucket == "invoice":
        return "Invoiced"
    if bucket == "legacy_billed":
        return "Legacy Billed"
    if bucket == "hold":
        return "Held for Billing"
    if bucket == "no_bill":
        return "Not Billable"
    if bucket == "other":
        return "Marked"
    return "Not Invoiced"


def time_status_for_invoice_marker(value: Any) -> str:
    bucket = invoice_bucket(value)
    if bucket in {"invoice", "legacy_billed", "no_bill"}:
        return "Billed"
    return "WIP"


def source_signature(row: dict[str, Any], description: str | None = None) -> tuple[Any, ...]:
    return (
        date_key(row.get("Date")),
        compact_description(description if description is not None else row.get("Description")),
        dec(row.get("Time (in hrs)")),
        dec(row.get("Hourly Rate/Flat Fee")),
        dec(row.get("Percentage")),
        dec(row.get("Amount to CS")),
    )


def time_signature(row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        date_key(row.get(sc.COL_TIME_DATE)),
        compact_description(row.get(sc.COL_TIME_DESC)),
        dec(row.get(sc.COL_TIME_HOURS)),
        dec(row.get(sc.COL_TIME_RATE)),
        dec(row.get(sc.COL_TIME_SHARE_PCT)),
        dec(row.get(sc.COL_TIME_NET)),
    )


def read_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, keep_vba=True, read_only=True)
    try:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = [clean(value) for value in next(rows)]
        result: list[dict[str, Any]] = []
        for row_number, values in enumerate(rows, start=2):
            row = {
                header: values[idx] if idx < len(values) else None
                for idx, header in enumerate(headers)
                if header
            }
            if any(clean(value) for value in row.values()):
                row["_row"] = row_number
                result.append(row)
        return result
    finally:
        wb.close()


def build_receivable_state(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    state: dict[str, dict[str, Any]] = {}
    for row in rows:
        invoice_num = clean(row.get("InvoiceNum"))
        if not is_actual_invoice(invoice_num):
            continue
        total = money_float(row.get("Total_Invoiced"))
        amount_paid = money_float(row.get("Amount_Paid"))
        credits_adj = money_float(row.get("Credits/Adj"))
        balance_due = money_float(row.get("Balance_Due"))
        status = clean(row.get("Status"))
        status_upper = status.upper()

        if status_upper in {"VOID", "VOIDED", "CANCELLED", "CANCELED"}:
            payment_status = "Void"
            amount_paid = 0.0
            balance_due = 0.0
        elif status_upper in {"PAID", "CLOSED"}:
            payment_status = "Paid"
            if amount_paid <= 0 and total > 0:
                amount_paid = max(0.0, total + credits_adj)
            balance_due = 0.0
        elif amount_paid > 0 and balance_due > 0:
            payment_status = "Partial"
        elif balance_due <= 0 and (amount_paid > 0 or total <= 0):
            payment_status = "Paid"
        else:
            payment_status = "Pending"

        state[invoice_num] = {
            "invoiceDate": date_key(row.get("Date")),
            "paymentStatus": payment_status,
            "invoiceTotal": round(float(total), 2),
            "invoiceAmountPaid": round(float(amount_paid), 2),
            "invoiceBalanceDue": round(float(max(0.0, balance_due)), 2),
            "receivableStatus": status,
        }
    return state


def payment_state_for_marker(invoice_ref: str, receivables: dict[str, dict[str, Any]]) -> dict[str, Any]:
    bucket = invoice_bucket(invoice_ref)
    if bucket == "invoice":
        return receivables.get(
            invoice_ref,
            {
                "invoiceDate": "",
                "paymentStatus": "Unknown",
                "invoiceTotal": 0.0,
                "invoiceAmountPaid": 0.0,
                "invoiceBalanceDue": 0.0,
                "receivableStatus": "",
            },
        )
    return {
        "invoiceDate": "",
        "paymentStatus": {
            "legacy_billed": "Unknown",
            "hold": "Not Invoiced",
            "no_bill": "Not Billable",
            "other": "Unknown",
        }.get(bucket, "Not Invoiced"),
        "invoiceTotal": 0.0,
        "invoiceAmountPaid": 0.0,
        "invoiceBalanceDue": 0.0,
        "receivableStatus": "",
    }


def build_source_lookup(source_rows: list[dict[str, Any]]) -> dict[tuple[Any, ...], deque[dict[str, Any]]]:
    lookup: dict[tuple[Any, ...], deque[dict[str, Any]]] = defaultdict(deque)
    for row in source_rows:
        lookup[source_signature(row)].append(row)
        sub_client = clean(row.get("Sub-Client"))
        if sub_client:
            lookup[source_signature(row, f"{clean(row.get('Description'))} {sub_client}".strip())].append(row)
    return lookup


def locate_headers(ws, required_headers: list[str]) -> dict[str, int]:
    headers = [clean(ws.cell(row=1, column=col).value) for col in range(1, ws.max_column + 1)]
    header_map = {header: idx + 1 for idx, header in enumerate(headers) if header}
    next_col = len(headers) + 1
    for header in required_headers:
        if header not in header_map:
            ws.cell(row=1, column=next_col, value=header)
            header_map[header] = next_col
            next_col += 1
    return header_map


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description="Sync canonical TimeEntries invoice/payment state from Dockets.xlsm.")
    parser.add_argument("--apply", action="store_true", help="Write changes to data/CSPM.xlsm.")
    args = parser.parse_args()

    if not SOURCE.exists():
        raise SystemExit(f"Source workbook not found: {SOURCE}")
    if not TARGET.exists():
        raise SystemExit(f"Target workbook not found: {TARGET}")

    source_rows = read_rows(SOURCE, "Dockets")
    receivable_state = build_receivable_state(read_rows(SOURCE, "Receivables"))
    source_lookup = build_source_lookup(source_rows)

    repo = ExcelRepo(AppPaths(REPO_ROOT))
    if args.apply:
        backup_dir = REPO_ROOT / "data" / "recovery_backups" / f"invoice_payment_state_sync_{datetime.now():%Y%m%d_%H%M%S}"
        backup_dir.mkdir(parents=True, exist_ok=False)
        shutil.copy2(TARGET, backup_dir / "CSPM.before_invoice_payment_state_sync.xlsm")
        repo.ensure_schema()

    wb = load_workbook(TARGET, keep_vba=True)
    matched = 0
    unmatched_time_entries: list[dict[str, Any]] = []
    unmatched_source_rows: list[int] = []
    status_counts: Counter[str] = Counter()
    payment_counts: Counter[str] = Counter()
    bucket_counts: Counter[str] = Counter()
    invoice_ref_counts: Counter[str] = Counter()

    try:
        ws = wb[sc.SHEET_TIME]
        headers = locate_headers(ws, sc.TABLE_COLUMNS[sc.TBL_TIME])
        used_source_ids: set[int] = set()

        for row_number in range(2, ws.max_row + 1):
            row = {header: ws.cell(row=row_number, column=col).value for header, col in headers.items()}
            if not clean(row.get(sc.COL_TIME_ENTRY_ID)) and not clean(row.get(sc.COL_TIME_DESC)):
                continue
            signature = time_signature(row)
            source_row = None
            while source_lookup.get(signature):
                candidate = source_lookup[signature].popleft()
                if id(candidate) not in used_source_ids:
                    source_row = candidate
                    used_source_ids.add(id(candidate))
                    break
            if source_row is None:
                unmatched_time_entries.append(
                    {
                        "targetRow": row_number,
                        "entryId": clean(row.get(sc.COL_TIME_ENTRY_ID)),
                        "date": date_key(row.get(sc.COL_TIME_DATE)),
                        "description": clean(row.get(sc.COL_TIME_DESC))[:120],
                    }
                )
                continue

            invoice_ref = clean(source_row.get("Invoice"))
            bucket = invoice_bucket(invoice_ref)
            state = payment_state_for_marker(invoice_ref, receivable_state)
            time_status = time_status_for_invoice_marker(invoice_ref)
            inv_status = invoice_status(invoice_ref)

            updates = {
                sc.COL_TIME_STATUS: time_status,
                sc.COL_TIME_INVOICE_REF: invoice_ref,
                sc.COL_TIME_INVOICE_STATUS: inv_status,
                sc.COL_TIME_PAYMENT_STATUS: state.get("paymentStatus", ""),
                sc.COL_TIME_INVOICE_TOTAL: state.get("invoiceTotal", 0.0),
                sc.COL_TIME_INVOICE_AMOUNT_PAID: state.get("invoiceAmountPaid", 0.0),
                sc.COL_TIME_INVOICE_BALANCE_DUE: state.get("invoiceBalanceDue", 0.0),
                sc.COL_TIME_INVOICE_DATE: state.get("invoiceDate", ""),
            }
            for header, value in updates.items():
                ws.cell(row=row_number, column=headers[header], value=value)

            matched += 1
            status_counts[time_status] += 1
            payment_counts[clean(state.get("paymentStatus"))] += 1
            bucket_counts[bucket] += 1
            if is_actual_invoice(invoice_ref):
                invoice_ref_counts[invoice_ref] += 1

        for queue in source_lookup.values():
            for row in queue:
                if id(row) not in used_source_ids:
                    unmatched_source_rows.append(int(row.get("_row") or 0))

        if args.apply:
            wb.save(TARGET)
    finally:
        wb.close()

    pending_receivables = [
        item for item in receivable_state.values()
        if item.get("paymentStatus") in {"Pending", "Partial"}
    ]
    summary = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "applied": bool(args.apply),
        "source": str(SOURCE),
        "target": str(TARGET),
        "sourceDocketRows": len(source_rows),
        "matchedTimeEntries": matched,
        "unmatchedTimeEntries": unmatched_time_entries,
        "unmatchedSourceRows": sorted(set(unmatched_source_rows)),
        "invoiceMarkerBuckets": dict(sorted(bucket_counts.items())),
        "timeStatusCounts": dict(sorted(status_counts.items())),
        "paymentStatusCounts": dict(sorted(payment_counts.items())),
        "uniqueActualInvoiceRefsOnDockets": len(invoice_ref_counts),
        "actualInvoiceDocketRows": sum(invoice_ref_counts.values()),
        "pendingOrPartialInvoiceCount": len(pending_receivables),
        "pendingOrPartialInvoiceBalance": round(
            sum(float(item.get("invoiceBalanceDue") or 0.0) for item in pending_receivables),
            2,
        ),
    }

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report_path = REPORT_DIR / f"invoice_payment_state_sync_{datetime.now():%Y%m%d_%H%M%S}.json"
    report_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))
    print(f"REPORT={report_path}")
    return 0 if not unmatched_time_entries else 2


if __name__ == "__main__":
    raise SystemExit(main())
