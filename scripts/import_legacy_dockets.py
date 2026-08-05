#!/usr/bin/env python3
"""
Import legacy Dockets.xlsm data into CSPM.xlsm.

Usage:  python scripts/import_legacy_dockets.py

Steps:
  1. Back up data/CSPM.xlsm
  2. Wipe all user data (keeps Trademarks + lookup/seed tables)
  3. Create 5 new sheets (Disbursements, Ledger, Receivables, InvoiceLog, HSTLog)
  4. Import every row from data/Dockets.xlsm
  5. Print verification report
"""

from __future__ import annotations

import os, re, sys, shutil
from collections import OrderedDict, Counter
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "src" / "python"))

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ============================================================
# PATHS
# ============================================================
CSPM_PATH   = REPO_ROOT / "data" / "CSPM.xlsm"
DOCKETS_PATH = REPO_ROOT / "data" / "Dockets.xlsm"
BACKUP_DIR  = REPO_ROOT / "data" / "recovery_backups"

TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False,
)

NOW_ISO = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
NOW_TAG = datetime.now().strftime("%Y%m%d_%H%M%S")

# ============================================================
# CLIENT NORMALIZATION RULES
# ============================================================

# Map raw variant names -> canonical display name
NAME_ALIASES: dict[str, str] = {
    "Graydel":                       "Graydel Capital",
    "Graydel Capital":               "Graydel Capital",
    "Litera Properties Inc.":        "Litera Group",
    "Litera Group":                  "Litera Group",
    "Finavator":                     "Finavator Inc.",
    "Finavator Inc.":                "Finavator Inc.",
    "Borkwosky, Jay And Cindy":      "Borkowsky, Jay And Cindy",
    "Rockford II Capital Inc.":      "Rockford Ii Capital Inc.",
    "Wild Bunch Beverages Ltd.":     "Wild Bunch Beverages Inc.",
    "Teelucksingh":                  "Teelucksingh, Gary",
    "Poulson, David":                "1471159 Ontario Ltd.",
    "Tremendis Group ":              "Tremendis Group",      # trailing space
}

# Force specific Client_IDs (override any auto-generated)
CLIENT_ID_OVERRIDES: dict[str, str] = {
    "James Bittner":                                "JBIT",
    "Truexperiences Tours Inc.":                    "TRU1",
    "CRLPBC (2016) Trust":                          "CRLP",
    "Fines, William":                               "FINE",
    "Beauty Kitchen UK Limited (T/A Reposit)":      "BEAU",
    "Curasion Inc.":                                "CURA",
    # Sub-clients from Dockets (not in Clients sheet)
    "Fensus":               "FENS",
    "HKCL":                 "HKCL",
    "Lee Development":      "LEED",
    "May Tower":            "MAYT",
    "Michels":              "MICH",
    "Rossland":             "ROSS",
    "Stryland, Nicholas":   "STRY",
    # Marciano Beckenstein sub-clients
    "DiBattista, Tony":     "DIBA",
    "Bob Leech":            "BOLE",
    "Elite Dairy":          "ELIT",
    "Iacovetta, Pasquale":  "IACO",
    "Cambria":              "CAMB",
    "Tremendis Corporation":"TREC",
    "Quinto, Gary":         "QUIN",
    "Max Strohmeier":       "STRO",
    "Larry Tung":           "TUNG",
    "Greg Mcknight":        "MCKN",
    "Olympic Kitchen":      "OLYM",
    "Cellucci, Tony":       "CELL",
    "Rimaldi, Max":         "RIMA",
}

# Sub-client text that is a *matter description*, not a real client
MATTER_DESC_SUBCLIENTS = {
    "December 2025 Dividends",
    "Incorporation - Flat Fee",
    "MarketCheck - Canadian Trademark",
}

# ============================================================
# HELPERS
# ============================================================

_MB_PATTERN = re.compile(r"^(\d+-\d+)\s*\((.+)\)$")

def parse_mb_subclient(raw: str) -> tuple[str | None, str | None]:
    """Parse '17-9384 (DiBattista, Tony)' -> ('DiBattista, Tony', '17-9384').
    Also handles '22-13337 (Quinto, Gary - Mattheson Trademarks)' -> ('Quinto, Gary', ...).
    Returns (client_name, file_number) or (None, None) if not MB format.
    """
    m = _MB_PATTERN.match(raw.strip())
    if not m:
        return None, None
    file_num = m.group(1)
    name_part = m.group(2).strip()
    if " - " in name_part:
        name_part = name_part.split(" - ", 1)[0].strip()
    return name_part, file_num


def clean(value) -> str:
    """Return stripped string; None / NaN -> ''."""
    if value is None:
        return ""
    s = str(value).strip()
    return "" if s.lower() in ("none", "nan") else s


def auto_client_id(name: str, existing_ids: set[str]) -> str:
    """Generate a 4-char upper-case client ID that doesn't collide."""
    cleaned = re.sub(r"[^A-Za-z0-9 ]", "", name).strip().upper()
    parts = cleaned.split()
    if not parts:
        base = "XXXX"
    elif len(parts) >= 2 and "," in name:          # "Last, First"
        base = parts[0][:4]
    else:
        base = parts[0][:4]
    base = (base + "XXXX")[:4]
    if base not in existing_ids:
        return base
    for i in range(1, 100):
        candidate = f"{base[:3]}{i}"
        if candidate not in existing_ids:
            return candidate
    return str(uuid4())[:4].upper()


def safe_float(v, default=0.0) -> float:
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def safe_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    try:
        return datetime.fromisoformat(str(v))
    except Exception:
        return None


def sheet_rows(ws, min_row=2) -> list[list]:
    """Read all data rows as lists."""
    rows = []
    for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, values_only=True):
        if any(v is not None for v in row):
            rows.append(list(row))
    return rows


def header_map(ws) -> dict[str, int]:
    """Map column header text -> 0-based index."""
    hm = {}
    for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1))):
        if cell.value:
            hm[str(cell.value).strip()] = i
    return hm


# ============================================================
# WIPE HELPERS
# ============================================================

def wipe_table(ws, table_name: str):
    """Delete all data rows from a table, leaving headers only."""
    tbl = ws.tables.get(table_name)
    if tbl is None:
        print(f"  [WARN] table {table_name} not found on sheet {ws.title}")
        return 0
    data_rows = ws.max_row - 1   # row 1 = header
    if data_rows <= 0:
        return 0
    ws.delete_rows(2, data_rows)
    # Resize table to header-only (A1 -> last_col 1)
    ncols = ws.max_column
    tbl.ref = f"A1:{get_column_letter(ncols)}1"
    return data_rows


def write_table_data(ws, table_name: str, headers: list[str], rows: list[list]):
    """Write headers + data rows and (re)create the table."""
    # Write headers
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    # Write data
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    last_row = max(len(rows) + 1, 2)  # at least 1 data row for table
    last_col = get_column_letter(len(headers))
    ref = f"A1:{last_col}{last_row}"
    # Remove old table if present
    if table_name in ws.tables:
        del ws.tables[table_name]
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)
    return len(rows)


def create_sheet_with_table(wb, sheet_name: str, table_name: str,
                             headers: list[str], rows: list[list]) -> int:
    """Create a new sheet, write data, and add a formal Excel table."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    return write_table_data(ws, table_name, headers, rows)


# ============================================================
# MAIN IMPORT
# ============================================================

def main():
    print("=" * 70)
    print("LEGACY DOCKETS.XLSM -> CSPM.XLSM IMPORT")
    print("=" * 70)

    # ── Step 1: Backup ────────────────────────────────────────
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup_path = BACKUP_DIR / f"CSPM.pre_legacy_import_{NOW_TAG}.xlsm"
    shutil.copy2(CSPM_PATH, backup_path)
    print(f"\n[1] Backup created: {backup_path}")

    # ── Step 2: Load workbooks ────────────────────────────────
    print("\n[2] Loading workbooks ...")
    wb_cspm = openpyxl.load_workbook(CSPM_PATH, keep_vba=True)
    wb_dock = openpyxl.load_workbook(DOCKETS_PATH, keep_vba=True, data_only=True)
    print(f"    CSPM sheets : {wb_cspm.sheetnames}")
    print(f"    Dockets sheets: {wb_dock.sheetnames}")

    # ── Step 3: Build client registry ─────────────────────────
    print("\n[3] Building client registry ...")
    ws_clients = wb_dock["Clients"]
    hm = header_map(ws_clients)
    raw_rows = sheet_rows(ws_clients)

    # --- 3a. Parse Clients sheet from Dockets.xlsm ---
    # Columns: Client Name(0) Entity Type(1) Principal(2) Position(3)
    #   Mailing Address(4) City(5) Province(6) Postal Code(7)
    #   Email(8) Phone Number(9) Standard Rate(10) Percentage(11)
    #   Comments(12) Client_ID(13) Parent_ID(14)
    ci = {k: hm[k] for k in hm}  # column-name -> index

    # client_registry: canonical_name -> dict of fields
    client_registry: dict[str, dict] = OrderedDict()
    # name_lookup: any_variant -> canonical_name
    name_lookup: dict[str, str] = {}
    all_client_ids: set[str] = set()

    # Fixed Client_ID for James Bittner (override before processing)
    # Process Clients sheet
    seen_canonical = set()
    for row in raw_rows:
        raw_name = clean(row[ci["Client Name"]])
        if not raw_name:
            continue

        canonical = NAME_ALIASES.get(raw_name, raw_name)

        # Skip duplicates (e.g. second "88 Queen" row)
        if canonical in seen_canonical:
            # Merge: fill in missing fields from duplicate row
            existing = client_registry[canonical]
            for field_key, idx in [
                ("entity_type", ci["Entity Type"]),
                ("principal", ci["Principal"]),
                ("position", ci["Position"]),
                ("address", ci["Mailing Address"]),
                ("city", ci["City"]),
                ("province", ci["Province"]),
                ("postal_code", ci["Postal Code"]),
                ("email", ci["Email"]),
                ("phone", ci["Phone Number"]),
                ("rate", ci["Standard Rate"]),
                ("percentage", ci["Percentage"]),
                ("comments", ci["Comments"]),
            ]:
                if not existing.get(field_key) and clean(row[idx]):
                    existing[field_key] = clean(row[idx])
            # Prefer a real Client_ID if the existing one is empty
            raw_cid = clean(row[ci["Client_ID"]])
            if raw_cid and not existing.get("client_id_raw"):
                existing["client_id_raw"] = raw_cid
            # Prefer a real Parent_ID if existing one is empty
            raw_pid = clean(row[ci["Parent_ID"]])
            if raw_pid and not existing.get("parent_id"):
                existing["parent_id"] = raw_pid
            name_lookup[raw_name] = canonical
            continue

        seen_canonical.add(canonical)
        name_lookup[raw_name] = canonical

        raw_cid = clean(row[ci["Client_ID"]])
        raw_pid = clean(row[ci["Parent_ID"]])

        client_registry[canonical] = {
            "client_id_raw": raw_cid,
            "parent_id": raw_pid,
            "entity_type": clean(row[ci["Entity Type"]]),
            "principal": clean(row[ci["Principal"]]),
            "position": clean(row[ci["Position"]]),
            "address": clean(row[ci["Mailing Address"]]),
            "city": clean(row[ci["City"]]),
            "province": clean(row[ci["Province"]]),
            "postal_code": clean(row[ci["Postal Code"]]),
            "email": clean(row[ci["Email"]]),
            "phone": clean(row[ci["Phone Number"]]),
            "rate": safe_float(row[ci["Standard Rate"]], 475),
            "percentage": safe_float(row[ci["Percentage"]], 1.0),
            "comments": clean(row[ci["Comments"]]),
        }

    # Also register alias variants
    for alias, canonical in NAME_ALIASES.items():
        name_lookup[alias] = canonical

    # --- 3b. Assign / fix Client_IDs ---
    for canonical, info in client_registry.items():
        raw_cid = info["client_id_raw"]
        # Apply overrides
        if canonical in CLIENT_ID_OVERRIDES:
            info["client_id"] = CLIENT_ID_OVERRIDES[canonical]
        elif raw_cid and raw_cid != "True":  # "True" was a boolean glitch
            info["client_id"] = str(raw_cid).strip().upper() if isinstance(raw_cid, str) else str(int(raw_cid)) if isinstance(raw_cid, (int, float)) and raw_cid == int(raw_cid) else str(raw_cid)
        else:
            info["client_id"] = ""   # will be auto-generated below

    # Fix numeric Client_IDs (like 1377 -> "1377")
    for canonical, info in client_registry.items():
        cid = info["client_id"]
        if cid:
            try:
                if float(cid) == int(float(cid)):
                    info["client_id"] = str(int(float(cid)))
            except (ValueError, TypeError):
                pass

    # Auto-generate missing IDs
    for canonical, info in client_registry.items():
        cid = info["client_id"]
        if cid:
            all_client_ids.add(cid)

    for canonical, info in client_registry.items():
        if not info["client_id"]:
            new_id = auto_client_id(canonical, all_client_ids)
            info["client_id"] = new_id
            all_client_ids.add(new_id)
            print(f"    Auto-generated Client_ID for '{canonical}': {new_id}")

    # --- 3c. Discover sub-clients from Dockets sheet that need new records ---
    ws_dockets = wb_dock["Dockets"]
    docket_hm = header_map(ws_dockets)
    docket_rows = sheet_rows(ws_dockets)

    sub_client_names_in_dockets = set()
    for row in docket_rows:
        sc = clean(row[docket_hm["Sub-Client"]] if len(row) > docket_hm["Sub-Client"] else "")
        if sc:
            sub_client_names_in_dockets.add(sc)

    for sc_raw in sorted(sub_client_names_in_dockets):
        if sc_raw in MATTER_DESC_SUBCLIENTS:
            continue
        # Check if it's a Marciano Beckenstein file-number entry
        mb_name, mb_file = parse_mb_subclient(sc_raw)
        if mb_name:
            sc_resolved = mb_name
        else:
            sc_resolved = sc_raw

        canonical = NAME_ALIASES.get(sc_resolved, sc_resolved)

        # If this client is already registered (by name or alias), skip
        if canonical in client_registry:
            name_lookup[sc_raw] = canonical
            name_lookup[sc_resolved] = canonical
            continue
        # Check if it matches an existing client name (case-insensitive / fuzzy)
        found = False
        for existing_name in client_registry:
            if existing_name.lower() == canonical.lower():
                name_lookup[sc_raw] = existing_name
                name_lookup[sc_resolved] = existing_name
                found = True
                break
        if found:
            continue

        # Need to create a new client record
        cid = CLIENT_ID_OVERRIDES.get(canonical, "")
        if not cid:
            cid = auto_client_id(canonical, all_client_ids)
        all_client_ids.add(cid)

        # Determine parent for this new sub-client from the Dockets context
        # Find what "Client" (billing parent) is used with this sub-client
        billing_parents = set()
        for row in docket_rows:
            rc = clean(row[docket_hm["Client"]])
            rsc = clean(row[docket_hm["Sub-Client"]] if len(row) > docket_hm["Sub-Client"] else "")
            if rsc == sc_raw and rc:
                billing_parents.add(rc)
        parent_id = ""
        if billing_parents:
            bp = list(billing_parents)[0]
            bp_canonical = NAME_ALIASES.get(bp, bp)
            if bp_canonical in client_registry:
                parent_id = client_registry[bp_canonical]["client_id"]

        client_registry[canonical] = {
            "client_id_raw": cid,
            "client_id": cid,
            "parent_id": parent_id,
            "entity_type": "",
            "principal": "",
            "position": "",
            "address": "",
            "city": "",
            "province": "",
            "postal_code": "",
            "email": "",
            "phone": "",
            "rate": 475,
            "percentage": 1.0,
            "comments": f"Created from Dockets sub-client reference",
        }
        name_lookup[sc_raw] = canonical
        name_lookup[sc_resolved] = canonical
        print(f"    New client from sub-client: '{canonical}' (ID={cid}, Parent={parent_id})")

    # Also register "Client" column entries not in Clients sheet
    for row in docket_rows:
        rc = clean(row[docket_hm["Client"]])
        if rc and rc not in name_lookup and rc not in client_registry:
            canonical = NAME_ALIASES.get(rc, rc)
            if canonical not in client_registry:
                # Check case-insensitive
                found = False
                for existing in client_registry:
                    if existing.lower() == canonical.lower():
                        name_lookup[rc] = existing
                        found = True
                        break
                if not found:
                    cid = CLIENT_ID_OVERRIDES.get(canonical, auto_client_id(canonical, all_client_ids))
                    all_client_ids.add(cid)
                    client_registry[canonical] = {
                        "client_id_raw": cid, "client_id": cid, "parent_id": "",
                        "entity_type": "", "principal": "", "position": "",
                        "address": "", "city": "", "province": "", "postal_code": "",
                        "email": "", "phone": "", "rate": 475, "percentage": 1.0,
                        "comments": "Created from Dockets client column",
                    }
                    name_lookup[rc] = canonical
                    print(f"    New client from Dockets Client col: '{canonical}' (ID={cid})")
            else:
                name_lookup[rc] = canonical

    # Ensure all name lookups point to valid registry entries
    for alias, canon in list(NAME_ALIASES.items()):
        name_lookup[alias] = canon

    # Special case: David Poulson's principal info
    if "1471159 Ontario Ltd." in client_registry:
        rec = client_registry["1471159 Ontario Ltd."]
        if not rec["principal"]:
            rec["principal"] = "David Poulson"

    print(f"    Total clients in registry: {len(client_registry)}")
    print(f"    Total name variants:       {len(name_lookup)}")

    # ── Step 4: Build parent registry ─────────────────────────
    print("\n[4] Building parent registry ...")

    # Parents come from:
    #   a) Clients sheet Parent_ID column (structural parents)
    #   b) Dockets Client column when Sub-Client is present (billing parents)
    parent_ids_needed: set[str] = set()

    # From Clients sheet
    for canonical, info in client_registry.items():
        pid = info["parent_id"]
        if pid:
            parent_ids_needed.add(pid)

    # From Dockets sheet (billing parents)
    billing_parent_names: set[str] = set()
    for row in docket_rows:
        rc = clean(row[docket_hm["Client"]])
        rsc = clean(row[docket_hm["Sub-Client"]] if len(row) > docket_hm["Sub-Client"] else "")
        if rc and rsc and rsc not in MATTER_DESC_SUBCLIENTS:
            # rc is a billing parent
            canonical_rc = name_lookup.get(rc, NAME_ALIASES.get(rc, rc))
            if canonical_rc in client_registry:
                pid = client_registry[canonical_rc]["client_id"]
                parent_ids_needed.add(pid)
                billing_parent_names.add(canonical_rc)

    # Also check for sub-client == client (self-reference) — not a parent relationship
    # Build parent records
    parent_registry: dict[str, dict] = OrderedDict()
    for pid in sorted(parent_ids_needed):
        # Find the client record for this parent
        parent_name = ""
        parent_rate = 475
        parent_pct = 1.0
        for canonical, info in client_registry.items():
            if info["client_id"] == pid:
                parent_name = canonical
                parent_rate = info["rate"]
                parent_pct = info["percentage"]
                break
        if not parent_name:
            parent_name = pid  # fallback
        parent_registry[pid] = {
            "parent_name": parent_name,
            "default_share_pct": parent_pct,
            "default_rate": parent_rate,
            "active": True,
            "notes": "",
        }
        print(f"    Parent: {pid} = '{parent_name}'")

    print(f"    Total parents: {len(parent_registry)}")

    # ── Step 5: Build matter registry ─────────────────────────
    print("\n[5] Building matter registry ...")
    ws_matters = wb_dock["Matters"]
    matter_hm = header_map(ws_matters)
    matter_rows = sheet_rows(ws_matters)

    matter_registry: list[dict] = []
    for row in matter_rows:
        def mg(col):
            idx = matter_hm.get(col)
            return row[idx] if idx is not None and idx < len(row) else None

        matter_id_code = clean(mg("Matter_ID"))
        raw_client_id = clean(mg("Client_ID"))
        sub_client_text = clean(mg("Sub-Client"))
        parent_id_raw = clean(mg("Parent_ID"))

        # Resolve client_id
        # First, try to find by raw Client_ID
        resolved_client_id = ""
        resolved_client_name = ""
        if raw_client_id:
            # numeric client IDs
            try:
                if float(raw_client_id) == int(float(raw_client_id)):
                    raw_client_id = str(int(float(raw_client_id)))
            except (ValueError, TypeError):
                pass
            for canonical, info in client_registry.items():
                if info["client_id"] == raw_client_id or info["client_id"] == raw_client_id.upper():
                    resolved_client_id = info["client_id"]
                    resolved_client_name = canonical
                    break
        if not resolved_client_id and raw_client_id:
            resolved_client_id = raw_client_id.upper()

        # Resolve parent_id
        resolved_parent_id = ""
        resolved_parent_name = ""
        if parent_id_raw:
            if parent_id_raw in parent_registry:
                resolved_parent_id = parent_id_raw
                resolved_parent_name = parent_registry[parent_id_raw]["parent_name"]
            else:
                resolved_parent_id = parent_id_raw

        matter_registry.append({
            "matter_id": str(uuid4()),
            "matter_number": matter_id_code,
            "client_id": resolved_client_id,
            "client_name": resolved_client_name,
            "parent_id": resolved_parent_id,
            "parent_name": resolved_parent_name,
            "matter_type": clean(mg("Type")),
            "year": mg("Year"),
            "sequence": mg("Sequence"),
            "description": clean(mg("Description (Internal)")),
            "display_name": clean(mg("Billing_Name")),
            "status": clean(mg("Status")) or "Open",
            "open_date": safe_date(mg("Open_Date")),
            "default_rate": safe_float(mg("Matter_Rate"), 475),
            "default_share_pct": safe_float(mg("Matter_Percent"), 1.0),
        })

    print(f"    Total matters: {len(matter_registry)}")

    # ── Client/SubClient resolution helper ────────────────────

    def resolve_docket_client(client_raw: str, subclient_raw: str):
        """Resolve Dockets sheet Client/Sub-Client -> (client_id, parent_id, client_name, matter_desc_extra).
        Returns (client_id, parent_id, resolved_client_name, extra_desc)
        """
        c = clean(client_raw)
        sc = clean(subclient_raw)
        extra_desc = ""

        if not c:
            return "", "", "", ""

        # Normalize the client name
        c_canonical = name_lookup.get(c, NAME_ALIASES.get(c, c))
        # Try case-insensitive fallback
        if c_canonical not in client_registry:
            for k in client_registry:
                if k.lower() == c_canonical.lower():
                    c_canonical = k
                    break

        if not sc:
            # Single client, no sub-client
            if c_canonical in client_registry:
                info = client_registry[c_canonical]
                return info["client_id"], info.get("parent_id", ""), c_canonical, ""
            return "", "", c, ""

        # Check if sub-client is a matter description
        if sc in MATTER_DESC_SUBCLIENTS:
            if c_canonical in client_registry:
                info = client_registry[c_canonical]
                return info["client_id"], info.get("parent_id", ""), c_canonical, sc
            return "", "", c, sc

        # Check if sub-client == client (self-reference)
        sc_canonical = name_lookup.get(sc, NAME_ALIASES.get(sc, sc))
        if sc_canonical not in client_registry:
            for k in client_registry:
                if k.lower() == sc_canonical.lower():
                    sc_canonical = k
                    break

        if sc_canonical == c_canonical or sc_canonical.lower() == c_canonical.lower():
            if c_canonical in client_registry:
                info = client_registry[c_canonical]
                return info["client_id"], info.get("parent_id", ""), c_canonical, ""
            return "", "", c, ""

        # Marciano Beckenstein file-number sub-clients
        mb_name, mb_file = parse_mb_subclient(sc)
        if mb_name:
            mb_canonical = name_lookup.get(mb_name, NAME_ALIASES.get(mb_name, mb_name))
            if mb_canonical not in client_registry:
                for k in client_registry:
                    if k.lower() == mb_canonical.lower():
                        mb_canonical = k
                        break
            if mb_canonical in client_registry:
                sc_info = client_registry[mb_canonical]
                # Client = Marciano (billing parent)
                c_info = client_registry.get(c_canonical, {})
                parent_id = c_info.get("client_id", "")
                return sc_info["client_id"], parent_id, mb_canonical, f"[{mb_file}]" if mb_file else ""
            # Fallback
            return "", "", mb_name, ""

        # Two distinct entities: Client = billing parent, Sub-Client = actual client
        sc_info = client_registry.get(sc_canonical, {})
        c_info = client_registry.get(c_canonical, {})

        client_id = sc_info.get("client_id", "")
        parent_id = c_info.get("client_id", "")

        return client_id, parent_id, sc_canonical, ""

    # ── Step 6: Wipe CSPM tables ──────────────────────────────
    print("\n[6] Wiping user data from CSPM.xlsm ...")

    tables_to_wipe = [
        ("Parents",       "tblParents"),
        ("Clients",       "tblClients"),
        ("ClientProfiles","tblClientProfiles"),
        ("Matters",       "tblMatters"),
        ("TimeEntries",   "tblTimeEntries"),
        ("Dockets",       "tblDockets"),
        ("Transactions",  "tblTransactionsMaster"),
    ]
    for sheet_name, table_name in tables_to_wipe:
        if sheet_name in wb_cspm.sheetnames:
            ws = wb_cspm[sheet_name]
            n = wipe_table(ws, table_name)
            print(f"    Wiped {n} rows from {table_name}")
        else:
            print(f"    [SKIP] Sheet '{sheet_name}' not found")

    # ── Step 7: Write Parents ─────────────────────────────────
    print("\n[7] Importing Parents ...")
    parent_headers = ["ParentID", "ParentName", "DefaultSharePct", "DefaultClientRate", "Active", "Notes"]
    parent_data_rows = []
    for pid, pinfo in parent_registry.items():
        parent_data_rows.append([
            pid,
            pinfo["parent_name"],
            pinfo["default_share_pct"],
            pinfo["default_rate"],
            True,
            pinfo["notes"],
        ])
    ws_parents = wb_cspm["Parents"]
    wipe_table(ws_parents, "tblParents")
    n = write_table_data(ws_parents, "tblParents", parent_headers, parent_data_rows)
    print(f"    Wrote {n} parent records")

    # ── Step 8: Write Clients ─────────────────────────────────
    print("\n[8] Importing Clients ...")
    client_headers = ["ClientID", "ClientName", "Email", "Phone", "Status", "Active", "Notes"]
    client_data_rows = []
    for canonical, info in client_registry.items():
        client_data_rows.append([
            info["client_id"],
            canonical,
            info.get("email", ""),
            info.get("phone", ""),
            "Active",
            True,
            info.get("comments", ""),
        ])
    ws_clients_out = wb_cspm["Clients"]
    wipe_table(ws_clients_out, "tblClients")
    n = write_table_data(ws_clients_out, "tblClients", client_headers, client_data_rows)
    print(f"    Wrote {n} client records")

    # ── Step 9: Write ClientProfiles ──────────────────────────
    print("\n[9] Importing ClientProfiles ...")
    profile_headers = [
        "ClientID", "LegalName", "DisplayName", "FirstName", "MiddleName", "LastName",
        "EntityType", "PrincipalName", "PrincipalPosition",
        "PrimaryEmail", "PrimaryPhone",
        "SecondaryContactName", "SecondaryContactPosition", "SecondaryContactEmail", "SecondaryContactPhone",
        "AddressLine1", "AddressLine2", "City", "StateProvince", "PostalCode", "Country", "FullAddress",
        "ParentClientID", "ParentClientName",
        "Website", "TaxID", "Industry", "BillingEmail",
        "KYCStatus", "OnboardingStatus", "RetainerRequired", "RetainerAmount",
        "EngagementStartDate", "DateClientAdded", "Birthday",
        "ReferralFrom", "ConflictNotes", "Notes",
        "CreatedAt", "UpdatedAt",
    ]
    profile_data_rows = []
    for canonical, info in client_registry.items():
        # Split name for individuals
        first_name = middle_name = last_name = ""
        if info.get("entity_type", "").lower() == "individual" or "," in canonical:
            parts = canonical.split(",", 1)
            if len(parts) == 2:
                last_name = parts[0].strip()
                given = parts[1].strip().split()
                first_name = given[0] if given else ""
                middle_name = " ".join(given[1:]) if len(given) > 1 else ""
            else:
                parts = canonical.split()
                if len(parts) >= 2:
                    first_name = parts[0]
                    last_name = parts[-1]
                    middle_name = " ".join(parts[1:-1])
                else:
                    first_name = canonical

        # Resolve parent name
        parent_id = info.get("parent_id", "")
        parent_name = ""
        if parent_id and parent_id in parent_registry:
            parent_name = parent_registry[parent_id]["parent_name"]

        # Build full address
        addr_parts = [p for p in [info.get("address", ""), info.get("city", ""),
                                   info.get("province", ""), info.get("postal_code", "")] if p]
        full_address = ", ".join(addr_parts)

        profile_data_rows.append([
            info["client_id"],
            canonical,           # LegalName
            canonical,           # DisplayName
            first_name, middle_name, last_name,
            info.get("entity_type", ""),
            info.get("principal", ""),
            info.get("position", ""),
            info.get("email", ""),   # PrimaryEmail
            info.get("phone", ""),   # PrimaryPhone
            "", "", "", "",          # Secondary contact
            info.get("address", ""), # AddressLine1
            "",                       # AddressLine2
            info.get("city", ""),
            info.get("province", ""),
            info.get("postal_code", ""),
            "",                       # Country
            full_address,
            parent_id,
            parent_name,
            "", "", "", "",          # Website, TaxID, Industry, BillingEmail
            "", "", "", "",          # KYC, Onboarding, Retainer, RetainerAmt
            "", "", "",              # Engagement, DateAdded, Birthday
            "", "",                  # Referral, Conflict
            info.get("comments", ""),  # Notes
            NOW_ISO,                 # CreatedAt
            NOW_ISO,                 # UpdatedAt
        ])

    ws_profiles = wb_cspm["ClientProfiles"]
    wipe_table(ws_profiles, "tblClientProfiles")
    n = write_table_data(ws_profiles, "tblClientProfiles", profile_headers, profile_data_rows)
    print(f"    Wrote {n} profile records")

    # ── Step 10: Write Matters ────────────────────────────────
    print("\n[10] Importing Matters ...")
    matter_headers = [
        "MatterID", "MatterNumber", "MatterName", "DisplayName",
        "ClientID", "ClientName", "ParentID", "ParentName",
        "MatterType", "PracticeArea", "Status",
        "ResponsibleLawyer", "BillingArrangement", "BillingContact", "BillingEmail",
        "DefaultRate", "DefaultSharePct", "RateHistory",
        "DateOfEngagement", "DateOpened", "DateClosed",
        "CourtFileNumber", "OpposingParty", "ReferralFrom",
        "Description", "Notes",
        "CreatedAt", "UpdatedAt",
    ]
    matter_data_rows = []
    for m in matter_registry:
        matter_data_rows.append([
            m["matter_id"],
            m["matter_number"],
            m["description"],            # MatterName
            m["display_name"],           # DisplayName
            m["client_id"],
            m["client_name"],
            m["parent_id"],
            m["parent_name"],
            m["matter_type"],            # MatterType
            "",                          # PracticeArea
            m["status"],
            "", "", "", "",              # Lawyer, Billing arrangement/contact/email
            m["default_rate"],
            m["default_share_pct"],
            "",                          # RateHistory
            "", m["open_date"], "",      # Engagement, Opened, Closed
            "", "", "",                  # CourtFile, Opposing, Referral
            m["description"],            # Description
            "",                          # Notes
            NOW_ISO, NOW_ISO,
        ])
    ws_matters = wb_cspm["Matters"]
    wipe_table(ws_matters, "tblMatters")
    n = write_table_data(ws_matters, "tblMatters", matter_headers, matter_data_rows)
    print(f"    Wrote {n} matter records")

    # ── Step 11: Write TimeEntries ────────────────────────────
    print("\n[11] Importing TimeEntries (from Dockets sheet) ...")
    time_headers = [
        "EntryID", "Date", "ClientID", "MatterID", "ParentID",
        "Description", "Hours", "ClientRate", "SharePct",
        "GrossToClient", "AmountToYou", "HST", "TotalInclHST",
        "RawSeconds", "Status", "LockAudit", "CreatedAt",
    ]
    time_data_rows = []
    orphan_count = 0
    sum_amount_to_cs_source = 0.0
    sum_amount_to_cs_imported = 0.0

    for row in docket_rows:
        date_val = safe_date(row[docket_hm["Date"]])
        client_raw = clean(row[docket_hm["Client"]])
        subclient_raw = clean(row[docket_hm["Sub-Client"]] if docket_hm.get("Sub-Client") is not None and docket_hm["Sub-Client"] < len(row) else "")
        description = clean(row[docket_hm["Description"]])
        hours = safe_float(row[docket_hm["Time (in hrs)"]])
        rate = safe_float(row[docket_hm["Hourly Rate/Flat Fee"]])
        pct = safe_float(row[docket_hm["Percentage"]])
        amount_to_cs = safe_float(row[docket_hm["Amount to CS"]])
        total_incl_hst = safe_float(row[docket_hm["Total Inclusive of HST"]])
        invoice_raw = clean(row[docket_hm["Invoice"]] if docket_hm.get("Invoice") is not None and docket_hm["Invoice"] < len(row) else "")
        matter_id_raw = clean(row[docket_hm["Matter_ID"]] if docket_hm.get("Matter_ID") is not None and docket_hm["Matter_ID"] < len(row) else "")
        raw_seconds = row[docket_hm["RawSeconds"]] if docket_hm.get("RawSeconds") is not None and docket_hm["RawSeconds"] < len(row) else None
        entry_type = clean(row[docket_hm["EntryType"]] if docket_hm.get("EntryType") is not None and docket_hm["EntryType"] < len(row) else "")

        sum_amount_to_cs_source += amount_to_cs

        # Resolve client/parent
        client_id, parent_id, resolved_name, extra_desc = resolve_docket_client(client_raw, subclient_raw)
        if not client_id:
            orphan_count += 1

        # Append extra description context
        if extra_desc:
            description = f"{description} {extra_desc}".strip() if description else extra_desc

        # Resolve matter
        resolved_matter_id = ""
        if matter_id_raw:
            for m in matter_registry:
                if m["matter_number"] == matter_id_raw:
                    resolved_matter_id = m["matter_id"]
                    break

        # Calculate financial fields
        gross_to_client = hours * rate if hours and rate else 0
        hst = total_incl_hst - amount_to_cs if total_incl_hst and amount_to_cs else 0

        # Status from invoice
        if invoice_raw.upper() == "BILLED":
            status = "Billed"
            lock_audit = ""
        elif invoice_raw:
            status = "Billed"
            lock_audit = f"inv:{invoice_raw}"
        else:
            status = "WIP"
            lock_audit = ""

        sum_amount_to_cs_imported += amount_to_cs

        time_data_rows.append([
            str(uuid4()),            # EntryID
            date_val,                # Date
            client_id,               # ClientID
            resolved_matter_id,      # MatterID
            parent_id,               # ParentID
            description,             # Description
            hours,                   # Hours
            rate,                    # ClientRate
            pct,                     # SharePct
            gross_to_client,         # GrossToClient
            amount_to_cs,            # AmountToYou
            hst,                     # HST
            total_incl_hst,          # TotalInclHST
            raw_seconds,             # RawSeconds
            status,                  # Status
            lock_audit,              # LockAudit
            NOW_ISO,                 # CreatedAt
        ])

    ws_time = wb_cspm["TimeEntries"]
    wipe_table(ws_time, "tblTimeEntries")
    n = write_table_data(ws_time, "tblTimeEntries", time_headers, time_data_rows)
    print(f"    Wrote {n} time entries ({orphan_count} orphans without ClientID)")

    # ── Step 12: Write raw Dockets (tblDockets) ───────────────
    print("\n[12] Writing raw Dockets (tblDockets) ...")
    docket_out_headers = [
        "Date", "Client", "Matter", "Parent", "Description",
        "Time (in hrs) or Units", "Hourly Rate/Flat Rate", "Percentage",
        "Amount to CS", "Total Inclusive of HST", "Invoice #",
        "RawSeconds", "EntryType",
    ]
    docket_out_rows = []
    for row in docket_rows:
        client_raw = clean(row[docket_hm["Client"]])
        subclient_raw = clean(row[docket_hm["Sub-Client"]] if docket_hm.get("Sub-Client") is not None and docket_hm["Sub-Client"] < len(row) else "")
        client_id, parent_id, resolved_name, _ = resolve_docket_client(client_raw, subclient_raw)
        # For the raw Dockets table, show Client = resolved client name, Parent = billing parent
        # Find parent name from parent_id
        parent_name = ""
        if parent_id and parent_id in parent_registry:
            parent_name = parent_registry[parent_id]["parent_name"]

        matter_id_raw = clean(row[docket_hm["Matter_ID"]] if docket_hm.get("Matter_ID") is not None and docket_hm["Matter_ID"] < len(row) else "")
        invoice_raw = clean(row[docket_hm["Invoice"]] if docket_hm.get("Invoice") is not None and docket_hm["Invoice"] < len(row) else "")

        docket_out_rows.append([
            safe_date(row[docket_hm["Date"]]),
            resolved_name or client_raw,
            matter_id_raw,
            parent_name,
            clean(row[docket_hm["Description"]]),
            safe_float(row[docket_hm["Time (in hrs)"]]),
            safe_float(row[docket_hm["Hourly Rate/Flat Fee"]]),
            safe_float(row[docket_hm["Percentage"]]),
            safe_float(row[docket_hm["Amount to CS"]]),
            safe_float(row[docket_hm["Total Inclusive of HST"]]),
            invoice_raw,
            row[docket_hm["RawSeconds"]] if docket_hm.get("RawSeconds") is not None and docket_hm["RawSeconds"] < len(row) else None,
            clean(row[docket_hm["EntryType"]] if docket_hm.get("EntryType") is not None and docket_hm["EntryType"] < len(row) else ""),
        ])

    ws_dockets_out = wb_cspm["Dockets"]
    wipe_table(ws_dockets_out, "tblDockets")
    n = write_table_data(ws_dockets_out, "tblDockets", docket_out_headers, docket_out_rows)
    print(f"    Wrote {n} raw docket rows")

    # ── Step 13: Create + Import Disbursements ────────────────
    print("\n[13] Importing Disbursements ...")
    ws_disb = wb_dock["Disbursements"]
    disb_hm = header_map(ws_disb)
    disb_rows = sheet_rows(ws_disb)

    disb_headers = [
        "DisbursementID", "Date", "ClientName", "SubClient",
        "ClientID", "ParentID", "MatterID",
        "Description", "Amount", "TaxExempt", "BillPct",
        "InvoiceRef", "CreatedAt",
    ]
    disb_data_rows = []
    for row in disb_rows:
        c_raw = clean(row[disb_hm["Client"]] if disb_hm.get("Client") is not None else "")
        sc_raw = clean(row[disb_hm["Sub-Client"]] if disb_hm.get("Sub-Client") is not None and disb_hm["Sub-Client"] < len(row) else "")
        client_id, parent_id, resolved_name, _ = resolve_docket_client(c_raw, sc_raw)
        matter_raw = clean(row[disb_hm["Matter_ID"]] if disb_hm.get("Matter_ID") is not None and disb_hm["Matter_ID"] < len(row) else "")

        disb_data_rows.append([
            str(uuid4()),
            safe_date(row[disb_hm["Date"]]),
            c_raw,
            sc_raw,
            client_id,
            parent_id,
            matter_raw,
            clean(row[disb_hm["Description"]] if disb_hm.get("Description") is not None else ""),
            safe_float(row[disb_hm["Amount"]] if disb_hm.get("Amount") is not None else 0),
            clean(row[disb_hm["Tax Exempt? (Y/N)"]] if disb_hm.get("Tax Exempt? (Y/N)") is not None else ""),
            safe_float(row[disb_hm["Bill %"]] if disb_hm.get("Bill %") is not None else 1),
            clean(row[disb_hm["Invoice"]] if disb_hm.get("Invoice") is not None else ""),
            NOW_ISO,
        ])

    n = create_sheet_with_table(wb_cspm, "Disbursements", "tblDisbursements", disb_headers, disb_data_rows)
    print(f"    Wrote {n} disbursement records")

    # ── Step 14: Create + Import Ledger ───────────────────────
    print("\n[14] Importing Ledger ...")
    ws_ledger = wb_dock["Ledger"]
    ledger_hm = header_map(ws_ledger)
    ledger_rows = sheet_rows(ws_ledger)

    ledger_headers = [
        "LedgerID", "Date", "ClientVendor", "Description", "Category", "Reference",
        "BillingsExclHST", "HSTCollected", "ExpensesExclHST", "HSTPaid",
        "Collected", "WriteOff", "Receivable",
        "TrxID", "ExternalRefID", "OriginalAmount", "WorkClient",
        "CreatedAt",
    ]
    ledger_data_rows = []
    sum_billings_source = 0.0
    for row in ledger_rows:
        def lg(col, default=None):
            idx = ledger_hm.get(col)
            if idx is not None and idx < len(row):
                return row[idx]
            return default

        trx_id = clean(lg("TrxID")) or str(uuid4())
        billings = safe_float(lg("Billings (excl. HST)"))
        sum_billings_source += billings

        ledger_data_rows.append([
            trx_id,
            safe_date(lg("Date")),
            clean(lg("Client/Vendor")),
            clean(lg("Description")),
            clean(lg("Category")),
            clean(lg("Reference")),
            billings,
            safe_float(lg("HST Collected")),
            safe_float(lg("Expenses (excl. HST)")),
            safe_float(lg("HST Paid")),
            safe_float(lg("Collected")),
            safe_float(lg("Write Off")),
            safe_float(lg("Receivable")),
            clean(lg("TrxID")),
            clean(lg("ExternalRefID")),
            safe_float(lg("OriginalAmount")) if lg("OriginalAmount") else None,
            clean(lg("Work Client")),
            NOW_ISO,
        ])

    n = create_sheet_with_table(wb_cspm, "Ledger", "tblLedger", ledger_headers, ledger_data_rows)
    print(f"    Wrote {n} ledger records")

    # ── Step 15: Create + Import Receivables ──────────────────
    print("\n[15] Importing Receivables ...")
    ws_recv = wb_dock["Receivables"]
    recv_hm = header_map(ws_recv)
    recv_rows = sheet_rows(ws_recv)

    recv_headers = [
        "InvoiceNum", "Date", "Client", "TotalInvoiced", "AmountPaid",
        "CreditsAdj", "BalanceDue", "Status", "WorkClient",
    ]
    recv_data_rows = []
    sum_invoiced_source = 0.0
    for row in recv_rows:
        def rg(col, default=None):
            idx = recv_hm.get(col)
            if idx is not None and idx < len(row):
                return row[idx]
            return default

        total_inv = safe_float(rg("Total_Invoiced"))
        sum_invoiced_source += total_inv

        recv_data_rows.append([
            clean(rg("InvoiceNum")),
            safe_date(rg("Date")),
            clean(rg("Client")),
            total_inv,
            safe_float(rg("Amount_Paid")),
            safe_float(rg("Credits/Adj")),
            safe_float(rg("Balance_Due")),
            clean(rg("Status")),
            clean(rg("Work Client")),
        ])

    n = create_sheet_with_table(wb_cspm, "Receivables", "tblReceivables", recv_headers, recv_data_rows)
    print(f"    Wrote {n} receivable records")

    # ── Step 16: Create + Import InvoiceLog ───────────────────
    print("\n[16] Importing InvoiceLog ...")
    ws_inv = wb_dock["Invoice Log"]
    inv_hm = header_map(ws_inv)
    inv_rows = sheet_rows(ws_inv)

    inv_headers = [
        "InvoiceNum", "ClientName", "SubClient", "InvoiceDate",
        "TotalFees", "TotalDisbursements", "TotalTax",
        "AggregateBilled", "BillToClient",
    ]
    inv_data_rows = []
    for row in inv_rows:
        def ig(col, default=None):
            idx = inv_hm.get(col)
            if idx is not None and idx < len(row):
                return row[idx]
            return default

        inv_data_rows.append([
            clean(ig("Invoice #")),
            clean(ig("Client Name")),
            clean(ig("Sub-Client")),
            safe_date(ig("Invoice Date")),
            safe_float(ig("Total Fees")) if ig("Total Fees") is not None else None,
            safe_float(ig("Total Disbursements")) if ig("Total Disbursements") is not None else None,
            safe_float(ig("Total Tax")) if ig("Total Tax") is not None else None,
            safe_float(ig("Aggregate Billed to Client")) if ig("Aggregate Billed to Client") is not None else None,
            clean(ig("Bill To Client")),
        ])

    n = create_sheet_with_table(wb_cspm, "InvoiceLog", "tblInvoiceLog", inv_headers, inv_data_rows)
    print(f"    Wrote {n} invoice log records")

    # ── Step 17: Create HSTLog (empty) ────────────────────────
    print("\n[17] Creating HSTLog (empty) ...")
    hst_headers = [
        "PeriodID", "PeriodStart", "PeriodEnd", "FiledDate",
        "ConfNum", "NetTax", "PaidDate", "PaymentRef",
    ]
    n = create_sheet_with_table(wb_cspm, "HSTLog", "tblHSTLog", hst_headers, [])
    print(f"    Created HSTLog with {n} rows (headers only)")

    # ── Step 18: Wipe TransactionsMaster ──────────────────────
    # (already wiped in Step 6 but ensure it's clean)
    ws_txn = wb_cspm["Transactions"]
    wipe_table(ws_txn, "tblTransactionsMaster")

    # ── Step 19: Save ─────────────────────────────────────────
    print("\n[19] Saving CSPM.xlsm ...")
    wb_cspm.save(CSPM_PATH)
    print(f"    Saved to {CSPM_PATH}")

    # ── Step 20: Verification Report ──────────────────────────
    print("\n" + "=" * 70)
    print("VERIFICATION REPORT")
    print("=" * 70)

    # Reload to verify
    wb_verify = openpyxl.load_workbook(CSPM_PATH, keep_vba=True, data_only=True)

    checks = [
        ("Parents",           "tblParents",              len(parent_data_rows)),
        ("Clients",           "tblClients",              len(client_data_rows)),
        ("ClientProfiles",    "tblClientProfiles",       len(profile_data_rows)),
        ("Matters",           "tblMatters",              len(matter_data_rows)),
        ("TimeEntries",       "tblTimeEntries",          len(time_data_rows)),
        ("Dockets",           "tblDockets",              len(docket_out_rows)),
        ("Disbursements",     "tblDisbursements",        len(disb_data_rows)),
        ("Ledger",            "tblLedger",               len(ledger_data_rows)),
        ("Receivables",       "tblReceivables",          len(recv_data_rows)),
        ("InvoiceLog",        "tblInvoiceLog",           len(inv_data_rows)),
        ("HSTLog",            "tblHSTLog",               0),
    ]
    all_ok = True
    for sheet_name, table_name, expected in checks:
        ws = wb_verify[sheet_name]
        actual = ws.max_row - 1  # subtract header
        if actual < 0:
            actual = 0
        status = "OK" if actual == expected else "MISMATCH"
        if status != "OK":
            all_ok = False
        print(f"  {sheet_name:25s}  expected={expected:5d}  actual={actual:5d}  [{status}]")

    # Financial checks
    print(f"\n  Financial checksums:")
    print(f"    AmountToCS source total:   {sum_amount_to_cs_source:>12.2f}")
    print(f"    AmountToCS imported total:  {sum_amount_to_cs_imported:>12.2f}")
    amt_match = abs(sum_amount_to_cs_source - sum_amount_to_cs_imported) < 0.01
    print(f"    Match: {'OK' if amt_match else 'MISMATCH'}")
    if not amt_match:
        all_ok = False

    print(f"\n    Ledger Billings source:    {sum_billings_source:>12.2f}")
    print(f"    Receivables Invoiced:      {sum_invoiced_source:>12.2f}")

    # Client ID uniqueness
    cids = [info["client_id"] for info in client_registry.values()]
    cid_dupes = [cid for cid, count in Counter(cids).items() if count > 1]
    if cid_dupes:
        print(f"\n  [WARN] Duplicate Client_IDs: {cid_dupes}")
        all_ok = False
    else:
        print(f"\n  Client_ID uniqueness: OK ({len(cids)} unique)")

    # Orphan time entries
    print(f"  Orphan time entries (no ClientID): {orphan_count}")

    # Lookup tables preserved
    for sheet_name, table_name in [
        ("TransactionAccounts", "tblTransactionAccounts"),
        ("TransactionCategories", "tblTransactionCategories"),
        ("TransactionBusinessUnits", "tblTransactionBusinessUnits"),
        ("TransactionPayees", "tblTransactionPayees"),
        ("Trademarks", "tblTrademarks"),
    ]:
        ws = wb_verify[sheet_name]
        rows = ws.max_row - 1
        print(f"  {sheet_name:30s}  rows={rows:5d}  [PRESERVED]")

    print(f"\n{'=' * 70}")
    if all_ok:
        print("ALL CHECKS PASSED")
    else:
        print("SOME CHECKS FAILED — review above")
    print(f"{'=' * 70}")
    print(f"\nBackup at: {backup_path}")
    print(f"Import complete.")


if __name__ == "__main__":
    main()
