#!/usr/bin/env python3
"""Build a safe full-import candidate from legacy Dockets.xlsm.

This script does not replace the active workbook. It creates an isolated
candidate under outputs/, uses the current repository importer for
integrity-safe client/matter/time rows, then restores the dedicated legacy
finance tables that report screens such as A/R Aging read directly.
"""

from __future__ import annotations

import argparse
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import json
from pathlib import Path
import re
import shutil
import sys
from typing import Any, Iterable


REPO_ROOT = Path(__file__).resolve().parents[1]
PYTHON_SRC = REPO_ROOT / "src" / "python"
if str(PYTHON_SRC) not in sys.path:
    sys.path.insert(0, str(PYTHON_SRC))

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from domain import schema_constants as sc
from repositories.excel_repo import ExcelRepo
from services.dockets_import_service import DocketsImportService
from services.paths import AppPaths
from services.workbook_integrity_service import WorkbookIntegrityService


RAW_DOCKETS_SHEET = "Dockets"
RAW_DOCKETS_TABLE = "tblDockets"
RAW_DOCKETS_HEADERS = [
    "Date",
    "Client",
    "Matter",
    "Parent",
    "Description",
    "Time (in hrs) or Units",
    "Hourly Rate/Flat Rate",
    "Percentage",
    "Amount to CS",
    "Total Inclusive of HST",
    "Invoice #",
    "RawSeconds",
    "EntryType",
]

RESET_TABLES = [
    (sc.SHEET_PARENTS, sc.TBL_PARENTS),
    (sc.SHEET_CLIENTS, sc.TBL_CLIENTS),
    (sc.SHEET_CLIENT_PROFILES, sc.TBL_CLIENT_PROFILES),
    (sc.SHEET_MATTERS, sc.TBL_MATTERS),
    (sc.SHEET_TIME, sc.TBL_TIME),
    (sc.SHEET_TRANSACTIONS, sc.TBL_TRANSACTIONS_MASTER),
    (sc.SHEET_DISBURSEMENTS, sc.TBL_DISBURSEMENTS),
    (sc.SHEET_LEDGER, sc.TBL_LEDGER),
    (sc.SHEET_RECEIVABLES, sc.TBL_RECEIVABLES),
    (sc.SHEET_INVOICE_LOG, sc.TBL_INVOICE_LOG),
    (sc.SHEET_HST_LOG, sc.TBL_HST_LOG),
]


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _utc_stamp() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "nan"}:
        return ""
    return re.sub(r"\s+", " ", text)


def _safe_float(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _money2(value: Any) -> Any:
    if value in (None, ""):
        return value
    try:
        normalized = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return value
    return float(normalized)


def _is_nonblank_row(row: dict[str, Any]) -> bool:
    return any(value not in (None, "") and _clean(value) for value in row.values())


def _table_style() -> TableStyleInfo:
    return TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )


def _write_table(
    wb: openpyxl.Workbook,
    sheet_name: str,
    table_name: str,
    headers: list[str],
    rows: list[dict[str, Any] | list[Any]],
    *,
    replace_sheet: bool = False,
) -> int:
    if replace_sheet and sheet_name in wb.sheetnames:
        del wb[sheet_name]

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for existing_table in list(ws.tables):
            del ws.tables[existing_table]
        max_row = max(int(ws.max_row or 1), len(rows) + 2)
        max_col = max(int(ws.max_column or 1), len(headers))
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                ws.cell(row_idx, col_idx).value = None
    else:
        ws = wb.create_sheet(sheet_name)

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    write_rows: Iterable[dict[str, Any] | list[Any]]
    write_rows = rows if rows else [{}]
    for row_idx, row_data in enumerate(write_rows, start=2):
        if isinstance(row_data, dict):
            values = [row_data.get(header, "") for header in headers]
        else:
            values = list(row_data)
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    last_row = max(2, len(rows) + 1)
    ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = _table_style()
    ws.add_table(table)
    return len(rows)


def _read_sheet_records(workbook: openpyxl.Workbook, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers: list[str] = []
    seen: dict[str, int] = {}
    for value in rows[0]:
        header = _clean(value)
        if not header:
            headers.append("")
            continue
        seen[header] = seen.get(header, 0) + 1
        headers.append(header if seen[header] == 1 else f"{header}_{seen[header]}")

    records: list[dict[str, Any]] = []
    for values in rows[1:]:
        record = {
            header: values[idx] if idx < len(values) else None
            for idx, header in enumerate(headers)
            if header
        }
        if _is_nonblank_row(record):
            records.append(record)
    return records


def _norm_name(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _clean(value).casefold())


def _load_name_indexes(workbook_path: Path) -> tuple[dict[str, str], dict[str, str]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, keep_vba=True)
    try:
        clients = _read_sheet_records(wb, sc.SHEET_CLIENTS)
        profiles = _read_sheet_records(wb, sc.SHEET_CLIENT_PROFILES)
        parents = _read_sheet_records(wb, sc.SHEET_PARENTS)
    finally:
        wb.close()

    client_ids: dict[str, str] = {}
    parent_ids: dict[str, str] = {}
    for row in clients:
        client_id = _clean(row.get(sc.COL_CLIENT_ID))
        client_name = _clean(row.get(sc.COL_CLIENT_NAME))
        if client_id and client_name:
            client_ids[_norm_name(client_name)] = client_id
    for row in profiles:
        client_id = _clean(row.get(sc.COL_PROFILE_CLIENT_ID))
        for key in (sc.COL_PROFILE_DISPLAY_NAME, sc.COL_PROFILE_LEGAL_NAME):
            name = _clean(row.get(key))
            if client_id and name:
                client_ids[_norm_name(name)] = client_id
    for row in parents:
        parent_id = _clean(row.get(sc.COL_PARENT_ID))
        parent_name = _clean(row.get(sc.COL_PARENT_NAME))
        if parent_id and parent_name:
            parent_ids[_norm_name(parent_name)] = parent_id
            client_ids.setdefault(_norm_name(parent_name), parent_id)
    return client_ids, parent_ids


def _resolve_client_parent_ids(
    client_name: Any,
    sub_client_name: Any,
    client_ids: dict[str, str],
    parent_ids: dict[str, str],
) -> tuple[str, str]:
    raw_client = _clean(client_name)
    raw_sub = _clean(sub_client_name)
    if raw_sub:
        client_id = client_ids.get(_norm_name(raw_sub), "")
        parent_id = parent_ids.get(_norm_name(raw_client), "") or client_ids.get(_norm_name(raw_client), "")
        return client_id, parent_id
    return client_ids.get(_norm_name(raw_client), ""), ""


def _reset_candidate_workbook(workbook_path: Path) -> None:
    wb = openpyxl.load_workbook(workbook_path, keep_vba=True)
    try:
        for sheet_name, table_name in RESET_TABLES:
            _write_table(
                wb,
                sheet_name,
                table_name,
                sc.TABLE_COLUMNS[table_name],
                [],
                replace_sheet=False,
            )
        _write_table(
            wb,
            RAW_DOCKETS_SHEET,
            RAW_DOCKETS_TABLE,
            RAW_DOCKETS_HEADERS,
            [],
            replace_sheet=False,
        )
        wb.save(workbook_path)
    finally:
        wb.close()


def _legacy_raw_rows(source_path: Path, candidate_path: Path) -> dict[str, int]:
    wb_source = openpyxl.load_workbook(source_path, data_only=True, keep_vba=True)
    wb_target = openpyxl.load_workbook(candidate_path, keep_vba=True)
    try:
        client_ids, parent_ids = _load_name_indexes(candidate_path)

        docket_rows = []
        for row in _read_sheet_records(wb_source, "Dockets"):
            raw_client = _clean(row.get("Client"))
            raw_sub = _clean(row.get("Sub-Client"))
            docket_rows.append(
                {
                    "Date": row.get("Date"),
                    "Client": raw_sub or raw_client,
                    "Matter": _clean(row.get("Matter_ID")),
                    "Parent": raw_client if raw_sub else "",
                    "Description": _clean(row.get("Description")),
                    "Time (in hrs) or Units": row.get("Time (in hrs)"),
                    "Hourly Rate/Flat Rate": row.get("Hourly Rate/Flat Fee"),
                    "Percentage": row.get("Percentage"),
                    "Amount to CS": _money2(row.get("Amount to CS")),
                    "Total Inclusive of HST": _money2(row.get("Total Inclusive of HST")),
                    "Invoice #": _clean(row.get("Invoice")),
                    "RawSeconds": row.get("RawSeconds"),
                    "EntryType": _clean(row.get("EntryType")),
                }
            )

        disbursement_rows = []
        for row in _read_sheet_records(wb_source, "Disbursements"):
            client_id, parent_id = _resolve_client_parent_ids(
                row.get("Client"),
                row.get("Sub-Client"),
                client_ids,
                parent_ids,
            )
            disbursement_rows.append(
                {
                    sc.COL_DISB_ID: "",
                    sc.COL_DISB_DATE: row.get("Date"),
                    sc.COL_DISB_CLIENT_NAME: _clean(row.get("Client")),
                    sc.COL_DISB_SUB_CLIENT: _clean(row.get("Sub-Client")),
                    sc.COL_DISB_CLIENT_ID: client_id,
                    sc.COL_DISB_PARENT_ID: parent_id,
                    sc.COL_DISB_MATTER_ID: _clean(row.get("Matter_ID")),
                    sc.COL_DISB_DESCRIPTION: _clean(row.get("Description")),
                    sc.COL_DISB_AMOUNT: row.get("Amount"),
                    sc.COL_DISB_TAX_EXEMPT: _clean(row.get("Tax Exempt? (Y/N)")),
                    sc.COL_DISB_BILL_PCT: row.get("Bill %"),
                    sc.COL_DISB_INVOICE_REF: _clean(row.get("Invoice")),
                    sc.COL_DISB_CREATED_AT: _utc_stamp(),
                }
            )

        ledger_rows = []
        for row in _read_sheet_records(wb_source, "Ledger"):
            ledger_rows.append(
                {
                    sc.COL_LEDGER_ID: _clean(row.get("TrxID")),
                    sc.COL_LEDGER_DATE: row.get("Date"),
                    sc.COL_LEDGER_CLIENT_VENDOR: _clean(row.get("Client/Vendor")),
                    sc.COL_LEDGER_DESCRIPTION: _clean(row.get("Description")),
                    sc.COL_LEDGER_CATEGORY: _clean(row.get("Category")),
                    sc.COL_LEDGER_REFERENCE: _clean(row.get("Reference")),
                    sc.COL_LEDGER_BILLINGS_EXCL_HST: row.get("Billings (excl. HST)"),
                    sc.COL_LEDGER_HST_COLLECTED: row.get("HST Collected"),
                    sc.COL_LEDGER_EXPENSES_EXCL_HST: row.get("Expenses (excl. HST)"),
                    sc.COL_LEDGER_HST_PAID: row.get("HST Paid"),
                    sc.COL_LEDGER_COLLECTED: row.get("Collected"),
                    sc.COL_LEDGER_WRITE_OFF: row.get("Write Off"),
                    sc.COL_LEDGER_RECEIVABLE: row.get("Receivable"),
                    sc.COL_LEDGER_TRX_ID: _clean(row.get("TrxID")),
                    sc.COL_LEDGER_EXTERNAL_REF_ID: _clean(row.get("ExternalRefID")),
                    sc.COL_LEDGER_ORIGINAL_AMOUNT: row.get("OriginalAmount"),
                    sc.COL_LEDGER_WORK_CLIENT: _clean(row.get("Work Client")),
                    sc.COL_LEDGER_CREATED_AT: _utc_stamp(),
                }
            )

        receivable_rows = []
        for row in _read_sheet_records(wb_source, "Receivables"):
            receivable_rows.append(
                {
                    sc.COL_RECV_INVOICE_NUM: _clean(row.get("InvoiceNum")),
                    sc.COL_RECV_DATE: row.get("Date"),
                    sc.COL_RECV_CLIENT: _clean(row.get("Client")),
                    sc.COL_RECV_TOTAL_INVOICED: row.get("Total_Invoiced"),
                    sc.COL_RECV_AMOUNT_PAID: row.get("Amount_Paid"),
                    sc.COL_RECV_CREDITS_ADJ: row.get("Credits/Adj"),
                    sc.COL_RECV_BALANCE_DUE: row.get("Balance_Due"),
                    sc.COL_RECV_STATUS: _clean(row.get("Status")),
                    sc.COL_RECV_WORK_CLIENT: _clean(row.get("Work Client")),
                }
            )

        invoice_rows = []
        for row in _read_sheet_records(wb_source, "Invoice Log"):
            invoice_rows.append(
                {
                    sc.COL_INV_INVOICE_NUM: _clean(row.get("Invoice #")),
                    sc.COL_INV_CLIENT_NAME: _clean(row.get("Client Name")),
                    sc.COL_INV_SUB_CLIENT: _clean(row.get("Sub-Client")),
                    sc.COL_INV_INVOICE_DATE: row.get("Invoice Date"),
                    sc.COL_INV_TOTAL_FEES: row.get("Total Fees"),
                    sc.COL_INV_TOTAL_DISBURSEMENTS: row.get("Total Disbursements"),
                    sc.COL_INV_TOTAL_TAX: row.get("Total Tax"),
                    sc.COL_INV_AGGREGATE_BILLED: row.get("Aggregate Billed to Client"),
                    sc.COL_INV_BILL_TO_CLIENT: _clean(row.get("Bill To Client")),
                }
            )

        hst_rows = []
        for row in _read_sheet_records(wb_source, "HST_Log"):
            hst_rows.append({header: row.get(header, "") for header in sc.TABLE_COLUMNS[sc.TBL_HST_LOG]})

        counts = {
            RAW_DOCKETS_TABLE: _write_table(
                wb_target,
                RAW_DOCKETS_SHEET,
                RAW_DOCKETS_TABLE,
                RAW_DOCKETS_HEADERS,
                docket_rows,
                replace_sheet=True,
            ),
            sc.TBL_DISBURSEMENTS: _write_table(
                wb_target,
                sc.SHEET_DISBURSEMENTS,
                sc.TBL_DISBURSEMENTS,
                sc.TABLE_COLUMNS[sc.TBL_DISBURSEMENTS],
                disbursement_rows,
                replace_sheet=True,
            ),
            sc.TBL_LEDGER: _write_table(
                wb_target,
                sc.SHEET_LEDGER,
                sc.TBL_LEDGER,
                sc.TABLE_COLUMNS[sc.TBL_LEDGER],
                ledger_rows,
                replace_sheet=True,
            ),
            sc.TBL_RECEIVABLES: _write_table(
                wb_target,
                sc.SHEET_RECEIVABLES,
                sc.TBL_RECEIVABLES,
                sc.TABLE_COLUMNS[sc.TBL_RECEIVABLES],
                receivable_rows,
                replace_sheet=True,
            ),
            sc.TBL_INVOICE_LOG: _write_table(
                wb_target,
                sc.SHEET_INVOICE_LOG,
                sc.TBL_INVOICE_LOG,
                sc.TABLE_COLUMNS[sc.TBL_INVOICE_LOG],
                invoice_rows,
                replace_sheet=True,
            ),
            sc.TBL_HST_LOG: _write_table(
                wb_target,
                sc.SHEET_HST_LOG,
                sc.TBL_HST_LOG,
                sc.TABLE_COLUMNS[sc.TBL_HST_LOG],
                hst_rows,
                replace_sheet=True,
            ),
        }
        wb_target.save(candidate_path)
        return counts
    finally:
        wb_source.close()
        wb_target.close()


def _summarize_import_result(result: dict[str, Any]) -> dict[str, Any]:
    return {
        "success": bool(result.get("success")),
        "clientsAdded": int(result.get("clientsAdded") or 0),
        "mattersAdded": int(result.get("mattersAdded") or 0),
        "docketsAdded": int(result.get("docketsAdded") or 0),
        "disbursementsAddedViaTransactions": int(result.get("disbursementsAdded") or 0),
        "ledgerAddedViaTransactions": int(result.get("ledgerAdded") or 0),
        "receivablesAddedViaTransactions": int(result.get("receivablesAdded") or 0),
        "invoiceLogAddedViaTransactions": int(result.get("invoiceLogAdded") or 0),
        "duplicatesFound": int(result.get("duplicatesFound") or 0),
        "duplicatesSkipped": int(result.get("duplicatesSkipped") or 0),
        "errorCount": len(result.get("errors") or []),
        "warningCount": len(result.get("warnings") or []),
        "warningSamples": list(result.get("warnings") or [])[:20],
        "errors": list(result.get("errors") or []),
        "performance": dict(result.get("performance") or {}),
    }


def _candidate_duplicate_response(payload: dict[str, Any]) -> dict[str, str]:
    kind = _clean((payload or {}).get("kind")).lower()
    if kind == "client":
        return {"action": "skip", "scope": "one"}
    return {"action": "add", "scope": "one"}


def build_candidate(
    *,
    project_root: Path,
    source_path: Path,
    baseline_path: Path,
    output_root: Path | None,
    as_of_date: str,
) -> dict[str, Any]:
    if not source_path.is_file():
        raise FileNotFoundError(f"Legacy source workbook not found: {source_path}")
    if not baseline_path.is_file():
        raise FileNotFoundError(f"Baseline CSPM workbook not found: {baseline_path}")

    candidate_root = output_root or project_root / "outputs" / f"full_dockets_import_candidate_{_timestamp()}"
    data_dir = candidate_root / "data"
    data_dir.mkdir(parents=True, exist_ok=False)
    candidate_workbook = data_dir / "CSPM.xlsm"
    source_copy = data_dir / "Dockets.xlsm"
    baseline_copy = data_dir / "CSPM.baseline_before_candidate.xlsm"
    shutil.copy2(baseline_path, candidate_workbook)
    shutil.copy2(baseline_path, baseline_copy)
    shutil.copy2(source_path, source_copy)

    _reset_candidate_workbook(candidate_workbook)

    paths = AppPaths(candidate_root)
    repo = ExcelRepo(paths)
    repo.ensure_schema()
    import_result = DocketsImportService(repo).import_legacy_workbook(
        str(source_copy),
        duplicate_callback=_candidate_duplicate_response,
    )
    if not import_result.get("success"):
        raise RuntimeError("Current repository import failed: " + "; ".join(import_result.get("errors") or []))

    direct_table_counts = _legacy_raw_rows(source_copy, candidate_workbook)

    repo = ExcelRepo(paths)
    ar_report = repo.ar_aging_report({"asOfDate": as_of_date})
    financial_dashboard = repo.financial_dashboard_report(int(as_of_date[:4]))
    integrity = WorkbookIntegrityService(paths).check(
        workbook_path=candidate_workbook,
        schema_path=project_root / "schema" / "workbook_schema.yml",
    )

    summary = {
        "ok": bool(integrity.ok and ar_report.get("ok")),
        "createdAtUtc": _utc_stamp(),
        "candidateRoot": str(candidate_root),
        "candidateWorkbook": str(candidate_workbook),
        "sourceCopy": str(source_copy),
        "baselineCopy": str(baseline_copy),
        "sourceWorkbook": str(source_path),
        "baselineWorkbook": str(baseline_path),
        "importResult": _summarize_import_result(import_result),
        "directTableCounts": direct_table_counts,
        "integrity": integrity.as_dict(),
        "arAging": {
            "ok": bool(ar_report.get("ok")),
            "asOfDate": ar_report.get("asOfDate"),
            "summary": dict(ar_report.get("summary") or {}),
            "cards": list(ar_report.get("cards") or []),
        },
        "financialDashboard": {
            "ok": bool(financial_dashboard.get("ok")),
            "year": financial_dashboard.get("year"),
            "summary": dict(financial_dashboard.get("summary") or {}),
            "quarters": list(financial_dashboard.get("quarters") or []),
            "arDetails": list(financial_dashboard.get("arDetails") or []),
            "topBillingClients": list(financial_dashboard.get("topBillingClients") or []),
        },
    }
    summary_path = candidate_root / "candidate_summary.json"
    summary_path.write_text(json.dumps(summary, indent=2, default=str) + "\n", encoding="utf-8")
    summary["summaryPath"] = str(summary_path)
    return summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build a full Dockets.xlsm import candidate workbook.")
    parser.add_argument("--project-root", default=str(REPO_ROOT))
    parser.add_argument("--source", default=str(REPO_ROOT / "data" / "Dockets.xlsm"))
    parser.add_argument("--baseline", default=str(REPO_ROOT / "data" / "CSPM.xlsm"))
    parser.add_argument("--output-root", default="")
    parser.add_argument("--as-of-date", default=date.today().isoformat())
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    project_root = Path(args.project_root).resolve()
    output_root = Path(args.output_root).resolve() if args.output_root else None
    summary = build_candidate(
        project_root=project_root,
        source_path=Path(args.source).resolve(),
        baseline_path=Path(args.baseline).resolve(),
        output_root=output_root,
        as_of_date=str(args.as_of_date or date.today().isoformat()),
    )
    import_result = summary["importResult"]
    ar_summary = summary["arAging"]["summary"]
    financial_summary = summary["financialDashboard"]["summary"]
    integrity_summary = summary["integrity"]["summary"]
    print(f"CANDIDATE={summary['candidateWorkbook']}")
    print(f"SUMMARY={summary['summaryPath']}")
    print(
        "IMPORT="
        f"clients:{import_result['clientsAdded']} "
        f"matters:{import_result['mattersAdded']} "
        f"time:{import_result['docketsAdded']} "
        f"warnings:{import_result['warningCount']}"
    )
    print(
        "DIRECT_TABLES="
        + " ".join(f"{key}:{value}" for key, value in summary["directTableCounts"].items())
    )
    print(
        "INTEGRITY="
        f"errors:{integrity_summary['errorCount']} "
        f"warnings:{integrity_summary['warningCount']} "
        f"rows:{integrity_summary['rowsChecked']}"
    )
    print(
        "AR_AGING="
        f"total:{ar_summary.get('totalAr', 0)} "
        f"invoices:{ar_summary.get('invoiceCount', 0)} "
        f"clients:{ar_summary.get('clientCount', 0)}"
    )
    print(
        "FINANCIAL_DASHBOARD="
        f"revenue:{financial_summary.get('revenueIncludingWip', 0)} "
        f"wip:{financial_summary.get('wipAmount', 0)} "
        f"ar:{financial_summary.get('dashboardAr', 0)} "
        f"collectibleAr:{financial_summary.get('collectibleAr', 0)}"
    )
    return 0 if summary["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
