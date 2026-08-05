import openpyxl
import os
import json

def analyze_workbook(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    report = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for table in sheet.tables.values():
            table_name = table.name
            ref = table.ref
            
            # Get columns
            min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(ref)
            columns = []
            for col in range(min_col, max_col + 1):
                cell_value = sheet.cell(row=min_row, column=col).value
                columns.append(str(cell_value) if cell_value else f"Col{col}")
            
            # Get rows
            row_count = max_row - min_row
            
            # Analyze IDs and Nulls
            id_col_index = -1
            if "ID" in columns:
                id_col_index = columns.index("ID")
            elif "id" in columns:
                id_col_index = columns.index("id")
                
            ids = []
            null_counts = {col: 0 for col in columns}
            
            for row in range(min_row + 1, max_row + 1):
                row_has_data = False
                for i, col in enumerate(columns):
                    val = sheet.cell(row=row, column=min_col + i).value
                    if val is None or str(val).strip() == "":
                        null_counts[col] += 1
                    else:
                        row_has_data = True
                        
                if row_has_data and id_col_index != -1:
                    id_val = sheet.cell(row=row, column=min_col + id_col_index).value
                    if id_val is not None and str(id_val).strip() != "":
                        ids.append(str(id_val).strip())
                        
            # Check duplicates
            unique_ids = set(ids)
            duplicate_ids = [id for id in unique_ids if ids.count(id) > 1]
            
            report[table_name] = {
                "sheet": sheet_name,
                "columns": columns,
                "row_count": row_count,
                "id_column": columns[id_col_index] if id_col_index != -1 else None,
                "missing_ids": row_count - len(ids) if id_col_index != -1 else None,
                "duplicate_ids": duplicate_ids,
                "null_counts": null_counts
            }
            
    return report

if __name__ == "__main__":
    filepath = r"C:\Projects\__CSPM\data\CSPM.xlsm"
    try:
        report = analyze_workbook(filepath)
        with open(r"C:\Projects\__CSPM\data_quality_report.json", "w") as f:
            json.dump(report, f, indent=2)
        print("Data quality report generated successfully.")
    except Exception as e:
        print(f"Error: {e}")
