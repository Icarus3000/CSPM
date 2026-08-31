from __future__ import annotations

import argparse
import hashlib
import json
import os
from pathlib import Path
import shutil
import sys
from tempfile import NamedTemporaryFile
from datetime import datetime, timezone
import zipfile

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_ROOT = Path(__file__).resolve().parents[2]
PYTHON_ROOT = PROJECT_ROOT / "src" / "python"
if str(PYTHON_ROOT) not in sys.path:
    sys.path.insert(0, str(PYTHON_ROOT))

from domain.workbook_contract import (  # noqa: E402
    CSPM_PRE_CONFLICT_TABLES,
    WorkbookTableContract,
    column_type,
)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def zip_inventory_hash(path: Path) -> str:
    with zipfile.ZipFile(path) as package:
        inventory = "|".join(
            f"{name}:{package.getinfo(name).file_size}"
            for name in sorted(package.namelist())
        )
    return hashlib.sha256(inventory.encode("utf-8")).hexdigest().lower()


def neutralize_props(workbook: Workbook) -> None:
    workbook.properties.creator = "System"
    workbook.properties.lastModifiedBy = "System"
    workbook.properties.title = ""
    workbook.properties.subject = ""
    workbook.properties.description = ""
    workbook.properties.keywords = ""
    workbook.properties.category = ""


def convert_to_xlsm(xlsx_path: Path, xlsm_path: Path) -> None:
    with NamedTemporaryFile(delete=False, suffix=".xlsm", dir=xlsm_path.parent) as temporary:
        temporary_path = Path(temporary.name)
    try:
        with zipfile.ZipFile(xlsx_path, "r") as source:
            with zipfile.ZipFile(temporary_path, "w") as target:
                for item in source.infolist():
                    content = source.read(item.filename)
                    if item.filename == "[Content_Types].xml":
                        content = content.replace(
                            b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                            b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                        )
                    target.writestr(item, content)
        os.replace(temporary_path, xlsm_path)
    finally:
        temporary_path.unlink(missing_ok=True)
        xlsx_path.unlink(missing_ok=True)


def create_workbook(path: Path, tables: tuple[WorkbookTableContract, ...]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    neutralize_props(workbook)
    for contract in tables:
        worksheet = workbook.create_sheet(title=contract.sheet)
        worksheet.append(list(contract.columns))
        worksheet.append([None] * len(contract.columns))
        table = Table(
            displayName=contract.table,
            ref=f"A1:{get_column_letter(len(contract.columns))}2",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    xlsx_path = path.with_suffix(".xlsx")
    workbook.save(xlsx_path)
    workbook.close()
    convert_to_xlsm(xlsx_path, path)


def schema_payload() -> dict:
    return {
        "schema_version": 2,
        "updated_utc": "2026-08-30T00:00:00Z",
        "status": "canonical-pre-conflict",
        "notes": [
            "This file is the canonical workbook contract for Excel-first mode.",
            "It is synchronized from runtime constants by the sanitized-template generator.",
            "Conflict tables are intentionally absent pending an approved Phase 2 candidate.",
        ],
        "workbook": {
            "path": "data/CSPM.xlsm",
            "mode": "tabular-database-only",
            "constraints": [
                "No data stored in formatting, color, or merged cells.",
                "Primary keys must be stable and portable to SQL.",
                "Date/time values should be parseable without locale-only assumptions.",
            ],
        },
        "tables": [
            {
                "sheet": contract.sheet,
                "table": contract.table,
                "primary_key": contract.primary_key,
                "identity_policy": contract.identity_policy,
                "columns": [
                    {"name": column, "type": column_type(column)}
                    for column in contract.columns
                ],
            }
            for contract in CSPM_PRE_CONFLICT_TABLES
        ],
        "migration_to_sql": {
            "strategy": "adapter-swap",
            "mapping_rule": "1:1 table and column names where possible",
            "preconditions": [
                "Excel service and repo code must use one unified naming contract.",
                "All business logic should operate on in-memory row objects, not cells.",
            ],
        },
    }


def specification_payload() -> dict:
    return {
        "filename": "CSPM.xlsm",
        "type": "macro-enabled-container-with-vba-prohibited",
        "worksheets": [contract.sheet for contract in CSPM_PRE_CONFLICT_TABLES],
        "tables": {
            contract.table: {
                "worksheet": contract.sheet,
                "columns": list(contract.columns),
            }
            for contract in CSPM_PRE_CONFLICT_TABLES
        },
    }


def inspect_sanitized_workbook(path: Path) -> dict:
    workbook = load_workbook(path, keep_vba=True, data_only=False)
    try:
        worksheets = [worksheet.title for worksheet in workbook.worksheets]
        tables: dict[str, dict] = {}
        non_empty_counts: dict[str, int] = {}
        confidential_rows = 0
        for worksheet in workbook.worksheets:
            non_empty_counts[worksheet.title] = sum(
                1
                for row in worksheet.iter_rows()
                for cell in row
                if cell.value not in (None, "")
            )
            for table in worksheet.tables.values():
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                headers = [
                    str(worksheet.cell(min_row, column).value or "").strip()
                    for column in range(min_col, max_col + 1)
                ]
                data_rows = 0
                for row_number in range(min_row + 1, max_row + 1):
                    if any(
                        worksheet.cell(row_number, column).value not in (None, "")
                        for column in range(min_col, max_col + 1)
                    ):
                        data_rows += 1
                confidential_rows += data_rows
                tables[table.name] = {
                    "worksheet": worksheet.title,
                    "columns": headers,
                    "dataRows": data_rows,
                }
    finally:
        for attribute in ("vba_archive", "_archive"):
            archive = getattr(workbook, attribute, None)
            if archive is not None:
                archive.close()
                setattr(workbook, attribute, None)
        workbook.close()

    with zipfile.ZipFile(path) as package:
        names = package.namelist()
        content_types = package.read("[Content_Types].xml")
        vba_present = "xl/vbaProject.bin" in names
        external_links = [
            name for name in names if name.startswith("xl/externalLinks/")
        ]
        comments = [name for name in names if "/comments" in name]
        connections = [name for name in names if name == "xl/connections.xml"]
        custom_xml = [name for name in names if name.startswith("customXml/")]
        embedded = [name for name in names if name.startswith("xl/embeddings/")]
        macro_enabled = (
            b"application/vnd.ms-excel.sheet.macroEnabled.main+xml"
            in content_types
        )
    if confidential_rows:
        raise RuntimeError("Sanitized CSPM template unexpectedly contains data rows.")
    if vba_present:
        raise RuntimeError("Sanitized CSPM template unexpectedly contains VBA.")
    if not macro_enabled:
        raise RuntimeError("Sanitized CSPM template lost its macro-enabled content type.")
    if external_links or connections or comments or embedded:
        raise RuntimeError("Sanitized CSPM template contains prohibited package content.")
    return {
        "worksheets": worksheets,
        "tables": tables,
        "nonEmptyCounts": non_empty_counts,
        "confidentialRows": confidential_rows,
        "vbaPresent": vba_present,
        "macroEnabled": macro_enabled,
        "externalLinks": external_links,
        "connections": connections,
        "comments": comments,
        "embedded": embedded,
        "customXml": custom_xml,
    }


def manifest_payload(
    template_path: Path,
    inspection: dict,
    generator_path: Path,
) -> dict:
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    return {
        "schema_version": "2.0",
        "template_name": "CSPM.xlsm",
        "purpose": "Sanitized pre-conflict clean-install workbook template",
        "creation_method": "openpyxl from synchronized runtime contract",
        "creator": "CSPM template governance",
        "creation_timestamp": now,
        "creation_timezone": "UTC",
        "generator_script_path": "scripts/template_governance/generate_sanitized_templates.py",
        "generator_script_SHA256": sha256_file(generator_path),
        "source_specifications": [
            "schema/workbook_schema.yml",
            "src/python/domain/schema_constants.py",
            "src/python/domain/ap_schema.py",
            "src/python/domain/workbook_contract.py",
            "src/templates/CSPM.template-spec.json",
        ],
        "approval_status": "CANDIDATE_AWAITING_CORY_APPROVAL",
        "approver": "",
        "approval_date": "",
        "approval_note": "Phase 1 synchronized candidate; human approval required.",
        "live_workbook_used": False,
        "client_data_present": False,
        "not_derived_from_live_workbook": True,
        "not_byte_identical_to_live_workbook": True,
        "template_sha256": sha256_file(template_path),
        "byte_count": template_path.stat().st_size,
        "vbaProject_sha256": "",
        "vba_policy": "prohibited",
        "zip_inventory_hash": zip_inventory_hash(template_path),
        "expected_worksheets": inspection["worksheets"],
        "worksheet_visibility": {
            worksheet: "visible" for worksheet in inspection["worksheets"]
        },
        "expected_tables": list(inspection["tables"]),
        "table_columns": {
            table: details["columns"]
            for table, details in inspection["tables"].items()
        },
        "expected_named_ranges": [],
        "worksheet_non_empty_cell_counts": inspection["nonEmptyCounts"],
        "external_link_count": len(inspection["externalLinks"]),
        "connection_count": len(inspection["connections"]),
        "comment_count": len(inspection["comments"]),
        "embedded_object_count": len(inspection["embedded"]),
        "custom_xml_inventory": inspection["customXml"],
        "vba_presence": inspection["vbaPresent"],
        "document_property_sanitization": "PASS",
        "confidentiality_scan_result": "PASS",
        "structural_validation_result": "PASS",
        "prohibited_external_links": True,
        "prohibited_external_connections": True,
        "expected_vba": "None - VBA is prohibited for clean installation templates.",
        "validation_method": "approve_phase9_candidates.py --review-only and build_release.py validate_templates",
        "reasoning_empty_collections": (
            "Empty link, connection, comment, object, custom XML, and data-row "
            "collections are required for this generated sanitized template."
        ),
    }


def update_approval(
    approval_path: Path,
    template_path: Path,
    inspection: dict,
) -> None:
    payload = json.loads(approval_path.read_text(encoding="utf-8"))
    item = next(
        row
        for row in payload["templates"]
        if row["approved_filename"] == "CSPM.xlsm"
    )
    item.update(
        {
            "human_approval_status": "CANDIDATE_AWAITING_CORY_APPROVAL",
            "sha256": sha256_file(template_path),
            "vba_policy": "prohibited",
            "vbaProject_sha256": "",
            "required_worksheets": inspection["worksheets"],
            "required_tables": {
                table: {
                    "worksheet": details["worksheet"],
                    "required_columns": details["columns"],
                }
                for table, details in inspection["tables"].items()
            },
        }
    )
    approval_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def generate_cspm(project_root: Path) -> dict:
    template_dir = project_root / "src" / "templates"
    template_dir.mkdir(parents=True, exist_ok=True)
    schema_path = project_root / "schema" / "workbook_schema.yml"
    spec_path = template_dir / "CSPM.template-spec.json"
    manifest_path = template_dir / "CSPM.template-manifest.json"
    approval_path = template_dir / "template_approval.json"
    template_path = template_dir / "CSPM.xlsm"

    schema_path.write_text(
        yaml.safe_dump(schema_payload(), sort_keys=False, allow_unicode=True),
        encoding="utf-8",
    )
    spec_path.write_text(
        json.dumps(specification_payload(), indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    create_workbook(template_path, CSPM_PRE_CONFLICT_TABLES)
    inspection = inspect_sanitized_workbook(template_path)
    manifest_path.write_text(
        json.dumps(
            manifest_payload(
                template_path,
                inspection,
                Path(__file__).resolve(),
            ),
            indent=2,
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )
    update_approval(approval_path, template_path, inspection)
    live_path = project_root / "data" / "CSPM.xlsm"
    if live_path.is_file() and sha256_file(live_path) == sha256_file(template_path):
        raise RuntimeError("Sanitized template unexpectedly matches the repository workbook.")
    return {
        "template": str(template_path),
        "sha256": sha256_file(template_path),
        "worksheets": len(inspection["worksheets"]),
        "tables": len(inspection["tables"]),
        "dataRows": inspection["confidentialRows"],
        "macroEnabled": inspection["macroEnabled"],
        "vbaPresent": inspection["vbaPresent"],
        "approvalStatus": "CANDIDATE_AWAITING_CORY_APPROVAL",
    }


def create_dockets_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    neutralize_props(workbook)
    for sheet_name in (
        "Clients",
        "Matters",
        "Dockets",
        "Disbursements",
        "Ledger",
        "Receivables",
        "Invoice Log",
    ):
        worksheet = workbook.create_sheet(title=sheet_name)
        worksheet.append([f"{sheet_name}Header1", f"{sheet_name}Header2"])
    xlsx_path = path.with_suffix(".xlsx")
    workbook.save(xlsx_path)
    workbook.close()
    convert_to_xlsm(xlsx_path, path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--project-root",
        type=Path,
        default=PROJECT_ROOT,
    )
    parser.add_argument(
        "--target",
        choices=("cspm", "dockets", "all"),
        default="all",
    )
    args = parser.parse_args()
    project_root = args.project_root.resolve()
    result: dict[str, object] = {}
    if args.target in {"cspm", "all"}:
        result["cspm"] = generate_cspm(project_root)
    if args.target in {"dockets", "all"}:
        dockets_path = project_root / "src" / "templates" / "Dockets.xlsm"
        create_dockets_workbook(dockets_path)
        result["dockets"] = {
            "template": str(dockets_path),
            "sha256": sha256_file(dockets_path),
        }
    print(json.dumps(result, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
