"""
FORENSIC: Dump actual cell values from source Dockets.xlsm
Check for formulas vs values, None cells, and compare files.
"""
import openpyxl
from pathlib import Path
from datetime import datetime

# ── 1. Check local file ───────────────────────────────────
local_path = Path(__file__).resolve().parent / "data" / "Dockets.xlsm"
print(f"LOCAL FILE: {local_path}")
print(f"  Size: {local_path.stat().st_size:,} bytes")
print(f"  Modified: {datetime.fromtimestamp(local_path.stat().st_mtime)}")

# ── 2. Check OneDrive production file ─────────────────────
onedrive_path = Path(r"C:\Users\cschn\OneDrive - LPN\__Invoices (1)\Dockets.xlsm")
if onedrive_path.exists():
    print(f"\nONEDRIVE FILE: {onedrive_path}")
    print(f"  Size: {onedrive_path.stat().st_size:,} bytes")
    print(f"  Modified: {datetime.fromtimestamp(onedrive_path.stat().st_mtime)}")
    if onedrive_path.stat().st_size != local_path.stat().st_size:
        print(f"  *** FILES ARE DIFFERENT SIZES! ***")
else:
    print(f"\nOneDrive file not found at {onedrive_path}")
    onedrive_path = None

# ── 3. Open local file WITH formulas (data_only=False) ────
print("\n" + "=" * 70)
print("LOCAL FILE: Dockets sheet — FORMULA mode (data_only=False)")
wb_formula = openpyxl.load_workbook(str(local_path), keep_vba=True, data_only=False)
ws_f = wb_formula["Dockets"]

# Get headers
headers = [str(c.value) if c.value else f"(col{i})" for i, c in enumerate(ws_f[1], 1)]
print(f"Headers: {headers}")

# Find "Amount to CS" column
amt_col = None
for i, h in enumerate(headers):
    if "Amount" in h and "CS" in h:
        amt_col = i + 1  # 1-based
        break
print(f"'Amount to CS' column index: {amt_col}")

# Check first 10 data rows — are they formulas or values?
print(f"\nFirst 10 rows — raw cell inspection:")
for row_idx in range(2, min(12, ws_f.max_row + 1)):
    date_cell = ws_f.cell(row=row_idx, column=1)
    amt_cell = ws_f.cell(row=row_idx, column=amt_col) if amt_col else None
    
    date_val = date_cell.value
    amt_val = amt_cell.value if amt_cell else None
    amt_type = type(amt_val).__name__
    
    # Check if it's a formula
    is_formula = isinstance(amt_val, str) and amt_val.startswith("=")
    
    print(f"  Row {row_idx:3d}: Date={str(date_val)[:10]:>10s}  "
          f"Amount={str(amt_val)[:40]:>40s}  "
          f"type={amt_type:>8s}  formula={is_formula}")

wb_formula.close()

# ── 4. Open local file WITH cached values (data_only=True) ─
print("\n" + "=" * 70)
print("LOCAL FILE: Dockets sheet — CACHED VALUES mode (data_only=True)")
wb_data = openpyxl.load_workbook(str(local_path), keep_vba=True, data_only=True)
ws_d = wb_data["Dockets"]

headers_d = [str(c.value) if c.value else f"(col{i})" for i, c in enumerate(ws_d[1], 1)]

# Sum Amount to CS with both modes
total_cached = 0.0
none_count = 0
zero_count = 0
has_value_count = 0
for row_idx in range(2, ws_d.max_row + 1):
    cell = ws_d.cell(row=row_idx, column=amt_col) if amt_col else None
    if cell:
        val = cell.value
        if val is None:
            none_count += 1
        elif val == 0:
            zero_count += 1
        else:
            try:
                total_cached += float(val)
                has_value_count += 1
            except:
                pass

print(f"  Rows with numeric value: {has_value_count}")
print(f"  Rows with None (formula not cached): {none_count}")
print(f"  Rows with zero: {zero_count}")
print(f"  Sum of cached 'Amount to CS': ${total_cached:,.2f}")

# Also check first 10 rows cached values
print(f"\nFirst 10 rows — cached values:")
for row_idx in range(2, min(12, ws_d.max_row + 1)):
    date_val = ws_d.cell(row=row_idx, column=1).value
    amt_val = ws_d.cell(row=row_idx, column=amt_col).value if amt_col else None
    
    print(f"  Row {row_idx:3d}: Date={str(date_val)[:10]:>10s}  "
          f"Amount={str(amt_val):>15s}  type={type(amt_val).__name__}")

wb_data.close()

# ── 5. If OneDrive file exists, compare ───────────────────
if onedrive_path and onedrive_path.exists():
    print("\n" + "=" * 70)
    print(f"ONEDRIVE FILE: Dockets sheet comparison")
    try:
        wb_od = openpyxl.load_workbook(str(onedrive_path), keep_vba=True, data_only=True, read_only=True)
        if "Dockets" in wb_od.sheetnames:
            ws_od = wb_od["Dockets"]
            od_headers = []
            od_rows = 0
            od_total = 0.0
            od_none = 0
            first = True
            for row in ws_od.iter_rows(values_only=True):
                if first:
                    od_headers = [str(h) if h else "" for h in row]
                    first = False
                    continue
                if not any(row):
                    continue
                od_rows += 1
                # Find Amount to CS in OneDrive headers
                for i, h in enumerate(od_headers):
                    if "Amount" in h and "CS" in h:
                        val = row[i] if i < len(row) else None
                        if val is None:
                            od_none += 1
                        else:
                            try:
                                od_total += float(val)
                            except:
                                pass
                        break
            
            print(f"  Headers: {od_headers[:15]}")
            print(f"  Total rows: {od_rows}")
            print(f"  None (formula) cells in Amount: {od_none}")
            print(f"  Sum of 'Amount to CS': ${od_total:,.2f}")
        wb_od.close()
    except Exception as e:
        print(f"  Error reading OneDrive file: {e}")

# ── 6. Also check ALL columns for any large totals ────────
print("\n" + "=" * 70)
print("COLUMN SUMS (local file, cached values):")
wb_data2 = openpyxl.load_workbook(str(local_path), keep_vba=True, data_only=True)
ws_d2 = wb_data2["Dockets"]
headers_d2 = [str(c.value) if c.value else f"(col{i})" for i, c in enumerate(ws_d2[1], 1)]

for col_idx in range(1, len(headers_d2) + 1):
    col_total = 0.0
    col_count = 0
    col_none = 0
    for row_idx in range(2, ws_d2.max_row + 1):
        val = ws_d2.cell(row=row_idx, column=col_idx).value
        if val is None:
            col_none += 1
        else:
            try:
                col_total += float(val)
                col_count += 1
            except:
                pass
    if col_count > 0 or col_none > 0:
        h = headers_d2[col_idx - 1]
        print(f"  {h:30s}  sum=${col_total:>12,.2f}  values={col_count:>4d}  none={col_none:>4d}")

wb_data2.close()
