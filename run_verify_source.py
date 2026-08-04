import openpyxl
wb = openpyxl.load_workbook('data/Dockets.xlsm', read_only=True, data_only=True)
ws = wb['Dockets']
headers = [str(c.value) if c.value else '' for c in ws[1]]
print('Headers:', headers[:15])
count = 0
total = 0.0
none_count = 0
amt_idx = None
for i, h in enumerate(headers):
    if 'Amount' in h and 'CS' in h:
        amt_idx = i
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    count += 1
    if amt_idx is not None:
        v = row[amt_idx]
        if v is None:
            none_count += 1
        else:
            try: total += float(v)
            except: pass
print(f"Rows: {count}")
print(f"Amount to CS sum: ${total:,.2f}")
print(f"None (uncached formula): {none_count}")

# Also compute from components
hrs_idx = None
rate_idx = None
pct_idx = None
for i, h in enumerate(headers):
    if 'Time' in h and ('hrs' in h or 'hour' in h.lower()):
        hrs_idx = i
    if 'Rate' in h or 'Fee' in h:
        rate_idx = i
    if 'Percentage' in h:
        pct_idx = i
print(f"\nComponent columns: hrs={hrs_idx}, rate={rate_idx}, pct={pct_idx}")

computed_total = 0.0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    hrs = float(row[hrs_idx] or 0) if hrs_idx is not None else 0
    rate = float(row[rate_idx] or 0) if rate_idx is not None else 0
    pct = float(row[pct_idx] or 0) if pct_idx is not None else 0
    if hrs > 0 and rate > 0 and pct > 0:
        computed_total += round(hrs * rate * (pct / 100.0), 2)
print(f"Computed from components (hrs*rate*pct/100): ${computed_total:,.2f}")
wb.close()
