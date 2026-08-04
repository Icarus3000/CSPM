"""
Refresh the Dockets sheet in CSPM.xlsm from the fresh source Dockets.xlsm.
Materializes formula values: if 'Amount to CS' is None, compute from components.
"""
import sys, os
from pathlib import Path
import openpyxl

src_path = Path(__file__).resolve().parent / "data" / "Dockets.xlsm"
cspm_path = Path(__file__).resolve().parent / "data" / "CSPM.xlsm"

print(f"Source: {src_path}")
print(f"Target: {cspm_path}")

# Read source data
src_wb = openpyxl.load_workbook(str(src_path), read_only=True, data_only=True)
src_ws = src_wb["Dockets"]
src_headers = [str(c.value) if c.value else "" for c in src_ws[1]]
print(f"Source headers: {src_headers[:15]}")

# Find key column indices
amt_idx = None
hrs_idx = None
rate_idx = None
pct_idx = None
hst_idx = None
for i, h in enumerate(src_headers):
    if "Amount" in h and "CS" in h:
        amt_idx = i
    if "Time" in h and ("hrs" in h or "hour" in h.lower()):
        hrs_idx = i
    if "Rate" in h or "Fee" in h:
        rate_idx = i
    if "Percentage" in h:
        pct_idx = i
    if "HST" in h or "Inclusive" in h:
        hst_idx = i

print(f"Column indices: amt={amt_idx}, hrs={hrs_idx}, rate={rate_idx}, pct={pct_idx}, hst={hst_idx}")

# Read all source rows
src_rows = []
for row in src_ws.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    src_rows.append(list(row))
print(f"Source rows: {len(src_rows)}")

src_wb.close()

# Materialized "Amount to CS" column headers for CSPM
# Map source -> CSPM headers
CSPM_HEADERS = [
    "Date", "Client", "Matter", "Parent", "Description",
    "Time (in hrs) or Units", "Hourly Rate/Flat Rate", "Percentage",
    "Amount to CS", "Total Inclusive of HST",
    "Invoice #", "RawSeconds", "EntryType"
]

# Source header -> CSPM header mapping
SRC_TO_CSPM = {
    "Date": "Date",
    "Client": "Client",
    "Sub-Client": "Matter",  # Sub-Client maps to Matter
    "Description": "Description",
    "Time (in hrs)": "Time (in hrs) or Units",
    "Hourly Rate/Flat Fee": "Hourly Rate/Flat Rate",
    "Percentage": "Percentage",
    "Amount to CS": "Amount to CS",
    "Total Inclusive of HST": "Total Inclusive of HST",
    "Invoice": "Invoice #",
    "Matter_ID": None,  # Skip - mapped differently
    "RawSeconds": "RawSeconds",
    "EntryType": "EntryType",
}

def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

# Build CSPM rows
cspm_rows = []
total_amt = 0.0
computed_count = 0
cached_count = 0
for src_row in src_rows:
    cspm_row = {}
    for src_i, src_h in enumerate(src_headers):
        if not src_h or src_i >= len(src_row):
            continue
        cspm_h = SRC_TO_CSPM.get(src_h)
        if cspm_h:
            cspm_row[cspm_h] = src_row[src_i]
    
    # Parent: use Client as Parent if no Sub-Client
    cspm_row.setdefault("Parent", cspm_row.get("Client", ""))
    
    # Materialize Amount to CS
    amt = safe_float(cspm_row.get("Amount to CS"))
    if amt == 0.0:
        hrs = safe_float(cspm_row.get("Time (in hrs) or Units"))
        rate = safe_float(cspm_row.get("Hourly Rate/Flat Rate"))
        pct = safe_float(cspm_row.get("Percentage"))
        if hrs > 0 and rate > 0 and pct > 0:
            amt = round(hrs * rate * (pct / 100.0), 2)
            computed_count += 1
    else:
        cached_count += 1
    cspm_row["Amount to CS"] = amt
    total_amt += amt
    
    # Materialize Total Inclusive of HST
    hst_val = safe_float(cspm_row.get("Total Inclusive of HST"))
    if hst_val == 0.0 and amt > 0:
        hst_val = round(amt * 1.13, 2)
    cspm_row["Total Inclusive of HST"] = hst_val
    
    cspm_rows.append(cspm_row)

print(f"\nMaterialized amounts: {cached_count} cached, {computed_count} computed")
print(f"Total Amount to CS: ${total_amt:,.2f}")

# Now write to CSPM.xlsm
tgt_wb = openpyxl.load_workbook(str(cspm_path), keep_vba=True)

# Remove old Dockets sheet if it exists
if "Dockets" in tgt_wb.sheetnames:
    del tgt_wb["Dockets"]
    print("Removed old Dockets sheet")

# Create new Dockets sheet
ws = tgt_wb.create_sheet("Dockets")

# Write headers
for col_idx, header in enumerate(CSPM_HEADERS, 1):
    ws.cell(row=1, column=col_idx, value=header)

# Write data rows
for row_idx, cspm_row in enumerate(cspm_rows, 2):
    for col_idx, header in enumerate(CSPM_HEADERS, 1):
        val = cspm_row.get(header)
        ws.cell(row=row_idx, column=col_idx, value=val)

print(f"Wrote {len(cspm_rows)} rows to Dockets sheet")

# Create Excel table
from openpyxl.worksheet.table import Table, TableStyleInfo
last_col_letter = openpyxl.utils.get_column_letter(len(CSPM_HEADERS))
table_ref = f"A1:{last_col_letter}{len(cspm_rows) + 1}"
table = Table(displayName="tblDockets", ref=table_ref)
style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True)
table.tableStyleInfo = style
ws.add_table(table)
print(f"Created table 'tblDockets': {table_ref}")

tgt_wb.save(str(cspm_path))
tgt_wb.close()
print(f"\nSaved {cspm_path}")
print("Done!")
