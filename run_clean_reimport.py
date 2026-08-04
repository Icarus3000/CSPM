"""
CLEAN RE-IMPORT SCRIPT
1. Refresh Dockets sheet in CSPM.xlsm from fresh OneDrive source (materialize formulas)
2. Clear ALL existing Transactions rows (they have wrong Class values)
3. Re-import Ledger with corrected Class mapping (Revenue/Operating Expense/Asset)
4. Run financial dashboard diagnostic to compare against benchmarks
"""
import sys, os
from pathlib import Path
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

root = Path(__file__).resolve().parent
sys.path.insert(0, str(root / "src" / "python"))

src_path = root / "data" / "Dockets.xlsm"
cspm_path = root / "data" / "CSPM.xlsm"

def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

# ══════════════════════════════════════════════════════════
# STEP 1: Refresh Dockets sheet with materialized formulas
# ══════════════════════════════════════════════════════════
print("=" * 60)
print("STEP 1: Refresh Dockets sheet from fresh source")
print("=" * 60)

src_wb = openpyxl.load_workbook(str(src_path), read_only=True, data_only=True)
src_ws = src_wb["Dockets"]
src_headers = [str(c.value) if c.value else "" for c in src_ws[1]]

SRC_TO_CSPM = {
    "Date": "Date", "Client": "Client", "Sub-Client": "Matter",
    "Description": "Description", "Time (in hrs)": "Time (in hrs) or Units",
    "Hourly Rate/Flat Fee": "Hourly Rate/Flat Rate", "Percentage": "Percentage",
    "Amount to CS": "Amount to CS", "Total Inclusive of HST": "Total Inclusive of HST",
    "Invoice": "Invoice #", "RawSeconds": "RawSeconds", "EntryType": "EntryType",
}
CSPM_HEADERS = [
    "Date", "Client", "Matter", "Parent", "Description",
    "Time (in hrs) or Units", "Hourly Rate/Flat Rate", "Percentage",
    "Amount to CS", "Total Inclusive of HST", "Invoice #", "RawSeconds", "EntryType"
]

cspm_rows = []
total_amt = 0.0
for row in src_ws.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    cspm_row = {}
    for src_i, src_h in enumerate(src_headers):
        if not src_h or src_i >= len(row):
            continue
        cspm_h = SRC_TO_CSPM.get(src_h)
        if cspm_h:
            cspm_row[cspm_h] = row[src_i]
    cspm_row.setdefault("Parent", cspm_row.get("Client", ""))
    
    # Materialize Amount to CS
    amt = safe_float(cspm_row.get("Amount to CS"))
    if amt == 0.0:
        hrs = safe_float(cspm_row.get("Time (in hrs) or Units"))
        rate = safe_float(cspm_row.get("Hourly Rate/Flat Rate"))
        pct = safe_float(cspm_row.get("Percentage"))
        if hrs > 0 and rate > 0 and pct > 0:
            amt = round(hrs * rate * (pct / 100.0), 2)
    cspm_row["Amount to CS"] = amt
    total_amt += amt
    
    hst_val = safe_float(cspm_row.get("Total Inclusive of HST"))
    if hst_val == 0.0 and amt > 0:
        hst_val = round(amt * 1.13, 2)
    cspm_row["Total Inclusive of HST"] = hst_val
    
    cspm_rows.append(cspm_row)

src_wb.close()
print(f"  Source rows: {len(cspm_rows)}")
print(f"  Total Amount to CS: ${total_amt:,.2f}")

# Write to CSPM.xlsm
tgt_wb = openpyxl.load_workbook(str(cspm_path), keep_vba=True)

if "Dockets" in tgt_wb.sheetnames:
    del tgt_wb["Dockets"]
ws = tgt_wb.create_sheet("Dockets")
for col_idx, header in enumerate(CSPM_HEADERS, 1):
    ws.cell(row=1, column=col_idx, value=header)
for row_idx, cspm_row in enumerate(cspm_rows, 2):
    for col_idx, header in enumerate(CSPM_HEADERS, 1):
        ws.cell(row=row_idx, column=col_idx, value=cspm_row.get(header))

last_col = openpyxl.utils.get_column_letter(len(CSPM_HEADERS))
table_ref = f"A1:{last_col}{len(cspm_rows) + 1}"
table = Table(displayName="tblDockets", ref=table_ref)
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(table)
print(f"  Created tblDockets: {table_ref}")

# ══════════════════════════════════════════════════════════
# STEP 2: Clear ALL existing Transactions rows
# ══════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("STEP 2: Clear existing Transactions rows")
print("=" * 60)

if "Transactions" in tgt_wb.sheetnames:
    txn_ws = tgt_wb["Transactions"]
    headers_row = [str(c.value) if c.value else "" for c in txn_ws[1]]
    old_count = txn_ws.max_row - 1
    print(f"  Existing transaction rows: {old_count}")
    
    # Delete data rows (keep headers)
    if old_count > 0:
        txn_ws.delete_rows(2, old_count)
        print(f"  Cleared {old_count} transaction rows")
    
    # Update table reference to header-only
    for tbl_name in list(txn_ws.tables):
        tbl = txn_ws.tables[tbl_name]
        # Set table to just header row + 1 empty row
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
        last_col_letter = openpyxl.utils.get_column_letter(max_col)
        tbl.ref = f"A1:{last_col_letter}2"
        print(f"  Reset table '{tbl_name}' to: {tbl.ref}")
else:
    print("  Transactions sheet not found!")

tgt_wb.save(str(cspm_path))
tgt_wb.close()
print("  Saved CSPM.xlsm")

# ══════════════════════════════════════════════════════════
# STEP 3: Re-import Ledger with corrected class mapping
# ══════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("STEP 3: Re-import Ledger")
print("=" * 60)

from services.paths import AppPaths
from repositories.excel_repo import ExcelRepo
from services.dockets_import_service import DocketsImportService

paths = AppPaths(root)
repo = ExcelRepo(paths)

def duplicate_callback(payload):
    kind = (payload.get("kind") or "").lower()
    if kind == "client":
        return {"action": "skip", "scope": "one"}
    return {"action": "add", "scope": "one"}

result = DocketsImportService(repo).import_legacy_workbook(
    str(src_path),
    duplicate_callback=duplicate_callback,
)

print(f"  Clients added: {result.get('clientsAdded', 0)}")
print(f"  Matters added: {result.get('mattersAdded', 0)}")
print(f"  Dockets added: {result.get('docketsAdded', 0)}")
print(f"  Ledger added:  {result.get('ledgerAdded', 0)}")
print(f"  Errors: {len(result.get('errors', []))}")

ledger_warns = [w for w in result.get("warnings", []) if "Ledger" in w]
print(f"  Ledger warnings: {len(ledger_warns)}")
for w in ledger_warns[:5]:
    print(f"    {w}")

# ══════════════════════════════════════════════════════════
# STEP 4: Run financial dashboard diagnostic
# ══════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("STEP 4: Financial Dashboard Benchmark Check")
print("=" * 60)

# Need fresh repo instance to pick up changes
repo2 = ExcelRepo(paths)
dashboard = repo2.financial_dashboard_report(2026)
summary = dashboard.get("summary", {})

rev_wip = summary.get("revenueIncludingWip", 0.0)
prod = summary.get("docketedAmount", 0.0)

target_rev_wip = 176098.21
target_prod = 174950.25

delta_rev = rev_wip - target_rev_wip
delta_prod = prod - target_prod

print(f"  {'Metric':25s} {'Python':>14s} {'Excel Target':>14s} {'Delta':>12s}")
print(f"  {'-'*25} {'-'*14} {'-'*14} {'-'*12}")
print(f"  {'Revenue + WIP':25s} ${rev_wip:>12,.2f}  ${target_rev_wip:>12,.2f}  ${delta_rev:>+10,.2f}")
print(f"  {'Adjusted Production':25s} ${prod:>12,.2f}  ${target_prod:>12,.2f}  ${delta_prod:>+10,.2f}")

print(f"\n  Additional details:")
print(f"    billedTimeAmount:  ${summary.get('billedTimeAmount', 0.0):>12,.2f}")
print(f"    wipAmount:         ${summary.get('wipAmount', 0.0):>12,.2f}")
print(f"    wipHours:          {summary.get('wipHours', 0.0):>12.1f}")
print(f"    expenses:          ${summary.get('expenses', 0.0):>12,.2f}")

if abs(delta_rev) < 500 and abs(delta_prod) < 500:
    print("\n  PASS -- Both metrics within $500 tolerance.")
else:
    print("\n  NEEDS INVESTIGATION -- Discrepancy exceeds $500 tolerance.")
