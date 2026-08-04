"""
CRITICAL: Check if Dockets.xlsm has hidden rows, active filters, or table range limits
that could cause the import to miss data.
"""
import openpyxl
from pathlib import Path

source_path = Path(__file__).resolve().parent / "data" / "Dockets.xlsm"
print(f"Source: {source_path}")

# Open WITHOUT read_only so we get full metadata
wb = openpyxl.load_workbook(str(source_path), keep_vba=True, data_only=True)

for sheet_name in ["Dockets", "Ledger"]:
    if sheet_name not in wb.sheetnames:
        print(f"\n*** Sheet '{sheet_name}' NOT FOUND ***")
        continue
    
    ws = wb[sheet_name]
    print(f"\n{'='*70}")
    print(f"Sheet: {sheet_name}")
    print(f"  Dimensions: {ws.dimensions}")
    print(f"  max_row: {ws.max_row}, max_column: {ws.max_column}")
    print(f"  min_row: {ws.min_row}, min_column: {ws.min_column}")
    
    # Check AutoFilter
    if ws.auto_filter and ws.auto_filter.ref:
        print(f"  *** AUTOFILTER ACTIVE: ref={ws.auto_filter.ref} ***")
        if ws.auto_filter.filterColumn:
            for fc in ws.auto_filter.filterColumn:
                print(f"      Filter on column {fc.colId}: {fc.filters}")
    else:
        print(f"  AutoFilter: None")
    
    # Check for hidden rows
    hidden_count = 0
    total_rows = 0
    hidden_with_data = 0
    for row_idx in range(2, ws.max_row + 1):
        total_rows += 1
        rd = ws.row_dimensions.get(row_idx)
        if rd and rd.hidden:
            hidden_count += 1
            # Check if this hidden row has data
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val is not None:
                hidden_with_data += 1
    
    print(f"  Total data rows: {total_rows}")
    print(f"  Hidden rows: {hidden_count}")
    print(f"  Hidden rows WITH data: {hidden_with_data}")
    
    # Check for hidden columns
    hidden_cols = []
    for col_idx in range(1, ws.max_column + 1):
        cd = ws.column_dimensions.get(openpyxl.utils.get_column_letter(col_idx))
        if cd and cd.hidden:
            header_val = ws.cell(row=1, column=col_idx).value
            hidden_cols.append(f"col{col_idx}({header_val})")
    if hidden_cols:
        print(f"  *** HIDDEN COLUMNS: {hidden_cols} ***")
    
    # Check named tables
    if hasattr(ws, 'tables') and ws.tables:
        for tbl_name in ws.tables:
            tbl = ws.tables[tbl_name]
            tbl_ref = tbl.ref if hasattr(tbl, 'ref') else str(tbl)
            print(f"  Table '{tbl_name}': ref={tbl_ref}")
            try:
                from openpyxl.utils import range_boundaries
                min_col, min_row, max_col, max_row = range_boundaries(tbl_ref)
                print(f"    Table range: rows {min_row}-{max_row}, cols {min_col}-{max_col}")
                print(f"    Sheet max_row: {ws.max_row}")
                if max_row < ws.max_row:
                    print(f"    *** TABLE RANGE SHORTER THAN SHEET! Missing {ws.max_row - max_row} rows ***")
            except Exception as e:
                print(f"    Could not parse table ref: {e}")
    
    # Count ALL rows (including hidden)
    print(f"\n  Row-by-row count (all rows, not just visible):")
    data_rows_all = 0
    data_rows_visible = 0
    for row_idx in range(2, ws.max_row + 1):
        first_cell = ws.cell(row=row_idx, column=1).value
        if first_cell is not None:
            data_rows_all += 1
            rd = ws.row_dimensions.get(row_idx)
            if not (rd and rd.hidden):
                data_rows_visible += 1
    
    print(f"    Data rows (ALL):     {data_rows_all}")
    print(f"    Data rows (VISIBLE): {data_rows_visible}")
    if data_rows_all != data_rows_visible:
        print(f"    *** {data_rows_all - data_rows_visible} ROWS HIDDEN WITH DATA! ***")

# Also check the OneDrive LPN source path
import os
onedrive_path = None
for candidate in [
    os.path.expanduser("~/OneDrive/Documents"),
    os.path.expanduser("~/OneDrive - Personal/Documents"),
    os.path.expanduser("~/OneDrive"),
]:
    p = Path(candidate)
    if p.exists():
        # Search for Dockets.xlsm
        for f in p.rglob("Dockets.xlsm"):
            print(f"\n*** Found OneDrive copy: {f} ***")
            print(f"    Size: {f.stat().st_size:,} bytes")
            onedrive_path = f
            break

local_size = source_path.stat().st_size
print(f"\nLocal data/Dockets.xlsm size: {local_size:,} bytes")
if onedrive_path:
    od_size = onedrive_path.stat().st_size
    print(f"OneDrive Dockets.xlsm size:   {od_size:,} bytes")
    if od_size != local_size:
        print(f"*** SIZE MISMATCH! OneDrive has {od_size - local_size:+,} bytes difference ***")
        print(f"*** The local copy may be STALE or DIFFERENT! ***")

wb.close()
