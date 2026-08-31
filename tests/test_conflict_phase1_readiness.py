from __future__ import annotations

import hashlib
import json
from pathlib import Path
import sys
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
import pytest
import yaml


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = PROJECT_ROOT / "src" / "python"
if str(SOURCE_ROOT) not in sys.path:
    sys.path.append(str(SOURCE_ROOT))

from domain import schema_constants as sc
from domain.workbook_contract import CSPM_PRE_CONFLICT_TABLES
from repositories.excel_repo import (
    ExcelRepo,
    TBL_TRANSACTION_ACCOUNTS,
    TBL_TRANSACTION_BUSINESS_UNITS,
    TBL_TRANSACTIONS_MASTER,
)
from repositories.finance_repo import FinanceRepo
from services.paths import AppPaths
from services.workbook_integrity_service import WorkbookIntegrityService


def _repo(tmp_path: Path) -> tuple[ExcelRepo, AppPaths]:
    paths = AppPaths(root=PROJECT_ROOT, override_data_dir=tmp_path)
    repo = ExcelRepo(paths)
    repo.ensure_schema()
    return repo, paths


def _account(code: str, *, active: int = 1, aliases: str = "") -> dict:
    return {
        sc.COL_TXN_ACCOUNT_CODE: code,
        sc.COL_TXN_ACCOUNT_NAME: f"{code}-NAME",
        sc.COL_TXN_ACCOUNT_KIND: "chequing",
        sc.COL_TXN_ACCOUNT_OWNER: "",
        sc.COL_TXN_ACCOUNT_ACTIVE: active,
        sc.COL_TXN_ACCOUNT_ALIASES: aliases,
    }


def _unit(name: str, *, active: int = 1) -> dict:
    return {
        sc.COL_TXN_BUSINESS_UNIT_NAME: name,
        sc.COL_TXN_BUSINESS_UNIT_OWNER: "",
        sc.COL_TXN_BUSINESS_UNIT_ACTIVE: active,
    }


def _transaction(
    identifier: str,
    *,
    account: str,
    unit: str = "UNIT-ACTIVE",
    status: str = "Pending",
    txn_class: str = "Business",
    txn_type: str = "Income",
    to_account: str = "",
) -> dict:
    return {
        sc.COL_TXN_ID: identifier,
        sc.COL_TXN_DATE: "2026-01-02",
        sc.COL_TXN_CLASS: txn_class,
        sc.COL_TXN_BUSINESS_UNIT: unit,
        sc.COL_TXN_TYPE: txn_type,
        sc.COL_TXN_FROM_ACCOUNT: account,
        sc.COL_TXN_TO_ACCOUNT: to_account,
        sc.COL_TXN_PAYEE: "SYNTHETIC-PAYEE",
        sc.COL_TXN_CATEGORY_CODE: "SYNTHETIC-CATEGORY",
        sc.COL_TXN_CATEGORY_NAME: "Synthetic category",
        sc.COL_TXN_MEMBER: "Joint",
        sc.COL_TXN_AMOUNT: 1.0,
        sc.COL_TXN_TAX_AMOUNT: 0.0,
        sc.COL_TXN_TAX_FLAG: "None",
        sc.COL_TXN_HST_EXEMPT: 1,
        sc.COL_TXN_GENERAL_OFFICE_EXPENSE: 0,
        sc.COL_TXN_SHADOW: 0,
        sc.COL_TXN_BILL_CLAIM_PCT: 0.0,
        sc.COL_TXN_TOTAL_CLAIM_AMOUNT: 0.0,
        sc.COL_TXN_STATUS: status,
        sc.COL_TXN_CURRENCY: "CAD",
        sc.COL_TXN_CLEARED_AT: "2026-01-02" if status in {"Cleared", "Reconciled"} else "",
        sc.COL_TXN_RECONCILED_AT: "2026-01-02" if status == "Reconciled" else "",
        sc.COL_TXN_VOID_REASON: "Synthetic void" if status == "Void" else "",
        sc.COL_TXN_CREATED_AT: "2026-01-02 10:00:00",
        sc.COL_TXN_UPDATED_AT: "2026-01-02 10:00:00",
    }


def _install_reference_fixture(
    repo: ExcelRepo,
    transactions: list[dict],
    *,
    accounts: list[dict] | None = None,
    units: list[dict] | None = None,
) -> None:
    repo._write_table_rows_bulk(
        {
            TBL_TRANSACTION_ACCOUNTS: accounts or [_account("ACCOUNT-ACTIVE")],
            TBL_TRANSACTION_BUSINESS_UNITS: units or [_unit("UNIT-ACTIVE")],
            TBL_TRANSACTIONS_MASTER: transactions,
        }
    )


def test_seven_reviewed_warning_shapes_remain_visible_and_sanitized(tmp_path: Path) -> None:
    repo, paths = _repo(tmp_path)
    rows = [
        _transaction(
            f"SYNTHETIC-HISTORICAL-{index}",
            account=f"SYNTHETIC-MISSING-ACCOUNT-{index}",
            unit="SYNTHETIC-MISSING-UNIT",
            status="Cleared",
        )
        for index in range(1, 4)
    ]
    rows.append(
        _transaction(
            "SYNTHETIC-CURRENT-AP",
            account="SYNTHETIC-MISSING-SYSTEM-ACCOUNT",
            status="Pending",
        )
    )
    _install_reference_fixture(repo, rows)

    report = WorkbookIntegrityService(paths).check()
    issues = [
        issue
        for issue in report.issues
        if issue.code in {"missing_account_reference", "missing_business_unit_reference"}
    ]

    assert len(issues) == 7
    assert sum(issue.severity == "warning" for issue in issues) == 6
    assert sum(issue.severity == "error" for issue in issues) == 1
    assert {issue.record_id for issue in issues} == {
        "SYNTHETIC-HISTORICAL-1",
        "SYNTHETIC-HISTORICAL-2",
        "SYNTHETIC-HISTORICAL-3",
        "SYNTHETIC-CURRENT-AP",
    }
    assert all(issue.value.startswith("REF-") for issue in issues)
    public_json = report.to_json()
    assert "SYNTHETIC-MISSING-ACCOUNT" not in public_json
    assert "SYNTHETIC-MISSING-UNIT" not in public_json
    assert "SYNTHETIC-MISSING-SYSTEM-ACCOUNT" not in public_json


def test_reference_policy_covers_active_retired_mapped_broken_and_blank_states(
    tmp_path: Path,
) -> None:
    repo, paths = _repo(tmp_path)
    accounts = [
        _account("ACCOUNT-ACTIVE", aliases="ACCOUNT-LEGACY-MAPPED"),
        _account("ACCOUNT-RETIRED", active=0),
    ]
    units = [_unit("UNIT-ACTIVE"), _unit("UNIT-RETIRED", active=0)]
    rows = [
        _transaction("TXN-ACTIVE", account="ACCOUNT-ACTIVE"),
        _transaction("TXN-RETIRED-HISTORY", account="ACCOUNT-RETIRED", status="Cleared"),
        _transaction("TXN-MAPPED-HISTORY", account="ACCOUNT-LEGACY-MAPPED", status="Cleared"),
        _transaction("TXN-BLANK-OPTIONAL", account="ACCOUNT-ACTIVE", to_account=""),
        _transaction("TXN-BROKEN-CURRENT", account="ACCOUNT-UNKNOWN"),
        _transaction("TXN-RETIRED-CURRENT", account="ACCOUNT-RETIRED"),
        _transaction("TXN-MAPPED-CURRENT", account="ACCOUNT-LEGACY-MAPPED"),
        _transaction("TXN-BROKEN-HISTORY", account="ACCOUNT-UNKNOWN-HISTORY", status="Cleared"),
        _transaction("TXN-UNIT-RETIRED-HISTORY", account="ACCOUNT-ACTIVE", unit="UNIT-RETIRED", status="Cleared"),
        _transaction("TXN-UNIT-RETIRED-CURRENT", account="ACCOUNT-ACTIVE", unit="UNIT-RETIRED"),
        _transaction("TXN-UNIT-MISSING", account="ACCOUNT-ACTIVE", unit=""),
        _transaction("TXN-ACCOUNT-MISSING", account=""),
    ]
    _install_reference_fixture(repo, rows, accounts=accounts, units=units)

    report = WorkbookIntegrityService(paths).check()
    by_record = {}
    for issue in report.issues:
        by_record.setdefault(issue.record_id, []).append(issue)

    assert "TXN-ACTIVE" not in by_record
    assert "TXN-RETIRED-HISTORY" not in by_record
    assert "TXN-MAPPED-HISTORY" not in by_record
    assert "TXN-BLANK-OPTIONAL" not in by_record
    assert by_record["TXN-BROKEN-CURRENT"][0].reference_state == "broken_current_reference"
    assert by_record["TXN-RETIRED-CURRENT"][0].code == "retired_account_reference"
    assert by_record["TXN-MAPPED-CURRENT"][0].code == "noncanonical_current_reference"
    assert by_record["TXN-BROKEN-HISTORY"][0].reference_state == "unmapped_historical_identifier"
    assert by_record["TXN-UNIT-RETIRED-CURRENT"][0].code == "retired_business_unit_reference"
    assert by_record["TXN-UNIT-MISSING"][0].reference_state == "blank_required"
    assert by_record["TXN-ACCOUNT-MISSING"][0].reference_state == "blank_required"


def test_shared_transaction_write_paths_reject_broken_or_retired_references(
    tmp_path: Path,
) -> None:
    repo, _paths = _repo(tmp_path)
    _install_reference_fixture(
        repo,
        [],
        accounts=[
            _account("ACCOUNT-ACTIVE", aliases="ACCOUNT-LEGACY-MAPPED"),
            _account("ACCOUNT-RETIRED", active=0),
        ],
        units=[_unit("UNIT-ACTIVE"), _unit("UNIT-RETIRED", active=0)],
    )
    payload = {
        "txnDate": "2026-01-02",
        "class": "Business",
        "businessUnit": "UNIT-ACTIVE",
        "type": "Income",
        "fromAccount": "ACCOUNT-LEGACY-MAPPED",
        "payee": "SYNTHETIC-PAYEE",
        "categoryCode": "SYNTHETIC-CATEGORY",
        "categoryName": "Synthetic category",
        "member": "Joint",
        "amount": 1.0,
        "taxAmount": 0.0,
        "taxFlag": "None",
        "hstExempt": 1,
        "status": "Pending",
        "currency": "CAD",
    }

    saved = repo.save_transaction(payload)
    assert saved["ok"] is True
    stored = next(row for row in repo.list_transactions() if row["transactionId"] == saved["transactionId"])
    assert stored["fromAccount"] == "ACCOUNT-ACTIVE"

    for field, invalid in (
        ("fromAccount", "ACCOUNT-UNKNOWN"),
        ("fromAccount", "ACCOUNT-RETIRED"),
        ("businessUnit", "UNIT-UNKNOWN"),
        ("businessUnit", "UNIT-RETIRED"),
    ):
        broken = dict(payload)
        broken[field] = invalid
        with pytest.raises(ValueError, match="configured|retired"):
            repo.save_transaction(broken)

    finance = FinanceRepo(repo)
    broken = dict(payload)
    broken["fromAccount"] = "ACCOUNT-UNKNOWN"
    with pytest.raises(ValueError, match="configured"):
        finance.save_transaction(broken)


def test_payment_write_paths_have_no_account_or_business_unit_fallbacks(tmp_path: Path) -> None:
    repo, _paths = _repo(tmp_path)
    with pytest.raises(ValueError, match="Select the active account"):
        repo.post_invoice_payment(
            {
                "invoice": "SYNTHETIC-INVOICE",
                "date": "2026-01-02",
                "mode": "Payment",
                "amount": 1.0,
                "method": "Synthetic method",
            }
        )

    source = (PROJECT_ROOT / "src/python/repositories/excel_repo.py").read_text(
        encoding="utf-8"
    )
    assert 'or method or "Operating Account"' not in source
    assert 'data.get("account") or method' not in source
    assert '"businessUnit": sc.SYSTEM_BUSINESS_UNIT_LEGAL_PRACTICE' in source


def test_schema_template_manifest_and_approval_share_one_pre_conflict_contract() -> None:
    schema = yaml.safe_load(
        (PROJECT_ROOT / "schema/workbook_schema.yml").read_text(encoding="utf-8")
    )
    spec = json.loads(
        (PROJECT_ROOT / "src/templates/CSPM.template-spec.json").read_text(encoding="utf-8")
    )
    manifest = json.loads(
        (PROJECT_ROOT / "src/templates/CSPM.template-manifest.json").read_text(encoding="utf-8")
    )
    approval = json.loads(
        (PROJECT_ROOT / "src/templates/template_approval.json").read_text(encoding="utf-8")
    )
    approval_item = next(
        item for item in approval["templates"] if item["approved_filename"] == "CSPM.xlsm"
    )
    contract = {
        item.table: (item.sheet, list(item.columns))
        for item in CSPM_PRE_CONFLICT_TABLES
    }
    schema_contract = {
        item["table"]: (item["sheet"], [column["name"] for column in item["columns"]])
        for item in schema["tables"]
    }
    spec_contract = {
        table: (definition["worksheet"], definition["columns"])
        for table, definition in spec["tables"].items()
    }
    approval_contract = {
        table: (definition["worksheet"], definition["required_columns"])
        for table, definition in approval_item["required_tables"].items()
    }

    assert schema_contract == contract
    assert spec_contract == contract
    assert manifest["expected_tables"] == list(contract)
    assert manifest["table_columns"] == {
        table: columns for table, (_sheet, columns) in contract.items()
    }
    assert approval_contract == contract
    assert approval_item["human_approval_status"] == "CANDIDATE_AWAITING_CORY_APPROVAL"
    assert manifest["approval_status"] == "CANDIDATE_AWAITING_CORY_APPROVAL"
    assert "tblMatterParties" in contract
    assert not any(table.startswith("tblConflict") for table in contract)


def test_sanitized_template_has_no_data_rows_and_preserves_xlsm_package_contract() -> None:
    path = PROJECT_ROOT / "src/templates/CSPM.xlsm"
    workbook = load_workbook(path, keep_vba=True, data_only=False)
    try:
        found_tables = {}
        for worksheet in workbook.worksheets:
            for table in worksheet.tables.values():
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                data_rows = [
                    row_number
                    for row_number in range(min_row + 1, max_row + 1)
                    if any(
                        worksheet.cell(row_number, column).value not in (None, "")
                        for column in range(min_col, max_col + 1)
                    )
                ]
                assert data_rows == []
                found_tables[table.name] = worksheet.title
        assert len(found_tables) == 23
    finally:
        for attribute in ("vba_archive", "_archive"):
            archive = getattr(workbook, attribute, None)
            if archive is not None:
                archive.close()
                setattr(workbook, attribute, None)
        workbook.close()

    with ZipFile(path) as package:
        names = set(package.namelist())
        assert "xl/vbaProject.bin" not in names
        assert not any(name.startswith("xl/externalLinks/") for name in names)
        assert b"application/vnd.ms-excel.sheet.macroEnabled.main+xml" in package.read(
            "[Content_Types].xml"
        )
    approval = json.loads(
        (PROJECT_ROOT / "src/templates/template_approval.json").read_text(encoding="utf-8")
    )
    approval_item = next(
        item for item in approval["templates"] if item["approved_filename"] == "CSPM.xlsm"
    )
    assert approval_item["sha256"] == hashlib.sha256(path.read_bytes()).hexdigest().upper()


def test_conflict_tables_and_a08_runtime_route_remain_inactive() -> None:
    schema_text = (PROJECT_ROOT / "schema/workbook_schema.yml").read_text(encoding="utf-8")
    pathways = (PROJECT_ROOT / "src/qml/standards/ModulePathways.js").read_text(
        encoding="utf-8"
    )
    assert "tblConflict" not in schema_text
    assert json.dumps("A08") not in pathways
    assert "ConflictInquiry" not in pathways
