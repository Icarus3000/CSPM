from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
import pytest

from domain import schema_constants as sc
from domain.ap_lifecycle import APValidationError
from domain.ap_schema import (
    AP_BILLS_HEADERS,
    AP_BILLS_SHEET,
    AP_BILLS_TABLE,
    AP_PAYMENTS_HEADERS,
    AP_PAYMENTS_SHEET,
    AP_PAYMENTS_TABLE,
)
from repositories.excel_repo import (
    ExcelRepo,
    TBL_LEDGER,
    TBL_RECEIVABLES,
    TBL_TIME,
    TBL_TRANSACTION_ACCOUNTS,
    TBL_TRANSACTION_BUSINESS_UNITS,
    TBL_TRANSACTION_CATEGORIES,
    TBL_TRANSACTIONS_MASTER,
)
from services.ap_setoff_service import APSetoffService, SETOFF_ACCOUNT, SETOFF_METHOD
from services.paths import AppPaths


def _write_table(repo: ExcelRepo, workbook, sheet_name: str, table_name: str, headers, rows) -> None:
    worksheet = workbook[sheet_name]
    repo._write_table(worksheet, worksheet.tables.get(table_name), table_name, list(headers), rows, None)


def _setup_repo(tmp_path: Path) -> ExcelRepo:
    paths = AppPaths(root=tmp_path, override_data_dir=tmp_path)
    repo = ExcelRepo(paths)
    repo.ensure_schema()

    path = paths.workbook_path()
    workbook = load_workbook(path, keep_vba=True, data_only=False)
    for sheet_name in (AP_BILLS_SHEET, AP_PAYMENTS_SHEET):
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)

    _write_table(
        repo,
        workbook,
        AP_BILLS_SHEET,
        AP_BILLS_TABLE,
        AP_BILLS_HEADERS,
        [{
            "APBillID": "AP-SET-1",
            "Vendor": "LIHDC Professional Corporation",
            "VendorInvoiceNumber": "Settlement-2026-07-01",
            "InvoiceDate": "2026-07-01",
            "DueDate": "2026-07-29",
            "Subtotal": 100.0,
            "TaxAmount": 0.0,
            "Total": 100.0,
            "AmountPaid": 0.0,
            "Balance": 100.0,
            "Status": "Unpaid",
            "Currency": "CAD",
            "ExpenseTransactionID": "TXN-AP-SET-1",
            "Notes": "Settlement fee",
        }],
    )
    _write_table(repo, workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS, [])
    _write_table(
        repo,
        workbook,
        TBL_RECEIVABLES.sheet,
        TBL_RECEIVABLES.table,
        sc.TABLE_COLUMNS[TBL_RECEIVABLES.table],
        [
            {
                sc.COL_RECV_INVOICE_NUM: "26-0042",
                sc.COL_RECV_DATE: "2026-04-30",
                sc.COL_RECV_CLIENT: "Entomo Farms",
                sc.COL_RECV_TOTAL_INVOICED: 60.0,
                sc.COL_RECV_AMOUNT_PAID: 0.0,
                sc.COL_RECV_CREDITS_ADJ: 0.0,
                sc.COL_RECV_BALANCE_DUE: 60.0,
                sc.COL_RECV_STATUS: "Open",
            },
            {
                sc.COL_RECV_INVOICE_NUM: "26-0054",
                sc.COL_RECV_DATE: "2026-05-31",
                sc.COL_RECV_CLIENT: "PLC Group",
                sc.COL_RECV_TOTAL_INVOICED: 70.0,
                sc.COL_RECV_AMOUNT_PAID: 0.0,
                sc.COL_RECV_CREDITS_ADJ: 0.0,
                sc.COL_RECV_BALANCE_DUE: 70.0,
                sc.COL_RECV_STATUS: "Open",
            },
        ],
    )
    _write_table(repo, workbook, TBL_LEDGER.sheet, TBL_LEDGER.table, sc.TABLE_COLUMNS[TBL_LEDGER.table], [])
    _write_table(repo, workbook, TBL_TIME.sheet, TBL_TIME.table, sc.TABLE_COLUMNS[TBL_TIME.table], [])
    _write_table(
        repo,
        workbook,
        TBL_TRANSACTIONS_MASTER.sheet,
        TBL_TRANSACTIONS_MASTER.table,
        sc.TABLE_COLUMNS[TBL_TRANSACTIONS_MASTER.table],
        [{
            sc.COL_TXN_ID: "TXN-AP-SET-1",
            sc.COL_TXN_DATE: "2026-07-01",
            sc.COL_TXN_CLASS: "Business",
            sc.COL_TXN_TYPE: "Expense",
            sc.COL_TXN_FROM_ACCOUNT: "AP_PAYABLE",
            sc.COL_TXN_PAYEE: "LIHDC Professional Corporation",
            sc.COL_TXN_AMOUNT: 100.0,
            sc.COL_TXN_STATUS: "Pending",
            sc.COL_TXN_CURRENCY: "CAD",
        }],
    )
    _write_table(
        repo,
        workbook,
        TBL_TRANSACTION_ACCOUNTS.sheet,
        TBL_TRANSACTION_ACCOUNTS.table,
        sc.TABLE_COLUMNS[TBL_TRANSACTION_ACCOUNTS.table],
        [{
            sc.COL_TXN_ACCOUNT_CODE: "AP_PAYABLE",
            sc.COL_TXN_ACCOUNT_NAME: "Accounts Payable",
            sc.COL_TXN_ACCOUNT_KIND: "clearing",
            sc.COL_TXN_ACCOUNT_OWNER: "Business",
            sc.COL_TXN_ACCOUNT_ACTIVE: 1,
        }],
    )
    _write_table(
        repo,
        workbook,
        TBL_TRANSACTION_BUSINESS_UNITS.sheet,
        TBL_TRANSACTION_BUSINESS_UNITS.table,
        sc.TABLE_COLUMNS[TBL_TRANSACTION_BUSINESS_UNITS.table],
        [{
            sc.COL_TXN_BUSINESS_UNIT_NAME: "Cory Business",
            sc.COL_TXN_BUSINESS_UNIT_OWNER: "Cory",
            sc.COL_TXN_BUSINESS_UNIT_ACTIVE: 1,
        }],
    )
    _write_table(
        repo,
        workbook,
        TBL_TRANSACTION_CATEGORIES.sheet,
        TBL_TRANSACTION_CATEGORIES.table,
        sc.TABLE_COLUMNS[TBL_TRANSACTION_CATEGORIES.table],
        [{
            sc.COL_TXN_CATEGORY_LKP_CODE: "EXP_LEGAL_FEES",
            sc.COL_TXN_CATEGORY_LKP_NAME: "Legal Fees Expense",
            sc.COL_TXN_CATEGORY_LKP_TYPE: "Expense",
            sc.COL_TXN_CATEGORY_LKP_CLASS_SCOPE: "Business",
            sc.COL_TXN_CATEGORY_LKP_TAX_FLAG_DEFAULT: "Business Deductible",
            sc.COL_TXN_CATEGORY_LKP_ACTIVE: 1,
        }],
    )
    repo._safe_save(workbook, path)
    return repo


def _rows(repo: ExcelRepo, tref) -> list[dict]:
    workbook = load_workbook(repo.paths.workbook_path(), keep_vba=True, data_only=False)
    try:
        return APSetoffService(repo)._read_finance(workbook, tref)
    finally:
        repo._close_workbook(workbook)


def test_setoff_posts_and_reverses_all_affected_records(tmp_path: Path) -> None:
    repo = _setup_repo(tmp_path)
    service = APSetoffService(repo)

    result = service.record({
        "APPaymentID": "APP-SET-1",
        "APBillID": "AP-SET-1",
        "PaymentDate": "2026-07-29",
        "Amount": 100.0,
        "Reference": "Settlement agreement",
        "Allocations": [
            {"InvoiceID": "26-0042", "Amount": 60.0},
            {"InvoiceID": "26-0054", "Amount": 40.0},
        ],
    })

    assert result["ok"] is True
    assert result["allocationCount"] == 2

    from repositories.ap_workbook_repository import APWorkbookRepository

    ap = APWorkbookRepository(repo.paths.workbook_path())
    bill = ap.get_bill("AP-SET-1")
    assert bill["Status"] == "Paid"
    assert bill["Balance"] == 0
    payment = ap.list_payments("AP-SET-1")[0]
    assert payment["Method"] == SETOFF_METHOD
    assert payment["FromAccount"] == SETOFF_ACCOUNT

    receivables = {row[sc.COL_RECV_INVOICE_NUM]: row for row in _rows(repo, TBL_RECEIVABLES)}
    assert receivables["26-0042"][sc.COL_RECV_STATUS] == "Paid"
    assert receivables["26-0042"][sc.COL_RECV_BALANCE_DUE] == 0
    assert receivables["26-0054"][sc.COL_RECV_STATUS] == "Partial"
    assert receivables["26-0054"][sc.COL_RECV_AMOUNT_PAID] == 40
    assert receivables["26-0054"][sc.COL_RECV_BALANCE_DUE] == 30
    assert len(_rows(repo, TBL_LEDGER)) == 2

    transactions = _rows(repo, TBL_TRANSACTIONS_MASTER)
    transaction = next(row for row in transactions if row[sc.COL_TXN_ID] == "TXN-AP-SET-1")
    assert transaction[sc.COL_TXN_FROM_ACCOUNT] == SETOFF_ACCOUNT
    assert transaction[sc.COL_TXN_STATUS] == "Cleared"

    reversal = service.reverse("APP-SET-1", "APP-REV-SET-1", "Entered in error")
    assert reversal["ok"] is True

    bill = ap.get_bill("AP-SET-1")
    assert bill["Status"] == "Unpaid"
    assert bill["Balance"] == 100
    payments = {row["APPaymentID"]: row for row in ap.list_payments("AP-SET-1")}
    assert payments["APP-SET-1"]["Status"] == "Reversed"
    assert payments["APP-REV-SET-1"]["Status"] == "Reversal"

    receivables = {row[sc.COL_RECV_INVOICE_NUM]: row for row in _rows(repo, TBL_RECEIVABLES)}
    assert receivables["26-0042"][sc.COL_RECV_AMOUNT_PAID] == 0
    assert receivables["26-0042"][sc.COL_RECV_BALANCE_DUE] == 60
    assert receivables["26-0054"][sc.COL_RECV_AMOUNT_PAID] == 0
    assert receivables["26-0054"][sc.COL_RECV_BALANCE_DUE] == 70
    assert len(_rows(repo, TBL_LEDGER)) == 4

    transaction = next(row for row in _rows(repo, TBL_TRANSACTIONS_MASTER) if row[sc.COL_TXN_ID] == "TXN-AP-SET-1")
    assert transaction[sc.COL_TXN_FROM_ACCOUNT] == "AP_PAYABLE"
    assert transaction[sc.COL_TXN_STATUS] == "Pending"


def test_setoff_validation_leaves_every_ledger_unchanged(tmp_path: Path) -> None:
    repo = _setup_repo(tmp_path)
    service = APSetoffService(repo)

    with pytest.raises(APValidationError, match="exceeds the open balance"):
        service.record({
            "APPaymentID": "APP-INVALID",
            "APBillID": "AP-SET-1",
            "PaymentDate": "2026-07-29",
            "Amount": 100.0,
            "Reference": "Settlement agreement",
            "Allocations": [
                {"InvoiceID": "26-0042", "Amount": 60.01},
                {"InvoiceID": "26-0054", "Amount": 39.99},
            ],
        })

    from repositories.ap_workbook_repository import APWorkbookRepository

    assert APWorkbookRepository(repo.paths.workbook_path()).list_payments("AP-SET-1") == []
    receivables = {row[sc.COL_RECV_INVOICE_NUM]: row for row in _rows(repo, TBL_RECEIVABLES)}
    assert receivables["26-0042"][sc.COL_RECV_AMOUNT_PAID] == 0
    assert receivables["26-0054"][sc.COL_RECV_AMOUNT_PAID] == 0
    assert _rows(repo, TBL_LEDGER) == []


def test_ap_expense_gateway_defaults_business_context_for_ap_form(tmp_path: Path) -> None:
    repo = _setup_repo(tmp_path)

    result = repo.save_ap_expense({
        "transactionId": "TXN-AP-FORM-1",
        "txnDate": "2026-07-01",
        "class": "Business",
        "type": "Expense",
        "payee": "LIHDC Professional Corporation",
        "categoryCode": "EXP_LEGAL_FEES",
        "amount": 100.0,
        "taxAmount": 13.0,
        "generalOfficeExpense": 1,
        "invoiceRef": "Settlement-2026-07-01",
    })

    assert result["ok"] is True
    saved = result["savedRow"]
    assert saved[sc.COL_TXN_FROM_ACCOUNT] == "AP_PAYABLE"
    assert saved[sc.COL_TXN_BUSINESS_UNIT] == "Cory Business"
    assert saved[sc.COL_TXN_TAX_FLAG] == "Business Deductible"
    assert saved[sc.COL_TXN_STATUS] == "Pending"
