from __future__ import annotations

import hashlib
import json
from pathlib import Path
import sys
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = PROJECT_ROOT / "src" / "python"
if str(SOURCE_ROOT) not in sys.path:
    sys.path.append(str(SOURCE_ROOT))

from domain import schema_constants as sc
from repositories.excel_repo import ExcelRepo
from services.paths import AppPaths
from services.workbook_integrity_service import WorkbookIntegrityService


def _repo(tmp_path: Path) -> tuple[ExcelRepo, AppPaths]:
    paths = AppPaths(root=PROJECT_ROOT, override_data_dir=tmp_path)
    repo = ExcelRepo(paths)
    repo.ensure_schema()
    return repo, paths


def _save_client(repo: ExcelRepo, name: str = "Synthetic Client") -> str:
    saved = repo.save_client_profile({"clientName": name, "status": "Active"})
    assert saved["ok"] is True
    return str(saved["clientId"])


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _close_workbook(workbook: object) -> None:
    for attribute in ("vba_archive", "_archive"):
        archive = getattr(workbook, attribute, None)
        if archive is not None:
            archive.close()
            setattr(workbook, attribute, None)
    workbook.close()  # type: ignore[attr-defined]


def _blank_matter_number(workbook_path: Path, matter_id: str) -> tuple[str, set[str]]:
    workbook = load_workbook(workbook_path, keep_vba=True, data_only=False)
    try:
        worksheet = workbook[sc.SHEET_MATTERS]
        table = worksheet.tables[sc.TBL_MATTERS]
        min_col, header_row, max_col, max_row = range_boundaries(table.ref)
        headers = [
            str(worksheet.cell(header_row, column).value or "").strip()
            for column in range(min_col, max_col + 1)
        ]
        id_column = min_col + headers.index(sc.COL_MATTER_ID)
        number_column = min_col + headers.index(sc.COL_MATTER_NUMBER)
        for row_number in range(header_row + 1, max_row + 1):
            if str(worksheet.cell(row_number, id_column).value or "").strip() == matter_id:
                worksheet.cell(row_number, number_column).value = ""
                break
        else:
            raise AssertionError("Synthetic matter was not found")
        table_ref = table.ref
        table_names = set(worksheet.tables.keys())
        workbook.save(workbook_path)
        return table_ref, table_names
    finally:
        _close_workbook(workbook)


def _add_seven_synthetic_reference_warnings(repo: ExcelRepo) -> None:
    repo._write_table_rows(
        sc.TBL_TRANSACTION_ACCOUNTS,
        [
            {
                sc.COL_TXN_ACCOUNT_CODE: "SYNTHETIC-VALID-ACCOUNT",
                sc.COL_TXN_ACCOUNT_NAME: "Synthetic valid account",
            }
        ],
    )
    repo._write_table_rows(
        sc.TBL_TRANSACTION_BUSINESS_UNITS,
        [{sc.COL_TXN_BUSINESS_UNIT_NAME: "SYNTHETIC-VALID-UNIT"}],
    )
    rows = []
    for index in range(1, 5):
        rows.append(
            {
                sc.COL_TXN_ID: f"SYNTHETIC-TXN-{index}",
                sc.COL_TXN_DATE: "2026-01-02",
                sc.COL_TXN_TYPE: "Transfer",
                sc.COL_TXN_AMOUNT: 1.0,
                sc.COL_TXN_FROM_ACCOUNT: f"SYNTHETIC-MISSING-ACCOUNT-{index}",
                sc.COL_TXN_BUSINESS_UNIT: (
                    f"SYNTHETIC-MISSING-UNIT-{index}" if index <= 3 else ""
                ),
            }
        )
    repo._write_table_rows(sc.TBL_TRANSACTIONS_MASTER, rows)


def test_missing_matter_number_is_detected_safely_without_changing_workbook(
    tmp_path: Path,
) -> None:
    repo, paths = _repo(tmp_path)
    confidential_client = "CONFIDENTIAL-SYNTHETIC-CLIENT-SENTINEL"
    confidential_matter = "CONFIDENTIAL-SYNTHETIC-MATTER-SENTINEL"
    client_id = _save_client(repo, confidential_client)
    saved = repo.save_matter_profile(
        {
            "clientId": client_id,
            "clientName": confidential_client,
            "matterName": confidential_matter,
            "dateOpened": "2026-01-02",
        }
    )
    assert saved["ok"] is True
    _add_seven_synthetic_reference_warnings(repo)

    workbook_path = paths.workbook_path()
    expected_ref, expected_tables = _blank_matter_number(workbook_path, str(saved["matterId"]))
    before_hash = _sha256(workbook_path)
    with ZipFile(workbook_path) as package:
        before_members = set(package.namelist())

    report = WorkbookIntegrityService(paths).check()

    assert report.error_count == 1
    required_issues = [issue for issue in report.issues if issue.code == "missing_required_value"]
    assert len(required_issues) == 1
    issue = required_issues[0]
    assert issue.table == sc.TBL_MATTERS
    assert issue.column == sc.COL_MATTER_NUMBER
    assert issue.value == ""
    assert report.warning_count == 7
    assert [issue.code for issue in report.issues if issue.severity == "warning"].count(
        "missing_account_reference"
    ) == 4
    assert [issue.code for issue in report.issues if issue.severity == "warning"].count(
        "missing_business_unit_reference"
    ) == 3

    diagnostic_json = report.to_json()
    assert confidential_client not in diagnostic_json
    assert confidential_matter not in diagnostic_json
    assert _sha256(workbook_path) == before_hash
    with ZipFile(workbook_path) as package:
        assert set(package.namelist()) == before_members

    reopened = load_workbook(workbook_path, keep_vba=True, data_only=False)
    try:
        matters = reopened[sc.SHEET_MATTERS]
        assert set(matters.tables.keys()) == expected_tables
        assert matters.tables[sc.TBL_MATTERS].ref == expected_ref
    finally:
        _close_workbook(reopened)


def test_supported_matter_profile_save_assigns_number_and_survives_reload(tmp_path: Path) -> None:
    repo, paths = _repo(tmp_path)
    client_id = _save_client(repo)

    saved = repo.save_matter_profile(
        {
            "clientId": client_id,
            "clientName": "Synthetic Client",
            "matterName": "Synthetic Profile Matter",
            "matterType": "General",
            "dateOpened": "2026-01-02",
        }
    )

    assert saved["ok"] is True
    assert str(saved["matterNumber"]).strip()
    reloaded_repo = ExcelRepo(paths)
    reloaded = reloaded_repo.get_matter_profile(str(saved["matterId"]))
    assert reloaded["ok"] is True
    assert reloaded["matter"]["matterNumber"] == saved["matterNumber"]
    assert WorkbookIntegrityService(paths).check().error_count == 0


def test_typed_time_entry_matter_creation_assigns_number_and_survives_reload(tmp_path: Path) -> None:
    repo, paths = _repo(tmp_path)

    saved = repo.add_time_entry(
        {
            "date": "2026-01-02",
            "clientName": "Synthetic Docket Client",
            "matterName": "Synthetic Docket Matter",
            "description": "Synthetic docket description",
            "hours": 1.0,
            "clientRate": 100.0,
            "sharePct": 100.0,
            "status": "Draft",
        }
    )

    assert saved["ok"] is True
    matters = ExcelRepo(paths).list_matter_directory()
    created = next(row for row in matters if row["matterName"] == "Synthetic Docket Matter")
    assert str(created["matterNumber"]).strip()
    assert WorkbookIntegrityService(paths).check().error_count == 0


def test_conflict_schema_and_a08_route_remain_inactive() -> None:
    schema_text = (PROJECT_ROOT / "schema" / "workbook_schema.yml").read_text(encoding="utf-8")
    pathways = (PROJECT_ROOT / "src" / "qml" / "standards" / "ModulePathways.js").read_text(
        encoding="utf-8"
    )

    assert "tblConflict" not in schema_text
    assert json.dumps("A08") not in pathways
    assert "ConflictInquiry" not in pathways
