from __future__ import annotations

from copy import copy
import json
from pathlib import Path
import shutil
import sys

import pytest
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_ROOT = PROJECT_ROOT / "scripts"
if str(SCRIPTS_ROOT) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_ROOT))

import repair_confirmed_orphan_records as repair  # noqa: E402


SYNTHETIC_MATTER_LABEL = "repair-fixture-sparse-matter"
SYNTHETIC_TIME_DESCRIPTION = "repair-fixture-fee"


def _table(workbook, table_name: str):
    return repair._find_table_object(workbook, table_name)


def _append_table_row(workbook, table_name: str, values: dict[str, object]) -> None:
    worksheet, table = _table(workbook, table_name)
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [repair._clean(worksheet.cell(min_row, column).value) for column in range(min_col, max_col + 1)]
    new_row = max_row + 1
    style_source_row = max(min_row, max_row)
    for offset, header in enumerate(headers):
        column = min_col + offset
        target = worksheet.cell(new_row, column)
        source = worksheet.cell(style_source_row, column)
        target._style = copy(source._style)
        target.value = values.get(header)
    table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_row}"


def _first_client_id(workbook) -> str:
    worksheet, table = _table(workbook, repair.sc.TBL_CLIENTS)
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [repair._clean(worksheet.cell(min_row, column).value) for column in range(min_col, max_col + 1)]
    client_column = min_col + headers.index(repair.sc.COL_CLIENT_ID)
    for row_number in range(min_row + 1, max_row + 1):
        value = repair._clean(worksheet.cell(row_number, client_column).value)
        if value:
            return value
    raise AssertionError("The test fixture requires one existing client identifier.")


def _inject_reviewed_records(path: Path) -> None:
    workbook = load_workbook(path, keep_vba=True, data_only=False, keep_links=True)
    try:
        client_id = _first_client_id(workbook)
        _append_table_row(
            workbook,
            repair.sc.TBL_MATTERS,
            {
                repair.sc.COL_MATTER_ID: repair.AUTHORIZED_MATTER_ID,
                repair.sc.COL_MATTER_NAME: SYNTHETIC_MATTER_LABEL,
                repair.sc.COL_MATTER_CLIENT_ID: client_id,
                repair.sc.COL_MATTER_STATUS: "Active",
            },
        )
        base_time = {
            repair.sc.COL_TIME_DATE: "2026-08-05",
            repair.sc.COL_TIME_CLIENT_ID: client_id,
            repair.sc.COL_TIME_MATTER_ID: repair.AUTHORIZED_MATTER_ID,
            repair.sc.COL_TIME_DESC: SYNTHETIC_TIME_DESCRIPTION,
            repair.sc.COL_TIME_HOURS: 0,
            repair.sc.COL_TIME_RATE: 0,
            repair.sc.COL_TIME_GROSS: 5000,
            repair.sc.COL_TIME_NET: 5000,
            repair.sc.COL_TIME_HST: 650,
            repair.sc.COL_TIME_INVOICE_STATUS: "Unbilled",
        }
        for entry_id in repair.AUTHORIZED_ENTRY_IDS:
            _append_table_row(
                workbook,
                repair.sc.TBL_TIME,
                {repair.sc.COL_TIME_ENTRY_ID: entry_id, **base_time},
            )
        workbook.save(path)
    finally:
        repair._close_workbook(workbook)


def _ready_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "CSPM.xlsm"
    shutil.copy2(repair.REPOSITORY_WORKBOOK, path)
    snapshot = repair.snapshot_workbook(path)
    status = repair.evaluate_preconditions(snapshot)["status"]
    if status == "already_repaired":
        _inject_reviewed_records(path)
    assert repair.evaluate_preconditions(repair.snapshot_workbook(path))["status"] == "ready"
    return path


def _set_target_time_value(path: Path, column_name: str, value: object) -> None:
    workbook = load_workbook(path, keep_vba=True, data_only=False, keep_links=True)
    try:
        worksheet, table = _table(workbook, repair.sc.TBL_TIME)
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [repair._clean(worksheet.cell(min_row, column).value) for column in range(min_col, max_col + 1)]
        key_column = min_col + headers.index(repair.sc.COL_TIME_ENTRY_ID)
        value_column = min_col + headers.index(column_name)
        for row_number in range(min_row + 1, max_row + 1):
            if repair._clean(worksheet.cell(row_number, key_column).value) == repair.AUTHORIZED_ENTRY_IDS[0]:
                worksheet.cell(row_number, value_column).value = value
                workbook.save(path)
                return
        raise AssertionError("Authorized test entry was not found.")
    finally:
        repair._close_workbook(workbook)


def _set_external_dependency(path: Path, table_name: str, column_name: str, label: str) -> None:
    workbook = load_workbook(path, keep_vba=True, data_only=False, keep_links=True)
    try:
        worksheet, table = _table(workbook, table_name)
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [repair._clean(worksheet.cell(min_row, column).value) for column in range(min_col, max_col + 1)]
        value_column = min_col + headers.index(column_name)
        for row_number in range(min_row + 1, max_row + 1):
            row_values = [repair._clean(worksheet.cell(row_number, column).value) for column in range(min_col, max_col + 1)]
            if repair.PROTECTED_INVOICE not in row_values:
                worksheet.cell(row_number, value_column).value = f"{label}:{repair.AUTHORIZED_ENTRY_IDS[0]}"
                workbook.save(path)
                return
        raise AssertionError(f"The {table_name} test fixture has no unprotected row.")
    finally:
        repair._close_workbook(workbook)


def _repair_candidate(path: Path):
    before = repair.snapshot_workbook(path)
    result = repair.apply_candidate_record_removal(path)
    after = repair.snapshot_workbook(path)
    return before, result, after


def test_only_authorized_entries_and_matter_are_removed(tmp_path: Path) -> None:
    path = _ready_workbook(tmp_path)
    before, result, after = _repair_candidate(path)

    assert result["removed"] == {repair.sc.TBL_TIME: 2, repair.sc.TBL_MATTERS: 1}
    for table_name in sorted(before.tables):
        expected_delta = repair.TARGET_TABLE_REMOVALS.get(table_name, 0)
        assert len(before.tables[table_name].rows) - len(after.tables[table_name].rows) == expected_delta
        assert repair._filtered_table_payload(before, table_name) == repair._filtered_table_payload(after, table_name)


def test_sparse_matter_removal_requires_dependents_first(tmp_path: Path) -> None:
    path = _ready_workbook(tmp_path)
    workbook = load_workbook(path, keep_vba=True, data_only=False, keep_links=True)
    try:
        with pytest.raises(repair.RepairRefused, match="still has a dependent"):
            repair._assert_matter_removal_safe(workbook)
        repair._delete_exact_table_rows(
            workbook,
            repair.sc.TBL_TIME,
            repair.sc.COL_TIME_ENTRY_ID,
            repair.AUTHORIZED_ENTRY_IDS,
        )
        repair._assert_matter_removal_safe(workbook)
        repair._delete_exact_table_rows(
            workbook,
            repair.sc.TBL_MATTERS,
            repair.sc.COL_MATTER_ID,
            [repair.AUTHORIZED_MATTER_ID],
        )
    finally:
        repair._close_workbook(workbook)


def test_unexpected_matter_dependent_blocks_repair(tmp_path: Path) -> None:
    path = _ready_workbook(tmp_path)
    workbook = load_workbook(path, keep_vba=True, data_only=False, keep_links=True)
    try:
        client_id = _first_client_id(workbook)
        _append_table_row(
            workbook,
            repair.sc.TBL_TIME,
            {
                repair.sc.COL_TIME_ENTRY_ID: "fixture-unexpected-dependent",
                repair.sc.COL_TIME_DATE: "2026-08-05",
                repair.sc.COL_TIME_CLIENT_ID: client_id,
                repair.sc.COL_TIME_MATTER_ID: repair.AUTHORIZED_MATTER_ID,
                repair.sc.COL_TIME_DESC: "fixture-dependent",
                repair.sc.COL_TIME_HOURS: 0,
                repair.sc.COL_TIME_RATE: 0,
                repair.sc.COL_TIME_GROSS: 0,
                repair.sc.COL_TIME_NET: 0,
                repair.sc.COL_TIME_HST: 0,
                repair.sc.COL_TIME_INVOICE_STATUS: "Unbilled",
            },
        )
        workbook.save(path)
    finally:
        repair._close_workbook(workbook)

    with pytest.raises(repair.RepairRefused, match="unexpected dependent"):
        repair.evaluate_preconditions(repair.snapshot_workbook(path))


@pytest.mark.parametrize(
    ("dependency_kind", "table_name", "column_name"),
    [
        ("invoice", repair.sc.TBL_TIME, repair.sc.COL_TIME_INVOICE_REF),
        ("payment", repair.sc.TBL_TIME, repair.sc.COL_TIME_PAYMENT_STATUS),
        ("credit", repair.sc.TBL_TRANSACTIONS_MASTER, repair.sc.COL_TXN_NOTES),
        ("adjustment", repair.sc.TBL_TRANSACTIONS_MASTER, repair.sc.COL_TXN_NOTES),
        ("transaction", repair.sc.TBL_TRANSACTIONS_MASTER, repair.sc.COL_TXN_NOTES),
        ("ledger", repair.sc.TBL_LEDGER, repair.sc.COL_LEDGER_EXTERNAL_REF_ID),
    ],
)
def test_financial_dependencies_block_repair(
    tmp_path: Path,
    dependency_kind: str,
    table_name: str,
    column_name: str,
) -> None:
    path = _ready_workbook(tmp_path)
    if table_name == repair.sc.TBL_TIME:
        _set_target_time_value(path, column_name, f"fixture-{dependency_kind}")
    else:
        _set_external_dependency(path, table_name, column_name, dependency_kind)

    with pytest.raises(repair.RepairRefused):
        repair.evaluate_preconditions(repair.snapshot_workbook(path))


def test_target_matter_number_error_disappears_after_candidate_repair(tmp_path: Path) -> None:
    path = _ready_workbook(tmp_path)
    before_integrity = repair._integrity_summary(path)
    repair.apply_candidate_record_removal(path)
    after_integrity = repair._integrity_summary(path)

    assert before_integrity["hasTargetMatterNumberError"] is True
    assert before_integrity["errorCount"] == 1
    assert after_integrity["hasTargetMatterNumberError"] is False
    assert after_integrity["errorCount"] == 0


def test_seven_independent_warnings_remain_visible(tmp_path: Path) -> None:
    path = _ready_workbook(tmp_path)
    repair.apply_candidate_record_removal(path)
    integrity = repair._integrity_summary(path)

    assert integrity["warningCount"] == 7
    assert integrity["warningCounts"] == repair.EXPECTED_WARNING_COUNTS


def test_protected_invoice_chain_is_value_for_value_unchanged(tmp_path: Path) -> None:
    path = _ready_workbook(tmp_path)
    before, _, after = _repair_candidate(path)

    assert repair._protected_invoice_payload(before)
    assert repair._digest_json(repair._protected_invoice_payload(before)) == repair._digest_json(
        repair._protected_invoice_payload(after)
    )


def test_unrelated_financial_control_totals_are_unchanged(tmp_path: Path) -> None:
    path = _ready_workbook(tmp_path)
    before, _, after = _repair_candidate(path)

    assert repair._financial_control_digest(before) == repair._financial_control_digest(after)
    assert repair._unrelated_data_digest(before) == repair._unrelated_data_digest(after)


def test_macro_container_and_workbook_structure_are_preserved(tmp_path: Path) -> None:
    path = _ready_workbook(tmp_path)
    before = repair.snapshot_workbook(path)
    repair.apply_candidate_record_removal(path)
    verification = repair.verify_repaired_candidate(before, path)

    assert verification["macrosPreserved"] is True
    assert verification["packageMembersPreserved"] is True
    assert verification["formulasPreserved"] is True
    assert verification["definedNamesPreserved"] is True
    assert verification["validationsPreserved"] is True
    assert verification["formattingPreserved"] is True
    assert verification["tableStructurePreserved"] is True


def test_unknown_live_and_cloud_targets_are_refused_without_change(tmp_path: Path) -> None:
    unknown = tmp_path / "unknown.xlsm"
    shutil.copy2(repair.REPOSITORY_WORKBOOK, unknown)
    boundaries = repair._load_boundary_paths()
    boundary_before = {name: repair.fingerprint(path, name) for name, path in boundaries.items()}

    with pytest.raises(repair.RepairRefused, match="Unknown workbook target refused"):
        repair.assert_repository_target(unknown, forbidden_paths=list(boundaries.values()))
    for boundary in boundaries.values():
        with pytest.raises(repair.RepairRefused):
            repair.assert_repository_target(boundary, forbidden_paths=list(boundaries.values()))

    boundary_after = {name: repair.fingerprint(path, name) for name, path in boundaries.items()}
    repair._assert_fingerprints_unchanged(boundary_before, boundary_after)


def test_repeated_candidate_execution_reports_already_repaired(tmp_path: Path) -> None:
    path = _ready_workbook(tmp_path)
    first = repair.apply_candidate_record_removal(path)
    first_hash = repair._sha256(path)
    second = repair.apply_candidate_record_removal(path)

    assert first["changed"] is True
    assert second == {"status": "already_repaired", "changed": False, "sha256": first_hash}
    assert repair._sha256(path) == first_hash


def test_public_preflight_omits_confidential_row_values(tmp_path: Path) -> None:
    path = _ready_workbook(tmp_path)
    snapshot = repair.snapshot_workbook(path)
    preconditions = repair.evaluate_preconditions(snapshot)
    fingerprint = repair.fingerprint(path, "repository_cspm")
    public = repair._public_preflight(
        fingerprint,
        preconditions,
        {"read_only_fixture": fingerprint},
        fingerprint,
    )
    serialized = json.dumps(public, sort_keys=True)

    assert SYNTHETIC_MATTER_LABEL not in serialized
    assert SYNTHETIC_TIME_DESCRIPTION not in serialized
    assert "MatterName" not in serialized
    assert "Description" not in serialized
