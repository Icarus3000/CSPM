import openpyxl
import sys
import traceback

file_path = r"C:\Projects\__CSPM\Dockets_temp.xlsm"

try:
    # Get sheet names
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    # Try to read the most likely sheets
    for sheet_name in wb.sheetnames:
        # Check all sheets since we aren't sure which one has the data
        ws = wb[sheet_name]
        
        # Get first 10 rows
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 10:
                break
            rows.append(row)
        
        if rows:
            print(f"\n--- Sheet: {sheet_name} ---")
            print("Row 1 (Possible Headers):")
            print(rows[0])
            print("Row 2 (Possible Data):")
            if len(rows) > 1:
                print(rows[1])
            print("Row 3 (Possible Data):")
            if len(rows) > 2:
                print(rows[2])
                
except Exception as e:
    print(f"Error reading Excel file: {e}")
    traceback.print_exc()
