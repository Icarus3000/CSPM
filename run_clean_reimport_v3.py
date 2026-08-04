"""
MASTER CLEAN RE-IMPORT v3
─────────────────────────
Goal: After this script, CSPM data alone should produce:
  Revenue + WIP    ≈ $176,098.21
  Adj. Production  ≈ $174,950.25

Steps:
  1. REMOVE the legacy "Dockets" sheet from CSPM.xlsm (force dashboard to use TimeEntries)
  2. CLEAR ALL TimeEntries rows
  3. CLEAR ALL Transactions rows
  4. RE-IMPORT Dockets from fresh source with:
     - Correct Client/Billing Client mapping (Excel Client = CSPM Billing Client/Parent)
     - One faux matter per (Client, Billing Client) pair
  5. RE-IMPORT Ledger with correct Type values
  6. Run financial dashboard benchmark
"""
import sys, os, math
from pathlib import Path
from datetime import datetime, date
from typing import Dict, Any, Optional, List
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import range_boundaries, get_column_letter

root = Path(__file__).resolve().parent
sys.path.insert(0, str(root / "src" / "python"))

from services.paths import AppPaths
from repositories.excel_repo import ExcelRepo
from domain import schema_constants as sc

src_path = root / "data" / "Dockets.xlsm"
cspm_path = root / "data" / "CSPM.xlsm"

def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        if isinstance(val, str):
            import re
            cleaned = re.sub(r'[^\d.\-]', '', val)
            return float(cleaned) if cleaned else default
        return float(val)
    except (ValueError, TypeError):
        return default

def clean(val):
    if val is None:
        return ""
    return str(val).strip()

def format_date(val):
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    if val:
        return str(val)[:10]
    return ""


# ══════════════════════════════════════════════════════════
# STEP 1+2+3: Remove Dockets sheet, clear TimeEntries & Transactions
# ══════════════════════════════════════════════════════════
print("=" * 60)
print("STEP 1-3: Remove Dockets sheet, Clear TimeEntries & Transactions")
print("=" * 60)

wb = openpyxl.load_workbook(str(cspm_path), keep_vba=True)

# Remove legacy Dockets sheet
if "Dockets" in wb.sheetnames:
    del wb["Dockets"]
    print("  Removed legacy 'Dockets' sheet")

# Clear TimeEntries
if "TimeEntries" in wb.sheetnames:
    ws = wb["TimeEntries"]
    old_count = ws.max_row - 1
    if old_count > 0:
        ws.delete_rows(2, old_count)
    # Update table ref
    for tbl_name in list(ws.tables):
        tbl = ws.tables[tbl_name]
        _, _, max_col, _ = range_boundaries(tbl.ref)
        tbl.ref = f"A1:{get_column_letter(max_col)}2"
    print(f"  Cleared {old_count} TimeEntries rows")

# Clear Transactions
if "Transactions" in wb.sheetnames:
    ws = wb["Transactions"]
    old_count = ws.max_row - 1
    if old_count > 0:
        ws.delete_rows(2, old_count)
    for tbl_name in list(ws.tables):
        tbl = ws.tables[tbl_name]
        _, _, max_col, _ = range_boundaries(tbl.ref)
        tbl.ref = f"A1:{get_column_letter(max_col)}2"
    print(f"  Cleared {old_count} Transactions rows")

wb.save(str(cspm_path))
wb.close()
print("  Saved CSPM.xlsm")


# ══════════════════════════════════════════════════════════
# STEP 4: Re-import using DocketsImportService
# ══════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("STEP 4: Re-import from Dockets.xlsm")
print("=" * 60)

from services.dockets_import_service import DocketsImportService

paths = AppPaths(root)
repo = ExcelRepo(paths)

def duplicate_callback(payload):
    kind = (payload.get("kind") or "").lower()
    if kind == "client":
        return {"action": "skip", "scope": "one"}
    return {"action": "add", "scope": "one"}

result = DocketsImportService(repo).import_legacy_workbook(
    str(src_path),
    duplicate_callback=duplicate_callback,
)

print(f"  Clients added: {result.get('clientsAdded', 0)}")
print(f"  Matters added: {result.get('mattersAdded', 0)}")
print(f"  Dockets added: {result.get('docketsAdded', 0)}")
print(f"  Ledger added:  {result.get('ledgerAdded', 0)}")
print(f"  Errors: {len(result.get('errors', []))}")
for e in result.get("errors", [])[:5]:
    print(f"    ERROR: {e}")

warns = result.get("warnings", [])
print(f"  Warnings: {len(warns)}")
for w in warns[:10]:
    print(f"    WARN: {w}")


# ══════════════════════════════════════════════════════════
# STEP 5: Verify TimeEntries data
# ══════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("STEP 5: Verify TimeEntries & Transactions data")
print("=" * 60)

# Check TimeEntries
wb_check = openpyxl.load_workbook(str(cspm_path), read_only=True, data_only=True)

if "TimeEntries" in wb_check.sheetnames:
    ws = wb_check["TimeEntries"]
    headers = [str(c.value) if c.value else "" for c in ws[1]]
    
    gross_idx = headers.index("GrossToClient") if "GrossToClient" in headers else -1
    date_idx = headers.index("Date") if "Date" in headers else -1
    status_idx = headers.index("Status") if "Status" in headers else -1
    inv_ref_idx = headers.index("InvoiceRef") if "InvoiceRef" in headers else -1
    
    count = 0
    yr_2026_count = 0
    yr_2026_gross = 0.0
    yr_2026_wip = 0.0
    yr_2026_billed = 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        count += 1
        d = row[date_idx] if date_idx >= 0 else None
        yr = None
        if isinstance(d, datetime):
            yr = d.year
        elif isinstance(d, str) and len(d) >= 4:
            try: yr = int(d[:4])
            except: pass
        
        if yr == 2026:
            yr_2026_count += 1
            g = safe_float(row[gross_idx]) if gross_idx >= 0 else 0
            yr_2026_gross += g
            
            status = str(row[status_idx]).lower() if status_idx >= 0 and row[status_idx] else ""
            inv = str(row[inv_ref_idx]).strip() if inv_ref_idx >= 0 and row[inv_ref_idx] else ""
            
            if status in ("billed", "merged") or (inv and inv != "None"):
                yr_2026_billed += g
            else:
                yr_2026_wip += g
    
    print(f"  TimeEntries total rows: {count}")
    print(f"  2026 rows: {yr_2026_count}")
    print(f"  2026 GrossToClient total: ${yr_2026_gross:,.2f}")
    print(f"  2026 Billed amount: ${yr_2026_billed:,.2f}")
    print(f"  2026 WIP amount: ${yr_2026_wip:,.2f}")

# Check Transactions
if "Transactions" in wb_check.sheetnames:
    ws = wb_check["Transactions"]
    headers = [str(c.value) if c.value else "" for c in ws[1]]
    
    type_idx = headers.index("Type") if "Type" in headers else -1
    amt_idx = headers.index("Amount") if "Amount" in headers else -1
    
    txn_count = 0
    type_totals = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        txn_count += 1
        typ = str(row[type_idx]) if type_idx >= 0 and row[type_idx] else "?"
        amt = safe_float(row[amt_idx]) if amt_idx >= 0 else 0
        type_totals[typ] = type_totals.get(typ, 0.0) + amt
    
    print(f"\n  Transactions total: {txn_count}")
    for typ, tot in sorted(type_totals.items()):
        print(f"    {typ}: ${tot:,.2f}")

wb_check.close()


# ══════════════════════════════════════════════════════════
# STEP 6: Run financial dashboard benchmark
# ══════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("STEP 6: Financial Dashboard Benchmark")
print("=" * 60)

repo2 = ExcelRepo(paths)
dashboard = repo2.financial_dashboard_report(2026)
summary = dashboard.get("summary", {})

rev_wip = summary.get("revenueIncludingWip", 0.0)
prod = summary.get("docketedAmount", 0.0)

target_rev_wip = 176098.21
target_prod = 174950.25

delta_rev = rev_wip - target_rev_wip
delta_prod = prod - target_prod

print(f"  {'Metric':25s} {'Python':>14s} {'Excel Target':>14s} {'Delta':>12s}")
print(f"  {'-'*25} {'-'*14} {'-'*14} {'-'*12}")
print(f"  {'Revenue + WIP':25s} ${rev_wip:>12,.2f}  ${target_rev_wip:>12,.2f}  ${delta_rev:>+10,.2f}")
print(f"  {'Adjusted Production':25s} ${prod:>12,.2f}  ${target_prod:>12,.2f}  ${delta_prod:>+10,.2f}")

print(f"\n  Dashboard details:")
for k in ["billedTimeAmount", "wipAmount", "wipHours", "ledgerBillings", "expenses", "bankedAmount"]:
    v = summary.get(k, "N/A")
    if isinstance(v, float):
        print(f"    {k:30s}: ${v:>12,.2f}")
    else:
        print(f"    {k:30s}: {v}")

if abs(delta_rev) < 500 and abs(delta_prod) < 500:
    print("\n  [PASS] -- Both metrics within $500 tolerance.")
elif abs(delta_rev) < 2000 and abs(delta_prod) < 2000:
    print("\n  [CLOSE] -- Within $2000, likely rounding/date boundary differences.")
else:
    print("\n  [FAIL] -- NEEDS INVESTIGATION -- Discrepancy exceeds tolerance.")
