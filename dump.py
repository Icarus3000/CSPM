import openpyxl
path = r'C:\Projects\__CSPM\data\CSPM.xlsm'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Matters']
headers = [cell.value for cell in ws[1]]
try:
    id_idx = headers.index('MatterID')
    no_idx = headers.index('MatterNumber')
    name_idx = headers.index('MatterName')
    for r in range(2, 6):
        print(f'ID: {ws[r][id_idx].value}, NO: {ws[r][no_idx].value}, NAME: {ws[r][name_idx].value}')
except Exception as e:
    print(e)
