"""
Check what Class/Type values exist in CSPM Transactions table
and what the Ledger sheet contains.
"""
import sys, os
from pathlib import Path
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "src", "python"))
from services.paths import AppPaths

paths = AppPaths(Path(__file__).resolve().parent)
wb_path = paths.workbook_path()

wb = openpyxl.load_workbook(str(wb_path), read_only=True, data_only=True)

# Check Transactions sheet
ws = wb["Transactions"]
headers = [str(cell.value) if cell.value else "" for cell in ws[1]]
print(f"Transaction headers: {headers}")

# Find column indices
class_idx = headers.index("Class") if "Class" in headers else -1
type_idx = headers.index("Type") if "Type" in headers else -1
amount_idx = headers.index("Amount") if "Amount" in headers else -1
client_idx = headers.index("Client") if "Client" in headers else -1
invoice_ref_idx = headers.index("InvoiceRef") if "InvoiceRef" in headers else -1
date_idx = headers.index("TxnDate") if "TxnDate" in headers else -1

print(f"  Class col: {class_idx}, Type col: {type_idx}, Amount col: {amount_idx}")
print(f"  InvoiceRef col: {invoice_ref_idx}, Client col: {client_idx}")

classes = {}
types = {}
class_type_combos = {}
revenue_income_total = 0.0
revenue_income_2026 = 0.0
row_count = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(c is not None for c in row[:5]):
        continue
    row_count += 1
    
    cls = str(row[class_idx]) if class_idx >= 0 and row[class_idx] else "(empty)"
    typ = str(row[type_idx]) if type_idx >= 0 and row[type_idx] else "(empty)"
    amt = 0.0
    try:
        amt = float(row[amount_idx]) if amount_idx >= 0 and row[amount_idx] else 0.0
    except:
        pass
    
    classes[cls] = classes.get(cls, 0) + 1
    types[typ] = types.get(typ, 0) + 1
    combo = f"{cls} / {typ}"
    if combo not in class_type_combos:
        class_type_combos[combo] = {"count": 0, "total_amt": 0.0}
    class_type_combos[combo]["count"] += 1
    class_type_combos[combo]["total_amt"] += amt
    
    if cls == "Revenue" and typ == "Income":
        revenue_income_total += amt
        d = str(row[date_idx]) if date_idx >= 0 and row[date_idx] else ""
        if "2026" in d:
            revenue_income_2026 += amt

print(f"\nTotal rows: {row_count}")
print(f"\nClasses: {classes}")
print(f"\nTypes: {types}")
print(f"\nClass/Type combos:")
for combo, data in sorted(class_type_combos.items()):
    print(f"  {combo:40s}  count={data['count']:>5d}  total=${data['total_amt']:>12,.2f}")

print(f"\nRevenue/Income total (all years): ${revenue_income_total:,.2f}")
print(f"Revenue/Income total (2026):      ${revenue_income_2026:,.2f}")

# Also check CSPM Ledger sheet
print("\n" + "=" * 70)
print("CSPM Ledger Sheet")
ws = wb["Ledger"]
headers = [str(cell.value) if cell.value else "" for cell in ws[1]]
print(f"Headers: {headers}")

billing_idx = headers.index("BillingsExclHST") if "BillingsExclHST" in headers else -1
date_idx2 = headers.index("Date") if "Date" in headers else -1
client_idx2 = headers.index("ClientVendor") if "ClientVendor" in headers else -1
ref_idx = headers.index("Reference") if "Reference" in headers else -1

total_billings = 0.0
billings_2026 = 0.0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(c is not None for c in row[:5]):
        continue
    amt = 0.0
    try:
        amt = float(row[billing_idx]) if billing_idx >= 0 and row[billing_idx] else 0.0
    except:
        pass
    total_billings += amt
    d = str(row[date_idx2]) if date_idx2 >= 0 and row[date_idx2] else ""
    if "2026" in d:
        billings_2026 += amt

print(f"Total Billings (all years): ${total_billings:,.2f}")
print(f"Total Billings (2026):      ${billings_2026:,.2f}")

wb.close()
