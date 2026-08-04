"""
Diagnose why the financial numbers are ~2x the target.
Check for double-counting between legacy Dockets sheet and TimeEntries.
"""
import sys, os
from pathlib import Path
root = Path(__file__).resolve().parent
sys.path.insert(0, str(root / "src" / "python"))

from services.paths import AppPaths
from repositories.excel_repo import ExcelRepo

paths = AppPaths(root)
repo = ExcelRepo(paths)

# 1. Check legacy docket rows
legacy_rows = repo._read_raw_excel_table_rows("Dockets", "tblDockets")
print(f"Legacy Dockets rows: {len(legacy_rows)}")

# Count by year
from datetime import datetime
year_counts = {}
year_totals = {}
for row in legacy_rows:
    d = row.get("Date")
    yr = None
    if isinstance(d, datetime):
        yr = d.year
    elif isinstance(d, str):
        try:
            yr = int(d[:4])
        except:
            pass
    yr_key = yr or "unknown"
    year_counts[yr_key] = year_counts.get(yr_key, 0) + 1
    
    amt = row.get("Amount to CS")
    try:
        v = float(amt) if amt is not None else 0.0
    except:
        v = 0.0
    year_totals[yr_key] = year_totals.get(yr_key, 0.0) + v

print(f"  By year: {dict(sorted(year_counts.items()))}")
print(f"  Totals:  ", {k: f"${v:,.2f}" for k, v in sorted(year_totals.items())})

# 2. Check TimeEntries rows
import openpyxl
from domain import schema_constants as sc
cspm_path = paths.workbook_path()
wb = openpyxl.load_workbook(str(cspm_path), read_only=True, data_only=True)
if "TimeEntries" in wb.sheetnames:
    ws = wb["TimeEntries"]
    headers = [str(c.value) if c.value else "" for c in ws[1]]
    te_count = 0
    te_2026 = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        te_count += 1
        d = row[headers.index("Date")] if "Date" in headers else None
        if isinstance(d, datetime) and d.year == 2026:
            te_2026 += 1
        elif isinstance(d, str) and "2026" in str(d):
            te_2026 += 1
    print(f"\nTimeEntries rows: {te_count}")
    print(f"  2026 rows: {te_2026}")
else:
    print("\nTimeEntries sheet NOT FOUND")
wb.close()

# 3. Check Transactions
wb2 = openpyxl.load_workbook(str(cspm_path), read_only=True, data_only=True)
if "Transactions" in wb2.sheetnames:
    ws = wb2["Transactions"]
    headers = [str(c.value) if c.value else "" for c in ws[1]]
    txn_count = 0
    type_counts = {}
    type_totals = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        txn_count += 1
        typ_idx = headers.index("Type") if "Type" in headers else -1
        amt_idx = headers.index("Amount") if "Amount" in headers else -1
        typ = row[typ_idx] if typ_idx >= 0 else "?"
        amt = float(row[amt_idx]) if amt_idx >= 0 and row[amt_idx] is not None else 0.0
        type_counts[typ] = type_counts.get(typ, 0) + 1
        type_totals[typ] = type_totals.get(typ, 0.0) + amt
    print(f"\nTransactions rows: {txn_count}")
    print(f"  By Type: {type_counts}")
    print(f"  Totals:  ", {k: f"${v:,.2f}" for k, v in type_totals.items()})
else:
    print("\nTransactions sheet NOT FOUND")
wb2.close()

# 4. Check dashboard internals
print("\n" + "=" * 60)
print("Dashboard internals (financial_dashboard_report):")
dashboard = repo.financial_dashboard_report(2026)
summary = dashboard.get("summary", {})
for k, v in sorted(summary.items()):
    if isinstance(v, (int, float)):
        print(f"  {k:30s}: {v:>14,.2f}")
    elif isinstance(v, str):
        print(f"  {k:30s}: {v}")
