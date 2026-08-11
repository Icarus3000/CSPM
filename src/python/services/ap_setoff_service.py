from __future__ import annotations

"""One-workbook posting and reversal for A/P payments settled against A/R."""

from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
import json
from typing import Any, Mapping

from openpyxl import load_workbook

from domain import schema_constants as sc
from domain.ap_lifecycle import APValidationError, build_bill_snapshot, clean_text
from domain.ap_schema import (
    AP_BILLS_HEADERS, AP_BILLS_SHEET, AP_BILLS_TABLE,
    AP_PAYMENTS_HEADERS, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE,
)
from repositories.ap_workbook_repository import _read_rows, _replace_rows
from repositories.excel_repo import (
    TBL_LEDGER, TBL_RECEIVABLES, TBL_TIME, TBL_TRANSACTION_ACCOUNTS,
    TBL_TRANSACTION_CATEGORIES, TBL_TRANSACTIONS_MASTER, with_db_lock,
)


SETOFF_ACCOUNT = "AR_SET_OFF"
SETOFF_ACCOUNT_NAME = "Accounts Receivable - Set-off"
SETOFF_METHOD = "Set-off"
ALLOCATION_MARKER = "CSPM_SET_OFF_ALLOCATIONS_V1:"
SETOFF_CATEGORY_CODE = "AP_SET_OFF"
SETOFF_CATEGORY_NAME = "Settlement set-off"


def _amount(value: Any, field: str = "amount") -> Decimal:
    try:
        result = Decimal(str(value if value not in (None, "") else 0)).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    except Exception as exc:
        raise APValidationError(f"Invalid {field}.") from exc
    return result


def _stamp() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


class APSetoffService:
    """Posts both sides of a set-off in memory and saves the workbook once."""

    def __init__(self, excel_repo: Any) -> None:
        self.repo = excel_repo

    def _read_finance(self, workbook: Any, tref: Any) -> list[dict[str, Any]]:
        ws = workbook[tref.sheet]
        if tref.table not in ws.tables:
            raise APValidationError(f"Required table is missing: {tref.table}")
        _headers, raw_rows = self.repo._rows_from_table(ws, ws.tables[tref.table])
        return [
            {header: self.repo._value_with_alias(tref.table, raw, header)
             for header in sc.TABLE_COLUMNS[tref.table]}
            for raw in raw_rows
        ]

    def _write_finance(self, workbook: Any, tref: Any, rows: list[dict[str, Any]]) -> None:
        ws = workbook[tref.sheet]
        table = ws.tables[tref.table]
        self.repo._write_table(
            ws, table, tref.table, sc.TABLE_COLUMNS[tref.table], rows,
            table.tableStyleInfo,
        )

    @staticmethod
    def _find(rows: list[dict[str, Any]], key: str, value: Any) -> dict[str, Any] | None:
        wanted = clean_text(value).casefold()
        return next((row for row in rows if clean_text(row.get(key)).casefold() == wanted), None)

    @staticmethod
    def _note(existing: Any, extra: str) -> str:
        existing = clean_text(existing)
        return f"{existing}\n{extra}" if existing else extra

    @staticmethod
    def _allocations(payload: Mapping[str, Any]) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        seen: set[str] = set()
        for item in list(payload.get("Allocations") or payload.get("allocations") or []):
            item = dict(item or {})
            invoice = clean_text(
                item.get("InvoiceNum") or item.get("InvoiceID") or item.get("invoice") or item.get("Invoice")
            )
            amount = _amount(item.get("Amount", item.get("amount")), "allocation amount")
            key = invoice.casefold()
            if not invoice or amount <= 0:
                raise APValidationError("Each set-off allocation needs an invoice and positive amount.")
            if key in seen:
                raise APValidationError(f"Invoice {invoice} appears more than once in this set-off.")
            seen.add(key)
            result.append({"invoice": invoice, "amount": amount})
        if not result:
            raise APValidationError("Enter at least one receivable allocation for the set-off.")
        return result

    @staticmethod
    def _stored_allocations(notes: Any) -> list[dict[str, Any]]:
        for line in str(notes or "").splitlines():
            if line.startswith(ALLOCATION_MARKER):
                try:
                    return APSetoffService._allocations({"allocations": json.loads(line[len(ALLOCATION_MARKER):])})
                except Exception as exc:
                    raise APValidationError("The set-off payment has unreadable allocation evidence.") from exc
        raise APValidationError("The set-off payment has no allocation evidence to reverse.")

    @staticmethod
    def _payment_lifecycle(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return [
            {"payment_id": row.get("APPaymentID"), "amount": row.get("Amount"),
             "reversed": clean_text(row.get("Status")).casefold() == "reversed"}
            for row in rows if clean_text(row.get("Status")).casefold() != "reversal"
        ]

    @staticmethod
    def _set_receivable(row: dict[str, Any], amount_delta: Decimal) -> tuple[Decimal, str]:
        total = _amount(row.get(sc.COL_RECV_TOTAL_INVOICED), "invoice total")
        paid = _amount(row.get(sc.COL_RECV_AMOUNT_PAID), "amount paid") + amount_delta
        credits = _amount(row.get(sc.COL_RECV_CREDITS_ADJ), "credits")
        if paid < 0:
            raise APValidationError(f"Invoice {row.get(sc.COL_RECV_INVOICE_NUM)} cannot be reversed below $0.00 paid.")
        balance = (total - paid - credits).quantize(Decimal("0.01"))
        if balance < Decimal("-0.01"):
            raise APValidationError(f"Invoice {row.get(sc.COL_RECV_INVOICE_NUM)} would be overpaid.")
        if abs(balance) <= Decimal("0.01"):
            balance = Decimal("0.00")
        row[sc.COL_RECV_AMOUNT_PAID] = float(paid)
        row[sc.COL_RECV_BALANCE_DUE] = float(balance)
        row[sc.COL_RECV_STATUS] = "Paid" if balance <= 0 else "Partial"
        return balance, str(row[sc.COL_RECV_STATUS])

    @staticmethod
    def _sync_time(rows: list[dict[str, Any]], receivable: Mapping[str, Any]) -> None:
        invoice = clean_text(receivable.get(sc.COL_RECV_INVOICE_NUM)).casefold()
        for row in rows:
            if clean_text(row.get(sc.COL_TIME_INVOICE_REF)).casefold() != invoice:
                continue
            row[sc.COL_TIME_PAYMENT_STATUS] = receivable.get(sc.COL_RECV_STATUS)
            row[sc.COL_TIME_INVOICE_TOTAL] = receivable.get(sc.COL_RECV_TOTAL_INVOICED)
            row[sc.COL_TIME_INVOICE_AMOUNT_PAID] = round(
                _amount(receivable.get(sc.COL_RECV_AMOUNT_PAID))
                + _amount(receivable.get(sc.COL_RECV_CREDITS_ADJ)), 2
            )
            row[sc.COL_TIME_INVOICE_BALANCE_DUE] = receivable.get(sc.COL_RECV_BALANCE_DUE)
            row[sc.COL_TIME_INVOICE_DATE] = receivable.get(sc.COL_RECV_DATE)

    @staticmethod
    def _ensure_setoff_account(rows: list[dict[str, Any]]) -> None:
        if any(clean_text(row.get(sc.COL_TXN_ACCOUNT_CODE)).casefold() == SETOFF_ACCOUNT.casefold() for row in rows):
            return
        rows.append({
            sc.COL_TXN_ACCOUNT_CODE: SETOFF_ACCOUNT,
            sc.COL_TXN_ACCOUNT_NAME: SETOFF_ACCOUNT_NAME,
            sc.COL_TXN_ACCOUNT_KIND: "clearing",
            sc.COL_TXN_ACCOUNT_OWNER: "Business",
            sc.COL_TXN_ACCOUNT_ACTIVE: 1,
            sc.COL_TXN_ACCOUNT_ALIASES: "A/R set-off; set-off clearing",
        })

    @staticmethod
    def _ensure_setoff_category(rows: list[dict[str, Any]]) -> None:
        """Expose reconstructed historical allocations without a bank category."""

        if any(
            clean_text(row.get(sc.COL_TXN_CATEGORY_LKP_CODE)).casefold()
            == SETOFF_CATEGORY_CODE.casefold()
            for row in rows
        ):
            return
        rows.append({
            sc.COL_TXN_CATEGORY_LKP_CODE: SETOFF_CATEGORY_CODE,
            sc.COL_TXN_CATEGORY_LKP_NAME: SETOFF_CATEGORY_NAME,
            sc.COL_TXN_CATEGORY_LKP_TYPE: "Transfer",
            sc.COL_TXN_CATEGORY_LKP_CLASS_SCOPE: "Business",
            sc.COL_TXN_CATEGORY_LKP_TAX_FLAG_DEFAULT: "None",
            sc.COL_TXN_CATEGORY_LKP_BILLABLE_ALLOWED: 0,
            sc.COL_TXN_CATEGORY_LKP_MEDICAL_ELIGIBLE: 0,
            sc.COL_TXN_CATEGORY_LKP_DEDUCTIBLE_ELIGIBLE: 0,
            sc.COL_TXN_CATEGORY_LKP_BUSINESS_DEDUCTIBLE_ELIGIBLE: 0,
            sc.COL_TXN_CATEGORY_LKP_ACTIVE: 1,
            sc.COL_TXN_CATEGORY_LKP_SORT_ORDER: 9999,
            sc.COL_TXN_CATEGORY_LKP_NOTES: "Non-cash A/R to A/P settlement set-off.",
        })

    @with_db_lock
    def reconstruct_historic_setoff(self, payload: Mapping[str, Any]) -> dict[str, Any]:
        """Restore one legacy set-off where receivables were imported but A/P was not.

        This is deliberately narrower than :meth:`record`: it verifies that
        each receivable already reflects the supplied allocation, then creates
        the missing governed A/P payment and ledger evidence without applying
        the allocation a second time.  The supplied legacy transaction rows
        are converted from false bank transfers into explicit A/R--A/P
        clearing transfers so no CIBC/AMEX/merchant details remain.
        """

        self.repo.ensure_schema()
        data = dict(payload or {})
        bill_id = clean_text(data.get("APBillID"))
        payment_id = clean_text(data.get("APPaymentID"))
        reference = clean_text(data.get("Reference"))
        payment_date = clean_text(data.get("PaymentDate"))
        amount = _amount(data.get("Amount"), "set-off amount")
        allocations = self._allocations(data)
        legacy_ids = [clean_text(value) for value in list(data.get("LegacyTransactionIDs") or [])]
        legacy_ids = [value for value in legacy_ids if value]
        if not bill_id or not payment_id or not reference:
            raise APValidationError("A/P bill, set-off payment ID, and reference are required.")
        if not legacy_ids or len(set(ident.casefold() for ident in legacy_ids)) != len(legacy_ids):
            raise APValidationError("Each reconstructed set-off needs one unique legacy transaction ID per allocation.")
        if len(legacy_ids) != len(allocations):
            raise APValidationError("Each reconstructed set-off allocation must name its legacy transaction ID.")
        if amount <= 0 or sum(item["amount"] for item in allocations) != amount:
            raise APValidationError("Set-off allocations must total the A/P payment amount exactly.")
        try:
            date.fromisoformat(payment_date)
        except Exception as exc:
            raise APValidationError("Set-off date must be YYYY-MM-DD.") from exc

        path = self.repo.paths.workbook_path()
        wb = load_workbook(path, keep_vba=True, data_only=False)
        saved = False
        try:
            bills = _read_rows(wb, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS)
            payments = _read_rows(wb, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS)
            receivables = self._read_finance(wb, TBL_RECEIVABLES)
            ledger = self._read_finance(wb, TBL_LEDGER)
            transactions = self._read_finance(wb, TBL_TRANSACTIONS_MASTER)
            accounts = self._read_finance(wb, TBL_TRANSACTION_ACCOUNTS)
            categories = self._read_finance(wb, TBL_TRANSACTION_CATEGORIES)

            if self._find(payments, "APPaymentID", payment_id):
                raise APValidationError(f"Duplicate A/P payment ID: {payment_id}")
            bill = self._find(bills, "APBillID", bill_id)
            if not bill:
                raise APValidationError("Selected A/P bill was not found.")
            if clean_text(bill.get("Status")).casefold() not in {"unpaid", "draft"}:
                raise APValidationError("The settlement A/P bill has already been posted or cannot be reconstructed.")
            if _amount(bill.get("AmountPaid"), "A/P bill amount paid") != Decimal("0.00"):
                raise APValidationError("The settlement A/P bill already has payment activity.")
            if _amount(bill.get("Total"), "A/P bill total") != amount:
                raise APValidationError("The supplied set-off amount does not equal the settlement A/P bill total.")
            linked_txn = self._find(transactions, sc.COL_TXN_ID, bill.get("ExpenseTransactionID"))
            if not linked_txn:
                raise APValidationError("The settlement A/P bill has no linked expense transaction to clear.")

            transaction_by_id = {
                clean_text(row.get(sc.COL_TXN_ID)).casefold(): row for row in transactions
            }
            receivable_by_invoice = {
                clean_text(row.get(sc.COL_RECV_INVOICE_NUM)).casefold(): row for row in receivables
            }
            for allocation, legacy_id in zip(allocations, legacy_ids):
                receivable = receivable_by_invoice.get(allocation["invoice"].casefold())
                if not receivable:
                    raise APValidationError(f"Invoice {allocation['invoice']} is missing from receivables.")
                # These receivables were already reduced by the legacy import.
                # Requiring an exact match prevents this repair from silently
                # combining an unrelated receipt with a settlement allocation.
                if _amount(receivable.get(sc.COL_RECV_AMOUNT_PAID), "invoice amount paid") != allocation["amount"]:
                    raise APValidationError(
                        f"Invoice {allocation['invoice']} does not have the expected historic set-off amount."
                    )
                legacy = transaction_by_id.get(legacy_id.casefold())
                if not legacy:
                    raise APValidationError(f"Legacy transaction {legacy_id} was not found.")
                if clean_text(legacy.get(sc.COL_TXN_INVOICE_REF)).casefold() != allocation["invoice"].casefold():
                    raise APValidationError(f"Legacy transaction {legacy_id} does not belong to invoice {allocation['invoice']}.")
                if _amount(legacy.get(sc.COL_TXN_AMOUNT), "legacy transaction amount") != allocation["amount"]:
                    raise APValidationError(f"Legacy transaction {legacy_id} does not match its set-off allocation amount.")
                if clean_text(legacy.get(sc.COL_TXN_TYPE)).casefold() != "transfer" or "set-off" not in clean_text(legacy.get(sc.COL_TXN_NOTES)).casefold():
                    raise APValidationError(f"Legacy transaction {legacy_id} is not a recognized historic set-off transfer.")

            allocation_json = json.dumps(
                [{"invoice": item["invoice"], "amount": float(item["amount"])} for item in allocations],
                separators=(",", ":"),
            )
            payment_notes = self._note(
                clean_text(data.get("Notes")),
                ALLOCATION_MARKER + allocation_json,
            )
            payments.append({
                "APPaymentID": payment_id, "APBillID": bill_id, "PaymentDate": payment_date,
                "Amount": float(amount), "FromAccount": SETOFF_ACCOUNT, "Method": SETOFF_METHOD,
                "Reference": reference, "Status": "Posted", "ReversalOfPaymentID": "",
                "ReversalReason": "", "Notes": payment_notes, "CreatedAt": _stamp(), "UpdatedAt": _stamp(),
            })
            snapshot = build_bill_snapshot(
                bill_id=bill.get("APBillID"), vendor=bill.get("Vendor"),
                vendor_invoice_number=bill.get("VendorInvoiceNumber"), invoice_date=bill.get("InvoiceDate"),
                subtotal=bill.get("Subtotal"), tax_amount=bill.get("TaxAmount"),
                payments=self._payment_lifecycle([
                    row for row in payments
                    if clean_text(row.get("APBillID")).casefold() == bill_id.casefold()
                ]),
            )
            bill.update({
                "AmountPaid": float(snapshot.total_paid), "Balance": float(snapshot.balance),
                "Status": snapshot.status.value, "UpdatedAt": _stamp(),
            })
            for index, allocation in enumerate(allocations, start=1):
                receivable = receivable_by_invoice[allocation["invoice"].casefold()]
                ledger.append({
                    sc.COL_LEDGER_ID: f"LED-SET-{payment_id}-{index}", sc.COL_LEDGER_DATE: payment_date,
                    sc.COL_LEDGER_CLIENT_VENDOR: receivable.get(sc.COL_RECV_CLIENT, ""),
                    sc.COL_LEDGER_DESCRIPTION: f"Set-off applied to invoice {allocation['invoice']} (A/P {bill_id})",
                    sc.COL_LEDGER_CATEGORY: SETOFF_METHOD, sc.COL_LEDGER_REFERENCE: allocation["invoice"],
                    sc.COL_LEDGER_BILLINGS_EXCL_HST: 0.0, sc.COL_LEDGER_HST_COLLECTED: 0.0,
                    sc.COL_LEDGER_EXPENSES_EXCL_HST: 0.0, sc.COL_LEDGER_HST_PAID: 0.0,
                    sc.COL_LEDGER_COLLECTED: float(allocation["amount"]), sc.COL_LEDGER_WRITE_OFF: 0.0,
                    sc.COL_LEDGER_RECEIVABLE: -float(allocation["amount"]), sc.COL_LEDGER_TRX_ID: payment_id,
                    sc.COL_LEDGER_EXTERNAL_REF_ID: payment_id, sc.COL_LEDGER_ORIGINAL_AMOUNT: float(allocation["amount"]),
                    sc.COL_LEDGER_WORK_CLIENT: receivable.get(sc.COL_RECV_WORK_CLIENT, ""), sc.COL_LEDGER_CREATED_AT: _stamp(),
                })
                legacy = transaction_by_id[legacy_ids[index - 1].casefold()]
                legacy.update({
                    sc.COL_TXN_FROM_ACCOUNT: SETOFF_ACCOUNT,
                    sc.COL_TXN_TO_ACCOUNT: "AP_PAYABLE",
                    sc.COL_TXN_PAYEE: clean_text(bill.get("Vendor")),
                    sc.COL_TXN_CATEGORY_CODE: SETOFF_CATEGORY_CODE,
                    sc.COL_TXN_CATEGORY_NAME: SETOFF_CATEGORY_NAME,
                    sc.COL_TXN_TAX_AMOUNT: 0.0,
                    sc.COL_TXN_TAX_FLAG: "None",
                    sc.COL_TXN_HST_EXEMPT: 1,
                    sc.COL_TXN_GENERAL_OFFICE_EXPENSE: 0,
                    sc.COL_TXN_NOTES: (
                        f"Set-off allocation for invoice {allocation['invoice']} against "
                        f"LIHDC settlement {payment_id}. No bank movement."
                    ),
                    sc.COL_TXN_STATUS: "Cleared",
                    sc.COL_TXN_CLEARED_AT: payment_date,
                    sc.COL_TXN_UPDATED_AT: _stamp(),
                })
            linked_txn[sc.COL_TXN_FROM_ACCOUNT] = SETOFF_ACCOUNT
            linked_txn[sc.COL_TXN_STATUS] = "Cleared"
            linked_txn[sc.COL_TXN_CLEARED_AT] = payment_date
            linked_txn[sc.COL_TXN_NOTES] = self._note(
                linked_txn.get(sc.COL_TXN_NOTES), f"Set-off {payment_id}: {reference}"
            )
            linked_txn[sc.COL_TXN_UPDATED_AT] = _stamp()
            self._ensure_setoff_account(accounts)
            self._ensure_setoff_category(categories)
            _replace_rows(wb, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS, bills)
            _replace_rows(wb, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS, payments)
            for tref, rows in (
                (TBL_LEDGER, ledger), (TBL_TRANSACTIONS_MASTER, transactions),
                (TBL_TRANSACTION_ACCOUNTS, accounts), (TBL_TRANSACTION_CATEGORIES, categories),
            ):
                self._write_finance(wb, tref, rows)
            self.repo._safe_save(wb, path)
            saved = True
        finally:
            if not saved:
                self.repo._close_workbook(wb)
        return {
            "ok": True, "APPaymentID": payment_id, "APBillID": bill_id,
            "allocationCount": len(allocations),
            "message": "Historic settlement set-off was reconstructed across A/P, ledger, and transaction records.",
        }

    @with_db_lock
    def retire_superseded_historic_settlement_expense(
        self,
        bill_id: str,
        payment_id: str,
        legacy_transaction_id: str,
    ) -> dict[str, Any]:
        """Void an imported bank-labelled settlement expense superseded by A/P.

        The governed A/P bill is the sole expense authority once its set-off
        payment exists. Keeping the old imported expense active would both
        display fictitious banking details and double-count the settlement.
        """

        self.repo.ensure_schema()
        bill_id = clean_text(bill_id)
        payment_id = clean_text(payment_id)
        legacy_transaction_id = clean_text(legacy_transaction_id)
        if not bill_id or not payment_id or not legacy_transaction_id:
            raise APValidationError("Settlement bill, payment, and legacy transaction IDs are required.")
        path = self.repo.paths.workbook_path()
        wb = load_workbook(path, keep_vba=True, data_only=False)
        saved = False
        try:
            bills = _read_rows(wb, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS)
            payments = _read_rows(wb, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS)
            transactions = self._read_finance(wb, TBL_TRANSACTIONS_MASTER)
            bill = self._find(bills, "APBillID", bill_id)
            payment = self._find(payments, "APPaymentID", payment_id)
            legacy = self._find(transactions, sc.COL_TXN_ID, legacy_transaction_id)
            if not bill or not payment or not legacy:
                raise APValidationError("The governed LIHDC settlement evidence is incomplete.")
            if clean_text(payment.get("APBillID")).casefold() != bill_id.casefold():
                raise APValidationError("The supplied set-off payment belongs to a different A/P bill.")
            if clean_text(payment.get("Method")).casefold() != SETOFF_METHOD.casefold():
                raise APValidationError("The supplied A/P payment is not a set-off.")
            if clean_text(bill.get("Status")).casefold() != "paid":
                raise APValidationError("The governed settlement A/P bill is not posted as paid.")
            existing_reason = clean_text(legacy.get(sc.COL_TXN_VOID_REASON))
            if (
                clean_text(legacy.get(sc.COL_TXN_STATUS)).casefold() == "void"
                and payment_id.casefold() in existing_reason.casefold()
            ):
                return {
                    "ok": True, "transactionId": legacy_transaction_id,
                    "message": "Superseded historic settlement expense was already voided.",
                }
            legacy_note = clean_text(legacy.get(sc.COL_TXN_NOTES)).casefold()
            legacy_invoice_ref = clean_text(legacy.get(sc.COL_TXN_INVOICE_REF)).casefold()
            if "settlement" not in legacy_note and "settlement" not in legacy_invoice_ref:
                raise APValidationError("Legacy transaction is not identified as the settlement expense.")
            legacy.update({
                sc.COL_TXN_TYPE: "Expense",
                sc.COL_TXN_FROM_ACCOUNT: SETOFF_ACCOUNT,
                sc.COL_TXN_TO_ACCOUNT: "",
                sc.COL_TXN_PAYEE: clean_text(bill.get("Vendor")),
                sc.COL_TXN_PARENT: "",
                sc.COL_TXN_CLIENT: "",
                sc.COL_TXN_MATTER: "",
                sc.COL_TXN_CATEGORY_CODE: clean_text(bill.get("CategoryCode")) or "EXP_LEGAL_FEES",
                sc.COL_TXN_CATEGORY_NAME: clean_text(bill.get("CategoryName")) or "Legal Fees Expense",
                sc.COL_TXN_TAX_FLAG: "Business Deductible",
                sc.COL_TXN_HST_EXEMPT: 0,
                sc.COL_TXN_GENERAL_OFFICE_EXPENSE: 1,
                sc.COL_TXN_STATUS: "Void",
                sc.COL_TXN_VOID_REASON: (
                    f"Superseded by governed LIHDC settlement A/P bill {bill_id} "
                    f"and set-off payment {payment_id}; no bank movement."
                ),
                sc.COL_TXN_NOTES: (
                    f"Historic settlement expense superseded by {payment_id}. "
                    "No bank movement; use the governed A/P settlement record."
                ),
                sc.COL_TXN_UPDATED_AT: _stamp(),
            })
            self._write_finance(wb, TBL_TRANSACTIONS_MASTER, transactions)
            self.repo._safe_save(wb, path)
            saved = True
        finally:
            if not saved:
                self.repo._close_workbook(wb)
        return {
            "ok": True, "transactionId": legacy_transaction_id,
            "message": "Superseded historic settlement expense was voided in favour of governed A/P evidence.",
        }

    @with_db_lock
    def record(self, payload: Mapping[str, Any]) -> dict[str, Any]:
        self.repo.ensure_schema()
        data = dict(payload or {})
        bill_id = clean_text(data.get("APBillID"))
        payment_id = clean_text(data.get("APPaymentID"))
        reference = clean_text(data.get("Reference"))
        payment_date = clean_text(data.get("PaymentDate"))
        amount = _amount(data.get("Amount"), "set-off amount")
        allocations = self._allocations(data)
        if not bill_id or not payment_id or not reference:
            raise APValidationError("A/P bill, set-off payment ID, and reference are required.")
        if amount <= 0 or sum(item["amount"] for item in allocations) != amount:
            raise APValidationError("Set-off allocations must total the A/P payment amount exactly.")
        try:
            date.fromisoformat(payment_date)
        except Exception as exc:
            raise APValidationError("Set-off date must be YYYY-MM-DD.") from exc

        path = self.repo.paths.workbook_path()
        wb = load_workbook(path, keep_vba=True, data_only=False)
        saved = False
        try:
            bills = _read_rows(wb, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS)
            payments = _read_rows(wb, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS)
            receivables = self._read_finance(wb, TBL_RECEIVABLES)
            ledger = self._read_finance(wb, TBL_LEDGER)
            transactions = self._read_finance(wb, TBL_TRANSACTIONS_MASTER)
            accounts = self._read_finance(wb, TBL_TRANSACTION_ACCOUNTS)
            time_rows = self._read_finance(wb, TBL_TIME)
            if self._find(payments, "APPaymentID", payment_id):
                raise APValidationError(f"Duplicate A/P payment ID: {payment_id}")
            bill = self._find(bills, "APBillID", bill_id)
            if not bill:
                raise APValidationError("Selected A/P bill was not found.")
            if clean_text(bill.get("Status")).casefold() in {"paid", "voided", "reversed"}:
                raise APValidationError("Selected A/P bill is not available for set-off.")
            linked_txn = self._find(transactions, sc.COL_TXN_ID, bill.get("ExpenseTransactionID"))
            if not linked_txn:
                raise APValidationError("The A/P bill has no linked expense transaction to clear.")
            related = [row for row in payments if clean_text(row.get("APBillID")).casefold() == bill_id.casefold()]
            snapshot = build_bill_snapshot(
                bill_id=bill.get("APBillID"), vendor=bill.get("Vendor"),
                vendor_invoice_number=bill.get("VendorInvoiceNumber"), invoice_date=bill.get("InvoiceDate"),
                subtotal=bill.get("Subtotal"), tax_amount=bill.get("TaxAmount"),
                payments=self._payment_lifecycle(related + [{"APPaymentID": payment_id, "Amount": amount, "Status": "Posted"}]),
            )
            touched: list[dict[str, Any]] = []
            for allocation in allocations:
                row = self._find(receivables, sc.COL_RECV_INVOICE_NUM, allocation["invoice"])
                if not row or clean_text(row.get(sc.COL_RECV_STATUS)).casefold() in {"void", "closed", "paid"}:
                    raise APValidationError(f"Invoice {allocation['invoice']} is not open for set-off.")
                if allocation["amount"] > _amount(row.get(sc.COL_RECV_BALANCE_DUE), "invoice balance"):
                    raise APValidationError(f"Allocation exceeds the open balance of invoice {allocation['invoice']}.")
                touched.append(row)

            allocation_json = json.dumps(
                [{"invoice": item["invoice"], "amount": float(item["amount"])} for item in allocations], separators=(",", ":")
            )
            note = clean_text(data.get("Notes"))
            payment_notes = self._note(note, ALLOCATION_MARKER + allocation_json)
            payments.append({
                "APPaymentID": payment_id, "APBillID": bill_id, "PaymentDate": payment_date,
                "Amount": float(amount), "FromAccount": SETOFF_ACCOUNT, "Method": SETOFF_METHOD,
                "Reference": reference, "Status": "Posted", "ReversalOfPaymentID": "",
                "ReversalReason": "", "Notes": payment_notes, "CreatedAt": _stamp(), "UpdatedAt": _stamp(),
            })
            bill.update({"AmountPaid": float(snapshot.total_paid), "Balance": float(snapshot.balance),
                         "Status": snapshot.status.value, "UpdatedAt": _stamp()})
            for index, allocation in enumerate(allocations, start=1):
                row = touched[index - 1]
                balance, status = self._set_receivable(row, allocation["amount"])
                self._sync_time(time_rows, row)
                ledger.append({
                    sc.COL_LEDGER_ID: f"LED-SET-{payment_id}-{index}", sc.COL_LEDGER_DATE: payment_date,
                    sc.COL_LEDGER_CLIENT_VENDOR: row.get(sc.COL_RECV_CLIENT, ""),
                    sc.COL_LEDGER_DESCRIPTION: f"Set-off applied to invoice {allocation['invoice']} (A/P {bill_id})",
                    sc.COL_LEDGER_CATEGORY: SETOFF_METHOD, sc.COL_LEDGER_REFERENCE: reference,
                    sc.COL_LEDGER_BILLINGS_EXCL_HST: 0.0, sc.COL_LEDGER_HST_COLLECTED: 0.0,
                    sc.COL_LEDGER_EXPENSES_EXCL_HST: 0.0, sc.COL_LEDGER_HST_PAID: 0.0,
                    sc.COL_LEDGER_COLLECTED: float(allocation["amount"]), sc.COL_LEDGER_WRITE_OFF: 0.0,
                    sc.COL_LEDGER_RECEIVABLE: -float(allocation["amount"]), sc.COL_LEDGER_TRX_ID: payment_id,
                    sc.COL_LEDGER_EXTERNAL_REF_ID: payment_id, sc.COL_LEDGER_ORIGINAL_AMOUNT: float(allocation["amount"]),
                    sc.COL_LEDGER_WORK_CLIENT: row.get(sc.COL_RECV_WORK_CLIENT, ""), sc.COL_LEDGER_CREATED_AT: _stamp(),
                })
            linked_txn[sc.COL_TXN_FROM_ACCOUNT] = SETOFF_ACCOUNT
            linked_txn[sc.COL_TXN_STATUS] = "Cleared"
            linked_txn[sc.COL_TXN_CLEARED_AT] = payment_date
            linked_txn[sc.COL_TXN_NOTES] = self._note(linked_txn.get(sc.COL_TXN_NOTES), f"Set-off {payment_id}: {reference}")
            linked_txn[sc.COL_TXN_UPDATED_AT] = _stamp()
            self._ensure_setoff_account(accounts)
            _replace_rows(wb, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS, bills)
            _replace_rows(wb, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS, payments)
            for tref, rows in ((TBL_RECEIVABLES, receivables), (TBL_LEDGER, ledger),
                               (TBL_TRANSACTIONS_MASTER, transactions), (TBL_TRANSACTION_ACCOUNTS, accounts),
                               (TBL_TIME, time_rows)):
                self._write_finance(wb, tref, rows)
            self.repo._safe_save(wb, path)
            saved = True
        finally:
            if not saved:
                self.repo._close_workbook(wb)
        return {"ok": True, "APPaymentID": payment_id, "APBillID": bill_id,
                "allocationCount": len(allocations), "message": "Set-off posted across A/P and receivables."}

    @with_db_lock
    def reverse(self, payment_id: str, reversal_id: str, reason: str) -> dict[str, Any]:
        payment_id, reversal_id, reason = clean_text(payment_id), clean_text(reversal_id), clean_text(reason)
        if not payment_id or not reversal_id or not reason:
            raise APValidationError("Set-off payment, reversal ID, and reason are required.")
        self.repo.ensure_schema()
        path = self.repo.paths.workbook_path()
        wb = load_workbook(path, keep_vba=True, data_only=False)
        saved = False
        try:
            bills = _read_rows(wb, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS)
            payments = _read_rows(wb, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS)
            receivables, ledger = self._read_finance(wb, TBL_RECEIVABLES), self._read_finance(wb, TBL_LEDGER)
            transactions, time_rows = self._read_finance(wb, TBL_TRANSACTIONS_MASTER), self._read_finance(wb, TBL_TIME)
            original = self._find(payments, "APPaymentID", payment_id)
            if not original or clean_text(original.get("Method")).casefold() != SETOFF_METHOD.casefold():
                raise APValidationError("Selected payment is not a set-off payment.")
            if clean_text(original.get("Status")).casefold() in {"reversed", "reversal"} or self._find(payments, "APPaymentID", reversal_id):
                raise APValidationError("This set-off has already been reversed or the reversal ID is duplicate.")
            allocations = self._stored_allocations(original.get("Notes"))
            bill = self._find(bills, "APBillID", original.get("APBillID"))
            if not bill:
                raise APValidationError("The original A/P bill was not found.")
            for item in allocations:
                if not self._find(receivables, sc.COL_RECV_INVOICE_NUM, item["invoice"]):
                    raise APValidationError(f"Invoice {item['invoice']} is missing; no reversal was posted.")
            original["Status"], original["UpdatedAt"] = "Reversed", _stamp()
            payments.append({"APPaymentID": reversal_id, "APBillID": bill.get("APBillID"), "PaymentDate": date.today().isoformat(),
                             "Amount": original.get("Amount"), "FromAccount": SETOFF_ACCOUNT, "Method": SETOFF_METHOD,
                             "Reference": original.get("Reference"), "Status": "Reversal", "ReversalOfPaymentID": payment_id,
                             "ReversalReason": reason, "Notes": f"Reversal of set-off {payment_id}: {reason}", "CreatedAt": _stamp(), "UpdatedAt": _stamp()})
            related = [row for row in payments if clean_text(row.get("APBillID")).casefold() == clean_text(bill.get("APBillID")).casefold()]
            snapshot = build_bill_snapshot(bill_id=bill.get("APBillID"), vendor=bill.get("Vendor"), vendor_invoice_number=bill.get("VendorInvoiceNumber"), invoice_date=bill.get("InvoiceDate"), subtotal=bill.get("Subtotal"), tax_amount=bill.get("TaxAmount"), payments=self._payment_lifecycle(related))
            bill.update({"AmountPaid": float(snapshot.total_paid), "Balance": float(snapshot.balance), "Status": snapshot.status.value, "UpdatedAt": _stamp()})
            for index, item in enumerate(allocations, start=1):
                row = self._find(receivables, sc.COL_RECV_INVOICE_NUM, item["invoice"])
                self._set_receivable(row, -item["amount"])
                self._sync_time(time_rows, row)
                ledger.append({sc.COL_LEDGER_ID: f"LED-REV-{reversal_id}-{index}", sc.COL_LEDGER_DATE: date.today().isoformat(), sc.COL_LEDGER_CLIENT_VENDOR: row.get(sc.COL_RECV_CLIENT, ""), sc.COL_LEDGER_DESCRIPTION: f"Reversal of set-off {payment_id} for invoice {item['invoice']}: {reason}", sc.COL_LEDGER_CATEGORY: SETOFF_METHOD, sc.COL_LEDGER_REFERENCE: original.get("Reference", ""), sc.COL_LEDGER_BILLINGS_EXCL_HST: 0.0, sc.COL_LEDGER_HST_COLLECTED: 0.0, sc.COL_LEDGER_EXPENSES_EXCL_HST: 0.0, sc.COL_LEDGER_HST_PAID: 0.0, sc.COL_LEDGER_COLLECTED: -float(item["amount"]), sc.COL_LEDGER_WRITE_OFF: 0.0, sc.COL_LEDGER_RECEIVABLE: float(item["amount"]), sc.COL_LEDGER_TRX_ID: reversal_id, sc.COL_LEDGER_EXTERNAL_REF_ID: payment_id, sc.COL_LEDGER_ORIGINAL_AMOUNT: -float(item["amount"]), sc.COL_LEDGER_WORK_CLIENT: row.get(sc.COL_RECV_WORK_CLIENT, ""), sc.COL_LEDGER_CREATED_AT: _stamp()})
            linked = self._find(transactions, sc.COL_TXN_ID, bill.get("ExpenseTransactionID"))
            if linked and _amount(bill.get("AmountPaid")) <= 0:
                linked[sc.COL_TXN_FROM_ACCOUNT], linked[sc.COL_TXN_STATUS], linked[sc.COL_TXN_CLEARED_AT] = "AP_PAYABLE", "Pending", ""
                linked[sc.COL_TXN_NOTES] = self._note(linked.get(sc.COL_TXN_NOTES), f"Set-off {payment_id} reversed: {reason}")
                linked[sc.COL_TXN_UPDATED_AT] = _stamp()
            _replace_rows(wb, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS, bills)
            _replace_rows(wb, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS, payments)
            for tref, rows in ((TBL_RECEIVABLES, receivables), (TBL_LEDGER, ledger), (TBL_TRANSACTIONS_MASTER, transactions), (TBL_TIME, time_rows)):
                self._write_finance(wb, tref, rows)
            self.repo._safe_save(wb, path)
            saved = True
        finally:
            if not saved:
                self.repo._close_workbook(wb)
        return {"ok": True, "APPaymentID": reversal_id, "ReversalOfPaymentID": payment_id,
                "message": "Set-off reversal posted across A/P and receivables."}
