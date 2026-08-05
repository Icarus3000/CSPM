from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import hashlib
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import yaml
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.datetime import from_excel

from domain.money import calc_amounts, normalize_pct
from services.paths import AppPaths


Severity = str


@dataclass(frozen=True)
class WorkbookIntegrityIssue:
    severity: Severity
    code: str
    message: str
    sheet: str = ""
    table: str = ""
    row: Optional[int] = None
    column: str = ""
    value: str = ""

    def as_dict(self) -> Dict[str, Any]:
        payload: Dict[str, Any] = {
            "severity": self.severity,
            "code": self.code,
            "message": self.message,
        }
        if self.sheet:
            payload["sheet"] = self.sheet
        if self.table:
            payload["table"] = self.table
        if self.row is not None:
            payload["row"] = self.row
        if self.column:
            payload["column"] = self.column
        if self.value:
            payload["value"] = self.value
        return payload


@dataclass
class WorkbookIntegrityReport:
    workbook_path: str
    schema_path: str
    workbook_sha256: str = ""
    ok: bool = False
    tables_checked: int = 0
    rows_checked: int = 0
    financial_totals: Dict[str, float] = field(default_factory=dict)
    issues: List[WorkbookIntegrityIssue] = field(default_factory=list)

    @property
    def error_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == "error")

    @property
    def warning_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == "warning")

    def as_dict(self) -> Dict[str, Any]:
        return {
            "ok": self.ok,
            "workbookPath": self.workbook_path,
            "schemaPath": self.schema_path,
            "workbookSha256": self.workbook_sha256,
            "summary": {
                "tablesChecked": self.tables_checked,
                "rowsChecked": self.rows_checked,
                "errorCount": self.error_count,
                "warningCount": self.warning_count,
                "financialTotals": self.financial_totals,
            },
            "issues": [issue.as_dict() for issue in self.issues],
        }

    def to_json(self) -> str:
        return json.dumps(self.as_dict(), ensure_ascii=False, indent=2)


@dataclass
class _TableRead:
    sheet: str
    table: str
    primary_key: str
    expected_columns: List[str]
    headers: List[str]
    rows: List[Dict[str, Any]]
    row_numbers: List[int]


class WorkbookIntegrityService:
    """
    Read-only health check for the Excel-first CSPM database.

    The checker deliberately does not call ExcelRepo.ensure_schema(); it should
    report the live workbook's condition without repairing it first.
    """

    MIN_YEAR = 1900
    MAX_YEAR = 2100
    MONEY_TOLERANCE = 0.02

    REQUIRED_FIELDS_BY_TABLE: Dict[str, List[str]] = {
        "tblParents": ["ParentID", "ParentName"],
        "tblClients": ["ClientID", "ClientName"],
        "tblClientProfiles": ["ClientID"],
        "tblMatters": ["MatterID", "MatterNumber", "ClientID", "MatterName"],
        "tblTimeEntries": ["EntryID", "Date", "MatterID", "Description"],
        "tblTrademarks": ["TrademarkID", "TrademarkText"],
        "tblTransactionsMaster": ["TransactionID", "TxnDate", "Type", "Amount"],
        "tblTransactionAccounts": ["AccountCode", "AccountName"],
        "tblTransactionCategories": ["CategoryCode", "CategoryName", "Type"],
        "tblTransactionBusinessUnits": ["BusinessUnit"],
        "tblTransactionPayees": ["PayeeName"],
    }

    UNIQUE_FIELDS_BY_TABLE: Dict[str, List[str]] = {
        "tblMatters": ["MatterNumber"],
        "tblTrademarks": ["ApplicationNumber", "RegistrationNumber"],
    }

    def __init__(self, paths: AppPaths):
        self.paths = paths

    def check(
        self,
        workbook_path: Optional[Path] = None,
        schema_path: Optional[Path] = None,
    ) -> WorkbookIntegrityReport:
        schema_file = Path(schema_path) if schema_path else self.paths.workbook_schema_path()
        workbook_file = Path(workbook_path) if workbook_path else self.paths.workbook_path()
        report = WorkbookIntegrityReport(
            workbook_path=str(workbook_file),
            schema_path=str(schema_file),
        )

        schema = self._load_schema(schema_file, report)
        if not workbook_file.exists():
            self._add(report, "error", "missing_workbook", f"Workbook not found: {workbook_file}")
            report.ok = False
            return report

        report.workbook_sha256 = self._sha256(workbook_file)

        if not schema:
            report.ok = False
            return report

        table_specs = self._table_specs(schema, report)
        if not table_specs:
            self._add(report, "error", "schema_has_no_tables", "Workbook schema has no table definitions.")
            report.ok = False
            return report

        try:
            workbook = load_workbook(workbook_file, keep_vba=True, data_only=False)
        except Exception as exc:
            self._add(
                report,
                "error",
                "workbook_open_failed",
                f"Workbook could not be opened by openpyxl: {exc}",
            )
            report.ok = False
            return report

        tables: Dict[str, _TableRead] = {}
        try:
            for spec in table_specs:
                table_read = self._check_table(workbook, spec, report)
                if table_read is not None:
                    tables[table_read.table] = table_read
                    report.tables_checked += 1
                    report.rows_checked += len(table_read.rows)

            self._check_references(tables, report)
            self._check_financial_totals(tables, report)
            self._check_date_relationships(tables, report)
        finally:
            self._close_workbook(workbook)

        report.ok = report.error_count == 0
        return report

    def _load_schema(self, schema_path: Path, report: WorkbookIntegrityReport) -> Dict[str, Any]:
        if not schema_path.exists():
            self._add(report, "error", "missing_schema", f"Workbook schema not found: {schema_path}")
            return {}
        try:
            payload = yaml.safe_load(schema_path.read_text(encoding="utf-8"))
        except Exception as exc:
            self._add(report, "error", "schema_read_failed", f"Workbook schema could not be read: {exc}")
            return {}
        if not isinstance(payload, dict):
            self._add(report, "error", "schema_invalid", "Workbook schema did not parse to an object.")
            return {}
        return payload

    def _table_specs(self, schema: Dict[str, Any], report: WorkbookIntegrityReport) -> List[Dict[str, Any]]:
        raw_tables = schema.get("tables")
        if not isinstance(raw_tables, list):
            self._add(report, "error", "schema_tables_invalid", "Schema 'tables' must be a list.")
            return []
        specs: List[Dict[str, Any]] = []
        for raw in raw_tables:
            if not isinstance(raw, dict):
                self._add(report, "error", "schema_table_invalid", "A table entry is not an object.")
                continue
            sheet = str(raw.get("sheet", "") or "").strip()
            table = str(raw.get("table", "") or "").strip()
            primary_key = str(raw.get("primary_key", "") or "").strip()
            columns = raw.get("columns")
            if not sheet or not table or not primary_key or not isinstance(columns, list):
                self._add(
                    report,
                    "error",
                    "schema_table_incomplete",
                    f"Schema table entry is incomplete: sheet={sheet!r}, table={table!r}",
                )
                continue
            column_names = []
            for column in columns:
                if not isinstance(column, dict):
                    continue
                name = str(column.get("name", "") or "").strip()
                if name:
                    column_names.append(name)
            if primary_key not in column_names:
                self._add(
                    report,
                    "error",
                    "schema_primary_key_missing_from_columns",
                    f"Primary key '{primary_key}' is not listed as a column.",
                    sheet=sheet,
                    table=table,
                    column=primary_key,
                )
            specs.append(
                {
                    "sheet": sheet,
                    "table": table,
                    "primary_key": primary_key,
                    "columns": column_names,
                    "column_types": {
                        str(col.get("name", "") or "").strip(): str(col.get("type", "") or "").strip()
                        for col in columns
                        if isinstance(col, dict)
                    },
                }
            )
        return specs

    def _check_table(self, workbook: Any, spec: Dict[str, Any], report: WorkbookIntegrityReport) -> Optional[_TableRead]:
        sheet_name = str(spec["sheet"])
        table_name = str(spec["table"])
        expected_columns = list(spec["columns"])
        primary_key = str(spec["primary_key"])

        if sheet_name not in workbook.sheetnames:
            self._add(
                report,
                "error",
                "missing_sheet",
                f"Required sheet '{sheet_name}' is missing.",
                sheet=sheet_name,
                table=table_name,
            )
            return None

        worksheet = workbook[sheet_name]
        table_obj = worksheet.tables.get(table_name) if hasattr(worksheet, "tables") else None
        if table_obj is None:
            self._add(
                report,
                "error",
                "missing_table",
                f"Required Excel table '{table_name}' is missing from sheet '{sheet_name}'.",
                sheet=sheet_name,
                table=table_name,
            )
            header_row, min_col, max_col = self._find_header_row(worksheet, expected_columns)
            if header_row <= 0:
                return _TableRead(sheet_name, table_name, primary_key, expected_columns, [], [], [])
            max_row = int(worksheet.max_row or header_row)
        else:
            min_col, header_row, max_col, max_row = range_boundaries(table_obj.ref)

        headers = [
            self._clean(worksheet.cell(row=header_row, column=col_idx).value)
            for col_idx in range(min_col, max_col + 1)
        ]
        header_counts: Dict[str, int] = {}
        for header in headers:
            if not header:
                continue
            header_counts[header] = header_counts.get(header, 0) + 1
        for header, count in sorted(header_counts.items()):
            if count > 1:
                self._add(
                    report,
                    "error",
                    "duplicate_header",
                    f"Header '{header}' appears {count} times.",
                    sheet=sheet_name,
                    table=table_name,
                    row=header_row,
                    column=header,
                )

        header_to_col = {
            header: min_col + offset
            for offset, header in enumerate(headers)
            if header
        }
        for column in expected_columns:
            if column not in header_to_col:
                self._add(
                    report,
                    "error",
                    "missing_column",
                    f"Required column '{column}' is missing.",
                    sheet=sheet_name,
                    table=table_name,
                    column=column,
                )

        rows: List[Dict[str, Any]] = []
        row_numbers: List[int] = []
        for row_idx in range(header_row + 1, int(max_row or header_row) + 1):
            row = {
                column: worksheet.cell(row=row_idx, column=header_to_col[column]).value
                for column in expected_columns
                if column in header_to_col
            }
            if not self._row_has_data(row):
                continue
            rows.append(row)
            row_numbers.append(row_idx)

        table_read = _TableRead(
            sheet=sheet_name,
            table=table_name,
            primary_key=primary_key,
            expected_columns=expected_columns,
            headers=headers,
            rows=rows,
            row_numbers=row_numbers,
        )

        self._check_required_fields(table_read, report)
        self._check_primary_key(table_read, report)
        self._check_unique_fields(table_read, report)
        self._check_column_types(table_read, dict(spec["column_types"]), report)
        return table_read

    def _find_header_row(self, worksheet: Any, expected_columns: List[str]) -> Tuple[int, int, int]:
        if not expected_columns:
            return 0, 0, 0
        anchor = expected_columns[0]
        max_scan_row = min(max(int(worksheet.max_row or 1), 1), 50)
        max_scan_col = min(max(int(worksheet.max_column or 1), len(expected_columns)), 256)
        for row_idx in range(1, max_scan_row + 1):
            labels: Dict[str, int] = {}
            for col_idx in range(1, max_scan_col + 1):
                value = self._clean(worksheet.cell(row=row_idx, column=col_idx).value)
                if value and value not in labels:
                    labels[value] = col_idx
            if anchor in labels:
                present_cols = [labels[col] for col in expected_columns if col in labels]
                if not present_cols:
                    return row_idx, 1, max_scan_col
                return row_idx, min(present_cols), max(present_cols)
        return 0, 0, 0

    def _check_required_fields(self, table_read: _TableRead, report: WorkbookIntegrityReport) -> None:
        required = self.REQUIRED_FIELDS_BY_TABLE.get(table_read.table, [table_read.primary_key])
        for row, row_idx in zip(table_read.rows, table_read.row_numbers):
            for column in required:
                if column not in table_read.expected_columns:
                    continue
                if self._is_empty(row.get(column)):
                    self._add(
                        report,
                        "error",
                        "missing_required_value",
                        f"Required value '{column}' is blank.",
                        sheet=table_read.sheet,
                        table=table_read.table,
                        row=row_idx,
                        column=column,
                    )

    def _check_primary_key(self, table_read: _TableRead, report: WorkbookIntegrityReport) -> None:
        key = table_read.primary_key
        seen: Dict[str, int] = {}
        for row, row_idx in zip(table_read.rows, table_read.row_numbers):
            value = row.get(key)
            text = self._clean(value)
            if text.startswith("="):
                self._add(
                    report,
                    "error",
                    "primary_key_formula",
                    "Primary keys must be stable values, not formulas.",
                    sheet=table_read.sheet,
                    table=table_read.table,
                    row=row_idx,
                    column=key,
                    value=text,
                )
                continue
            if not text:
                continue
            if text in seen:
                self._add(
                    report,
                    "error",
                    "duplicate_primary_key",
                    f"Primary key '{text}' duplicates row {seen[text]}.",
                    sheet=table_read.sheet,
                    table=table_read.table,
                    row=row_idx,
                    column=key,
                    value=text,
                )
            else:
                seen[text] = row_idx

    def _check_unique_fields(self, table_read: _TableRead, report: WorkbookIntegrityReport) -> None:
        for column in self.UNIQUE_FIELDS_BY_TABLE.get(table_read.table, []):
            seen: Dict[str, int] = {}
            for row, row_idx in zip(table_read.rows, table_read.row_numbers):
                text = self._clean(row.get(column))
                if not text:
                    continue
                if text in seen:
                    self._add(
                        report,
                        "error",
                        "duplicate_sequence_value",
                        f"Sequence-like value '{text}' in '{column}' duplicates row {seen[text]}.",
                        sheet=table_read.sheet,
                        table=table_read.table,
                        row=row_idx,
                        column=column,
                        value=text,
                    )
                else:
                    seen[text] = row_idx

    def _check_column_types(
        self,
        table_read: _TableRead,
        column_types: Dict[str, str],
        report: WorkbookIntegrityReport,
    ) -> None:
        for row, row_idx in zip(table_read.rows, table_read.row_numbers):
            for column, expected_type in column_types.items():
                if column not in row or self._is_empty(row.get(column)):
                    continue
                value = row.get(column)
                if expected_type in {"number", "int"} and self._to_decimal(value) is None:
                    self._add(
                        report,
                        "error",
                        "invalid_number_value",
                        f"Expected a {expected_type} value in '{column}'.",
                        sheet=table_read.sheet,
                        table=table_read.table,
                        row=row_idx,
                        column=column,
                        value=self._safe_value(value),
                    )
                if expected_type == "bool-int" and not self._is_bool_int(value):
                    self._add(
                        report,
                        "error",
                        "invalid_bool_int_value",
                        f"Expected a 0/1 or true/false value in '{column}'.",
                        sheet=table_read.sheet,
                        table=table_read.table,
                        row=row_idx,
                        column=column,
                        value=self._safe_value(value),
                    )
                if expected_type in {"date", "datetime"}:
                    parsed = self._parse_date(value)
                    if parsed is None:
                        self._add(
                            report,
                            "error",
                            "invalid_date_value",
                            f"Expected a parseable {expected_type} value in '{column}'.",
                            sheet=table_read.sheet,
                            table=table_read.table,
                            row=row_idx,
                            column=column,
                            value=self._safe_value(value),
                        )
                    elif parsed.year < self.MIN_YEAR or parsed.year > self.MAX_YEAR:
                        self._add(
                            report,
                            "error",
                            "date_out_of_range",
                            f"Date in '{column}' is outside the supported year range {self.MIN_YEAR}-{self.MAX_YEAR}.",
                            sheet=table_read.sheet,
                            table=table_read.table,
                            row=row_idx,
                            column=column,
                            value=parsed.isoformat(),
                        )

    def _check_references(self, tables: Dict[str, _TableRead], report: WorkbookIntegrityReport) -> None:
        clients = self._value_set(tables.get("tblClients"), "ClientID")
        parents = self._value_set(tables.get("tblParents"), "ParentID")
        matters = self._value_set(tables.get("tblMatters"), "MatterID")
        matter_numbers = self._value_set(tables.get("tblMatters"), "MatterNumber")
        accounts = self._value_set(tables.get("tblTransactionAccounts"), "AccountCode")
        categories = self._value_set(tables.get("tblTransactionCategories"), "CategoryCode")
        business_units = self._value_set(tables.get("tblTransactionBusinessUnits"), "BusinessUnit")

        self._check_ref(tables.get("tblClientProfiles"), "ClientID", clients, report, "missing_client_reference")
        self._check_ref(tables.get("tblMatters"), "ClientID", clients, report, "missing_client_reference")
        self._check_ref(tables.get("tblMatters"), "ParentID", parents, report, "missing_parent_reference", "warning")
        self._check_ref(tables.get("tblTimeEntries"), "ClientID", clients, report, "missing_client_reference")
        self._check_ref(tables.get("tblTimeEntries"), "MatterID", matters, report, "missing_matter_reference")
        self._check_ref(tables.get("tblTimeEntries"), "ParentID", parents, report, "missing_parent_reference", "warning")
        self._check_ref(tables.get("tblTrademarks"), "MatterNumber", matter_numbers, report, "missing_matter_number_reference", "warning")
        self._check_ref(tables.get("tblTransactionsMaster"), "CategoryCode", categories, report, "missing_category_reference", "warning")
        self._check_ref(tables.get("tblTransactionsMaster"), "FromAccount", accounts, report, "missing_account_reference", "warning")
        self._check_ref(tables.get("tblTransactionsMaster"), "ToAccount", accounts, report, "missing_account_reference", "warning")
        self._check_ref(tables.get("tblTransactionsMaster"), "BusinessUnit", business_units, report, "missing_business_unit_reference", "warning")
        self._check_matter_client_ownership(
            tables.get("tblMatters"),
            tables.get("tblTimeEntries"),
            report,
        )

    def _check_matter_client_ownership(
        self,
        matters: Optional[_TableRead],
        time_entries: Optional[_TableRead],
        report: WorkbookIntegrityReport,
    ) -> None:
        if matters is None or time_entries is None:
            return
        matter_clients = {
            self._clean(row.get("MatterID")).lower(): self._clean(row.get("ClientID"))
            for row in matters.rows
            if self._clean(row.get("MatterID")) and self._clean(row.get("ClientID"))
        }
        for row, row_idx in zip(time_entries.rows, time_entries.row_numbers):
            matter_id = self._clean(row.get("MatterID"))
            client_id = self._clean(row.get("ClientID"))
            matter_client_id = matter_clients.get(matter_id.lower(), "")
            if not matter_id or not client_id or not matter_client_id:
                continue
            if client_id.lower() == matter_client_id.lower():
                continue
            self._add(
                report,
                "error",
                "matter_client_mismatch",
                f"Time entry ClientID '{client_id}' does not match referenced matter ClientID '{matter_client_id}'.",
                sheet=time_entries.sheet,
                table=time_entries.table,
                row=row_idx,
                column="MatterID",
                value=matter_id,
            )

    def _check_ref(
        self,
        table_read: Optional[_TableRead],
        column: str,
        allowed_values: set[str],
        report: WorkbookIntegrityReport,
        code: str,
        severity: Severity = "error",
    ) -> None:
        if table_read is None or not allowed_values:
            return
        for row, row_idx in zip(table_read.rows, table_read.row_numbers):
            value = self._clean(row.get(column))
            if not value or value in allowed_values:
                continue
            self._add(
                report,
                severity,
                code,
                f"Value '{value}' in '{column}' does not exist in the referenced table.",
                sheet=table_read.sheet,
                table=table_read.table,
                row=row_idx,
                column=column,
                value=value,
            )

    def _check_financial_totals(self, tables: Dict[str, _TableRead], report: WorkbookIntegrityReport) -> None:
        time_table = tables.get("tblTimeEntries")
        time_gross = Decimal("0")
        time_net = Decimal("0")
        txn_amount = Decimal("0")

        if time_table is not None:
            for row, row_idx in zip(time_table.rows, time_table.row_numbers):
                hours = self._to_decimal(row.get("Hours"))
                rate = self._to_decimal(row.get("ClientRate"))
                share_pct = self._to_decimal(row.get("SharePct"))
                if hours is None or rate is None or share_pct is None:
                    continue
                expected = calc_amounts(float(hours), float(rate), normalize_pct(share_pct))
                self._check_money(row, row_idx, time_table, "GrossToClient", expected["gross_to_client"], report)
                self._check_money(row, row_idx, time_table, "AmountToYou", expected["amount_to_you"], report)
                self._check_money(row, row_idx, time_table, "HST", expected["hst_on_you"], report)
                self._check_money(row, row_idx, time_table, "TotalInclHST", expected["total_you_incl_hst"], report)
                time_gross += Decimal(str(expected["gross_to_client"]))
                time_net += Decimal(str(expected["amount_to_you"]))

        txn_table = tables.get("tblTransactionsMaster")
        if txn_table is not None:
            for row in txn_table.rows:
                amount = self._to_decimal(row.get("Amount"))
                if amount is not None:
                    txn_amount += amount

        report.financial_totals = {
            "timeGrossToClient": float(round(time_gross, 2)),
            "timeAmountToYou": float(round(time_net, 2)),
            "transactionsAmount": float(round(txn_amount, 2)),
        }

    def _check_money(
        self,
        row: Dict[str, Any],
        row_idx: int,
        table_read: _TableRead,
        column: str,
        expected: float,
        report: WorkbookIntegrityReport,
    ) -> None:
        if self._is_empty(row.get(column)):
            return
        actual = self._to_decimal(row.get(column))
        if actual is None:
            return
        expected_decimal = Decimal(str(expected))
        if abs(actual - expected_decimal) > Decimal(str(self.MONEY_TOLERANCE)):
            self._add(
                report,
                "error",
                "financial_mismatch",
                f"Financial field '{column}' is {actual} but should be {expected_decimal}.",
                sheet=table_read.sheet,
                table=table_read.table,
                row=row_idx,
                column=column,
                value=str(actual),
            )

    def _check_date_relationships(self, tables: Dict[str, _TableRead], report: WorkbookIntegrityReport) -> None:
        matters = tables.get("tblMatters")
        if matters is not None:
            for row, row_idx in zip(matters.rows, matters.row_numbers):
                opened = self._parse_date(row.get("DateOpened"))
                closed = self._parse_date(row.get("DateClosed"))
                if opened and closed and closed < opened:
                    self._add(
                        report,
                        "error",
                        "date_order_invalid",
                        "DateClosed is earlier than DateOpened.",
                        sheet=matters.sheet,
                        table=matters.table,
                        row=row_idx,
                        column="DateClosed",
                        value=closed.isoformat(),
                    )

        trademarks = tables.get("tblTrademarks")
        if trademarks is not None:
            deadline_columns = [
                "RenewalDeadline",
                "OfficeActionResponseDeadline",
                "OppositionDeadline",
                "SOUDeadline",
                "Section8Deadline",
                "Section15Deadline",
                "Section9Deadline",
                "OppositionPeriodEndDate",
                "UpcomingLocalDeadlineOfficeActionDate",
            ]
            for row, row_idx in zip(trademarks.rows, trademarks.row_numbers):
                filing = self._parse_date(row.get("FilingDate"))
                if not filing:
                    continue
                for column in deadline_columns:
                    deadline = self._parse_date(row.get(column))
                    if deadline and deadline < filing:
                        self._add(
                            report,
                            "warning",
                            "deadline_before_filing_date",
                            f"{column} is earlier than FilingDate.",
                            sheet=trademarks.sheet,
                            table=trademarks.table,
                            row=row_idx,
                            column=column,
                            value=deadline.isoformat(),
                        )

    def _value_set(self, table_read: Optional[_TableRead], column: str) -> set[str]:
        if table_read is None:
            return set()
        return {
            self._clean(row.get(column))
            for row in table_read.rows
            if self._clean(row.get(column))
        }

    def _add(
        self,
        report: WorkbookIntegrityReport,
        severity: Severity,
        code: str,
        message: str,
        sheet: str = "",
        table: str = "",
        row: Optional[int] = None,
        column: str = "",
        value: str = "",
    ) -> None:
        report.issues.append(
            WorkbookIntegrityIssue(
                severity=severity,
                code=code,
                message=message,
                sheet=sheet,
                table=table,
                row=row,
                column=column,
                value=value,
            )
        )

    def _row_has_data(self, row: Dict[str, Any]) -> bool:
        return any(not self._is_empty(value) for value in row.values())

    def _is_empty(self, value: Any) -> bool:
        return value is None or (isinstance(value, str) and value.strip() == "")

    def _clean(self, value: Any) -> str:
        return str(value or "").strip()

    def _safe_value(self, value: Any) -> str:
        text = self._clean(value)
        if len(text) > 120:
            return text[:117] + "..."
        return text

    def _to_decimal(self, value: Any) -> Optional[Decimal]:
        if self._is_empty(value):
            return None
        if isinstance(value, bool):
            return Decimal(1 if value else 0)
        try:
            text = str(value).strip().replace(",", "").replace("$", "").replace("%", "")
            if text.startswith("="):
                return None
            return Decimal(text)
        except (InvalidOperation, ValueError, TypeError):
            return None

    def _is_bool_int(self, value: Any) -> bool:
        if isinstance(value, bool):
            return True
        text = str(value).strip().lower()
        return text in {"0", "1", "true", "false", "yes", "no", "y", "n"}

    def _parse_date(self, value: Any) -> Optional[date]:
        if self._is_empty(value):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            try:
                return from_excel(value).date()
            except Exception:
                return None
        text = str(value).strip()
        if text.startswith("="):
            return None
        formats = [
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%m/%d/%Y",
            "%m-%d-%Y",
            "%d/%m/%Y",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text).date()
        except ValueError:
            return None

    def _sha256(self, path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    def _close_workbook(self, workbook: Any) -> None:
        if workbook is None:
            return
        for attr_name in ("vba_archive", "_archive"):
            try:
                archive = getattr(workbook, attr_name, None)
                if archive is not None:
                    try:
                        archive.close()
                    except Exception:
                        pass
                    try:
                        setattr(workbook, attr_name, None)
                    except Exception:
                        pass
            except Exception:
                pass
        try:
            workbook.close()
        except Exception:
            pass
