import logging
import difflib
import re
import time
import os
import shutil
import tempfile
from collections import Counter
from contextlib import nullcontext
from datetime import date, datetime
from typing import Any, Callable, Dict, List, Optional, Tuple
import openpyxl

from repositories.excel_repo import (
    ExcelRepo,
    TBL_CLIENTS,
    TBL_LEDGER,
    TBL_MATTERS,
    TBL_TIME,
    TBL_TRANSACTIONS_MASTER,
)
from domain import schema_constants as sc

logger = logging.getLogger(__name__)

# Type alias for progress callbacks: (phase_label, current_row, total_rows)
ProgressCallback = Optional[Callable[[str, int, int], None]]
DuplicateCallback = Optional[Callable[[Dict[str, Any]], Dict[str, Any]]]

DUPLICATE_ACTION_SKIP = "skip"
DUPLICATE_ACTION_ADD = "add"
DUPLICATE_ACTION_OVERWRITE = "overwrite"
DUPLICATE_ACTIONS = {
    DUPLICATE_ACTION_SKIP,
    DUPLICATE_ACTION_ADD,
    DUPLICATE_ACTION_OVERWRITE,
}
DECISION_SCOPE_ALL = "all"
DECISION_SCOPE_ONE = "one"
LEGACY_UNASSIGNED_MATTER_NAME = "Legacy Unassigned - Review Required"
LEGACY_UNMAPPED_MATTER_PREFIX = "Legacy Matter"


class DocketsImportService:
    """
    Handles importing data from a legacy Dockets.xlsm file into the CSPM active workbook.
    Imports: Clients, Matters, Dockets (time entries), Disbursements, Ledger, Receivables, Invoice Log.
    """

    IMPORT_SHEETS = ["Clients", "Matters", "Dockets", "Disbursements", "Ledger", "Receivables", "Invoice Log"]
    DATA_TYPE_CATEGORIES = [
        "Clients", "Matters", "Dockets", "Disbursements",
        "Billings", "Payments", "Expenses", "Write-offs",
        "Receivables", "Invoice Log",
    ]
    LEDGER_SUB_TYPES = {"Billings", "Payments", "Expenses", "Write-offs"}
    DATE_FILTER_COLUMNS = {
        "Dockets": "Date",
        "Disbursements": "Date",
        "Ledger": "Date",
        "Receivables": "Date",
        "Invoice Log": "Invoice Date",
    }
    ANALYSIS_PHASE_PREFIX = "Analysis:"
    PLAN_PHASE_PREFIX = "Plan:"
    ANALYSIS_TOTAL_STEPS = len(IMPORT_SHEETS) + 4
    CLIENT_NAME_COLUMNS = ["Client Name", "Client_Name", "Client", "Display Name", "Name"]
    CLIENT_FIRST_NAME_COLUMNS = ["First Name", "FirstName", "Given Name", "GivenName"]
    CLIENT_MIDDLE_NAME_COLUMNS = ["Middle Name", "MiddleName", "Middle Initial", "MiddleInitial"]
    CLIENT_LAST_NAME_COLUMNS = ["Last Name", "LastName", "Surname", "Family Name", "FamilyName"]

    def __init__(self, excel_repo: ExcelRepo):
        self.excel_repo = excel_repo
        self._client_name_cache: Dict[str, str] = {}
        self._parent_name_cache: Dict[str, str] = {}
        self._duplicate_maps: Dict[str, Dict[str, Dict[str, Any]]] = {}
        # Exact client/matter display labels evolved during the CSPM migration.
        # Keep an occurrence-aware, commercial-field index for the read-only
        # import review so existing historical dockets are not re-offered just
        # because their client/matter labels were normalized differently.
        self._time_core_duplicate_counts: Counter = Counter()
        self._duplicate_policy_all: str = ""
        self._legacy_client_id_to_cspm_id: Dict[str, str] = {}
        self._legacy_matter_map: Dict[str, Dict[str, str]] = {}
        self._legacy_unassigned_warning_clients: set[str] = set()
        self._legacy_unmapped_warning_ids: set[str] = set()
        self._repo_client_id_to_name: Dict[str, str] = {}
        self._repo_matter_id_to_name: Dict[str, str] = {}
        self._source_rows_cache: Dict[Tuple[int, str], List[Dict[str, Any]]] = {}
        self._source_receivable_state_cache: Dict[int, Dict[str, Dict[str, Any]]] = {}

    def _create_shadow_copy(self, file_path: str) -> str:
        """Create a temporary shadow copy of the workbook to bypass Excel file locks."""
        fd, temp_path = tempfile.mkstemp(suffix=".xlsm", prefix="cspm_shadow_")
        os.close(fd)
        # Fast path: standard copy (works if file is not locked)
        try:
            shutil.copy2(file_path, temp_path)
            return temp_path
        except (PermissionError, OSError):
            pass
        # Fallback: read with sharing mode (works when Excel has file open)
        try:
            with open(file_path, 'rb') as src:
                data = src.read()
            with open(temp_path, 'wb') as dst:
                dst.write(data)
            return temp_path
        except Exception as e:
            try:
                os.unlink(temp_path)
            except Exception:
                pass
            raise RuntimeError(
                f"Could not read source file '{file_path}'. "
                f"It may be exclusively locked by another application. Error: {e}"
            )

    def list_source_clients(self, file_path: str) -> List[str]:
        """Quick scan of source workbook to extract unique client names for filter UI."""
        shadow_path = None
        wb = None
        clients = set()
        try:
            shadow_path = self._create_shadow_copy(file_path)
            wb = openpyxl.load_workbook(shadow_path, read_only=True, data_only=True)
            # Scan Clients sheet
            for row in self._parse_sheet(wb, "Clients"):
                name = self._legacy_client_name_from_row(row)
                if name:
                    clients.add(name.strip())
            # Scan Dockets sheet Client/Sub-Client columns
            for row in self._parse_sheet(wb, "Dockets"):
                raw_client = self._clean(row.get("Client"))
                raw_subclient = self._clean(row.get("Sub-Client"))
                if raw_subclient:
                    clients.add(raw_subclient.strip())
                elif raw_client:
                    clients.add(raw_client.strip())
            # Scan Ledger sheet Client/Vendor column
            for row in self._parse_sheet(wb, "Ledger"):
                cv = self._clean(row.get("Client/Vendor"))
                if cv:
                    clients.add(cv.strip())
        except Exception as e:
            logger.warning(f"Could not scan source clients: {e}")
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass
            if shadow_path and os.path.exists(shadow_path):
                try:
                    os.unlink(shadow_path)
                except Exception:
                    pass
        return sorted(clients, key=str.lower)

    # ── Fuzzy matching ────────────────────────────────────────────────────────

    def _fuzzy_match(self, raw_name: str, existing_names: List[str], threshold: float = 0.8) -> Optional[str]:
        if not raw_name:
            return None
        raw_lower = raw_name.strip().lower()
        for name in existing_names:
            if name.lower() == raw_lower:
                return name
        matches = difflib.get_close_matches(raw_name, existing_names, n=1, cutoff=threshold)
        return matches[0] if matches else None

    def _resolve_client_name(self, raw_name: str, existing_clients: List[str]) -> str:
        if not raw_name or not raw_name.strip():
            return ""
        key = raw_name.strip().lower()
        if key in self._client_name_cache:
            return self._client_name_cache[key]
        matched = self._fuzzy_match(raw_name.strip(), existing_clients)
        if matched:
            self._client_name_cache[key] = matched
            return matched
        canonical = raw_name.strip()
        existing_clients.append(canonical)
        self._client_name_cache[key] = canonical
        return canonical

    def _resolve_parent_name(self, raw_name: str, existing_parents: List[str]) -> str:
        if not raw_name or not raw_name.strip():
            return ""
        key = raw_name.strip().lower()
        if key in self._parent_name_cache:
            return self._parent_name_cache[key]
        matched = self._fuzzy_match(raw_name.strip(), existing_parents)
        if matched:
            self._parent_name_cache[key] = matched
            return matched
        canonical = raw_name.strip()
        existing_parents.append(canonical)
        self._parent_name_cache[key] = canonical
        return canonical

    # ── Client / Sub-Client mapping ───────────────────────────────────────────

    def _map_client_parent(self, raw_client: str, raw_subclient: str,
                           existing_clients: List[str], existing_parents: List[str]):
        client = raw_client.strip() if raw_client else ""
        subclient = raw_subclient.strip() if raw_subclient else ""
        if subclient:
            cspm_parent = self._resolve_parent_name(client, existing_parents)
            cspm_client = self._resolve_client_name(subclient, existing_clients)
        else:
            cspm_parent = ""
            cspm_client = self._resolve_client_name(client, existing_clients)
        return cspm_client, cspm_parent

    # ── Date filtering ────────────────────────────────────────────────────────

    def _normalise_legacy_id(self, value) -> str:
        text = self._clean(value)
        if not text:
            return ""
        if text.endswith(".0"):
            text = text[:-2]
        import re
        return re.sub(r'[\W_]+', '', text).lower()

    def _pick_first_text(self, row: Dict[str, Any], keys: List[str]) -> str:
        for key in keys:
            value = self._clean(row.get(key))
            if value:
                return value
        return ""

    def _legacy_client_name_from_row(self, row: Dict[str, Any]) -> str:
        return self._pick_first_text(row, self.CLIENT_NAME_COLUMNS)

    def _split_individual_name_parts(self, raw_name: str) -> Tuple[str, str, str]:
        text = " ".join(self._clean(raw_name).split())
        if not text:
            return "", "", ""
        if "," in text:
            last_part, rest = text.split(",", 1)
            last_name = " ".join(last_part.split())
            given_parts = " ".join(rest.split()).split()
            if given_parts and last_name:
                return given_parts[0], " ".join(given_parts[1:]), last_name
        parts = text.split()
        if len(parts) == 1:
            return parts[0], "", ""
        if len(parts) == 2:
            return parts[0], "", parts[1]
        return parts[0], " ".join(parts[1:-1]), parts[-1]

    def _legacy_client_name_parts_from_row(
        self,
        row: Dict[str, Any],
        fallback_name: str,
        entity_type: str,
    ) -> Tuple[str, str, str]:
        first_name = self._pick_first_text(row, self.CLIENT_FIRST_NAME_COLUMNS)
        middle_name = self._pick_first_text(row, self.CLIENT_MIDDLE_NAME_COLUMNS)
        last_name = self._pick_first_text(row, self.CLIENT_LAST_NAME_COLUMNS)
        if entity_type.strip().lower() == "individual" and not (first_name and last_name):
            split_first, split_middle, split_last = self._split_individual_name_parts(fallback_name)
            if not first_name:
                first_name = split_first
            if not middle_name:
                middle_name = split_middle
            if not last_name:
                last_name = split_last
        return first_name, middle_name, last_name

    def _legacy_client_name_for_id(self, legacy_id) -> str:
        key = self._normalise_legacy_id(legacy_id)
        if not key:
            return ""
        return getattr(self, "_legacy_id_to_name", {}).get(key, "")

    def _build_legacy_client_index(self, client_rows: List[Dict[str, Any]]) -> None:
        self._legacy_id_to_name: Dict[str, str] = {}
        for row in client_rows:
            legacy_id = self._normalise_legacy_id(row.get("Client_ID"))
            client_name = self._legacy_client_name_from_row(row)
            if legacy_id and client_name:
                self._legacy_id_to_name[legacy_id] = client_name

    def _coerce_datetime(self, value) -> Optional[datetime]:
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day)
        text = self._clean(value)
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        return None

    def _passes_date_filter(self, row_date, mode: str,
                            start_date: Optional[datetime], end_date: Optional[datetime]) -> bool:
        if mode == "all":
            return True
        parsed_date = self._coerce_datetime(row_date)
        if not parsed_date:
            return mode != "date_range"
        if mode == "date_range":
            if start_date and parsed_date < start_date:
                return False
            if end_date:
                end_of_day = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59, 999999)
                if parsed_date > end_of_day:
                    return False
        return True

    # ── Sheet parsing helpers ─────────────────────────────────────────────────

    def _parse_sheet(self, wb, sheet_name: str) -> List[Dict[str, Any]]:
        cache_key = (id(wb), sheet_name)
        if cache_key in self._source_rows_cache:
            return self._source_rows_cache[cache_key]
        if sheet_name not in wb.sheetnames:
            self._source_rows_cache[cache_key] = []
            return []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            self._source_rows_cache[cache_key] = []
            return []
        headers = rows[0]
        col_map = {}
        for i, h in enumerate(headers):
            if h:
                col_map[i] = str(h).strip()
        result = []
        for row in rows[1:]:
            if not any(row):
                continue
            record = {}
            for idx, header_name in col_map.items():
                record[header_name] = row[idx] if idx < len(row) else None
            result.append(record)
        self._source_rows_cache[cache_key] = result
        return result

    def _clean(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _safe_float(self, value, default: float = 0.0) -> float:
        if value is None:
            return default
        try:
            if isinstance(value, str):
                import re
                cleaned = re.sub(r'[^\d\.\-]', '', value)
                return float(cleaned) if cleaned else default
            return float(value)
        except (ValueError, TypeError):
            return default

    def _is_actual_invoice_ref(self, value) -> bool:
        return bool(re.match(r"^\d{2}-\d{4}(?:-[A-Z])?$", self._clean(value)))

    def _invoice_marker_bucket(self, value) -> str:
        text = self._clean(value)
        upper = text.upper()
        if not text:
            return "blank"
        if self._is_actual_invoice_ref(text):
            return "invoice"
        if upper == "BILLED":
            return "legacy_billed"
        if "HOLD" in upper or "FORGOT" in upper:
            return "hold"
        if "FREE" in upper or "DO NOT BILL" in upper or "WRITE OFF" in upper:
            return "no_bill"
        return "other"

    def _invoice_status_for_marker(self, value) -> str:
        bucket = self._invoice_marker_bucket(value)
        if bucket == "invoice":
            return "Invoiced"
        if bucket == "legacy_billed":
            return "Legacy Billed"
        if bucket == "hold":
            return "Held for Billing"
        if bucket == "no_bill":
            return "Not Billable"
        if bucket == "other":
            return "Marked"
        return "Not Invoiced"

    def _build_receivable_state_map(self, wb) -> Dict[str, Dict[str, Any]]:
        cache_key = id(wb)
        if cache_key in self._source_receivable_state_cache:
            return self._source_receivable_state_cache[cache_key]

        state: Dict[str, Dict[str, Any]] = {}
        for row in self._parse_sheet(wb, "Receivables"):
            invoice_num = self._clean(row.get("InvoiceNum"))
            if not self._is_actual_invoice_ref(invoice_num):
                continue
            status = self._clean(row.get("Status"))
            total = self._safe_float(row.get("Total_Invoiced"))
            amount_paid = self._safe_float(row.get("Amount_Paid"))
            credits_adj = self._safe_float(row.get("Credits/Adj"))
            balance_due = self._safe_float(row.get("Balance_Due"))
            status_upper = status.upper()

            if status_upper in {"VOID", "VOIDED", "CANCELLED", "CANCELED"}:
                payment_status = "Void"
                amount_paid = 0.0
                balance_due = 0.0
            elif status_upper in {"PAID", "CLOSED"}:
                payment_status = "Paid"
                if amount_paid <= 0 and total > 0:
                    amount_paid = max(0.0, total + credits_adj)
                balance_due = 0.0
            elif amount_paid > 0 and balance_due > 0:
                payment_status = "Partial"
            elif balance_due <= 0 and (amount_paid > 0 or total <= 0):
                payment_status = "Paid"
            else:
                payment_status = "Pending"

            state[invoice_num] = {
                "invoiceDate": self._format_date(row.get("Date")),
                "paymentStatus": payment_status,
                "invoiceTotal": round(total, 2),
                "invoiceAmountPaid": round(amount_paid, 2),
                "invoiceBalanceDue": round(max(0.0, balance_due), 2),
                "receivableStatus": status,
            }

        self._source_receivable_state_cache[cache_key] = state
        return state

    def _payment_state_for_invoice(self, wb, invoice_ref: str) -> Dict[str, Any]:
        bucket = self._invoice_marker_bucket(invoice_ref)
        if bucket != "invoice":
            fallback = {
                "legacy_billed": "Unknown",
                "hold": "Not Invoiced",
                "no_bill": "Not Billable",
                "other": "Unknown",
            }.get(bucket, "Not Invoiced")
            return {
                "paymentStatus": fallback,
                "invoiceTotal": 0.0,
                "invoiceAmountPaid": 0.0,
                "invoiceBalanceDue": 0.0,
                "invoiceDate": "",
                "receivableStatus": "",
            }
        return self._build_receivable_state_map(wb).get(
            invoice_ref,
            {
                "paymentStatus": "Unknown",
                "invoiceTotal": 0.0,
                "invoiceAmountPaid": 0.0,
                "invoiceBalanceDue": 0.0,
                "invoiceDate": "",
                "receivableStatus": "",
            },
        )

    def _format_date(self, value) -> str:
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if value is None:
            return ""
        text = str(value).strip()
        if not text:
            return ""
        
        formats = ["%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y"]
        for fmt in formats:
            try:
                parsed_date = datetime.strptime(text, fmt)
                return parsed_date.strftime("%Y-%m-%d")
            except ValueError:
                continue
        return text

    def _emit(self, cb: ProgressCallback, phase: str, current: int, total: int):
        if cb:
            try:
                cb(phase, current, total)
            except Exception:
                pass

    # Duplicate handling

    def _normalise_duplicate_part(self, value) -> str:
        if isinstance(value, datetime):
            value = value.strftime("%Y-%m-%d")
        elif isinstance(value, date):
            value = value.strftime("%Y-%m-%d")
        text = self._clean(value)
        if not text:
            return ""
        return " ".join(text.split()).lower()

    def _duplicate_key(self, *parts) -> str:
        return "||".join(self._normalise_duplicate_part(part) for part in parts)

    def _numeric_key(self, value, places: int = 2) -> str:
        try:
            return f"{float(value):.{places}f}"
        except (TypeError, ValueError):
            return f"{0.0:.{places}f}"

    def _repo_text(self, row: Dict[str, Any], keys: List[str]) -> str:
        for key in keys:
            value = self._clean(row.get(key))
            if value:
                return value
        return ""

    def _remember_legacy_matter_mapping(
        self,
        legacy_matter_id: Any,
        *,
        matter_id: str,
        matter_name: str,
        client_name: str,
        parent_name: str,
    ) -> None:
        key = self._normalise_legacy_id(legacy_matter_id)
        if not key:
            return
        self._legacy_matter_map[key] = {
            "matterId": self._clean(matter_id),
            "matterName": self._clean(matter_name),
            "clientName": self._clean(client_name),
            "parentName": self._clean(parent_name),
        }

    def _read_repo_table(self, tref) -> List[Dict[str, Any]]:
        try:
            reader = getattr(self.excel_repo, "_read_table_rows", None)
            if callable(reader):
                return list(reader(tref) or [])
        except Exception as exc:
            logger.debug("Could not read duplicate map table %s: %s", getattr(tref, "table", tref), exc)
        return []

    def _client_duplicate_key(self, client_name: str) -> str:
        return self._duplicate_key(client_name)

    def _matter_duplicate_key(self, client_name: str, matter_name: str) -> str:
        return self._duplicate_key(client_name, matter_name)

    def _time_duplicate_key(self, payload: Dict[str, Any]) -> str:
        hours = self._safe_float(payload.get("hours"))

        return self._duplicate_key(
            payload.get("date"),
            payload.get("clientName") or payload.get("client"),
            payload.get("matterName") or payload.get("matter"),
            payload.get("description"),
            self._numeric_key(hours, 4),
            self._numeric_key(payload.get("clientRate") or payload.get("rate"), 2),
        )

    def _normalise_share_percent(self, value: Any) -> float:
        """Treat legacy fractional and CSPM whole-number shares as equivalent."""
        share = self._safe_float(value)
        if share != 0 and abs(share) <= 1:
            return share * 100.0
        return share

    def _time_core_duplicate_key(self, payload: Dict[str, Any]) -> str:
        """Return a client/matter-independent commercial signature for a docket.

        This is deliberately restricted to the immutable time/fee economics.
        It is used only for review-safe duplicate detection; it does not replace
        CSPM's normal, more detailed entry-creation duplicate policy.
        """
        return self._duplicate_key(
            payload.get("date"),
            payload.get("description"),
            self._numeric_key(payload.get("hours"), 4),
            self._numeric_key(payload.get("clientRate") or payload.get("rate"), 2),
            self._numeric_key(self._normalise_share_percent(payload.get("sharePct")), 4),
            self._numeric_key(payload.get("amountToYou"), 2),
        )

    def _transaction_duplicate_key(self, payload: Dict[str, Any]) -> str:
        return self._duplicate_key(
            payload.get("txnDate") or payload.get("date"),
            payload.get("class"),
            payload.get("type"),
            payload.get("client") or payload.get("clientName"),
            payload.get("matter") or payload.get("matterName"),
            self._numeric_key(payload.get("amount"), 2),
            self._numeric_key(payload.get("taxAmount"), 2),
            payload.get("invoiceRef") or payload.get("invoice"),
            payload.get("expenseDetails") or payload.get("details"),
            payload.get("notes"),
        )

    def _remember_duplicate(self, kind: str, key: str, record: Dict[str, Any]) -> None:
        if not key:
            return
        item = dict(record or {})
        item["kind"] = kind
        self._duplicate_maps.setdefault(kind, {})[key] = item

    def _result_id_for_kind(self, kind: str, result: Dict[str, Any], payload: Dict[str, Any]) -> str:
        if kind == "client":
            return self._clean(result.get("clientId") or payload.get("clientId"))
        if kind == "matter":
            return self._clean(result.get("matterId") or payload.get("matterId"))
        if kind == "time":
            return self._clean(result.get("entryId") or payload.get("entryId"))
        if kind == "transaction":
            return self._clean(result.get("transactionId") or payload.get("transactionId"))
        return ""

    def _register_saved_duplicate_key(
        self,
        kind: str,
        key: str,
        payload: Dict[str, Any],
        result: Dict[str, Any],
    ) -> None:
        record_id = self._result_id_for_kind(kind, result, payload)
        label = self._duplicate_label(kind, payload)
        summary = self._duplicate_summary(kind, payload)
        self._remember_duplicate(
            kind,
            key,
            {
                "id": record_id,
                "label": label,
                "summary": summary,
            },
        )

    def _duplicate_label(self, kind: str, payload: Dict[str, Any]) -> str:
        if kind == "client":
            return self._clean(payload.get("clientName"))
        if kind == "matter":
            return self._clean(payload.get("matterName"))
        if kind == "time":
            return self._clean(payload.get("description")) or "Time entry"
        if kind == "transaction":
            return (
                self._clean(payload.get("invoiceRef"))
                or self._clean(payload.get("notes"))
                or self._clean(payload.get("expenseDetails"))
                or "Transaction"
            )
        return "Record"

    def _duplicate_summary(self, kind: str, payload: Dict[str, Any]) -> str:
        if kind == "client":
            parent = self._clean(payload.get("parentClientName"))
            name = self._clean(payload.get("clientName"))
            return f"{name} | Parent: {parent or 'none'}"
        if kind == "matter":
            return (
                f"{self._clean(payload.get('clientName'))} / "
                f"{self._clean(payload.get('matterName'))}"
            )
        if kind == "time":
            return (
                f"{self._clean(payload.get('date'))} | "
                f"{self._clean(payload.get('clientName'))} / "
                f"{self._clean(payload.get('matterName')) or 'client-only'} | "
                f"{self._clean(payload.get('description'))} | "
                f"{self._numeric_key(payload.get('hours'), 2)} hrs"
            )
        if kind == "transaction":
            return (
                f"{self._clean(payload.get('txnDate'))} | "
                f"{self._clean(payload.get('type'))} | "
                f"{self._clean(payload.get('client')) or 'no client'} | "
                f"${self._numeric_key(payload.get('amount'), 2)}"
            )
        return self._duplicate_label(kind, payload)

    def _build_duplicate_maps(self, existing_clients: List[str]) -> None:
        self._duplicate_maps = {
            "client": {},
            "matter": {},
            "time": {},
            "transaction": {},
        }
        self._time_core_duplicate_counts = Counter()
        self._duplicate_policy_all = ""
        self._legacy_client_id_to_cspm_id = {}
        self._legacy_matter_map = {}
        self._legacy_unassigned_warning_clients = set()
        self._legacy_unmapped_warning_ids = set()
        self._repo_client_id_to_name = {}
        self._repo_matter_id_to_name = {}

        client_rows = self._read_repo_table(TBL_CLIENTS)
        for row in client_rows:
            client_name = self._repo_text(row, [sc.COL_CLIENT_NAME, "clientName", "ClientName"])
            client_id = self._repo_text(row, [sc.COL_CLIENT_ID, "clientId", "ClientID"])
            if client_id and client_name:
                self._repo_client_id_to_name[client_id.lower()] = client_name
            key = self._client_duplicate_key(client_name)
            if key:
                self._remember_duplicate(
                    "client",
                    key,
                    {
                        "id": client_id,
                        "label": client_name,
                        "summary": client_name,
                    },
                )

        for client_name in existing_clients:
            key = self._client_duplicate_key(client_name)
            if key and key not in self._duplicate_maps["client"]:
                self._remember_duplicate(
                    "client",
                    key,
                    {
                        "id": "",
                        "label": client_name,
                        "summary": client_name,
                    },
                )

        matter_rows = self._read_repo_table(TBL_MATTERS)
        for row in matter_rows:
            matter_name = self._repo_text(row, [sc.COL_MATTER_NAME, "matterName", "MatterName"])
            matter_id = self._repo_text(row, [sc.COL_MATTER_ID, "matterId", "MatterID"])
            matter_num = self._repo_text(row, [sc.COL_MATTER_NUMBER, "matterNumber", "MatterNumber"])
            client_name = self._repo_text(row, [sc.COL_MATTER_CLIENT_NAME, "clientName", "ClientName"])
            client_id = self._repo_text(row, [sc.COL_MATTER_CLIENT_ID, "clientId", "ClientID"])
            if not client_name and client_id:
                client_name = self._repo_client_id_to_name.get(client_id.lower(), "")
            if matter_id and matter_name:
                self._repo_matter_id_to_name[matter_id.lower()] = matter_name
            key = self._matter_duplicate_key(client_name, matter_name)
            if key:
                self._remember_duplicate(
                    "matter",
                    key,
                    {
                        "id": matter_id,
                        "label": matter_name,
                        "summary": f"{client_name} / {matter_name}",
                    },
                )
            if matter_num:
                self._remember_duplicate(
                    "matter",
                    f"num:{matter_num.lower()}",
                    {
                        "id": matter_id,
                        "label": matter_name,
                        "summary": f"{client_name} / {matter_name}",
                    },
                )

        time_rows = self._read_repo_table(TBL_TIME)
        for row in time_rows:
            client_id = self._repo_text(row, [sc.COL_TIME_CLIENT_ID, "clientId", "ClientID"])
            matter_id = self._repo_text(row, [sc.COL_TIME_MATTER_ID, "matterId", "MatterID"])
            client_name = self._repo_client_id_to_name.get(client_id.lower(), client_id)
            matter_name = self._repo_matter_id_to_name.get(matter_id.lower(), matter_id)
            payload = {
                "date": row.get(sc.COL_TIME_DATE),
                "clientName": client_name,
                "matterName": matter_name,
                "description": row.get(sc.COL_TIME_DESC),
                "hours": row.get(sc.COL_TIME_HOURS),
                "rawSeconds": row.get(sc.COL_TIME_SECONDS),
                "clientRate": row.get(sc.COL_TIME_RATE),
                "sharePct": row.get(sc.COL_TIME_SHARE_PCT),
                "amountToYou": row.get(sc.COL_TIME_NET),
            }
            key = self._time_duplicate_key(payload)
            if key:
                self._remember_duplicate(
                    "time",
                    key,
                    {
                        "id": self._repo_text(row, [sc.COL_TIME_ENTRY_ID, "entryId", "EntryID"]),
                        "label": self._duplicate_label("time", payload),
                        "summary": self._duplicate_summary("time", payload),
                    },
                )
            core_key = self._time_core_duplicate_key(payload)
            if core_key:
                self._time_core_duplicate_counts[core_key] += 1

        # 4. Ledger
        for row in self._read_repo_table(TBL_TRANSACTIONS_MASTER):
            payload = {
                "date": row.get(sc.COL_TXN_DATE),
                "class": row.get(sc.COL_TXN_CLASS),
                "type": row.get(sc.COL_TXN_TYPE),
                "amount": row.get(sc.COL_TXN_AMOUNT),
                "description": row.get(sc.COL_TXN_NOTES),
                "account": row.get(sc.COL_TXN_FROM_ACCOUNT) or row.get(sc.COL_TXN_TO_ACCOUNT),
            }
            key = self._transaction_duplicate_key(payload)
            self._duplicate_maps.setdefault("ledger", {})[key] = {
                "id": row.get(sc.COL_TXN_ID),
                "label": f"{row.get(sc.COL_TXN_DATE)} - {row.get(sc.COL_TXN_NOTES)}",
                "summary": f"{row.get(sc.COL_TXN_CLASS)} / {row.get(sc.COL_TXN_TYPE)}",
                "score": 100,
            }
            
        for row in self._read_repo_table(TBL_LEDGER):
            client_vendor = self._clean(row.get(sc.COL_LEDGER_CLIENT_VENDOR))
            billings = self._safe_float(row.get(sc.COL_LEDGER_BILLINGS_EXCL_HST))
            hst_collected = self._safe_float(row.get(sc.COL_LEDGER_HST_COLLECTED))
            expenses = self._safe_float(row.get(sc.COL_LEDGER_EXPENSES_EXCL_HST))
            hst_paid = self._safe_float(row.get(sc.COL_LEDGER_HST_PAID))
            collected = self._safe_float(row.get(sc.COL_LEDGER_COLLECTED))
            write_off = self._safe_float(row.get(sc.COL_LEDGER_WRITE_OFF))

            if billings > 0:
                txn_type, txn_class, amount, tax_amount = "Income", "Business", billings, hst_collected
            elif expenses > 0:
                txn_type, txn_class, amount, tax_amount = "Expense", "Business", expenses, hst_paid
            elif collected > 0:
                txn_type, txn_class, amount, tax_amount = "Transfer", "Business", collected, 0
            elif write_off > 0:
                txn_type, txn_class, amount, tax_amount = "Adjustment", "Business", -write_off, 0
            else:
                txn_type, txn_class, amount, tax_amount = "", "", 0.0, 0.0

            cspm_client = self._resolve_client_name(client_vendor, existing_clients) if client_vendor else ""

            payload = {
                "txnDate": self._format_date(row.get(sc.COL_LEDGER_DATE)),
                "class": txn_class,
                "type": txn_type,
                "client": cspm_client,
                "amount": amount,
                "taxAmount": tax_amount,
                "notes": self._clean(row.get(sc.COL_LEDGER_DESCRIPTION)),
                "invoiceRef": self._clean(row.get(sc.COL_LEDGER_REFERENCE)),
            }
            key = self._transaction_duplicate_key(payload)
            self._duplicate_maps.setdefault("transaction", {})[key] = {
                "id": row.get(sc.COL_LEDGER_ID),
                "label": f"{self._format_date(row.get(sc.COL_LEDGER_DATE))} - {payload['notes']}",
                "summary": "Legacy Ledger Entry",
                "score": 100,
            }

        transaction_rows = self._read_repo_table(TBL_TRANSACTIONS_MASTER)
        for row in transaction_rows:
            payload = {
                "txnDate": row.get(sc.COL_TXN_DATE),
                "class": row.get(sc.COL_TXN_CLASS),
                "type": row.get(sc.COL_TXN_TYPE),
                "parent": row.get(sc.COL_TXN_PARENT),
                "client": row.get(sc.COL_TXN_CLIENT),
                "matter": row.get(sc.COL_TXN_MATTER),
                "amount": row.get(sc.COL_TXN_AMOUNT),
                "taxAmount": row.get(sc.COL_TXN_TAX_AMOUNT),
                "invoiceRef": row.get(sc.COL_TXN_INVOICE_REF),
                "expenseDetails": row.get(sc.COL_TXN_EXPENSE_DETAILS),
                "notes": row.get(sc.COL_TXN_NOTES),
            }
            key = self._transaction_duplicate_key(payload)
            if key:
                self._remember_duplicate(
                    "transaction",
                    key,
                    {
                        "id": self._repo_text(row, [sc.COL_TXN_ID, "transactionId", "TransactionID"]),
                        "label": self._duplicate_label("transaction", payload),
                        "summary": self._duplicate_summary("transaction", payload),
                    },
                )

    def _resolve_duplicate_action(
        self,
        kind: str,
        key: str,
        sheet: str,
        row_number: int,
        payload: Dict[str, Any],
        results: Dict[str, Any],
        duplicate_callback: DuplicateCallback,
    ) -> Tuple[str, Optional[Dict[str, Any]]]:
        existing = self._duplicate_maps.get(kind, {}).get(key)
        if not existing:
            return DUPLICATE_ACTION_ADD, None

        results["duplicatesFound"] += 1
        action = self._duplicate_policy_all
        if action not in DUPLICATE_ACTIONS:
            if duplicate_callback:
                prompt = {
                    "kind": kind,
                    "sheet": sheet,
                    "rowNumber": row_number,
                    "recordLabel": self._duplicate_label(kind, payload),
                    "existingId": self._clean(existing.get("id")),
                    "existingSummary": self._clean(existing.get("summary")),
                    "incomingSummary": self._duplicate_summary(kind, payload),
                    "actions": [
                        DUPLICATE_ACTION_SKIP,
                        DUPLICATE_ACTION_ADD,
                        DUPLICATE_ACTION_OVERWRITE,
                    ],
                }
                results["duplicatePrompts"] += 1
                response = duplicate_callback(prompt) or {}
                action = self._clean(response.get("action")).lower()
                scope = self._clean(response.get("scope")).lower()
                if action not in DUPLICATE_ACTIONS:
                    action = DUPLICATE_ACTION_SKIP
                if scope == DECISION_SCOPE_ALL:
                    self._duplicate_policy_all = action
            else:
                action = DUPLICATE_ACTION_ADD

        if action == DUPLICATE_ACTION_SKIP:
            results["duplicatesSkipped"] += 1
            label = self._duplicate_label(kind, payload)
            results["warnings"].append(f"{sheet} row {row_number}: skipped duplicate '{label}'")
        return action, existing

    def _apply_duplicate_write_result(
        self,
        results: Dict[str, Any],
        action: str,
        existing: Optional[Dict[str, Any]],
        ok: bool,
    ) -> None:
        if not existing or not ok:
            return
        if action == DUPLICATE_ACTION_ADD:
            results["duplicatesAdded"] += 1
        elif action == DUPLICATE_ACTION_OVERWRITE:
            results["duplicatesOverwritten"] += 1

    def _save_transaction_with_duplicate_decision(
        self,
        payload: Dict[str, Any],
        sheet: str,
        row_number: int,
        results: Dict[str, Any],
        duplicate_callback: DuplicateCallback,
    ) -> Dict[str, Any]:
        key = self._transaction_duplicate_key(payload)
        action, duplicate_existing = self._resolve_duplicate_action(
            "transaction",
            key,
            sheet,
            row_number,
            payload,
            results,
            duplicate_callback,
        )
        if action == DUPLICATE_ACTION_SKIP:
            return {"ok": False, "skippedDuplicate": True, "message": "Duplicate skipped"}
        if action == DUPLICATE_ACTION_OVERWRITE:
            existing_id = self._clean((duplicate_existing or {}).get("id"))
            if existing_id:
                payload["transactionId"] = existing_id

        result = self.excel_repo.save_transaction(payload)
        if result.get("ok"):
            self._apply_duplicate_write_result(results, action, duplicate_existing, True)
            self._register_saved_duplicate_key("transaction", key, payload, result)
        return result

    # ── Row counting (pre-scan) ───────────────────────────────────────────────

    def _count_rows_in_workbook(
        self,
        wb,
        mode: str,
        start_date: Optional[datetime],
        end_date: Optional[datetime],
        progress_callback: ProgressCallback = None,
    ) -> Dict[str, int]:
        counts = {}
        for index, sheet_name in enumerate(self.IMPORT_SHEETS):
            if progress_callback:
                self._emit(
                    progress_callback,
                    f"{self.ANALYSIS_PHASE_PREFIX} Scanning {sheet_name}",
                    3 + index,
                    self.ANALYSIS_TOTAL_STEPS,
                )
            rows = self._parse_sheet(wb, sheet_name)
            date_column = self.DATE_FILTER_COLUMNS.get(sheet_name, "")
            if date_column:
                count = sum(
                    1
                    for row in rows
                    if self._passes_date_filter(row.get(date_column), mode, start_date, end_date)
                )
            else:
                count = len(rows)
            counts[sheet_name] = count
            if progress_callback:
                self._emit(progress_callback, f"{self.PLAN_PHASE_PREFIX} {sheet_name}", 0, count)
        counts["total"] = sum(counts.values())
        return counts

    def count_rows(
        self,
        file_path: str,
        mode: str = "all",
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
    ) -> Dict[str, int]:
        """Quick pre-scan to count rows per sheet for progress estimation."""
        try:
            self._source_rows_cache = {}
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            counts = self._count_rows_in_workbook(wb, mode, start_date, end_date)
            wb.close()
            return counts
        except Exception as e:
            logger.error(f"Row count failed: {e}")
            return {"total": 0}

    # ── Main entry point ──────────────────────────────────────────────────────

    def import_legacy_workbook(
        self,
        file_path: str,
        mode: str = "all",
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        progress_callback: ProgressCallback = None,
        duplicate_callback: DuplicateCallback = None,
        allowed_rows: Optional[Dict[str, List[int]]] = None,
        data_types: Optional[List[str]] = None,
        client_filter: Optional[str] = None,
    ) -> Dict[str, Any]:
        logger.info(f"Starting import from {file_path} with mode={mode}")
        import_started = time.perf_counter()
        phase_seconds: Dict[str, float] = {}
        batch_metrics: Dict[str, Any] = {
            "enabled": False,
            "saveCount": 0,
            "postSaveVerified": False,
        }
        results = {
            "success": False,
            "clientsAdded": 0,
            "mattersAdded": 0,
            "docketsAdded": 0,
            "disbursementsAdded": 0,
            "ledgerAdded": 0,
            "receivablesAdded": 0,
            "invoiceLogAdded": 0,
            "duplicatesFound": 0,
            "duplicatePrompts": 0,
            "duplicatesSkipped": 0,
            "duplicatesAdded": 0,
            "duplicatesOverwritten": 0,
            "addedRecords": [],
            "errors": [],
            "warnings": [],
            "performance": {
                "phasesSeconds": phase_seconds,
                "batchWrite": batch_metrics,
                "totalSeconds": 0.0,
            },
        }

        wb = None
        shadow_path = None
        try:
            self._source_rows_cache = {}
            self._emit(
                progress_callback,
                f"{self.ANALYSIS_PHASE_PREFIX} Opening source workbook",
                0,
                self.ANALYSIS_TOTAL_STEPS,
            )
            phase_started = time.perf_counter()
            shadow_path = self._create_shadow_copy(file_path)
            wb = openpyxl.load_workbook(shadow_path, read_only=True, data_only=True)
            phase_seconds["openSourceWorkbook"] = round(time.perf_counter() - phase_started, 4)
            self._emit(
                progress_callback,
                f"{self.ANALYSIS_PHASE_PREFIX} Reading current CSPM directories",
                1,
                self.ANALYSIS_TOTAL_STEPS,
            )
            phase_started = time.perf_counter()
            existing_clients = list(self.excel_repo.list_client_names())
            existing_parents = list(self.excel_repo.list_parent_names())
            phase_seconds["readCurrentDirectories"] = round(time.perf_counter() - phase_started, 4)
            self._emit(
                progress_callback,
                f"{self.ANALYSIS_PHASE_PREFIX} Comparing existing CSPM records",
                2,
                self.ANALYSIS_TOTAL_STEPS,
            )
            phase_started = time.perf_counter()
            self._build_duplicate_maps(existing_clients)
            phase_seconds["buildDuplicateMaps"] = round(time.perf_counter() - phase_started, 4)
            phase_started = time.perf_counter()
            self._count_rows_in_workbook(
                wb,
                mode,
                start_date,
                end_date,
                progress_callback=progress_callback,
            )
            phase_seconds["scanSourceSheets"] = round(time.perf_counter() - phase_started, 4)

            # Build legacy Client_ID -> Client Name lookup for parent resolution.
            self._emit(
                progress_callback,
                f"{self.ANALYSIS_PHASE_PREFIX} Mapping source relationships",
                self.ANALYSIS_TOTAL_STEPS - 1,
                self.ANALYSIS_TOTAL_STEPS,
            )
            phase_started = time.perf_counter()
            client_rows = self._parse_sheet(wb, "Clients")
            self._build_legacy_client_index(client_rows)
            phase_seconds["mapSourceRelationships"] = round(time.perf_counter() - phase_started, 4)
            self._emit(
                progress_callback,
                f"{self.ANALYSIS_PHASE_PREFIX} Import plan ready",
                self.ANALYSIS_TOTAL_STEPS,
                self.ANALYSIS_TOTAL_STEPS,
            )

            batch_factory = getattr(self.excel_repo, "import_batch", None)
            batch_context = batch_factory() if callable(batch_factory) else nullcontext(batch_metrics)
            with batch_context as active_batch_metrics:
                if isinstance(active_batch_metrics, dict):
                    batch_metrics = active_batch_metrics
                    results["performance"]["batchWrite"] = batch_metrics

                phase_started = time.perf_counter()
                self._import_clients(wb, existing_clients, existing_parents, results, progress_callback, duplicate_callback, allowed_rows)
                phase_seconds["importClients"] = round(time.perf_counter() - phase_started, 4)

                phase_started = time.perf_counter()
                self._import_matters(wb, existing_clients, existing_parents, results, progress_callback, duplicate_callback, allowed_rows)
                phase_seconds["importMatters"] = round(time.perf_counter() - phase_started, 4)

                phase_started = time.perf_counter()
                self._import_dockets(wb, mode, start_date, end_date, existing_clients, existing_parents, results, progress_callback, duplicate_callback, allowed_rows)
                phase_seconds["importDockets"] = round(time.perf_counter() - phase_started, 4)

                phase_started = time.perf_counter()
                self._import_ledger(wb, mode, start_date, end_date, existing_clients, results, progress_callback, duplicate_callback, allowed_rows)

                # ── Pass 5.5: Deduced Receivables ─────────────────────────────────
                deduced_invoices = self._aggregate_ledger_invoices(wb, mode, start_date, end_date, existing_clients, allowed_rows)
                for inv in deduced_invoices:
                    res_ar = self.excel_repo.save_receivable(inv)
                    if res_ar.get("ok"):
                        results.setdefault("receivablesAdded", 0)
                        results["receivablesAdded"] += 1
                    res_log = self.excel_repo.save_invoice_log(inv)
                    if res_log.get("ok"):
                        results.setdefault("invoiceLogAdded", 0)
                        results["invoiceLogAdded"] += 1

                phase_seconds["importLedger"] = round(time.perf_counter() - phase_started, 4)

                self._emit(progress_callback, "Finalizing validated import batch", 0, 1)

            self._emit(progress_callback, "Finalizing validated import batch", 1, 1)
            results["success"] = True
            logger.info(
                f"Import complete: {results['clientsAdded']} clients, "
                f"{results['mattersAdded']} matters, {results['docketsAdded']} dockets, "
                f"{results['ledgerAdded']} ledger entries"
            )
        except Exception as e:
            logger.error(f"Import failed: {e}", exc_info=True)
            results["errors"].append(str(e))
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
            if shadow_path and os.path.exists(shadow_path):
                try:
                    os.unlink(shadow_path)
                except Exception:
                    pass
            results["performance"]["totalSeconds"] = round(time.perf_counter() - import_started, 4)

        return results


    def analyze_legacy_workbook(
        self,
        file_path: str,
        mode: str = "all",
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        data_types: Optional[List[str]] = None,
        client_filter: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Build a read-only row review report for a legacy Dockets workbook."""
        started = time.perf_counter()
        rows: List[Dict[str, Any]] = []
        warnings: List[str] = []
        result: Dict[str, Any] = {
            "success": False,
            "sourcePath": file_path,
            "mode": mode or "all",
            "startDate": start_date.strftime("%Y-%m-%d") if start_date else "",
            "endDate": end_date.strftime("%Y-%m-%d") if end_date else "",
            "dataTypes": data_types or [],
            "clientFilter": client_filter or "",
            "rowCounts": {},
            "summary": {},
            "rows": rows,
            "warnings": warnings,
            "errors": [],
            "performance": {"totalSeconds": 0.0},
        }
        wb = None
        shadow_path = None
        try:
            self._source_rows_cache = {}
            self._source_receivable_state_cache = {}
            
            shadow_path = self._create_shadow_copy(file_path)
            wb = openpyxl.load_workbook(shadow_path, read_only=True, data_only=True)
            existing_clients = list(self.excel_repo.list_client_names())
            existing_parents = list(self.excel_repo.list_parent_names())
            self._build_duplicate_maps(existing_clients)
            result["rowCounts"] = self._count_rows_in_workbook(wb, mode, start_date, end_date)
            self._build_legacy_client_index(self._parse_sheet(wb, "Clients"))

            sheets_to_analyze = set(data_types) if data_types else set(self.DATA_TYPE_CATEGORIES)
            
            if client_filter and ("Dockets" in sheets_to_analyze or bool(self.LEDGER_SUB_TYPES.intersection(sheets_to_analyze))):
                sheets_to_analyze.add("Clients")
                sheets_to_analyze.add("Matters")

            if "Clients" in sheets_to_analyze:
                self._analyze_clients(wb, existing_clients, existing_parents, rows, warnings)
            if "Matters" in sheets_to_analyze:
                self._analyze_matters(wb, existing_clients, existing_parents, rows, warnings)
            if "Dockets" in sheets_to_analyze:
                self._analyze_dockets(wb, mode, start_date, end_date, existing_clients, existing_parents, rows, warnings)
            if bool(self.LEDGER_SUB_TYPES.intersection(sheets_to_analyze)):
                self._analyze_ledger(wb, mode, start_date, end_date, existing_clients, rows, warnings)
                # Filter Ledger rows down to only the requested sub-types
                if data_types:
                    requested_sub_types = self.LEDGER_SUB_TYPES.intersection(set(data_types))
                    if requested_sub_types:
                        rows[:] = [
                            r for r in rows 
                            if r.get("sheet") != "Ledger" or r.get("payload", {}).get("subType") in requested_sub_types
                        ]

            if bool(self.LEDGER_SUB_TYPES.intersection(sheets_to_analyze)):
                deduced_invoices = self._aggregate_ledger_invoices(wb, mode, start_date, end_date, existing_clients)
                for i, inv in enumerate(deduced_invoices):
                    rows.append({
                        "sheet": "Ledger (A/R Deduced)",
                        "row": i + 2,
                        "action": "add",
                        "title": inv["invoiceNum"],
                        "details": f"Deduced A/R: Bal {inv['balanceDue']:.2f}",
                        "payload": inv
                    })

            if client_filter:
                rows[:] = self._apply_client_filter(rows, client_filter, existing_clients)

            import uuid
            for r in rows:
                action = r.get("action", "")
                r["selectable"] = action != "skip"
                r["defaultSelected"] = action != "skip"
                r["status"] = "already_imported" if action == "skip" else "new"
                if "rowId" not in r:
                    r["rowId"] = f"{r.get('sheet', 'U')}:{r.get('row', '0')}:{uuid.uuid4().hex[:8]}"
                if "client" not in r:
                    payload = r.get("payload") or {}
                    r["client"] = str(payload.get("Client") or payload.get("clientName") or payload.get("Client_ID") or r.get("title", ""))
                if "matter" not in r:
                    payload = r.get("payload") or {}
                    r["matter"] = str(payload.get("Matter_ID") or payload.get("Description (Internal)") or "")

            result["summary"] = self._analysis_summary(rows)
            result["success"] = True
        except Exception as e:
            logger.error(f"Analysis failed: {e}", exc_info=True)
            result["errors"].append(str(e))
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
            if shadow_path and os.path.exists(shadow_path):
                try:
                    os.unlink(shadow_path)
                except Exception:
                    pass
            result["performance"]["totalSeconds"] = round(time.perf_counter() - started, 4)

        return result

    def _apply_client_filter(self, rows: List[Dict[str, Any]], client_filter: str, existing_clients: List[str]) -> List[Dict[str, Any]]:
        """Filter analysis rows to only include those matching the client filter."""
        filter_lower = client_filter.strip().lower()
        matching_clients = set()
        matching_clients.add(filter_lower)
        for name in existing_clients:
            if filter_lower in name.lower() or name.lower() in filter_lower:
                matching_clients.add(name.lower())
        fuzzy_match = self._fuzzy_match(client_filter, existing_clients, threshold=0.7)
        if fuzzy_match:
            matching_clients.add(fuzzy_match.lower())
        
        referenced_clients = set()
        for r in rows:
            if r.get("sheet") in ("Dockets", "Ledger", "Ledger (A/R Deduced)"):
                title = (r.get("title") or "").lower()
                payload = r.get("payload") or {}
                client_name = (
                    self._clean(payload.get("clientName")) or
                    self._clean(payload.get("client")) or
                    self._clean(payload.get("Client")) or
                    self._clean(payload.get("Sub-Client")) or
                    ""
                ).lower()
                if any(mc in title or mc in client_name for mc in matching_clients):
                    if client_name:
                        referenced_clients.add(client_name)
        
        filtered = []
        for r in rows:
            sheet = r.get("sheet", "")
            title = (r.get("title") or "").lower()
            payload = r.get("payload") or {}
            
            if sheet in ("Clients", "Matters"):
                client_name = title
                if any(mc in client_name for mc in matching_clients) or \
                   any(rc in client_name for rc in referenced_clients):
                    filtered.append(r)
            elif sheet in ("Dockets", "Ledger", "Ledger (A/R Deduced)"):
                client_name = (
                    self._clean(payload.get("clientName")) or
                    self._clean(payload.get("client")) or
                    self._clean(payload.get("Client")) or
                    self._clean(payload.get("Sub-Client")) or
                    ""
                ).lower()
                if any(mc in client_name for mc in matching_clients):
                    filtered.append(r)
            else:
                filtered.append(r)
        
        return filtered

    def _analysis_summary(self, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
        print(f"DEBUG rows: {[{'sheet': r.get('sheet'), 'action': r.get('action')} for r in rows]}")
        summary = {
            "totalRows": len(rows),
            "selectableRows": sum(1 for r in rows if r.get("action") != "skip"),
            "selectedRows": sum(1 for r in rows if r.get("action") != "skip"),
            "newRows": 0,
            "alreadyImportedRows": 0,
            "conflictRows": 0,
            "skippedRows": 0,
            "breakdown": {
                "Clients": 0,
                "Matters": 0,
                "Dockets": 0,
                "Ledger (Income/Expenses)": 0,
                "Invoices / A-R": 0
            }
        }
        for row in rows:
            action = row.get("action", "")
            sheet = row.get("sheet", "")
            if action == "add":
                summary["newRows"] += 1
                if sheet == "Clients":
                    summary["breakdown"]["Clients"] += 1
                elif sheet == "Matters":
                    summary["breakdown"]["Matters"] += 1
                elif sheet == "Dockets":
                    summary["breakdown"]["Dockets"] += 1
                elif sheet == "Ledger":
                    summary["breakdown"]["Ledger (Income/Expenses)"] += 1
                elif sheet.startswith("Ledger (A/R"):
                    summary["breakdown"]["Invoices / A-R"] += 1
            elif action == "skip":
                summary["alreadyImportedRows"] += 1
            elif action in ("conflict", "changed"):
                summary["conflictRows"] += 1
            else:
                summary["skippedRows"] += 1
        return summary

    def _analyze_clients(self, wb, existing_clients, existing_parents, rows, warnings):
        sheet_rows = self._parse_sheet(wb, "Clients")
        for i, row in enumerate(sheet_rows):
            try:
                raw_name = self._legacy_client_name_from_row(row)
                if not raw_name:
                    continue
                cspm_client = self._resolve_client_name(raw_name, existing_clients)
                
                key = self._client_duplicate_key(cspm_client)
                existing = self._duplicate_maps.get("client", {}).get(key)
                
                rows.append({
                    "sheet": "Clients",
                    "row": i + 2,
                    "action": "skip" if existing else "add",
                    "title": raw_name,
                    "details": "Client already exists" if existing else "New client",
                    "payload": dict(row),
                    "client": raw_name,
                    "matter": "",
                    "existingId": existing.get("id", "") if existing else ""
                })
            except Exception as e:
                warnings.append(f"Clients row {i+2} error: {e}")

    def _analyze_matters(self, wb, existing_clients, existing_parents, rows, warnings):
        sheet_rows = self._parse_sheet(wb, "Matters")
        for i, row in enumerate(sheet_rows):
            try:
                raw_name = self._legacy_client_name_for_id(row.get("Client_ID"))
                matter_id_legacy_raw = self._clean(row.get("Matter_ID"))
                matter_id_legacy = self._normalise_legacy_id(matter_id_legacy_raw)
                description = self._clean(row.get("Description (Internal)"))
                matter_name = description if description else matter_id_legacy_raw
                
                if not raw_name or not matter_name:
                    continue
                cspm_client = self._resolve_client_name(raw_name, existing_clients)
                
                key = self._matter_duplicate_key(cspm_client, matter_name)
                num_key = f"num:{matter_id_legacy_raw.lower()}" if matter_id_legacy_raw else ""
                if num_key and num_key in self._duplicate_maps.get("matter", {}) and key not in self._duplicate_maps.get("matter", {}):
                    key = num_key
                existing = self._duplicate_maps.get("matter", {}).get(key)
                action = "skip" if existing else "add"
                
                if matter_id_legacy:
                    self._remember_legacy_matter_mapping(
                        matter_id_legacy,
                        matter_id=self._clean((existing or {}).get("id")),
                        matter_name=matter_name,
                        client_name=cspm_client,
                        parent_name="",
                    )

                rows.append({
                    "sheet": "Matters",
                    "row": i + 2,
                    "action": action,
                    "title": f"Matter: {matter_name} ({raw_name})",
                    "details": "Matter already exists" if existing else f"New matter for {cspm_client or raw_name}",
                    "payload": dict(row),
                    "client": cspm_client or raw_name,
                    "matter": matter_name,
                    "existingId": existing.get("id", "") if isinstance(existing, dict) else ""
                })
            except Exception as e:
                warnings.append(f"Matters row {i+2} error: {e}")

    def _analyze_dockets(self, wb, mode, start_date, end_date, existing_clients, existing_parents, rows, warnings):
        sheet_rows = self._parse_sheet(wb, "Dockets")
        for i, row in enumerate(sheet_rows):
            try:
                raw_client = self._clean(row.get("Client"))
                raw_subclient = self._clean(row.get("Sub-Client"))
                raw_name = raw_subclient if raw_subclient else raw_client
                if not raw_name:
                    continue
                docket_date = row.get("Date")
                if not self._passes_date_filter(docket_date, mode, start_date, end_date):
                    continue

                hours = self._safe_float(row.get("Time (in hrs)"))
                raw_seconds = hours * 3600.0

                matter_id_legacy_raw = self._clean(row.get("Matter_ID"))
                matter_id_legacy = self._normalise_legacy_id(matter_id_legacy_raw)
                matter_mapping = self._legacy_matter_map.get(matter_id_legacy, {})
                
                cspm_client, cspm_parent = self._map_client_parent(
                    raw_client, raw_subclient, existing_clients, existing_parents
                )
                if matter_mapping:
                    cspm_client = self._clean(matter_mapping.get("clientName")) or cspm_client
                    
                if matter_mapping:
                    matter_name = self._clean(matter_mapping.get("matterName"))
                elif matter_id_legacy:
                    matter_name = f"{LEGACY_UNMAPPED_MATTER_PREFIX} {matter_id_legacy_raw}"
                else:
                    matter_name = LEGACY_UNASSIGNED_MATTER_NAME

                payload = {
                    "date": self._format_date(docket_date),
                    "clientName": cspm_client,
                    "matterName": matter_name,
                    "description": self._clean(row.get("Description")),
                    "hours": hours,
                    "rawSeconds": raw_seconds,
                    "clientRate": self._safe_float(row.get("Hourly Rate/Flat Fee")),
                    "sharePct": self._safe_float(row.get("Percentage")),
                    "amountToYou": self._safe_float(row.get("Amount to CS")),
                }
                key = self._time_duplicate_key(payload)
                existing = self._duplicate_maps.get("time", {}).get(key)
                core_key = self._time_core_duplicate_key(payload)
                core_match_count = int(self._time_core_duplicate_counts.get(core_key, 0))
                core_match = core_match_count > 0
                if core_match:
                    # Consume one target occurrence. This preserves genuine
                    # repeated entries: a second identical legacy row remains
                    # importable if CSPM has only one matching occurrence.
                    self._time_core_duplicate_counts[core_key] -= 1
                action = "skip" if (existing or core_match) else "add"
                
                rows.append({
                    "sheet": "Dockets",
                    "row": i + 2,
                    "action": action,
                    "title": self._clean(row.get("Description")) or f"Docket for {raw_name}",
                    "details": "Duplicate docket" if (existing or core_match) else f"Date: {self._format_date(docket_date)}",
                    "payload": dict(row),
                    "client": cspm_client,
                    "matter": matter_name,
                    # The QML safe-selection action may select only rows that
                    # have passed this occurrence-aware core reconciliation.
                    "safeDocketCandidate": not core_match and not existing,
                    "safeDocketHours": hours,
                    "safeDocketAmount": self._safe_float(row.get("Amount to CS")),
                })
            except Exception as e:
                warnings.append(f"Dockets row {i+2} error: {e}")

    def _analyze_ledger(self, wb, mode, start_date, end_date, existing_clients, rows, warnings):
        sheet_rows = self._parse_sheet(wb, "Ledger")
        for i, row in enumerate(sheet_rows):
            try:
                client_vendor = self._clean(row.get("Client/Vendor"))
                if not client_vendor:
                    continue
                txn_date = row.get("Date")
                if not self._passes_date_filter(txn_date, mode, start_date, end_date):
                    continue

                billings = self._safe_float(row.get("Billings (excl. HST)"))
                hst_collected = self._safe_float(row.get("HST Collected"))
                expenses = self._safe_float(row.get("Expenses (excl. HST)"))
                hst_paid = self._safe_float(row.get("HST Paid"))
                collected = self._safe_float(row.get("Collected"))
                write_off = self._safe_float(row.get("Write Off"))

                if billings > 0:
                    txn_type, txn_class, amount, tax_amount = "Income", "Business", billings, hst_collected
                    sub_type = "Billings"
                elif expenses > 0:
                    txn_type, txn_class, amount, tax_amount = "Expense", "Business", expenses, hst_paid
                    sub_type = "Expenses"
                elif collected > 0:
                    txn_type, txn_class, amount, tax_amount = "Transfer", "Business", collected, 0
                    sub_type = "Payments"
                elif write_off > 0:
                    txn_type, txn_class, amount, tax_amount = "Adjustment", "Business", -write_off, 0
                    sub_type = "Write-offs"
                else:
                    txn_type, txn_class, amount, tax_amount = "", "", 0.0, 0.0
                    sub_type = ""

                cspm_client = self._resolve_client_name(client_vendor, existing_clients) if client_vendor else ""

                payload = {
                    "txnDate": self._format_date(txn_date),
                    "class": txn_class,
                    "type": txn_type,
                    "subType": sub_type,
                    "client": cspm_client,
                    "amount": amount,
                    "taxAmount": tax_amount,
                    "notes": self._clean(row.get("Description")),
                    "invoiceRef": self._clean(row.get("Reference")),
                }
                key = self._transaction_duplicate_key(payload)
                existing = self._duplicate_maps.get("transaction", {}).get(key)
                action = "skip" if existing else "add"

                rows.append({
                    "sheet": "Ledger",
                    "row": i + 2,
                    "action": action,
                    "title": f"Ledger for {client_vendor}",
                    "details": "Duplicate ledger entry" if existing else f"Date: {self._format_date(txn_date)}",
                    "payload": dict(row)
                })
            except Exception as e:
                warnings.append(f"Ledger row {i+2} error: {e}")

    # ── Pass 1: Clients ───────────────────────────────────────────────────────

    def _import_clients(self, wb, existing_clients, existing_parents, results, cb, duplicate_callback, allowed_rows=None):
        rows = self._parse_sheet(wb, "Clients")
        total = len(rows)
        for i, row in enumerate(rows):
            try:
                raw_name = self._legacy_client_name_from_row(row)
                if not raw_name:
                    self._emit(cb, "Clients", i + 1, total)
                    continue
                cspm_client = self._resolve_client_name(raw_name, existing_clients)

                client_id_legacy = self._normalise_legacy_id(row.get("Client_ID"))
                parent_id_legacy = self._normalise_legacy_id(row.get("Parent_ID"))
                parent_name = ""
                if parent_id_legacy and parent_id_legacy != client_id_legacy:
                    raw_parent = self._legacy_client_name_for_id(parent_id_legacy)
                    if raw_parent:
                        parent_name = self._resolve_parent_name(raw_parent, existing_parents)

                entity_type = self._clean(row.get("Entity Type"))
                first_name, middle_name, last_name = self._legacy_client_name_parts_from_row(
                    row,
                    cspm_client,
                    entity_type,
                )
                payload = {
                    "clientName": cspm_client,
                    "parentClientName": parent_name,
                    "entityType": entity_type,
                    "firstName": first_name,
                    "middleName": middle_name,
                    "lastName": last_name,
                    "principalName": self._clean(row.get("Principal")),
                    "principalPosition": self._clean(row.get("Position")),
                    "addressLine1": self._clean(row.get("Mailing Address")),
                    "city": self._clean(row.get("City")),
                    "stateProvince": self._clean(row.get("Province")),
                    "postalCode": self._clean(row.get("Postal Code")),
                    "primaryEmail": self._clean(row.get("Email")),
                    "primaryPhone": self._clean(row.get("Phone Number")),
                    "notes": self._clean(row.get("Comments")),
                    "status": "Active",
                    "active": 1,
                }

                std_rate = self._clean(row.get("Standard Rate"))
                pct = self._clean(row.get("Percentage"))
                extra_notes = []
                if std_rate:
                    extra_notes.append(f"Legacy Rate: {std_rate}")
                if pct:
                    extra_notes.append(f"Legacy %: {pct}")
                if extra_notes:
                    existing_notes = payload["notes"]
                    if existing_notes:
                        payload["notes"] = existing_notes + " | " + " | ".join(extra_notes)
                    else:
                        payload["notes"] = " | ".join(extra_notes)

                key = self._client_duplicate_key(cspm_client)
                if allowed_rows is not None and (i + 2) not in allowed_rows.get("Clients", []):
                    duplicate_existing = self._duplicate_maps.get("client", {}).get(key)
                    existing_id = self._clean((duplicate_existing or {}).get("id"))
                    if client_id_legacy and existing_id:
                        self._legacy_client_id_to_cspm_id[client_id_legacy] = existing_id
                    self._emit(cb, "Clients", i + 1, total)
                    continue

                action, duplicate_existing = self._resolve_duplicate_action(
                    "client",
                    key,
                    "Clients",
                    i + 2,
                    payload,
                    results,
                    duplicate_callback,
                )
                if action == DUPLICATE_ACTION_SKIP:
                    existing_id = self._clean((duplicate_existing or {}).get("id"))
                    if client_id_legacy and existing_id:
                        self._legacy_client_id_to_cspm_id[client_id_legacy] = existing_id
                    self._emit(cb, "Clients", i + 1, total)
                    continue
                if action == DUPLICATE_ACTION_OVERWRITE:
                    existing_id = self._clean((duplicate_existing or {}).get("id"))
                    if existing_id:
                        payload["clientId"] = existing_id
                elif duplicate_existing:
                    payload["forceDuplicate"] = True

                result = self.excel_repo.save_client_profile(payload)
                if result.get("ok"):
                    results["clientsAdded"] += 1
                    results["addedRecords"].append(f"Client: {cspm_client}")
                    self._apply_duplicate_write_result(results, action, duplicate_existing, True)
                    self._register_saved_duplicate_key("client", key, payload, result)
                    saved_client_id = self._clean(result.get("clientId") or payload.get("clientId"))
                    if client_id_legacy and saved_client_id:
                        self._legacy_client_id_to_cspm_id[client_id_legacy] = saved_client_id
                else:
                    msg = result.get("message", "Unknown error saving client")
                    results["warnings"].append(f"Client '{cspm_client}': {msg}")
            except Exception as e:
                results["warnings"].append(f"Client row error: {e}")
            self._emit(cb, "Clients", i + 1, total)

    # ── Pass 2: Matters ───────────────────────────────────────────────────────

    def _import_matters(self, wb, existing_clients, existing_parents, results, cb, duplicate_callback, allowed_rows=None):
        rows = self._parse_sheet(wb, "Matters")
        total = len(rows)
        for i, row in enumerate(rows):
            try:
                matter_id_legacy_raw = self._clean(row.get("Matter_ID"))
                matter_id_legacy = self._normalise_legacy_id(matter_id_legacy_raw)
                raw_subclient = self._clean(row.get("Sub-Client"))
                description = self._clean(row.get("Description (Internal)"))
                status = self._clean(row.get("Status")) or "Open"
                open_date = self._format_date(row.get("Open_Date"))
                matter_type = self._clean(row.get("Type")) or "General"
                billing_name = self._clean(row.get("Billing_Name"))
                matter_rate = self._safe_float(row.get("Matter_Rate"))
                matter_pct = self._safe_float(row.get("Matter_Percent"))

                # Resolve client name: prefer Sub-Client, then Client_ID lookup, then Billing_Name
                cspm_client = ""
                cspm_parent = ""
                client_id_legacy = self._normalise_legacy_id(row.get("Client_ID"))
                parent_id_legacy = self._normalise_legacy_id(row.get("Parent_ID"))

                if raw_subclient:
                    cspm_client = self._resolve_client_name(raw_subclient, existing_clients)
                else:
                    legacy_client_name = self._legacy_client_name_for_id(client_id_legacy)
                    if legacy_client_name:
                        cspm_client = self._resolve_client_name(legacy_client_name, existing_clients)

                if not cspm_client and billing_name:
                    cspm_client = self._resolve_client_name(billing_name, existing_clients)

                if parent_id_legacy and parent_id_legacy != client_id_legacy:
                    legacy_parent_name = self._legacy_client_name_for_id(parent_id_legacy)
                    if legacy_parent_name:
                        cspm_parent = self._resolve_parent_name(legacy_parent_name, existing_parents)

                if not cspm_client:
                    results["warnings"].append(f"Matter '{matter_id_legacy}': no client name resolved, skipping")
                    self._emit(cb, "Matters", i + 1, total)
                    continue

                matter_name = description if description else matter_id_legacy_raw

                payload = {
                    "matterNumber": matter_id_legacy_raw,
                    "matterName": matter_name,
                    "clientName": cspm_client,
                    "parentName": cspm_parent,
                    "matterType": matter_type,
                    "status": status,
                    "dateOpened": open_date,
                    "description": description,
                    "defaultRate": matter_rate,
                    "defaultSharePct": matter_pct,
                    "billingContact": billing_name,
                }
                mapped_client_id = self._legacy_client_id_to_cspm_id.get(client_id_legacy)
                if mapped_client_id:
                    payload["clientId"] = mapped_client_id

                key = self._matter_duplicate_key(cspm_client, matter_name)
                num_key = f"num:{matter_id_legacy_raw.lower()}" if matter_id_legacy_raw else ""
                if num_key and num_key in self._duplicate_maps.get("matter", {}) and key not in self._duplicate_maps.get("matter", {}):
                    key = num_key

                if allowed_rows is not None and (i + 2) not in allowed_rows.get("Matters", []):
                    duplicate_existing = self._duplicate_maps.get("matter", {}).get(key)
                    self._remember_legacy_matter_mapping(
                        matter_id_legacy,
                        matter_id=self._clean((duplicate_existing or {}).get("id")),
                        matter_name=matter_name,
                        client_name=cspm_client,
                        parent_name=cspm_parent,
                    )
                    self._emit(cb, "Matters", i + 1, total)
                    continue

                action, duplicate_existing = self._resolve_duplicate_action(
                    "matter",
                    key,
                    "Matters",
                    i + 2,
                    payload,
                    results,
                    duplicate_callback,
                )
                if action == DUPLICATE_ACTION_SKIP:
                    self._remember_legacy_matter_mapping(
                        matter_id_legacy,
                        matter_id=self._clean((duplicate_existing or {}).get("id")),
                        matter_name=matter_name,
                        client_name=cspm_client,
                        parent_name=cspm_parent,
                    )
                    self._emit(cb, "Matters", i + 1, total)
                    continue
                if action == DUPLICATE_ACTION_OVERWRITE:
                    existing_id = self._clean((duplicate_existing or {}).get("id"))
                    if existing_id:
                        payload["matterId"] = existing_id
                elif duplicate_existing:
                    payload["forceDuplicate"] = True

                result = self.excel_repo.save_matter_profile(payload)
                if result.get("ok"):
                    results["mattersAdded"] += 1
                    results["addedRecords"].append(f"Matter: {matter_name}")
                    self._apply_duplicate_write_result(results, action, duplicate_existing, True)
                    self._register_saved_duplicate_key("matter", key, payload, result)
                    if num_key and num_key != key:
                        self._register_saved_duplicate_key("matter", num_key, payload, result)
                    saved_row = dict(result.get("savedRow") or {})
                    self._remember_legacy_matter_mapping(
                        matter_id_legacy,
                        matter_id=self._clean(result.get("matterId") or payload.get("matterId")),
                        matter_name=self._clean(saved_row.get(sc.COL_MATTER_NAME)) or matter_name,
                        client_name=self._clean(saved_row.get(sc.COL_MATTER_CLIENT_NAME)) or cspm_client,
                        parent_name=self._clean(saved_row.get(sc.COL_MATTER_PARENT_NAME)) or cspm_parent,
                    )
                else:
                    msg = result.get("message", "Unknown error saving matter")
                    results["warnings"].append(f"Matter '{matter_name}': {msg}")
            except Exception as e:
                results["warnings"].append(f"Matter row error: {e}")
            self._emit(cb, "Matters", i + 1, total)

    # ── Pass 3: Dockets (Time Entries) ────────────────────────────────────────

    def _import_dockets(self, wb, mode, start_date, end_date, existing_clients, existing_parents, results, cb, duplicate_callback, allowed_rows=None):
        rows = self._parse_sheet(wb, "Dockets")
        total = len(rows)
        for i, row in enumerate(rows):
            if allowed_rows is not None and (i + 2) not in allowed_rows.get("Dockets", []):
                self._emit(cb, "Dockets", i + 1, total)
                continue
            try:
                row_date = row.get("Date")
                if not self._passes_date_filter(row_date, mode, start_date, end_date):
                    self._emit(cb, "Dockets", i + 1, total)
                    continue

                raw_client = self._clean(row.get("Client"))
                raw_subclient = self._clean(row.get("Sub-Client"))
                matter_id_legacy_raw = self._clean(row.get("Matter_ID"))
                matter_id_legacy = self._normalise_legacy_id(matter_id_legacy_raw)
                matter_mapping = self._legacy_matter_map.get(matter_id_legacy, {})
                cspm_client, cspm_parent = self._map_client_parent(
                    raw_client, raw_subclient, existing_clients, existing_parents
                )
                if matter_mapping:
                    cspm_client = self._clean(matter_mapping.get("clientName")) or cspm_client
                    cspm_parent = self._clean(matter_mapping.get("parentName")) or cspm_parent
                if not cspm_client:
                    self._emit(cb, "Dockets", i + 1, total)
                    continue

                description = self._clean(row.get("Description"))
                hours = self._safe_float(row.get("Time (in hrs)"))
                rate = self._safe_float(row.get("Hourly Rate/Flat Fee"))
                share_pct = self._safe_float(row.get("Percentage"))
                invoice_ref = self._clean(row.get("Invoice"))
                
                raw_seconds = hours * 3600.0

                invoice_bucket = self._invoice_marker_bucket(invoice_ref)
                status = "Draft"
                if invoice_bucket in {"invoice", "legacy_billed", "no_bill"}:
                    status = "Billed"
                payment_state = self._payment_state_for_invoice(wb, invoice_ref)
                if matter_mapping:
                    matter_name = self._clean(matter_mapping.get("matterName"))
                elif matter_id_legacy:
                    matter_name = f"{LEGACY_UNMAPPED_MATTER_PREFIX} {matter_id_legacy_raw}"
                    if matter_id_legacy not in self._legacy_unmapped_warning_ids:
                        self._legacy_unmapped_warning_ids.add(matter_id_legacy)
                        results["warnings"].append(
                            f"Docket Matter_ID '{matter_id_legacy_raw}' was not found in the imported Matters "
                            f"mapping and was assigned to '{matter_name}' for review."
                        )
                else:
                    matter_name = LEGACY_UNASSIGNED_MATTER_NAME
                    warning_key = cspm_client.lower()
                    if warning_key not in self._legacy_unassigned_warning_clients:
                        self._legacy_unassigned_warning_clients.add(warning_key)
                        results["warnings"].append(
                            f"Dockets for '{cspm_client}' without Matter_ID were assigned to "
                            f"'{LEGACY_UNASSIGNED_MATTER_NAME}' for review."
                        )

                payload = {
                    "clientName": cspm_client,
                    "parentName": cspm_parent,
                    "matterName": matter_name,
                    "date": self._format_date(row_date),
                    "description": description,
                    "hours": hours,
                    "clientRate": rate,
                    "sharePct": share_pct,
                    "rawSeconds": raw_seconds,
                    "status": status,
                    "forceDuplicate": True,
                    "useExactHours": True,
                    "invoiceRef": invoice_ref,
                    "invoiceStatus": self._invoice_status_for_marker(invoice_ref),
                    "paymentStatus": payment_state.get("paymentStatus", ""),
                    "invoiceTotal": payment_state.get("invoiceTotal", 0.0),
                    "invoiceAmountPaid": payment_state.get("invoiceAmountPaid", 0.0),
                    "invoiceBalanceDue": payment_state.get("invoiceBalanceDue", 0.0),
                    "invoiceDate": payment_state.get("invoiceDate", ""),
                }

                key = self._time_duplicate_key(payload)
                action, duplicate_existing = self._resolve_duplicate_action(
                    "time",
                    key,
                    "Dockets",
                    i + 2,
                    payload,
                    results,
                    duplicate_callback,
                )
                if action == DUPLICATE_ACTION_SKIP:
                    self._emit(cb, "Dockets", i + 1, total)
                    continue
                if action == DUPLICATE_ACTION_OVERWRITE:
                    existing_id = self._clean((duplicate_existing or {}).get("id"))
                    if existing_id:
                        payload["entryId"] = existing_id
                        payload["rawSecondsMode"] = "replace"
                elif duplicate_existing:
                    payload["forceDuplicate"] = True

                result = self.excel_repo.add_time_entry(payload)
                if result.get("ok"):
                    results["docketsAdded"] += 1
                    results["addedRecords"].append(f"Time Entry: {self._duplicate_label('time', payload)}")
                    self._apply_duplicate_write_result(results, action, duplicate_existing, True)
                    self._register_saved_duplicate_key("time", key, payload, result)
                else:
                    msg = result.get("message", "Unknown error")
                    results["warnings"].append(f"Docket '{description[:40]}': {msg}")
            except Exception as e:
                results["warnings"].append(f"Docket row error: {e}")
            self._emit(cb, "Dockets", i + 1, total)

    # ── Pass 4: Disbursements ─────────────────────────────────────────────────

    def _import_ledger(self, wb, mode, start_date, end_date, existing_clients, results, cb, duplicate_callback, allowed_rows=None):
        rows = self._parse_sheet(wb, "Ledger")
        total = len(rows)
        for i, row in enumerate(rows):
            if allowed_rows is not None and (i + 2) not in allowed_rows.get("Ledger", []):
                self._emit(cb, "Ledger", i + 1, total)
                continue
            try:
                row_date = row.get("Date")
                if not self._passes_date_filter(row_date, mode, start_date, end_date):
                    self._emit(cb, "Ledger", i + 1, total)
                    continue

                client_vendor = self._clean(row.get("Client/Vendor"))
                description = self._clean(row.get("Description"))
                category = self._clean(row.get("Category"))
                reference = self._clean(row.get("Reference"))
                billings = self._safe_float(row.get("Billings (excl. HST)"))
                hst_collected = self._safe_float(row.get("HST Collected"))
                expenses = self._safe_float(row.get("Expenses (excl. HST)"))
                hst_paid = self._safe_float(row.get("HST Paid"))
                collected = self._safe_float(row.get("Collected"))
                write_off = self._safe_float(row.get("Write Off"))

                if billings > 0:
                    txn_type, txn_class, amount, tax_amount = "Income", "Business", billings, hst_collected
                elif expenses > 0:
                    txn_type, txn_class, amount, tax_amount = "Expense", "Business", expenses, hst_paid
                elif collected > 0:
                    txn_type, txn_class, amount, tax_amount = "Transfer", "Business", collected, 0
                elif write_off > 0:
                    txn_type, txn_class, amount, tax_amount = "Adjustment", "Business", -write_off, 0
                else:
                    self._emit(cb, "Ledger", i + 1, total)
                    continue

                cspm_client = self._resolve_client_name(client_vendor, existing_clients) if client_vendor else ""

                
                # Adjust Type for "Adjustment"
                if txn_type == "Adjustment":
                    txn_type = "Expense" if amount and amount < 0 else "Income"

                payload = {
                    "txnDate": self._format_date(row_date),
                    "class": txn_class,
                    "type": txn_type,
                    "client": cspm_client,
                    "categoryName": category,
                    "categoryCode": "INC_LEGAL_FEES" if txn_type == "Income" else "EXP_OFFICE_SUPPLIES",
                    "amount": abs(amount) if amount else 0.0,
                    "taxAmount": abs(tax_amount) if tax_amount else 0.0,
                    "notes": description,
                    "invoiceRef": reference,
                    "status": "Cleared",
                    "fromAccount": "CIBC_CHEQUING",
                    "toAccount": "CIBC_CHEQUING" if txn_type not in ["Transfer", "Debt Repayment"] else "AMEX",
                    "payee": "Costco", # Dummy payee to satisfy validation
                }
                
                payload["businessUnit"] = "Cory Business"

                result = self._save_transaction_with_duplicate_decision(
                    payload,
                    "Ledger",
                    i + 2,
                    results,
                    duplicate_callback,
                )
                if result.get("ok"):
                    results["ledgerAdded"] += 1
                elif not result.get("skippedDuplicate"):
                    msg = result.get("message", "Unknown error")
                    results["warnings"].append(f"Ledger '{description[:40]}': {msg}")
            except Exception as e:
                results["warnings"].append(f"Ledger row error: {e}")
            self._emit(cb, "Ledger", i + 1, total)

    def _aggregate_ledger_invoices(self, wb, mode, start_date, end_date, existing_clients, allowed_rows=None) -> List[Dict[str, Any]]:
        ledger_rows = self._parse_sheet(wb, "Ledger")
        invoices_map = {}
        
        for i, row in enumerate(ledger_rows):
            if allowed_rows is not None and (i + 2) not in allowed_rows.get("Ledger", []):
                continue
            row_date = row.get("Date")
            if not self._passes_date_filter(row_date, mode, start_date, end_date):
                continue

            reference = self._clean(row.get("Reference"))
            if not reference:
                continue

            client_vendor = self._clean(row.get("Client/Vendor"))
            cspm_client = self._resolve_client_name(client_vendor, existing_clients) if client_vendor else ""

            billings = self._safe_float(row.get("Billings (excl. HST)"))
            hst_collected = self._safe_float(row.get("HST Collected"))
            expenses = self._safe_float(row.get("Expenses (excl. HST)"))
            hst_paid = self._safe_float(row.get("HST Paid"))
            collected = self._safe_float(row.get("Collected"))
            write_off = self._safe_float(row.get("Write Off"))

            if reference not in invoices_map:
                invoices_map[reference] = {
                    "invoiceNum": reference,
                    "date": self._format_date(row_date),
                    "client": cspm_client,
                    "totalInvoiced": 0.0,
                    "totalFees": 0.0,
                    "totalDisbursements": 0.0,
                    "totalTax": 0.0,
                    "amountPaid": 0.0,
                    "creditsAdj": 0.0
                }
            
            inv = invoices_map[reference]
            
            if billings > 0:
                inv["totalInvoiced"] += billings + hst_collected
                inv["totalFees"] += billings
                inv["totalTax"] += hst_collected
            if expenses > 0:
                inv["totalInvoiced"] += expenses + hst_paid
                inv["totalDisbursements"] += expenses
                inv["totalTax"] += hst_paid
            if collected > 0:
                inv["amountPaid"] += collected
            if write_off > 0:
                inv["creditsAdj"] += write_off

        results = []
        for ref, inv in invoices_map.items():
            bal = round(inv["totalInvoiced"] - inv["amountPaid"] - inv["creditsAdj"], 2)
            inv["balanceDue"] = bal
            inv["status"] = "Open" if bal > 0 else "Closed"
            inv["clientName"] = inv["client"]
            inv["subClient"] = inv["client"]
            inv["invoiceDate"] = inv["date"]
            inv["aggregateBilled"] = inv["totalInvoiced"]
            inv["billToClient"] = inv["client"]
            inv["workClient"] = inv["client"]
            results.append(inv)
            
        return results

    def export_analysis_to_excel(self, rows: List[Dict[str, Any]], export_dir: str) -> str:
        import os
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from datetime import datetime

        wb = Workbook()
        ws = wb.active
        ws.title = "Analysis Review"

        headers = ["Row", "Sheet", "Action", "Title", "Details", "Payload"]
        ws.append(headers)

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for row in rows:
            sheet_name = str(row.get("sheet", ""))
            row_num = row.get("row", "")
            action = str(row.get("action", ""))
            title = str(row.get("title", ""))
            details = str(row.get("details", ""))
            payload = str(row.get("payload", ""))

            ws.append([row_num, sheet_name, action, title, details, payload])

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Legacy_Dockets_Analysis_{timestamp}.xlsx"
        filepath = os.path.join(export_dir, filename)
        
        os.makedirs(export_dir, exist_ok=True)
        wb.save(filepath)
        return filepath
