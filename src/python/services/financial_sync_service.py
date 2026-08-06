from __future__ import annotations

"""Auditable, source-controlled synchronization from historic Dockets.xlsm.

The legacy workbook is a complete historical snapshot, not an incremental feed.
This service therefore works in two deliberately separate phases:

* ``preview`` reads both workbooks and produces a deterministic plan.  It never
  writes either workbook.
* ``build_candidate`` copies CSPM to an isolated candidate, applies the plan to
  that copy, and proves the candidate against the source totals.  Promotion of a
  candidate is a separate, explicit operation.

The service purposely does not call the generic legacy importer.  That importer
is useful for one-off data intake but it must infer receivables from a ledger;
for a synchronization the source Receivables, Invoice Log, A/P Bills and A/P
Payments tables are the financial authority.
"""

from collections import Counter, defaultdict, deque
from copy import deepcopy
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import hashlib
import json
import os
from pathlib import Path
import re
import shutil
import tempfile
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from domain import schema_constants as sc
from domain.ap_schema import (
    AP_BILLS_HEADERS,
    AP_BILLS_SHEET,
    AP_BILLS_TABLE,
    AP_PAYMENTS_HEADERS,
    AP_PAYMENTS_SHEET,
    AP_PAYMENTS_TABLE,
)
from repositories.excel_repo import ExcelRepo
from services.paths import AppPaths
from services.workbook_integrity_service import WorkbookIntegrityService


MONEY_TOLERANCE = Decimal("0.02")
HOURS_TOLERANCE = Decimal("0.01")
INVOICE_RE = re.compile(r"^\d{2}-\d{4}(?:-[A-Z])?$")

RAW_DOCKETS_SHEET = "Dockets"
RAW_DOCKETS_TABLE = "tblDockets"
RAW_DOCKETS_HEADERS = [
    "Date", "Client", "Matter", "Parent", "Description",
    "Time (in hrs) or Units", "Hourly Rate/Flat Rate", "Percentage",
    "Amount to CS", "Total Inclusive of HST", "Invoice #", "RawSeconds",
    "EntryType",
]


class FinancialSyncError(RuntimeError):
    """A synchronization gate failed before a live workbook could be changed."""


def _text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"none", "nan"} else re.sub(r"\s+", " ", text)


def _key(value: Any) -> str:
    return _text(value).casefold()


def _date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text:
        return None
    for candidate in (text[:10], text):
        try:
            return datetime.fromisoformat(candidate.replace("Z", "+00:00")).date()
        except ValueError:
            pass
    for fmt in ("%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%B %d, %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _iso_date(value: Any) -> str:
    parsed = _date(value)
    return parsed.isoformat() if parsed else _text(value)


def _money(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0.00")
    text = str(value).replace("$", "").replace(",", "").strip()
    try:
        return Decimal(text).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def _number(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    text = str(value).replace("$", "").replace(",", "").strip()
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _money_float(value: Any) -> float:
    return float(_money(value))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _stamp() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _stable_id(prefix: str, *parts: Any) -> str:
    raw = "|".join(_text(part).casefold() for part in parts).encode("utf-8")
    return f"{prefix}-{hashlib.sha256(raw).hexdigest()[:20].upper()}"


def _is_actual_invoice(value: Any) -> bool:
    return bool(INVOICE_RE.match(_text(value).upper()))


def _invoice_marker_status(value: Any) -> tuple[str, str]:
    marker = _text(value)
    upper = marker.upper()
    if _is_actual_invoice(marker):
        return "Billed", "Invoiced"
    if upper == "BILLED":
        return "Billed", "Legacy Billed"
    if "HOLD" in upper or "FORGOT" in upper:
        return "Draft", "Held for Billing"
    if "FREE" in upper or "DO NOT BILL" in upper or "WRITE OFF" in upper:
        return "Billed", "Not Billable"
    return "Draft", "Not Invoiced"


def _normalise_share(value: Any) -> Decimal:
    share = _number(value)
    if Decimal("0") < abs(share) <= Decimal("1"):
        share *= Decimal("100")
    return share.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _source_table_rows(workbook: Any, sheet: str) -> list[dict[str, Any]]:
    """Read a source worksheet by header so table names may differ by history."""
    if sheet not in workbook.sheetnames:
        return []
    worksheet = workbook[sheet]
    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [_text(value) for value in values[0]]
    rows: list[dict[str, Any]] = []
    for row in values[1:]:
        record = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers)) if headers[index]}
        if any(value not in (None, "") for value in record.values()):
            rows.append(record)
    return rows


def _target_table_rows(workbook: Any, sheet: str, table: str) -> list[dict[str, Any]]:
    if sheet not in workbook.sheetnames or table not in workbook[sheet].tables:
        return []
    worksheet = workbook[sheet]
    table_obj = worksheet.tables[table]
    cells = list(worksheet[table_obj.ref])
    if not cells:
        return []
    headers = [_text(cell.value) for cell in cells[0]]
    rows: list[dict[str, Any]] = []
    for source_row in cells[1:]:
        record = {headers[index]: cell.value for index, cell in enumerate(source_row) if headers[index]}
        if any(value not in (None, "") for value in record.values()):
            rows.append(record)
    return rows


def _read_shadow(source_path: Path) -> Path:
    """Read Excel-open source workbooks without asking the user to close them."""
    fd, raw_path = tempfile.mkstemp(prefix="cspm_financial_sync_", suffix=source_path.suffix or ".xlsm")
    os.close(fd)
    shadow = Path(raw_path)
    try:
        shutil.copy2(source_path, shadow)
        return shadow
    except (PermissionError, OSError):
        try:
            with source_path.open("rb") as incoming, shadow.open("wb") as outgoing:
                shutil.copyfileobj(incoming, outgoing)
            return shadow
        except Exception:
            shadow.unlink(missing_ok=True)
            raise


class FinancialSyncService:
    """Build and validate a financial synchronization candidate.

    A source snapshot owns all legacy financial history through its last docket
    date.  Native CSPM entries strictly after that date are retained.  Any
    historical target-only row is *retired in the candidate* and recorded in
    the plan; it is never silently discarded from the live file because live
    promotion first creates a recoverable backup package.
    """

    SOURCE_SHEETS = (
        "Clients", "Matters", "Dockets", "Disbursements", "Ledger",
        "Receivables", "Invoice Log", "HST_Log", AP_BILLS_SHEET, AP_PAYMENTS_SHEET,
    )

    def __init__(self, paths: AppPaths) -> None:
        self.paths = paths

    @property
    def target_path(self) -> Path:
        return self.paths.workbook_path()

    def _read_source(self, source_path: Path) -> dict[str, list[dict[str, Any]]]:
        if not source_path.is_file():
            raise FinancialSyncError(f"Historic source workbook was not found: {source_path}")
        shadow = _read_shadow(source_path)
        workbook = None
        try:
            workbook = load_workbook(shadow, keep_vba=True, data_only=True)
            return {sheet: _source_table_rows(workbook, sheet) for sheet in self.SOURCE_SHEETS}
        finally:
            if workbook is not None:
                workbook.close()
            shadow.unlink(missing_ok=True)

    def _read_target(self, target_path: Path) -> dict[str, list[dict[str, Any]]]:
        if not target_path.is_file():
            raise FinancialSyncError(f"CSPM workbook was not found: {target_path}")
        workbook = load_workbook(target_path, keep_vba=True, data_only=True, read_only=False)
        try:
            return {
                "Parents": _target_table_rows(workbook, sc.SHEET_PARENTS, sc.TBL_PARENTS),
                "Clients": _target_table_rows(workbook, sc.SHEET_CLIENTS, sc.TBL_CLIENTS),
                "ClientProfiles": _target_table_rows(workbook, sc.SHEET_CLIENT_PROFILES, sc.TBL_CLIENT_PROFILES),
                "Matters": _target_table_rows(workbook, sc.SHEET_MATTERS, sc.TBL_MATTERS),
                "TimeEntries": _target_table_rows(workbook, sc.SHEET_TIME, sc.TBL_TIME),
                "Dockets": _source_table_rows(workbook, RAW_DOCKETS_SHEET),
                "Disbursements": _target_table_rows(workbook, sc.SHEET_DISBURSEMENTS, sc.TBL_DISBURSEMENTS),
                "Ledger": _target_table_rows(workbook, sc.SHEET_LEDGER, sc.TBL_LEDGER),
                "Receivables": _target_table_rows(workbook, sc.SHEET_RECEIVABLES, sc.TBL_RECEIVABLES),
                "Invoice Log": _target_table_rows(workbook, sc.SHEET_INVOICE_LOG, sc.TBL_INVOICE_LOG),
                "HST_Log": _target_table_rows(workbook, sc.SHEET_HST_LOG, sc.TBL_HST_LOG),
                "Transactions": _target_table_rows(workbook, sc.SHEET_TRANSACTIONS, sc.TBL_TRANSACTIONS_MASTER),
                "TransactionAccounts": _target_table_rows(workbook, sc.SHEET_TRANSACTION_ACCOUNTS, sc.TBL_TRANSACTION_ACCOUNTS),
                "TransactionCategories": _target_table_rows(workbook, sc.SHEET_TRANSACTION_CATEGORIES, sc.TBL_TRANSACTION_CATEGORIES),
                AP_BILLS_SHEET: _target_table_rows(workbook, AP_BILLS_SHEET, AP_BILLS_TABLE),
                AP_PAYMENTS_SHEET: _target_table_rows(workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE),
            }
        finally:
            workbook.close()

    @staticmethod
    def _max_date(rows: Iterable[Mapping[str, Any]], column: str) -> date | None:
        parsed = [_date(row.get(column)) for row in rows]
        dates = [value for value in parsed if value is not None]
        return max(dates) if dates else None

    @staticmethod
    def _source_receivables(rows: Sequence[Mapping[str, Any]]) -> dict[str, dict[str, Any]]:
        result: dict[str, dict[str, Any]] = {}
        for row in rows:
            number = _text(row.get("InvoiceNum"))
            if number:
                result[number.upper()] = dict(row)
        return result

    @staticmethod
    def _canonical_source_receivables(source: Mapping[str, Sequence[Mapping[str, Any]]]) -> tuple[list[dict[str, Any]], list[str]]:
        """Normalize legacy A/R without copying its subtotal/summary artefacts.

        Historic Receivables uses the opposite sign for credits and contains a
        few repeated invoice rows created by old adjustments.  We retain the
        source's final A/R state but express it in CSPM's canonical convention
        ``balance = total - paid - credits``.  The incoming ledger and invoice
        log are cross-checks for choosing the representative duplicate row.
        """
        by_invoice: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for raw in source.get("Receivables", []):
            number = _text(raw.get("InvoiceNum")).upper()
            if _is_actual_invoice(number):
                by_invoice[number].append(dict(raw))
        ledger_by_invoice: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
        for row in source.get("Ledger", []):
            number = _text(row.get("Reference")).upper()
            if _is_actual_invoice(number):
                ledger_by_invoice[number].append(row)
        invoice_log_by_invoice: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
        for row in source.get("Invoice Log", []):
            number = _text(row.get("Invoice #")).upper()
            if _is_actual_invoice(number):
                invoice_log_by_invoice[number].append(row)

        normalized: list[dict[str, Any]] = []
        warnings: list[str] = []
        for invoice, candidates in sorted(by_invoice.items()):
            ledger_rows = ledger_by_invoice.get(invoice, [])
            ledger_balance = sum((_money(row.get("Receivable")) for row in ledger_rows), Decimal("0"))
            # Where duplicate historic A/R rows exist, take the one whose stated
            # ending balance agrees with the ledger.  It is stronger evidence
            # than relying on row order in a manually maintained workbook.
            chosen = min(
                candidates,
                key=lambda row: (abs(_money(row.get("Balance_Due")) - ledger_balance), -_money(row.get("Total_Invoiced"))),
            )
            positive_ledger_invoice_total = sum(
                (max(Decimal("0"), _money(row.get("Billings (excl. HST)")) + _money(row.get("HST Collected"))) for row in ledger_rows),
                Decimal("0"),
            )
            invoice_log_total = max(
                (_money(row.get("Aggregate Billed to Client")) for row in invoice_log_by_invoice.get(invoice, [])),
                default=Decimal("0"),
            )
            total = max(_money(chosen.get("Total_Invoiced")), positive_ledger_invoice_total, invoice_log_total)
            paid = max(
                _money(chosen.get("Amount_Paid")),
                sum((_money(row.get("Collected")) for row in ledger_rows), Decimal("0")),
            )
            stated_balance = _money(chosen.get("Balance_Due"))
            source_status = _key(chosen.get("Status"))
            # Historic source can show a few cents of formula residue on paid
            # invoices.  A paid status is definitive; normalize the residue to
            # zero and preserve the original in the audit warning.
            if source_status in {"paid", "closed"} or abs(stated_balance) <= MONEY_TOLERANCE:
                balance = Decimal("0.00")
            else:
                balance = max(Decimal("0.00"), stated_balance)
            credits = (total - paid - balance).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            if len(candidates) > 1:
                warnings.append(f"Consolidated {len(candidates)} historic Receivables rows for {invoice} using ledger-supported state.")
            if abs(stated_balance - balance) > MONEY_TOLERANCE:
                warnings.append(f"Normalized historic A/R residue for {invoice}: stated {stated_balance:.2f}, canonical {balance:.2f}.")
            normalized.append({
                "InvoiceNum": invoice,
                "Date": chosen.get("Date"),
                "Client": _text(chosen.get("Client")),
                "Total_Invoiced": float(total),
                "Amount_Paid": float(paid),
                "Credits/Adj": float(credits),
                "Balance_Due": float(balance),
                "Status": "PAID" if balance == 0 else ("PARTIAL" if paid or credits else "PENDING"),
                "Work Client": _text(chosen.get("Work Client")),
            })
        return normalized, warnings

    def _canonicalize_source(self, source: Mapping[str, Sequence[Mapping[str, Any]]]) -> tuple[dict[str, list[dict[str, Any]]], list[str]]:
        canonical = {name: [dict(row) for row in rows] for name, rows in source.items()}
        receivables, warnings = self._canonical_source_receivables(canonical)
        canonical["Receivables"] = receivables
        return canonical, warnings

    @staticmethod
    def _target_receivables(rows: Sequence[Mapping[str, Any]]) -> dict[str, dict[str, Any]]:
        result: dict[str, dict[str, Any]] = {}
        for row in rows:
            number = _text(row.get(sc.COL_RECV_INVOICE_NUM))
            if number:
                result[number.upper()] = dict(row)
        return result

    @staticmethod
    def _receivable_state(row: Mapping[str, Any], *, source: bool) -> dict[str, Decimal | str]:
        prefix = "" if source else ""
        keys = {
            "total": "Total_Invoiced" if source else sc.COL_RECV_TOTAL_INVOICED,
            "paid": "Amount_Paid" if source else sc.COL_RECV_AMOUNT_PAID,
            "credits": "Credits/Adj" if source else sc.COL_RECV_CREDITS_ADJ,
            "balance": "Balance_Due" if source else sc.COL_RECV_BALANCE_DUE,
            "status": "Status" if source else sc.COL_RECV_STATUS,
        }
        total, paid, credits, balance = (_money(row.get(keys[name])) for name in ("total", "paid", "credits", "balance"))
        return {"total": total, "paid": paid, "credits": credits, "balance": balance, "status": _text(row.get(keys["status"]))}

    @staticmethod
    def _docket_signature(row: Mapping[str, Any], *, source: bool) -> tuple[str, str, str, str]:
        return (
            _iso_date(row.get("Date")),
            _key(row.get("Description")),
            str(_number(row.get("Time (in hrs)" if source else sc.COL_TIME_HOURS)).quantize(Decimal("0.0001"))),
            str(_number(row.get("Hourly Rate/Flat Fee" if source else sc.COL_TIME_RATE)).quantize(Decimal("0.0001"))),
        )

    @staticmethod
    def _source_metrics(source: Mapping[str, Sequence[Mapping[str, Any]]]) -> dict[str, float]:
        docket_rows = source.get("Dockets", [])
        ledger_rows = source.get("Ledger", [])
        receivable_rows = source.get("Receivables", [])
        hours = sum((_number(row.get("Time (in hrs)")) for row in docket_rows), Decimal("0"))
        gross = sum((
            _money(row.get("Amount to CS"))
            if _number(row.get("Time (in hrs)")) == 0 and _money(row.get("Amount to CS")) != 0
            else (_number(row.get("Time (in hrs)")) * _number(row.get("Hourly Rate/Flat Fee")))
            for row in docket_rows
        ), Decimal("0"))
        net = sum((_money(row.get("Amount to CS")) for row in docket_rows), Decimal("0"))
        billings = sum((_money(row.get("Billings (excl. HST)")) for row in ledger_rows), Decimal("0"))
        hst_collected = sum((_money(row.get("HST Collected")) for row in ledger_rows), Decimal("0"))
        expenses = sum((_money(row.get("Expenses (excl. HST)")) for row in ledger_rows), Decimal("0"))
        hst_paid = sum((_money(row.get("HST Paid")) for row in ledger_rows), Decimal("0"))
        collected = sum((_money(row.get("Collected")) for row in ledger_rows), Decimal("0"))
        setoff = sum(
            (_money(row.get("Collected")) for row in ledger_rows
             if "set-off" in _text(row.get("Description")).casefold() or "setoff" in _text(row.get("Description")).casefold()),
            Decimal("0"),
        )
        ar = sum((_money(row.get("Balance_Due")) for row in receivable_rows), Decimal("0"))
        return {
            "productivityHours": float(hours.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            "productivityGross": float(gross.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            "productivityNet": float(net.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            "ledgerBillings": float(billings),
            "ledgerHstCollected": float(hst_collected),
            "ledgerExpenses": float(expenses),
            "ledgerHstPaid": float(hst_paid),
            "ledgerCollected": float(collected),
            "ledgerSetoffCollected": float(setoff),
            "ledgerCashCollected": float(collected - setoff),
            "receivablesAr": float(ar),
            "netHst": float(hst_collected - hst_paid),
        }

    @staticmethod
    def _target_metrics(target: Mapping[str, Sequence[Mapping[str, Any]]]) -> dict[str, float]:
        time_rows = target.get("TimeEntries", [])
        ledger_rows = target.get("Ledger", [])
        receivable_rows = target.get("Receivables", [])
        hours = sum((_number(row.get(sc.COL_TIME_HOURS)) for row in time_rows), Decimal("0"))
        gross = sum((_money(row.get(sc.COL_TIME_GROSS)) for row in time_rows), Decimal("0"))
        net = sum((_money(row.get(sc.COL_TIME_NET)) for row in time_rows), Decimal("0"))
        billings = sum((_money(row.get(sc.COL_LEDGER_BILLINGS_EXCL_HST)) for row in ledger_rows), Decimal("0"))
        hst_collected = sum((_money(row.get(sc.COL_LEDGER_HST_COLLECTED)) for row in ledger_rows), Decimal("0"))
        expenses = sum((_money(row.get(sc.COL_LEDGER_EXPENSES_EXCL_HST)) for row in ledger_rows), Decimal("0"))
        hst_paid = sum((_money(row.get(sc.COL_LEDGER_HST_PAID)) for row in ledger_rows), Decimal("0"))
        collected = sum((_money(row.get(sc.COL_LEDGER_COLLECTED)) for row in ledger_rows), Decimal("0"))
        setoff = sum(
            (_money(row.get(sc.COL_LEDGER_COLLECTED)) for row in ledger_rows
             if "set-off" in _text(row.get(sc.COL_LEDGER_DESCRIPTION)).casefold() or "setoff" in _text(row.get(sc.COL_LEDGER_DESCRIPTION)).casefold()),
            Decimal("0"),
        )
        ar = sum((_money(row.get(sc.COL_RECV_BALANCE_DUE)) for row in receivable_rows), Decimal("0"))
        return {
            "productivityHours": float(hours.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            "productivityGross": float(gross.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            "productivityNet": float(net.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            "ledgerBillings": float(billings),
            "ledgerHstCollected": float(hst_collected),
            "ledgerExpenses": float(expenses),
            "ledgerHstPaid": float(hst_paid),
            "ledgerCollected": float(collected),
            "ledgerSetoffCollected": float(setoff),
            "ledgerCashCollected": float(collected - setoff),
            "receivablesAr": float(ar),
            "netHst": float(hst_collected - hst_paid),
        }

    @staticmethod
    def _metric_deltas(expected: Mapping[str, Any], actual: Mapping[str, Any]) -> dict[str, float]:
        return {key: round(float(actual.get(key, 0) or 0) - float(expected.get(key, 0) or 0), 2)
                for key in sorted(set(expected) | set(actual))}

    def preview(self, source_path: str | Path, target_path: str | Path | None = None) -> dict[str, Any]:
        """Return a source-of-truth plan without modifying either workbook."""
        source_file = Path(source_path)
        target_file = Path(target_path) if target_path else self.target_path
        source_raw = self._read_source(source_file)
        source, source_normalization_warnings = self._canonicalize_source(source_raw)
        target = self._read_target(target_file)
        errors: list[str] = []
        warnings: list[str] = list(source_normalization_warnings)
        actions: dict[str, int] = defaultdict(int)

        source_docket_cutoff = self._max_date(source["Dockets"], "Date")
        source_ledger_cutoff = self._max_date(source["Ledger"], "Date")
        source_invoice_cutoff = max(
            [candidate for candidate in (
                self._max_date(source["Receivables"], "Date"),
                self._max_date(source["Invoice Log"], "Invoice Date"),
            ) if candidate is not None],
            default=None,
        )
        if source_docket_cutoff is None or source_ledger_cutoff is None:
            errors.append("The historic source has no usable Dockets or Ledger cutoff date.")

        source_receivables = self._source_receivables(source["Receivables"])
        target_receivables = self._target_receivables(target["Receivables"])
        receivable_differences: list[dict[str, Any]] = []
        for invoice, source_row in sorted(source_receivables.items()):
            source_state = self._receivable_state(source_row, source=True)
            expected_balance = source_state["total"] - source_state["paid"] - source_state["credits"]
            if abs(expected_balance - source_state["balance"]) > MONEY_TOLERANCE:
                errors.append(
                    f"Source receivable {invoice} does not balance: total less paid/credits differs from balance."
                )
            target_row = target_receivables.get(invoice)
            if target_row is None:
                actions["receivablesAdd"] += 1
                receivable_differences.append({"invoice": invoice, "action": "add", "source": {k: float(v) if isinstance(v, Decimal) else v for k, v in source_state.items()}})
                continue
            target_state = self._receivable_state(target_row, source=False)
            fields = [name for name in ("total", "paid", "credits", "balance")
                      if abs(source_state[name] - target_state[name]) > MONEY_TOLERANCE]
            if fields or _key(source_state["status"]) != _key(target_state["status"]):
                actions["receivablesCorrect"] += 1
                receivable_differences.append({
                    "invoice": invoice, "action": "correct", "fields": fields + (["status"] if _key(source_state["status"]) != _key(target_state["status"]) else []),
                    "source": {k: float(v) if isinstance(v, Decimal) else v for k, v in source_state.items()},
                    "target": {k: float(v) if isinstance(v, Decimal) else v for k, v in target_state.items()},
                })
            else:
                actions["receivablesAlreadyMatch"] += 1

        target_only_receivables: list[str] = []
        for invoice, target_row in target_receivables.items():
            if invoice in source_receivables:
                continue
            item_date = _date(target_row.get(sc.COL_RECV_DATE))
            if (not _is_actual_invoice(invoice)) or (source_docket_cutoff and item_date and item_date <= source_docket_cutoff):
                target_only_receivables.append(invoice)
                actions["receivablesRetire"] += 1
            else:
                actions["receivablesPreserveNative"] += 1

        source_ap = { _text(row.get("APBillID")).upper(): row for row in source[AP_BILLS_SHEET] if _text(row.get("APBillID")) }
        target_ap = { _text(row.get("APBillID")).upper(): row for row in target[AP_BILLS_SHEET] if _text(row.get("APBillID")) }
        source_ap_keys = {
            self._ap_business_key(row): _text(row.get("APBillID"))
            for row in source_ap.values() if self._ap_business_key(row)
        }
        settlement_rows: list[dict[str, Any]] = []
        for bill_id, source_bill in source_ap.items():
            total = _money(source_bill.get("Total"))
            paid = _money(source_bill.get("AmountPaid"))
            balance = _money(source_bill.get("Balance"))
            if abs(total - paid - balance) > MONEY_TOLERANCE:
                errors.append(f"Source A/P bill {bill_id} does not balance.")
            if bill_id not in target_ap:
                actions["apBillsAdd"] += 1
            elif self._ap_equal(source_bill, target_ap[bill_id]):
                actions["apBillsAlreadyMatch"] += 1
            else:
                actions["apBillsCorrect"] += 1
            if "lihdc" in _key(source_bill.get("Vendor")) and "settlement" in _key(source_bill.get("VendorInvoiceNumber")):
                settlement_rows.append({
                    "apBillId": _text(source_bill.get("APBillID")), "total": float(total),
                    "amountPaid": float(paid), "balance": float(balance), "status": _text(source_bill.get("Status")),
                })
        for target_bill in target_ap.values():
            target_id = _text(target_bill.get("APBillID")).upper()
            if target_id in source_ap:
                continue
            duplicate_source_id = source_ap_keys.get(self._ap_business_key(target_bill))
            if duplicate_source_id:
                actions["apBillsSuperseded"] += 1
                warnings.append(
                    f"Target A/P bill {target_id} is superseded by authoritative source bill {duplicate_source_id}."
                )
            else:
                actions["apBillsPreserveNative"] += 1

        source_time_signatures = Counter(self._docket_signature(row, source=True) for row in source["Dockets"])
        target_time_signatures = Counter(self._docket_signature(row, source=False) for row in target["TimeEntries"])
        actions["timeEntriesAddOrUpdate"] = sum(source_time_signatures.values())
        actions["timeEntriesRetire"] = sum(
            1 for row in target["TimeEntries"]
            if (_date(row.get(sc.COL_TIME_DATE)) or date.min) <= (source_docket_cutoff or date.min)
            and target_time_signatures[self._docket_signature(row, source=False)] > source_time_signatures[self._docket_signature(row, source=False)]
        )

        source_metrics = self._source_metrics(source)
        target_metrics = self._target_metrics(target)
        delta = self._metric_deltas(source_metrics, target_metrics)
        summary = {
            "sourceDocketCutoff": source_docket_cutoff.isoformat() if source_docket_cutoff else "",
            "sourceLedgerCutoff": source_ledger_cutoff.isoformat() if source_ledger_cutoff else "",
            "sourceInvoiceCutoff": source_invoice_cutoff.isoformat() if source_invoice_cutoff else "",
            "sourceCounts": {name: len(rows) for name, rows in source.items()},
            "sourceRawReceivableCount": len(source_raw["Receivables"]),
            "targetCounts": {name: len(rows) for name, rows in target.items()},
            "actions": dict(sorted(actions.items())),
            "receivableDifferences": receivable_differences,
            "targetOnlyReceivables": sorted(target_only_receivables),
            "settlementEvidence": settlement_rows,
            "sourceMetrics": source_metrics,
            "targetMetrics": target_metrics,
            "targetDelta": delta,
        }
        return {
            "ok": not errors,
            "phase": "preview",
            "createdAtUtc": _stamp(),
            "sourcePath": str(source_file),
            "sourceSha256": _sha256(source_file),
            "targetPath": str(target_file),
            "targetSha256": _sha256(target_file),
            "errors": errors,
            "warnings": warnings,
            "summary": summary,
        }

    @staticmethod
    def _ap_business_key(row: Mapping[str, Any]) -> str:
        vendor = _key(row.get("Vendor"))
        invoice_date = _iso_date(row.get("InvoiceDate"))
        total = f"{_money(row.get('Total')):.2f}"
        # Vendors often label the same settlement differently (for example
        # "Settlement - 2026-07-01" versus "Settlement agreement 2026-07-01").
        # Vendor/date/total is the safely comparable business identity here;
        # the full label is still retained in the audit report and source row.
        return "|".join((vendor, invoice_date, total)) if vendor and invoice_date and total else ""

    @staticmethod
    def _ap_equal(left: Mapping[str, Any], right: Mapping[str, Any]) -> bool:
        fields = ("Vendor", "VendorInvoiceNumber", "InvoiceDate", "DueDate", "Status", "Currency")
        if any(_key(left.get(field)) != _key(right.get(field)) for field in fields):
            return False
        return all(abs(_money(left.get(field)) - _money(right.get(field))) <= MONEY_TOLERANCE
                   for field in ("Subtotal", "TaxAmount", "Total", "AmountPaid", "Balance"))

    def _write_table(self, workbook: Any, sheet: str, table_name: str, headers: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> None:
        if sheet not in workbook.sheetnames:
            worksheet = workbook.create_sheet(sheet)
        else:
            worksheet = workbook[sheet]
        old = worksheet.tables.get(table_name)
        if old is not None:
            old_ref = old.ref
            del worksheet.tables[table_name]
            # Clear the old rectangle but leave workbook formatting elsewhere intact.
            from openpyxl.utils.cell import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(old_ref)
            for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.value = None
        start_row, start_col = 1, 1
        for column, header in enumerate(headers, start=start_col):
            worksheet.cell(start_row, column).value = header
        for row_index, row in enumerate(rows, start=start_row + 1):
            for column, header in enumerate(headers, start=start_col):
                worksheet.cell(row_index, column).value = row.get(header, "")
        last_row = max(start_row + 1, start_row + len(rows))
        last_column = start_col + len(headers) - 1
        reference = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(last_column)}{last_row}"
        table = Table(displayName=table_name, ref=reference)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        worksheet.add_table(table)

    @staticmethod
    def _map_legacy_dockets(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
        return [{
            "Date": row.get("Date"),
            "Client": _text(row.get("Sub-Client")) or _text(row.get("Client")),
            "Matter": _text(row.get("Matter_ID")),
            "Parent": _text(row.get("Client")) if _text(row.get("Sub-Client")) else "",
            "Description": _text(row.get("Description")),
            "Time (in hrs) or Units": row.get("Time (in hrs)"),
            "Hourly Rate/Flat Rate": row.get("Hourly Rate/Flat Fee"),
            "Percentage": float(_normalise_share(row.get("Percentage"))),
            "Amount to CS": _money_float(row.get("Amount to CS")),
            "Total Inclusive of HST": _money_float(row.get("Total Inclusive of HST")),
            "Invoice #": _text(row.get("Invoice")),
            "RawSeconds": row.get("RawSeconds"),
            "EntryType": _text(row.get("EntryType")),
        } for row in rows]

    @staticmethod
    def _map_disbursements(rows: Sequence[Mapping[str, Any]], receivables: Mapping[str, Mapping[str, Any]]) -> list[dict[str, Any]]:
        result = []
        for index, row in enumerate(rows, start=1):
            invoice = _text(row.get("Invoice"))
            receivable = receivables.get(invoice.upper(), {})
            state = FinancialSyncService._receivable_state(receivable, source=True) if receivable else {}
            result.append({
                sc.COL_DISB_ID: _stable_id("LEG-DISB", row.get("Date"), row.get("Client"), row.get("Description"), row.get("Amount"), index),
                sc.COL_DISB_DATE: _iso_date(row.get("Date")),
                sc.COL_DISB_CLIENT_NAME: _text(row.get("Client")),
                sc.COL_DISB_SUB_CLIENT: _text(row.get("Sub-Client")),
                sc.COL_DISB_CLIENT_ID: "", sc.COL_DISB_PARENT_ID: "",
                sc.COL_DISB_MATTER_ID: _text(row.get("Matter_ID")),
                sc.COL_DISB_DESCRIPTION: _text(row.get("Description")),
                sc.COL_DISB_AMOUNT: _money_float(row.get("Amount")),
                sc.COL_DISB_TAX_EXEMPT: _text(row.get("Tax Exempt? (Y/N)")),
                sc.COL_DISB_BILL_PCT: _number(row.get("Bill %")),
                sc.COL_DISB_INVOICE_REF: invoice,
                sc.COL_DISB_PAYMENT_STATUS: state.get("status", ""),
                sc.COL_DISB_INVOICE_TOTAL: float(state.get("total", 0)),
                sc.COL_DISB_INVOICE_AMOUNT_PAID: float(state.get("paid", 0)) + float(state.get("credits", 0)),
                sc.COL_DISB_INVOICE_BALANCE_DUE: float(state.get("balance", 0)),
                sc.COL_DISB_CREATED_AT: _stamp(),
            })
        return result

    @staticmethod
    def _map_ledger(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
        result = []
        for index, row in enumerate(rows, start=1):
            explicit_id = _text(row.get("TrxID"))
            ledger_id = explicit_id or _stable_id(
                "LEG-LED", row.get("Date"), row.get("Client/Vendor"), row.get("Description"),
                row.get("Reference"), row.get("Billings (excl. HST)"), row.get("Expenses (excl. HST)"),
                row.get("Collected"), index,
            )
            result.append({
                sc.COL_LEDGER_ID: ledger_id,
                sc.COL_LEDGER_DATE: _iso_date(row.get("Date")),
                sc.COL_LEDGER_CLIENT_VENDOR: _text(row.get("Client/Vendor")),
                sc.COL_LEDGER_DESCRIPTION: _text(row.get("Description")),
                sc.COL_LEDGER_CATEGORY: _text(row.get("Category")),
                sc.COL_LEDGER_REFERENCE: _text(row.get("Reference")),
                sc.COL_LEDGER_BILLINGS_EXCL_HST: _money_float(row.get("Billings (excl. HST)")),
                sc.COL_LEDGER_HST_COLLECTED: _money_float(row.get("HST Collected")),
                sc.COL_LEDGER_EXPENSES_EXCL_HST: _money_float(row.get("Expenses (excl. HST)")),
                sc.COL_LEDGER_HST_PAID: _money_float(row.get("HST Paid")),
                sc.COL_LEDGER_COLLECTED: _money_float(row.get("Collected")),
                sc.COL_LEDGER_WRITE_OFF: _money_float(row.get("Write Off")),
                sc.COL_LEDGER_RECEIVABLE: _money_float(row.get("Receivable")),
                sc.COL_LEDGER_TRX_ID: explicit_id,
                sc.COL_LEDGER_EXTERNAL_REF_ID: _text(row.get("ExternalRefID")),
                sc.COL_LEDGER_ORIGINAL_AMOUNT: _money_float(row.get("OriginalAmount")),
                sc.COL_LEDGER_WORK_CLIENT: _text(row.get("Work Client")),
                sc.COL_LEDGER_CREATED_AT: _stamp(),
            })
        return result

    @staticmethod
    def _map_receivables(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
        return [{
            sc.COL_RECV_INVOICE_NUM: _text(row.get("InvoiceNum")),
            sc.COL_RECV_DATE: _iso_date(row.get("Date")),
            sc.COL_RECV_CLIENT: _text(row.get("Client")),
            sc.COL_RECV_TOTAL_INVOICED: _money_float(row.get("Total_Invoiced")),
            sc.COL_RECV_AMOUNT_PAID: _money_float(row.get("Amount_Paid")),
            sc.COL_RECV_CREDITS_ADJ: _money_float(row.get("Credits/Adj")),
            sc.COL_RECV_BALANCE_DUE: _money_float(row.get("Balance_Due")),
            sc.COL_RECV_STATUS: _text(row.get("Status")),
            sc.COL_RECV_WORK_CLIENT: _text(row.get("Work Client")),
        } for row in rows if _text(row.get("InvoiceNum"))]

    @staticmethod
    def _map_invoice_log(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
        return [{
            sc.COL_INV_INVOICE_NUM: _text(row.get("Invoice #")),
            sc.COL_INV_CLIENT_NAME: _text(row.get("Client Name")),
            sc.COL_INV_SUB_CLIENT: _text(row.get("Sub-Client")),
            sc.COL_INV_INVOICE_DATE: _iso_date(row.get("Invoice Date")),
            sc.COL_INV_TOTAL_FEES: _money_float(row.get("Total Fees")),
            sc.COL_INV_TOTAL_DISBURSEMENTS: _money_float(row.get("Total Disbursements")),
            sc.COL_INV_TOTAL_TAX: _money_float(row.get("Total Tax")),
            sc.COL_INV_AGGREGATE_BILLED: _money_float(row.get("Aggregate Billed to Client")),
            sc.COL_INV_BILL_TO_CLIENT: _text(row.get("Bill To Client")),
            sc.COL_INV_FILE_PATH: "",
        } for row in rows if _text(row.get("Invoice #"))]

    @staticmethod
    def _map_hst(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
        return [{header: row.get(header, "") for header in sc.TABLE_COLUMNS[sc.TBL_HST_LOG]} for row in rows]

    @staticmethod
    def _map_ap_rows(rows: Sequence[Mapping[str, Any]], headers: Sequence[str]) -> list[dict[str, Any]]:
        return [{header: row.get(header, "") for header in headers} for row in rows if any(row.get(header) not in (None, "") for header in headers)]

    @staticmethod
    def _legacy_transaction(row: Mapping[str, Any], index: int) -> dict[str, Any] | None:
        billings = _money(row.get("Billings (excl. HST)"))
        expense = _money(row.get("Expenses (excl. HST)"))
        collected = _money(row.get("Collected"))
        write_off = _money(row.get("Write Off"))
        hst_collected = _money(row.get("HST Collected"))
        hst_paid = _money(row.get("HST Paid"))
        description = _text(row.get("Description"))
        source_id = _text(row.get("TrxID"))
        if billings:
            kind, amount, tax, category = "Income", billings, hst_collected, "INC_LEGAL_FEES"
            from_account, to_account = "ACCOUNTS_RECEIVABLE", "LEGAL_REVENUE"
        elif expense:
            kind, amount, tax, category = "Expense", expense, hst_paid, "EXP_LEGAL_FEES"
            from_account, to_account = "AP_PAYABLE", ""
        elif collected:
            kind, amount, tax, category = "Transfer", collected, Decimal("0"), "AR_RECEIPT"
            is_setoff = "set-off" in description.casefold() or "setoff" in description.casefold()
            from_account, to_account = ("AR_SET_OFF", "AP_PAYABLE") if is_setoff else ("CASH_RECEIPTS", "ACCOUNTS_RECEIVABLE")
        elif write_off:
            kind, amount, tax, category = "Adjustment", -write_off, Decimal("0"), "AR_WRITE_OFF"
            from_account, to_account = "AR_WRITE_OFF", "ACCOUNTS_RECEIVABLE"
        else:
            return None
        transaction_id = source_id or _stable_id(
            "LEG-TXN", row.get("Date"), row.get("Client/Vendor"), description, row.get("Reference"), kind, amount, index,
        )
        return {
            sc.COL_TXN_ID: transaction_id,
            sc.COL_TXN_DATE: _iso_date(row.get("Date")),
            sc.COL_TXN_CLASS: "Business",
            sc.COL_TXN_BUSINESS_UNIT: "Cory Business",
            sc.COL_TXN_TYPE: kind,
            sc.COL_TXN_FROM_ACCOUNT: from_account,
            sc.COL_TXN_TO_ACCOUNT: to_account,
            sc.COL_TXN_PAYEE: _text(row.get("Client/Vendor")),
            sc.COL_TXN_PARENT: "",
            sc.COL_TXN_CLIENT: _text(row.get("Client/Vendor")),
            sc.COL_TXN_MATTER: "",
            sc.COL_TXN_CATEGORY_CODE: category,
            sc.COL_TXN_CATEGORY_NAME: _text(row.get("Category")) or category,
            sc.COL_TXN_MEMBER: "Cory",
            sc.COL_TXN_AMOUNT: float(amount),
            sc.COL_TXN_TAX_AMOUNT: float(tax),
            sc.COL_TXN_TAX_FLAG: "Business Deductible" if kind == "Expense" else "",
            sc.COL_TXN_HST_EXEMPT: 0 if tax else 1,
            sc.COL_TXN_GENERAL_OFFICE_EXPENSE: 1 if kind == "Expense" else 0,
            sc.COL_TXN_SHADOW: 0,
            sc.COL_TXN_INVOICE_REF: _text(row.get("Reference")),
            sc.COL_TXN_BILL_CLAIM_PCT: 0,
            sc.COL_TXN_TOTAL_CLAIM_AMOUNT: 0,
            sc.COL_TXN_EXPENSE_DETAILS: description,
            sc.COL_TXN_NOTES: "Historical sync: " + description,
            sc.COL_TXN_STATUS: "Cleared",
            sc.COL_TXN_CURRENCY: "CAD",
            sc.COL_TXN_VOID_REASON: "",
            sc.COL_TXN_CLEARED_AT: _iso_date(row.get("Date")),
            sc.COL_TXN_RECONCILED_AT: "",
            sc.COL_TXN_CREATED_AT: _stamp(),
            sc.COL_TXN_UPDATED_AT: _stamp(),
        }

    @staticmethod
    def _ensure_transaction_lookups(
        account_rows: Sequence[Mapping[str, Any]],
        category_rows: Sequence[Mapping[str, Any]],
        transactions: Sequence[Mapping[str, Any]],
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        """Add only the ledger-clearing lookups required by synced rows."""
        accounts = [dict(row) for row in account_rows]
        categories = [dict(row) for row in category_rows]
        known_accounts = {_key(row.get(sc.COL_TXN_ACCOUNT_CODE)) for row in accounts}
        account_names = {
            "accounts_receivable": "Accounts Receivable",
            "legal_revenue": "Legal Revenue",
            "ap_payable": "Accounts Payable",
            "cash_receipts": "Cash Receipts Clearing",
            "ar_set_off": "A/R Set-off Clearing",
            "ar_write_off": "A/R Write-off Clearing",
        }
        needed_accounts = {
            _key(row.get(sc.COL_TXN_FROM_ACCOUNT)) for row in transactions
        } | {
            _key(row.get(sc.COL_TXN_TO_ACCOUNT)) for row in transactions
        }
        for account in sorted(value for value in needed_accounts if value and value not in known_accounts):
            accounts.append({
                sc.COL_TXN_ACCOUNT_CODE: account.upper(),
                sc.COL_TXN_ACCOUNT_NAME: account_names.get(account, account.replace("_", " ").title()),
                sc.COL_TXN_ACCOUNT_KIND: "clearing",
                sc.COL_TXN_ACCOUNT_OWNER: "Business",
                sc.COL_TXN_ACCOUNT_ACTIVE: 1,
                sc.COL_TXN_ACCOUNT_ALIASES: "Historic financial synchronization",
            })
            known_accounts.add(account)
        known_categories = {_key(row.get(sc.COL_TXN_CATEGORY_LKP_CODE)) for row in categories}
        category_names = {
            "ar_receipt": "Accounts Receivable Receipt",
            "ar_write_off": "Accounts Receivable Write-off",
            "exp_legal_fees": "Legal Fees Expense",
            "inc_legal_fees": "Legal Fees Revenue",
        }
        category_types = {
            _key(row.get(sc.COL_TXN_CATEGORY_CODE)): _text(row.get(sc.COL_TXN_TYPE))
            for row in transactions if _text(row.get(sc.COL_TXN_CATEGORY_CODE))
        }
        for category in sorted(value for value in category_types if value and value not in known_categories):
            kind = category_types[category] or "Transfer"
            categories.append({
                sc.COL_TXN_CATEGORY_LKP_CODE: category.upper(),
                sc.COL_TXN_CATEGORY_LKP_NAME: category_names.get(category, category.replace("_", " ").title()),
                sc.COL_TXN_CATEGORY_LKP_TYPE: kind,
                sc.COL_TXN_CATEGORY_LKP_CLASS_SCOPE: "Business",
                sc.COL_TXN_CATEGORY_LKP_TAX_FLAG_DEFAULT: "Business Deductible" if kind == "Expense" else "None",
                sc.COL_TXN_CATEGORY_LKP_BILLABLE_ALLOWED: 0,
                sc.COL_TXN_CATEGORY_LKP_MEDICAL_ELIGIBLE: 0,
                sc.COL_TXN_CATEGORY_LKP_DEDUCTIBLE_ELIGIBLE: 0,
                sc.COL_TXN_CATEGORY_LKP_BUSINESS_DEDUCTIBLE_ELIGIBLE: 1 if kind == "Expense" else 0,
                sc.COL_TXN_CATEGORY_LKP_ACTIVE: 1,
                sc.COL_TXN_CATEGORY_LKP_SORT_ORDER: 900,
                sc.COL_TXN_CATEGORY_LKP_NOTES: "Added for historic financial synchronization.",
            })
            known_categories.add(category)
        return accounts, categories

    def _sync_directories(self, target: dict[str, list[dict[str, Any]]], source: Mapping[str, Sequence[Mapping[str, Any]]]) -> tuple[dict[str, list[dict[str, Any]]], dict[str, str], dict[str, str], dict[str, str], dict[str, str]]:
        """Upsert source identities without deleting CSPM-only directory metadata."""
        clients = deepcopy(target["Clients"])
        profiles = deepcopy(target["ClientProfiles"])
        matters = deepcopy(target["Matters"])
        clients_by_name = {_key(row.get(sc.COL_CLIENT_NAME)): row for row in clients if _text(row.get(sc.COL_CLIENT_NAME))}
        client_ids = {_key(row.get(sc.COL_CLIENT_NAME)): _text(row.get(sc.COL_CLIENT_ID)) for row in clients if _text(row.get(sc.COL_CLIENT_NAME))}

        source_client_rows = list(source["Clients"])
        source_client_by_legacy_id = {_key(row.get("Client_ID")): row for row in source_client_rows if _text(row.get("Client_ID"))}
        names: set[str] = set()
        for row in source_client_rows:
            if _text(row.get("Client Name")):
                names.add(_text(row.get("Client Name")))
        for row in source["Dockets"]:
            names.add(_text(row.get("Sub-Client")) or _text(row.get("Client")))
        for row in source["Receivables"]:
            names.add(_text(row.get("Client")))
        for name in sorted(name for name in names if name):
            normalized = _key(name)
            if normalized not in client_ids:
                client_id = _stable_id("LEG-CLIENT", name)
                client_ids[normalized] = client_id
                clients.append({
                    sc.COL_CLIENT_ID: client_id, sc.COL_CLIENT_NAME: name,
                    sc.COL_CLIENT_EMAIL: "", sc.COL_CLIENT_PHONE: "", sc.COL_CLIENT_STATUS: "Active",
                    sc.COL_CLIENT_ACTIVE: 1, sc.COL_CLIENT_NOTES: "Imported from historic Dockets.xlsm synchronization.",
                })

        profile_ids = {_key(row.get(sc.COL_PROFILE_CLIENT_ID)) for row in profiles}
        for source_row in source_client_rows:
            name = _text(source_row.get("Client Name"))
            client_id = client_ids.get(_key(name))
            if not client_id or _key(client_id) in profile_ids:
                continue
            profiles.append({
                sc.COL_PROFILE_CLIENT_ID: client_id,
                sc.COL_PROFILE_LEGAL_NAME: name,
                sc.COL_PROFILE_DISPLAY_NAME: name,
                sc.COL_PROFILE_ENTITY_TYPE: _text(source_row.get("Entity Type")),
                sc.COL_PROFILE_PRINCIPAL_NAME: _text(source_row.get("Principal")),
                sc.COL_PROFILE_PRINCIPAL_POSITION: _text(source_row.get("Position")),
                sc.COL_PROFILE_PRIMARY_EMAIL: _text(source_row.get("Email")),
                sc.COL_PROFILE_PRIMARY_PHONE: _text(source_row.get("Phone Number")),
                sc.COL_PROFILE_ADDR1: _text(source_row.get("Mailing Address")),
                sc.COL_PROFILE_CITY: _text(source_row.get("City")),
                sc.COL_PROFILE_STATE: _text(source_row.get("Province")),
                sc.COL_PROFILE_POSTAL: _text(source_row.get("Postal Code")),
                sc.COL_PROFILE_NOTES: _text(source_row.get("Comments")),
                sc.COL_PROFILE_CREATED: _stamp(), sc.COL_PROFILE_UPDATED: _stamp(),
            })

        matters_by_number = {_key(row.get(sc.COL_MATTER_NUMBER)): row for row in matters if _text(row.get(sc.COL_MATTER_NUMBER))}
        matter_ids = {_key(row.get(sc.COL_MATTER_NUMBER)): _text(row.get(sc.COL_MATTER_ID)) for row in matters if _text(row.get(sc.COL_MATTER_NUMBER))}
        matter_client_ids = {
            _key(row.get(sc.COL_MATTER_NUMBER)): _text(row.get(sc.COL_MATTER_CLIENT_ID))
            for row in matters if _text(row.get(sc.COL_MATTER_NUMBER))
        }
        matter_client_by_id = {
            _key(row.get(sc.COL_MATTER_ID)): _text(row.get(sc.COL_MATTER_CLIENT_ID))
            for row in matters if _text(row.get(sc.COL_MATTER_ID))
        }
        for source_row in source["Matters"]:
            number = _text(source_row.get("Matter_ID"))
            if not number or _key(number) in matter_ids:
                continue
            source_client = source_client_by_legacy_id.get(_key(source_row.get("Client_ID")), {})
            name = _text(source_row.get("Sub-Client")) or _text(source_client.get("Client Name"))
            client_id = client_ids.get(_key(name), "")
            matter_id = _stable_id("LEG-MATTER", number)
            matter_ids[_key(number)] = matter_id
            matter_client_ids[_key(number)] = client_id
            matter_client_by_id[_key(matter_id)] = client_id
            matters.append({
                sc.COL_MATTER_ID: matter_id, sc.COL_MATTER_NUMBER: number,
                sc.COL_MATTER_NAME: _text(source_row.get("Description (Internal)")) or number,
                "DisplayName": _text(source_row.get("Description (Internal)")) or number,
                sc.COL_MATTER_CLIENT_ID: client_id, sc.COL_MATTER_CLIENT_NAME: name,
                sc.COL_MATTER_PARENT_ID: "", sc.COL_MATTER_PARENT_NAME: "",
                sc.COL_MATTER_TYPE: _text(source_row.get("Type")),
                sc.COL_MATTER_STATUS: _text(source_row.get("Status")) or "Active",
                sc.COL_MATTER_DEF_RATE: _money_float(source_row.get("Matter_Rate")),
                sc.COL_MATTER_DEF_SHARE: float(_normalise_share(source_row.get("Matter_Percent"))),
                sc.COL_MATTER_OPEN_DATE: _iso_date(source_row.get("Open_Date")),
                sc.COL_MATTER_DESCRIPTION: _text(source_row.get("Description (Internal)")),
                sc.COL_MATTER_CREATED: _stamp(), sc.COL_MATTER_UPDATED: _stamp(),
            })
        return {"Clients": clients, "ClientProfiles": profiles, "Matters": matters}, client_ids, matter_ids, matter_client_ids, matter_client_by_id

    def _sync_time_entries(
        self,
        target_rows: Sequence[Mapping[str, Any]],
        source_rows: Sequence[Mapping[str, Any]],
        source_receivables: Mapping[str, Mapping[str, Any]],
        source_cutoff: date | None,
        client_ids: Mapping[str, str],
        matter_ids: Mapping[str, str],
        matter_client_ids: Mapping[str, str],
        matter_client_by_id: Mapping[str, str],
    ) -> list[dict[str, Any]]:
        by_signature: dict[tuple[str, str, str, str], deque[dict[str, Any]]] = defaultdict(deque)
        for row in target_rows:
            by_signature[self._docket_signature(row, source=False)].append(dict(row))
        result: list[dict[str, Any]] = []
        matched_target_ids: set[str] = set()
        for index, source_row in enumerate(source_rows, start=1):
            signature = self._docket_signature(source_row, source=True)
            existing = by_signature[signature].popleft() if by_signature.get(signature) else {}
            entry_id = _text(existing.get(sc.COL_TIME_ENTRY_ID)) or _stable_id("LEG-TIME", signature, index)
            matched_target_ids.add(entry_id.casefold())
            invoice_ref = _text(source_row.get("Invoice"))
            status, invoice_status = _invoice_marker_status(invoice_ref)
            receivable = source_receivables.get(invoice_ref.upper(), {})
            state = self._receivable_state(receivable, source=True) if receivable else {}
            client_name = _text(source_row.get("Sub-Client")) or _text(source_row.get("Client"))
            resolved_matter_id = matter_ids.get(_key(source_row.get("Matter_ID")), _text(existing.get(sc.COL_TIME_MATTER_ID)))
            source_total = _money(source_row.get("Total Inclusive of HST"))
            source_net = _money(source_row.get("Amount to CS"))
            hours_value = _number(source_row.get("Time (in hrs)"))
            if source_total == 0 and source_net:
                source_total = source_net
            is_direct_fee = hours_value == 0 and source_net != 0
            formula_net = (
                hours_value * _number(source_row.get("Hourly Rate/Flat Fee"))
                * (_normalise_share(source_row.get("Percentage")) / Decimal("100"))
            ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            is_legacy_override = not is_direct_fee and (
                abs(source_net - formula_net) > MONEY_TOLERANCE
                or abs(source_total - (source_net * Decimal("1.13")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)) > MONEY_TOLERANCE
            )
            result.append({
                sc.COL_TIME_ENTRY_ID: entry_id,
                sc.COL_TIME_DATE: _iso_date(source_row.get("Date")),
                sc.COL_TIME_CLIENT_ID: matter_client_ids.get(
                    _key(source_row.get("Matter_ID")),
                    matter_client_by_id.get(_key(resolved_matter_id), client_ids.get(_key(client_name), _text(existing.get(sc.COL_TIME_CLIENT_ID)))),
                ),
                sc.COL_TIME_MATTER_ID: resolved_matter_id,
                sc.COL_TIME_PARENT_ID: _text(existing.get(sc.COL_TIME_PARENT_ID)),
                sc.COL_TIME_DESC: _text(source_row.get("Description")),
                sc.COL_TIME_HOURS: float(hours_value),
                sc.COL_TIME_RATE: float(_number(source_row.get("Hourly Rate/Flat Fee"))),
                sc.COL_TIME_SHARE_PCT: float(_normalise_share(source_row.get("Percentage"))),
                sc.COL_TIME_GROSS: float(source_net if is_direct_fee else (hours_value * _number(source_row.get("Hourly Rate/Flat Fee"))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
                # Amount to CS is already the lawyer's amount.  Never apply its share twice.
                sc.COL_TIME_NET: float(source_net),
                sc.COL_TIME_HST: float((source_total - source_net).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
                sc.COL_TIME_TOTAL: float(source_total),
                sc.COL_TIME_SECONDS: int(_number(source_row.get("RawSeconds")) or (_number(source_row.get("Time (in hrs)")) * Decimal("3600"))),
                sc.COL_TIME_STATUS: status,
                sc.COL_TIME_INVOICE_REF: invoice_ref,
                sc.COL_TIME_INVOICE_STATUS: invoice_status,
                sc.COL_TIME_PAYMENT_STATUS: state.get("status", ""),
                sc.COL_TIME_INVOICE_TOTAL: float(state.get("total", 0)),
                sc.COL_TIME_INVOICE_AMOUNT_PAID: float(state.get("paid", 0)) + float(state.get("credits", 0)),
                sc.COL_TIME_INVOICE_BALANCE_DUE: float(state.get("balance", 0)),
                sc.COL_TIME_INVOICE_DATE: _iso_date(receivable.get("Date")) if receivable else "",
                sc.COL_TIME_LOCK_AUDIT: (
                    "EntryType:Fee | Financial sync source: Dockets.xlsm" if is_direct_fee
                    else ("EntryType:LegacyOverride | Financial sync source: Dockets.xlsm" if is_legacy_override
                          else "Financial sync source: Dockets.xlsm")
                ),
                sc.COL_TIME_CREATED: _text(existing.get(sc.COL_TIME_CREATED)) or _stamp(),
            })
        for target_row in target_rows:
            entry_id = _text(target_row.get(sc.COL_TIME_ENTRY_ID))
            if entry_id.casefold() in matched_target_ids:
                continue
            item_date = _date(target_row.get(sc.COL_TIME_DATE))
            if source_cutoff and item_date and item_date <= source_cutoff:
                continue  # Retired snapshot history, already captured in preview audit.
            result.append(dict(target_row))
        return result

    @staticmethod
    def _preserve_after(rows: Sequence[Mapping[str, Any]], column: str, cutoff: date | None) -> list[dict[str, Any]]:
        if cutoff is None:
            return [dict(row) for row in rows]
        return [dict(row) for row in rows if (_date(row.get(column)) or date.min) > cutoff]

    def _apply_to_candidate(self, source: Mapping[str, Sequence[Mapping[str, Any]]], candidate_path: Path) -> dict[str, int]:
        target = self._read_target(candidate_path)
        source_docket_cutoff = self._max_date(source["Dockets"], "Date")
        source_ledger_cutoff = self._max_date(source["Ledger"], "Date")
        source_receivables = self._source_receivables(source["Receivables"])
        directories, client_ids, matter_ids, matter_client_ids, matter_client_by_id = self._sync_directories(target, source)

        mapped_receivables = self._map_receivables(source["Receivables"])
        mapped_invoice_log = self._map_invoice_log(source["Invoice Log"])
        mapped_ledger = self._map_ledger(source["Ledger"])
        mapped_transactions = [
            transaction for index, row in enumerate(source["Ledger"], start=1)
            if (transaction := self._legacy_transaction(row, index)) is not None
        ]
        mapped_ap_bills = self._map_ap_rows(source[AP_BILLS_SHEET], AP_BILLS_HEADERS)
        mapped_ap_payments = self._map_ap_rows(source[AP_PAYMENTS_SHEET], AP_PAYMENTS_HEADERS)
        source_ap_keys = {self._ap_business_key(row) for row in mapped_ap_bills if self._ap_business_key(row)}

        # Preserve native current A/P only when it is not a duplicate of a source bill.
        preserved_ap_bills = [
            row for row in target[AP_BILLS_SHEET]
            if self._ap_business_key(row) not in source_ap_keys
            and (_date(row.get("InvoiceDate")) or date.min) > (source_docket_cutoff or date.min)
        ]
        all_ap_bills = mapped_ap_bills + preserved_ap_bills
        active_bill_ids = {_key(row.get("APBillID")) for row in all_ap_bills}
        preserved_ap_payments = [
            row for row in target[AP_PAYMENTS_SHEET]
            if _key(row.get("APBillID")) in active_bill_ids
            and _key(row.get("APPaymentID")) not in {_key(item.get("APPaymentID")) for item in mapped_ap_payments}
        ]

        # Current native transactions and rows postdating the source snapshot survive.
        preserved_transactions = self._preserve_after(target["Transactions"], sc.COL_TXN_DATE, source_ledger_cutoff)
        all_transactions = mapped_transactions + preserved_transactions
        transaction_accounts, transaction_categories = self._ensure_transaction_lookups(
            target["TransactionAccounts"], target["TransactionCategories"], all_transactions,
        )
        preserved_ledger = self._preserve_after(target["Ledger"], sc.COL_LEDGER_DATE, source_ledger_cutoff)
        preserved_receivables = [
            row for row in target["Receivables"]
            if _is_actual_invoice(row.get(sc.COL_RECV_INVOICE_NUM))
            and (_date(row.get(sc.COL_RECV_DATE)) or date.min) > (source_docket_cutoff or date.min)
            and _key(row.get(sc.COL_RECV_INVOICE_NUM)) not in {_key(item.get(sc.COL_RECV_INVOICE_NUM)) for item in mapped_receivables}
        ]
        preserved_invoice_log = [
            row for row in target["Invoice Log"]
            if _is_actual_invoice(row.get(sc.COL_INV_INVOICE_NUM))
            and (_date(row.get(sc.COL_INV_INVOICE_DATE)) or date.min) > (source_docket_cutoff or date.min)
            and _key(row.get(sc.COL_INV_INVOICE_NUM)) not in {_key(item.get(sc.COL_INV_INVOICE_NUM)) for item in mapped_invoice_log}
        ]
        preserved_disbursements = self._preserve_after(target["Disbursements"], sc.COL_DISB_DATE, source_docket_cutoff)

        workbook = load_workbook(candidate_path, keep_vba=True, data_only=False)
        try:
            self._write_table(workbook, sc.SHEET_CLIENTS, sc.TBL_CLIENTS, sc.TABLE_COLUMNS[sc.TBL_CLIENTS], directories["Clients"])
            self._write_table(workbook, sc.SHEET_CLIENT_PROFILES, sc.TBL_CLIENT_PROFILES, sc.TABLE_COLUMNS[sc.TBL_CLIENT_PROFILES], directories["ClientProfiles"])
            self._write_table(workbook, sc.SHEET_MATTERS, sc.TBL_MATTERS, sc.TABLE_COLUMNS[sc.TBL_MATTERS], directories["Matters"])
            self._write_table(workbook, RAW_DOCKETS_SHEET, RAW_DOCKETS_TABLE, RAW_DOCKETS_HEADERS, self._map_legacy_dockets(source["Dockets"]))
            self._write_table(workbook, sc.SHEET_TIME, sc.TBL_TIME, sc.TABLE_COLUMNS[sc.TBL_TIME], self._sync_time_entries(target["TimeEntries"], source["Dockets"], source_receivables, source_docket_cutoff, client_ids, matter_ids, matter_client_ids, matter_client_by_id))
            self._write_table(workbook, sc.SHEET_DISBURSEMENTS, sc.TBL_DISBURSEMENTS, sc.TABLE_COLUMNS[sc.TBL_DISBURSEMENTS], self._map_disbursements(source["Disbursements"], source_receivables) + preserved_disbursements)
            self._write_table(workbook, sc.SHEET_LEDGER, sc.TBL_LEDGER, sc.TABLE_COLUMNS[sc.TBL_LEDGER], mapped_ledger + preserved_ledger)
            self._write_table(workbook, sc.SHEET_RECEIVABLES, sc.TBL_RECEIVABLES, sc.TABLE_COLUMNS[sc.TBL_RECEIVABLES], mapped_receivables + preserved_receivables)
            self._write_table(workbook, sc.SHEET_INVOICE_LOG, sc.TBL_INVOICE_LOG, sc.TABLE_COLUMNS[sc.TBL_INVOICE_LOG], mapped_invoice_log + preserved_invoice_log)
            self._write_table(workbook, sc.SHEET_HST_LOG, sc.TBL_HST_LOG, sc.TABLE_COLUMNS[sc.TBL_HST_LOG], self._map_hst(source["HST_Log"]))
            self._write_table(workbook, sc.SHEET_TRANSACTIONS, sc.TBL_TRANSACTIONS_MASTER, sc.TABLE_COLUMNS[sc.TBL_TRANSACTIONS_MASTER], all_transactions)
            self._write_table(workbook, sc.SHEET_TRANSACTION_ACCOUNTS, sc.TBL_TRANSACTION_ACCOUNTS, sc.TABLE_COLUMNS[sc.TBL_TRANSACTION_ACCOUNTS], transaction_accounts)
            self._write_table(workbook, sc.SHEET_TRANSACTION_CATEGORIES, sc.TBL_TRANSACTION_CATEGORIES, sc.TABLE_COLUMNS[sc.TBL_TRANSACTION_CATEGORIES], transaction_categories)
            self._write_table(workbook, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS, all_ap_bills)
            self._write_table(workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS, mapped_ap_payments + preserved_ap_payments)
            workbook.save(candidate_path)
        finally:
            workbook.close()
        return {
            "clients": len(directories["Clients"]), "profiles": len(directories["ClientProfiles"]), "matters": len(directories["Matters"]),
            "timeEntries": len(self._sync_time_entries(target["TimeEntries"], source["Dockets"], source_receivables, source_docket_cutoff, client_ids, matter_ids, matter_client_ids, matter_client_by_id)),
            "ledger": len(mapped_ledger) + len(preserved_ledger), "receivables": len(mapped_receivables) + len(preserved_receivables),
            "invoiceLog": len(mapped_invoice_log) + len(preserved_invoice_log), "transactions": len(all_transactions),
            "apBills": len(all_ap_bills), "apPayments": len(mapped_ap_payments) + len(preserved_ap_payments),
        }

    def _candidate_reconciliation(self, source: Mapping[str, Sequence[Mapping[str, Any]]], candidate_path: Path) -> dict[str, Any]:
        candidate = self._read_target(candidate_path)
        source_metrics = self._source_metrics(source)
        candidate_metrics = self._target_metrics(candidate)
        deltas = self._metric_deltas(source_metrics, candidate_metrics)
        cutoff = self._max_date(source["Dockets"], "Date")
        # Native post-cutoff activity is deliberately retained, so compare only source-owned tables.
        source_ar = sum((_money(row.get("Balance_Due")) for row in source["Receivables"]), Decimal("0"))
        candidate_legacy_ar = sum(
            (_money(row.get(sc.COL_RECV_BALANCE_DUE)) for row in candidate["Receivables"]
             if (_date(row.get(sc.COL_RECV_DATE)) or date.min) <= (cutoff or date.max)), Decimal("0")
        )
        source_ledger = self._source_metrics(source)
        candidate_legacy_ledger_rows = [row for row in candidate["Ledger"] if (_date(row.get(sc.COL_LEDGER_DATE)) or date.min) <= (self._max_date(source["Ledger"], "Date") or date.max)]
        candidate_legacy_metrics = self._target_metrics({"TimeEntries": [], "Ledger": candidate_legacy_ledger_rows, "Receivables": []})
        candidate_legacy_time_rows = [
            row for row in candidate["TimeEntries"]
            if (_date(row.get(sc.COL_TIME_DATE)) or date.min) <= (cutoff or date.max)
        ]
        candidate_legacy_productivity = self._target_metrics({"TimeEntries": candidate_legacy_time_rows, "Ledger": [], "Receivables": []})
        controlled_deltas = {
            "receivablesAr": float(candidate_legacy_ar - source_ar),
            "ledgerBillings": round(candidate_legacy_metrics["ledgerBillings"] - source_ledger["ledgerBillings"], 2),
            "ledgerExpenses": round(candidate_legacy_metrics["ledgerExpenses"] - source_ledger["ledgerExpenses"], 2),
            "ledgerHstCollected": round(candidate_legacy_metrics["ledgerHstCollected"] - source_ledger["ledgerHstCollected"], 2),
            "ledgerHstPaid": round(candidate_legacy_metrics["ledgerHstPaid"] - source_ledger["ledgerHstPaid"], 2),
            "ledgerCollected": round(candidate_legacy_metrics["ledgerCollected"] - source_ledger["ledgerCollected"], 2),
            "productivityHours": round(candidate_legacy_productivity["productivityHours"] - source_metrics["productivityHours"], 2),
            "productivityGross": round(candidate_legacy_productivity["productivityGross"] - source_metrics["productivityGross"], 2),
            "productivityNet": round(candidate_legacy_productivity["productivityNet"] - source_metrics["productivityNet"], 2),
        }
        failed = {
            key: value for key, value in controlled_deltas.items()
            if abs(Decimal(str(value))) > (HOURS_TOLERANCE if key == "productivityHours" else MONEY_TOLERANCE)
        }
        return {
            "ok": not failed,
            "sourceMetrics": source_metrics,
            "candidateMetrics": candidate_metrics,
            "candidateDeltaIncludingPreservedNative": deltas,
            "sourceOwnedDeltas": controlled_deltas,
            "failedChecks": failed,
        }

    @staticmethod
    def _write_audit_files(output_dir: Path, report: Mapping[str, Any]) -> tuple[Path, Path]:
        output_dir.mkdir(parents=True, exist_ok=True)
        json_path = output_dir / "financial_sync_audit.json"
        markdown_path = output_dir / "financial_sync_audit.md"
        json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
        summary = report.get("summary") if isinstance(report.get("summary"), Mapping) else {}
        checks = report.get("reconciliation") if isinstance(report.get("reconciliation"), Mapping) else {}
        lines = [
            "# Financial Synchronization Audit",
            "",
            f"Created: {report.get('createdAtUtc', '')}",
            f"Source: `{report.get('sourcePath', '')}`",
            f"Target: `{report.get('targetPath', '')}`",
            f"Candidate: `{report.get('candidatePath', '')}`",
            "",
            "## Plan",
            "",
        ]
        for key, value in sorted(dict(summary.get("actions") or {}).items()):
            lines.append(f"- {key}: {value}")
        lines += ["", "## Reconciliation", ""]
        for key, value in sorted(dict(checks.get("sourceOwnedDeltas") or {}).items()):
            lines.append(f"- {key}: {value:+.2f}")
        errors = list(report.get("errors") or [])
        warnings = list(report.get("warnings") or [])
        if errors:
            lines += ["", "## Errors", ""] + [f"- {item}" for item in errors]
        if warnings:
            lines += ["", "## Warnings", ""] + [f"- {item}" for item in warnings]
        markdown_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return json_path, markdown_path

    @staticmethod
    def _year_time_metrics(rows: Sequence[Mapping[str, Any]], year: int) -> dict[str, float]:
        selected = [row for row in rows if (_date(row.get(sc.COL_TIME_DATE)) or date.min).year == year]
        return {
            "hours": float(sum((_number(row.get(sc.COL_TIME_HOURS)) for row in selected), Decimal("0")).quantize(Decimal("0.01"))),
            "gross": float(sum((_money(row.get(sc.COL_TIME_GROSS)) for row in selected), Decimal("0"))),
        }

    @staticmethod
    def _year_transaction_metrics(rows: Sequence[Mapping[str, Any]], year: int) -> dict[str, float]:
        selected = [row for row in rows if (_date(row.get(sc.COL_TXN_DATE)) or date.min).year == year]
        expenses = sum((_money(row.get(sc.COL_TXN_AMOUNT)) for row in selected if _key(row.get(sc.COL_TXN_TYPE)) == "expense"), Decimal("0"))
        transfers = sum((
            _money(row.get(sc.COL_TXN_AMOUNT)) for row in selected
            if _key(row.get(sc.COL_TXN_TYPE)) == "transfer"
            and _key(row.get(sc.COL_TXN_FROM_ACCOUNT)) != "ar_set_off"
            and _key(row.get(sc.COL_TXN_TO_ACCOUNT)) != "ar_set_off"
        ), Decimal("0"))
        return {"expenses": float(expenses), "banked": float(transfers)}

    def _dashboard_reconciliation(self, candidate_path: Path, source: Mapping[str, Sequence[Mapping[str, Any]]]) -> dict[str, Any]:
        year = (self._max_date(source["Dockets"], "Date") or date.today()).year
        candidate_paths = AppPaths(candidate_path.parent, override_data_dir=candidate_path.parent)
        repo = ExcelRepo(candidate_paths)
        financial = repo.financial_dashboard_report(year)
        productivity = repo.get_productivity_dashboard_data()
        target = self._read_target(candidate_path)
        expected_finance = self._year_transaction_metrics(target["Transactions"], year)
        expected_productivity = self._year_time_metrics(target["TimeEntries"], year)
        financial_summary = dict(financial.get("summary") or {})
        kpis = dict(productivity.get("kpiData") or {})
        actual = {
            "dashboardAr": float(financial_summary.get("dashboardAr") or 0),
            "expenses": float(financial_summary.get("expenses") or 0),
            "bankedAmount": float(financial_summary.get("bankedAmount") or 0),
            "productivityHours": float(kpis.get("hours") or 0),
            "productivityGross": float(kpis.get("gross") or 0),
        }
        expected = {
            "dashboardAr": self._target_metrics(target)["receivablesAr"],
            "expenses": expected_finance["expenses"],
            "bankedAmount": expected_finance["banked"],
            "productivityHours": expected_productivity["hours"],
            "productivityGross": expected_productivity["gross"],
        }
        deltas = self._metric_deltas(expected, actual)
        failed = {
            key: value for key, value in deltas.items()
            if abs(Decimal(str(value))) > (HOURS_TOLERANCE if key == "productivityHours" else MONEY_TOLERANCE)
        }
        return {
            "ok": bool(financial.get("ok") and productivity.get("ok") and not failed),
            "year": year,
            "expected": expected,
            "actual": actual,
            "deltas": deltas,
            "failedChecks": failed,
        }

    def build_candidate(
        self,
        source_path: str | Path,
        output_dir: str | Path,
        target_path: str | Path | None = None,
    ) -> dict[str, Any]:
        """Produce and prove an isolated candidate; this never changes live CSPM."""
        source_file = Path(source_path)
        target_file = Path(target_path) if target_path else self.target_path
        preview = self.preview(source_file, target_file)
        if not preview["ok"]:
            return preview
        source, source_normalization_warnings = self._canonicalize_source(self._read_source(source_file))
        candidate_root = Path(output_dir)
        candidate_root.mkdir(parents=True, exist_ok=True)
        # The directory is named as a candidate; keeping the workbook's normal
        # name allows the native report engine to evaluate it without any path
        # or configuration special case.
        candidate_path = candidate_root / "CSPM.xlsm"
        shutil.copy2(target_file, candidate_path)
        counts = self._apply_to_candidate(source, candidate_path)
        reconciliation = self._candidate_reconciliation(source, candidate_path)
        integrity = WorkbookIntegrityService(self.paths).check(
            workbook_path=candidate_path,
            schema_path=self.paths.workbook_schema_path(),
        ).as_dict()
        report = dict(preview)
        for warning in source_normalization_warnings:
            if warning not in report["warnings"]:
                report["warnings"].append(warning)
        report.update({
            "phase": "candidate",
            "candidatePath": str(candidate_path),
            "candidateSha256": _sha256(candidate_path),
            "candidateCounts": counts,
            "integrity": integrity,
            "reconciliation": reconciliation,
            "dashboardReconciliation": self._dashboard_reconciliation(candidate_path, source),
        })
        report["ok"] = bool(
            preview["ok"] and integrity.get("ok") and reconciliation.get("ok")
            and report["dashboardReconciliation"].get("ok")
        )
        json_path, markdown_path = self._write_audit_files(candidate_root, report)
        report["auditJsonPath"], report["auditMarkdownPath"] = str(json_path), str(markdown_path)
        return report

    def promote_candidate(self, candidate_path: str | Path, audit_json_path: str | Path) -> dict[str, Any]:
        """Atomically promote a proven candidate after a fresh-change guard.

        This method is intentionally separate from ``build_candidate``.  It
        refuses a candidate whose audit did not pass, or whose live target has
        changed since the candidate was built.  The caller must close CSPM
        first; an operating-system file lock will leave the live workbook
        unchanged and return a clear error.
        """
        candidate = Path(candidate_path)
        audit_path = Path(audit_json_path)
        target = self.target_path
        if not candidate.is_file() or not audit_path.is_file():
            raise FinancialSyncError("Candidate workbook and its audit report are both required for promotion.")
        try:
            audit = json.loads(audit_path.read_text(encoding="utf-8"))
        except Exception as exc:
            raise FinancialSyncError(f"Candidate audit report could not be read: {exc}") from exc
        if not audit.get("ok"):
            raise FinancialSyncError("Candidate did not pass all reconciliation gates and cannot be promoted.")
        expected_target_sha = _text(audit.get("targetSha256"))
        current_target_sha = _sha256(target)
        if not expected_target_sha or current_target_sha != expected_target_sha:
            raise FinancialSyncError(
                "The live CSPM workbook changed after this candidate was built. Build a fresh candidate before promotion."
            )
        backup_root = self.paths.backups_dir() / f"financial_sync_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        backup_root.mkdir(parents=True, exist_ok=False)
        backup_workbook = backup_root / "CSPM.before-financial-sync.xlsm"
        backup_audit = backup_root / "financial_sync_audit.json"
        staged_target = target.with_name(f"{target.stem}.financial-sync-staged{target.suffix}")
        try:
            shutil.copy2(target, backup_workbook)
            shutil.copy2(audit_path, backup_audit)
            shutil.copy2(candidate, staged_target)
            # os.replace is atomic on the same volume.  If Excel/CSPM has the
            # workbook open, Windows rejects it and the original stays intact.
            os.replace(staged_target, target)
        except PermissionError as exc:
            staged_target.unlink(missing_ok=True)
            raise FinancialSyncError("Close CSPM before promoting the synchronization candidate; the live workbook is in use.") from exc
        except Exception:
            staged_target.unlink(missing_ok=True)
            raise
        return {
            "ok": True,
            "targetPath": str(target),
            "targetSha256": _sha256(target),
            "backupPath": str(backup_workbook),
            "backupSha256": _sha256(backup_workbook),
            "auditPath": str(backup_audit),
            "message": "Financial synchronization candidate promoted with a recoverable pre-sync backup.",
        }
