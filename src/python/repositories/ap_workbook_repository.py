from __future__ import annotations

from datetime import datetime
from decimal import Decimal
import logging
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from domain.ap_lifecycle import APValidationError, build_bill_snapshot, clean_text, money
from domain.ap_schema import (
    AP_BILLS_HEADERS,
    AP_BILLS_SHEET,
    AP_BILLS_TABLE,
    AP_PAYMENTS_HEADERS,
    AP_PAYMENTS_SHEET,
    AP_PAYMENTS_TABLE,
)

logger = logging.getLogger("cspm.ap.repository")


class APWorkbookRepositoryError(RuntimeError):
    pass


def _close_workbook(workbook) -> None:
    for attribute in ("vba_archive", "_archive"):
        archive = getattr(workbook, attribute, None)
        if archive is not None:
            try:
                archive.close()
            except Exception:
                pass
            try:
                setattr(workbook, attribute, None)
            except Exception:
                pass
    try:
        workbook.close()
    except Exception:
        pass


def _open(path: Path):
    return load_workbook(path, keep_vba=path.suffix.lower() == ".xlsm", data_only=False, keep_links=True)


def _table_context(workbook, sheet_name: str, table_name: str, expected_headers: tuple[str, ...]):
    if sheet_name not in workbook.sheetnames:
        raise APWorkbookRepositoryError(f"Missing AP worksheet: {sheet_name}")
    worksheet = workbook[sheet_name]
    if table_name not in worksheet.tables:
        raise APWorkbookRepositoryError(f"Missing AP table: {table_name}")
    table = worksheet.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = tuple(str(worksheet.cell(min_row, column).value or "").strip() for column in range(min_col, max_col + 1))
    if headers != expected_headers:
        # Auto-migrate: if the existing headers are a leading subset of the
        # expected headers, append the missing columns to the worksheet and
        # expand the table reference.
        if expected_headers[:len(headers)] == headers and len(headers) < len(expected_headers):
            from openpyxl.utils import get_column_letter
            missing = expected_headers[len(headers):]
            for offset, col_name in enumerate(missing):
                col_idx = max_col + 1 + offset
                worksheet.cell(min_row, col_idx).value = col_name
            new_max_col = max_col + len(missing)
            table.ref = (f"{get_column_letter(min_col)}{min_row}:"
                         f"{get_column_letter(new_max_col)}{max_row}")
            max_col = new_max_col
            logger.info("Auto-migrated %s: appended columns %s", table_name, missing)
        else:
            raise APWorkbookRepositoryError(f"Unexpected headers in {table_name}: {headers}")
    return worksheet, table, min_col, min_row, max_col, max_row


def _read_rows(workbook, sheet_name: str, table_name: str, headers: tuple[str, ...]) -> list[dict[str, Any]]:
    worksheet, _, min_col, min_row, max_col, max_row = _table_context(workbook, sheet_name, table_name, headers)
    rows = []
    for row_number in range(min_row + 1, max_row + 1):
        values = [worksheet.cell(row_number, column).value for column in range(min_col, max_col + 1)]
        if any(value not in (None, "") for value in values):
            rows.append(dict(zip(headers, values)))
    return rows


def _replace_rows(workbook, sheet_name: str, table_name: str, headers: tuple[str, ...], rows: list[Mapping[str, Any]]) -> None:
    worksheet, table, min_col, min_row, max_col, max_row = _table_context(workbook, sheet_name, table_name, headers)
    for row_number in range(min_row + 1, max(max_row, min_row + 1) + 1):
        for column in range(min_col, max_col + 1):
            worksheet.cell(row_number, column).value = None
    for offset, row in enumerate(rows, start=1):
        for column_offset, header in enumerate(headers):
            worksheet.cell(min_row + offset, min_col + column_offset).value = row.get(header, "")
    last_row = min_row + max(1, len(rows))
    from openpyxl.utils import get_column_letter
    table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(min_col + len(headers) - 1)}{last_row}"


def _now() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def _active_payments(payments: list[Mapping[str, Any]]) -> list[dict[str, Any]]:
    """Return only unreversed outgoing payments, never reversal evidence rows."""
    active: list[dict[str, Any]] = []
    for payment in payments:
        status = clean_text(payment.get("Status")).casefold()
        if status not in {"reversed", "reversal"}:
            active.append(dict(payment))
    return active


def _payment_lifecycle(payments: list[Mapping[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "payment_id": row.get("APPaymentID"),
            "amount": row.get("Amount"),
            "reversed": clean_text(row.get("Status")).casefold() == "reversed",
        }
        for row in payments
        if clean_text(row.get("Status")).casefold() != "reversal"
    ]


class APWorkbookRepository:
    def __init__(self, workbook_path: str | Path):
        self.path = Path(workbook_path)
        if not self.path.is_file():
            raise APWorkbookRepositoryError(f"Workbook not found: {self.path}")

    def list_bills(self) -> list[dict[str, Any]]:
        workbook = _open(self.path)
        try:
            return _read_rows(workbook, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS)
        finally:
            _close_workbook(workbook)

    def list_payments(self, bill_id: str | None = None) -> list[dict[str, Any]]:
        workbook = _open(self.path)
        try:
            rows = _read_rows(workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS)
            if bill_id is None:
                return rows
            key = clean_text(bill_id).casefold()
            return [row for row in rows if clean_text(row.get("APBillID")).casefold() == key]
        finally:
            _close_workbook(workbook)

    def get_bill(self, bill_id: str) -> dict[str, Any] | None:
        key = clean_text(bill_id).casefold()
        for row in self.list_bills():
            if clean_text(row.get("APBillID")).casefold() == key:
                return row
        return None

    def list_active_payments(self, bill_id: str) -> list[dict[str, Any]]:
        return _active_payments(self.list_payments(bill_id))

    def create_bill(self, bill: Mapping[str, Any]) -> dict[str, Any]:
        bill_id = clean_text(bill.get("APBillID"))
        if not bill_id:
            raise APValidationError("APBillID is required.")
        workbook = _open(self.path)
        try:
            bills = _read_rows(workbook, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS)
            payments = _read_rows(workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS)
            if any(clean_text(row.get("APBillID")).casefold() == bill_id.casefold() for row in bills):
                raise APValidationError(f"Duplicate AP bill ID: {bill_id}")
            snapshot = build_bill_snapshot(
                bill_id=bill_id,
                vendor=bill.get("Vendor"),
                vendor_invoice_number=bill.get("VendorInvoiceNumber"),
                invoice_date=bill.get("InvoiceDate"),
                subtotal=bill.get("Subtotal"),
                tax_amount=bill.get("TaxAmount"),
                payments=(),
                posted=bool(bill.get("Posted", True)),
            )
            if any(clean_text(row.get("DuplicateKey")) == snapshot.duplicate_key for row in bills):
                raise APValidationError("A matching vendor invoice already exists.")
            created = clean_text(bill.get("CreatedAt")) or _now()
            row = {header: "" for header in AP_BILLS_HEADERS}
            row.update({
                "APBillID": snapshot.bill_id,
                "Vendor": snapshot.vendor,
                "VendorInvoiceNumber": snapshot.vendor_invoice_number,
                "InvoiceDate": snapshot.invoice_date,
                "DueDate": clean_text(bill.get("DueDate")),
                "Subtotal": float(snapshot.subtotal),
                "TaxAmount": float(snapshot.tax_amount),
                "Total": float(snapshot.total),
                "AmountPaid": 0.0,
                "Balance": float(snapshot.balance),
                "Status": snapshot.status.value,
                "Currency": clean_text(bill.get("Currency")) or "CAD",
                "ExpenseTransactionID": clean_text(bill.get("ExpenseTransactionID")),
                "DuplicateKey": snapshot.duplicate_key,
                "ExpenseTreatment": clean_text(bill.get("ExpenseTreatment")),
                "CategoryCode": clean_text(bill.get("CategoryCode")),
                "CategoryName": clean_text(bill.get("CategoryName")),
                "SourceAccount": clean_text(bill.get("SourceAccount") or bill.get("FromAccount")),
                "Notes": clean_text(bill.get("Notes")),
                "CreatedAt": created,
                "UpdatedAt": created,
            })
            # Keep all governed V2 evidence supplied by the orchestration
            # service.  Unknown input keys are intentionally ignored, but every
            # declared workbook column is durable and survives a later edit.
            for header in AP_BILLS_HEADERS:
                if header in bill and header not in {"APBillID", "AmountPaid", "Balance", "Status", "DuplicateKey", "CreatedAt", "UpdatedAt"}:
                    row[header] = bill.get(header, row.get(header, ""))
            bills.append(row)
            _replace_rows(workbook, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS, bills)
            _replace_rows(workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS, payments)
            workbook.save(self.path)
        finally:
            _close_workbook(workbook)
        persisted = self.get_bill(bill_id)
        if persisted is None:
            raise APWorkbookRepositoryError("AP bill failed post-save verification.")
        return persisted

    def update_bill(self, bill: Mapping[str, Any]) -> dict[str, Any]:
        """Update one AP bill in place while retaining its identity and payment history."""
        bill_id = clean_text(bill.get("APBillID"))
        if not bill_id:
            raise APValidationError("APBillID is required.")
        workbook = _open(self.path)
        try:
            bills = _read_rows(workbook, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS)
            payments = _read_rows(workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS)
            existing = next(
                (row for row in bills if clean_text(row.get("APBillID")).casefold() == bill_id.casefold()),
                None,
            )
            if existing is None:
                raise APValidationError(f"Unknown AP bill ID: {bill_id}")

            related_payments = [
                row for row in payments
                if clean_text(row.get("APBillID")).casefold() == bill_id.casefold()
            ]
            active_payments = _active_payments(related_payments)
            snapshot = build_bill_snapshot(
                bill_id=bill_id,
                vendor=bill.get("Vendor"),
                vendor_invoice_number=bill.get("VendorInvoiceNumber"),
                invoice_date=bill.get("InvoiceDate"),
                subtotal=bill.get("Subtotal"),
                tax_amount=bill.get("TaxAmount"),
                payments=_payment_lifecycle(related_payments),
                posted=bool(bill.get("Posted", True)),
            )
            if active_payments:
                existing_subtotal = money(existing.get("Subtotal"), "existing subtotal")
                existing_tax = money(existing.get("TaxAmount"), "existing tax amount")
                if snapshot.subtotal != existing_subtotal or snapshot.tax_amount != existing_tax:
                    raise APValidationError(
                        "Reverse active payments before changing a bill subtotal or tax amount."
                    )
            if any(
                clean_text(row.get("DuplicateKey")) == snapshot.duplicate_key
                and clean_text(row.get("APBillID")).casefold() != bill_id.casefold()
                for row in bills
            ):
                raise APValidationError("A matching vendor invoice already exists.")

            created = clean_text(existing.get("CreatedAt")) or _now()
            existing.update({
                "Vendor": snapshot.vendor,
                "VendorInvoiceNumber": snapshot.vendor_invoice_number,
                "InvoiceDate": snapshot.invoice_date,
                "DueDate": clean_text(bill.get("DueDate")),
                "Subtotal": float(snapshot.subtotal),
                "TaxAmount": float(snapshot.tax_amount),
                "Total": float(snapshot.total),
                "AmountPaid": float(snapshot.total_paid),
                "Balance": float(snapshot.balance),
                "Status": snapshot.status.value,
                "Currency": clean_text(bill.get("Currency")) or clean_text(existing.get("Currency")) or "CAD",
                "ExpenseTransactionID": clean_text(bill.get("ExpenseTransactionID"))
                    or clean_text(existing.get("ExpenseTransactionID")),
                "DuplicateKey": snapshot.duplicate_key,
                "ExpenseTreatment": clean_text(bill.get("ExpenseTreatment")) or clean_text(existing.get("ExpenseTreatment")),
                "CategoryCode": clean_text(bill.get("CategoryCode")) or clean_text(existing.get("CategoryCode")),
                "CategoryName": clean_text(bill.get("CategoryName")) or clean_text(existing.get("CategoryName")),
                "SourceAccount": clean_text(bill.get("SourceAccount") or bill.get("FromAccount")) or clean_text(existing.get("SourceAccount")),
                "Notes": clean_text(bill.get("Notes")),
                "CreatedAt": created,
                "UpdatedAt": _now(),
            })
            for header in AP_BILLS_HEADERS:
                if header in bill and header not in {"APBillID", "AmountPaid", "Balance", "Status", "DuplicateKey", "CreatedAt", "UpdatedAt"}:
                    existing[header] = bill.get(header, existing.get(header, ""))
            _replace_rows(workbook, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS, bills)
            _replace_rows(workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS, payments)
            workbook.save(self.path)
        finally:
            _close_workbook(workbook)
        persisted = self.get_bill(bill_id)
        if persisted is None:
            raise APWorkbookRepositoryError("AP bill failed post-save verification after update.")
        return persisted

    def delete_bill(self, bill_id: str) -> dict[str, Any]:
        """Permanently remove an eligible bill and all payment rows that reference it."""
        bill_id = clean_text(bill_id)
        if not bill_id:
            raise APValidationError("APBillID is required.")
        workbook = _open(self.path)
        try:
            bills = _read_rows(workbook, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS)
            payments = _read_rows(workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS)
            existing = next(
                (row for row in bills if clean_text(row.get("APBillID")).casefold() == bill_id.casefold()),
                None,
            )
            if existing is None:
                raise APValidationError(f"Unknown AP bill ID: {bill_id}")
            related_payments = [
                row for row in payments
                if clean_text(row.get("APBillID")).casefold() == bill_id.casefold()
            ]
            if _active_payments(related_payments):
                raise APValidationError(
                    "Reverse the active AP payments before permanently deleting this bill."
                )
            status = clean_text(existing.get("Status")).casefold()
            if money(existing.get("AmountPaid"), "amount paid") != Decimal("0.00") or status not in {"unpaid", "draft"}:
                raise APValidationError(
                    "Only unpaid bills with no active payments can be permanently deleted."
                )
            retained_bills = [
                row for row in bills
                if clean_text(row.get("APBillID")).casefold() != bill_id.casefold()
            ]
            retained_payments = [
                row for row in payments
                if clean_text(row.get("APBillID")).casefold() != bill_id.casefold()
            ]
            _replace_rows(workbook, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS, retained_bills)
            _replace_rows(workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS, retained_payments)
            workbook.save(self.path)
        finally:
            _close_workbook(workbook)
        if self.get_bill(bill_id) is not None:
            raise APWorkbookRepositoryError("AP bill failed post-delete verification.")
        return {
            "ok": True,
            "APBillID": bill_id,
            "ExpenseTransactionID": clean_text(existing.get("ExpenseTransactionID")),
            "message": "Supplier bill deleted.",
        }

    def post_payment(self, payment: Mapping[str, Any]) -> dict[str, Any]:
        payment_id = clean_text(payment.get("APPaymentID"))
        bill_id = clean_text(payment.get("APBillID"))
        if not payment_id or not bill_id:
            raise APValidationError("APPaymentID and APBillID are required.")
        amount_value = money(payment.get("Amount"), "payment amount")
        if amount_value <= Decimal("0.00"):
            raise APValidationError("AP payment amount must be greater than zero.")
        workbook = _open(self.path)
        try:
            bills = _read_rows(workbook, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS)
            payments = _read_rows(workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS)
            if any(clean_text(row.get("APPaymentID")).casefold() == payment_id.casefold() for row in payments):
                raise APValidationError(f"Duplicate AP payment ID: {payment_id}")
            target = next((row for row in bills if clean_text(row.get("APBillID")).casefold() == bill_id.casefold()), None)
            if target is None:
                raise APValidationError(f"Unknown AP bill ID: {bill_id}")
            related = [row for row in payments if clean_text(row.get("APBillID")).casefold() == bill_id.casefold()]
            lifecycle_payments = [{"payment_id": row.get("APPaymentID"), "amount": row.get("Amount"), "reversed": clean_text(row.get("Status")).casefold() == "reversed"} for row in related]
            lifecycle_payments.append({"payment_id": payment_id, "amount": amount_value})
            snapshot = build_bill_snapshot(
                bill_id=target.get("APBillID"), vendor=target.get("Vendor"),
                vendor_invoice_number=target.get("VendorInvoiceNumber"), invoice_date=target.get("InvoiceDate"),
                subtotal=target.get("Subtotal"), tax_amount=target.get("TaxAmount"), payments=lifecycle_payments,
            )
            created = clean_text(payment.get("CreatedAt")) or _now()
            row = {header: "" for header in AP_PAYMENTS_HEADERS}
            row.update({
                "APPaymentID": payment_id, "APBillID": bill_id,
                "PaymentDate": clean_text(payment.get("PaymentDate")), "Amount": float(amount_value),
                "FromAccount": clean_text(payment.get("FromAccount")), "Method": clean_text(payment.get("Method")),
                "Reference": clean_text(payment.get("Reference")), "Status": "Posted",
                "Notes": clean_text(payment.get("Notes")), "CreatedAt": created, "UpdatedAt": created,
            })
            for header in AP_PAYMENTS_HEADERS:
                if header in payment and header not in {"APPaymentID", "APBillID", "Status", "CreatedAt", "UpdatedAt"}:
                    row[header] = payment.get(header, row.get(header, ""))
            payments.append(row)
            target.update({"AmountPaid": float(snapshot.total_paid), "Balance": float(snapshot.balance), "Status": snapshot.status.value, "UpdatedAt": _now()})
            _replace_rows(workbook, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS, bills)
            _replace_rows(workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS, payments)
            workbook.save(self.path)
        finally:
            _close_workbook(workbook)
        persisted = next((row for row in self.list_payments(bill_id) if clean_text(row.get("APPaymentID")).casefold() == payment_id.casefold()), None)
        if persisted is None:
            raise APWorkbookRepositoryError("AP payment failed post-save verification.")
        return persisted

    def reverse_payment(self, payment_id: str, reversal_id: str, reason: str) -> dict[str, Any]:
        payment_key = clean_text(payment_id).casefold()
        reversal_id = clean_text(reversal_id)
        if not reversal_id or not clean_text(reason):
            raise APValidationError("Reversal ID and reason are required.")
        workbook = _open(self.path)
        try:
            bills = _read_rows(workbook, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS)
            payments = _read_rows(workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS)
            original = next((row for row in payments if clean_text(row.get("APPaymentID")).casefold() == payment_key), None)
            if original is None:
                raise APValidationError(f"Unknown AP payment ID: {payment_id}")
            if clean_text(original.get("Status")).casefold() == "reversed":
                raise APValidationError("The AP payment is already reversed.")
            if any(clean_text(row.get("APPaymentID")).casefold() == reversal_id.casefold() for row in payments):
                raise APValidationError(f"Duplicate AP payment ID: {reversal_id}")
            original["Status"] = "Reversed"
            original["UpdatedAt"] = _now()
            reversal = {header: "" for header in AP_PAYMENTS_HEADERS}
            created = _now()
            reversal.update({
                "APPaymentID": reversal_id, "APBillID": original.get("APBillID"),
                "PaymentDate": created[:10], "Amount": original.get("Amount"),
                "FromAccount": original.get("FromAccount"), "Method": "Reversal",
                "Reference": original.get("Reference"), "Status": "Reversal",
                "ReversalOfPaymentID": original.get("APPaymentID"), "ReversalReason": clean_text(reason),
                "CreatedAt": created, "UpdatedAt": created,
            })
            for header in AP_PAYMENTS_HEADERS:
                if header in original and header not in {"APPaymentID", "APBillID", "Status", "CreatedAt", "UpdatedAt", "ReversalOfPaymentID", "ReversalReason"}:
                    reversal[header] = original.get(header, reversal.get(header, ""))
            payments.append(reversal)
            bill_id = clean_text(original.get("APBillID"))
            target = next(row for row in bills if clean_text(row.get("APBillID")).casefold() == bill_id.casefold())
            related = [row for row in payments if clean_text(row.get("APBillID")).casefold() == bill_id.casefold() and clean_text(row.get("Status")).casefold() != "reversal"]
            lifecycle = [{"payment_id": row.get("APPaymentID"), "amount": row.get("Amount"), "reversed": clean_text(row.get("Status")).casefold() == "reversed"} for row in related]
            snapshot = build_bill_snapshot(bill_id=target.get("APBillID"), vendor=target.get("Vendor"), vendor_invoice_number=target.get("VendorInvoiceNumber"), invoice_date=target.get("InvoiceDate"), subtotal=target.get("Subtotal"), tax_amount=target.get("TaxAmount"), payments=lifecycle)
            target.update({"AmountPaid": float(snapshot.total_paid), "Balance": float(snapshot.balance), "Status": snapshot.status.value, "UpdatedAt": _now()})
            _replace_rows(workbook, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS, bills)
            _replace_rows(workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS, payments)
            workbook.save(self.path)
        finally:
            _close_workbook(workbook)
        return next(row for row in self.list_payments(bill_id) if clean_text(row.get("APPaymentID")).casefold() == reversal_id.casefold())

    def update_payment_metadata(self, payment_id: str, metadata: Mapping[str, Any]) -> dict[str, Any]:
        """Add immutable linkage evidence after the cash-transfer transaction saves.

        This deliberately cannot change payment amount, bill, or lifecycle
        status; it is only for transaction/document identifiers produced by the
        governed orchestration layer.
        """
        key = clean_text(payment_id).casefold()
        if not key:
            raise APValidationError("APPaymentID is required.")
        workbook = _open(self.path)
        try:
            bills = _read_rows(workbook, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS)
            payments = _read_rows(workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS)
            row = next((item for item in payments if clean_text(item.get("APPaymentID")).casefold() == key), None)
            if row is None:
                raise APValidationError(f"Unknown AP payment ID: {payment_id}")
            for header in AP_PAYMENTS_HEADERS:
                if header in metadata:
                    row[header] = metadata[header]
            row["UpdatedAt"] = _now()
            _replace_rows(workbook, AP_BILLS_SHEET, AP_BILLS_TABLE, AP_BILLS_HEADERS, bills)
            _replace_rows(workbook, AP_PAYMENTS_SHEET, AP_PAYMENTS_TABLE, AP_PAYMENTS_HEADERS, payments)
            workbook.save(self.path)
        finally:
            _close_workbook(workbook)
        return next(row for row in self.list_payments() if clean_text(row.get("APPaymentID")).casefold() == key)
