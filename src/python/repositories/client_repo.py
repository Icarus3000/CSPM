from __future__ import annotations

import sys
import csv
import math
import json
import os
from dataclasses import dataclass
from datetime import datetime, UTC
from pathlib import Path
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple

import typing
if typing.TYPE_CHECKING:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils.cell import get_column_letter, range_boundaries
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.worksheet import Worksheet
else:
    Workbook = None
    
    MergedCell = None
    get_column_letter = None
    range_boundaries = None
    Table = None
    TableStyleInfo = None
    Worksheet = None

def _lazy_load_heavy_libs():
    global Workbook, load_workbook, MergedCell, get_column_letter, range_boundaries, Table, TableStyleInfo, Worksheet
    if Workbook is None:
        from openpyxl import Workbook, load_workbook
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils.cell import get_column_letter, range_boundaries
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.worksheet.worksheet import Worksheet


from uuid import uuid4

# Support direct execution via:
# python src/python/repositories/excel_repo.py
if __package__ in (None, ""):
    _THIS_FILE = Path(__file__).resolve()
    _PYTHON_ROOT = _THIS_FILE.parents[1]
    _PYTHON_ROOT_STR = str(_PYTHON_ROOT)
    if _PYTHON_ROOT_STR not in sys.path:
        sys.path.insert(0, _PYTHON_ROOT_STR)


from domain.money import calc_amounts, normalize_pct
from domain import schema_constants as sc
from services.paths import AppPaths
from repositories.excel_repo import ExcelRepo, TableRef

import logging
logger = logging.getLogger(__name__)

TBL_PARENTS = TableRef(sc.SHEET_PARENTS, sc.TBL_PARENTS)
TBL_CLIENTS = TableRef(sc.SHEET_CLIENTS, sc.TBL_CLIENTS)
TBL_CLIENT_PROFILES = TableRef(sc.SHEET_CLIENT_PROFILES, sc.TBL_CLIENT_PROFILES)
TBL_MATTERS = TableRef(sc.SHEET_MATTERS, sc.TBL_MATTERS)
TBL_TIME = TableRef(sc.SHEET_TIME, sc.TBL_TIME)
TBL_TRADEMARKS = TableRef(sc.SHEET_TRADEMARKS, sc.TBL_TRADEMARKS)
TBL_TRANSACTIONS_MASTER = TableRef(sc.SHEET_TRANSACTIONS, sc.TBL_TRANSACTIONS_MASTER)
TBL_TRANSACTION_ACCOUNTS = TableRef(sc.SHEET_TRANSACTION_ACCOUNTS, sc.TBL_TRANSACTION_ACCOUNTS)
TBL_TRANSACTION_CATEGORIES = TableRef(sc.SHEET_TRANSACTION_CATEGORIES, sc.TBL_TRANSACTION_CATEGORIES)
TBL_TRANSACTION_BUSINESS_UNITS = TableRef(sc.SHEET_TRANSACTION_BUSINESS_UNITS, sc.TBL_TRANSACTION_BUSINESS_UNITS)
TBL_TRANSACTION_PAYEES = TableRef(sc.SHEET_TRANSACTION_PAYEES, sc.TBL_TRANSACTION_PAYEES)

TABLES_IN_ORDER = [
    TBL_PARENTS,
    TBL_CLIENTS,
    TBL_CLIENT_PROFILES,
    TBL_MATTERS,
    TBL_TIME,
    TBL_TRADEMARKS,
    TBL_TRANSACTIONS_MASTER,
    TBL_TRANSACTION_ACCOUNTS,
    TBL_TRANSACTION_CATEGORIES,
    TBL_TRANSACTION_BUSINESS_UNITS,
    TBL_TRANSACTION_PAYEES,
]
TABLE_COLUMNS = sc.TABLE_COLUMNS
TABLE_ALIASES = sc.TABLE_ALIASES
TABLE_META_CACHE_SCHEMA_VERSION = 1


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_us_phone(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    digits = re.sub(r"\D+", "", text)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) != 10:
        return text
    return f"{digits[0:3]}-{digits[3:6]}-{digits[6:10]}"


def _looks_like_email_address(value: Any) -> bool:
    text = _clean_text(value)
    if not text:
        return True
    return bool(re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", text))


def _normalize_search_text(value: Any) -> str:
    text = _clean_text(value).lower()
    if not text:
        return ""
    return re.sub(r"\s+", " ", text)


def _matter_token_fragment(value: Any, width: int = 3, fallback: str = "XXX") -> str:
    """Legacy compat shim – kept for any call-sites not yet migrated."""
    token = re.sub(r"[^A-Za-z0-9]+", "", _clean_text(value).upper())
    if not token:
        token = _clean_text(fallback).upper() or "X"
    if len(token) >= width:
        return token[:width]
    return (token + ("X" * width))[:width]


# ---------------------------------------------------------------------------
# Legacy matter-number helpers
# ---------------------------------------------------------------------------

# Hardcoded 3-char matter-type codes.  Keys are normalised lowercase.
_MATTER_TYPE_CODES: Dict[str, str] = {
    "trademark": "TMK",
    "trademarks": "TMK",
    "tmk": "TMK",
    "commercial": "COM",
    "com": "COM",
    "corporate": "CRP",
    "crp": "CRP",
    "corporation": "CRP",
    "estate": "EST",
    "estates": "EST",
    "est": "EST",
    "tax": "TAX",
    "taxation": "TAX",
    "taxes": "TAX",
    "audit": "AUD",
    "audits": "AUD",
    "aud": "AUD",
    "income": "INC",
    "inc": "INC",
    "general": "GEN",
    "gen": "GEN",
}

# Words that should be stripped from the *end* of a company name before coding
_COMPANY_NOISE_SUFFIXES = {
    "inc", "ltd", "llc", "llp", "lp", "corp", "co", "plc", "pte", "pty",
    "sa", "srl", "sas", "gmbh", "ag", "nv", "bv", "ab",
}


def _legacy_matter_type_code(matter_type: Any) -> str:
    """Return the canonical 3-char matter-type code for the given string."""
    key = re.sub(r"[^a-z]", "", _clean_text(matter_type).lower())
    return _MATTER_TYPE_CODES.get(key, "GEN")


def _legacy_raw_client_code(name: Any, entity_type: Any = "") -> str:
    """
    Derive an *uncollision-checked* 4-character client code from *name*.

    Rules
    -----
    Individual  (EntityType == "Individual"):
        last[:4].  If len(last) < 4, pad with first[:needed].
    Company / anything else:
        Strip trailing noise suffixes, then take first-word[:4].
    In both cases, keep alphanumerics (including digits), pad with "X",
    return exactly 4 uppercase characters.
    """
    cleaned = re.sub(r"[^A-Za-z0-9 ]+", "", _clean_text(name)).strip()
    parts = cleaned.split()
    if not parts:
        return "XXXX"

    is_individual = _clean_text(entity_type).lower() == "individual"

    if is_individual:
        last = parts[-1].upper() if len(parts) >= 2 else parts[0].upper()
        raw = last[:4]
        if len(raw) < 4 and len(parts) >= 2:
            first = parts[0].upper()
            raw = raw + first[: 4 - len(raw)]
    else:
        # company: try the first meaningful word
        meaningful_parts = [p for p in parts if p.lower() not in _COMPANY_NOISE_SUFFIXES]
        main = (meaningful_parts[0] if meaningful_parts else parts[0]).upper()
        raw = main[:4]

    return (raw + "XXXX")[:4]


def _matter_year_two_digits(date_text: Any) -> str:
    text = _clean_text(date_text)
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", text)
    if match:
        return match.group(1)[2:]
    return datetime.now().strftime("%y")


def _normalize_choice(value: Any, allowed: List[str], default_value: str = "") -> str:
    normalized_allowed: Dict[str, str] = {
        _clean_text(choice).lower(): _clean_text(choice)
        for choice in allowed
        if _clean_text(choice)
    }
    if not normalized_allowed:
        return _clean_text(value) or _clean_text(default_value)
    raw = _clean_text(value)
    if not raw:
        return _clean_text(default_value) if _clean_text(default_value) else next(iter(normalized_allowed.values()))
    return normalized_allowed.get(raw.lower(), "")


def _is_valid_iso_date(text: Any) -> bool:
    value = _clean_text(text)
    if not value:
        return False
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", value):
        return False
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return True
    except Exception:
        return False


def _to_code_token(value: Any, fallback: str = "X") -> str:
    token = re.sub(r"[^A-Za-z0-9]+", "_", _clean_text(value).strip()).strip("_").upper()
    if token:
        return token
    return _clean_text(fallback).upper() or "X"


def _search_terms(value: Any) -> List[str]:
    normalized = _normalize_search_text(value)
    if not normalized:
        return []

    terms: List[str] = []
    seen = set()
    for raw_token in re.findall(r"[a-z0-9@._:/#\-]+", normalized, flags=re.IGNORECASE):
        token = raw_token.strip(" .,_-")
        if not token:
            continue
        if token in seen:
            continue
        seen.add(token)
        terms.append(token)

    if " " in normalized and normalized not in seen:
        terms.insert(0, normalized)
    return terms


def _boolean_query_tokens(value: Any) -> List[str]:
    raw = _clean_text(value)
    if not raw:
        return []
    return [tok for tok in re.findall(r"\(|\)|AND|OR|NOT|[^\s()]+", raw, flags=re.IGNORECASE) if tok]


def _boolean_to_postfix(tokens: List[str]) -> List[str]:
    precedence = {"OR": 1, "AND": 2, "NOT": 3}
    output: List[str] = []
    stack: List[str] = []

    for raw_tok in tokens:
        tok = str(raw_tok or "").strip()
        if not tok:
            continue
        upper = tok.upper()
        if upper in ("AND", "OR", "NOT"):
            if upper == "NOT":
                while stack and stack[-1] in precedence and precedence[stack[-1]] > precedence[upper]:
                    output.append(stack.pop())
            else:
                while stack and stack[-1] in precedence and precedence[stack[-1]] >= precedence[upper]:
                    output.append(stack.pop())
            stack.append(upper)
            continue
        if tok == "(":
            stack.append(tok)
            continue
        if tok == ")":
            while stack and stack[-1] != "(":
                output.append(stack.pop())
            if stack and stack[-1] == "(":
                stack.pop()
            continue
        output.append(_normalize_search_text(tok))

    while stack:
        top = stack.pop()
        if top in ("(", ")"):
            continue
        output.append(top)
    return output


def _match_boolean_postfix(postfix: List[str], haystack: str) -> bool:
    if not postfix:
        return False
    stack: List[bool] = []

    for tok in postfix:
        upper = tok.upper()
        if upper == "NOT":
            value = stack.pop() if stack else False
            stack.append(not value)
            continue
        if upper in ("AND", "OR"):
            right = stack.pop() if stack else False
            left = stack.pop() if stack else False
            stack.append(left and right if upper == "AND" else left or right)
            continue
        needle = _normalize_search_text(tok)
        stack.append(bool(needle) and needle in haystack)

    return stack[-1] if stack else False


def _normalize_index(value: Any, fallback: int) -> int:
    try:
        parsed = int(value)
        if parsed > 0:
            return parsed
    except Exception:
        pass
    return fallback


def _safe_range_boundaries(ref: str) -> Tuple[int, int, int, int]:
    _lazy_load_heavy_libs()
    min_col_raw, min_row_raw, max_col_raw, max_row_raw = range_boundaries(ref)
    min_col = _normalize_index(min_col_raw, 1)
    min_row = _normalize_index(min_row_raw, 1)
    max_col = _normalize_index(max_col_raw, min_col)
    max_row = _normalize_index(max_row_raw, min_row)
    if max_col < min_col:
        max_col = min_col
    if max_row < min_row:
        max_row = min_row
    return min_col, min_row, max_col, max_row


def _set_cell_value(ws: Worksheet, row_idx: int, col_idx: int, value: Any) -> None:
    _lazy_load_heavy_libs()
    cell_obj = ws.cell(row=row_idx, column=col_idx)
    if isinstance(cell_obj, MergedCell):
        return
    cell_obj.value = value

class ClientRepo:
    def __init__(self, db: ExcelRepo):
        self.db = db

    def list_parent_names(self) -> List[str]:
        rows = self.db._read_table_rows(TBL_PARENTS)
        names = [_clean_text(r.get(sc.COL_PARENT_NAME, "")) for r in rows]
        return sorted([name for name in names if name])

    def list_client_names(self) -> List[str]:
        rows = [self.db._canonicalize_client_row(r) for r in self.db._read_table_rows(TBL_CLIENTS)]
        names = [_clean_text(r.get(sc.COL_CLIENT_NAME, "")) for r in rows]
        return sorted([name for name in names if name])

    def list_active_client_names(self) -> List[str]:
        rows = [self.db._canonicalize_client_row(r) for r in self.db._read_table_rows(TBL_CLIENTS)]
        names: List[str] = []
        for row in rows:
            if not self._is_client_row_active(row):
                continue
            name = _clean_text(row.get(sc.COL_CLIENT_NAME, ""))
            if name:
                names.append(name)
        return sorted(names)

    def list_client_directory(self) -> List[Dict[str, Any]]:
        def _build_search_text(*sources: Dict[str, Any]) -> str:
            chunks: List[str] = []
            for source in sources:
                if not source:
                    continue
                for value in source.values():
                    if value is None:
                        continue
                    text = _clean_text(value)
                    if text:
                        chunks.append(text)
            return " ".join(chunks)

        client_rows = [self.db._canonicalize_client_row(r) for r in self.db._read_table_rows(TBL_CLIENTS)]
        profile_rows = [self.db._canonicalize_client_profile_row(r) for r in self.db._read_table_rows(TBL_CLIENT_PROFILES)]

        profile_by_id: Dict[str, Dict[str, Any]] = {}
        for row in profile_rows:
            client_id = _clean_text(row.get(sc.COL_PROFILE_CLIENT_ID))
            if client_id:
                profile_by_id[client_id] = row

        directory: List[Dict[str, Any]] = []
        seen_ids = set()
        for client_row in client_rows:
            client_id = _clean_text(client_row.get(sc.COL_CLIENT_ID))
            if client_id:
                seen_ids.add(client_id.lower())
            profile_row = profile_by_id.get(client_id, {})
            status = _clean_text(client_row.get(sc.COL_CLIENT_STATUS)) or "Active"
            active_flag = 1 if self._is_client_row_active(client_row) else 0
            search_text = _build_search_text(client_row, profile_row)
            directory.append(
                {
                    "clientId": client_id,
                    "clientName": _clean_text(client_row.get(sc.COL_CLIENT_NAME)),
                    "displayName": _clean_text(profile_row.get(sc.COL_PROFILE_DISPLAY_NAME))
                    or _clean_text(client_row.get(sc.COL_CLIENT_NAME)),
                    "legalName": _clean_text(profile_row.get(sc.COL_PROFILE_LEGAL_NAME)),
                    "firstName": _clean_text(profile_row.get(sc.COL_PROFILE_FIRST_NAME)),
                    "middleName": _clean_text(profile_row.get(sc.COL_PROFILE_MIDDLE_NAME)),
                    "lastName": _clean_text(profile_row.get(sc.COL_PROFILE_LAST_NAME)),
                    "entityType": _clean_text(profile_row.get(sc.COL_PROFILE_ENTITY_TYPE)),
                    "status": status,
                    "active": active_flag,
                    "primaryEmail": _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_EMAIL))
                    or _clean_text(client_row.get(sc.COL_CLIENT_EMAIL)),
                    "primaryPhone": _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_PHONE))
                    or _clean_text(client_row.get(sc.COL_CLIENT_PHONE)),
                    "parentClientName": _clean_text(profile_row.get(sc.COL_PROFILE_PARENT_NAME)),
                    "onboardingStatus": _clean_text(profile_row.get(sc.COL_PROFILE_ONBOARDING_STATUS)),
                    "kycStatus": _clean_text(profile_row.get(sc.COL_PROFILE_KYC_STATUS)),
                    "updatedAt": _clean_text(profile_row.get(sc.COL_PROFILE_UPDATED)),
                    "searchText": search_text,
                }
            )

        # Include orphan profile rows if the profile exists but the client row is missing.
        for profile_row in profile_rows:
            client_id = _clean_text(profile_row.get(sc.COL_PROFILE_CLIENT_ID))
            if not client_id or client_id.lower() in seen_ids:
                continue
            search_text = _build_search_text(profile_row)
            directory.append(
                {
                    "clientId": client_id,
                    "clientName": _clean_text(profile_row.get(sc.COL_PROFILE_DISPLAY_NAME))
                    or _clean_text(profile_row.get(sc.COL_PROFILE_LEGAL_NAME)),
                    "displayName": _clean_text(profile_row.get(sc.COL_PROFILE_DISPLAY_NAME))
                    or _clean_text(profile_row.get(sc.COL_PROFILE_LEGAL_NAME)),
                    "legalName": _clean_text(profile_row.get(sc.COL_PROFILE_LEGAL_NAME)),
                    "firstName": _clean_text(profile_row.get(sc.COL_PROFILE_FIRST_NAME)),
                    "middleName": _clean_text(profile_row.get(sc.COL_PROFILE_MIDDLE_NAME)),
                    "lastName": _clean_text(profile_row.get(sc.COL_PROFILE_LAST_NAME)),
                    "entityType": _clean_text(profile_row.get(sc.COL_PROFILE_ENTITY_TYPE)),
                    "status": "Active",
                    "active": 1,
                    "primaryEmail": _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_EMAIL)),
                    "primaryPhone": _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_PHONE)),
                    "parentClientName": _clean_text(profile_row.get(sc.COL_PROFILE_PARENT_NAME)),
                    "onboardingStatus": _clean_text(profile_row.get(sc.COL_PROFILE_ONBOARDING_STATUS)),
                    "kycStatus": _clean_text(profile_row.get(sc.COL_PROFILE_KYC_STATUS)),
                    "updatedAt": _clean_text(profile_row.get(sc.COL_PROFILE_UPDATED)),
                    "searchText": search_text,
                }
            )

        directory.sort(
            key=lambda row: (
                _clean_text(row.get("displayName")).lower(),
                _clean_text(row.get("clientName")).lower(),
                _clean_text(row.get("clientId")).lower(),
            )
        )
        return directory

    def get_client_profile(self, client_key: str) -> Dict[str, Any]:
        lookup = _clean_text(client_key)
        if not lookup:
            return {"ok": False, "message": "Client key is required.", "client": {}}

        client_rows = [self.db._canonicalize_client_row(r) for r in self.db._read_table_rows(TBL_CLIENTS)]
        profile_rows = [self.db._canonicalize_client_profile_row(r) for r in self.db._read_table_rows(TBL_CLIENT_PROFILES)]
        profile_by_id: Dict[str, Dict[str, Any]] = {
            _clean_text(r.get(sc.COL_PROFILE_CLIENT_ID)): r
            for r in profile_rows
            if _clean_text(r.get(sc.COL_PROFILE_CLIENT_ID))
        }

        lookup_lower = lookup.lower()
        client_row: Optional[Dict[str, Any]] = None
        for row in client_rows:
            if _clean_text(row.get(sc.COL_CLIENT_ID)).lower() == lookup_lower:
                client_row = row
                break
        if client_row is None:
            for row in client_rows:
                if _clean_text(row.get(sc.COL_CLIENT_NAME)).lower() == lookup_lower:
                    client_row = row
                    break

        profile_row: Optional[Dict[str, Any]] = None
        if client_row is not None:
            profile_row = profile_by_id.get(_clean_text(client_row.get(sc.COL_CLIENT_ID)))

        if profile_row is None:
            for row in profile_rows:
                if _clean_text(row.get(sc.COL_PROFILE_CLIENT_ID)).lower() == lookup_lower:
                    profile_row = row
                    break
                if _clean_text(row.get(sc.COL_PROFILE_DISPLAY_NAME)).lower() == lookup_lower:
                    profile_row = row
                    break
                if _clean_text(row.get(sc.COL_PROFILE_LEGAL_NAME)).lower() == lookup_lower:
                    profile_row = row
                    break

        if client_row is None and profile_row is not None:
            profile_id = _clean_text(profile_row.get(sc.COL_PROFILE_CLIENT_ID))
            for row in client_rows:
                if _clean_text(row.get(sc.COL_CLIENT_ID)) == profile_id:
                    client_row = row
                    break

        if client_row is None and profile_row is None:
            return {"ok": False, "message": f"Client not found: {lookup}", "client": {}}

        client_id = _clean_text(
            (client_row or {}).get(sc.COL_CLIENT_ID) or (profile_row or {}).get(sc.COL_PROFILE_CLIENT_ID)
        )
        client_name = _clean_text((client_row or {}).get(sc.COL_CLIENT_NAME))
        display_name = _clean_text((profile_row or {}).get(sc.COL_PROFILE_DISPLAY_NAME)) or client_name
        legal_name = _clean_text((profile_row or {}).get(sc.COL_PROFILE_LEGAL_NAME)) or display_name or client_name
        status = _clean_text((client_row or {}).get(sc.COL_CLIENT_STATUS)) or "Active"
        active = 1 if self._is_client_row_active(client_row or {}) else 0

        full_address = _clean_text((profile_row or {}).get(sc.COL_PROFILE_FULL_ADDRESS))
        if not full_address:
            full_address = self.db._format_full_address(
                line1=_clean_text((profile_row or {}).get(sc.COL_PROFILE_ADDR1)),
                line2=_clean_text((profile_row or {}).get(sc.COL_PROFILE_ADDR2)),
                city=_clean_text((profile_row or {}).get(sc.COL_PROFILE_CITY)),
                state_province=_clean_text((profile_row or {}).get(sc.COL_PROFILE_STATE)),
                postal_code=_clean_text((profile_row or {}).get(sc.COL_PROFILE_POSTAL)),
                country=_clean_text((profile_row or {}).get(sc.COL_PROFILE_COUNTRY)),
            )

        client_payload = {
            "clientId": client_id,
            "clientName": client_name,
            "displayName": display_name,
            "legalName": legal_name,
            "status": status,
            "active": active,
            "firstName": _clean_text((profile_row or {}).get(sc.COL_PROFILE_FIRST_NAME)),
            "middleName": _clean_text((profile_row or {}).get(sc.COL_PROFILE_MIDDLE_NAME)),
            "lastName": _clean_text((profile_row or {}).get(sc.COL_PROFILE_LAST_NAME)),
            "entityType": _clean_text((profile_row or {}).get(sc.COL_PROFILE_ENTITY_TYPE)),
            "principalName": _clean_text((profile_row or {}).get(sc.COL_PROFILE_PRINCIPAL_NAME)),
            "principalPosition": _clean_text((profile_row or {}).get(sc.COL_PROFILE_PRINCIPAL_POSITION)),
            "primaryEmail": _clean_text((profile_row or {}).get(sc.COL_PROFILE_PRIMARY_EMAIL))
            or _clean_text((client_row or {}).get(sc.COL_CLIENT_EMAIL)),
            "primaryPhone": _clean_text((profile_row or {}).get(sc.COL_PROFILE_PRIMARY_PHONE))
            or _clean_text((client_row or {}).get(sc.COL_CLIENT_PHONE)),
            "secondaryContactName": _clean_text((profile_row or {}).get(sc.COL_PROFILE_SECONDARY_CONTACT)),
            "secondaryContactPosition": _clean_text((profile_row or {}).get(sc.COL_PROFILE_SECONDARY_POSITION)),
            "secondaryContactEmail": _clean_text((profile_row or {}).get(sc.COL_PROFILE_SECONDARY_EMAIL)),
            "secondaryContactPhone": _clean_text((profile_row or {}).get(sc.COL_PROFILE_SECONDARY_PHONE)),
            "addressLine1": _clean_text((profile_row or {}).get(sc.COL_PROFILE_ADDR1)),
            "addressLine2": _clean_text((profile_row or {}).get(sc.COL_PROFILE_ADDR2)),
            "city": _clean_text((profile_row or {}).get(sc.COL_PROFILE_CITY)),
            "stateProvince": _clean_text((profile_row or {}).get(sc.COL_PROFILE_STATE)),
            "postalCode": _clean_text((profile_row or {}).get(sc.COL_PROFILE_POSTAL)),
            "country": _clean_text((profile_row or {}).get(sc.COL_PROFILE_COUNTRY)),
            "fullAddress": full_address,
            "parentClientId": _clean_text((profile_row or {}).get(sc.COL_PROFILE_PARENT_ID)),
            "parentClientName": _clean_text((profile_row or {}).get(sc.COL_PROFILE_PARENT_NAME)),
            "website": _clean_text((profile_row or {}).get(sc.COL_PROFILE_WEBSITE)),
            "taxId": _clean_text((profile_row or {}).get(sc.COL_PROFILE_TAX_ID)),
            "industry": _clean_text((profile_row or {}).get(sc.COL_PROFILE_INDUSTRY)),
            "billingEmail": _clean_text((profile_row or {}).get(sc.COL_PROFILE_BILLING_EMAIL)),
            "kycStatus": _clean_text((profile_row or {}).get(sc.COL_PROFILE_KYC_STATUS)),
            "onboardingStatus": _clean_text((profile_row or {}).get(sc.COL_PROFILE_ONBOARDING_STATUS)),
            "retainerRequired": self.db._to_bool_int(
                (profile_row or {}).get(sc.COL_PROFILE_RETAINER_REQUIRED),
                default=0,
            ),
            "retainerAmount": round(
                float(self.db._parse_float((profile_row or {}).get(sc.COL_PROFILE_RETAINER_AMOUNT)) or 0.0),
                2,
            ),
            "engagementStartDate": _clean_text((profile_row or {}).get(sc.COL_PROFILE_ENGAGEMENT_START)),
            "dateClientAdded": _clean_text((profile_row or {}).get(sc.COL_PROFILE_DATE_CLIENT_ADDED)),
            "birthday": _clean_text((profile_row or {}).get(sc.COL_PROFILE_BIRTHDAY)),
            "referralFrom": _clean_text((profile_row or {}).get(sc.COL_PROFILE_REFERRAL_FROM)),
            "conflictNotes": _clean_text((profile_row or {}).get(sc.COL_PROFILE_CONFLICT_NOTES)),
            "notes": _clean_text((profile_row or {}).get(sc.COL_PROFILE_NOTES))
            or _clean_text((client_row or {}).get(sc.COL_CLIENT_NOTES)),
            "createdAt": _clean_text((profile_row or {}).get(sc.COL_PROFILE_CREATED)),
            "updatedAt": _clean_text((profile_row or {}).get(sc.COL_PROFILE_UPDATED)),
        }
        return {"ok": True, "message": "", "client": client_payload}

    def list_matter_names(self) -> List[str]:
        rows = [self.db._canonicalize_matter_row(r) for r in self.db._read_table_rows(TBL_MATTERS)]
        names = [_clean_text(r.get(sc.COL_MATTER_NAME, "")) for r in rows]
        return sorted([name for name in names if name])

    def list_active_matter_names(self) -> List[str]:
        rows = [self.db._canonicalize_matter_row(r) for r in self.db._read_table_rows(TBL_MATTERS)]
        names: List[str] = []
        for row in rows:
            if not self._is_matter_row_active(row):
                continue
            name = _clean_text(row.get(sc.COL_MATTER_NAME, ""))
            if name:
                names.append(name)
        return sorted(names)


    def preview_matter_number(
        self,
        client_name: str,
        matter_type: str,
        date_opened: str,
        existing_matter_id: str = "",
    ) -> str:
        try:
            return self._build_matter_number(
                client_name=client_name,
                matter_type=matter_type,
                date_opened=date_opened,
                existing_matter_id=existing_matter_id,
                parent_name="",
                entity_type="",
            )
        except Exception:
            return ""

    def list_matter_directory(self) -> List[Dict[str, Any]]:
        def _build_search_text(*sources: Dict[str, Any]) -> str:
            chunks: List[str] = []
            for source in sources:
                if not source:
                    continue
                for value in source.values():
                    if value is None:
                        continue
                    text = _clean_text(value)
                    if text:
                        chunks.append(text)
            return " ".join(chunks)

        matter_rows = [self.db._canonicalize_matter_row(r) for r in self.db._read_table_rows(TBL_MATTERS)]
        client_rows = [self.db._canonicalize_client_row(r) for r in self.db._read_table_rows(TBL_CLIENTS)]
        parent_rows = [self.db._canonicalize_parent_row(r) for r in self.db._read_table_rows(TBL_PARENTS)]

        client_name_by_id: Dict[str, str] = {}
        for row in client_rows:
            cid = _clean_text(row.get(sc.COL_CLIENT_ID))
            cname = _clean_text(row.get(sc.COL_CLIENT_NAME))
            if cid and cname:
                client_name_by_id[cid] = cname

        parent_name_by_id: Dict[str, str] = {}
        for row in parent_rows:
            pid = _clean_text(row.get(sc.COL_PARENT_ID))
            pname = _clean_text(row.get(sc.COL_PARENT_NAME))
            if pid and pname:
                parent_name_by_id[pid] = pname

        directory: List[Dict[str, Any]] = []
        for matter_row in matter_rows:
            matter_id = _clean_text(matter_row.get(sc.COL_MATTER_ID))
            matter_number = _clean_text(matter_row.get(sc.COL_MATTER_NUMBER))
            matter_name = _clean_text(matter_row.get(sc.COL_MATTER_NAME))
            display_name = _clean_text(matter_row.get(sc.COL_MATTER_DISPLAY_NAME)) or matter_name
            client_id = _clean_text(matter_row.get(sc.COL_MATTER_CLIENT_ID))
            parent_id = _clean_text(matter_row.get(sc.COL_MATTER_PARENT_ID))
            client_name = _clean_text(matter_row.get(sc.COL_MATTER_CLIENT_NAME)) or client_name_by_id.get(
                client_id, ""
            )
            parent_name = _clean_text(matter_row.get(sc.COL_MATTER_PARENT_NAME)) or parent_name_by_id.get(
                parent_id, ""
            )
            status = _clean_text(matter_row.get(sc.COL_MATTER_STATUS)) or "Open"
            active = 1 if self._is_matter_row_active(matter_row) else 0
            search_text = _build_search_text(
                matter_row,
                {"clientName": client_name, "parentName": parent_name},
            )
            directory.append(
                {
                    "matterId": matter_id,
                    "matterNumber": _clean_text(matter_row.get(sc.COL_MATTER_NUMBER)),
                    "matterName": matter_name,
                    "displayName": display_name,
                    "clientId": client_id,
                    "clientName": client_name,
                    "parentId": parent_id,
                    "parentName": parent_name,
                    "matterType": _clean_text(matter_row.get(sc.COL_MATTER_TYPE)),
                    "practiceArea": _clean_text(matter_row.get(sc.COL_MATTER_PRACTICE_AREA)),
                    "responsibleLawyer": _clean_text(matter_row.get(sc.COL_MATTER_RESPONSIBLE_LAWYER)),
                    "billingArrangement": _clean_text(
                        matter_row.get(sc.COL_MATTER_BILLING_ARRANGEMENT)
                    ),
                    "status": status,
                    "active": active,
                    "dateOpened": _clean_text(matter_row.get(sc.COL_MATTER_OPEN_DATE)),
                    "updatedAt": _clean_text(matter_row.get(sc.COL_MATTER_UPDATED)),
                    "searchText": search_text,
                }
            )

        directory.sort(
            key=lambda row: (
                _clean_text(row.get("displayName")).lower(),
                _clean_text(row.get("matterName")).lower(),
                _clean_text(row.get("matterId")).lower(),
            )
        )
        return directory

    def list_active_matter_directory(self) -> List[Dict[str, Any]]:
        # reuse existing directory builder and filter by active flag
        all_rows = self.list_matter_directory()
        return [r for r in all_rows if r.get("active") == 1]

    def get_matter_profile(self, matter_key: str) -> Dict[str, Any]:
        lookup = _clean_text(matter_key)
        if not lookup:
            return {"ok": False, "message": "Matter key is required.", "matter": {}}

        matter_rows = [self.db._canonicalize_matter_row(r) for r in self.db._read_table_rows(TBL_MATTERS)]
        client_rows = [self.db._canonicalize_client_row(r) for r in self.db._read_table_rows(TBL_CLIENTS)]
        parent_rows = [self.db._canonicalize_parent_row(r) for r in self.db._read_table_rows(TBL_PARENTS)]

        client_name_by_id: Dict[str, str] = {}
        for row in client_rows:
            cid = _clean_text(row.get(sc.COL_CLIENT_ID))
            cname = _clean_text(row.get(sc.COL_CLIENT_NAME))
            if cid and cname:
                client_name_by_id[cid] = cname

        parent_name_by_id: Dict[str, str] = {}
        for row in parent_rows:
            pid = _clean_text(row.get(sc.COL_PARENT_ID))
            pname = _clean_text(row.get(sc.COL_PARENT_NAME))
            if pid and pname:
                parent_name_by_id[pid] = pname

        lookup_lc = lookup.lower()
        selected: Optional[Dict[str, Any]] = None
        for row in matter_rows:
            if _clean_text(row.get(sc.COL_MATTER_ID)).lower() == lookup_lc:
                selected = row
                break
        if selected is None:
            for row in matter_rows:
                if _clean_text(row.get(sc.COL_MATTER_NAME)).lower() == lookup_lc:
                    selected = row
                    break
        if selected is None:
            for row in matter_rows:
                if _clean_text(row.get(sc.COL_MATTER_DISPLAY_NAME)).lower() == lookup_lc:
                    selected = row
                    break
        if selected is None:
            for row in matter_rows:
                if _clean_text(row.get(sc.COL_MATTER_COURT_FILE_NO)).lower() == lookup_lc:
                    selected = row
                    break
        if selected is None:
            for row in matter_rows:
                if _clean_text(row.get(sc.COL_MATTER_NUMBER)).lower() == lookup_lc:
                    selected = row
                    break

        if selected is None:
            return {"ok": False, "message": f"Matter not found: {lookup}", "matter": {}}

        client_id = _clean_text(selected.get(sc.COL_MATTER_CLIENT_ID))
        parent_id = _clean_text(selected.get(sc.COL_MATTER_PARENT_ID))
        client_name = _clean_text(selected.get(sc.COL_MATTER_CLIENT_NAME)) or client_name_by_id.get(client_id, "")
        parent_name = _clean_text(selected.get(sc.COL_MATTER_PARENT_NAME)) or parent_name_by_id.get(parent_id, "")

        matter_payload = {
            "matterId": _clean_text(selected.get(sc.COL_MATTER_ID)),
            "matterNumber": _clean_text(selected.get(sc.COL_MATTER_NUMBER)),
            "matterName": _clean_text(selected.get(sc.COL_MATTER_NAME)),
            "displayName": _clean_text(selected.get(sc.COL_MATTER_DISPLAY_NAME))
            or _clean_text(selected.get(sc.COL_MATTER_NAME)),
            "clientId": client_id,
            "clientName": client_name,
            "parentId": parent_id,
            "parentName": parent_name,
            "matterType": _clean_text(selected.get(sc.COL_MATTER_TYPE)),
            "practiceArea": _clean_text(selected.get(sc.COL_MATTER_PRACTICE_AREA)),
            "status": _clean_text(selected.get(sc.COL_MATTER_STATUS)) or "Open",
            "responsibleLawyer": _clean_text(selected.get(sc.COL_MATTER_RESPONSIBLE_LAWYER)),
            "billingArrangement": _clean_text(selected.get(sc.COL_MATTER_BILLING_ARRANGEMENT)),
            "billingContact": _clean_text(selected.get(sc.COL_MATTER_BILLING_CONTACT)),
            "billingEmail": _clean_text(selected.get(sc.COL_MATTER_BILLING_EMAIL)),
            "defaultRate": round(float(self.db._parse_float(selected.get(sc.COL_MATTER_DEF_RATE)) or 0.0), 2),
            "defaultSharePct": round(
                normalize_pct(selected.get(sc.COL_MATTER_DEF_SHARE), default_pct=100.0),
                2,
            ),
            "rateHistory": _clean_text(selected.get(sc.COL_MATTER_RATE_HISTORY)) or "[]",
            "dateOfEngagement": _clean_text(selected.get(sc.COL_MATTER_ENGAGEMENT_DATE)),
            "dateOpened": _clean_text(selected.get(sc.COL_MATTER_OPEN_DATE)),
            "dateClosed": _clean_text(selected.get(sc.COL_MATTER_CLOSE_DATE)),
            "courtFileNumber": _clean_text(selected.get(sc.COL_MATTER_COURT_FILE_NO)),
            "opposingParty": _clean_text(selected.get(sc.COL_MATTER_OPPOSING_PARTY)),
            "referralFrom": _clean_text(selected.get(sc.COL_MATTER_REFERRAL_FROM)),
            "description": _clean_text(selected.get(sc.COL_MATTER_DESCRIPTION)),
            "notes": _clean_text(selected.get(sc.COL_MATTER_NOTES)),
            "active": 1 if self._is_matter_row_active(selected) else 0,
            "createdAt": _clean_text(selected.get(sc.COL_MATTER_CREATED)),
            "updatedAt": _clean_text(selected.get(sc.COL_MATTER_UPDATED)),
        }
        return {"ok": True, "message": "", "matter": matter_payload}

    def search_global_entities(self, query: str, mode: str = "any", limit: int = 250) -> Dict[str, Any]:
        raw_query = _clean_text(query)
        normalized_query = _normalize_search_text(raw_query)
        normalized_mode = _clean_text(mode).lower()
        search_mode = "boolean" if normalized_mode == "boolean" else "any"
        try:
            max_results = int(limit)
        except Exception:
            max_results = 250
        max_results = max(1, min(max_results, 1000))

        facets: Dict[str, int] = {
            "client": 0,
            "matter": 0,
            "parent": 0,
            "invoice": 0,
            "transaction": 0,
            "account": 0,
            "category": 0,
            "business_unit": 0,
            "payee": 0,
            "tickler": 0,
            "deadline": 0,
            "docket": 0,
            "trademark": 0,
        }
        entity_labels = {
            "client": "Client",
            "matter": "Matter",
            "parent": "Parent",
            "invoice": "Invoice",
            "transaction": "Transaction",
            "account": "Account",
            "category": "Category",
            "business_unit": "Business Unit",
            "payee": "Payee",
            "tickler": "Tickler",
            "deadline": "Deadline",
            "docket": "Docket",
            "trademark": "Trademark",
        }

        empty_payload = {
            "ok": True,
            "query": raw_query,
            "mode": search_mode,
            "results": [],
            "facets": dict(facets),
            "total": 0,
            "returnedCount": 0,
        }
        if not normalized_query:
            return empty_payload

        any_terms = _search_terms(normalized_query)
        boolean_tokens = _boolean_query_tokens(raw_query)
        boolean_postfix = _boolean_to_postfix(boolean_tokens) if search_mode == "boolean" else []
        boolean_terms = [t for t in _search_terms(raw_query) if t not in ("and", "or", "not")]
        if search_mode == "boolean" and not boolean_postfix:
            search_mode = "any"

        try:
            self.db.ensure_schema()
            client_rows = [self.db._canonicalize_client_row(r) for r in self.db._read_table_rows(TBL_CLIENTS)]
            profile_rows = [
                self.db._canonicalize_client_profile_row(r)
                for r in self.db._read_table_rows(TBL_CLIENT_PROFILES)
            ]
            parent_rows = [self.db._canonicalize_parent_row(r) for r in self.db._read_table_rows(TBL_PARENTS)]
            matter_rows = [self.db._canonicalize_matter_row(r) for r in self.db._read_table_rows(TBL_MATTERS)]
            time_rows = [self.db._canonicalize_time_row(r) for r in self.db._read_table_rows(TBL_TIME)]
            trademark_rows = [self.db._canonicalize_trademark_row(r) for r in self.db._read_table_rows(TBL_TRADEMARKS)]
            txn_rows = [
                self.db._canonicalize_transaction_row(r)
                for r in self.db._read_table_rows(TBL_TRANSACTIONS_MASTER)
            ]
            account_rows = [
                self.db._canonicalize_transaction_account_row(r)
                for r in self.db._read_table_rows(TBL_TRANSACTION_ACCOUNTS)
            ]
            category_rows = [
                self.db._canonicalize_transaction_category_row(r)
                for r in self.db._read_table_rows(TBL_TRANSACTION_CATEGORIES)
            ]
            business_unit_rows = [
                self.db._canonicalize_transaction_business_unit_row(r)
                for r in self.db._read_table_rows(TBL_TRANSACTION_BUSINESS_UNITS)
            ]
            payee_rows = [
                self.db._canonicalize_transaction_payee_row(r)
                for r in self.db._read_table_rows(TBL_TRANSACTION_PAYEES)
            ]
        except Exception as exc:
            return {
                "ok": False,
                "query": raw_query,
                "mode": search_mode,
                "results": [],
                "facets": dict(facets),
                "total": 0,
                "returnedCount": 0,
                "message": str(exc),
            }

        profile_by_client_id: Dict[str, Dict[str, Any]] = {}
        for row in profile_rows:
            client_id = _clean_text(row.get(sc.COL_PROFILE_CLIENT_ID))
            if client_id:
                profile_by_client_id[client_id] = row

        client_name_by_id: Dict[str, str] = {}
        for row in client_rows:
            client_id = _clean_text(row.get(sc.COL_CLIENT_ID))
            client_name = _clean_text(row.get(sc.COL_CLIENT_NAME))
            if client_id and client_name:
                client_name_by_id[client_id] = client_name

        parent_name_by_id: Dict[str, str] = {}
        for row in parent_rows:
            parent_id = _clean_text(row.get(sc.COL_PARENT_ID))
            parent_name = _clean_text(row.get(sc.COL_PARENT_NAME))
            if parent_id and parent_name:
                parent_name_by_id[parent_id] = parent_name

        matter_name_by_id: Dict[str, str] = {}
        for row in matter_rows:
            matter_id = _clean_text(row.get(sc.COL_MATTER_ID))
            matter_name = _clean_text(row.get(sc.COL_MATTER_DISPLAY_NAME)) or _clean_text(
                row.get(sc.COL_MATTER_NAME)
            )
            if matter_id and matter_name:
                matter_name_by_id[matter_id] = matter_name

        results: List[Dict[str, Any]] = []

        def _field_weight(label: str) -> int:
            normalized_label = _normalize_search_text(label)
            if any(token in normalized_label for token in ("notes", "description", "details", "audit")):
                return 4
            if normalized_label in {
                "client name",
                "display name",
                "legal name",
                "principal name",
                "parent name",
                "matter name",
                "account name",
                "category name",
                "business unit",
                "payee",
                "payee name",
                "trademark text",
            }:
                return 34
            if any(
                token in normalized_label
                for token in (
                    "email",
                    "phone",
                    "responsible lawyer",
                    "billing contact",
                    "member",
                    "owner",
                    "applicant",
                )
            ):
                return 20
            if any(token in normalized_label for token in ("id", "number", "code", "ref")):
                return 18
            if any(token in normalized_label for token in ("status", "type", "class", "scope", "tax")):
                return 10
            return 12

        def _match_fields(field_pairs: List[Tuple[str, Any]], title_text: str) -> Optional[Tuple[int, List[str]]]:
            normalized_fields: List[Tuple[str, str]] = []
            for label, raw_value in field_pairs:
                value_text = _clean_text(raw_value)
                if not value_text:
                    continue
                normalized_value = _normalize_search_text(value_text)
                if not normalized_value:
                    continue
                normalized_fields.append((str(label), normalized_value))
            if not normalized_fields:
                return None

            haystack = " | ".join(v for _, v in normalized_fields if v)
            if not haystack:
                return None

            active_terms = any_terms
            if search_mode == "boolean" and boolean_postfix:
                if not _match_boolean_postfix(boolean_postfix, haystack):
                    return None
                active_terms = boolean_terms
            else:
                if not any(term and term in haystack for term in any_terms):
                    return None

            matched_terms = 0
            seen_terms = set()
            for term in active_terms:
                if not term or term in seen_terms:
                    continue
                if term in haystack:
                    matched_terms += 1
                    seen_terms.add(term)

            matched_fields: List[str] = []
            seen_fields = set()
            for label, value_text in normalized_fields:
                for term in active_terms:
                    if term and term in value_text:
                        key = label.lower()
                        if key not in seen_fields:
                            seen_fields.add(key)
                            matched_fields.append(label)
                        break

            title_lc = _normalize_search_text(title_text)
            score = matched_terms * 6
            for term in active_terms:
                if not term:
                    continue
                if title_lc.startswith(term):
                    score += 28
                elif term in title_lc:
                    score += 14

            for label, value_text in normalized_fields:
                weight = _field_weight(label)
                for term in active_terms:
                    if not term or term not in value_text:
                        continue
                    score += weight
                    if value_text == term:
                        score += 16
                    elif value_text.startswith(term):
                        score += 8
                    break

            return score, matched_fields[:4]

        def _append_result(
            *,
            entity_type: str,
            entity_id: str,
            title: str,
            subtitle: str,
            status: str,
            route_tile_index: int,
            route_node_id: str,
            route_node_title: str,
            field_pairs: List[Tuple[str, Any]],
            extra: Optional[Dict[str, Any]] = None,
        ) -> None:
            title_text = _clean_text(title) or _clean_text(entity_id) or "Untitled"
            match_meta = _match_fields(field_pairs, title_text)
            if not match_meta:
                return
            score, matched_fields = match_meta

            row = {
                "entityType": entity_type,
                "entityTypeLabel": entity_labels.get(entity_type, entity_type.title()),
                "entityId": _clean_text(entity_id),
                "title": title_text,
                "subtitle": _clean_text(subtitle),
                "status": _clean_text(status),
                "routeTileIndex": int(route_tile_index),
                "routeNodeId": _clean_text(route_node_id),
                "routeNodeTitle": _clean_text(route_node_title),
                "matchedFields": matched_fields,
                "_score": int(score),
            }
            if extra:
                row.update(extra)
            results.append(row)
            if entity_type in facets:
                facets[entity_type] += 1

        seen_client_ids = set()
        for client_row in client_rows:
            client_id = _clean_text(client_row.get(sc.COL_CLIENT_ID))
            if client_id:
                seen_client_ids.add(client_id.lower())
            profile_row = profile_by_client_id.get(client_id, {})

            client_name = _clean_text(client_row.get(sc.COL_CLIENT_NAME))
            display_name = _clean_text(profile_row.get(sc.COL_PROFILE_DISPLAY_NAME)) or client_name
            legal_name = _clean_text(profile_row.get(sc.COL_PROFILE_LEGAL_NAME))
            status = _clean_text(client_row.get(sc.COL_CLIENT_STATUS)) or "Active"
            primary_email = _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_EMAIL)) or _clean_text(
                client_row.get(sc.COL_CLIENT_EMAIL)
            )
            primary_phone = _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_PHONE)) or _clean_text(
                client_row.get(sc.COL_CLIENT_PHONE)
            )

            subtitle_bits = [x for x in (legal_name, primary_email, status) if x]
            field_pairs: List[Tuple[str, Any]] = [
                ("Client ID", client_id),
                ("Client Name", client_name),
                ("Display Name", display_name),
                ("Legal Name", legal_name),
                ("First Name", profile_row.get(sc.COL_PROFILE_FIRST_NAME)),
                ("Middle Name", profile_row.get(sc.COL_PROFILE_MIDDLE_NAME)),
                ("Last Name", profile_row.get(sc.COL_PROFILE_LAST_NAME)),
                ("Client Status", status),
                ("Entity Type", profile_row.get(sc.COL_PROFILE_ENTITY_TYPE)),
                ("Principal Name", profile_row.get(sc.COL_PROFILE_PRINCIPAL_NAME)),
                ("Principal Position", profile_row.get(sc.COL_PROFILE_PRINCIPAL_POSITION)),
                ("Primary Email", primary_email),
                ("Primary Phone", primary_phone),
                ("Secondary Contact", profile_row.get(sc.COL_PROFILE_SECONDARY_CONTACT)),
                ("Secondary Position", profile_row.get(sc.COL_PROFILE_SECONDARY_POSITION)),
                ("Secondary Email", profile_row.get(sc.COL_PROFILE_SECONDARY_EMAIL)),
                ("Secondary Phone", profile_row.get(sc.COL_PROFILE_SECONDARY_PHONE)),
                ("Address Line 1", profile_row.get(sc.COL_PROFILE_ADDR1)),
                ("Address Line 2", profile_row.get(sc.COL_PROFILE_ADDR2)),
                ("City", profile_row.get(sc.COL_PROFILE_CITY)),
                ("State/Province", profile_row.get(sc.COL_PROFILE_STATE)),
                ("Postal Code", profile_row.get(sc.COL_PROFILE_POSTAL)),
                ("Country", profile_row.get(sc.COL_PROFILE_COUNTRY)),
                ("Full Address", profile_row.get(sc.COL_PROFILE_FULL_ADDRESS)),
                ("Parent Client ID", profile_row.get(sc.COL_PROFILE_PARENT_ID)),
                ("Parent Client Name", profile_row.get(sc.COL_PROFILE_PARENT_NAME)),
                ("Website", profile_row.get(sc.COL_PROFILE_WEBSITE)),
                ("Tax ID", profile_row.get(sc.COL_PROFILE_TAX_ID)),
                ("Industry", profile_row.get(sc.COL_PROFILE_INDUSTRY)),
                ("Billing Email", profile_row.get(sc.COL_PROFILE_BILLING_EMAIL)),
                ("KYC Status", profile_row.get(sc.COL_PROFILE_KYC_STATUS)),
                ("Onboarding Status", profile_row.get(sc.COL_PROFILE_ONBOARDING_STATUS)),
                ("Retainer Required", profile_row.get(sc.COL_PROFILE_RETAINER_REQUIRED)),
                ("Retainer Amount", profile_row.get(sc.COL_PROFILE_RETAINER_AMOUNT)),
                ("Engagement Start Date", profile_row.get(sc.COL_PROFILE_ENGAGEMENT_START)),
                ("Date Client Added", profile_row.get(sc.COL_PROFILE_DATE_CLIENT_ADDED)),
                ("Birthday", profile_row.get(sc.COL_PROFILE_BIRTHDAY)),
                ("Referral From", profile_row.get(sc.COL_PROFILE_REFERRAL_FROM)),
                ("Conflict Notes", profile_row.get(sc.COL_PROFILE_CONFLICT_NOTES)),
                ("Client Notes", profile_row.get(sc.COL_PROFILE_NOTES) or client_row.get(sc.COL_CLIENT_NOTES)),
            ]
            _append_result(
                entity_type="client",
                entity_id=client_id,
                title=display_name or client_name or legal_name or client_id,
                subtitle=" | ".join(subtitle_bits[:3]),
                status=status,
                route_tile_index=0,
                route_node_id="A03",
                route_node_title="Client Profile 360",
                field_pairs=field_pairs,
                extra={
                    "clientId": client_id,
                    "clientName": client_name,
                    "legalName": legal_name,
                    "firstName": _clean_text(profile_row.get(sc.COL_PROFILE_FIRST_NAME)),
                    "middleName": _clean_text(profile_row.get(sc.COL_PROFILE_MIDDLE_NAME)),
                    "lastName": _clean_text(profile_row.get(sc.COL_PROFILE_LAST_NAME)),
                    "principalName": _clean_text(profile_row.get(sc.COL_PROFILE_PRINCIPAL_NAME)),
                    "primaryEmail": primary_email,
                    "primaryPhone": primary_phone,
                },
            )

        for profile_row in profile_rows:
            client_id = _clean_text(profile_row.get(sc.COL_PROFILE_CLIENT_ID))
            if client_id and client_id.lower() in seen_client_ids:
                continue
            display_name = _clean_text(profile_row.get(sc.COL_PROFILE_DISPLAY_NAME))
            legal_name = _clean_text(profile_row.get(sc.COL_PROFILE_LEGAL_NAME))
            status = "Active"
            subtitle_bits = [
                x
                for x in (
                    legal_name,
                    _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_EMAIL)),
                    status,
                )
                if x
            ]
            field_pairs = [
                ("Client ID", client_id),
                ("Display Name", display_name),
                ("Legal Name", legal_name),
                ("First Name", profile_row.get(sc.COL_PROFILE_FIRST_NAME)),
                ("Middle Name", profile_row.get(sc.COL_PROFILE_MIDDLE_NAME)),
                ("Last Name", profile_row.get(sc.COL_PROFILE_LAST_NAME)),
                ("Entity Type", profile_row.get(sc.COL_PROFILE_ENTITY_TYPE)),
                ("Primary Email", profile_row.get(sc.COL_PROFILE_PRIMARY_EMAIL)),
                ("Primary Phone", profile_row.get(sc.COL_PROFILE_PRIMARY_PHONE)),
                ("Secondary Contact", profile_row.get(sc.COL_PROFILE_SECONDARY_CONTACT)),
                ("Secondary Position", profile_row.get(sc.COL_PROFILE_SECONDARY_POSITION)),
                ("Secondary Email", profile_row.get(sc.COL_PROFILE_SECONDARY_EMAIL)),
                ("Secondary Phone", profile_row.get(sc.COL_PROFILE_SECONDARY_PHONE)),
                ("Full Address", profile_row.get(sc.COL_PROFILE_FULL_ADDRESS)),
                ("Parent Client Name", profile_row.get(sc.COL_PROFILE_PARENT_NAME)),
                ("Website", profile_row.get(sc.COL_PROFILE_WEBSITE)),
                ("Tax ID", profile_row.get(sc.COL_PROFILE_TAX_ID)),
                ("Industry", profile_row.get(sc.COL_PROFILE_INDUSTRY)),
                ("Billing Email", profile_row.get(sc.COL_PROFILE_BILLING_EMAIL)),
                ("KYC Status", profile_row.get(sc.COL_PROFILE_KYC_STATUS)),
                ("Onboarding Status", profile_row.get(sc.COL_PROFILE_ONBOARDING_STATUS)),
                ("Date Client Added", profile_row.get(sc.COL_PROFILE_DATE_CLIENT_ADDED)),
                ("Birthday", profile_row.get(sc.COL_PROFILE_BIRTHDAY)),
                ("Referral From", profile_row.get(sc.COL_PROFILE_REFERRAL_FROM)),
                ("Conflict Notes", profile_row.get(sc.COL_PROFILE_CONFLICT_NOTES)),
                ("Client Notes", profile_row.get(sc.COL_PROFILE_NOTES)),
            ]
            _append_result(
                entity_type="client",
                entity_id=client_id,
                title=display_name or legal_name or client_id,
                subtitle=" | ".join(subtitle_bits[:3]),
                status=status,
                route_tile_index=0,
                route_node_id="A03",
                route_node_title="Client Profile 360",
                field_pairs=field_pairs,
                extra={
                    "clientId": client_id,
                    "clientName": display_name or legal_name,
                    "legalName": legal_name,
                    "firstName": _clean_text(profile_row.get(sc.COL_PROFILE_FIRST_NAME)),
                    "middleName": _clean_text(profile_row.get(sc.COL_PROFILE_MIDDLE_NAME)),
                    "lastName": _clean_text(profile_row.get(sc.COL_PROFILE_LAST_NAME)),
                    "principalName": _clean_text(profile_row.get(sc.COL_PROFILE_PRINCIPAL_NAME)),
                    "primaryEmail": _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_EMAIL)),
                    "primaryPhone": _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_PHONE)),
                },
            )

        for parent_row in parent_rows:
            parent_id = _clean_text(parent_row.get(sc.COL_PARENT_ID))
            parent_name = _clean_text(parent_row.get(sc.COL_PARENT_NAME))
            parent_active = self.db._to_bool_int(parent_row.get(sc.COL_PARENT_ACTIVE), default=1)
            status = "Active" if parent_active else "Inactive"
            _append_result(
                entity_type="parent",
                entity_id=parent_id,
                title=parent_name or parent_id,
                subtitle=" | ".join(
                    x
                    for x in (
                        ("Parent ID: " + parent_id) if parent_id else "",
                        status,
                    )
                    if x
                ),
                status=status,
                route_tile_index=0,
                route_node_id="A05",
                route_node_title="Parent-Child Link Manager",
                field_pairs=[
                    ("Parent ID", parent_id),
                    ("Parent Name", parent_name),
                    ("Default Share %", parent_row.get(sc.COL_PARENT_DEF_SHARE)),
                    ("Default Rate", parent_row.get(sc.COL_PARENT_DEF_RATE)),
                    ("Notes", parent_row.get(sc.COL_PARENT_NOTES)),
                    ("Status", status),
                ],
                extra={"parentId": parent_id, "parentName": parent_name},
            )

        for matter_row in matter_rows:
            matter_id = _clean_text(matter_row.get(sc.COL_MATTER_ID))
            matter_number = _clean_text(matter_row.get(sc.COL_MATTER_NUMBER))
            matter_name = _clean_text(matter_row.get(sc.COL_MATTER_NAME))
            display_name = _clean_text(matter_row.get(sc.COL_MATTER_DISPLAY_NAME)) or matter_name
            status = _clean_text(matter_row.get(sc.COL_MATTER_STATUS)) or "Open"
            client_ref = _clean_text(matter_row.get(sc.COL_MATTER_CLIENT_ID))
            parent_ref = _clean_text(matter_row.get(sc.COL_MATTER_PARENT_ID))
            client_name = _clean_text(matter_row.get(sc.COL_MATTER_CLIENT_NAME)) or client_name_by_id.get(
                client_ref, client_ref
            )
            parent_name = _clean_text(matter_row.get(sc.COL_MATTER_PARENT_NAME)) or parent_name_by_id.get(
                parent_ref, parent_ref
            )
            matter_type = _clean_text(matter_row.get(sc.COL_MATTER_TYPE))
            practice_area = _clean_text(matter_row.get(sc.COL_MATTER_PRACTICE_AREA))
            billing_email = _clean_text(matter_row.get(sc.COL_MATTER_BILLING_EMAIL))
            responsible_lawyer = _clean_text(matter_row.get(sc.COL_MATTER_RESPONSIBLE_LAWYER))

            subtitle_bits = [x for x in (client_name, matter_type or practice_area, status) if x]
            _append_result(
                entity_type="matter",
                entity_id=matter_id,
                title=display_name or matter_name or matter_id,
                subtitle=" | ".join(subtitle_bits[:3]),
                status=status,
                route_tile_index=0,
                route_node_id="A11",
                route_node_title="Matter Profile 360",
                field_pairs=[
                    ("Matter ID", matter_id),
                    ("Matter Number", matter_number),
                    ("Matter Name", matter_name),
                    ("Display Name", display_name),
                    ("Client ID/Ref", client_ref),
                    ("Client Name", client_name),
                    ("Parent ID/Ref", parent_ref),
                    ("Parent Name", parent_name),
                    ("Matter Type", matter_type),
                    ("Practice Area", practice_area),
                    ("Responsible Lawyer", responsible_lawyer),
                    ("Billing Arrangement", matter_row.get(sc.COL_MATTER_BILLING_ARRANGEMENT)),
                    ("Billing Contact", matter_row.get(sc.COL_MATTER_BILLING_CONTACT)),
                    ("Billing Email", billing_email),
                    ("Default Rate", matter_row.get(sc.COL_MATTER_DEF_RATE)),
                    ("Default Share %", matter_row.get(sc.COL_MATTER_DEF_SHARE)),
                    ("Date Of Engagement", matter_row.get(sc.COL_MATTER_ENGAGEMENT_DATE)),
                    ("Date Opened", matter_row.get(sc.COL_MATTER_OPEN_DATE)),
                    ("Date Closed", matter_row.get(sc.COL_MATTER_CLOSE_DATE)),
                    ("Court File Number", matter_row.get(sc.COL_MATTER_COURT_FILE_NO)),
                    ("Opposing Party", matter_row.get(sc.COL_MATTER_OPPOSING_PARTY)),
                    ("Referral From", matter_row.get(sc.COL_MATTER_REFERRAL_FROM)),
                    ("Description", matter_row.get(sc.COL_MATTER_DESCRIPTION)),
                    ("Status", status),
                    ("Notes", matter_row.get(sc.COL_MATTER_NOTES)),
                ],
                extra={
                    "matterId": matter_id,
                    "matterNumber": matter_number,
                    "matterName": matter_name,
                    "displayName": display_name,
                    "clientId": client_ref,
                    "clientName": client_name,
                    "parentId": parent_ref,
                    "parentName": parent_name,
                    "matterType": matter_type,
                    "practiceArea": practice_area,
                    "responsibleLawyer": responsible_lawyer,
                    "billingEmail": billing_email,
                },
            )

        for time_row in time_rows:
            entry_id = _clean_text(time_row.get(sc.COL_TIME_ENTRY_ID))
            date_text = _clean_text(time_row.get(sc.COL_TIME_DATE))
            description = _clean_text(time_row.get(sc.COL_TIME_DESC))
            status = _clean_text(time_row.get(sc.COL_TIME_STATUS)) or "WIP"
            lock_audit = _clean_text(time_row.get(sc.COL_TIME_LOCK_AUDIT))
            client_ref = _clean_text(time_row.get(sc.COL_TIME_CLIENT_ID))
            matter_ref = _clean_text(time_row.get(sc.COL_TIME_MATTER_ID))
            parent_ref = _clean_text(time_row.get(sc.COL_TIME_PARENT_ID))

            client_name = client_name_by_id.get(client_ref, client_ref)
            matter_name = matter_name_by_id.get(matter_ref, matter_ref)
            parent_name = parent_name_by_id.get(parent_ref, parent_ref)

            desc_lc = _normalize_search_text(description)
            status_lc = _normalize_search_text(status)
            entity_type = "docket"
            route_tile_index = 1
            route_node_id = "B01"
            route_node_title = "Time Docket Entry"
            if "deadline" in desc_lc or "deadline" in status_lc:
                entity_type = "deadline"
                route_node_id = "B07"
                route_node_title = "Deadline Master Calendar"
            elif "tickler" in desc_lc or "reminder" in desc_lc:
                entity_type = "tickler"
                route_node_id = "B11"
                route_node_title = "Tickler Scheduler"
            elif "invoice" in desc_lc or status_lc.startswith("invoice"):
                entity_type = "invoice"
                route_tile_index = 2
                route_node_id = "C04"
                route_node_title = "Invoice Directory"

            title = " - ".join(x for x in (date_text, description) if x) or entry_id or "Docket Entry"
            subtitle_bits = [x for x in (client_name, matter_name, status) if x]
            _append_result(
                entity_type=entity_type,
                entity_id=entry_id,
                title=title,
                subtitle=" | ".join(subtitle_bits[:3]),
                status=status,
                route_tile_index=route_tile_index,
                route_node_id=route_node_id,
                route_node_title=route_node_title,
                field_pairs=[
                    ("Entry ID", entry_id),
                    ("Date", date_text),
                    ("Description", description),
                    ("Status", status),
                    ("Client ID/Ref", client_ref),
                    ("Client Name", client_name),
                    ("Matter ID/Ref", matter_ref),
                    ("Matter Name", matter_name),
                    ("Parent ID/Ref", parent_ref),
                    ("Parent Name", parent_name),
                    ("Hours", time_row.get(sc.COL_TIME_HOURS)),
                    ("Client Rate", time_row.get(sc.COL_TIME_RATE)),
                    ("Share %", time_row.get(sc.COL_TIME_SHARE_PCT)),
                    ("Gross to Client", time_row.get(sc.COL_TIME_GROSS)),
                    ("Amount to You", time_row.get(sc.COL_TIME_NET)),
                    ("HST", time_row.get(sc.COL_TIME_HST)),
                    ("Total", time_row.get(sc.COL_TIME_TOTAL)),
                    ("Lock Audit", lock_audit),
                    ("Created At", time_row.get(sc.COL_TIME_CREATED)),
                ],
                extra={
                    "entryId": entry_id,
                    "clientId": client_ref,
                    "matterId": matter_ref,
                    "parentId": parent_ref,
                },
            )

        for trademark_row in trademark_rows:
            payload = self.db._trademark_row_to_payload(trademark_row)
            trademark_id = _clean_text(payload.get("trademarkId"))
            title = _clean_text(payload.get("trademarkText")) or _clean_text(payload.get("applicationNumber")) or trademark_id
            if not title:
                continue

            jurisdiction = _clean_text(payload.get("jurisdiction"))
            jurisdiction_other = _clean_text(payload.get("jurisdictionOther"))
            jurisdiction_label = jurisdiction
            if jurisdiction.lower() == "other" and jurisdiction_other:
                jurisdiction_label = f"{jurisdiction} ({jurisdiction_other})"
            status = _clean_text(payload.get("status")) or "Open"

            subtitle_bits = [
                jurisdiction_label,
                _clean_text(payload.get("applicationNumber")) or _clean_text(payload.get("registrationNumber")),
                status,
            ]
            _append_result(
                entity_type="trademark",
                entity_id=trademark_id,
                title=title,
                subtitle=" | ".join(x for x in subtitle_bits if x),
                status=status,
                route_tile_index=1,
                route_node_id="B16",
                route_node_title="Trademark Filing",
                field_pairs=[
                    ("Trademark ID", payload.get("trademarkId")),
                    ("Trademark Text", payload.get("trademarkText")),
                    ("Jurisdiction", payload.get("jurisdiction")),
                    ("Jurisdiction Other", payload.get("jurisdictionOther")),
                    ("Client Name", payload.get("clientName")),
                    ("Matter Number", payload.get("matterNumber")),
                    ("Application Number", payload.get("applicationNumber")),
                    ("Registration Number", payload.get("registrationNumber")),
                    ("Current Status", payload.get("currentStatus")),
                    ("CIPO Status", payload.get("cipoStatus")),
                    ("USPTO Status", payload.get("usptoStatusIndicator")),
                    ("Mark Type", payload.get("markType")),
                    ("Nice Classes", payload.get("niceClasses")),
                    ("Goods Services", payload.get("goodsServices")),
                    ("Applicant/Owner", payload.get("applicantNameAddress"))
                    or payload.get("ownerNameAddress"),
                    ("Registry Link", payload.get("registryLink")),
                    ("Updated At", payload.get("updatedAt")),
                ],
                extra={
                    "trademarkId": payload.get("trademarkId"),
                    "trademarkText": payload.get("trademarkText"),
                    "jurisdiction": payload.get("jurisdiction"),
                    "applicationNumber": payload.get("applicationNumber"),
                    "registrationNumber": payload.get("registrationNumber"),
                    "registryLink": payload.get("registryLink"),
                },
            )

        for txn_row in txn_rows:
            txn_id = _clean_text(txn_row.get(sc.COL_TXN_ID))
            txn_date = _clean_text(txn_row.get(sc.COL_TXN_DATE))
            txn_type = _clean_text(txn_row.get(sc.COL_TXN_TYPE))
            txn_class = _clean_text(txn_row.get(sc.COL_TXN_CLASS))
            business_unit = _clean_text(txn_row.get(sc.COL_TXN_BUSINESS_UNIT))
            payee = _clean_text(txn_row.get(sc.COL_TXN_PAYEE))
            category_name = _clean_text(txn_row.get(sc.COL_TXN_CATEGORY_NAME))
            category_code = _clean_text(txn_row.get(sc.COL_TXN_CATEGORY_CODE))
            amount = _clean_text(txn_row.get(sc.COL_TXN_AMOUNT))
            currency = _clean_text(txn_row.get(sc.COL_TXN_CURRENCY)) or "CAD"
            status = _clean_text(txn_row.get(sc.COL_TXN_STATUS)) or "Pending"
            title_bits = [x for x in (txn_date, payee or category_name or txn_type, amount and f"{amount} {currency}") if x]
            subtitle_bits = [x for x in (business_unit, txn_class, txn_type, status) if x]
            _append_result(
                entity_type="transaction",
                entity_id=txn_id,
                title=" - ".join(title_bits) or txn_id or "Transaction",
                subtitle=" | ".join(subtitle_bits[:4]),
                status=status,
                route_tile_index=2,
                route_node_id="C11",
                route_node_title="Transactions Master",
                field_pairs=[
                    ("Transaction ID", txn_id),
                    ("Date", txn_date),
                    ("Class", txn_class),
                    ("Business Unit", business_unit),
                    ("Type", txn_type),
                    ("From Account", txn_row.get(sc.COL_TXN_FROM_ACCOUNT)),
                    ("To Account", txn_row.get(sc.COL_TXN_TO_ACCOUNT)),
                    ("Payee", payee),
                    ("Parent", txn_row.get(sc.COL_TXN_PARENT)),
                    ("Client Name", txn_row.get(sc.COL_TXN_CLIENT)),
                    ("Matter Name", txn_row.get(sc.COL_TXN_MATTER)),
                    ("Category Code", category_code),
                    ("Category Name", category_name),
                    ("Member", txn_row.get(sc.COL_TXN_MEMBER)),
                    ("Amount", amount),
                    ("Tax Amount", txn_row.get(sc.COL_TXN_TAX_AMOUNT)),
                    ("Tax Flag", txn_row.get(sc.COL_TXN_TAX_FLAG)),
                    ("Invoice Ref", txn_row.get(sc.COL_TXN_INVOICE_REF)),
                    ("Expense Details", txn_row.get(sc.COL_TXN_EXPENSE_DETAILS)),
                    ("Notes", txn_row.get(sc.COL_TXN_NOTES)),
                    ("Status", status),
                    ("Currency", currency),
                ],
                extra={
                    "transactionId": txn_id,
                    "txnDate": txn_date,
                    "businessUnit": business_unit,
                    "type": txn_type,
                    "payee": payee,
                    "categoryCode": category_code,
                    "categoryName": category_name,
                    "amount": amount,
                    "currency": currency,
                },
            )

        for account_row in account_rows:
            account_code = _clean_text(account_row.get(sc.COL_TXN_ACCOUNT_CODE))
            account_name = _clean_text(account_row.get(sc.COL_TXN_ACCOUNT_NAME))
            account_kind = _clean_text(account_row.get(sc.COL_TXN_ACCOUNT_KIND))
            owner = _clean_text(account_row.get(sc.COL_TXN_ACCOUNT_OWNER))
            active = self.db._to_bool_int(account_row.get(sc.COL_TXN_ACCOUNT_ACTIVE), default=1)
            status = "Active" if active else "Inactive"
            _append_result(
                entity_type="account",
                entity_id=account_code,
                title=account_name or account_code,
                subtitle=" | ".join(x for x in (account_code, account_kind, owner, status) if x),
                status=status,
                route_tile_index=2,
                route_node_id="C11",
                route_node_title="Transactions Master",
                field_pairs=[
                    ("Account Code", account_code),
                    ("Account Name", account_name),
                    ("Account Kind", account_kind),
                    ("Owner", owner),
                    ("Aliases", account_row.get(sc.COL_TXN_ACCOUNT_ALIASES)),
                    ("Status", status),
                ],
                extra={
                    "accountCode": account_code,
                    "accountName": account_name,
                    "accountKind": account_kind,
                    "owner": owner,
                },
            )

        for category_row in category_rows:
            category_code = _clean_text(category_row.get(sc.COL_TXN_CATEGORY_LKP_CODE))
            category_name = _clean_text(category_row.get(sc.COL_TXN_CATEGORY_LKP_NAME))
            category_type = _clean_text(category_row.get(sc.COL_TXN_CATEGORY_LKP_TYPE))
            class_scope = _clean_text(category_row.get(sc.COL_TXN_CATEGORY_LKP_CLASS_SCOPE))
            active = self.db._to_bool_int(category_row.get(sc.COL_TXN_CATEGORY_LKP_ACTIVE), default=1)
            status = "Active" if active else "Inactive"
            _append_result(
                entity_type="category",
                entity_id=category_code,
                title=category_name or category_code,
                subtitle=" | ".join(x for x in (category_code, category_type, class_scope, status) if x),
                status=status,
                route_tile_index=2,
                route_node_id="C11",
                route_node_title="Transactions Master",
                field_pairs=[
                    ("Category Code", category_code),
                    ("Category Name", category_name),
                    ("Type", category_type),
                    ("Class Scope", class_scope),
                    ("Tax Flag", category_row.get(sc.COL_TXN_CATEGORY_LKP_TAX_FLAG_DEFAULT)),
                    ("Notes", category_row.get(sc.COL_TXN_CATEGORY_LKP_NOTES)),
                    ("Status", status),
                ],
                extra={
                    "categoryCode": category_code,
                    "categoryName": category_name,
                    "type": category_type,
                    "classScope": class_scope,
                },
            )

        for unit_row in business_unit_rows:
            business_unit = _clean_text(unit_row.get(sc.COL_TXN_BUSINESS_UNIT_NAME))
            owner = _clean_text(unit_row.get(sc.COL_TXN_BUSINESS_UNIT_OWNER))
            active = self.db._to_bool_int(unit_row.get(sc.COL_TXN_BUSINESS_UNIT_ACTIVE), default=1)
            status = "Active" if active else "Inactive"
            _append_result(
                entity_type="business_unit",
                entity_id=business_unit,
                title=business_unit,
                subtitle=" | ".join(x for x in (owner, status) if x),
                status=status,
                route_tile_index=2,
                route_node_id="C11",
                route_node_title="Transactions Master",
                field_pairs=[
                    ("Business Unit", business_unit),
                    ("Owner", owner),
                    ("Status", status),
                ],
                extra={"businessUnit": business_unit, "owner": owner},
            )

        for payee_row in payee_rows:
            payee_name = _clean_text(payee_row.get(sc.COL_TXN_PAYEE_NAME))
            default_category = _clean_text(payee_row.get(sc.COL_TXN_PAYEE_DEFAULT_CATEGORY_CODE))
            active = self.db._to_bool_int(payee_row.get(sc.COL_TXN_PAYEE_ACTIVE), default=1)
            status = "Active" if active else "Inactive"
            _append_result(
                entity_type="payee",
                entity_id=payee_name,
                title=payee_name,
                subtitle=" | ".join(x for x in (default_category, status) if x),
                status=status,
                route_tile_index=2,
                route_node_id="C11",
                route_node_title="Transactions Master",
                field_pairs=[
                    ("Payee Name", payee_name),
                    ("Default Category Code", default_category),
                    ("Status", status),
                ],
                extra={"payeeName": payee_name, "defaultCategoryCode": default_category},
            )

        results.sort(
            key=lambda row: (
                -int(row.get("_score", 0)),
                _normalize_search_text(row.get("entityType", "")),
                _normalize_search_text(row.get("title", "")),
                _normalize_search_text(row.get("entityId", "")),
            )
        )

        total_count = len(results)
        if total_count > max_results:
            results = results[:max_results]

        for row in results:
            if "_score" in row:
                del row["_score"]

        return {
            "ok": True,
            "query": raw_query,
            "mode": search_mode,
            "results": results,
            "facets": facets,
            "total": total_count,
            "returnedCount": len(results),
        }

    def home_dashboard_summary(self) -> Dict[str, Any]:
        def _empty(ok: bool = True) -> Dict[str, Any]:
            return {
                "ok": bool(ok),
                "asOfDate": datetime.now().strftime("%Y-%m-%d"),
                "deadlinesCount": 0,
                "unbilledDraftCount": 0,
                "clientMeetingCount": 0,
                "queueCount": 0,
                "activeClientCount": 0,
                # matter summary added per UX requirement
                "activeMatterCount": 0,
            }

        try:
            client_rows = self.db._read_table_rows(TBL_CLIENTS)
            matter_rows = self.db._read_table_rows(TBL_MATTERS)
            time_rows = self.db._read_table_rows(TBL_TIME)
            deadline_rows = self.db._load_deadlines()
        except Exception:
            try:
                self.db.ensure_schema()
                client_rows = self.db._read_table_rows(TBL_CLIENTS)
                matter_rows = self.db._read_table_rows(TBL_MATTERS)
                time_rows = self.db._read_table_rows(TBL_TIME)
                deadline_rows = self.db._load_deadlines()
            except Exception:
                return _empty(ok=False)

        today_iso = datetime.now().strftime("%Y-%m-%d")

        active_client_count = 0
        for row in client_rows:
            status = _clean_text(row.get(sc.COL_CLIENT_STATUS)).lower()
            active_flag = self.db._to_bool_int(row.get(sc.COL_CLIENT_ACTIVE), default=1)
            if active_flag == 1 and status not in ("inactive", "closed", "archived"):
                active_client_count += 1

        # matter activity count uses canonical helper
        active_matter_count = 0
        for row in matter_rows:
            if self._is_matter_row_active(row):
                active_matter_count += 1

        deadlines_count = sum(1 for d in deadline_rows if not d.get("completed", False))
        unbilled_draft_count = 0
        client_meeting_count = 0
        for row in time_rows:
            description = _clean_text(row.get(sc.COL_TIME_DESC)).lower()
            status = _clean_text(row.get(sc.COL_TIME_STATUS)).lower()
            if status in ("", "wip", "draft", "unbilled", "pending", "open", "ready for billing"):
                unbilled_draft_count += 1

            date_text = _clean_text(row.get(sc.COL_TIME_DATE))
            if date_text == today_iso and (
                "meeting" in description
                or "conference" in description
                or "call" in description
            ):
                client_meeting_count += 1

        queue_count = deadlines_count + unbilled_draft_count + client_meeting_count
        return {
            "ok": True,
            "asOfDate": today_iso,
            "deadlinesCount": int(deadlines_count),
            "unbilledDraftCount": int(unbilled_draft_count),
            "clientMeetingCount": int(client_meeting_count),
            "queueCount": int(queue_count),
            "activeClientCount": int(active_client_count),
            "activeMatterCount": int(active_matter_count),
        }

    def save_client_profile(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.db.ensure_schema()
        normalized = self.db._normalize_client_profile_payload(payload)

        parent_row: Optional[Dict[str, Any]] = None
        if normalized["parentClientName"]:
            parent_row = self.db._get_or_create_parent(normalized["parentClientName"])

        requested_client_id = _clean_text(normalized.get("clientId"))
        client_row: Optional[Dict[str, Any]] = None
        if requested_client_id:
            client_rows = [self.db._canonicalize_client_row(r) for r in self.db._read_table_rows(TBL_CLIENTS)]
            for row in client_rows:
                if _clean_text(row.get(sc.COL_CLIENT_ID)).lower() == requested_client_id.lower():
                    client_row = row
                    break
        if client_row is None:
            client_row = self.db._get_or_create_client(normalized["clientName"])

        client_id = requested_client_id or _clean_text(client_row.get(sc.COL_CLIENT_ID))
        if not client_id:
            client_id = self.db._new_id("C")
        parent_id = _clean_text(parent_row.get(sc.COL_PARENT_ID)) if parent_row else normalized["parentClientId"]

        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        existing_profile = self.db._find_client_profile(client_id)
        created_at = _clean_text(existing_profile.get(sc.COL_PROFILE_CREATED)) if existing_profile else now_stamp
        if not created_at:
            created_at = now_stamp

        client_master_row = {
            sc.COL_CLIENT_ID: client_id,
            sc.COL_CLIENT_NAME: normalized["clientName"],
            sc.COL_CLIENT_EMAIL: normalized["primaryEmail"],
            sc.COL_CLIENT_PHONE: normalized["primaryPhone"],
            sc.COL_CLIENT_STATUS: normalized["status"],
            sc.COL_CLIENT_ACTIVE: normalized["active"],
            sc.COL_CLIENT_NOTES: normalized["notes"],
        }
        self.db._upsert_row_by_key(TBL_CLIENTS, sc.COL_CLIENT_ID, client_id, client_master_row)

        full_address = normalized["fullAddress"] or self.db._format_full_address(
            line1=normalized["addressLine1"],
            line2=normalized["addressLine2"],
            city=normalized["city"],
            state_province=normalized["stateProvince"],
            postal_code=normalized["postalCode"],
            country=normalized["country"],
        )

        profile_row = {
            sc.COL_PROFILE_CLIENT_ID: client_id,
            sc.COL_PROFILE_LEGAL_NAME: normalized["legalName"],
            sc.COL_PROFILE_DISPLAY_NAME: normalized["displayName"] or normalized["clientName"],
            sc.COL_PROFILE_FIRST_NAME: normalized["firstName"],
            sc.COL_PROFILE_MIDDLE_NAME: normalized["middleName"],
            sc.COL_PROFILE_LAST_NAME: normalized["lastName"],
            sc.COL_PROFILE_ENTITY_TYPE: normalized["entityType"],
            sc.COL_PROFILE_PRINCIPAL_NAME: normalized["principalName"],
            sc.COL_PROFILE_PRINCIPAL_POSITION: normalized["principalPosition"],
            sc.COL_PROFILE_PRIMARY_EMAIL: normalized["primaryEmail"],
            sc.COL_PROFILE_PRIMARY_PHONE: normalized["primaryPhone"],
            sc.COL_PROFILE_SECONDARY_CONTACT: normalized["secondaryContactName"],
            sc.COL_PROFILE_SECONDARY_POSITION: normalized["secondaryContactPosition"],
            sc.COL_PROFILE_SECONDARY_EMAIL: normalized["secondaryContactEmail"],
            sc.COL_PROFILE_SECONDARY_PHONE: normalized["secondaryContactPhone"],
            sc.COL_PROFILE_ADDR1: normalized["addressLine1"],
            sc.COL_PROFILE_ADDR2: normalized["addressLine2"],
            sc.COL_PROFILE_CITY: normalized["city"],
            sc.COL_PROFILE_STATE: normalized["stateProvince"],
            sc.COL_PROFILE_POSTAL: normalized["postalCode"],
            sc.COL_PROFILE_COUNTRY: normalized["country"],
            sc.COL_PROFILE_FULL_ADDRESS: full_address,
            sc.COL_PROFILE_PARENT_ID: parent_id,
            sc.COL_PROFILE_PARENT_NAME: normalized["parentClientName"],
            sc.COL_PROFILE_WEBSITE: normalized["website"],
            sc.COL_PROFILE_TAX_ID: normalized["taxId"],
            sc.COL_PROFILE_INDUSTRY: normalized["industry"],
            sc.COL_PROFILE_BILLING_EMAIL: normalized["billingEmail"],
            sc.COL_PROFILE_KYC_STATUS: normalized["kycStatus"],
            sc.COL_PROFILE_ONBOARDING_STATUS: normalized["onboardingStatus"],
            sc.COL_PROFILE_RETAINER_REQUIRED: normalized["retainerRequired"],
            sc.COL_PROFILE_RETAINER_AMOUNT: round(float(normalized["retainerAmount"]), 2),
            sc.COL_PROFILE_ENGAGEMENT_START: normalized["engagementStartDate"],
            sc.COL_PROFILE_DATE_CLIENT_ADDED: normalized["dateClientAdded"],
            sc.COL_PROFILE_BIRTHDAY: normalized["birthday"],
            sc.COL_PROFILE_REFERRAL_FROM: normalized["referralFrom"],
            sc.COL_PROFILE_CONFLICT_NOTES: normalized["conflictNotes"],
            sc.COL_PROFILE_NOTES: normalized["notes"],
            sc.COL_PROFILE_CREATED: created_at,
            sc.COL_PROFILE_UPDATED: now_stamp,
        }

        self.db._upsert_row_by_key(TBL_CLIENT_PROFILES, sc.COL_PROFILE_CLIENT_ID, client_id, profile_row)
        persisted = self.db._find_client_profile(client_id)
        verified = self.db._compare_client_profile_rows_loose(profile_row, persisted)

        return {
            "ok": bool(verified),
            "verifiedExact": bool(verified),
            "clientId": client_id,
            "savedRow": profile_row,
            "message": "" if verified else "Client profile verification failed.",
        }
    def delete_matter_profile(self, matter_id: str) -> Dict[str, Any]:
        return self.db.delete_matter_profile(matter_id)

    def save_matter_profile(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.db.ensure_schema()
        normalized = self.db._normalize_matter_profile_payload(payload)

        client_id = normalized["clientId"]
        client_name = normalized["clientName"]
        if client_name:
            client_row = self.db._get_or_create_client(client_name)
            client_id = _clean_text(client_row.get(sc.COL_CLIENT_ID))
            client_name = _clean_text(client_row.get(sc.COL_CLIENT_NAME))
        elif client_id:
            client_rows = [self.db._canonicalize_client_row(r) for r in self.db._read_table_rows(TBL_CLIENTS)]
            for row in client_rows:
                if _clean_text(row.get(sc.COL_CLIENT_ID)).lower() == client_id.lower():
                    client_name = _clean_text(row.get(sc.COL_CLIENT_NAME))
                    break

        if not client_id:
            raise ValueError("Client is required.")

        parent_id = normalized["parentId"]
        parent_name = normalized["parentName"]
        if parent_name:
            parent_row = self.db._get_or_create_parent(parent_name)
            parent_id = _clean_text(parent_row.get(sc.COL_PARENT_ID))
            parent_name = _clean_text(parent_row.get(sc.COL_PARENT_NAME))
        elif parent_id:
            parent_rows = [self.db._canonicalize_parent_row(r) for r in self.db._read_table_rows(TBL_PARENTS)]
            for row in parent_rows:
                if _clean_text(row.get(sc.COL_PARENT_ID)).lower() == parent_id.lower():
                    parent_name = _clean_text(row.get(sc.COL_PARENT_NAME))
                    break

        existing_row: Optional[Dict[str, Any]] = None
        matter_id = normalized["matterId"]
        if matter_id:
            existing_row = self.db._find_matter_row(matter_id)
        if existing_row is None:
            matter_rows = [self.db._canonicalize_matter_row(r) for r in self.db._read_table_rows(TBL_MATTERS)]
            target_name_lc = normalized["matterName"].lower()
            for row in matter_rows:
                if _clean_text(row.get(sc.COL_MATTER_NAME)).lower() != target_name_lc:
                    continue
                row_client_id = _clean_text(row.get(sc.COL_MATTER_CLIENT_ID))
                if client_id and row_client_id and row_client_id.lower() != client_id.lower():
                    continue
                existing_row = row
                break
        if existing_row is not None:
            matter_id = _clean_text(existing_row.get(sc.COL_MATTER_ID))
        if not matter_id:
            matter_id = self.db._new_id("M")

        # Resolve entity_type from the client profile so individual-vs-company
        # naming rules work correctly.
        entity_type = ""
        if client_id:
            profile_rows = self.db._read_table_rows(TBL_CLIENT_PROFILES)
            for pr in profile_rows:
                if _clean_text(pr.get(sc.COL_PROFILE_CLIENT_ID, "")).lower() == client_id.lower():
                    entity_type = _clean_text(pr.get(sc.COL_PROFILE_ENTITY_TYPE, ""))
                    break

        matter_number = _clean_text(normalized.get("matterNumber"))
        if not matter_number and existing_row is not None:
            matter_number = _clean_text(existing_row.get(sc.COL_MATTER_NUMBER))
        if not matter_number:
            matter_number = self._build_matter_number(
                client_name=client_name,
                matter_type=normalized["matterType"],
                date_opened=normalized["dateOpened"],
                existing_matter_id=matter_id,
                parent_name=parent_name,
                entity_type=entity_type,
            )
        matter_number = _clean_text(matter_number).upper()
        if self.db._matter_number_in_use(matter_number, ignore_matter_id=matter_id):
            if _clean_text(normalized.get("matterNumber")):
                raise ValueError(f"Matter number already exists: {matter_number}")
            matter_number = self._build_matter_number(
                client_name=client_name,
                matter_type=normalized["matterType"],
                date_opened=normalized["dateOpened"],
                existing_matter_id=matter_id,
                parent_name=parent_name,
                entity_type=entity_type,
            )

        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        created_at = _clean_text((existing_row or {}).get(sc.COL_MATTER_CREATED)) or now_stamp

        matter_row = {
            sc.COL_MATTER_ID: matter_id,
            sc.COL_MATTER_NUMBER: matter_number,
            sc.COL_MATTER_NAME: normalized["matterName"],
            sc.COL_MATTER_DISPLAY_NAME: normalized["displayName"] or normalized["matterName"],
            sc.COL_MATTER_CLIENT_ID: client_id,
            sc.COL_MATTER_CLIENT_NAME: client_name,
            sc.COL_MATTER_PARENT_ID: parent_id,
            sc.COL_MATTER_PARENT_NAME: parent_name,
            sc.COL_MATTER_TYPE: normalized["matterType"],
            sc.COL_MATTER_PRACTICE_AREA: normalized["practiceArea"],
            sc.COL_MATTER_STATUS: normalized["status"],
            sc.COL_MATTER_RESPONSIBLE_LAWYER: normalized["responsibleLawyer"],
            sc.COL_MATTER_BILLING_ARRANGEMENT: normalized["billingArrangement"],
            sc.COL_MATTER_BILLING_CONTACT: normalized["billingContact"],
            sc.COL_MATTER_BILLING_EMAIL: normalized["billingEmail"],
            sc.COL_MATTER_DEF_RATE: round(float(normalized["defaultRate"]), 2),
            sc.COL_MATTER_DEF_SHARE: round(float(normalized["defaultSharePct"]), 2),
            sc.COL_MATTER_ENGAGEMENT_DATE: normalized["dateOfEngagement"],
            sc.COL_MATTER_OPEN_DATE: normalized["dateOpened"],
            sc.COL_MATTER_CLOSE_DATE: normalized["dateClosed"],
            sc.COL_MATTER_COURT_FILE_NO: normalized["courtFileNumber"],
            sc.COL_MATTER_OPPOSING_PARTY: normalized["opposingParty"],
            sc.COL_MATTER_REFERRAL_FROM: normalized["referralFrom"],
            sc.COL_MATTER_DESCRIPTION: normalized["description"],
            sc.COL_MATTER_NOTES: normalized["notes"],
            sc.COL_MATTER_CREATED: created_at,
            sc.COL_MATTER_UPDATED: now_stamp,
        }

        self.db._upsert_row_by_key(TBL_MATTERS, sc.COL_MATTER_ID, matter_id, matter_row)
        persisted = self.db._find_matter_row(matter_id)
        verified = self.db._compare_matter_profile_rows_loose(matter_row, persisted)

        return {
            "ok": bool(verified),
            "verifiedExact": bool(verified),
            "matterId": matter_id,
            "matterNumber": matter_number,
            "savedRow": matter_row,
            "message": "" if verified else "Matter profile verification failed.",
        }

    def run_conflict_check(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        data = dict(payload or {})
        client_key = _clean_text(data.get("clientKey") or data.get("client") or data.get("clientName"))
        matter_key = _clean_text(data.get("matterKey") or data.get("matter") or data.get("matterName"))
        subject_text = _clean_text(data.get("subjectText") or data.get("query"))
        extra_text = _clean_text(data.get("extraTerms"))

        client_profile: Dict[str, Any] = {}
        matter_profile: Dict[str, Any] = {}
        if client_key:
            loaded = self.get_client_profile(client_key)
            if bool(loaded.get("ok")):
                client_profile = dict(loaded.get("client") or {})
        if matter_key:
            loaded = self.get_matter_profile(matter_key)
            if bool(loaded.get("ok")):
                matter_profile = dict(loaded.get("matter") or {})

        term_candidates = [
            subject_text,
            client_key,
            matter_key,
            _clean_text(client_profile.get("displayName")),
            _clean_text(client_profile.get("legalName")),
            _clean_text(client_profile.get("primaryEmail")),
            _clean_text(client_profile.get("primaryPhone")),
            _clean_text(client_profile.get("secondaryContactEmail")),
            _clean_text(client_profile.get("secondaryContactPhone")),
            _clean_text(client_profile.get("taxId")),
            _clean_text(client_profile.get("parentClientName")),
            _clean_text(matter_profile.get("displayName")),
            _clean_text(matter_profile.get("matterName")),
            _clean_text(matter_profile.get("matterNumber")),
            _clean_text(matter_profile.get("opposingParty")),
            _clean_text(matter_profile.get("clientName")),
        ]
        if extra_text:
            term_candidates.extend(_search_terms(extra_text))

        terms: List[str] = []
        seen_terms = set()
        for raw in term_candidates:
            text = _clean_text(raw)
            if not text:
                continue
            key = text.lower()
            if key in seen_terms:
                continue
            seen_terms.add(key)
            terms.append(text)
            if len(terms) >= 12:
                break

        if not terms:
            return {
                "ok": False,
                "message": "Enter at least one client, matter, or subject term.",
                "riskLevel": "none",
                "termsUsed": [],
                "matches": [],
                "totalMatches": 0,
                "summary": {"client": 0, "matter": 0, "parent": 0},
                "checkedAtUtc": datetime.now(UTC).replace(tzinfo=None).isoformat() + "Z",
            }

        selected_client_id = _clean_text(client_profile.get("clientId")).lower()
        selected_matter_id = _clean_text(matter_profile.get("matterId")).lower()
        aggregate: Dict[str, Dict[str, Any]] = {}
        summary = {"client": 0, "matter": 0, "parent": 0}

        for term in terms:
            search = self.search_global_entities(term, mode="any", limit=250)
            for row in search.get("results", []):
                entity_type = _clean_text(row.get("entityType")).lower()
                if entity_type not in ("client", "matter", "parent"):
                    continue
                entity_id = _clean_text(row.get("entityId"))
                title = _clean_text(row.get("title")) or entity_id
                if not title:
                    continue
                key = f"{entity_type}|{entity_id.lower()}|{title.lower()}"
                existing = aggregate.get(key)
                if existing is None:
                    existing = {
                        "entityType": entity_type,
                        "entityId": entity_id,
                        "title": title,
                        "subtitle": _clean_text(row.get("subtitle")),
                        "status": _clean_text(row.get("status")),
                        "routeTileIndex": int(row.get("routeTileIndex", 0)),
                        "routeNodeId": _clean_text(row.get("routeNodeId")),
                        "routeNodeTitle": _clean_text(row.get("routeNodeTitle")),
                        "matchedFields": list(row.get("matchedFields") or []),
                        "_hitTerms": [],
                        "_hits": 0,
                    }
                    aggregate[key] = existing

                term_lc = term.lower()
                if term_lc not in existing["_hitTerms"]:
                    existing["_hitTerms"].append(term_lc)
                    existing["_hits"] += 1

        matches: List[Dict[str, Any]] = []
        for row in aggregate.values():
            entity_type = _clean_text(row.get("entityType")).lower()
            entity_id_lc = _clean_text(row.get("entityId")).lower()
            if entity_type == "client" and selected_client_id and entity_id_lc == selected_client_id:
                continue
            if entity_type == "matter" and selected_matter_id and entity_id_lc == selected_matter_id:
                continue

            score = int(row.get("_hits", 0)) * 10
            if entity_type == "client":
                score += 4
            elif entity_type == "matter":
                score += 2
            else:
                score += 1

            clean_row = {
                "entityType": entity_type,
                "entityId": _clean_text(row.get("entityId")),
                "title": _clean_text(row.get("title")),
                "subtitle": _clean_text(row.get("subtitle")),
                "status": _clean_text(row.get("status")),
                "routeTileIndex": int(row.get("routeTileIndex", 0)),
                "routeNodeId": _clean_text(row.get("routeNodeId")),
                "routeNodeTitle": _clean_text(row.get("routeNodeTitle")),
                "matchedFields": list(row.get("matchedFields") or []),
                "matchedTerms": list(row.get("_hitTerms") or []),
                "score": score,
            }
            matches.append(clean_row)
            if entity_type in summary:
                summary[entity_type] += 1

        matches.sort(
            key=lambda row: (
                -int(row.get("score", 0)),
                _normalize_search_text(row.get("entityType")),
                _normalize_search_text(row.get("title")),
                _normalize_search_text(row.get("entityId")),
            )
        )
        for row in matches:
            row.pop("score", None)

        top_hits = max((len(row.get("matchedTerms") or []) for row in matches), default=0)
        total = len(matches)
        if total <= 0:
            risk_level = "none"
        elif top_hits >= 3 or total >= 8:
            risk_level = "high"
        elif top_hits >= 2 or total >= 3:
            risk_level = "medium"
        else:
            risk_level = "low"

        return {
            "ok": True,
            "message": f"{total} potential conflict match(es) found.",
            "riskLevel": risk_level,
            "termsUsed": terms,
            "matches": matches[:120],
            "totalMatches": total,
            "summary": summary,
            "checkedAtUtc": datetime.now(UTC).replace(tzinfo=None).isoformat() + "Z",
        }

    def reassign_matter(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.db.ensure_schema()
        data = dict(payload or {})
        matter_key = _clean_text(data.get("matterKey") or data.get("matterId") or data.get("matterName"))
        from_lawyer = _clean_text(data.get("fromLawyer"))
        to_lawyer = _clean_text(
            data.get("toLawyer")
            or data.get("newResponsibleLawyer")
            or data.get("responsibleLawyer")
        )
        reason = _clean_text(data.get("reason"))
        actor = _clean_text(data.get("actor"))

        if not matter_key:
            return {"ok": False, "message": "Matter key is required.", "changed": False}
        if not to_lawyer:
            return {"ok": False, "message": "New responsible lawyer is required.", "changed": False}

        matter_row = self.db._find_matter_row(matter_key)
        if not matter_row:
            return {"ok": False, "message": f"Matter not found: {matter_key}", "changed": False}

        matter_id = _clean_text(matter_row.get(sc.COL_MATTER_ID))
        matter_name = _clean_text(matter_row.get(sc.COL_MATTER_DISPLAY_NAME)) or _clean_text(
            matter_row.get(sc.COL_MATTER_NAME)
        )
        current_lawyer = _clean_text(matter_row.get(sc.COL_MATTER_RESPONSIBLE_LAWYER))

        if from_lawyer and current_lawyer and from_lawyer.lower() != current_lawyer.lower():
            return {
                "ok": False,
                "message": f"Current lawyer mismatch (expected '{from_lawyer}', found '{current_lawyer}').",
                "changed": False,
                "matterId": matter_id,
            }

        if current_lawyer.lower() == to_lawyer.lower():
            return {
                "ok": True,
                "changed": False,
                "matterId": matter_id,
                "matterName": matter_name,
                "fromLawyer": current_lawyer,
                "toLawyer": to_lawyer,
                "message": "Matter already assigned to that lawyer.",
                "updatedAtUtc": datetime.now(UTC).replace(tzinfo=None).isoformat() + "Z",
            }

        rows = [self.db._canonicalize_matter_row(r) for r in self.db._read_table_rows(TBL_MATTERS)]
        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        audit_note = (
            f"[Reassigned {now_stamp}] {current_lawyer or '[unassigned]'} -> {to_lawyer}"
            + (f"; Reason: {reason}" if reason else "")
            + (f"; Actor: {actor}" if actor else "")
        )

        changed_rows = 0
        for row in rows:
            row_id = _clean_text(row.get(sc.COL_MATTER_ID))
            if row_id.lower() != matter_id.lower():
                continue
            row[sc.COL_MATTER_RESPONSIBLE_LAWYER] = to_lawyer
            row[sc.COL_MATTER_UPDATED] = now_stamp
            row[sc.COL_MATTER_NOTES] = self.db._append_note_line(row.get(sc.COL_MATTER_NOTES), audit_note)
            changed_rows += 1
            break

        if changed_rows <= 0:
            return {"ok": False, "message": "Matter reassignment failed to apply.", "changed": False}

        self.db._replace_table_rows(TBL_MATTERS, rows)
        persisted = self.db._find_matter_row(matter_id)
        persisted_lawyer = _clean_text((persisted or {}).get(sc.COL_MATTER_RESPONSIBLE_LAWYER))
        verified = persisted_lawyer.lower() == to_lawyer.lower()

        return {
            "ok": bool(verified),
            "changed": True,
            "verifiedExact": bool(verified),
            "matterId": matter_id,
            "matterName": matter_name,
            "fromLawyer": current_lawyer,
            "toLawyer": to_lawyer,
            "message": "" if verified else "Matter reassignment verification failed.",
            "updatedAtUtc": datetime.now(UTC).replace(tzinfo=None).isoformat() + "Z",
            "auditNote": audit_note,
        }

    def merge_duplicate_entities(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        data = dict(payload or {})
        merge_type = _clean_text(data.get("mergeType") or "client").lower()
        if merge_type in ("matter", "matters"):
            return self.db._merge_duplicate_matters(data)
        return self.db._merge_duplicate_clients(data)

    def _canonicalize_parent_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        parent_id = _clean_text(self.db._value_with_alias(TBL_PARENTS.table, row, sc.COL_PARENT_ID))
        if not parent_id:
            parent_id = self.db._new_id("P")

        parent_name = _clean_text(self.db._value_with_alias(TBL_PARENTS.table, row, sc.COL_PARENT_NAME))
        share_pct = normalize_pct(self.db._value_with_alias(TBL_PARENTS.table, row, sc.COL_PARENT_DEF_SHARE), default_pct=100.0)
        default_rate = float(self.db._parse_float(self.db._value_with_alias(TBL_PARENTS.table, row, sc.COL_PARENT_DEF_RATE)) or 0.0)
        active = self.db._to_bool_int(self.db._value_with_alias(TBL_PARENTS.table, row, sc.COL_PARENT_ACTIVE), default=1)
        notes = _clean_text(self.db._value_with_alias(TBL_PARENTS.table, row, sc.COL_PARENT_NOTES))

        return {
            sc.COL_PARENT_ID: parent_id,
            sc.COL_PARENT_NAME: parent_name,
            sc.COL_PARENT_DEF_SHARE: round(share_pct, 2),
            sc.COL_PARENT_DEF_RATE: round(default_rate, 2),
            sc.COL_PARENT_ACTIVE: active,
            sc.COL_PARENT_NOTES: notes,
        }

    def _canonicalize_client_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        client_id = _clean_text(self.db._value_with_alias(TBL_CLIENTS.table, row, sc.COL_CLIENT_ID))
        if not client_id:
            client_id = self.db._new_id("C")

        status = _clean_text(self.db._value_with_alias(TBL_CLIENTS.table, row, sc.COL_CLIENT_STATUS))
        if not status:
            status = "Active"

        return {
            sc.COL_CLIENT_ID: client_id,
            sc.COL_CLIENT_NAME: _clean_text(self.db._value_with_alias(TBL_CLIENTS.table, row, sc.COL_CLIENT_NAME)),
            sc.COL_CLIENT_EMAIL: _clean_text(self.db._value_with_alias(TBL_CLIENTS.table, row, sc.COL_CLIENT_EMAIL)),
            sc.COL_CLIENT_PHONE: _clean_text(self.db._value_with_alias(TBL_CLIENTS.table, row, sc.COL_CLIENT_PHONE)),
            sc.COL_CLIENT_STATUS: status,
            sc.COL_CLIENT_ACTIVE: self.db._to_bool_int(self.db._value_with_alias(TBL_CLIENTS.table, row, sc.COL_CLIENT_ACTIVE), default=1),
            sc.COL_CLIENT_NOTES: _clean_text(self.db._value_with_alias(TBL_CLIENTS.table, row, sc.COL_CLIENT_NOTES)),
        }

    def _canonicalize_client_profile_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        client_id = _clean_text(
            self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_CLIENT_ID)
        )
        if not client_id:
            client_id = self.db._new_id("C")

        display_name = _clean_text(
            self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_DISPLAY_NAME)
        )
        legal_name = _clean_text(
            self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_LEGAL_NAME)
        )
        if not display_name:
            display_name = legal_name

        address_line1 = _clean_text(
            self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_ADDR1)
        )
        address_line2 = _clean_text(
            self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_ADDR2)
        )
        city = _clean_text(self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_CITY))
        state_province = _clean_text(
            self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_STATE)
        )
        postal_code = _clean_text(
            self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_POSTAL)
        )
        country = _clean_text(
            self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_COUNTRY)
        )
        full_address = _clean_text(
            self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_FULL_ADDRESS)
        )
        if not full_address:
            full_address = self.db._format_full_address(
                line1=address_line1,
                line2=address_line2,
                city=city,
                state_province=state_province,
                postal_code=postal_code,
                country=country,
            )

        onboarding_status = _clean_text(
            self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_ONBOARDING_STATUS)
        )
        if not onboarding_status:
            onboarding_status = "Prospect"

        kyc_status = _clean_text(
            self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_KYC_STATUS)
        )
        if not kyc_status:
            kyc_status = "Pending"

        created_at = _clean_text(
            self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_CREATED)
        )
        if not created_at:
            created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        updated_at = _clean_text(
            self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_UPDATED)
        )
        if not updated_at:
            updated_at = created_at

        return {
            sc.COL_PROFILE_CLIENT_ID: client_id,
            sc.COL_PROFILE_LEGAL_NAME: legal_name,
            sc.COL_PROFILE_DISPLAY_NAME: display_name,
            sc.COL_PROFILE_FIRST_NAME: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_FIRST_NAME)
            ),
            sc.COL_PROFILE_MIDDLE_NAME: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_MIDDLE_NAME)
            ),
            sc.COL_PROFILE_LAST_NAME: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_LAST_NAME)
            ),
            sc.COL_PROFILE_ENTITY_TYPE: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_ENTITY_TYPE)
            ),
            sc.COL_PROFILE_PRINCIPAL_NAME: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_PRINCIPAL_NAME)
            ),
            sc.COL_PROFILE_PRINCIPAL_POSITION: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_PRINCIPAL_POSITION)
            ),
            sc.COL_PROFILE_PRIMARY_EMAIL: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_PRIMARY_EMAIL)
            ),
            sc.COL_PROFILE_PRIMARY_PHONE: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_PRIMARY_PHONE)
            ),
            sc.COL_PROFILE_SECONDARY_CONTACT: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_SECONDARY_CONTACT)
            ),
            sc.COL_PROFILE_SECONDARY_POSITION: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_SECONDARY_POSITION)
            ),
            sc.COL_PROFILE_SECONDARY_EMAIL: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_SECONDARY_EMAIL)
            ),
            sc.COL_PROFILE_SECONDARY_PHONE: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_SECONDARY_PHONE)
            ),
            sc.COL_PROFILE_ADDR1: address_line1,
            sc.COL_PROFILE_ADDR2: address_line2,
            sc.COL_PROFILE_CITY: city,
            sc.COL_PROFILE_STATE: state_province,
            sc.COL_PROFILE_POSTAL: postal_code,
            sc.COL_PROFILE_COUNTRY: country,
            sc.COL_PROFILE_FULL_ADDRESS: full_address,
            sc.COL_PROFILE_PARENT_ID: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_PARENT_ID)
            ),
            sc.COL_PROFILE_PARENT_NAME: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_PARENT_NAME)
            ),
            sc.COL_PROFILE_WEBSITE: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_WEBSITE)
            ),
            sc.COL_PROFILE_TAX_ID: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_TAX_ID)
            ),
            sc.COL_PROFILE_INDUSTRY: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_INDUSTRY)
            ),
            sc.COL_PROFILE_BILLING_EMAIL: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_BILLING_EMAIL)
            ),
            sc.COL_PROFILE_KYC_STATUS: kyc_status,
            sc.COL_PROFILE_ONBOARDING_STATUS: onboarding_status,
            sc.COL_PROFILE_RETAINER_REQUIRED: self.db._to_bool_int(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_RETAINER_REQUIRED),
                default=0,
            ),
            sc.COL_PROFILE_RETAINER_AMOUNT: round(
                float(
                    self.db._parse_float(
                        self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_RETAINER_AMOUNT)
                    )
                    or 0.0
                ),
                2,
            ),
            sc.COL_PROFILE_ENGAGEMENT_START: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_ENGAGEMENT_START)
            ),
            sc.COL_PROFILE_DATE_CLIENT_ADDED: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_DATE_CLIENT_ADDED)
            ),
            sc.COL_PROFILE_BIRTHDAY: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_BIRTHDAY)
            ),
            sc.COL_PROFILE_REFERRAL_FROM: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_REFERRAL_FROM)
            ),
            sc.COL_PROFILE_CONFLICT_NOTES: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_CONFLICT_NOTES)
            ),
            sc.COL_PROFILE_NOTES: _clean_text(
                self.db._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_NOTES)
            ),
            sc.COL_PROFILE_CREATED: created_at,
            sc.COL_PROFILE_UPDATED: updated_at,
        }

    def _canonicalize_matter_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        matter_id = _clean_text(self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_ID))
        if not matter_id:
            matter_id = self.db._new_id("M")

        matter_number = _clean_text(
            self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_NUMBER)
        ).upper()
        if not matter_number:
            matter_number = matter_id

        matter_name = _clean_text(self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_NAME))
        display_name = _clean_text(self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_DISPLAY_NAME))
        if not display_name:
            display_name = matter_name

        status = _clean_text(self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_STATUS))
        if not status:
            status = "Open"

        matter_type = _clean_text(self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_TYPE))
        if not matter_type:
            matter_type = "General"

        billing_arrangement = _clean_text(
            self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_BILLING_ARRANGEMENT)
        )
        if not billing_arrangement:
            billing_arrangement = "Hourly"

        created_at = _clean_text(self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_CREATED))
        if not created_at:
            created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        updated_at = _clean_text(self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_UPDATED))
        if not updated_at:
            updated_at = created_at

        return {
            sc.COL_MATTER_ID: matter_id,
            sc.COL_MATTER_NUMBER: matter_number,
            sc.COL_MATTER_NAME: matter_name,
            sc.COL_MATTER_DISPLAY_NAME: display_name,
            sc.COL_MATTER_CLIENT_ID: _clean_text(
                self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_CLIENT_ID)
            ),
            sc.COL_MATTER_CLIENT_NAME: _clean_text(
                self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_CLIENT_NAME)
            ),
            sc.COL_MATTER_PARENT_ID: _clean_text(
                self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_PARENT_ID)
            ),
            sc.COL_MATTER_PARENT_NAME: _clean_text(
                self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_PARENT_NAME)
            ),
            sc.COL_MATTER_TYPE: matter_type,
            sc.COL_MATTER_PRACTICE_AREA: _clean_text(
                self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_PRACTICE_AREA)
            ),
            sc.COL_MATTER_STATUS: status,
            sc.COL_MATTER_RESPONSIBLE_LAWYER: _clean_text(
                self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_RESPONSIBLE_LAWYER)
            ),
            sc.COL_MATTER_BILLING_ARRANGEMENT: billing_arrangement,
            sc.COL_MATTER_BILLING_CONTACT: _clean_text(
                self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_BILLING_CONTACT)
            ),
            sc.COL_MATTER_BILLING_EMAIL: _clean_text(
                self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_BILLING_EMAIL)
            ),
            sc.COL_MATTER_DEF_RATE: round(
                float(self.db._parse_float(self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_DEF_RATE)) or 0.0),
                2,
            ),
            sc.COL_MATTER_DEF_SHARE: round(
                normalize_pct(self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_DEF_SHARE), default_pct=100.0),
                2,
            ),
            sc.COL_MATTER_ENGAGEMENT_DATE: _clean_text(
                self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_ENGAGEMENT_DATE)
            ),
            sc.COL_MATTER_OPEN_DATE: _clean_text(
                self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_OPEN_DATE)
            ),
            sc.COL_MATTER_CLOSE_DATE: _clean_text(
                self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_CLOSE_DATE)
            ),
            sc.COL_MATTER_COURT_FILE_NO: _clean_text(
                self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_COURT_FILE_NO)
            ),
            sc.COL_MATTER_OPPOSING_PARTY: _clean_text(
                self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_OPPOSING_PARTY)
            ),
            sc.COL_MATTER_REFERRAL_FROM: _clean_text(
                self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_REFERRAL_FROM)
            ),
            sc.COL_MATTER_DESCRIPTION: _clean_text(
                self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_DESCRIPTION)
            ),
            sc.COL_MATTER_NOTES: _clean_text(self.db._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_NOTES)),
            sc.COL_MATTER_CREATED: created_at,
            sc.COL_MATTER_UPDATED: updated_at,
        }

    def _is_client_row_active(self, row: Dict[str, Any]) -> bool:
        status = _clean_text(row.get(sc.COL_CLIENT_STATUS)).lower()
        active_flag = self.db._to_bool_int(row.get(sc.COL_CLIENT_ACTIVE), default=1)
        return active_flag == 1 and status not in ("inactive", "closed", "archived")

    def _is_matter_row_active(self, row: Dict[str, Any]) -> bool:
        status = _clean_text(row.get(sc.COL_MATTER_STATUS)).lower()
        return status not in ("inactive", "closed", "archived")

    def _build_matter_number(
        self,
        client_name: str,
        matter_type: str,
        date_opened: str,
        existing_matter_id: str = "",
        parent_name: str = "",
        entity_type: str = "",
    ) -> str:
        """
        Build a Legacy-format matter number.

        Format (no parent):   CLNT-TYP-YY-NNNN
        Format (with parent): CLNT-PRNT-TYP-YY-NNNN

        where:
          CLNT  = 4-char client code
          PRNT  = 4-char parent-entity code (only when parent is set)
          TYP   = 3-char matter-type code  (TMK / COM / CRP / EST / TAX / AUD / GEN)
          YY    = 2-digit year
          NNNN  = 4-digit globally-incrementing sequence within the year
        """
        year_code   = _matter_year_two_digits(date_opened)
        client_code = self.db._legacy_client_code(client_name, entity_type, existing_matter_id)
        type_code   = _legacy_matter_type_code(matter_type)

        # Parent code – only inject when a parent name is present
        parent_code = ""
        if _clean_text(parent_name):
            parent_code = _legacy_raw_client_code(parent_name, entity_type="")

        # Idempotency: if the existing matter already carries a valid number,
        # keep it as long as the key segments still match.
        if _clean_text(existing_matter_id):
            existing_row = self.db._find_matter_row(existing_matter_id)
            if existing_row:
                existing_num = _clean_text(existing_row.get(sc.COL_MATTER_NUMBER, "")).upper()
                if parent_code:
                    expected_prefix = f"{client_code}-{parent_code}-{type_code}-{year_code}-"
                else:
                    expected_prefix = f"{client_code}-{type_code}-{year_code}-"
                if existing_num.startswith(expected_prefix):
                    return existing_num

        next_seq = self.db._next_matter_sequence_global(
            year_two_digits=year_code,
            existing_matter_id=existing_matter_id,
        )

        if parent_code:
            return f"{client_code}-{parent_code}-{type_code}-{year_code}-{next_seq:04d}"
        return f"{client_code}-{type_code}-{year_code}-{next_seq:04d}"
