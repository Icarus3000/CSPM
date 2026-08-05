from __future__ import annotations

import sys
import csv
import math
import json
import os
from dataclasses import dataclass
from datetime import datetime, UTC
from pathlib import Path
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple

import typing
if typing.TYPE_CHECKING:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils.cell import get_column_letter, range_boundaries
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.worksheet import Worksheet
else:
    Workbook = None
    
    MergedCell = None
    get_column_letter = None
    range_boundaries = None
    Table = None
    TableStyleInfo = None
    Worksheet = None

def _lazy_load_heavy_libs():
    global Workbook, load_workbook, MergedCell, get_column_letter, range_boundaries, Table, TableStyleInfo, Worksheet
    if Workbook is None:
        from openpyxl import Workbook, load_workbook
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils.cell import get_column_letter, range_boundaries
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.worksheet.worksheet import Worksheet


from uuid import uuid4

# Support direct execution via:
# python src/python/repositories/excel_repo.py
if __package__ in (None, ""):
    _THIS_FILE = Path(__file__).resolve()
    _PYTHON_ROOT = _THIS_FILE.parents[1]
    _PYTHON_ROOT_STR = str(_PYTHON_ROOT)
    if _PYTHON_ROOT_STR not in sys.path:
        sys.path.insert(0, _PYTHON_ROOT_STR)


from domain.money import calc_amounts, normalize_pct
from domain import schema_constants as sc
from services.paths import AppPaths
from repositories.excel_repo import ExcelRepo, TableRef

import logging
logger = logging.getLogger(__name__)

TBL_PARENTS = TableRef(sc.SHEET_PARENTS, sc.TBL_PARENTS)
TBL_CLIENTS = TableRef(sc.SHEET_CLIENTS, sc.TBL_CLIENTS)
TBL_CLIENT_PROFILES = TableRef(sc.SHEET_CLIENT_PROFILES, sc.TBL_CLIENT_PROFILES)
TBL_MATTERS = TableRef(sc.SHEET_MATTERS, sc.TBL_MATTERS)
TBL_TIME = TableRef(sc.SHEET_TIME, sc.TBL_TIME)
TBL_TRADEMARKS = TableRef(sc.SHEET_TRADEMARKS, sc.TBL_TRADEMARKS)
TBL_TRANSACTIONS_MASTER = TableRef(sc.SHEET_TRANSACTIONS, sc.TBL_TRANSACTIONS_MASTER)
TBL_TRANSACTION_ACCOUNTS = TableRef(sc.SHEET_TRANSACTION_ACCOUNTS, sc.TBL_TRANSACTION_ACCOUNTS)
TBL_TRANSACTION_CATEGORIES = TableRef(sc.SHEET_TRANSACTION_CATEGORIES, sc.TBL_TRANSACTION_CATEGORIES)
TBL_TRANSACTION_BUSINESS_UNITS = TableRef(sc.SHEET_TRANSACTION_BUSINESS_UNITS, sc.TBL_TRANSACTION_BUSINESS_UNITS)
TBL_TRANSACTION_PAYEES = TableRef(sc.SHEET_TRANSACTION_PAYEES, sc.TBL_TRANSACTION_PAYEES)

TABLES_IN_ORDER = [
    TBL_PARENTS,
    TBL_CLIENTS,
    TBL_CLIENT_PROFILES,
    TBL_MATTERS,
    TBL_TIME,
    TBL_TRADEMARKS,
    TBL_TRANSACTIONS_MASTER,
    TBL_TRANSACTION_ACCOUNTS,
    TBL_TRANSACTION_CATEGORIES,
    TBL_TRANSACTION_BUSINESS_UNITS,
    TBL_TRANSACTION_PAYEES,
]
TABLE_COLUMNS = sc.TABLE_COLUMNS
TABLE_ALIASES = sc.TABLE_ALIASES
TABLE_META_CACHE_SCHEMA_VERSION = 1


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_us_phone(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    digits = re.sub(r"\D+", "", text)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) != 10:
        return text
    return f"{digits[0:3]}-{digits[3:6]}-{digits[6:10]}"


def _looks_like_email_address(value: Any) -> bool:
    text = _clean_text(value)
    if not text:
        return True
    return bool(re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", text))


def _normalize_search_text(value: Any) -> str:
    text = _clean_text(value).lower()
    if not text:
        return ""
    return re.sub(r"\s+", " ", text)


def _matter_token_fragment(value: Any, width: int = 3, fallback: str = "XXX") -> str:
    """Legacy compat shim – kept for any call-sites not yet migrated."""
    token = re.sub(r"[^A-Za-z0-9]+", "", _clean_text(value).upper())
    if not token:
        token = _clean_text(fallback).upper() or "X"
    if len(token) >= width:
        return token[:width]
    return (token + ("X" * width))[:width]


# ---------------------------------------------------------------------------
# Legacy matter-number helpers
# ---------------------------------------------------------------------------

# Hardcoded 3-char matter-type codes.  Keys are normalised lowercase.
_MATTER_TYPE_CODES: Dict[str, str] = {
    "trademark": "TMK",
    "trademarks": "TMK",
    "tmk": "TMK",
    "commercial": "COM",
    "com": "COM",
    "corporate": "CRP",
    "crp": "CRP",
    "corporation": "CRP",
    "estate": "EST",
    "estates": "EST",
    "est": "EST",
    "tax": "TAX",
    "taxation": "TAX",
    "taxes": "TAX",
    "audit": "AUD",
    "audits": "AUD",
    "aud": "AUD",
    "income": "INC",
    "inc": "INC",
    "general": "GEN",
    "gen": "GEN",
}

# Words that should be stripped from the *end* of a company name before coding
_COMPANY_NOISE_SUFFIXES = {
    "inc", "ltd", "llc", "llp", "lp", "corp", "co", "plc", "pte", "pty",
    "sa", "srl", "sas", "gmbh", "ag", "nv", "bv", "ab",
}


def _legacy_matter_type_code(matter_type: Any) -> str:
    """Return the canonical 3-char matter-type code for the given string."""
    key = re.sub(r"[^a-z]", "", _clean_text(matter_type).lower())
    return _MATTER_TYPE_CODES.get(key, "GEN")


def _legacy_raw_client_code(name: Any, entity_type: Any = "") -> str:
    """
    Derive an *uncollision-checked* 4-character client code from *name*.

    Rules
    -----
    Individual  (EntityType == "Individual"):
        last[:4].  If len(last) < 4, pad with first[:needed].
    Company / anything else:
        Strip trailing noise suffixes, then take first-word[:4].
    In both cases, keep alphanumerics (including digits), pad with "X",
    return exactly 4 uppercase characters.
    """
    cleaned = re.sub(r"[^A-Za-z0-9 ]+", "", _clean_text(name)).strip()
    parts = cleaned.split()
    if not parts:
        return "XXXX"

    is_individual = _clean_text(entity_type).lower() == "individual"

    if is_individual:
        last = parts[-1].upper() if len(parts) >= 2 else parts[0].upper()
        raw = last[:4]
        if len(raw) < 4 and len(parts) >= 2:
            first = parts[0].upper()
            raw = raw + first[: 4 - len(raw)]
    else:
        # company: try the first meaningful word
        meaningful_parts = [p for p in parts if p.lower() not in _COMPANY_NOISE_SUFFIXES]
        main = (meaningful_parts[0] if meaningful_parts else parts[0]).upper()
        raw = main[:4]

    return (raw + "XXXX")[:4]


def _matter_year_two_digits(date_text: Any) -> str:
    text = _clean_text(date_text)
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", text)
    if match:
        return match.group(1)[2:]
    return datetime.now().strftime("%y")


def _normalize_choice(value: Any, allowed: List[str], default_value: str = "") -> str:
    normalized_allowed: Dict[str, str] = {
        _clean_text(choice).lower(): _clean_text(choice)
        for choice in allowed
        if _clean_text(choice)
    }
    if not normalized_allowed:
        return _clean_text(value) or _clean_text(default_value)
    raw = _clean_text(value)
    if not raw:
        return _clean_text(default_value) if _clean_text(default_value) else next(iter(normalized_allowed.values()))
    return normalized_allowed.get(raw.lower(), "")


def _is_valid_iso_date(text: Any) -> bool:
    value = _clean_text(text)
    if not value:
        return False
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", value):
        return False
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return True
    except Exception:
        return False


def _to_code_token(value: Any, fallback: str = "X") -> str:
    token = re.sub(r"[^A-Za-z0-9]+", "_", _clean_text(value).strip()).strip("_").upper()
    if token:
        return token
    return _clean_text(fallback).upper() or "X"


def _search_terms(value: Any) -> List[str]:
    normalized = _normalize_search_text(value)
    if not normalized:
        return []

    terms: List[str] = []
    seen = set()
    for raw_token in re.findall(r"[a-z0-9@._:/#\-]+", normalized, flags=re.IGNORECASE):
        token = raw_token.strip(" .,_-")
        if not token:
            continue
        if token in seen:
            continue
        seen.add(token)
        terms.append(token)

    if " " in normalized and normalized not in seen:
        terms.insert(0, normalized)
    return terms


def _boolean_query_tokens(value: Any) -> List[str]:
    raw = _clean_text(value)
    if not raw:
        return []
    return [tok for tok in re.findall(r"\(|\)|AND|OR|NOT|[^\s()]+", raw, flags=re.IGNORECASE) if tok]


def _boolean_to_postfix(tokens: List[str]) -> List[str]:
    precedence = {"OR": 1, "AND": 2, "NOT": 3}
    output: List[str] = []
    stack: List[str] = []

    for raw_tok in tokens:
        tok = str(raw_tok or "").strip()
        if not tok:
            continue
        upper = tok.upper()
        if upper in ("AND", "OR", "NOT"):
            if upper == "NOT":
                while stack and stack[-1] in precedence and precedence[stack[-1]] > precedence[upper]:
                    output.append(stack.pop())
            else:
                while stack and stack[-1] in precedence and precedence[stack[-1]] >= precedence[upper]:
                    output.append(stack.pop())
            stack.append(upper)
            continue
        if tok == "(":
            stack.append(tok)
            continue
        if tok == ")":
            while stack and stack[-1] != "(":
                output.append(stack.pop())
            if stack and stack[-1] == "(":
                stack.pop()
            continue
        output.append(_normalize_search_text(tok))

    while stack:
        top = stack.pop()
        if top in ("(", ")"):
            continue
        output.append(top)
    return output


def _match_boolean_postfix(postfix: List[str], haystack: str) -> bool:
    if not postfix:
        return False
    stack: List[bool] = []

    for tok in postfix:
        upper = tok.upper()
        if upper == "NOT":
            value = stack.pop() if stack else False
            stack.append(not value)
            continue
        if upper in ("AND", "OR"):
            right = stack.pop() if stack else False
            left = stack.pop() if stack else False
            stack.append(left and right if upper == "AND" else left or right)
            continue
        needle = _normalize_search_text(tok)
        stack.append(bool(needle) and needle in haystack)

    return stack[-1] if stack else False


def _normalize_index(value: Any, fallback: int) -> int:
    try:
        parsed = int(value)
        if parsed > 0:
            return parsed
    except Exception:
        pass
    return fallback


def _safe_range_boundaries(ref: str) -> Tuple[int, int, int, int]:
    _lazy_load_heavy_libs()
    min_col_raw, min_row_raw, max_col_raw, max_row_raw = range_boundaries(ref)
    min_col = _normalize_index(min_col_raw, 1)
    min_row = _normalize_index(min_row_raw, 1)
    max_col = _normalize_index(max_col_raw, min_col)
    max_row = _normalize_index(max_row_raw, min_row)
    if max_col < min_col:
        max_col = min_col
    if max_row < min_row:
        max_row = min_row
    return min_col, min_row, max_col, max_row


def _set_cell_value(ws: Worksheet, row_idx: int, col_idx: int, value: Any) -> None:
    _lazy_load_heavy_libs()
    cell_obj = ws.cell(row=row_idx, column=col_idx)
    if isinstance(cell_obj, MergedCell):
        return
    cell_obj.value = value

class FinanceRepo:
    def __init__(self, db: ExcelRepo):
        self.db = db

    def list_transaction_accounts(self, include_inactive: bool = False) -> List[Dict[str, Any]]:
        rows = [self.db._canonicalize_transaction_account_row(r) for r in self.db._read_table_rows(TBL_TRANSACTION_ACCOUNTS)]
        if not rows:
            seed_rows = self.db._load_seed_csv_rows("transactions_master.accounts.seed.csv")
            rows = [self.db._canonicalize_transaction_account_row(r) for r in seed_rows]
        out: List[Dict[str, Any]] = []
        for row in rows:
            active = self.db._to_bool_int(row.get(sc.COL_TXN_ACCOUNT_ACTIVE), default=1)
            if not include_inactive and active != 1:
                continue
            alias_text = _clean_text(row.get(sc.COL_TXN_ACCOUNT_ALIASES))
            aliases = [a.strip() for a in alias_text.split(",") if a.strip()]
            out.append(
                {
                    "accountCode": _clean_text(row.get(sc.COL_TXN_ACCOUNT_CODE)),
                    "accountName": _clean_text(row.get(sc.COL_TXN_ACCOUNT_NAME)),
                    "accountKind": _clean_text(row.get(sc.COL_TXN_ACCOUNT_KIND)),
                    "owner": _clean_text(row.get(sc.COL_TXN_ACCOUNT_OWNER)),
                    "active": active,
                    "aliases": aliases,
                }
            )
        out.sort(
            key=lambda row: (
                _normalize_search_text(row.get("accountName")),
                _normalize_search_text(row.get("accountCode")),
            )
        )
        return out

    def list_transaction_categories(
        self,
        txn_type: str = "",
        txn_class: str = "",
        include_inactive: bool = False,
    ) -> List[Dict[str, Any]]:
        type_filter = _normalize_search_text(txn_type)
        class_filter = _normalize_search_text(txn_class)
        rows = [
            self.db._canonicalize_transaction_category_row(r)
            for r in self.db._read_table_rows(TBL_TRANSACTION_CATEGORIES)
        ]
        if not rows:
            seed_rows = self.db._load_seed_csv_rows("transactions_master.categories.seed.csv")
            rows = [self.db._canonicalize_transaction_category_row(r) for r in seed_rows]
        out: List[Dict[str, Any]] = []
        for row in rows:
            active = self.db._to_bool_int(row.get(sc.COL_TXN_CATEGORY_LKP_ACTIVE), default=1)
            if not include_inactive and active != 1:
                continue
            row_type = _clean_text(row.get(sc.COL_TXN_CATEGORY_LKP_TYPE))
            row_scope = _clean_text(row.get(sc.COL_TXN_CATEGORY_LKP_CLASS_SCOPE))
            if type_filter and not self._txn_scope_match(row_type, type_filter):
                continue
            if class_filter and not self._txn_scope_match(row_scope, class_filter):
                continue
            out.append(
                {
                    "categoryCode": _clean_text(row.get(sc.COL_TXN_CATEGORY_LKP_CODE)),
                    "categoryName": _clean_text(row.get(sc.COL_TXN_CATEGORY_LKP_NAME)),
                    "type": row_type,
                    "classScope": row_scope,
                    "taxFlagDefault": _clean_text(row.get(sc.COL_TXN_CATEGORY_LKP_TAX_FLAG_DEFAULT)),
                    "billableAllowed": self.db._to_bool_int(
                        row.get(sc.COL_TXN_CATEGORY_LKP_BILLABLE_ALLOWED), default=0
                    ),
                    "medicalEligible": self.db._to_bool_int(
                        row.get(sc.COL_TXN_CATEGORY_LKP_MEDICAL_ELIGIBLE), default=0
                    ),
                    "deductibleEligible": self.db._to_bool_int(
                        row.get(sc.COL_TXN_CATEGORY_LKP_DEDUCTIBLE_ELIGIBLE), default=0
                    ),
                    "businessDeductibleEligible": self.db._to_bool_int(
                        row.get(sc.COL_TXN_CATEGORY_LKP_BUSINESS_DEDUCTIBLE_ELIGIBLE), default=0
                    ),
                    "active": active,
                    "sortOrder": self.db._parse_int(row.get(sc.COL_TXN_CATEGORY_LKP_SORT_ORDER)) or 0,
                    "notes": _clean_text(row.get(sc.COL_TXN_CATEGORY_LKP_NOTES)),
                }
            )
        out.sort(
            key=lambda row: (
                int(row.get("sortOrder", 0)),
                _normalize_search_text(row.get("categoryName")),
                _normalize_search_text(row.get("categoryCode")),
            )
        )
        return out

    def list_transaction_business_units(self, include_inactive: bool = False) -> List[Dict[str, Any]]:
        rows = [
            self.db._canonicalize_transaction_business_unit_row(r)
            for r in self.db._read_table_rows(TBL_TRANSACTION_BUSINESS_UNITS)
        ]
        if not rows:
            seed_rows = self.db._load_seed_csv_rows("transactions_master.business_units.seed.csv")
            rows = [self.db._canonicalize_transaction_business_unit_row(r) for r in seed_rows]
        out: List[Dict[str, Any]] = []
        for row in rows:
            active = self.db._to_bool_int(row.get(sc.COL_TXN_BUSINESS_UNIT_ACTIVE), default=1)
            if not include_inactive and active != 1:
                continue
            out.append(
                {
                    "businessUnit": _clean_text(row.get(sc.COL_TXN_BUSINESS_UNIT_NAME)),
                    "owner": _clean_text(row.get(sc.COL_TXN_BUSINESS_UNIT_OWNER)),
                    "active": active,
                }
            )
        out.sort(key=lambda row: _normalize_search_text(row.get("businessUnit")))
        return out

    def list_transaction_payees(self, include_inactive: bool = False) -> List[Dict[str, Any]]:
        rows = [self.db._canonicalize_transaction_payee_row(r) for r in self.db._read_table_rows(TBL_TRANSACTION_PAYEES)]
        if not rows:
            seed_rows = self.db._load_seed_csv_rows("transactions_master.payees.seed.csv")
            rows = [self.db._canonicalize_transaction_payee_row(r) for r in seed_rows]
        out: List[Dict[str, Any]] = []
        for row in rows:
            active = self.db._to_bool_int(row.get(sc.COL_TXN_PAYEE_ACTIVE), default=1)
            if not include_inactive and active != 1:
                continue
            out.append(
                {
                    "payeeName": _clean_text(row.get(sc.COL_TXN_PAYEE_NAME)),
                    "defaultCategoryCode": _clean_text(row.get(sc.COL_TXN_PAYEE_DEFAULT_CATEGORY_CODE)),
                    "active": active,
                }
            )
        out.sort(key=lambda row: _normalize_search_text(row.get("payeeName")))
        return out

    def list_transactions(self, filters: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        f = dict(filters or {})
        query = _normalize_search_text(f.get("query", ""))
        class_filter = _normalize_search_text(f.get("class") or f.get("txnClass"))
        type_filter = _normalize_search_text(f.get("type") or f.get("txnType"))
        status_filter = _normalize_search_text(f.get("status"))
        business_unit_filter = _normalize_search_text(f.get("businessUnit"))
        parent_filter = _normalize_search_text(f.get("parent"))
        client_filter = _normalize_search_text(f.get("client"))
        matter_filter = _normalize_search_text(f.get("matter"))
        include_void = self.db._to_bool_int(f.get("includeVoid"), default=1) == 1

        date_from = _clean_text(
            f.get("dateFrom") or f.get("fromDate") or f.get("startDate")
        )
        date_to = _clean_text(
            f.get("dateTo") or f.get("toDate") or f.get("endDate")
        )
        if date_from and not _is_valid_iso_date(date_from):
            raise ValueError("dateFrom must be in YYYY-MM-DD format.")
        if date_to and not _is_valid_iso_date(date_to):
            raise ValueError("dateTo must be in YYYY-MM-DD format.")

        rows = [
            self.db._canonicalize_transaction_row(r)
            for r in self.db._read_table_rows(TBL_TRANSACTIONS_MASTER)
        ]
        out: List[Dict[str, Any]] = []
        for row in rows:
            row_date = _clean_text(row.get(sc.COL_TXN_DATE))
            row_class = _normalize_search_text(row.get(sc.COL_TXN_CLASS))
            row_type = _normalize_search_text(row.get(sc.COL_TXN_TYPE))
            row_status = _normalize_search_text(row.get(sc.COL_TXN_STATUS))
            row_business_unit = _normalize_search_text(row.get(sc.COL_TXN_BUSINESS_UNIT))
            row_parent = _normalize_search_text(row.get(sc.COL_TXN_PARENT))
            row_client = _normalize_search_text(row.get(sc.COL_TXN_CLIENT))
            row_matter = _normalize_search_text(row.get(sc.COL_TXN_MATTER))

            if class_filter and row_class != class_filter:
                continue
            if type_filter and row_type != type_filter:
                continue
            if status_filter and row_status != status_filter:
                continue
            if business_unit_filter and business_unit_filter not in row_business_unit:
                continue
            if parent_filter and parent_filter not in row_parent:
                continue
            if client_filter and client_filter not in row_client:
                continue
            if matter_filter and matter_filter not in row_matter:
                continue
            if not include_void and row_status == "void":
                continue

            if date_from and row_date and row_date < date_from:
                continue
            if date_to and row_date and row_date > date_to:
                continue

            if query:
                haystack = " | ".join(_normalize_search_text(v) for v in row.values() if _clean_text(v))
                if query not in haystack:
                    continue

            out.append(
                {
                    "transactionId": _clean_text(row.get(sc.COL_TXN_ID)),
                    "txnDate": row_date,
                    "class": _clean_text(row.get(sc.COL_TXN_CLASS)),
                    "businessUnit": _clean_text(row.get(sc.COL_TXN_BUSINESS_UNIT)),
                    "type": _clean_text(row.get(sc.COL_TXN_TYPE)),
                    "fromAccount": _clean_text(row.get(sc.COL_TXN_FROM_ACCOUNT)),
                    "toAccount": _clean_text(row.get(sc.COL_TXN_TO_ACCOUNT)),
                    "payee": _clean_text(row.get(sc.COL_TXN_PAYEE)),
                    "parent": _clean_text(row.get(sc.COL_TXN_PARENT)),
                    "client": _clean_text(row.get(sc.COL_TXN_CLIENT)),
                    "matter": _clean_text(row.get(sc.COL_TXN_MATTER)),
                    "categoryCode": _clean_text(row.get(sc.COL_TXN_CATEGORY_CODE)),
                    "categoryName": _clean_text(row.get(sc.COL_TXN_CATEGORY_NAME)),
                    "member": _clean_text(row.get(sc.COL_TXN_MEMBER)),
                    "amount": round(float(self.db._parse_float(row.get(sc.COL_TXN_AMOUNT)) or 0.0), 2),
                    "taxAmount": round(float(self.db._parse_float(row.get(sc.COL_TXN_TAX_AMOUNT)) or 0.0), 2),
                    "taxFlag": _clean_text(row.get(sc.COL_TXN_TAX_FLAG)),
                    "hstExempt": self.db._to_bool_int(row.get(sc.COL_TXN_HST_EXEMPT), default=0),
                    "generalOfficeExpense": self.db._to_bool_int(
                        row.get(sc.COL_TXN_GENERAL_OFFICE_EXPENSE), default=0
                    ),
                    "shadow": self.db._to_bool_int(row.get(sc.COL_TXN_SHADOW), default=0),
                    "invoiceRef": _clean_text(row.get(sc.COL_TXN_INVOICE_REF)),
                    "billClaimPct": round(float(self.db._parse_float(row.get(sc.COL_TXN_BILL_CLAIM_PCT)) or 0.0), 2),
                    "totalClaimAmount": round(
                        float(self.db._parse_float(row.get(sc.COL_TXN_TOTAL_CLAIM_AMOUNT)) or 0.0), 2
                    ),
                    "expenseDetails": _clean_text(row.get(sc.COL_TXN_EXPENSE_DETAILS)),
                    "notes": _clean_text(row.get(sc.COL_TXN_NOTES)),
                    "status": _clean_text(row.get(sc.COL_TXN_STATUS)),
                    "currency": _clean_text(row.get(sc.COL_TXN_CURRENCY)),
                    "voidReason": _clean_text(row.get(sc.COL_TXN_VOID_REASON)),
                    "clearedAt": _clean_text(row.get(sc.COL_TXN_CLEARED_AT)),
                    "reconciledAt": _clean_text(row.get(sc.COL_TXN_RECONCILED_AT)),
                    "createdAt": _clean_text(row.get(sc.COL_TXN_CREATED_AT)),
                    "updatedAt": _clean_text(row.get(sc.COL_TXN_UPDATED_AT)),
                }
            )

        out.sort(
            key=lambda row: (
                _clean_text(row.get("txnDate")),
                _clean_text(row.get("updatedAt")),
                _clean_text(row.get("transactionId")),
            ),
            reverse=True,
        )
        return out

    def save_transaction(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.db.ensure_schema()
        normalized = self._normalize_transaction_payload(payload)

        txn_id = _clean_text(normalized.get("transactionId"))
        existing_row: Optional[Dict[str, Any]] = self._find_transaction_row(txn_id) if txn_id else None
        self._validate_transaction_status_transition(
            old_status=_clean_text((existing_row or {}).get(sc.COL_TXN_STATUS)),
            new_status=_clean_text(normalized.get("status")),
            has_existing=existing_row is not None,
        )
        if not txn_id:
            txn_id = self.db._new_id("TXN")

        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        created_at = (
            _clean_text(normalized.get("createdAt"))
            or _clean_text((existing_row or {}).get(sc.COL_TXN_CREATED_AT))
            or now_stamp
        )

        transaction_row = {
            sc.COL_TXN_ID: txn_id,
            sc.COL_TXN_DATE: normalized["txnDate"],
            sc.COL_TXN_CLASS: normalized["class"],
            sc.COL_TXN_BUSINESS_UNIT: normalized["businessUnit"],
            sc.COL_TXN_TYPE: normalized["type"],
            sc.COL_TXN_FROM_ACCOUNT: normalized["fromAccount"],
            sc.COL_TXN_TO_ACCOUNT: normalized["toAccount"],
            sc.COL_TXN_PAYEE: normalized["payee"],
            sc.COL_TXN_PARENT: normalized["parent"],
            sc.COL_TXN_CLIENT: normalized["client"],
            sc.COL_TXN_MATTER: normalized["matter"],
            sc.COL_TXN_CATEGORY_CODE: normalized["categoryCode"],
            sc.COL_TXN_CATEGORY_NAME: normalized["categoryName"],
            sc.COL_TXN_MEMBER: normalized["member"],
            sc.COL_TXN_AMOUNT: round(float(normalized["amount"]), 2),
            sc.COL_TXN_TAX_AMOUNT: round(float(normalized["taxAmount"]), 2),
            sc.COL_TXN_TAX_FLAG: normalized["taxFlag"],
            sc.COL_TXN_HST_EXEMPT: normalized["hstExempt"],
            sc.COL_TXN_GENERAL_OFFICE_EXPENSE: normalized["generalOfficeExpense"],
            sc.COL_TXN_SHADOW: normalized["shadow"],
            sc.COL_TXN_INVOICE_REF: normalized["invoiceRef"],
            sc.COL_TXN_BILL_CLAIM_PCT: round(float(normalized["billClaimPct"]), 2),
            sc.COL_TXN_TOTAL_CLAIM_AMOUNT: round(float(normalized["totalClaimAmount"]), 2),
            sc.COL_TXN_EXPENSE_DETAILS: normalized["expenseDetails"],
            sc.COL_TXN_NOTES: normalized["notes"],
            sc.COL_TXN_STATUS: normalized["status"],
            sc.COL_TXN_CURRENCY: normalized["currency"],
            sc.COL_TXN_VOID_REASON: normalized["voidReason"],
            sc.COL_TXN_CLEARED_AT: normalized["clearedAt"],
            sc.COL_TXN_RECONCILED_AT: normalized["reconciledAt"],
            sc.COL_TXN_CREATED_AT: created_at,
            sc.COL_TXN_UPDATED_AT: now_stamp,
        }

        self.db._upsert_row_by_key(TBL_TRANSACTIONS_MASTER, sc.COL_TXN_ID, txn_id, transaction_row)
        persisted = self._find_transaction_row(txn_id)
        verified = self.db._compare_transaction_rows_loose(transaction_row, persisted)

        warnings = list(normalized.get("warnings", []))
        if verified and warnings:
            message = "Transaction saved with warnings: " + "; ".join(warnings)
        elif verified:
            message = ""
        else:
            message = "Transaction verification failed."

        return {
            "ok": bool(verified),
            "verifiedExact": bool(verified),
            "transactionId": txn_id,
            "savedRow": transaction_row,
            "warnings": warnings,
            "message": message,
        }

    def _canonicalize_transaction_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        txn_id = _clean_text(self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_ID))
        if not txn_id:
            txn_id = self.db._new_id("TXN")

        txn_date = _clean_text(self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_DATE))
        if not _is_valid_iso_date(txn_date):
            txn_date = datetime.now().strftime("%Y-%m-%d")

        txn_class = _normalize_choice(
            self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_CLASS),
            self.db.TXN_CLASS_OPTIONS,
            "Family",
        )
        if not txn_class:
            txn_class = "Family"

        txn_type = _normalize_choice(
            self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_TYPE),
            self.db.TXN_TYPE_OPTIONS,
            "Expense",
        )
        if not txn_type:
            txn_type = "Expense"

        member = _normalize_choice(
            self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_MEMBER),
            self.db.TXN_MEMBER_OPTIONS,
            "Joint",
        )
        if not member:
            member = "Joint"

        tax_flag = _normalize_choice(
            self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_TAX_FLAG),
            self.db.TXN_TAX_FLAG_OPTIONS,
            "None",
        )
        if not tax_flag:
            tax_flag = "None"

        status = _normalize_choice(
            self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_STATUS),
            self.db.TXN_STATUS_OPTIONS,
            "Pending",
        )
        if not status:
            status = "Pending"

        currency = _normalize_choice(
            self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_CURRENCY),
            self.db.TXN_CURRENCY_OPTIONS,
            "CAD",
        )
        if not currency:
            currency = "CAD"

        amount = float(
            self.db._parse_float(self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_AMOUNT))
            or 0.0
        )
        tax_amount = float(
            self.db._parse_float(self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_TAX_AMOUNT))
            or 0.0
        )
        bill_claim_pct = float(
            self.db._parse_float(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_BILL_CLAIM_PCT)
            )
            or 0.0
        )
        if bill_claim_pct < 0:
            bill_claim_pct = 0.0
        if bill_claim_pct > 100:
            bill_claim_pct = 100.0

        hst_exempt = self.db._to_bool_int(
            self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_HST_EXEMPT),
            default=0,
        )
        if hst_exempt == 1:
            tax_amount = 0.0

        total_claim_existing = self.db._parse_float(
            self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_TOTAL_CLAIM_AMOUNT)
        )
        total_claim_amount = (
            float(total_claim_existing)
            if total_claim_existing is not None
            else round((amount + tax_amount) * (bill_claim_pct / 100.0), 2)
        )

        created_at = _clean_text(
            self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_CREATED_AT)
        )
        if not created_at:
            created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        updated_at = _clean_text(
            self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_UPDATED_AT)
        )
        if not updated_at:
            updated_at = created_at

        return {
            sc.COL_TXN_ID: txn_id,
            sc.COL_TXN_DATE: txn_date,
            sc.COL_TXN_CLASS: txn_class,
            sc.COL_TXN_BUSINESS_UNIT: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_BUSINESS_UNIT)
            ),
            sc.COL_TXN_TYPE: txn_type,
            sc.COL_TXN_FROM_ACCOUNT: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_FROM_ACCOUNT)
            ),
            sc.COL_TXN_TO_ACCOUNT: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_TO_ACCOUNT)
            ),
            sc.COL_TXN_PAYEE: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_PAYEE)
            ),
            sc.COL_TXN_PARENT: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_PARENT)
            ),
            sc.COL_TXN_CLIENT: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_CLIENT)
            ),
            sc.COL_TXN_MATTER: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_MATTER)
            ),
            sc.COL_TXN_CATEGORY_CODE: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_CATEGORY_CODE)
            ),
            sc.COL_TXN_CATEGORY_NAME: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_CATEGORY_NAME)
            ),
            sc.COL_TXN_MEMBER: member,
            sc.COL_TXN_AMOUNT: round(amount, 2),
            sc.COL_TXN_TAX_AMOUNT: round(tax_amount, 2),
            sc.COL_TXN_TAX_FLAG: tax_flag,
            sc.COL_TXN_HST_EXEMPT: hst_exempt,
            sc.COL_TXN_GENERAL_OFFICE_EXPENSE: self.db._to_bool_int(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_GENERAL_OFFICE_EXPENSE),
                default=0,
            ),
            sc.COL_TXN_SHADOW: self.db._to_bool_int(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_SHADOW),
                default=0,
            ),
            sc.COL_TXN_INVOICE_REF: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_INVOICE_REF)
            ),
            sc.COL_TXN_BILL_CLAIM_PCT: round(bill_claim_pct, 2),
            sc.COL_TXN_TOTAL_CLAIM_AMOUNT: round(float(total_claim_amount), 2),
            sc.COL_TXN_EXPENSE_DETAILS: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_EXPENSE_DETAILS)
            ),
            sc.COL_TXN_NOTES: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_NOTES)
            ),
            sc.COL_TXN_STATUS: status,
            sc.COL_TXN_CURRENCY: currency,
            sc.COL_TXN_VOID_REASON: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_VOID_REASON)
            ),
            sc.COL_TXN_CLEARED_AT: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_CLEARED_AT)
            ),
            sc.COL_TXN_RECONCILED_AT: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_RECONCILED_AT)
            ),
            sc.COL_TXN_CREATED_AT: created_at,
            sc.COL_TXN_UPDATED_AT: updated_at,
        }

    def _canonicalize_transaction_account_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        account_code = _clean_text(
            self.db._value_with_alias(TBL_TRANSACTION_ACCOUNTS.table, row, sc.COL_TXN_ACCOUNT_CODE)
        )
        account_name = _clean_text(
            self.db._value_with_alias(TBL_TRANSACTION_ACCOUNTS.table, row, sc.COL_TXN_ACCOUNT_NAME)
        )
        if not account_code:
            account_code = _to_code_token(account_name, fallback="ACCOUNT")
        if not account_name:
            account_name = account_code

        return {
            sc.COL_TXN_ACCOUNT_CODE: account_code,
            sc.COL_TXN_ACCOUNT_NAME: account_name,
            sc.COL_TXN_ACCOUNT_KIND: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTION_ACCOUNTS.table, row, sc.COL_TXN_ACCOUNT_KIND)
            ).lower(),
            sc.COL_TXN_ACCOUNT_OWNER: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTION_ACCOUNTS.table, row, sc.COL_TXN_ACCOUNT_OWNER)
            ),
            sc.COL_TXN_ACCOUNT_ACTIVE: self.db._to_bool_int(
                self.db._value_with_alias(TBL_TRANSACTION_ACCOUNTS.table, row, sc.COL_TXN_ACCOUNT_ACTIVE),
                default=1,
            ),
            sc.COL_TXN_ACCOUNT_ALIASES: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTION_ACCOUNTS.table, row, sc.COL_TXN_ACCOUNT_ALIASES)
            ),
        }

    def _canonicalize_transaction_category_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        code = _clean_text(
            self.db._value_with_alias(TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_CODE)
        )
        name = _clean_text(
            self.db._value_with_alias(TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_NAME)
        )
        if not code:
            code = _to_code_token(name, fallback="CATEGORY")
        if not name:
            name = code

        sort_order = self.db._parse_int(
            self.db._value_with_alias(TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_SORT_ORDER)
        )
        if sort_order is None:
            sort_order = 0

        return {
            sc.COL_TXN_CATEGORY_LKP_CODE: code,
            sc.COL_TXN_CATEGORY_LKP_NAME: name,
            sc.COL_TXN_CATEGORY_LKP_TYPE: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_TYPE)
            ),
            sc.COL_TXN_CATEGORY_LKP_CLASS_SCOPE: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_CLASS_SCOPE)
            ),
            sc.COL_TXN_CATEGORY_LKP_TAX_FLAG_DEFAULT: _clean_text(
                self.db._value_with_alias(
                    TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_TAX_FLAG_DEFAULT
                )
            ),
            sc.COL_TXN_CATEGORY_LKP_BILLABLE_ALLOWED: self.db._to_bool_int(
                self.db._value_with_alias(
                    TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_BILLABLE_ALLOWED
                ),
                default=0,
            ),
            sc.COL_TXN_CATEGORY_LKP_MEDICAL_ELIGIBLE: self.db._to_bool_int(
                self.db._value_with_alias(
                    TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_MEDICAL_ELIGIBLE
                ),
                default=0,
            ),
            sc.COL_TXN_CATEGORY_LKP_DEDUCTIBLE_ELIGIBLE: self.db._to_bool_int(
                self.db._value_with_alias(
                    TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_DEDUCTIBLE_ELIGIBLE
                ),
                default=0,
            ),
            sc.COL_TXN_CATEGORY_LKP_BUSINESS_DEDUCTIBLE_ELIGIBLE: self.db._to_bool_int(
                self.db._value_with_alias(
                    TBL_TRANSACTION_CATEGORIES.table,
                    row,
                    sc.COL_TXN_CATEGORY_LKP_BUSINESS_DEDUCTIBLE_ELIGIBLE,
                ),
                default=0,
            ),
            sc.COL_TXN_CATEGORY_LKP_ACTIVE: self.db._to_bool_int(
                self.db._value_with_alias(TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_ACTIVE),
                default=1,
            ),
            sc.COL_TXN_CATEGORY_LKP_SORT_ORDER: int(sort_order),
            sc.COL_TXN_CATEGORY_LKP_NOTES: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_NOTES)
            ),
        }

    def _canonicalize_transaction_business_unit_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        unit_name = _clean_text(
            self.db._value_with_alias(TBL_TRANSACTION_BUSINESS_UNITS.table, row, sc.COL_TXN_BUSINESS_UNIT_NAME)
        )
        if not unit_name:
            unit_name = "Unassigned"
        return {
            sc.COL_TXN_BUSINESS_UNIT_NAME: unit_name,
            sc.COL_TXN_BUSINESS_UNIT_OWNER: _clean_text(
                self.db._value_with_alias(TBL_TRANSACTION_BUSINESS_UNITS.table, row, sc.COL_TXN_BUSINESS_UNIT_OWNER)
            ),
            sc.COL_TXN_BUSINESS_UNIT_ACTIVE: self.db._to_bool_int(
                self.db._value_with_alias(TBL_TRANSACTION_BUSINESS_UNITS.table, row, sc.COL_TXN_BUSINESS_UNIT_ACTIVE),
                default=1,
            ),
        }

    def _canonicalize_transaction_payee_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        payee_name = _clean_text(self.db._value_with_alias(TBL_TRANSACTION_PAYEES.table, row, sc.COL_TXN_PAYEE_NAME))
        if not payee_name:
            payee_name = "Unknown Payee"
        return {
            sc.COL_TXN_PAYEE_NAME: payee_name,
            sc.COL_TXN_PAYEE_DEFAULT_CATEGORY_CODE: _clean_text(
                self.db._value_with_alias(
                    TBL_TRANSACTION_PAYEES.table, row, sc.COL_TXN_PAYEE_DEFAULT_CATEGORY_CODE
                )
            ),
            sc.COL_TXN_PAYEE_ACTIVE: self.db._to_bool_int(
                self.db._value_with_alias(TBL_TRANSACTION_PAYEES.table, row, sc.COL_TXN_PAYEE_ACTIVE),
                default=1,
            ),
        }

    def _txn_scope_match(self, source_scope: str, expected: str) -> bool:
        expected_norm = _normalize_search_text(expected)
        source_norm = _normalize_search_text(source_scope)
        if not expected_norm:
            return True
        if not source_norm:
            return True
        if source_norm in ("all", "any", "both", "*"):
            return True
        tokens = [tok.strip() for tok in re.split(r"[|,;/]", source_norm) if tok.strip()]
        if not tokens:
            tokens = [source_norm]
        return expected_norm in tokens

    def _normalize_transaction_payload(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        warnings: List[str] = []
        today_iso = datetime.now().strftime("%Y-%m-%d")

        txn_id = self.db._pick_text(payload, ["transactionId", "txnId", sc.COL_TXN_ID, "TransactionID"])
        txn_date = self.db._pick_text(payload, ["txnDate", "date", "dateText", sc.COL_TXN_DATE])
        if not txn_date:
            txn_date = today_iso
        if not _is_valid_iso_date(txn_date):
            raise ValueError("TxnDate must be in YYYY-MM-DD format.")

        txn_class = _normalize_choice(
            self.db._pick_text(payload, ["class", "txnClass", sc.COL_TXN_CLASS]),
            self.db.TXN_CLASS_OPTIONS,
            "Family",
        )
        if not txn_class:
            raise ValueError(f"Class must be one of: {', '.join(self.db.TXN_CLASS_OPTIONS)}")

        txn_type = _normalize_choice(
            self.db._pick_text(payload, ["type", "txnType", sc.COL_TXN_TYPE]),
            self.db.TXN_TYPE_OPTIONS,
            "Expense",
        )
        if not txn_type:
            raise ValueError(f"Type must be one of: {', '.join(self.db.TXN_TYPE_OPTIONS)}")

        business_unit = self.db._pick_text(payload, ["businessUnit", "bu", sc.COL_TXN_BUSINESS_UNIT])
        if txn_class == "Business" and not business_unit:
            raise ValueError("BusinessUnit is required when Class is Business.")

        from_account = self.db._pick_text(payload, ["fromAccount", "sourceAccount", sc.COL_TXN_FROM_ACCOUNT])
        to_account = self.db._pick_text(payload, ["toAccount", "destinationAccount", sc.COL_TXN_TO_ACCOUNT])
        if not from_account:
            raise ValueError("FromAccount is required.")

        txn_type_lc = txn_type.lower()
        movement_type = txn_type_lc in self.db.TXN_MOVEMENT_TYPES
        if movement_type and not to_account:
            raise ValueError("ToAccount is required for Transfer and Debt Repayment.")
        if movement_type and to_account and from_account.lower() == to_account.lower():
            raise ValueError("FromAccount and ToAccount must differ for Transfer and Debt Repayment.")

        payee = self.db._pick_text(payload, ["payee", "vendor", sc.COL_TXN_PAYEE])
        if txn_type_lc != "transfer" and not payee:
            raise ValueError("Payee is required for non-transfer transactions.")

        member = _normalize_choice(
            self.db._pick_text(payload, ["member", sc.COL_TXN_MEMBER]),
            self.db.TXN_MEMBER_OPTIONS,
            "Joint",
        )
        if not member:
            raise ValueError(f"Member must be one of: {', '.join(self.db.TXN_MEMBER_OPTIONS)}")

        tax_flag = _normalize_choice(
            self.db._pick_text(payload, ["taxFlag", "taxCategory", sc.COL_TXN_TAX_FLAG]),
            self.db.TXN_TAX_FLAG_OPTIONS,
            "None",
        )
        if not tax_flag:
            raise ValueError(f"TaxFlag must be one of: {', '.join(self.db.TXN_TAX_FLAG_OPTIONS)}")

        status = _normalize_choice(
            self.db._pick_text(payload, ["status", sc.COL_TXN_STATUS]),
            self.db.TXN_STATUS_OPTIONS,
            "Pending",
        )
        if not status:
            raise ValueError(f"Status must be one of: {', '.join(self.db.TXN_STATUS_OPTIONS)}")

        currency = _normalize_choice(
            self.db._pick_text(payload, ["currency", sc.COL_TXN_CURRENCY]),
            self.db.TXN_CURRENCY_OPTIONS,
            "CAD",
        )
        if not currency:
            raise ValueError(f"Currency must be one of: {', '.join(self.db.TXN_CURRENCY_OPTIONS)}")

        amount = self.db._pick_float(payload, ["amount", sc.COL_TXN_AMOUNT])
        if amount is None:
            raise ValueError("Amount is required.")
        amount = float(amount)
        if amount <= 0:
            raise ValueError("Amount must be greater than 0.")

        tax_amount = self.db._pick_float(payload, ["taxAmount", "tax", "hstAmount", sc.COL_TXN_TAX_AMOUNT])
        if tax_amount is None:
            tax_amount = 0.0
        tax_amount = float(tax_amount)
        if tax_amount < 0:
            raise ValueError("TaxAmount must be >= 0.")

        hst_exempt = self.db._to_bool_int(self.db._pick_value(payload, ["hstExempt", sc.COL_TXN_HST_EXEMPT]), default=0)
        if hst_exempt == 1 and abs(tax_amount) > 0.0001:
            tax_amount = 0.0
            warnings.append("HSTExempt=1 forced TaxAmount to 0.00")

        general_office_expense = self.db._to_bool_int(
            self.db._pick_value(payload, ["generalOfficeExpense", "officeExpense", sc.COL_TXN_GENERAL_OFFICE_EXPENSE]),
            default=0,
        )
        shadow = self.db._to_bool_int(self.db._pick_value(payload, ["shadow", sc.COL_TXN_SHADOW]), default=0)

        parent = self.db._pick_text(payload, ["parent", "parentName", sc.COL_TXN_PARENT])
        client = self.db._pick_text(payload, ["client", "clientName", sc.COL_TXN_CLIENT])
        matter = self.db._pick_text(payload, ["matter", "matterName", sc.COL_TXN_MATTER])

        bill_claim_pct = self.db._pick_float(payload, ["billClaimPct", "claimPct", sc.COL_TXN_BILL_CLAIM_PCT])
        if bill_claim_pct is None:
            bill_claim_pct = 0.0
        bill_claim_pct = float(bill_claim_pct)
        if bill_claim_pct < 0 or bill_claim_pct > 100:
            raise ValueError("BillClaimPct must be between 0 and 100.")

        if general_office_expense == 1:
            if parent or client or matter:
                raise ValueError("GeneralOfficeExpense cannot include Parent/Client/Matter context.")
            if abs(bill_claim_pct) > 0.0001:
                raise ValueError("GeneralOfficeExpense cannot have BillClaimPct > 0.")
            bill_claim_pct = 0.0

        if txn_type_lc == "expense" and bill_claim_pct > 0:
            if not (parent and client and matter):
                raise ValueError("Billable expense requires Parent, Client, and Matter when BillClaimPct > 0.")

        total_claim_amount = round((amount + tax_amount) * (bill_claim_pct / 100.0), 2)

        category_code = self.db._pick_text(payload, ["categoryCode", "catCode", sc.COL_TXN_CATEGORY_CODE])
        category_name = self.db._pick_text(payload, ["categoryName", "category", sc.COL_TXN_CATEGORY_NAME])
        category_by_code, category_by_name = self._transaction_category_lookup_maps()
        if category_code and not category_name:
            found = category_by_code.get(category_code.lower())
            if found:
                category_name = _clean_text(found.get("categoryName"))
        if category_name and not category_code:
            found = category_by_name.get(category_name.lower())
            if found:
                category_code = _clean_text(found.get("categoryCode"))
        if category_code and category_by_code.get(category_code.lower()):
            canonical_name = _clean_text(category_by_code[category_code.lower()].get("categoryName"))
            if canonical_name:
                category_name = canonical_name
        if not category_code:
            raise ValueError("CategoryCode is required.")
        if not category_name:
            raise ValueError("CategoryName is required.")

        if txn_type_lc == "debt repayment":
            account_kind_map = self.db._transaction_account_kind_lookup()
            if account_kind_map:
                to_kind = _clean_text(account_kind_map.get(to_account.lower(), "")).lower()
                if not to_kind:
                    raise ValueError("Debt Repayment ToAccount must exist in transaction account lookup.")
                if to_kind not in self.db.TXN_DEBT_DEST_ACCOUNT_KINDS:
                    allowed = ", ".join(sorted(self.db.TXN_DEBT_DEST_ACCOUNT_KINDS))
                    raise ValueError(f"Debt Repayment ToAccount must be a debt account kind ({allowed}).")

        invoice_ref = self.db._pick_text(payload, ["invoiceRef", "invoice", sc.COL_TXN_INVOICE_REF])
        expense_details = self.db._pick_text(payload, ["expenseDetails", "details", sc.COL_TXN_EXPENSE_DETAILS])
        notes = self.db._pick_text(payload, ["notes", "notesText", sc.COL_TXN_NOTES])

        void_reason = self.db._pick_text(payload, ["voidReason", "voidNotes", sc.COL_TXN_VOID_REASON])
        cleared_at = self.db._pick_text(payload, ["clearedAt", "clearedDate", sc.COL_TXN_CLEARED_AT])
        reconciled_at = self.db._pick_text(payload, ["reconciledAt", "reconciledDate", sc.COL_TXN_RECONCILED_AT])

        if status == "Void" and not void_reason:
            raise ValueError("VoidReason is required when Status is Void.")

        if status in ("Cleared", "Reconciled") and not cleared_at:
            cleared_at = today_iso
        if status == "Reconciled" and not reconciled_at:
            reconciled_at = today_iso

        if cleared_at and not _is_valid_iso_date(cleared_at):
            raise ValueError("ClearedAt must be in YYYY-MM-DD format.")
        if reconciled_at and not _is_valid_iso_date(reconciled_at):
            raise ValueError("ReconciledAt must be in YYYY-MM-DD format.")

        if status == "Reconciled" and (not cleared_at or not reconciled_at):
            raise ValueError("Reconciled status requires ClearedAt and ReconciledAt.")

        return {
            "transactionId": txn_id,
            "txnDate": txn_date,
            "class": txn_class,
            "businessUnit": business_unit,
            "type": txn_type,
            "fromAccount": from_account,
            "toAccount": to_account,
            "payee": payee,
            "parent": parent,
            "client": client,
            "matter": matter,
            "categoryCode": category_code,
            "categoryName": category_name,
            "member": member,
            "amount": amount,
            "taxAmount": tax_amount,
            "taxFlag": tax_flag,
            "hstExempt": hst_exempt,
            "generalOfficeExpense": general_office_expense,
            "shadow": shadow,
            "invoiceRef": invoice_ref,
            "billClaimPct": bill_claim_pct,
            "totalClaimAmount": total_claim_amount,
            "expenseDetails": expense_details,
            "notes": notes,
            "status": status,
            "currency": currency,
            "voidReason": void_reason,
            "clearedAt": cleared_at,
            "reconciledAt": reconciled_at,
            "createdAt": self.db._pick_text(payload, ["createdAt", sc.COL_TXN_CREATED_AT]),
            "warnings": warnings,
        }

    def _validate_transaction_status_transition(
        self,
        old_status: str,
        new_status: str,
        has_existing: bool,
    ) -> None:
        if not has_existing:
            return

        previous = _clean_text(old_status)
        target = _clean_text(new_status)
        if not previous or not target or previous == target:
            return

        allowed_transitions = {
            "Pending": {"Pending", "Cleared", "Void"},
            "Cleared": {"Cleared", "Reconciled", "Void"},
            "Reconciled": {"Reconciled", "Void"},
            "Void": {"Void"},
        }
        if previous not in allowed_transitions:
            return

        allowed = allowed_transitions[previous]
        if target not in allowed:
            raise ValueError(
                f"Invalid Status transition: {previous} -> {target}. "
                f"Allowed next statuses: {', '.join(sorted(allowed))}."
            )

    def _find_transaction_row(self, transaction_key: str) -> Optional[Dict[str, Any]]:
        target = _clean_text(transaction_key)
        if not target:
            return None
        target_lc = target.lower()
        rows = [
            self.db._canonicalize_transaction_row(r)
            for r in self.db._read_table_rows(TBL_TRANSACTIONS_MASTER)
        ]
        for row in rows:
            if _clean_text(row.get(sc.COL_TXN_ID)).lower() == target_lc:
                return row
        return None


    def _transaction_category_lookup_maps(self) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
        by_code: Dict[str, Dict[str, Any]] = {}
        by_name: Dict[str, Dict[str, Any]] = {}
        rows = self.list_transaction_categories(txn_type="", txn_class="", include_inactive=True)
        for row in rows:
            code = _clean_text(row.get("categoryCode")).lower()
            name = _clean_text(row.get("categoryName")).lower()
            if code:
                by_code[code] = row
            if name:
                by_name[name] = row
        return by_code, by_name

