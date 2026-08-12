from __future__ import annotations

import sys
import csv
import math
import json
import os
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, timedelta, UTC
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import re
import zipfile
import threading
from functools import wraps
from urllib.parse import quote_plus
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple
import collections

_DB_LOCK = threading.RLock()

import gc
def with_db_lock(func):
    def wrapper(self, *args, **kwargs):
        with _DB_LOCK:
            was_enabled = gc.isenabled()
            if was_enabled:
                gc.disable()
            try:
                return func(self, *args, **kwargs)
            finally:
                if was_enabled:
                    gc.enable()
    return wrapper


def with_financial_write_batch(func):
    """Commit a multi-table financial mutation in one workbook replacement.

    A payment touches the transactions, ledger, receivables, time, and
    disbursements tables.  Persisting those tables one at a time made a single
    payment perform several macro-workbook rewrites, leaving more chances for
    a transient Windows/OneDrive lock and briefly exposing partial state.
    """
    @wraps(func)
    def wrapper(self, *args, **kwargs):
        # Lightweight repository doubles used by focused logic tests do not
        # have an on-disk workbook.  Keep their existing in-memory behavior.
        if not hasattr(self, "import_batch") or not hasattr(self, "_import_batch_active"):
            return func(self, *args, **kwargs)

        with _DB_LOCK:
            was_enabled = gc.isenabled()
            if was_enabled:
                gc.disable()
            try:
                # A caller can already be coordinating a larger atomic batch.
                if self._import_batch_active:
                    return func(self, *args, **kwargs)
                with self.import_batch() as metrics:
                    result = func(self, *args, **kwargs)
                logger.info(
                    "[excel_repo] financial write committed in one save: tables=%s saveSeconds=%s",
                    metrics.get("dirtyTables", 0),
                    metrics.get("saveSeconds", 0.0),
                )
                return result
            finally:
                if was_enabled:
                    gc.enable()
    return wrapper

import typing
if typing.TYPE_CHECKING:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils.cell import get_column_letter, range_boundaries
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.worksheet import Worksheet
else:
    Workbook = None
    
    MergedCell = None
    get_column_letter = None
    range_boundaries = None
    Table = None
    TableStyleInfo = None
    Worksheet = None

def _lazy_load_heavy_libs():
    global Workbook, load_workbook, MergedCell, get_column_letter, range_boundaries, Table, TableStyleInfo, Worksheet
    if Workbook is None:
        from openpyxl import Workbook, load_workbook
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils.cell import get_column_letter, range_boundaries
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.worksheet.worksheet import Worksheet


from uuid import uuid4

# Support direct execution via:
# python src/python/repositories/excel_repo.py
if __package__ in (None, ""):
    _THIS_FILE = Path(__file__).resolve()
    _PYTHON_ROOT = _THIS_FILE.parents[1]
    _PYTHON_ROOT_STR = str(_PYTHON_ROOT)
    if _PYTHON_ROOT_STR not in sys.path:
        sys.path.insert(0, _PYTHON_ROOT_STR)


from domain.money import calc_amounts, normalize_pct
from domain import schema_constants as sc
from services.paths import AppPaths

import logging
logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class TableRef:
    sheet: str
    table: str


TBL_PARENTS = TableRef(sc.SHEET_PARENTS, sc.TBL_PARENTS)
TBL_CLIENTS = TableRef(sc.SHEET_CLIENTS, sc.TBL_CLIENTS)
TBL_CLIENT_PROFILES = TableRef(sc.SHEET_CLIENT_PROFILES, sc.TBL_CLIENT_PROFILES)
TBL_MATTERS = TableRef(sc.SHEET_MATTERS, sc.TBL_MATTERS)
TBL_MATTER_PARTIES = TableRef(sc.SHEET_MATTER_PARTIES, sc.TBL_MATTER_PARTIES)
TBL_TIME = TableRef(sc.SHEET_TIME, sc.TBL_TIME)
TBL_TRADEMARKS = TableRef(sc.SHEET_TRADEMARKS, sc.TBL_TRADEMARKS)
TBL_DISBURSEMENTS = TableRef(sc.SHEET_DISBURSEMENTS, sc.TBL_DISBURSEMENTS)
TBL_LEDGER = TableRef(sc.SHEET_LEDGER, sc.TBL_LEDGER)
TBL_RECEIVABLES = TableRef(sc.SHEET_RECEIVABLES, sc.TBL_RECEIVABLES)
TBL_INVOICE_LOG = TableRef(sc.SHEET_INVOICE_LOG, sc.TBL_INVOICE_LOG)
TBL_TRANSACTIONS_MASTER = TableRef(sc.SHEET_TRANSACTIONS, sc.TBL_TRANSACTIONS_MASTER)
TBL_TRANSACTION_ACCOUNTS = TableRef(sc.SHEET_TRANSACTION_ACCOUNTS, sc.TBL_TRANSACTION_ACCOUNTS)
TBL_TRANSACTION_CATEGORIES = TableRef(sc.SHEET_TRANSACTION_CATEGORIES, sc.TBL_TRANSACTION_CATEGORIES)
TBL_TRANSACTION_BUSINESS_UNITS = TableRef(sc.SHEET_TRANSACTION_BUSINESS_UNITS, sc.TBL_TRANSACTION_BUSINESS_UNITS)
TBL_TRANSACTION_PAYEES = TableRef(sc.SHEET_TRANSACTION_PAYEES, sc.TBL_TRANSACTION_PAYEES)
TBL_DRAFT_INVOICES = TableRef(sc.SHEET_DRAFT_INVOICES, sc.TBL_DRAFT_INVOICES)
TBL_CORP_ENTITIES = TableRef(sc.SHEET_CORP_ENTITIES, sc.TBL_CORP_ENTITIES)
TBL_CORP_RELATIONSHIPS = TableRef(sc.SHEET_CORP_RELATIONSHIPS, sc.TBL_CORP_RELATIONSHIPS)
TBL_CORP_TRANSACTIONS = TableRef(sc.SHEET_CORP_TRANSACTIONS, sc.TBL_CORP_TRANSACTIONS)

TABLES_IN_ORDER = [
    TBL_PARENTS,
    TBL_CLIENTS,
    TBL_CLIENT_PROFILES,
    TBL_MATTERS,
    TBL_MATTER_PARTIES,
    TBL_TIME,
    TBL_TRADEMARKS,
    TBL_DISBURSEMENTS,
    TBL_LEDGER,
    TBL_RECEIVABLES,
    TBL_INVOICE_LOG,
    TBL_TRANSACTIONS_MASTER,
    TBL_TRANSACTION_ACCOUNTS,
    TBL_TRANSACTION_CATEGORIES,
    TBL_TRANSACTION_BUSINESS_UNITS,
    TBL_TRANSACTION_PAYEES,
    TBL_DRAFT_INVOICES,
    TBL_CORP_ENTITIES,
    TBL_CORP_RELATIONSHIPS,
    TBL_CORP_TRANSACTIONS,
]
TABLE_COLUMNS = sc.TABLE_COLUMNS
TABLE_ALIASES = sc.TABLE_ALIASES
TABLE_META_CACHE_SCHEMA_VERSION = 2


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_us_phone(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    digits = re.sub(r"\D+", "", text)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) != 10:
        return text
    return f"{digits[0:3]}-{digits[3:6]}-{digits[6:10]}"


def _is_individual_entity_type(value: Any) -> bool:
    return _clean_text(value).lower() == "individual"


def _split_person_name_parts(value: Any) -> Tuple[str, str, str]:
    text = re.sub(r"\s+", " ", _clean_text(value)).strip()
    if not text:
        return "", "", ""

    if "," in text:
        last_part, rest = text.split(",", 1)
        last_name = re.sub(r"\s+", " ", last_part).strip()
        given_parts = re.sub(r"\s+", " ", rest).strip().split()
        if given_parts and last_name:
            first_name = given_parts[0]
            middle_name = " ".join(given_parts[1:])
            return first_name, middle_name, last_name

    parts = text.split()
    if len(parts) == 1:
        return parts[0], "", ""
    if len(parts) == 2:
        return parts[0], "", parts[1]
    return parts[0], " ".join(parts[1:-1]), parts[-1]


def _join_person_name_parts(first_name: Any, middle_name: Any, last_name: Any) -> str:
    return " ".join(
        part
        for part in (
            _clean_text(first_name),
            _clean_text(middle_name),
            _clean_text(last_name),
        )
        if part
    )


def _looks_like_email_address(value: Any) -> bool:
    text = _clean_text(value)
    if not text:
        return True
    return bool(re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", text))


def _normalize_search_text(value: Any) -> str:
    text = _clean_text(value).lower()
    if not text:
        return ""
    return re.sub(r"\s+", " ", text)


def _report_name_keys(value: Any) -> set[str]:
    text = _clean_text(value).lower()
    if not text:
        return set()

    candidates = {text}
    if "," in text:
        last, first = [part.strip() for part in text.split(",", 1)]
        if first and last:
            candidates.add(f"{first} {last}")

    keys: set[str] = set()
    for candidate in candidates:
        normalized = re.sub(r"\bdoing business as\b", " dba ", candidate)
        normalized = re.sub(r"\bd/b/a\b", " dba ", normalized)
        normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
        normalized = re.sub(r"\s+", " ", normalized).strip()
        if not normalized:
            continue
        keys.add(normalized)
        if " dba " in normalized:
            before, after = normalized.split(" dba ", 1)
            if before:
                keys.add(before.strip())
            if after:
                keys.add(after.strip())
            keys.add(re.sub(r"\s+", " ", f"{before} {after}").strip())

    for key in list(keys):
        if "millennium" in key:
            keys.add(key.replace("millennium", "millenium"))
        if "millenium" in key:
            keys.add(key.replace("millenium", "millennium"))

    return {key for key in keys if key}


def _matter_token_fragment(value: Any, width: int = 3, fallback: str = "XXX") -> str:
    """Legacy compat shim – kept for any call-sites not yet migrated."""
    token = re.sub(r"[^A-Za-z0-9]+", "", _clean_text(value).upper())
    if not token:
        token = _clean_text(fallback).upper() or "X"
    if len(token) >= width:
        return token[:width]
    return (token + ("X" * width))[:width]


# ---------------------------------------------------------------------------
# Legacy matter-number helpers
# ---------------------------------------------------------------------------

# Hardcoded 3-char matter-type codes.  Keys are normalised lowercase.
_MATTER_TYPE_CODES: Dict[str, str] = {
    "trademark": "TMK",
    "trademarks": "TMK",
    "tmk": "TMK",
    "commercial": "COM",
    "com": "COM",
    "corporate": "CRP",
    "crp": "CRP",
    "corporation": "CRP",
    "corporatereorganization": "CRP",
    "incorporationorganization": "CRP",
    "corporatemaintenanceannualreturns": "CRP",
    "amalgamationdissolution": "CRP",
    "estate": "EST",
    "estates": "EST",
    "est": "EST",
    "tax": "TAX",
    "taxation": "TAX",
    "taxes": "TAX",
    "audit": "AUD",
    "audits": "AUD",
    "aud": "AUD",
    "income": "INC",
    "inc": "INC",
    "general": "GEN",
    "gen": "GEN",
}

# Words that should be stripped from the *end* of a company name before coding
_COMPANY_NOISE_SUFFIXES = {
    "inc", "ltd", "llc", "llp", "lp", "corp", "co", "plc", "pte", "pty",
    "sa", "srl", "sas", "gmbh", "ag", "nv", "bv", "ab",
}


def _legacy_matter_type_code(matter_type: Any) -> str:
    """Return the canonical 3-char matter-type code for the given string."""
    key = re.sub(r"[^a-z]", "", _clean_text(matter_type).lower())
    return _MATTER_TYPE_CODES.get(key, "GEN")


def _legacy_raw_client_code(name: Any, entity_type: Any = "") -> str:
    """
    Derive an *uncollision-checked* 4-character client code from *name*.

    Rules
    -----
    Individual  (EntityType == "Individual"):
        last[:4].  If len(last) < 4, pad with first[:needed].
    Company / anything else:
        Strip trailing noise suffixes, then take first-word[:4].
    In both cases, keep alphanumerics (including digits), pad with "X",
    return exactly 4 uppercase characters.
    """
    cleaned = re.sub(r"[^A-Za-z0-9 ]+", "", _clean_text(name)).strip()
    parts = cleaned.split()
    if not parts:
        return "XXXX"

    is_individual = _clean_text(entity_type).lower() == "individual"

    if is_individual:
        last = parts[-1].upper() if len(parts) >= 2 else parts[0].upper()
        raw = last[:4]
        if len(raw) < 4 and len(parts) >= 2:
            first = parts[0].upper()
            raw = raw + first[: 4 - len(raw)]
    else:
        # company: try the first meaningful word
        meaningful_parts = [p for p in parts if p.lower() not in _COMPANY_NOISE_SUFFIXES]
        main = (meaningful_parts[0] if meaningful_parts else parts[0]).upper()
        raw = main[:4]

    return (raw + "XXXX")[:4]


def _matter_year_two_digits(date_text: Any) -> str:
    text = _clean_text(date_text)
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", text)
    if match:
        return match.group(1)[2:]
    return datetime.now().strftime("%y")


def _normalize_choice(value: Any, allowed: List[str], default_value: str = "") -> str:
    normalized_allowed: Dict[str, str] = {
        _clean_text(choice).lower(): _clean_text(choice)
        for choice in allowed
        if _clean_text(choice)
    }
    if not normalized_allowed:
        return _clean_text(value) or _clean_text(default_value)
    raw = _clean_text(value)
    if not raw:
        return _clean_text(default_value) if _clean_text(default_value) else next(iter(normalized_allowed.values()))
    return normalized_allowed.get(raw.lower(), "")


def _is_valid_iso_date(text: Any) -> bool:
    value = _clean_text(text)
    if not value:
        return False
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", value):
        return False
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return True
    except Exception:
        return False


def _to_code_token(value: Any, fallback: str = "X") -> str:
    token = re.sub(r"[^A-Za-z0-9]+", "_", _clean_text(value).strip()).strip("_").upper()
    if token:
        return token
    return _clean_text(fallback).upper() or "X"


def _search_terms(value: Any) -> List[str]:
    normalized = _normalize_search_text(value)
    if not normalized:
        return []

    terms: List[str] = []
    seen = set()
    for raw_token in re.findall(r"[a-z0-9@._:/#\-]+", normalized, flags=re.IGNORECASE):
        token = raw_token.strip(" .,_-")
        if not token:
            continue
        if token in seen:
            continue
        seen.add(token)
        terms.append(token)

    if " " in normalized and normalized not in seen:
        terms.insert(0, normalized)
    return terms


def _boolean_query_tokens(value: Any) -> List[str]:
    raw = _clean_text(value)
    if not raw:
        return []
    return [tok for tok in re.findall(r"\(|\)|AND|OR|NOT|[^\s()]+", raw, flags=re.IGNORECASE) if tok]


def _boolean_to_postfix(tokens: List[str]) -> List[str]:
    precedence = {"OR": 1, "AND": 2, "NOT": 3}
    output: List[str] = []
    stack: List[str] = []

    for raw_tok in tokens:
        tok = str(raw_tok or "").strip()
        if not tok:
            continue
        upper = tok.upper()
        if upper in ("AND", "OR", "NOT"):
            if upper == "NOT":
                while stack and stack[-1] in precedence and precedence[stack[-1]] > precedence[upper]:
                    output.append(stack.pop())
            else:
                while stack and stack[-1] in precedence and precedence[stack[-1]] >= precedence[upper]:
                    output.append(stack.pop())
            stack.append(upper)
            continue
        if tok == "(":
            stack.append(tok)
            continue
        if tok == ")":
            while stack and stack[-1] != "(":
                output.append(stack.pop())
            if stack and stack[-1] == "(":
                stack.pop()
            continue
        output.append(_normalize_search_text(tok))

    while stack:
        top = stack.pop()
        if top in ("(", ")"):
            continue
        output.append(top)
    return output


def _match_boolean_postfix(postfix: List[str], haystack: str) -> bool:
    if not postfix:
        return False
    stack: List[bool] = []

    for tok in postfix:
        upper = tok.upper()
        if upper == "NOT":
            value = stack.pop() if stack else False
            stack.append(not value)
            continue
        if upper in ("AND", "OR"):
            right = stack.pop() if stack else False
            left = stack.pop() if stack else False
            stack.append(left and right if upper == "AND" else left or right)
            continue
        needle = _normalize_search_text(tok)
        stack.append(bool(needle) and needle in haystack)

    return stack[-1] if stack else False


def _normalize_index(value: Any, fallback: int) -> int:
    try:
        parsed = int(value)
        if parsed > 0:
            return parsed
    except Exception:
        pass
    return fallback


def _safe_range_boundaries(ref: str) -> Tuple[int, int, int, int]:
    _lazy_load_heavy_libs()
    min_col_raw, min_row_raw, max_col_raw, max_row_raw = range_boundaries(ref)
    min_col = _normalize_index(min_col_raw, 1)
    min_row = _normalize_index(min_row_raw, 1)
    max_col = _normalize_index(max_col_raw, min_col)
    max_row = _normalize_index(max_row_raw, min_row)
    if max_col < min_col:
        max_col = min_col
    if max_row < min_row:
        max_row = min_row
    return min_col, min_row, max_col, max_row


def _set_cell_value(ws: Worksheet, row_idx: int, col_idx: int, value: Any) -> None:
    _lazy_load_heavy_libs()
    cell_obj = ws.cell(row=row_idx, column=col_idx)
    if isinstance(cell_obj, MergedCell):
        return
    cell_obj.value = value


class ExcelRepo:
    def invalidate_after_external_workbook_replace(self) -> None:
        """Drop caches after an atomic workbook replacement by a helper process."""

        self._row_cache.clear()
        self._row_cache_mtime = 0.0
        self._ensured_schema_signature = ""
        self._invalidate_table_meta_cache()

    def _safe_save(self, workbook, filepath):
        """
        Anti-Corruption protocol for OneDrive/SharePoint collisions.
        Saves to a temporary file first, then atomically replaces the active file.
        """
        import os

        self._assert_write_permitted()

        target_path = Path(filepath)
        # A unique sibling temp file prevents a stale/interrupted save from
        # colliding with the next save attempt.  Keeping it beside the target
        # preserves the atomic-replace guarantee because both files are on the
        # same volume.
        temp_path = target_path.with_name(f".{target_path.name}.{uuid4().hex}.tmp")
        try:
            # 1. Save to an isolated temp file (OneDrive ignores this)
            workbook.save(str(temp_path))
            self._normalize_package_for_target_extension(temp_path, target_path)

            # openpyxl keeps the source archive open for keep_vba workbooks.
            # Close it before replacing the target so Windows does not deny
            # the rename because this process is still holding the old file.
            self._close_workbook(workbook)

            # 2. Atomically overwrite the real file
            import time
            retry_delays = (0.10, 0.15, 0.20, 0.30, 0.45, 0.65, 0.85, 1.0, 1.0, 1.0)
            last_error: Optional[BaseException] = None
            for attempt in range(len(retry_delays) + 1):
                try:
                    if os.path.exists(str(target_path)):
                        os.replace(str(temp_path), str(target_path))
                    else:
                        os.rename(str(temp_path), str(target_path))
                    break
                except (OSError, PermissionError) as e:
                    last_error = e
                    if attempt == len(retry_delays):
                        raise RuntimeError(
                            "CSPM could not replace its data workbook after several retries. "
                            "Close any copy of CSPM.xlsm open in Excel or another program, then retry the action. "
                            f"The change was not saved. {e}"
                        ) from e
                    time.sleep(retry_delays[attempt])
            
            # Invalidate memory cache so subsequent reads refetch
            self._row_cache.clear()
            self._row_cache_mtime = 0.0
            self._ensured_schema_signature = ""
            self._invalidate_table_meta_cache()
            self._persist_table_meta_cache()

        except Exception as e:
            logger.error("[excel_repo] _safe_save failed atomically: %s", e)
            raise
        finally:
            if os.path.exists(str(temp_path)):
                try:
                    os.remove(str(temp_path))
                except Exception:
                    pass

    def _normalize_package_for_target_extension(self, saved_path: Path, target_path: Path) -> None:
        if target_path.suffix.lower() != ".xlsm":
            return
        if not saved_path.exists() or not zipfile.is_zipfile(saved_path):
            return
        macro_content_type = "application/vnd.ms-excel.sheet.macroEnabled.main+xml"
        workbook_part = 'PartName="/xl/workbook.xml"'
        temp_zip_path = saved_path.with_name(saved_path.name + ".ziptmp")
        changed = False
        try:
            with zipfile.ZipFile(saved_path, "r") as src, zipfile.ZipFile(
                temp_zip_path, "w", compression=zipfile.ZIP_DEFLATED
            ) as dst:
                for info in src.infolist():
                    data = src.read(info.filename)
                    if info.filename == "[Content_Types].xml":
                        text = data.decode("utf-8")
                        next_text = re.sub(
                            rf'({re.escape(workbook_part)}\s+ContentType=")[^"]+(")',
                            rf"\1{macro_content_type}\2",
                            text,
                            count=1,
                        )
                        if next_text != text:
                            data = next_text.encode("utf-8")
                            changed = True
                    dst.writestr(info, data)
            if changed:
                max_retries = 10
                for attempt in range(max_retries):
                    try:
                        os.replace(str(temp_zip_path), str(saved_path))
                        break
                    except (OSError, PermissionError) as e:
                        if attempt == max_retries - 1:
                            logger.warning("[excel_repo] Unable to replace xlsm package after retries: %s", e)
                        import time
                        time.sleep(0.5)
            else:
                try:
                    temp_zip_path.unlink()
                except Exception:
                    pass
        except Exception as exc:
            logger.warning("[excel_repo] Unable to normalize xlsm package content type: %s", exc)
            try:
                if temp_zip_path.exists():
                    temp_zip_path.unlink()
            except Exception:
                pass

    """
    Canonical Excel adapter for the workbook contract in schema/workbook_schema.yml.
    """

    TXN_CLASS_OPTIONS = ["Family", "Business"]
    TXN_TYPE_OPTIONS = ["Income", "Expense", "Debt Repayment", "Transfer"]
    TXN_MEMBER_OPTIONS = ["Joint", "Deborah", "Cory", "Alexa", "Emma", "Maya"]
    TXN_TAX_FLAG_OPTIONS = ["None", "HST - Biz", "Business Deductible", "Medical", "Deductible"]
    TXN_STATUS_OPTIONS = ["Pending", "Cleared", "Reconciled", "Void"]
    TXN_CURRENCY_OPTIONS = ["CAD", "USD"]
    TXN_MOVEMENT_TYPES = {"transfer", "debt repayment"}
    TXN_DEBT_DEST_ACCOUNT_KINDS = {"credit-card", "loc", "mortgage", "loan"}
    TIME_STATUS_OPTIONS = ["Draft", "Ready for Billing", "Billed"]
    TM_JURISDICTION_OPTIONS = ["CIPO", "USPTO", "Other"]
    TM_MARK_TYPE_OPTIONS = ["Standard Character", "Design", "3D", "Sound", "Color", "Other"]

    def __init__(self, paths: AppPaths, write_guard: Optional[Callable[[], None]] = None):
        self.paths = paths
        # The application controller supplies this for shared-cloud sessions.
        # Isolated repair/candidate repositories intentionally leave it unset.
        self._write_guard = write_guard
        self._missing_tables = set()  # Priority 5: Cache missing table warnings
        import time
        self._time = time
        self._row_cache: Dict[str, List[Dict[str, Any]]] = {}
        self._row_cache_mtime: float = 0.0
        # ``ensure_schema`` is deliberately thorough because it protects the
        # workbook on first access.  Re-running that full macro-workbook scan
        # for every repository call, however, makes unrelated tab changes
        # needlessly slow.  This in-process signature records a verified
        # schema until CSPM.xlsm is changed or replaced.
        self._ensured_schema_signature: str = ""
        self._table_meta_cache_path: Path = self.paths.excel_metadata_cache_path()
        self._table_meta_cache_loaded: bool = False
        self._table_meta_cache_signature: str = ""
        self._table_meta_cache_tables: Dict[str, Dict[str, Any]] = {}
        self._table_meta_schema_requires_migration: Optional[bool] = None
        self._table_meta_cache_dirty: bool = False
        self._import_batch_active: bool = False
        self._import_batch_rows: Dict[str, List[Dict[str, Any]]] = {}
        self._import_batch_dirty_tables: Dict[str, TableRef] = {}
        # Dockets.xlsm remains the authoritative historical source for some
        # invoice-to-matter links that predate the normalized time table.
        # Keep the small, read-only projection in memory so opening a
        # statement does not re-open the legacy workbook for every refresh.
        self._legacy_docket_rows_cache: List[Dict[str, Any]] = []
        self._legacy_docket_rows_signature: str = ""

    def _assert_write_permitted(self) -> None:
        if self._write_guard is not None:
            self._write_guard()

    def _close_workbook(self, workbook) -> None:
        if workbook is None:
            return
        try:
            vba_archive = getattr(workbook, "vba_archive", None)
            if vba_archive is not None:
                try:
                    vba_archive.close()
                except Exception:
                    pass
                try:
                    workbook.vba_archive = None
                except Exception:
                    pass
        except Exception:
            pass
        try:
            archive = getattr(workbook, "_archive", None)
            if archive is not None:
                try:
                    archive.close()
                except Exception:
                    pass
                try:
                    workbook._archive = None
                except Exception:
                    pass
        except Exception:
            pass
        try:
            workbook.close()
        except Exception:
            pass

    def _table_cache_key(self, tref: TableRef) -> str:
        return f"{tref.sheet}::{tref.table}"

    def _workbook_signature(self, path: Path) -> str:
        try:
            stat = path.stat()
            return f"{int(stat.st_size)}:{int(stat.st_mtime_ns)}"
        except Exception:
            return ""

    def _load_table_meta_cache(self) -> None:
        if self._table_meta_cache_loaded:
            return
        self._table_meta_cache_loaded = True
        self._table_meta_cache_signature = ""
        self._table_meta_cache_tables = {}
        self._table_meta_schema_requires_migration = None
        self._table_meta_cache_dirty = False
        try:
            if not self._table_meta_cache_path.exists():
                return
            raw = json.loads(self._table_meta_cache_path.read_text(encoding="utf-8"))
            if not isinstance(raw, dict):
                return
            if int(raw.get("schemaVersion", 0) or 0) != TABLE_META_CACHE_SCHEMA_VERSION:
                return
            signature = str(raw.get("workbookSignature", "") or "")
            tables = raw.get("tables")
            if not isinstance(tables, dict):
                tables = {}
            cleaned_tables: Dict[str, Dict[str, Any]] = {}
            for key, value in tables.items():
                if not isinstance(key, str) or not isinstance(value, dict):
                    continue
                cleaned_tables[key] = dict(value)
            self._table_meta_cache_signature = signature
            self._table_meta_cache_tables = cleaned_tables
            if isinstance(raw.get("schemaRequiresMigration"), bool):
                self._table_meta_schema_requires_migration = bool(raw.get("schemaRequiresMigration"))
        except Exception:
            self._table_meta_cache_signature = ""
            self._table_meta_cache_tables = {}
            self._table_meta_schema_requires_migration = None
            self._table_meta_cache_dirty = False

    def _ensure_table_meta_signature(self, signature: str) -> None:
        self._load_table_meta_cache()
        if signature != self._table_meta_cache_signature:
            self._table_meta_cache_signature = signature
            self._table_meta_cache_tables = {}
            self._table_meta_schema_requires_migration = None
            self._table_meta_cache_dirty = True

    def _invalidate_table_meta_cache(self) -> None:
        self._table_meta_cache_signature = ""
        self._table_meta_cache_tables = {}
        self._table_meta_schema_requires_migration = None
        self._table_meta_cache_dirty = True

    def _persist_table_meta_cache(self) -> None:
        if not self._table_meta_cache_dirty:
            return
        self.paths.state_dir().mkdir(parents=True, exist_ok=True)
        payload = {
            "schemaVersion": TABLE_META_CACHE_SCHEMA_VERSION,
            "workbookSignature": self._table_meta_cache_signature,
            "tables": self._table_meta_cache_tables,
            "schemaRequiresMigration": self._table_meta_schema_requires_migration,
        }
        tmp_path = self._table_meta_cache_path.with_suffix(self._table_meta_cache_path.suffix + ".tmp")
        try:
            tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            os.replace(str(tmp_path), str(self._table_meta_cache_path))
            self._table_meta_cache_dirty = False
        except Exception:
            try:
                if tmp_path.exists():
                    tmp_path.unlink()
            except Exception:
                pass

    def _get_cached_table_meta(self, tref: TableRef, workbook_signature: str) -> Optional[Dict[str, Any]]:
        if not workbook_signature:
            return None
        self._load_table_meta_cache()
        if self._table_meta_cache_signature != workbook_signature:
            return None
        cached = self._table_meta_cache_tables.get(self._table_cache_key(tref))
        if not isinstance(cached, dict):
            return None
        return dict(cached)

    def _set_cached_table_meta(self, tref: TableRef, workbook_signature: str, payload: Dict[str, Any]) -> None:
        if not workbook_signature or not isinstance(payload, dict):
            return
        self._ensure_table_meta_signature(workbook_signature)
        cache_key = self._table_cache_key(tref)
        next_payload = dict(payload)
        if self._table_meta_cache_tables.get(cache_key) != next_payload:
            self._table_meta_cache_tables[cache_key] = next_payload
            self._table_meta_cache_dirty = True

    def _get_cached_schema_requires_migration(self, workbook_signature: str) -> Optional[bool]:
        if not workbook_signature:
            return None
        self._load_table_meta_cache()
        if self._table_meta_cache_signature != workbook_signature:
            return None
        return self._table_meta_schema_requires_migration

    def _set_cached_schema_requires_migration(self, workbook_signature: str, required: bool) -> None:
        if not workbook_signature:
            return
        self._ensure_table_meta_signature(workbook_signature)
        required_bool = bool(required)
        if self._table_meta_schema_requires_migration is not required_bool:
            self._table_meta_schema_requires_migration = required_bool
            self._table_meta_cache_dirty = True

    def warm_startup_metadata_cache(self) -> Dict[str, Any]:
        """Warm the data used by the first client, WIP, and invoice screens.

        This used to open the macro workbook once *per* table.  A user who
        opened Invoice Builder while that background task was running could
        therefore wait behind several full Excel parses.  The bulk reader
        takes one immutable snapshot instead, which also leaves a useful
        in-memory cache for the first preview.
        """
        tables = (
            TBL_CLIENTS,
            TBL_CLIENT_PROFILES,
            TBL_MATTERS,
            TBL_PARENTS,
            TBL_TIME,
            TBL_DRAFT_INVOICES,
            TBL_INVOICE_LOG,
        )
        try:
            rows_by_table = self._read_table_rows_bulk(list(tables))
        except Exception:
            logger.exception("Could not warm startup workbook tables.")
            return {"ok": False, "tablesWarmed": 0, "tablesFailed": len(tables)}

        warmed = sum(1 for tref in tables if tref.table in rows_by_table)
        failed = len(tables) - warmed
        return {"ok": failed == 0, "tablesWarmed": warmed, "tablesFailed": failed}

    @contextmanager
    def import_batch(self):
        """Defer validated import writes, save once, then verify every changed table."""

        if self._import_batch_active:
            raise RuntimeError("An Excel import batch is already active.")

        started = self._time.perf_counter()
        self.ensure_schema()
        loaded_rows = self._load_import_batch_rows()

        metrics: Dict[str, Any] = {
            "enabled": True,
            "tablesLoaded": len(loaded_rows),
            "rowsLoaded": sum(len(rows) for rows in loaded_rows.values()),
            "dirtyTables": 0,
            "rowsWritten": 0,
            "saveCount": 0,
            "postSaveVerified": False,
            "prepareSeconds": round(self._time.perf_counter() - started, 4),
            "saveSeconds": 0.0,
            "verifySeconds": 0.0,
            "totalSeconds": 0.0,
        }
        self._import_batch_active = True
        self._import_batch_rows = loaded_rows
        self._import_batch_dirty_tables = {}
        try:
            yield metrics
            self._flush_import_batch(metrics)
        finally:
            metrics["totalSeconds"] = round(self._time.perf_counter() - started, 4)
            self._import_batch_active = False
            self._import_batch_rows = {}
            self._import_batch_dirty_tables = {}

    def _load_import_batch_rows(self) -> Dict[str, List[Dict[str, Any]]]:
        _lazy_load_heavy_libs()
        path = self.paths.workbook_path()
        workbook = load_workbook(path, keep_vba=True, data_only=False)
        loaded_rows: Dict[str, List[Dict[str, Any]]] = {}
        try:
            for tref in TABLES_IN_ORDER:
                worksheet = workbook[tref.sheet]
                if not hasattr(worksheet, "tables") or tref.table not in worksheet.tables:
                    raise RuntimeError(f"Import batch could not find table {tref.table}.")
                _, raw_rows = self._rows_from_table(worksheet, worksheet.tables[tref.table])
                loaded_rows[self._table_cache_key(tref)] = [
                    self._canonical_batch_row(tref, row)
                    for row in raw_rows
                ]
        finally:
            self._close_workbook(workbook)
        return loaded_rows

    def _canonical_batch_row(self, tref: TableRef, row: Dict[str, Any]) -> Dict[str, Any]:
        canonical = self._canonicalize_row(tref, dict(row or {}))
        return {
            header: canonical.get(header, "")
            for header in TABLE_COLUMNS[tref.table]
        }

    def _batch_table_rows(self, tref: TableRef) -> List[Dict[str, Any]]:
        key = self._table_cache_key(tref)
        if key not in self._import_batch_rows:
            raise RuntimeError(f"Import batch table was not prepared: {tref.table}")
        return self._import_batch_rows[key]

    def _mark_import_batch_dirty(self, tref: TableRef) -> None:
        self._import_batch_dirty_tables[self._table_cache_key(tref)] = tref

    def _batch_append_row(self, tref: TableRef, row: Dict[str, Any]) -> None:
        self._batch_table_rows(tref).append(self._canonical_batch_row(tref, row))
        self._mark_import_batch_dirty(tref)

    def _batch_upsert_row(
        self,
        tref: TableRef,
        key_column: str,
        key_value: Any,
        row: Dict[str, Any],
    ) -> None:
        rows = self._batch_table_rows(tref)
        target_key = _clean_text(key_value).lower()
        canonical = self._canonical_batch_row(tref, row)
        for index, existing in enumerate(rows):
            if _clean_text(existing.get(key_column)).lower() == target_key:
                rows[index] = canonical
                self._mark_import_batch_dirty(tref)
                return
        rows.append(canonical)
        self._mark_import_batch_dirty(tref)

    def _batch_replace_rows(self, tref: TableRef, rows: List[Dict[str, Any]]) -> None:
        self._import_batch_rows[self._table_cache_key(tref)] = [
            self._canonical_batch_row(tref, row)
            for row in list(rows or [])
        ]
        self._mark_import_batch_dirty(tref)

    def _flush_import_batch(self, metrics: Dict[str, Any]) -> None:
        dirty_tables = [
            tref
            for tref in TABLES_IN_ORDER
            if self._table_cache_key(tref) in self._import_batch_dirty_tables
        ]
        metrics["dirtyTables"] = len(dirty_tables)
        metrics["rowsWritten"] = sum(
            len(self._batch_table_rows(tref))
            for tref in dirty_tables
        )
        if not dirty_tables:
            metrics["postSaveVerified"] = True
            return

        _lazy_load_heavy_libs()
        path = self.paths.workbook_path()
        save_started = self._time.perf_counter()
        workbook = load_workbook(path, keep_vba=True)
        try:
            for tref in dirty_tables:
                worksheet = workbook[tref.sheet]
                existing_table = (
                    worksheet.tables[tref.table]
                    if hasattr(worksheet, "tables") and tref.table in worksheet.tables
                    else None
                )
                style = existing_table.tableStyleInfo if existing_table is not None else None
                self._write_table(
                    worksheet,
                    existing_table=existing_table,
                    table_name=tref.table,
                    headers=TABLE_COLUMNS[tref.table],
                    rows=self._batch_table_rows(tref),
                    style=style,
                )
            self._safe_save(workbook, path)
            metrics["saveCount"] = 1
        finally:
            self._close_workbook(workbook)
        metrics["saveSeconds"] = round(self._time.perf_counter() - save_started, 4)

        verify_started = self._time.perf_counter()
        self._verify_import_batch_tables(path, dirty_tables)
        metrics["verifySeconds"] = round(self._time.perf_counter() - verify_started, 4)
        metrics["postSaveVerified"] = True

    def _verify_import_batch_tables(self, path: Path, dirty_tables: List[TableRef]) -> None:
        workbook = load_workbook(path, keep_vba=True, data_only=False)
        try:
            for tref in dirty_tables:
                worksheet = workbook[tref.sheet]
                if not hasattr(worksheet, "tables") or tref.table not in worksheet.tables:
                    raise RuntimeError(f"Post-save verification could not find table {tref.table}.")
                _, raw_rows = self._rows_from_table(worksheet, worksheet.tables[tref.table])
                actual_rows = [
                    self._canonical_batch_row(tref, row)
                    for row in raw_rows
                ]
                expected_rows = self._batch_table_rows(tref)
                actual_sig = self._batch_rows_signature(tref, actual_rows)
                expected_sig = self._batch_rows_signature(tref, expected_rows)
                if actual_sig != expected_sig:
                    import logging
                    log = logging.getLogger("app.import")
                    log.error(f"Mismatch in {tref.table}")
                    mismatches = []
                    if len(actual_sig) != len(expected_sig):
                        log.error(f"Length mismatch: {len(actual_sig)} vs {len(expected_sig)}")
                        mismatches.append({"index": -1, "expected": f"length {len(expected_sig)}", "actual": f"length {len(actual_sig)}"})
                    else:
                        for i, (act, exp) in enumerate(zip(actual_sig, expected_sig)):
                            if act != exp:
                                mismatches.append({"index": i, "expected": exp, "actual": act})
                    for m in mismatches:
                        log.error(f"Mismatch in {tref.table} Row {m['index']}:\nExpected: {m['expected']}\nActual:   {m['actual']}")
                    raise RuntimeError(f"Post-save verification failed for import batch table {tref.table}. Details: {mismatches[0] if mismatches else 'Length mismatch'}")
        finally:
            self._close_workbook(workbook)

    def _batch_rows_signature(
        self,
        tref: TableRef,
        rows: List[Dict[str, Any]],
    ) -> List[Tuple[str, ...]]:
        headers = TABLE_COLUMNS[tref.table]
        return [
            tuple(self._batch_value_signature(row.get(header)) for header in headers)
            for row in rows
        ]

    def _batch_value_signature(self, value: Any) -> str:
        if value in (None, ""):
            return ""
        if isinstance(value, datetime):
            return value.isoformat(sep=" ", timespec="microseconds")
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, bool):
            return "1" if value else "0"
        if isinstance(value, (int, float)):
            number = float(value)
            if math.isnan(number) or math.isinf(number):
                return str(number)
            if number == 0.0:
                value = 0.0
            return format(Decimal(str(value)).normalize(), "f")
        return str(value).strip()

    def ensure_schema(self) -> Dict[str, Any]:
        import time
        _lazy_load_heavy_libs()
        if self._import_batch_active:
            return {
                "createdWorkbook": False,
                "changed": False,
                "tableChanges": [],
                "seedChanges": [],
                "batchMode": True,
            }
        path = self.paths.workbook_path()
        path.parent.mkdir(parents=True, exist_ok=True)

        current_signature = self._workbook_signature(path)
        if current_signature and current_signature == self._ensured_schema_signature:
            return {
                "createdWorkbook": False,
                "changed": False,
                "tableChanges": [],
                "seedChanges": [],
                "cached": True,
            }

        created_workbook = False
        if not path.exists():
            wb = Workbook()
            created_workbook = True
        else:
            wb = load_workbook(path, keep_vba=True)
            time.sleep(0.02)  # Explicit GIL yield after massive C-extension XML parse

        changed = False
        table_changes: List[Dict[str, str]] = []
        try:
            for tref in TABLES_IN_ORDER:
                did_change, mode = self._ensure_sheet_schema(wb, tref)
                if did_change:
                    changed = True
                    table_changes.append({"table": tref.table, "mode": mode})
                
                # Relinquish Python interpreter lock momentarily to guarantee GUI framerate
                time.sleep(0.01)

            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                default_ws = wb["Sheet"]
                if int(default_ws.max_row or 0) <= 1 and int(default_ws.max_column or 0) <= 1:
                    if default_ws["A1"].value in (None, ""):
                        wb.remove(default_ws)
                        changed = True
                        table_changes.append({"table": "Sheet", "mode": "removed_default_sheet"})

            if changed:
                self._safe_save(wb, path)
        finally:
            self._close_workbook(wb)

        # Use the post-save signature when a repair changed the workbook.
        # A later external edit naturally has a different size/mtime
        # signature and forces a fresh validation.
        self._ensured_schema_signature = self._workbook_signature(path)

        return {
            "createdWorkbook": created_workbook,
            "changed": changed,
            "tableChanges": table_changes,
            "seedChanges": [],
        }

    def schema_requires_migration(self) -> bool:
        import time
        _lazy_load_heavy_libs()
        path = self.paths.workbook_path()
        if not path.exists():
            return False
        workbook_signature = self._workbook_signature(path)
        cached_required = self._get_cached_schema_requires_migration(workbook_signature)
        if isinstance(cached_required, bool):
            return cached_required
        required = False
        try:
            # Huge optimization: read_only=True skips the massive XML caching process
            # and drops the GIL immediately, preventing the UI from freezing.
            wb = load_workbook(path, read_only=True, keep_vba=False, data_only=True)
            time.sleep(0.01)  # Explicit GIL yield after XML parse
        except Exception:
            required = True
            self._set_cached_schema_requires_migration(workbook_signature, required)
            self._persist_table_meta_cache()
            return required
        try:
            for tref in TABLES_IN_ORDER:
                headers = TABLE_COLUMNS[tref.table]
                if tref.sheet not in wb.sheetnames:
                    required = True
                    break
                ws = wb[tref.sheet]
                header_map: Dict[str, int] = {}
                header_row = 0
                try:
                    if hasattr(ws, "tables") and ws.tables and tref.table in ws.tables:
                        table_obj = ws.tables[tref.table]
                        min_col, min_row, max_col, _ = _safe_range_boundaries(table_obj.ref)
                        header_row = min_row
                        for col_idx in range(min_col, max_col + 1):
                            label = _clean_text(ws.cell(row=min_row, column=col_idx).value)
                            if not label or label in header_map:
                                continue
                            header_map[label] = col_idx
                except Exception:
                    header_row = 0
                    header_map = {}
                if header_row <= 0:
                    header_row, header_map = self._locate_sheet_header_map(ws, tref, create_missing=False)
                if header_row <= 0:
                    required = True
                    break
                missing_required = [h for h in headers if h not in header_map]
                if missing_required:
                    required = True
                    break
                
                # Relinquish lock to unblock Qt GUI frame rendering
                time.sleep(0.01)
        finally:
            self._close_workbook(wb)
        self._set_cached_schema_requires_migration(workbook_signature, required)
        self._persist_table_meta_cache()
        return required

    def list_parent_names(self) -> List[str]:
        rows = self._read_table_rows(TBL_PARENTS)
        names = [_clean_text(r.get(sc.COL_PARENT_NAME, "")) for r in rows]
        return sorted([name for name in names if name])

    def list_client_names(self) -> List[str]:
        rows = [self._canonicalize_client_row(r) for r in self._read_table_rows(TBL_CLIENTS)]
        names = [_clean_text(r.get(sc.COL_CLIENT_NAME, "")) for r in rows]
        return sorted([name for name in names if name])

    def list_active_client_names(self) -> List[str]:
        rows = [self._canonicalize_client_row(r) for r in self._read_table_rows(TBL_CLIENTS)]
        names: List[str] = []
        for row in rows:
            if not self._is_client_row_active(row):
                continue
            name = _clean_text(row.get(sc.COL_CLIENT_NAME, ""))
            if name:
                names.append(name)
        return sorted(names)

    def list_client_directory(self) -> List[Dict[str, Any]]:
        def _build_search_text(*sources: Dict[str, Any]) -> str:
            chunks: List[str] = []
            for source in sources:
                if not source:
                    continue
                for value in source.values():
                    if value is None:
                        continue
                    text = _clean_text(value)
                    if text:
                        chunks.append(text)
            return " ".join(chunks)

        client_rows = [self._canonicalize_client_row(r) for r in self._read_table_rows(TBL_CLIENTS)]
        profile_rows = [self._canonicalize_client_profile_row(r) for r in self._read_table_rows(TBL_CLIENT_PROFILES)]

        profile_by_id: Dict[str, Dict[str, Any]] = {}
        for row in profile_rows:
            client_id = _clean_text(row.get(sc.COL_PROFILE_CLIENT_ID))
            if client_id:
                profile_by_id[client_id] = row

        directory: List[Dict[str, Any]] = []
        seen_ids = set()
        for client_row in client_rows:
            client_id = _clean_text(client_row.get(sc.COL_CLIENT_ID))
            if client_id:
                seen_ids.add(client_id.lower())
            profile_row = profile_by_id.get(client_id, {})
            status = _clean_text(client_row.get(sc.COL_CLIENT_STATUS)) or "Active"
            active_flag = 1 if self._is_client_row_active(client_row) else 0
            search_text = _build_search_text(client_row, profile_row)
            directory.append(
                {
                    "clientId": client_id,
                    "clientName": _clean_text(client_row.get(sc.COL_CLIENT_NAME)),
                    "displayName": _clean_text(profile_row.get(sc.COL_PROFILE_DISPLAY_NAME))
                    or _clean_text(client_row.get(sc.COL_CLIENT_NAME)),
                    "legalName": _clean_text(profile_row.get(sc.COL_PROFILE_LEGAL_NAME)),
                    "firstName": _clean_text(profile_row.get(sc.COL_PROFILE_FIRST_NAME)),
                    "middleName": _clean_text(profile_row.get(sc.COL_PROFILE_MIDDLE_NAME)),
                    "lastName": _clean_text(profile_row.get(sc.COL_PROFILE_LAST_NAME)),
                    "entityType": _clean_text(profile_row.get(sc.COL_PROFILE_ENTITY_TYPE)),
                    "status": status,
                    "active": active_flag,
                    "primaryEmail": _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_EMAIL))
                    or _clean_text(client_row.get(sc.COL_CLIENT_EMAIL)),
                    "primaryPhone": _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_PHONE))
                    or _clean_text(client_row.get(sc.COL_CLIENT_PHONE)),
                    "parentClientName": _clean_text(profile_row.get(sc.COL_PROFILE_PARENT_NAME)),
                    "onboardingStatus": _clean_text(profile_row.get(sc.COL_PROFILE_ONBOARDING_STATUS)),
                    "kycStatus": _clean_text(profile_row.get(sc.COL_PROFILE_KYC_STATUS)),
                    "updatedAt": _clean_text(profile_row.get(sc.COL_PROFILE_UPDATED)),
                    "searchText": search_text,
                }
            )

        # Include orphan profile rows if the profile exists but the client row is missing.
        for profile_row in profile_rows:
            client_id = _clean_text(profile_row.get(sc.COL_PROFILE_CLIENT_ID))
            if not client_id or client_id.lower() in seen_ids:
                continue
            search_text = _build_search_text(profile_row)
            directory.append(
                {
                    "clientId": client_id,
                    "clientName": _clean_text(profile_row.get(sc.COL_PROFILE_DISPLAY_NAME))
                    or _clean_text(profile_row.get(sc.COL_PROFILE_LEGAL_NAME)),
                    "displayName": _clean_text(profile_row.get(sc.COL_PROFILE_DISPLAY_NAME))
                    or _clean_text(profile_row.get(sc.COL_PROFILE_LEGAL_NAME)),
                    "legalName": _clean_text(profile_row.get(sc.COL_PROFILE_LEGAL_NAME)),
                    "firstName": _clean_text(profile_row.get(sc.COL_PROFILE_FIRST_NAME)),
                    "middleName": _clean_text(profile_row.get(sc.COL_PROFILE_MIDDLE_NAME)),
                    "lastName": _clean_text(profile_row.get(sc.COL_PROFILE_LAST_NAME)),
                    "entityType": _clean_text(profile_row.get(sc.COL_PROFILE_ENTITY_TYPE)),
                    "status": "Active",
                    "active": 1,
                    "primaryEmail": _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_EMAIL)),
                    "primaryPhone": _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_PHONE)),
                    "parentClientName": _clean_text(profile_row.get(sc.COL_PROFILE_PARENT_NAME)),
                    "onboardingStatus": _clean_text(profile_row.get(sc.COL_PROFILE_ONBOARDING_STATUS)),
                    "kycStatus": _clean_text(profile_row.get(sc.COL_PROFILE_KYC_STATUS)),
                    "updatedAt": _clean_text(profile_row.get(sc.COL_PROFILE_UPDATED)),
                    "searchText": search_text,
                }
            )

        directory.sort(
            key=lambda row: (
                _clean_text(row.get("displayName")).lower(),
                _clean_text(row.get("clientName")).lower(),
                _clean_text(row.get("clientId")).lower(),
            )
        )
        return directory

    def get_client_profile(self, client_key: str) -> Dict[str, Any]:
        lookup = _clean_text(client_key)
        if not lookup:
            return {"ok": False, "message": "Client key is required.", "client": {}}

        client_rows = [self._canonicalize_client_row(r) for r in self._read_table_rows(TBL_CLIENTS)]
        profile_rows = [self._canonicalize_client_profile_row(r) for r in self._read_table_rows(TBL_CLIENT_PROFILES)]
        profile_by_id: Dict[str, Dict[str, Any]] = {
            _clean_text(r.get(sc.COL_PROFILE_CLIENT_ID)): r
            for r in profile_rows
            if _clean_text(r.get(sc.COL_PROFILE_CLIENT_ID))
        }

        lookup_lower = lookup.lower()
        client_row: Optional[Dict[str, Any]] = None
        for row in client_rows:
            if _clean_text(row.get(sc.COL_CLIENT_ID)).lower() == lookup_lower:
                client_row = row
                break
        if client_row is None:
            for row in client_rows:
                if _clean_text(row.get(sc.COL_CLIENT_NAME)).lower() == lookup_lower:
                    client_row = row
                    break

        profile_row: Optional[Dict[str, Any]] = None
        if client_row is not None:
            profile_row = profile_by_id.get(_clean_text(client_row.get(sc.COL_CLIENT_ID)))

        if profile_row is None:
            for row in profile_rows:
                if _clean_text(row.get(sc.COL_PROFILE_CLIENT_ID)).lower() == lookup_lower:
                    profile_row = row
                    break
                if _clean_text(row.get(sc.COL_PROFILE_DISPLAY_NAME)).lower() == lookup_lower:
                    profile_row = row
                    break
                if _clean_text(row.get(sc.COL_PROFILE_LEGAL_NAME)).lower() == lookup_lower:
                    profile_row = row
                    break

        if client_row is None and profile_row is not None:
            profile_id = _clean_text(profile_row.get(sc.COL_PROFILE_CLIENT_ID))
            for row in client_rows:
                if _clean_text(row.get(sc.COL_CLIENT_ID)) == profile_id:
                    client_row = row
                    break

        if client_row is None and profile_row is None:
            return {"ok": False, "message": f"Client not found: {lookup}", "client": {}}

        client_id = _clean_text(
            (client_row or {}).get(sc.COL_CLIENT_ID) or (profile_row or {}).get(sc.COL_PROFILE_CLIENT_ID)
        )
        client_name = _clean_text((client_row or {}).get(sc.COL_CLIENT_NAME))
        display_name = _clean_text((profile_row or {}).get(sc.COL_PROFILE_DISPLAY_NAME)) or client_name
        legal_name = _clean_text((profile_row or {}).get(sc.COL_PROFILE_LEGAL_NAME)) or display_name or client_name
        status = _clean_text((client_row or {}).get(sc.COL_CLIENT_STATUS)) or "Active"
        active = 1 if self._is_client_row_active(client_row or {}) else 0

        full_address = _clean_text((profile_row or {}).get(sc.COL_PROFILE_FULL_ADDRESS))
        if not full_address:
            full_address = self._format_full_address(
                line1=_clean_text((profile_row or {}).get(sc.COL_PROFILE_ADDR1)),
                line2=_clean_text((profile_row or {}).get(sc.COL_PROFILE_ADDR2)),
                city=_clean_text((profile_row or {}).get(sc.COL_PROFILE_CITY)),
                state_province=_clean_text((profile_row or {}).get(sc.COL_PROFILE_STATE)),
                postal_code=_clean_text((profile_row or {}).get(sc.COL_PROFILE_POSTAL)),
                country=_clean_text((profile_row or {}).get(sc.COL_PROFILE_COUNTRY)),
            )

        client_payload = {
            "clientId": client_id,
            "clientName": client_name,
            "displayName": display_name,
            "legalName": legal_name,
            "status": status,
            "active": active,
            "firstName": _clean_text((profile_row or {}).get(sc.COL_PROFILE_FIRST_NAME)),
            "middleName": _clean_text((profile_row or {}).get(sc.COL_PROFILE_MIDDLE_NAME)),
            "lastName": _clean_text((profile_row or {}).get(sc.COL_PROFILE_LAST_NAME)),
            "entityType": _clean_text((profile_row or {}).get(sc.COL_PROFILE_ENTITY_TYPE)),
            "principalName": _clean_text((profile_row or {}).get(sc.COL_PROFILE_PRINCIPAL_NAME)),
            "principalPosition": _clean_text((profile_row or {}).get(sc.COL_PROFILE_PRINCIPAL_POSITION)),
            "primaryEmail": _clean_text((profile_row or {}).get(sc.COL_PROFILE_PRIMARY_EMAIL))
            or _clean_text((client_row or {}).get(sc.COL_CLIENT_EMAIL)),
            "primaryPhone": _clean_text((profile_row or {}).get(sc.COL_PROFILE_PRIMARY_PHONE))
            or _clean_text((client_row or {}).get(sc.COL_CLIENT_PHONE)),
            "secondaryContactName": _clean_text((profile_row or {}).get(sc.COL_PROFILE_SECONDARY_CONTACT)),
            "secondaryContactPosition": _clean_text((profile_row or {}).get(sc.COL_PROFILE_SECONDARY_POSITION)),
            "secondaryContactEmail": _clean_text((profile_row or {}).get(sc.COL_PROFILE_SECONDARY_EMAIL)),
            "secondaryContactPhone": _clean_text((profile_row or {}).get(sc.COL_PROFILE_SECONDARY_PHONE)),
            "addressLine1": _clean_text((profile_row or {}).get(sc.COL_PROFILE_ADDR1)),
            "addressLine2": _clean_text((profile_row or {}).get(sc.COL_PROFILE_ADDR2)),
            "city": _clean_text((profile_row or {}).get(sc.COL_PROFILE_CITY)),
            "stateProvince": _clean_text((profile_row or {}).get(sc.COL_PROFILE_STATE)),
            "postalCode": _clean_text((profile_row or {}).get(sc.COL_PROFILE_POSTAL)),
            "country": _clean_text((profile_row or {}).get(sc.COL_PROFILE_COUNTRY)),
            "fullAddress": full_address,
            "parentClientId": _clean_text((profile_row or {}).get(sc.COL_PROFILE_PARENT_ID)),
            "parentClientName": _clean_text((profile_row or {}).get(sc.COL_PROFILE_PARENT_NAME)),
            "website": _clean_text((profile_row or {}).get(sc.COL_PROFILE_WEBSITE)),
            "taxId": _clean_text((profile_row or {}).get(sc.COL_PROFILE_TAX_ID)),
            "industry": _clean_text((profile_row or {}).get(sc.COL_PROFILE_INDUSTRY)),
            "billingEmail": _clean_text((profile_row or {}).get(sc.COL_PROFILE_BILLING_EMAIL)),
            "kycStatus": _clean_text((profile_row or {}).get(sc.COL_PROFILE_KYC_STATUS)),
            "onboardingStatus": _clean_text((profile_row or {}).get(sc.COL_PROFILE_ONBOARDING_STATUS)),
            "retainerRequired": self._to_bool_int(
                (profile_row or {}).get(sc.COL_PROFILE_RETAINER_REQUIRED),
                default=0,
            ),
            "retainerAmount": round(
                float(self._parse_float((profile_row or {}).get(sc.COL_PROFILE_RETAINER_AMOUNT)) or 0.0),
                2,
            ),
            "engagementStartDate": _clean_text((profile_row or {}).get(sc.COL_PROFILE_ENGAGEMENT_START)),
            "dateClientAdded": _clean_text((profile_row or {}).get(sc.COL_PROFILE_DATE_CLIENT_ADDED)),
            "birthday": _clean_text((profile_row or {}).get(sc.COL_PROFILE_BIRTHDAY)),
            "referralFrom": _clean_text((profile_row or {}).get(sc.COL_PROFILE_REFERRAL_FROM)),
            "conflictNotes": _clean_text((profile_row or {}).get(sc.COL_PROFILE_CONFLICT_NOTES)),
            "notes": _clean_text((profile_row or {}).get(sc.COL_PROFILE_NOTES))
            or _clean_text((client_row or {}).get(sc.COL_CLIENT_NOTES)),
            "createdAt": _clean_text((profile_row or {}).get(sc.COL_PROFILE_CREATED)),
            "updatedAt": _clean_text((profile_row or {}).get(sc.COL_PROFILE_UPDATED)),
        }
        return {"ok": True, "message": "", "client": client_payload}

    def list_matter_names(self) -> List[str]:
        rows = [self._canonicalize_matter_row(r) for r in self._read_table_rows(TBL_MATTERS)]
        names = [_clean_text(r.get(sc.COL_MATTER_NAME, "")) for r in rows]
        return sorted([name for name in names if name])

    def list_active_matter_names(self) -> List[str]:
        rows = [self._canonicalize_matter_row(r) for r in self._read_table_rows(TBL_MATTERS)]
        names: List[str] = []
        for row in rows:
            if not self._is_matter_row_active(row):
                continue
            name = _clean_text(row.get(sc.COL_MATTER_NAME, ""))
            if name:
                names.append(name)
        return sorted(names)

    def list_matter_parties(self, matter_id: str) -> List[Dict[str, Any]]:
        """Return independent client records attached to one matter.

        This relationship deliberately lives at the matter level.  It never
        derives a family, household, or parent-client relationship between the
        records in the Client Directory.
        """
        target_id = _clean_text(matter_id).casefold()
        if not target_id:
            return []
        tables = self._read_table_rows_bulk([TBL_MATTER_PARTIES, TBL_CLIENT_PROFILES])
        profiles = {
            _clean_text(row.get(sc.COL_PROFILE_CLIENT_ID)).casefold(): self._canonicalize_client_profile_row(row)
            for row in tables.get(TBL_CLIENT_PROFILES.table, [])
            if _clean_text(row.get(sc.COL_PROFILE_CLIENT_ID))
        }
        parties: List[Dict[str, Any]] = []
        for raw in tables.get(TBL_MATTER_PARTIES.table, []):
            party = self._canonicalize_matter_party_row(raw)
            if _clean_text(party.get(sc.COL_MATTER_PARTY_MATTER_ID)).casefold() != target_id:
                continue
            client_id = _clean_text(party.get(sc.COL_MATTER_PARTY_CLIENT_ID))
            profile = profiles.get(client_id.casefold(), {})
            display_name = (
                _clean_text(profile.get(sc.COL_PROFILE_DISPLAY_NAME))
                or _clean_text(profile.get(sc.COL_PROFILE_LEGAL_NAME))
                or _clean_text(party.get(sc.COL_MATTER_PARTY_CLIENT_NAME))
            )
            full_address = _clean_text(profile.get(sc.COL_PROFILE_FULL_ADDRESS)) or self._format_full_address(
                line1=_clean_text(profile.get(sc.COL_PROFILE_ADDR1)),
                line2=_clean_text(profile.get(sc.COL_PROFILE_ADDR2)),
                city=_clean_text(profile.get(sc.COL_PROFILE_CITY)),
                state_province=_clean_text(profile.get(sc.COL_PROFILE_STATE)),
                postal_code=_clean_text(profile.get(sc.COL_PROFILE_POSTAL)),
                country=_clean_text(profile.get(sc.COL_PROFILE_COUNTRY)),
            )
            parties.append({
                "matterPartyId": _clean_text(party.get(sc.COL_MATTER_PARTY_ID)),
                "matterId": _clean_text(party.get(sc.COL_MATTER_PARTY_MATTER_ID)),
                "clientId": client_id,
                "clientName": display_name,
                "role": _clean_text(party.get(sc.COL_MATTER_PARTY_ROLE)),
                "isFileAnchor": self._to_bool_int(party.get(sc.COL_MATTER_PARTY_IS_FILE_ANCHOR), default=0),
                "isBillingRecipient": self._to_bool_int(
                    party.get(sc.COL_MATTER_PARTY_IS_BILLING_RECIPIENT), default=1
                ),
                "sortOrder": int(self._parse_int(party.get(sc.COL_MATTER_PARTY_SORT_ORDER)) or 0),
                "notes": _clean_text(party.get(sc.COL_MATTER_PARTY_NOTES)),
                "primaryEmail": _clean_text(profile.get(sc.COL_PROFILE_PRIMARY_EMAIL)),
                "billingEmail": _clean_text(profile.get(sc.COL_PROFILE_BILLING_EMAIL)),
                "fullAddress": full_address,
            })
        parties.sort(key=lambda party: (int(party.get("sortOrder", 0) or 0), party["clientName"].casefold()))
        return parties

    def list_invoice_bill_to_options(self, work_entry_ids: List[str]) -> List[Dict[str, Any]]:
        """Return eligible bill-to clients when selected WIP includes a joint matter.

        An empty result preserves the established single-client invoice flow.
        The invoice service repeats this lookup before saving, so a caller
        cannot bypass a required bill-to choice by calling the draft API.
        """
        selected_ids = {_clean_text(value) for value in list(work_entry_ids or []) if _clean_text(value)}
        if not selected_ids:
            return []
        tables = self._read_table_rows_bulk([TBL_TIME, TBL_DISBURSEMENTS])
        matter_ids: set[str] = set()
        for row in tables.get(TBL_TIME.table, []):
            if _clean_text(row.get(sc.COL_TIME_ENTRY_ID)) in selected_ids:
                matter_id = _clean_text(row.get(sc.COL_TIME_MATTER_ID))
                if matter_id:
                    matter_ids.add(matter_id)
        for row in tables.get(TBL_DISBURSEMENTS.table, []):
            if _clean_text(row.get(sc.COL_DISB_ID)) in selected_ids:
                matter_id = _clean_text(row.get(sc.COL_DISB_MATTER_ID))
                if matter_id:
                    matter_ids.add(matter_id)
        if not matter_ids:
            return []

        common_recipient_ids: Optional[set[str]] = None
        candidate_by_id: Dict[str, Dict[str, Any]] = {}
        for matter_id in sorted(matter_ids):
            parties = self.list_matter_parties(matter_id)
            # Existing single-client matters and joint records which explicitly
            # nominate one invoice recipient do not need a prompt.
            eligible = [party for party in parties if int(party.get("isBillingRecipient", 0) or 0) == 1]
            if len(eligible) <= 1:
                continue
            recipient_ids = {
                _clean_text(party.get("clientId")).casefold()
                for party in eligible
                if _clean_text(party.get("clientId"))
            }
            if len(recipient_ids) <= 1:
                continue
            for party in eligible:
                party_id = _clean_text(party.get("clientId")).casefold()
                if party_id:
                    candidate_by_id[party_id] = dict(party)
            common_recipient_ids = recipient_ids if common_recipient_ids is None else common_recipient_ids & recipient_ids

        if common_recipient_ids is None:
            return []
        if not common_recipient_ids:
            raise ValueError(
                "The selected joint matters have no common invoice recipient. Create separate drafts or update the matter parties."
            )
        result = [candidate_by_id[client_id] for client_id in common_recipient_ids if client_id in candidate_by_id]
        result.sort(key=lambda party: (party.get("sortOrder", 0), party.get("clientName", "").casefold()))
        return result


    def preview_matter_number(
        self,
        client_name: str,
        matter_type: str,
        date_opened: str,
        existing_matter_id: str = "",
    ) -> str:
        try:
            return self._build_matter_number(
                client_name=client_name,
                matter_type=matter_type,
                date_opened=date_opened,
                existing_matter_id=existing_matter_id,
                parent_name="",
                entity_type="",
            )
        except Exception:
            return ""

    def list_matter_directory(self) -> List[Dict[str, Any]]:
        def _build_search_text(*sources: Dict[str, Any]) -> str:
            chunks: List[str] = []
            for source in sources:
                if not source:
                    continue
                for value in source.values():
                    if value is None:
                        continue
                    text = _clean_text(value)
                    if text:
                        chunks.append(text)
            return " ".join(chunks)

        matter_rows = [self._canonicalize_matter_row(r) for r in self._read_table_rows(TBL_MATTERS)]
        client_rows = [self._canonicalize_client_row(r) for r in self._read_table_rows(TBL_CLIENTS)]
        parent_rows = [self._canonicalize_parent_row(r) for r in self._read_table_rows(TBL_PARENTS)]

        client_name_by_id: Dict[str, str] = {}
        for row in client_rows:
            cid = _clean_text(row.get(sc.COL_CLIENT_ID))
            cname = _clean_text(row.get(sc.COL_CLIENT_NAME))
            if cid and cname:
                client_name_by_id[cid] = cname

        parent_name_by_id: Dict[str, str] = {}
        for row in parent_rows:
            pid = _clean_text(row.get(sc.COL_PARENT_ID))
            pname = _clean_text(row.get(sc.COL_PARENT_NAME))
            if pid and pname:
                parent_name_by_id[pid] = pname

        directory: List[Dict[str, Any]] = []
        for matter_row in matter_rows:
            matter_id = _clean_text(matter_row.get(sc.COL_MATTER_ID))
            matter_number = _clean_text(matter_row.get(sc.COL_MATTER_NUMBER))
            matter_name = _clean_text(matter_row.get(sc.COL_MATTER_NAME))
            display_name = _clean_text(matter_row.get(sc.COL_MATTER_DISPLAY_NAME)) or matter_name
            client_id = _clean_text(matter_row.get(sc.COL_MATTER_CLIENT_ID))
            parent_id = _clean_text(matter_row.get(sc.COL_MATTER_PARENT_ID))
            client_name = _clean_text(matter_row.get(sc.COL_MATTER_CLIENT_NAME)) or client_name_by_id.get(
                client_id, ""
            )
            parent_name = _clean_text(matter_row.get(sc.COL_MATTER_PARENT_NAME)) or parent_name_by_id.get(
                parent_id, ""
            )
            status = _clean_text(matter_row.get(sc.COL_MATTER_STATUS)) or "Open"
            active = 1 if self._is_matter_row_active(matter_row) else 0
            search_text = _build_search_text(
                matter_row,
                {"clientName": client_name, "parentName": parent_name},
            )
            directory.append(
                {
                    "matterId": matter_id,
                    "matterNumber": _clean_text(matter_row.get(sc.COL_MATTER_NUMBER)),
                    "matterName": matter_name,
                    "displayName": display_name,
                    "clientId": client_id,
                    "clientName": client_name,
                    "parentId": parent_id,
                    "parentName": parent_name,
                    "matterType": _clean_text(matter_row.get(sc.COL_MATTER_TYPE)),
                    "practiceArea": _clean_text(matter_row.get(sc.COL_MATTER_PRACTICE_AREA)),
                    "responsibleLawyer": _clean_text(matter_row.get(sc.COL_MATTER_RESPONSIBLE_LAWYER)),
                    "billingArrangement": _clean_text(
                        matter_row.get(sc.COL_MATTER_BILLING_ARRANGEMENT)
                    ),
                    "status": status,
                    "active": active,
                    "dateOpened": _clean_text(matter_row.get(sc.COL_MATTER_OPEN_DATE)),
                    "updatedAt": _clean_text(matter_row.get(sc.COL_MATTER_UPDATED)),
                    "searchText": search_text,
                }
            )

        directory.sort(
            key=lambda row: (
                _clean_text(row.get("displayName")).lower(),
                _clean_text(row.get("matterName")).lower(),
                _clean_text(row.get("matterId")).lower(),
            )
        )
        return directory

    def list_active_matter_directory(self) -> List[Dict[str, Any]]:
        # reuse existing directory builder and filter by active flag
        all_rows = self.list_matter_directory()
        return [r for r in all_rows if r.get("active") == 1]

    def get_matter_profile(self, matter_key: str) -> Dict[str, Any]:
        lookup = _clean_text(matter_key)
        if not lookup:
            return {"ok": False, "message": "Matter key is required.", "matter": {}}

        matter_rows = [self._canonicalize_matter_row(r) for r in self._read_table_rows(TBL_MATTERS)]
        client_rows = [self._canonicalize_client_row(r) for r in self._read_table_rows(TBL_CLIENTS)]
        parent_rows = [self._canonicalize_parent_row(r) for r in self._read_table_rows(TBL_PARENTS)]

        client_name_by_id: Dict[str, str] = {}
        for row in client_rows:
            cid = _clean_text(row.get(sc.COL_CLIENT_ID))
            cname = _clean_text(row.get(sc.COL_CLIENT_NAME))
            if cid and cname:
                client_name_by_id[cid] = cname

        parent_name_by_id: Dict[str, str] = {}
        for row in parent_rows:
            pid = _clean_text(row.get(sc.COL_PARENT_ID))
            pname = _clean_text(row.get(sc.COL_PARENT_NAME))
            if pid and pname:
                parent_name_by_id[pid] = pname

        lookup_lc = lookup.lower()
        selected: Optional[Dict[str, Any]] = None
        for row in matter_rows:
            if _clean_text(row.get(sc.COL_MATTER_ID)).lower() == lookup_lc:
                selected = row
                break
        if selected is None:
            for row in matter_rows:
                if _clean_text(row.get(sc.COL_MATTER_NAME)).lower() == lookup_lc:
                    selected = row
                    break
        if selected is None:
            for row in matter_rows:
                if _clean_text(row.get(sc.COL_MATTER_DISPLAY_NAME)).lower() == lookup_lc:
                    selected = row
                    break
        if selected is None:
            for row in matter_rows:
                if _clean_text(row.get(sc.COL_MATTER_COURT_FILE_NO)).lower() == lookup_lc:
                    selected = row
                    break
        if selected is None:
            for row in matter_rows:
                if _clean_text(row.get(sc.COL_MATTER_NUMBER)).lower() == lookup_lc:
                    selected = row
                    break

        if selected is None:
            return {"ok": False, "message": f"Matter not found: {lookup}", "matter": {}}

        client_id = _clean_text(selected.get(sc.COL_MATTER_CLIENT_ID))
        parent_id = _clean_text(selected.get(sc.COL_MATTER_PARENT_ID))
        client_name = _clean_text(selected.get(sc.COL_MATTER_CLIENT_NAME)) or client_name_by_id.get(client_id, "")
        parent_name = _clean_text(selected.get(sc.COL_MATTER_PARENT_NAME)) or parent_name_by_id.get(parent_id, "")

        matter_payload = {
            "matterId": _clean_text(selected.get(sc.COL_MATTER_ID)),
            "matterNumber": _clean_text(selected.get(sc.COL_MATTER_NUMBER)),
            "matterName": _clean_text(selected.get(sc.COL_MATTER_NAME)),
            "displayName": _clean_text(selected.get(sc.COL_MATTER_DISPLAY_NAME))
            or _clean_text(selected.get(sc.COL_MATTER_NAME)),
            "clientId": client_id,
            "clientName": client_name,
            "parentId": parent_id,
            "parentName": parent_name,
            "matterType": _clean_text(selected.get(sc.COL_MATTER_TYPE)),
            "practiceArea": _clean_text(selected.get(sc.COL_MATTER_PRACTICE_AREA)),
            "status": _clean_text(selected.get(sc.COL_MATTER_STATUS)) or "Open",
            "responsibleLawyer": _clean_text(selected.get(sc.COL_MATTER_RESPONSIBLE_LAWYER)),
            "billingArrangement": _clean_text(selected.get(sc.COL_MATTER_BILLING_ARRANGEMENT)),
            "billingContact": _clean_text(selected.get(sc.COL_MATTER_BILLING_CONTACT)),
            "billingEmail": _clean_text(selected.get(sc.COL_MATTER_BILLING_EMAIL)),
            "defaultRate": round(float(self._parse_float(selected.get(sc.COL_MATTER_DEF_RATE)) or 0.0), 2),
            "defaultSharePct": round(
                normalize_pct(selected.get(sc.COL_MATTER_DEF_SHARE), default_pct=100.0),
                2,
            ),
            "dateOfEngagement": _clean_text(selected.get(sc.COL_MATTER_ENGAGEMENT_DATE)),
            "dateOpened": _clean_text(selected.get(sc.COL_MATTER_OPEN_DATE)),
            "dateClosed": _clean_text(selected.get(sc.COL_MATTER_CLOSE_DATE)),
            "courtFileNumber": _clean_text(selected.get(sc.COL_MATTER_COURT_FILE_NO)),
            "opposingParty": _clean_text(selected.get(sc.COL_MATTER_OPPOSING_PARTY)),
            "referralFrom": _clean_text(selected.get(sc.COL_MATTER_REFERRAL_FROM)),
            "description": _clean_text(selected.get(sc.COL_MATTER_DESCRIPTION)),
            "notes": _clean_text(selected.get(sc.COL_MATTER_NOTES)),
            "representationMode": _clean_text(selected.get(sc.COL_MATTER_REPRESENTATION_MODE)) or "Single Client",
            "jointNoConfidentialityConfirmed": self._to_bool_int(
                selected.get(sc.COL_MATTER_JOINT_NO_CONFIDENTIALITY_CONFIRMED), default=0
            ),
            "jointInstructionsRequireAll": self._to_bool_int(
                selected.get(sc.COL_MATTER_JOINT_INSTRUCTIONS_REQUIRE_ALL), default=0
            ),
            "jointEngagementDocument": _clean_text(selected.get(sc.COL_MATTER_JOINT_ENGAGEMENT_DOCUMENT)),
            "parties": self.list_matter_parties(_clean_text(selected.get(sc.COL_MATTER_ID))),
            "active": 1 if self._is_matter_row_active(selected) else 0,
            "createdAt": _clean_text(selected.get(sc.COL_MATTER_CREATED)),
            "updatedAt": _clean_text(selected.get(sc.COL_MATTER_UPDATED)),
        }
        matter_payload["isJointRetainer"] = (
            matter_payload["representationMode"].casefold() == "joint retainer"
            or len(matter_payload["parties"]) > 1
        )
        return {"ok": True, "message": "", "matter": matter_payload}

    def list_transaction_accounts(self, include_inactive: bool = False) -> List[Dict[str, Any]]:
        rows = [self._canonicalize_transaction_account_row(r) for r in self._read_table_rows(TBL_TRANSACTION_ACCOUNTS)]
        if not rows:
            seed_rows = self._load_seed_csv_rows("transactions_master.accounts.seed.csv")
            rows = [self._canonicalize_transaction_account_row(r) for r in seed_rows]
        out: List[Dict[str, Any]] = []
        for row in rows:
            active = self._to_bool_int(row.get(sc.COL_TXN_ACCOUNT_ACTIVE), default=1)
            if not include_inactive and active != 1:
                continue
            alias_text = _clean_text(row.get(sc.COL_TXN_ACCOUNT_ALIASES))
            aliases = [a.strip() for a in alias_text.split(",") if a.strip()]
            out.append(
                {
                    "accountCode": _clean_text(row.get(sc.COL_TXN_ACCOUNT_CODE)),
                    "accountName": _clean_text(row.get(sc.COL_TXN_ACCOUNT_NAME)),
                    "accountKind": _clean_text(row.get(sc.COL_TXN_ACCOUNT_KIND)),
                    "owner": _clean_text(row.get(sc.COL_TXN_ACCOUNT_OWNER)),
                    "active": active,
                    "aliases": aliases,
                }
            )
        out.sort(
            key=lambda row: (
                _normalize_search_text(row.get("accountName")),
                _normalize_search_text(row.get("accountCode")),
            )
        )
        return out

    def list_transaction_categories(
        self,
        txn_type: str = "",
        txn_class: str = "",
        include_inactive: bool = False,
    ) -> List[Dict[str, Any]]:
        type_filter = _normalize_search_text(txn_type)
        class_filter = _normalize_search_text(txn_class)
        rows = [
            self._canonicalize_transaction_category_row(r)
            for r in self._read_table_rows(TBL_TRANSACTION_CATEGORIES)
        ]
        if not rows:
            seed_rows = self._load_seed_csv_rows("transactions_master.categories.seed.csv")
            rows = [self._canonicalize_transaction_category_row(r) for r in seed_rows]
        out: List[Dict[str, Any]] = []
        for row in rows:
            active = self._to_bool_int(row.get(sc.COL_TXN_CATEGORY_LKP_ACTIVE), default=1)
            if not include_inactive and active != 1:
                continue
            row_type = _clean_text(row.get(sc.COL_TXN_CATEGORY_LKP_TYPE))
            row_scope = _clean_text(row.get(sc.COL_TXN_CATEGORY_LKP_CLASS_SCOPE))
            if type_filter and not self._txn_scope_match(row_type, type_filter):
                continue
            if class_filter and not self._txn_scope_match(row_scope, class_filter):
                continue
            out.append(
                {
                    "categoryCode": _clean_text(row.get(sc.COL_TXN_CATEGORY_LKP_CODE)),
                    "categoryName": _clean_text(row.get(sc.COL_TXN_CATEGORY_LKP_NAME)),
                    "type": row_type,
                    "classScope": row_scope,
                    "taxFlagDefault": _clean_text(row.get(sc.COL_TXN_CATEGORY_LKP_TAX_FLAG_DEFAULT)),
                    "billableAllowed": self._to_bool_int(
                        row.get(sc.COL_TXN_CATEGORY_LKP_BILLABLE_ALLOWED), default=0
                    ),
                    "medicalEligible": self._to_bool_int(
                        row.get(sc.COL_TXN_CATEGORY_LKP_MEDICAL_ELIGIBLE), default=0
                    ),
                    "deductibleEligible": self._to_bool_int(
                        row.get(sc.COL_TXN_CATEGORY_LKP_DEDUCTIBLE_ELIGIBLE), default=0
                    ),
                    "businessDeductibleEligible": self._to_bool_int(
                        row.get(sc.COL_TXN_CATEGORY_LKP_BUSINESS_DEDUCTIBLE_ELIGIBLE), default=0
                    ),
                    "active": active,
                    "sortOrder": self._parse_int(row.get(sc.COL_TXN_CATEGORY_LKP_SORT_ORDER)) or 0,
                    "notes": _clean_text(row.get(sc.COL_TXN_CATEGORY_LKP_NOTES)),
                }
            )
        out.sort(
            key=lambda row: (
                int(row.get("sortOrder", 0)),
                _normalize_search_text(row.get("categoryName")),
                _normalize_search_text(row.get("categoryCode")),
            )
        )
        return out

    def list_transaction_business_units(self, include_inactive: bool = False) -> List[Dict[str, Any]]:
        rows = [
            self._canonicalize_transaction_business_unit_row(r)
            for r in self._read_table_rows(TBL_TRANSACTION_BUSINESS_UNITS)
        ]
        if not rows:
            seed_rows = self._load_seed_csv_rows("transactions_master.business_units.seed.csv")
            rows = [self._canonicalize_transaction_business_unit_row(r) for r in seed_rows]
        out: List[Dict[str, Any]] = []
        for row in rows:
            active = self._to_bool_int(row.get(sc.COL_TXN_BUSINESS_UNIT_ACTIVE), default=1)
            if not include_inactive and active != 1:
                continue
            out.append(
                {
                    "businessUnit": _clean_text(row.get(sc.COL_TXN_BUSINESS_UNIT_NAME)),
                    "owner": _clean_text(row.get(sc.COL_TXN_BUSINESS_UNIT_OWNER)),
                    "active": active,
                }
            )
        out.sort(key=lambda row: _normalize_search_text(row.get("businessUnit")))
        return out

    def list_transaction_payees(self, include_inactive: bool = False) -> List[Dict[str, Any]]:
        rows = [self._canonicalize_transaction_payee_row(r) for r in self._read_table_rows(TBL_TRANSACTION_PAYEES)]
        if not rows:
            seed_rows = self._load_seed_csv_rows("transactions_master.payees.seed.csv")
            rows = [self._canonicalize_transaction_payee_row(r) for r in seed_rows]
        out: List[Dict[str, Any]] = []
        for row in rows:
            active = self._to_bool_int(row.get(sc.COL_TXN_PAYEE_ACTIVE), default=1)
            if not include_inactive and active != 1:
                continue
            out.append(
                {
                    "payeeName": _clean_text(row.get(sc.COL_TXN_PAYEE_NAME)),
                    "defaultCategoryCode": _clean_text(row.get(sc.COL_TXN_PAYEE_DEFAULT_CATEGORY_CODE)),
                    "active": active,
                }
            )
        out.sort(key=lambda row: _normalize_search_text(row.get("payeeName")))
        return out

    def list_transactions(self, filters: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        f = dict(filters or {})
        query = _normalize_search_text(f.get("query", ""))
        class_filter = _normalize_search_text(f.get("class") or f.get("txnClass"))
        type_filter = _normalize_search_text(f.get("type") or f.get("txnType"))
        status_filter = _normalize_search_text(f.get("status"))
        business_unit_filter = _normalize_search_text(f.get("businessUnit"))
        parent_filter = _normalize_search_text(f.get("parent"))
        client_filter = _normalize_search_text(f.get("client"))
        matter_filter = _normalize_search_text(f.get("matter"))
        include_void = self._to_bool_int(f.get("includeVoid"), default=1) == 1

        date_from = _clean_text(
            f.get("dateFrom") or f.get("fromDate") or f.get("startDate")
        )
        date_to = _clean_text(
            f.get("dateTo") or f.get("toDate") or f.get("endDate")
        )
        if date_from and not _is_valid_iso_date(date_from):
            raise ValueError("dateFrom must be in YYYY-MM-DD format.")
        if date_to and not _is_valid_iso_date(date_to):
            raise ValueError("dateTo must be in YYYY-MM-DD format.")

        rows = [
            self._canonicalize_transaction_row(r)
            for r in self._read_table_rows(TBL_TRANSACTIONS_MASTER)
        ]
        out: List[Dict[str, Any]] = []
        for row in rows:
            row_date = _clean_text(row.get(sc.COL_TXN_DATE))
            row_class = _normalize_search_text(row.get(sc.COL_TXN_CLASS))
            row_type = _normalize_search_text(row.get(sc.COL_TXN_TYPE))
            row_status = _normalize_search_text(row.get(sc.COL_TXN_STATUS))
            row_business_unit = _normalize_search_text(row.get(sc.COL_TXN_BUSINESS_UNIT))
            row_parent = _normalize_search_text(row.get(sc.COL_TXN_PARENT))
            row_client = _normalize_search_text(row.get(sc.COL_TXN_CLIENT))
            row_matter = _normalize_search_text(row.get(sc.COL_TXN_MATTER))

            if class_filter and row_class != class_filter:
                continue
            if type_filter and row_type != type_filter:
                continue
            if status_filter and row_status != status_filter:
                continue
            if business_unit_filter and business_unit_filter not in row_business_unit:
                continue
            if parent_filter and parent_filter not in row_parent:
                continue
            if client_filter and client_filter not in row_client:
                continue
            if matter_filter and matter_filter not in row_matter:
                continue
            if not include_void and row_status == "void":
                continue

            if date_from and row_date and row_date < date_from:
                continue
            if date_to and row_date and row_date > date_to:
                continue

            if query:
                haystack = " | ".join(_normalize_search_text(v) for v in row.values() if _clean_text(v))
                if query not in haystack:
                    continue

            out.append(
                {
                    "transactionId": _clean_text(row.get(sc.COL_TXN_ID)),
                    "txnDate": row_date,
                    "class": _clean_text(row.get(sc.COL_TXN_CLASS)),
                    "businessUnit": _clean_text(row.get(sc.COL_TXN_BUSINESS_UNIT)),
                    "type": _clean_text(row.get(sc.COL_TXN_TYPE)),
                    "fromAccount": _clean_text(row.get(sc.COL_TXN_FROM_ACCOUNT)),
                    "toAccount": _clean_text(row.get(sc.COL_TXN_TO_ACCOUNT)),
                    "payee": _clean_text(row.get(sc.COL_TXN_PAYEE)),
                    "parent": _clean_text(row.get(sc.COL_TXN_PARENT)),
                    "client": _clean_text(row.get(sc.COL_TXN_CLIENT)),
                    "matter": _clean_text(row.get(sc.COL_TXN_MATTER)),
                    "categoryCode": _clean_text(row.get(sc.COL_TXN_CATEGORY_CODE)),
                    "categoryName": _clean_text(row.get(sc.COL_TXN_CATEGORY_NAME)),
                    "member": _clean_text(row.get(sc.COL_TXN_MEMBER)),
                    "amount": round(float(self._parse_float(row.get(sc.COL_TXN_AMOUNT)) or 0.0), 2),
                    "taxAmount": round(float(self._parse_float(row.get(sc.COL_TXN_TAX_AMOUNT)) or 0.0), 2),
                    "taxFlag": _clean_text(row.get(sc.COL_TXN_TAX_FLAG)),
                    "hstExempt": self._to_bool_int(row.get(sc.COL_TXN_HST_EXEMPT), default=0),
                    "generalOfficeExpense": self._to_bool_int(
                        row.get(sc.COL_TXN_GENERAL_OFFICE_EXPENSE), default=0
                    ),
                    "shadow": self._to_bool_int(row.get(sc.COL_TXN_SHADOW), default=0),
                    "invoiceRef": _clean_text(row.get(sc.COL_TXN_INVOICE_REF)),
                    "billClaimPct": round(float(self._parse_float(row.get(sc.COL_TXN_BILL_CLAIM_PCT)) or 0.0), 2),
                    "totalClaimAmount": round(
                        float(self._parse_float(row.get(sc.COL_TXN_TOTAL_CLAIM_AMOUNT)) or 0.0), 2
                    ),
                    "expenseDetails": _clean_text(row.get(sc.COL_TXN_EXPENSE_DETAILS)),
                    "notes": _clean_text(row.get(sc.COL_TXN_NOTES)),
                    "status": _clean_text(row.get(sc.COL_TXN_STATUS)),
                    "currency": _clean_text(row.get(sc.COL_TXN_CURRENCY)),
                    "voidReason": _clean_text(row.get(sc.COL_TXN_VOID_REASON)),
                    "clearedAt": _clean_text(row.get(sc.COL_TXN_CLEARED_AT)),
                    "reconciledAt": _clean_text(row.get(sc.COL_TXN_RECONCILED_AT)),
                    "createdAt": _clean_text(row.get(sc.COL_TXN_CREATED_AT)),
                    "updatedAt": _clean_text(row.get(sc.COL_TXN_UPDATED_AT)),
                }
            )

        out.sort(
            key=lambda row: (
                _clean_text(row.get("txnDate")),
                _clean_text(row.get("updatedAt")),
                _clean_text(row.get("transactionId")),
            ),
            reverse=True,
        )
        return out

    def save_transaction(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.ensure_schema()
        normalized = self._normalize_transaction_payload(payload)

        txn_id = _clean_text(normalized.get("transactionId"))
        existing_row: Optional[Dict[str, Any]] = self._find_transaction_row(txn_id) if txn_id else None
        self._validate_transaction_status_transition(
            old_status=_clean_text((existing_row or {}).get(sc.COL_TXN_STATUS)),
            new_status=_clean_text(normalized.get("status")),
            has_existing=existing_row is not None,
        )
        if not txn_id:
            txn_id = self._new_id("TXN")

        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        created_at = (
            _clean_text(normalized.get("createdAt"))
            or _clean_text((existing_row or {}).get(sc.COL_TXN_CREATED_AT))
            or now_stamp
        )

        transaction_row = {
            sc.COL_TXN_ID: txn_id,
            sc.COL_TXN_DATE: normalized["txnDate"],
            sc.COL_TXN_CLASS: normalized["class"],
            sc.COL_TXN_BUSINESS_UNIT: normalized["businessUnit"],
            sc.COL_TXN_TYPE: normalized["type"],
            sc.COL_TXN_FROM_ACCOUNT: normalized["fromAccount"],
            sc.COL_TXN_TO_ACCOUNT: normalized["toAccount"],
            sc.COL_TXN_PAYEE: normalized["payee"],
            sc.COL_TXN_PARENT: normalized["parent"],
            sc.COL_TXN_CLIENT: normalized["client"],
            sc.COL_TXN_MATTER: normalized["matter"],
            sc.COL_TXN_CATEGORY_CODE: normalized["categoryCode"],
            sc.COL_TXN_CATEGORY_NAME: normalized["categoryName"],
            sc.COL_TXN_MEMBER: normalized["member"],
            sc.COL_TXN_AMOUNT: round(float(normalized["amount"]), 2),
            sc.COL_TXN_TAX_AMOUNT: round(float(normalized["taxAmount"]), 2),
            sc.COL_TXN_TAX_FLAG: normalized["taxFlag"],
            sc.COL_TXN_HST_EXEMPT: normalized["hstExempt"],
            sc.COL_TXN_GENERAL_OFFICE_EXPENSE: normalized["generalOfficeExpense"],
            sc.COL_TXN_SHADOW: normalized["shadow"],
            sc.COL_TXN_INVOICE_REF: normalized["invoiceRef"],
            sc.COL_TXN_BILL_CLAIM_PCT: round(float(normalized["billClaimPct"]), 2),
            sc.COL_TXN_TOTAL_CLAIM_AMOUNT: round(float(normalized["totalClaimAmount"]), 2),
            sc.COL_TXN_EXPENSE_DETAILS: normalized["expenseDetails"],
            sc.COL_TXN_NOTES: normalized["notes"],
            sc.COL_TXN_STATUS: normalized["status"],
            sc.COL_TXN_CURRENCY: normalized["currency"],
            sc.COL_TXN_VOID_REASON: normalized["voidReason"],
            sc.COL_TXN_CLEARED_AT: normalized["clearedAt"],
            sc.COL_TXN_RECONCILED_AT: normalized["reconciledAt"],
            sc.COL_TXN_CREATED_AT: created_at,
            sc.COL_TXN_UPDATED_AT: now_stamp,
        }

        self._upsert_row_by_key(TBL_TRANSACTIONS_MASTER, sc.COL_TXN_ID, txn_id, transaction_row)
        persisted = self._find_transaction_row(txn_id)
        verified = self._compare_transaction_rows_loose(transaction_row, persisted)

        warnings = list(normalized.get("warnings", []))
        if verified and warnings:
            message = "Transaction saved with warnings: " + "; ".join(warnings)
        elif verified:
            message = ""
        else:
            message = "Transaction verification failed."

        return {
            "ok": bool(verified),
            "verifiedExact": bool(verified),
            "transactionId": txn_id,
            "savedRow": transaction_row,
            "warnings": warnings,
            "message": message,
        }

    def save_ap_expense(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """Save the expense linked to an unpaid A/P bill.

        A supplier bill is not a cash payment.  Until it is settled, its linked
        transaction remains pending against the A/P clearing account.  A later
        set-off (or ordinary A/P payment) supplies the actual settlement path.
        """
        data = dict(payload or {})
        self.ensure_schema()
        if not _clean_text(data.get("fromAccount") or data.get("sourceAccount")):
            data["fromAccount"] = "AP_PAYABLE"
        if not _clean_text(data.get("businessUnit") or data.get("bu")):
            active_units = self.list_transaction_business_units()
            preferred = next(
                (
                    _clean_text(row.get("businessUnit"))
                    for row in active_units
                    if _clean_text(row.get("businessUnit")).casefold() == "cory business"
                ),
                "",
            )
            if not preferred and len(active_units) == 1:
                preferred = _clean_text(active_units[0].get("businessUnit"))
            if not preferred:
                raise ValueError(
                    "A/P business expense needs a BusinessUnit. Configure an active 'Cory Business' unit "
                    "or provide BusinessUnit explicitly."
                )
            data["businessUnit"] = preferred
        if not _clean_text(data.get("taxFlag") or data.get("taxCategory")):
            category_code = _clean_text(data.get("categoryCode") or data.get("catCode"))
            matching_category = next(
                (
                    row for row in self.list_transaction_categories("Expense", "", False)
                    if _clean_text(row.get("categoryCode")).casefold() == category_code.casefold()
                ),
                None,
            )
            data["taxFlag"] = _clean_text((matching_category or {}).get("taxFlagDefault")) or "None"
        data.setdefault("status", "Pending")
        return self.save_transaction(data)

    def save_receivable(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """Save a Receivables master record directly (used for data generation during legacy import)."""
        self.ensure_schema()
        invoice = _clean_text(payload.get("invoiceNum"))
        if not invoice:
            return {"ok": False, "message": "Missing invoice number."}

        row = {
            sc.COL_RECV_INVOICE_NUM: invoice,
            sc.COL_RECV_DATE: payload.get("date", ""),
            sc.COL_RECV_CLIENT: payload.get("client", ""),
            sc.COL_RECV_TOTAL_INVOICED: round(float(payload.get("totalInvoiced", 0.0)), 2),
            sc.COL_RECV_AMOUNT_PAID: round(float(payload.get("amountPaid", 0.0)), 2),
            sc.COL_RECV_CREDITS_ADJ: round(float(payload.get("creditsAdj", 0.0)), 2),
            sc.COL_RECV_BALANCE_DUE: round(float(payload.get("balanceDue", 0.0)), 2),
            sc.COL_RECV_STATUS: payload.get("status", ""),
            sc.COL_RECV_WORK_CLIENT: payload.get("workClient", "")
        }

        self._upsert_row_by_key(TBL_RECEIVABLES, sc.COL_RECV_INVOICE_NUM, invoice, row)
        return {"ok": True, "message": f"Saved legacy receivable {invoice}.", "invoice": invoice, "savedRow": row}

    def get_receivable(self, invoice_num: str) -> Dict[str, Any]:
        """Fetch a single receivable record."""
        self.ensure_schema()
        row = self._find_row_by_key(TBL_RECEIVABLES, sc.COL_RECV_INVOICE_NUM, invoice_num)
        if not row:
            return {"ok": False, "message": f"Invoice {invoice_num} not found in Receivables."}
        return {"ok": True, "receivable": self._canonicalize_row(TBL_RECEIVABLES, row)}

    def update_receivable(self, invoice_num: str, changes: Dict[str, Any]) -> Dict[str, Any]:
        self.ensure_schema()
        row = self._find_row_by_key(TBL_RECEIVABLES, sc.COL_RECV_INVOICE_NUM, invoice_num)
        if not row:
            return {"ok": False, "message": f"Invoice {invoice_num} not found."}

        updated = False
        
        # We need to map standard camelCase names to column names
        mapping = {
            "client": sc.COL_RECV_CLIENT,
            "date": sc.COL_RECV_DATE,
            "invoiceNum": sc.COL_RECV_INVOICE_NUM,
            "totalInvoiced": sc.COL_RECV_TOTAL_INVOICED,
            "amountPaid": sc.COL_RECV_AMOUNT_PAID,
            "creditsAdj": sc.COL_RECV_CREDITS_ADJ,
            "status": sc.COL_RECV_STATUS,
            "workClient": sc.COL_RECV_WORK_CLIENT
        }

        # Handle updating the primary key if invoiceNum changed
        old_invoice_num = invoice_num
        new_invoice_num = invoice_num
        
        if "invoiceNum" in changes and str(changes["invoiceNum"]).strip() != invoice_num:
            new_invoice_num = str(changes["invoiceNum"]).strip()
            if not new_invoice_num:
                return {"ok": False, "message": "Invoice number cannot be empty."}
            row[sc.COL_RECV_INVOICE_NUM] = new_invoice_num
            updated = True
            
            # Since primary key changed, we must delete the old row and insert the new one
            self._delete_row_by_key_hard(TBL_RECEIVABLES, sc.COL_RECV_INVOICE_NUM, old_invoice_num)

        for key, col in mapping.items():
            if key in changes and key != "invoiceNum":
                val = changes[key]
                if key in ["amountPaid", "creditsAdj", "totalInvoiced"]:
                    try:
                        row[col] = round(float(val), 2)
                    except (ValueError, TypeError):
                        pass
                else:
                    row[col] = _clean_text(val)
                updated = True

        if updated:
            # Recompute Balance Due
            total = float(row.get(sc.COL_RECV_TOTAL_INVOICED) or 0)
            paid = float(row.get(sc.COL_RECV_AMOUNT_PAID) or 0)
            credits_adj = float(row.get(sc.COL_RECV_CREDITS_ADJ) or 0)
            row[sc.COL_RECV_BALANCE_DUE] = round(total - paid - credits_adj, 2)
            
            self._upsert_row_by_key(TBL_RECEIVABLES, sc.COL_RECV_INVOICE_NUM, new_invoice_num, row)
            return {"ok": True, "message": "Receivable updated successfully.", "savedRow": self._canonicalize_row(TBL_RECEIVABLES, row)}
            
        return {"ok": True, "message": "No changes made."}

    def save_invoice_log(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """Save an Invoice Log record directly (used for data generation during legacy import)."""
        self.ensure_schema()
        invoice = _clean_text(payload.get("invoiceNum"))
        if not invoice:
            return {"ok": False, "message": "Missing invoice number."}
            
        existing = self._find_row_by_key(TBL_INVOICE_LOG, sc.COL_INV_INVOICE_NUM, invoice) or {}
        
        row = {
            sc.COL_INV_INVOICE_NUM: invoice,
            sc.COL_INV_CLIENT_NAME: payload.get("clientName", existing.get(sc.COL_INV_CLIENT_NAME, "")),
            sc.COL_INV_SUB_CLIENT: payload.get("subClient", existing.get(sc.COL_INV_SUB_CLIENT, "")),
            sc.COL_INV_INVOICE_DATE: payload.get("invoiceDate") or existing.get(sc.COL_INV_INVOICE_DATE, ""),
            sc.COL_INV_TOTAL_FEES: round(float(payload.get("totalFees", existing.get(sc.COL_INV_TOTAL_FEES) or 0.0)), 2),
            sc.COL_INV_TOTAL_DISBURSEMENTS: round(float(payload.get("totalDisbursements", existing.get(sc.COL_INV_TOTAL_DISBURSEMENTS) or 0.0)), 2),
            sc.COL_INV_TOTAL_TAX: round(float(payload.get("totalTax", existing.get(sc.COL_INV_TOTAL_TAX) or 0.0)), 2),
            sc.COL_INV_AGGREGATE_BILLED: round(float(payload.get("aggregateBilled", existing.get(sc.COL_INV_AGGREGATE_BILLED) or 0.0)), 2),
            sc.COL_INV_BILL_TO_CLIENT: payload.get("billToClient", existing.get(sc.COL_INV_BILL_TO_CLIENT, ""))
        }
        
        self._upsert_row_by_key(TBL_INVOICE_LOG, sc.COL_INV_INVOICE_NUM, invoice, row)
        
        # Sync the updated aggregate amount (and date) to the Receivables tracker if it exists there
        recv_row = self._find_row_by_key(TBL_RECEIVABLES, sc.COL_RECV_INVOICE_NUM, invoice)
        if recv_row:
            recv_update = {}
            if "aggregateBilled" in payload:
                recv_update["totalInvoiced"] = round(float(payload.get("aggregateBilled", 0.0)), 2)
            if payload.get("invoiceDate"):
                recv_update["date"] = payload.get("invoiceDate")
            
            if recv_update:
                self.update_receivable(invoice, recv_update)
            
        return {"ok": True, "invoice": invoice}

    def reverse_and_reissue_invoice(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """
        Reverse an existing invoice and immediately reissue it as amended (-A).
        This executes the strict audit flow without un-billing WIP.
        """
        self.ensure_schema()
        orig_invoice = _clean_text(payload.get("invoiceNum"))
        if not orig_invoice:
            return {"ok": False, "message": "Missing invoice number."}

        # 1. Void original in Receivables
        recv_rows = self._read_table_rows(TBL_RECEIVABLES)
        orig_recv = None
        for r in recv_rows:
            if r.get(sc.COL_RECV_INVOICE_NUM) == orig_invoice:
                orig_recv = dict(r)
                r[sc.COL_RECV_STATUS] = "Void"
        if not orig_recv:
            return {"ok": False, "message": f"Original invoice {orig_invoice} not found in Receivables."}
        
        self._write_table_rows(TBL_RECEIVABLES, recv_rows)

        # 2. Add -V (Void) and -A (Amended) entries to Invoice Log
        invoice_log = self._read_table_rows(TBL_INVOICE_LOG)
        orig_log = next((i for i in invoice_log if i.get(sc.COL_INV_INVOICE_NUM) == orig_invoice), None)
        if not orig_log:
            return {"ok": False, "message": f"Original invoice {orig_invoice} not found in Invoice Log."}
            
        # Create -V entry
        void_inv = dict(orig_log)
        void_inv[sc.COL_INV_INVOICE_NUM] = f"{orig_invoice}-V"
        for amount_col in [sc.COL_INV_TOTAL_FEES, sc.COL_INV_TOTAL_DISBURSEMENTS, sc.COL_INV_TOTAL_TAX, sc.COL_INV_AGGREGATE_BILLED]:
            val = void_inv.get(amount_col)
            if val:
                try:
                    void_inv[amount_col] = -float(val)
                except ValueError:
                    pass
        invoice_log.append(void_inv)
        
        # Create -A entry in Invoice Log
        amended_inv = dict(orig_log)
        amended_inv[sc.COL_INV_INVOICE_NUM] = f"{orig_invoice}-A"
        amended_inv[sc.COL_INV_INVOICE_DATE] = payload.get("invoiceDate") or orig_log.get(sc.COL_INV_INVOICE_DATE, "")
        amended_inv[sc.COL_INV_CLIENT_NAME] = payload.get("clientName", orig_log.get(sc.COL_INV_CLIENT_NAME, ""))
        amended_inv[sc.COL_INV_TOTAL_FEES] = round(float(payload.get("totalFees", 0.0)), 2)
        amended_inv[sc.COL_INV_TOTAL_DISBURSEMENTS] = round(float(payload.get("totalDisbursements", 0.0)), 2)
        amended_inv[sc.COL_INV_TOTAL_TAX] = round(float(payload.get("totalTax", 0.0)), 2)
        amended_inv[sc.COL_INV_AGGREGATE_BILLED] = round(float(payload.get("aggregateBilled", 0.0)), 2)
        invoice_log.append(amended_inv)
        
        self._write_table_rows(TBL_INVOICE_LOG, invoice_log)
        
        # 3. Create -A entry in Receivables
        amended_recv = dict(orig_recv)
        amended_recv[sc.COL_RECV_INVOICE_NUM] = f"{orig_invoice}-A"
        amended_recv[sc.COL_RECV_STATUS] = "Active"
        amended_recv[sc.COL_RECV_DATE] = amended_inv[sc.COL_INV_INVOICE_DATE]
        amended_recv[sc.COL_RECV_CLIENT] = amended_inv[sc.COL_INV_CLIENT_NAME]
        
        total_invoiced = amended_inv[sc.COL_INV_AGGREGATE_BILLED]
        amount_paid = float(orig_recv.get(sc.COL_RECV_AMOUNT_PAID, 0.0) or 0.0)
        credits_adj = float(orig_recv.get(sc.COL_RECV_CREDITS_ADJ, 0.0) or 0.0)
        balance_due = total_invoiced - amount_paid + credits_adj
        
        amended_recv[sc.COL_RECV_TOTAL_INVOICED] = round(total_invoiced, 2)
        amended_recv[sc.COL_RECV_BALANCE_DUE] = round(balance_due, 2)
        
        recv_rows = self._read_table_rows(TBL_RECEIVABLES)
        recv_rows.append(amended_recv)
        self._write_table_rows(TBL_RECEIVABLES, recv_rows)
        
        return {"ok": True, "invoice": f"{orig_invoice}-A", "message": f"Successfully reversed and re-issued as {orig_invoice}-A"}

    def list_open_payment_invoices(self, filters: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """Return open invoice rows shaped for the Payment Entry workspace."""

        payload = dict(filters or {})
        report = self.ar_aging_report(
            {
                "asOfDate": payload.get("asOfDate") or payload.get("as_of_date") or date.today().isoformat(),
                "query": payload.get("query") or payload.get("searchText") or payload.get("search") or "",
                "groupBy": payload.get("groupBy") or payload.get("group_by") or "client",
            }
        )
        if not report.get("ok"):
            return []

        rows: List[Dict[str, Any]] = []
        for row in list(report.get("rows") or []):
            invoice = _clean_text(row.get("invoice"))
            if not invoice:
                continue
            client = _clean_text(row.get("client"))
            billing_client = _clean_text(row.get("billingClient"))
            balance = self._money_round(row.get("balance"))
            invoice_total = self._money_round(row.get("invoiceTotal"))
            paid = self._money_round(row.get("paid"))
            rows.append(
                {
                    "invoice": invoice,
                    "invoiceNumber": invoice,
                    "date": _clean_text(row.get("date")),
                    "client": client,
                    "billingClient": billing_client,
                    "workClient": _clean_text(row.get("workClient")) or client,
                    "status": _clean_text(row.get("status")),
                    "ageDays": int(row.get("ageDays") or 0),
                    "bucketLabel": _clean_text(row.get("bucketLabel")),
                    "invoiceTotal": invoice_total,
                    "paid": paid,
                    "credits": self._money_round(row.get("credits")),
                    "balance": balance,
                    "display": f"{invoice} | {client or billing_client} | ${balance:,.2f}",
                }
            )
        return rows

    def _list_ap_setoff_invoice_activity(self, invoice_ref: str) -> List[Dict[str, Any]]:
        """Return governed A/P set-off evidence allocated to one receivable.

        A set-off payment is an A/P record with multiple receivable
        allocations, not a cash receipt. Reading it here keeps the Invoice
        Directory from falling back to a historical transfer-row surrogate.
        """

        paths = getattr(self, "paths", None)
        if not paths:
            return []
        target = _clean_text(invoice_ref).casefold()
        if not target:
            return []
        try:
            from repositories.ap_workbook_repository import APWorkbookRepository
            payments = APWorkbookRepository(paths.workbook_path()).list_payments()
        except Exception:
            logger.exception("Invoice payment history could not read A/P set-off records")
            return []

        marker = "CSPM_SET_OFF_ALLOCATIONS_V1:"
        activity: List[Dict[str, Any]] = []
        for payment in payments:
            if _clean_text(payment.get("Method")).casefold() != "set-off":
                continue
            if _clean_text(payment.get("Status")).casefold() in {"reversed", "reversal"}:
                continue
            allocations = []
            for line in str(payment.get("Notes") or "").splitlines():
                if line.startswith(marker):
                    try:
                        allocations = json.loads(line[len(marker):])
                    except (TypeError, ValueError, json.JSONDecodeError):
                        allocations = []
                    break
            for allocation in allocations:
                allocation = dict(allocation or {})
                invoice = _clean_text(allocation.get("invoice") or allocation.get("InvoiceID"))
                if invoice.casefold() != target:
                    continue
                amount = self._money_round(allocation.get("amount", allocation.get("Amount")))
                if amount <= 0:
                    continue
                payment_id = _clean_text(payment.get("APPaymentID"))
                reference = _clean_text(payment.get("Reference"))
                activity.append(
                    {
                        "date": self._date_iso(payment.get("PaymentDate")),
                        "type": "Settlement set-off",
                        "displayLabel": "Settlement set-off" + (f" — {reference}" if reference else ""),
                        "paymentId": payment_id,
                        "transactionId": "",
                        "ledgerId": "",
                        "editable": True,
                        "reference": payment_id or reference,
                        "method": "Accounts Receivable - Set-off",
                        "amount": amount,
                        "notes": _clean_text(payment.get("Notes")).split(marker, 1)[0].rstrip(),
                        "source": "A/P Set-off",
                        "openTarget": "ap_setoff",
                        "apBillId": _clean_text(payment.get("APBillID")),
                        "apPaymentId": payment_id,
                    }
                )
        return activity

    def list_invoice_payment_history(self, invoice_ref: str) -> List[Dict[str, Any]]:
        """Return the dated payment activity recorded against one invoice.

        The ledger is the authoritative receivables record.  A payment posted
        through CSPM creates both a financial transaction and a ledger entry,
        so returning both naively makes one payment appear twice.  Keep the
        ledger entry, then only add older transaction-only payments that have
        no matching ledger transaction.
        """

        target = _clean_text(invoice_ref)
        if not target:
            return []
        target_lc = target.lower()
        rows: List[Dict[str, Any]] = []
        load_setoff_activity = getattr(self, "_list_ap_setoff_invoice_activity", None)
        setoff_activity = load_setoff_activity(target) if callable(load_setoff_activity) else []
        canonical_setoff_ids = {
            _clean_text(item.get("paymentId")).casefold()
            for item in setoff_activity if _clean_text(item.get("paymentId"))
        }
        canonical_setoff_signatures = {
            (self._date_iso(item.get("date")), self._money_round(item.get("amount")))
            for item in setoff_activity
        }
        ledger_transaction_ids = set()
        for row in self._read_table_rows(TBL_LEDGER):
            if _clean_text(row.get(sc.COL_LEDGER_REFERENCE)).lower() != target_lc:
                continue
            collected = self._money_round(row.get(sc.COL_LEDGER_COLLECTED))
            write_off = self._money_round(row.get(sc.COL_LEDGER_WRITE_OFF))
            if collected <= 0 and write_off <= 0:
                continue
            transaction_id = _clean_text(row.get(sc.COL_LEDGER_TRX_ID))
            external_reference = _clean_text(row.get(sc.COL_LEDGER_EXTERNAL_REF_ID))
            if transaction_id.casefold() in canonical_setoff_ids:
                # The governed A/P record below owns the user-facing set-off
                # detail; this ledger row remains its audit evidence.
                continue
            if transaction_id:
                ledger_transaction_ids.add(transaction_id.lower())
            # Older ledger rows sometimes stored the generated transaction ID
            # in ExternalRef instead of TransactionID.
            if external_reference.lower().startswith("txn_"):
                ledger_transaction_ids.add(external_reference.lower())
            rows.append(
                {
                    "date": self._date_iso(row.get(sc.COL_LEDGER_DATE)),
                    "type": "Payment" if collected > 0 else "Write-off/Adjustment",
                    "paymentId": transaction_id
                    or _clean_text(row.get(sc.COL_LEDGER_ID)),
                    "transactionId": transaction_id,
                    "ledgerId": _clean_text(row.get(sc.COL_LEDGER_ID)),
                    "editable": True,
                    "reference": transaction_id
                    or external_reference
                    or _clean_text(row.get(sc.COL_LEDGER_ID)),
                    "method": _clean_text(row.get(sc.COL_LEDGER_CATEGORY)),
                    "amount": collected if collected > 0 else write_off,
                    "notes": _clean_text(row.get(sc.COL_LEDGER_DESCRIPTION)),
                    "source": "Ledger",
                }
            )

        # Preserve genuinely historic payment transactions that predate the
        # ledger workflow, together with an explicitly recorded settlement
        # set-off.  Do not show invoice-creation revenue or a duplicate
        # transaction that already has ledger evidence.
        for row in self.list_transactions({"invoiceRef": target}):
            if _clean_text(row.get("invoiceRef")).lower() != target_lc:
                continue
            transaction_id = _clean_text(row.get("transactionId"))
            if transaction_id and transaction_id.lower() in ledger_transaction_ids:
                continue
            transaction_type = _clean_text(row.get("type")).lower()
            notes = _clean_text(row.get("notes"))
            note_key = notes.lower()
            is_payment_note = (
                "payment" in note_key
                or "applied to invoice" in note_key
                or "received from" in note_key
            )
            is_settlement_set_off = (
                transaction_type == "transfer"
                and "set-off" in note_key
            )
            if not (
                (transaction_type == "income" and is_payment_note)
                or is_settlement_set_off
            ):
                continue
            amount = self._money_round(row.get("amount"))
            if amount <= 0:
                continue
            if is_settlement_set_off and (
                self._date_iso(row.get("txnDate")), amount
            ) in canonical_setoff_signatures:
                # The reconstructed A/P payment is authoritative. Do not
                # show its former legacy-transfer representation a second time.
                continue
            activity_type = "Settlement set-off" if is_settlement_set_off else "Payment"
            display_label = activity_type
            if is_settlement_set_off:
                # The operational note normally identifies the settlement
                # source in parentheses, for example "Set-off — LIHDC
                # settlement".  Surface that useful context instead of an
                # inaccurate blanket "No payments recorded" message.
                set_off_detail = ""
                marker = "set-off"
                marker_index = note_key.find(marker)
                if marker_index >= 0:
                    set_off_detail = notes[marker_index + len(marker):].strip(" —:-()")
                if set_off_detail:
                    display_label = f"{activity_type} — {set_off_detail}"
            rows.append(
                {
                    "date": _clean_text(row.get("txnDate")),
                    "type": activity_type,
                    "displayLabel": display_label,
                    "paymentId": transaction_id,
                    "transactionId": transaction_id,
                    "ledgerId": "",
                    "editable": True,
                    "reference": transaction_id,
                    "method": _clean_text(row.get("fromAccount")),
                    "amount": amount,
                    "notes": notes,
                    "source": "Transactions",
                    # Ledger-backed rows continue to open the Payment Editor.
                    # A set-off is a transfer record, so it must open its
                    # editable Transaction Master record instead.
                    "openTarget": "transaction" if is_settlement_set_off else "payment",
                    "transactionPayload": dict(row),
                }
            )

        rows.extend(setoff_activity)

        rows.sort(key=lambda item: (_clean_text(item.get("date")), _clean_text(item.get("reference"))), reverse=True)
        return rows

    def _invoice_payment_record(self, payment_id: str) -> Dict[str, Any]:
        """Resolve a payment ID to its linked ledger and transaction evidence."""

        target = _clean_text(payment_id)
        if not target:
            raise ValueError("Payment ID is required.")
        target_lc = target.lower()
        ledger_rows = self._read_table_rows(TBL_LEDGER)
        transaction_rows = self._read_table_rows(TBL_TRANSACTIONS_MASTER)
        ledger_index = -1
        ledger_row: Dict[str, Any] = {}

        for index, row in enumerate(ledger_rows):
            identifiers = (
                _clean_text(row.get(sc.COL_LEDGER_ID)),
                _clean_text(row.get(sc.COL_LEDGER_TRX_ID)),
                _clean_text(row.get(sc.COL_LEDGER_EXTERNAL_REF_ID)),
            )
            if any(identifier.lower() == target_lc for identifier in identifiers if identifier):
                ledger_index = index
                ledger_row = dict(row)
                break

        transaction_id = _clean_text(ledger_row.get(sc.COL_LEDGER_TRX_ID)) or target
        transaction_index = -1
        transaction_row: Dict[str, Any] = {}
        for index, row in enumerate(transaction_rows):
            if _clean_text(row.get(sc.COL_TXN_ID)).lower() == transaction_id.lower():
                transaction_index = index
                transaction_row = dict(row)
                transaction_id = _clean_text(row.get(sc.COL_TXN_ID))
                break

        if ledger_index < 0 and transaction_index < 0:
            raise ValueError(f"Payment {target} was not found.")

        invoice = _clean_text(ledger_row.get(sc.COL_LEDGER_REFERENCE)) or _clean_text(
            transaction_row.get(sc.COL_TXN_INVOICE_REF)
        )
        if not invoice:
            raise ValueError(f"Payment {target} is not linked to an invoice.")

        return {
            "paymentId": transaction_id
            or _clean_text(ledger_row.get(sc.COL_LEDGER_ID))
            or target,
            "invoice": invoice,
            "ledgerIndex": ledger_index,
            "ledgerRow": ledger_row,
            "ledgerRows": ledger_rows,
            "transactionId": transaction_id,
            "transactionIndex": transaction_index,
            "transactionRow": transaction_row,
            "transactionRows": transaction_rows,
        }

    def get_invoice_payment_entry(self, payment_id: str) -> Dict[str, Any]:
        """Return one saved payment in the editable Payment Entry shape."""

        try:
            record = self._invoice_payment_record(payment_id)
            invoice = _clean_text(record.get("invoice"))
            ledger_row = dict(record.get("ledgerRow") or {})
            transaction_row = dict(record.get("transactionRow") or {})
            collected = self._money_round(ledger_row.get(sc.COL_LEDGER_COLLECTED))
            write_off = self._money_round(ledger_row.get(sc.COL_LEDGER_WRITE_OFF))
            if not ledger_row:
                collected = self._money_round(transaction_row.get(sc.COL_TXN_AMOUNT))
            if collected <= 0 and write_off <= 0:
                raise ValueError(f"Payment {payment_id} has no editable amount.")

            receivable = None
            for row in self._read_table_rows(TBL_RECEIVABLES):
                if _clean_text(row.get(sc.COL_RECV_INVOICE_NUM)).lower() == invoice.lower():
                    receivable = dict(row)
                    break
            if receivable is None:
                raise ValueError(f"Invoice {invoice} was not found in Receivables.")

            reference = _clean_text(ledger_row.get(sc.COL_LEDGER_EXTERNAL_REF_ID))
            if reference.lower() == _clean_text(record.get("transactionId")).lower():
                reference = ""
            notes = _clean_text(transaction_row.get(sc.COL_TXN_NOTES)) or _clean_text(
                ledger_row.get(sc.COL_LEDGER_DESCRIPTION)
            )
            return {
                "ok": True,
                "paymentId": _clean_text(record.get("paymentId")),
                "transactionId": _clean_text(record.get("transactionId")),
                "ledgerId": _clean_text(ledger_row.get(sc.COL_LEDGER_ID)),
                "invoice": invoice,
                "invoiceRow": self._payment_invoice_snapshot(receivable),
                "date": self._date_iso(
                    ledger_row.get(sc.COL_LEDGER_DATE)
                    or transaction_row.get(sc.COL_TXN_DATE)
                ),
                "mode": "Write-off / Adjustment" if write_off > 0 else "Payment",
                "method": _clean_text(ledger_row.get(sc.COL_LEDGER_CATEGORY))
                or _clean_text(transaction_row.get(sc.COL_TXN_FROM_ACCOUNT)),
                "depositAccount": _clean_text(transaction_row.get(sc.COL_TXN_FROM_ACCOUNT)),
                "reference": reference,
                "amount": collected,
                "adjustmentAmount": write_off,
                "adjustmentReason": "",
                "notes": notes,
            }
        except Exception as exc:
            return {"ok": False, "message": str(exc)}

    @with_financial_write_batch
    def update_invoice_payment(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """Amend one recorded payment and keep A/R, ledger, and WIP in sync."""

        self.ensure_schema()
        data = dict(payload or {})
        record = self._invoice_payment_record(
            _clean_text(data.get("paymentId") or data.get("transactionId"))
        )
        invoice = _clean_text(record.get("invoice"))
        payment_date = _clean_text(data.get("date") or data.get("paymentDate"))
        if not _is_valid_iso_date(payment_date):
            raise ValueError("Payment date must be in YYYY-MM-DD format.")

        ledger_row = dict(record.get("ledgerRow") or {})
        transaction_row = dict(record.get("transactionRow") or {})
        old_payment = self._money_round(ledger_row.get(sc.COL_LEDGER_COLLECTED))
        old_adjustment = self._money_round(ledger_row.get(sc.COL_LEDGER_WRITE_OFF))
        if not ledger_row:
            old_payment = self._money_round(transaction_row.get(sc.COL_TXN_AMOUNT))

        # Editing preserves the original payment kind.  A payment can be
        # corrected without accidentally converting cash received into a
        # write-off (or vice versa).
        if old_payment > 0:
            payment_amount = self._money_round(data.get("amount"))
            adjustment_amount = 0.0
            if payment_amount <= 0:
                raise ValueError("Payment amount must be greater than 0.")
        else:
            payment_amount = 0.0
            adjustment_amount = self._money_round(
                data.get("adjustmentAmount") or data.get("amount")
            )
            if adjustment_amount <= 0:
                raise ValueError("Adjustment amount must be greater than 0.")

        receivable_rows = self._read_table_rows(TBL_RECEIVABLES)
        receivable_index = -1
        receivable: Dict[str, Any] = {}
        for index, row in enumerate(receivable_rows):
            if _clean_text(row.get(sc.COL_RECV_INVOICE_NUM)).lower() == invoice.lower():
                receivable_index = index
                receivable = dict(row)
                break
        if receivable_index < 0:
            raise ValueError(f"Invoice {invoice} was not found in Receivables.")

        invoice_total = self._money_round(receivable.get(sc.COL_RECV_TOTAL_INVOICED))
        base_paid = self._money_round(receivable.get(sc.COL_RECV_AMOUNT_PAID)) - old_payment
        base_credits = self._money_round(receivable.get(sc.COL_RECV_CREDITS_ADJ)) - old_adjustment
        if base_paid < -0.01 or base_credits < -0.01:
            raise ValueError("This payment no longer matches the invoice balance. Refresh and try again.")
        next_paid = self._money_round(max(0.0, base_paid) + payment_amount)
        next_credits = self._money_round(max(0.0, base_credits) + adjustment_amount)
        next_balance = self._money_round(invoice_total - next_paid - next_credits)
        if next_balance < -0.01:
            raise ValueError(
                f"Updated amount exceeds the invoice total by ${abs(next_balance):,.2f}."
            )
        if abs(next_balance) <= 0.01:
            next_balance = 0.0

        method = _clean_text(data.get("method"))
        deposit_account = _clean_text(data.get("depositAccount") or data.get("account"))
        reference = _clean_text(data.get("reference"))
        notes = _clean_text(data.get("notes"))
        adjustment_reason = _clean_text(data.get("adjustmentReason"))
        transaction_id = _clean_text(record.get("transactionId"))
        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        transaction_rows = list(record.get("transactionRows") or [])
        transaction_index = int(record.get("transactionIndex", -1))
        if payment_amount > 0:
            if transaction_index >= 0:
                updated_transaction = dict(transaction_rows[transaction_index])
                updated_transaction[sc.COL_TXN_DATE] = payment_date
                updated_transaction[sc.COL_TXN_TYPE] = "Income"
                updated_transaction[sc.COL_TXN_FROM_ACCOUNT] = deposit_account or method or _clean_text(
                    updated_transaction.get(sc.COL_TXN_FROM_ACCOUNT)
                )
                updated_transaction[sc.COL_TXN_AMOUNT] = payment_amount
                updated_transaction[sc.COL_TXN_INVOICE_REF] = invoice
                updated_transaction[sc.COL_TXN_NOTES] = notes or f"{method or 'Payment'} applied to invoice {invoice}"
                updated_transaction[sc.COL_TXN_CLEARED_AT] = payment_date
                updated_transaction[sc.COL_TXN_UPDATED_AT] = now_stamp
                transaction_rows[transaction_index] = updated_transaction
                transaction_id = _clean_text(updated_transaction.get(sc.COL_TXN_ID))
            else:
                result = self.save_transaction(
                    {
                        "txnDate": payment_date,
                        "class": "Business",
                        "businessUnit": "Legal Practice",
                        "type": "Income",
                        "fromAccount": deposit_account or method or "Operating Account",
                        "payee": _clean_text(receivable.get(sc.COL_RECV_CLIENT)),
                        "client": _clean_text(receivable.get(sc.COL_RECV_WORK_CLIENT))
                        or _clean_text(receivable.get(sc.COL_RECV_CLIENT)),
                        "categoryCode": "INC_LEGAL_FEES",
                        "categoryName": "Legal Fees Revenue",
                        "member": "Cory",
                        "amount": payment_amount,
                        "taxAmount": 0.0,
                        "taxFlag": "None",
                        "hstExempt": 1,
                        "invoiceRef": invoice,
                        "status": "Cleared",
                        "currency": "CAD",
                        "notes": notes or f"{method or 'Payment'} applied to invoice {invoice}",
                        "clearedAt": payment_date,
                    }
                )
                if not result.get("ok"):
                    raise ValueError(_clean_text(result.get("message")) or "Payment transaction was not saved.")
                transaction_id = _clean_text(result.get("transactionId"))
        if transaction_index >= 0:
            self._replace_table_rows(TBL_TRANSACTIONS_MASTER, transaction_rows)

        ledger_rows = list(record.get("ledgerRows") or [])
        ledger_index = int(record.get("ledgerIndex", -1))
        ledger_description = (
            f"Payment applied to invoice {invoice}"
            if payment_amount > 0
            else f"Credit/Adj applied to invoice {invoice}"
        )
        if payment_amount > 0 and method:
            ledger_description += f" ({method})"
        if notes:
            ledger_description += f" - {notes}"
        elif adjustment_reason:
            ledger_description += f" - {adjustment_reason}"
        updated_ledger = dict(ledger_row) if ledger_row else {
            sc.COL_LEDGER_ID: self._new_id("LED"),
            sc.COL_LEDGER_CLIENT_VENDOR: _clean_text(receivable.get(sc.COL_RECV_CLIENT)),
            sc.COL_LEDGER_BILLINGS_EXCL_HST: 0.0,
            sc.COL_LEDGER_HST_COLLECTED: 0.0,
            sc.COL_LEDGER_EXPENSES_EXCL_HST: 0.0,
            sc.COL_LEDGER_HST_PAID: 0.0,
            sc.COL_LEDGER_WORK_CLIENT: _clean_text(receivable.get(sc.COL_RECV_WORK_CLIENT)),
            sc.COL_LEDGER_CREATED_AT: now_stamp,
        }
        updated_ledger[sc.COL_LEDGER_DATE] = payment_date
        updated_ledger[sc.COL_LEDGER_DESCRIPTION] = ledger_description
        updated_ledger[sc.COL_LEDGER_CATEGORY] = method or (
            "Payment" if payment_amount > 0 else "Credit/Adj"
        )
        updated_ledger[sc.COL_LEDGER_REFERENCE] = invoice
        updated_ledger[sc.COL_LEDGER_COLLECTED] = payment_amount
        updated_ledger[sc.COL_LEDGER_WRITE_OFF] = adjustment_amount
        updated_ledger[sc.COL_LEDGER_RECEIVABLE] = self._money_round(
            -(payment_amount + adjustment_amount)
        )
        updated_ledger[sc.COL_LEDGER_TRX_ID] = transaction_id
        updated_ledger[sc.COL_LEDGER_EXTERNAL_REF_ID] = reference
        updated_ledger[sc.COL_LEDGER_ORIGINAL_AMOUNT] = self._money_round(
            payment_amount + adjustment_amount
        )
        if ledger_index >= 0:
            ledger_rows[ledger_index] = updated_ledger
        else:
            ledger_rows.append(updated_ledger)
        self._replace_table_rows(TBL_LEDGER, ledger_rows)

        updated_receivable = dict(receivable)
        updated_receivable[sc.COL_RECV_AMOUNT_PAID] = next_paid
        updated_receivable[sc.COL_RECV_CREDITS_ADJ] = next_credits
        updated_receivable[sc.COL_RECV_BALANCE_DUE] = next_balance
        updated_receivable[sc.COL_RECV_STATUS] = "Paid" if next_balance <= 0 else "Partial"
        receivable_rows[receivable_index] = updated_receivable
        self._replace_table_rows(TBL_RECEIVABLES, receivable_rows)

        payment_status = updated_receivable[sc.COL_RECV_STATUS]
        for table, invoice_column in (
            (TBL_TIME, sc.COL_TIME_INVOICE_REF),
            (TBL_DISBURSEMENTS, sc.COL_DISB_INVOICE_REF),
        ):
            rows = self._read_table_rows(table)
            updated_rows: List[Dict[str, Any]] = []
            for row in rows:
                next_row = dict(row)
                if _clean_text(next_row.get(invoice_column)).lower() == invoice.lower():
                    next_row[sc.COL_TIME_PAYMENT_STATUS if table == TBL_TIME else sc.COL_DISB_PAYMENT_STATUS] = payment_status
                    next_row[sc.COL_TIME_INVOICE_TOTAL if table == TBL_TIME else sc.COL_DISB_INVOICE_TOTAL] = invoice_total
                    next_row[sc.COL_TIME_INVOICE_AMOUNT_PAID if table == TBL_TIME else sc.COL_DISB_INVOICE_AMOUNT_PAID] = self._money_round(next_paid + next_credits)
                    next_row[sc.COL_TIME_INVOICE_BALANCE_DUE if table == TBL_TIME else sc.COL_DISB_INVOICE_BALANCE_DUE] = next_balance
                updated_rows.append(next_row)
            self._replace_table_rows(table, updated_rows)

        return {
            "ok": True,
            "mode": "Payment" if payment_amount > 0 else "Adjustment",
            "invoice": invoice,
            "paymentId": transaction_id or _clean_text(updated_ledger.get(sc.COL_LEDGER_ID)),
            "transactionId": transaction_id,
            "ledgerId": _clean_text(updated_ledger.get(sc.COL_LEDGER_ID)),
            "amount": self._money_round(payment_amount + adjustment_amount),
            "paymentAmount": payment_amount,
            "adjustmentAmount": adjustment_amount,
            "afterBalance": next_balance,
            "invoiceRow": self._payment_invoice_snapshot(updated_receivable),
            "message": f"Payment for invoice {invoice} updated.",
        }

    @with_financial_write_batch
    def post_invoice_payment(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """Post a payment or write-off against one open receivable invoice."""

        self.ensure_schema()
        data = dict(payload or {})
        invoice = _clean_text(data.get("invoice") or data.get("invoiceNumber") or data.get("invoiceRef"))
        if not invoice:
            raise ValueError("Invoice is required.")

        payment_date = _clean_text(data.get("date") or data.get("paymentDate") or data.get("txnDate"))
        if not payment_date:
            payment_date = date.today().isoformat()
        if not _is_valid_iso_date(payment_date):
            raise ValueError("Payment date must be in YYYY-MM-DD format.")

        mode_raw = _clean_text(data.get("mode") or data.get("type") or "Payment").lower()
        
        raw_amt = self._money_round(data.get("amount") or 0.0)
        payment_amt = self._money_round(data.get("paymentAmount") or 0.0)
        adj_amt = self._money_round(data.get("adjustmentAmount") or 0.0)
        
        if raw_amt > 0:
            if "adjustment" in mode_raw or "write-off" in mode_raw or "writeoff" in mode_raw:
                adj_amt = adj_amt or raw_amt
            else:
                payment_amt = payment_amt or raw_amt
        
        if payment_amt <= 0 and adj_amt <= 0:
            raise ValueError("Payment or adjustment amount must be greater than 0.")

        method = _clean_text(data.get("method") or data.get("paymentMethod"))
        reference = _clean_text(data.get("reference") or data.get("ref") or data.get("cheque") or data.get("chequeNumber"))
        deposit_account = _clean_text(data.get("depositAccount") or data.get("account") or method or "Operating Account")
        notes = _clean_text(data.get("notes"))
        adj_reason = _clean_text(data.get("adjustmentReason"))

        receivable_rows = self._read_table_rows(TBL_RECEIVABLES)
        target_row: Optional[Dict[str, Any]] = None
        for row in receivable_rows:
            if _clean_text(row.get(sc.COL_RECV_INVOICE_NUM)).lower() == invoice.lower():
                target_row = dict(row)
                break
        if target_row is None:
            raise ValueError(f"Invoice {invoice} was not found in Receivables.")

        status_key = _clean_text(target_row.get(sc.COL_RECV_STATUS)).lower()
        if status_key in {"void", "cancelled", "canceled", "closed", "paid"}:
            raise ValueError(f"Invoice {invoice} is not open for payment.")

        before_paid = self._money_round(target_row.get(sc.COL_RECV_AMOUNT_PAID))
        before_credits = self._money_round(target_row.get(sc.COL_RECV_CREDITS_ADJ))
        before_balance = self._money_round(target_row.get(sc.COL_RECV_BALANCE_DUE))
        invoice_total = self._money_round(target_row.get(sc.COL_RECV_TOTAL_INVOICED))
        if before_balance <= 0:
            raise ValueError(f"Invoice {invoice} has no remaining balance.")
        
        total_applied = self._money_round(payment_amt + adj_amt)
        if total_applied - before_balance > 0.01:
            raise ValueError(
                f"Total applied amount ${total_applied:,.2f} exceeds the remaining invoice balance ${before_balance:,.2f}."
            )

        transaction_result: Dict[str, Any] = {}
        transaction_id = ""
        ledger_id = ""
        
        if payment_amt > 0:
            transaction_payload = {
                "txnDate": payment_date,
                "class": "Business",
                "businessUnit": "Legal Practice",
                "type": "Income",
                "fromAccount": deposit_account,
                "payee": _clean_text(target_row.get(sc.COL_RECV_CLIENT)),
                "client": _clean_text(target_row.get(sc.COL_RECV_WORK_CLIENT))
                or _clean_text(target_row.get(sc.COL_RECV_CLIENT)),
                "categoryCode": "INC_LEGAL_FEES",
                "categoryName": "Legal Fees Revenue",
                "member": "Cory",
                "amount": payment_amt,
                "taxAmount": 0.0,
                "taxFlag": "None",
                "hstExempt": 1,
                "invoiceRef": invoice,
                "status": "Cleared",
                "currency": "CAD",
                "notes": notes or f"{method or 'Payment'} applied to invoice {invoice}",
                "clearedAt": payment_date,
            }
            self._normalize_transaction_payload(transaction_payload)
            transaction_result = self.save_transaction(transaction_payload)
            if not transaction_result.get("ok"):
                raise ValueError(_clean_text(transaction_result.get("message")) or "Payment transaction was not saved.")
            transaction_id = _clean_text(transaction_result.get("transactionId"))
            
            ledger_id = self._new_id("LED")
            ledger_description = (
                f"Payment applied to invoice {invoice}"
                + (f" ({method})" if method else "")
                + (f" - {notes}" if notes else "")
            )
            self._append_row_to_table(
                TBL_LEDGER,
                {
                    sc.COL_LEDGER_ID: ledger_id,
                    sc.COL_LEDGER_DATE: payment_date,
                    sc.COL_LEDGER_CLIENT_VENDOR: _clean_text(target_row.get(sc.COL_RECV_CLIENT)),
                    sc.COL_LEDGER_DESCRIPTION: ledger_description,
                    sc.COL_LEDGER_CATEGORY: method or "Payment",
                    sc.COL_LEDGER_REFERENCE: invoice,
                    sc.COL_LEDGER_BILLINGS_EXCL_HST: 0.0,
                    sc.COL_LEDGER_HST_COLLECTED: 0.0,
                    sc.COL_LEDGER_EXPENSES_EXCL_HST: 0.0,
                    sc.COL_LEDGER_HST_PAID: 0.0,
                    sc.COL_LEDGER_COLLECTED: payment_amt,
                    sc.COL_LEDGER_WRITE_OFF: 0.0,
                    sc.COL_LEDGER_RECEIVABLE: -payment_amt,
                    sc.COL_LEDGER_TRX_ID: transaction_id,
                    sc.COL_LEDGER_EXTERNAL_REF_ID: reference,
                    sc.COL_LEDGER_ORIGINAL_AMOUNT: payment_amt,
                    sc.COL_LEDGER_WORK_CLIENT: _clean_text(target_row.get(sc.COL_RECV_WORK_CLIENT)),
                    sc.COL_LEDGER_CREATED_AT: datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                },
            )
            
        if adj_amt > 0:
            ledger_id = ledger_id or self._new_id("LED")
            transaction_id = transaction_id or self._new_id("TXN")
            ledger_description = (
                f"Credit/Adj applied to invoice {invoice}"
                + (f" - {adj_reason}" if adj_reason else "")
            )
            self._append_row_to_table(
                TBL_LEDGER,
                {
                    sc.COL_LEDGER_ID: ledger_id,
                    sc.COL_LEDGER_DATE: payment_date,
                    sc.COL_LEDGER_CLIENT_VENDOR: _clean_text(target_row.get(sc.COL_RECV_CLIENT)),
                    sc.COL_LEDGER_DESCRIPTION: ledger_description,
                    sc.COL_LEDGER_CATEGORY: "Credit/Adj",
                    sc.COL_LEDGER_REFERENCE: invoice,
                    sc.COL_LEDGER_BILLINGS_EXCL_HST: 0.0,
                    sc.COL_LEDGER_HST_COLLECTED: 0.0,
                    sc.COL_LEDGER_EXPENSES_EXCL_HST: 0.0,
                    sc.COL_LEDGER_HST_PAID: 0.0,
                    sc.COL_LEDGER_COLLECTED: 0.0,
                    sc.COL_LEDGER_WRITE_OFF: adj_amt,
                    sc.COL_LEDGER_RECEIVABLE: -adj_amt,
                    sc.COL_LEDGER_TRX_ID: transaction_id,
                    sc.COL_LEDGER_EXTERNAL_REF_ID: reference,
                    sc.COL_LEDGER_ORIGINAL_AMOUNT: adj_amt,
                    sc.COL_LEDGER_WORK_CLIENT: _clean_text(target_row.get(sc.COL_RECV_WORK_CLIENT)),
                    sc.COL_LEDGER_CREATED_AT: datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                },
            )

        next_paid = before_paid + payment_amt
        next_credits = before_credits + adj_amt
        next_balance = self._money_round(invoice_total - next_paid - next_credits)
        if abs(next_balance) <= 0.01:
            next_balance = 0.0
        next_status = "Paid" if next_balance <= 0 else "Partial"

        updated_receivable = dict(target_row)
        updated_receivable[sc.COL_RECV_AMOUNT_PAID] = self._money_round(next_paid)
        updated_receivable[sc.COL_RECV_CREDITS_ADJ] = self._money_round(next_credits)
        updated_receivable[sc.COL_RECV_BALANCE_DUE] = self._money_round(next_balance)
        updated_receivable[sc.COL_RECV_STATUS] = next_status
        self._upsert_row_by_key(TBL_RECEIVABLES, sc.COL_RECV_INVOICE_NUM, invoice, updated_receivable)

        time_rows = self._read_table_rows(TBL_TIME)
        updated_time_rows: List[Dict[str, Any]] = []
        touched_time_entries = 0
        for row in time_rows:
            next_row = dict(row)
            if _clean_text(next_row.get(sc.COL_TIME_INVOICE_REF)).lower() == invoice.lower():
                next_row[sc.COL_TIME_PAYMENT_STATUS] = next_status
                next_row[sc.COL_TIME_INVOICE_TOTAL] = invoice_total
                next_row[sc.COL_TIME_INVOICE_AMOUNT_PAID] = self._money_round(next_paid + next_credits)
                next_row[sc.COL_TIME_INVOICE_BALANCE_DUE] = self._money_round(next_balance)
                next_row[sc.COL_TIME_INVOICE_DATE] = _clean_text(target_row.get(sc.COL_RECV_DATE))
                touched_time_entries += 1
            updated_time_rows.append(next_row)
        if touched_time_entries:
            self._replace_table_rows(TBL_TIME, updated_time_rows)

        updated_invoice = self._payment_invoice_snapshot(updated_receivable)
        
        mode_str = "Payment"
        if payment_amt > 0 and adj_amt > 0:
            mode_str = "Payment & Adjustment"
        elif adj_amt > 0:
            mode_str = "Adjustment"
            
        return {
            "ok": True,
            "mode": mode_str,
            "invoice": invoice,
            "amount": total_applied,
            "paymentId": transaction_id,
            "transactionId": transaction_id,
            "ledgerId": "",
            "beforeBalance": before_balance,
            "afterBalance": self._money_round(next_balance),
            "touchedTimeEntries": touched_time_entries,
            "invoiceRow": updated_invoice,
            "message": f"Transaction posted to invoice {invoice}.",
        }

    def _payment_invoice_snapshot(self, row: Dict[str, Any]) -> Dict[str, Any]:
        invoice = _clean_text(row.get(sc.COL_RECV_INVOICE_NUM))
        invoice_total = self._money_round(row.get(sc.COL_RECV_TOTAL_INVOICED))
        paid = self._money_round(row.get(sc.COL_RECV_AMOUNT_PAID))
        credits = self._money_round(row.get(sc.COL_RECV_CREDITS_ADJ))
        balance = self._money_round(row.get(sc.COL_RECV_BALANCE_DUE))
        client = _clean_text(row.get(sc.COL_RECV_WORK_CLIENT)) or _clean_text(row.get(sc.COL_RECV_CLIENT))
        return {
            "invoice": invoice,
            "invoiceNumber": invoice,
            "date": self._date_iso(row.get(sc.COL_RECV_DATE)),
            "client": client,
            "billingClient": _clean_text(row.get(sc.COL_RECV_CLIENT)),
            "workClient": _clean_text(row.get(sc.COL_RECV_WORK_CLIENT)) or client,
            "status": _clean_text(row.get(sc.COL_RECV_STATUS)),
            "invoiceTotal": invoice_total,
            "paid": self._money_round(paid + credits),
            "cashPaid": paid,
            "credits": credits,
            "balance": balance,
            "display": f"{invoice} | {client} | ${balance:,.2f}",
        }

    def _money_round(self, value: Any) -> float:
        try:
            return float(Decimal(str(value or 0.0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
        except Exception:
            return 0.0

    def _deadlines_path(self) -> Path:
        """JSON file where deadlines are persisted."""
        try:
            base = Path(self.paths.data_dir())
        except Exception:
            base = (
                Path(self.paths.workbook_path()).parent
                if hasattr(self.paths, "workbook_path")
                else Path(".")
            )
        p = base / "deadlines.json"
        p.parent.mkdir(parents=True, exist_ok=True)
        return p

    def _load_deadlines(self) -> list:
        p = self._deadlines_path()
        if not p.exists():
            return []
        try:
            import json
            raw = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return []
        if not isinstance(raw, list):
            return []
        normalized: List[Dict[str, Any]] = []
        for item in raw:
            if isinstance(item, dict):
                normalized.append(self._normalize_deadline_entry(item))
        return self._sort_deadline_entries(normalized)

    def _write_deadlines(self, entries: list) -> None:
        p = self._deadlines_path()
        import json
        p.write_text(json.dumps(entries, indent=2), encoding="utf-8")

    def _now_utc_iso(self) -> str:
        return datetime.now(UTC).replace(microsecond=0, tzinfo=None).isoformat() + "Z"

    def _normalize_deadline_entry(self, entry: Dict[str, Any]) -> Dict[str, Any]:
        source = dict(entry or {})
        now_iso = self._now_utc_iso()
        entry_id = _clean_text(source.get("id")) or str(uuid4())
        date_text = _clean_text(source.get("date"))
        if not _is_valid_iso_date(date_text):
            date_text = datetime.now(UTC).strftime("%Y-%m-%d")
        description = _clean_text(source.get("description"))
        created_at = _clean_text(source.get("createdAt")) or now_iso
        updated_at = _clean_text(source.get("updatedAt")) or now_iso
        raw_escalated = source.get("escalated")
        if isinstance(raw_escalated, bool):
            escalated = raw_escalated
        elif isinstance(raw_escalated, (int, float)):
            escalated = int(raw_escalated) != 0
        else:
            escalated = _clean_text(raw_escalated).lower() in {"1", "true", "yes", "y", "on", "escalated", "urgent"}

        raw_completed = source.get("completed")
        if isinstance(raw_completed, bool):
            completed = raw_completed
        elif isinstance(raw_completed, (int, float)):
            completed = int(raw_completed) != 0
        else:
            completed_text = _clean_text(raw_completed).lower()
            completed = completed_text in {"1", "true", "yes", "y", "on", "done", "complete", "completed", "closed"}
            if not completed and _clean_text(source.get("status")).lower() in {"done", "complete", "completed", "closed"}:
                completed = True

        assignment_type = _clean_text(
            source.get("assignmentType")
            or source.get("assignment")
            or source.get("scope")
        )
        matter_name = _clean_text(source.get("matterName") or source.get("matter"))
        matter_id = _clean_text(source.get("matterId"))
        client_name = _clean_text(source.get("clientName") or source.get("client"))
        if not assignment_type:
            assignment_type = "Matter" if matter_name else "General"
        assignment_type = "Matter" if assignment_type.lower() == "matter" else "General"
        if assignment_type != "Matter":
            matter_name = ""
            matter_id = ""

        completed_at = _clean_text(source.get("completedAt"))
        if completed and not completed_at:
            completed_at = now_iso
        if not completed:
            completed_at = ""

        source_type = _clean_text(source.get("sourceType") or source.get("source"))
        source_id = _clean_text(source.get("sourceId"))
        source_field = _clean_text(source.get("sourceField"))
        source_label = _clean_text(source.get("sourceLabel"))
        raw_entry_type = _clean_text(source.get("entryType") or source.get("category") or source.get("type"))
        entry_type_lc = raw_entry_type.lower()
        if entry_type_lc in {"task", "reminder", "tickler", "todo", "to-do"}:
            entry_type = "Task"
        elif entry_type_lc in {"information", "information only", "info", "info-only"}:
            entry_type = "Information"
        else:
            entry_type = "Deadline"

        # Task-specific fields
        work_date = _clean_text(source.get("workDate") or source.get("workOnDate") or source.get("startDate"))
        if work_date and not _is_valid_iso_date(work_date):
            work_date = ""

        raw_priority = _clean_text(source.get("priority") or source.get("urgency")).lower()
        if raw_priority in {"high", "urgent", "critical"}:
            priority = "High"
        elif raw_priority in {"low", "minor"}:
            priority = "Low"
        else:
            priority = "Normal"

        reminder_note = _clean_text(source.get("reminderNote") or source.get("reminder") or source.get("note"))

        # Recurrence fields (Phase 2 prep — stored now, expanded later)
        raw_recurrence = _clean_text(source.get("recurrence") or source.get("repeat")).lower()
        if raw_recurrence in {"daily", "weekly", "biweekly", "monthly", "yearly", "custom"}:
            recurrence = raw_recurrence.capitalize()
            if recurrence == "Biweekly":
                recurrence = "Biweekly"
        else:
            recurrence = "None"
        recurrence_interval = 1
        try:
            recurrence_interval = max(1, int(source.get("recurrenceInterval") or 1))
        except (ValueError, TypeError):
            recurrence_interval = 1
        recurrence_end_date = _clean_text(source.get("recurrenceEndDate") or source.get("repeatUntil"))
        if recurrence_end_date and not _is_valid_iso_date(recurrence_end_date):
            recurrence_end_date = ""
        recurrence_parent_id = _clean_text(source.get("recurrenceParentId"))

        raw_auto_generated = source.get("autoGenerated")
        if isinstance(raw_auto_generated, bool):
            auto_generated = raw_auto_generated
        elif isinstance(raw_auto_generated, (int, float)):
            auto_generated = int(raw_auto_generated) != 0
        else:
            auto_generated = _clean_text(raw_auto_generated).lower() in {"1", "true", "yes", "y", "on", "auto", "generated"}
        return {
            "id": entry_id,
            "date": date_text,
            "description": description,
            "escalated": bool(escalated),
            "completed": bool(completed),
            "completedAt": completed_at,
            "assignmentType": assignment_type,
            "matterId": matter_id,
            "matterName": matter_name,
            "clientName": client_name,
            "entryType": entry_type,
            "workDate": work_date,
            "priority": priority,
            "reminderNote": reminder_note,
            "recurrence": recurrence,
            "recurrenceInterval": recurrence_interval,
            "recurrenceEndDate": recurrence_end_date,
            "recurrenceParentId": recurrence_parent_id,
            "sourceType": source_type,
            "sourceId": source_id,
            "sourceField": source_field,
            "sourceLabel": source_label,
            "autoGenerated": bool(auto_generated),
            "createdAt": created_at,
            "updatedAt": updated_at,
        }

    def _sort_deadline_entries(self, entries: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        def _sort_key(e: Dict[str, Any]):
            # For tasks with a workDate, use the earlier of workDate and date
            # so they surface when work is due, not just when the deadline hits.
            due_date = _clean_text(e.get("date")) or "9999-12-31"
            work_date = _clean_text(e.get("workDate"))
            effective_date = min(due_date, work_date) if work_date else due_date
            return (
                effective_date,
                1 if bool(e.get("completed", False)) else 0,
                0 if bool(e.get("escalated", False)) else 1,
                # Priority ordering: High < Normal < Low (High surfaces first)
                0 if _clean_text(e.get("priority")).lower() == "high" else (2 if _clean_text(e.get("priority")).lower() == "low" else 1),
                0 if _clean_text(e.get("assignmentType")).lower() == "matter" else 1,
                _clean_text(e.get("matterName")).lower(),
                _clean_text(e.get("description")).lower(),
                _clean_text(e.get("id")),
            )
        return sorted(entries, key=_sort_key)

    def list_deadline_entries(self) -> list:
        """Return all deadline records."""
        if not bool(getattr(self, "_deadline_generation_bootstrap_done", False)):
            self._sync_all_trademark_generated_deadlines()
            self._deadline_generation_bootstrap_done = True
        return self._sort_deadline_entries(self._load_deadlines())

    def create_deadline_entry(self, entry: Dict[str, Any]) -> Dict[str, Any]:
        """Add a new deadline and return the saved object (with id)."""
        entries = self._load_deadlines()
        new = self._normalize_deadline_entry(entry)
        now_iso = self._now_utc_iso()
        new["createdAt"] = now_iso
        new["updatedAt"] = now_iso
        entries.append(new)
        entries = self._sort_deadline_entries(entries)
        self._write_deadlines(entries)
        return dict(new)

    def update_deadline_entry(self, entry_id: str, changes: Dict[str, Any]) -> Dict[str, Any] | None:
        entries = self._load_deadlines()
        target_id = _clean_text(entry_id)
        for idx, existing in enumerate(entries):
            if _clean_text(existing.get("id")) == target_id:
                merged = dict(existing)
                merged.update(dict(changes or {}))
                normalized = self._normalize_deadline_entry(merged)
                normalized["id"] = target_id or normalized["id"]
                normalized["createdAt"] = _clean_text(existing.get("createdAt")) or normalized.get("createdAt")
                normalized["updatedAt"] = self._now_utc_iso()
                entries[idx] = normalized
                entries = self._sort_deadline_entries(entries)
                self._write_deadlines(entries)
                return dict(normalized)
        return None

    def delete_deadline_entry(self, entry_id: str) -> bool:
        entries = self._load_deadlines()
        target_id = _clean_text(entry_id)
        new_list = [e for e in entries if _clean_text(e.get("id")) != target_id]
        if len(new_list) != len(entries):
            self._write_deadlines(new_list)
            return True
        return False

    def _resolve_trademark_deadline_assignment(
        self,
        trademark_row: Dict[str, Any],
    ) -> Tuple[str, str, str, str]:
        matter_ref = _clean_text(trademark_row.get(sc.COL_TM_MATTER_NUMBER))
        if not matter_ref:
            return ("General", "", "", "")
        lookup = matter_ref.lower()
        for row in self.list_matter_directory():
            matter_id = _clean_text(row.get("matterId"))
            matter_number = _clean_text(row.get("matterNumber"))
            matter_name = _clean_text(row.get("matterName"))
            display_name = _clean_text(row.get("displayName")) or matter_name
            client_name = _clean_text(row.get("clientName"))
            keys = {
                matter_id.lower(),
                matter_number.lower(),
                matter_name.lower(),
                display_name.lower(),
            }
            if lookup in keys:
                return ("Matter", matter_id, display_name or matter_name, client_name)
        return ("General", "", "", "")

    def _build_trademark_deadline_candidates(
        self,
        trademark_row: Dict[str, Any],
    ) -> List[Dict[str, Any]]:
        trademark_id = _clean_text(trademark_row.get(sc.COL_TM_ID))
        if not trademark_id:
            return []

        title = (
            _clean_text(trademark_row.get(sc.COL_TM_TRADEMARK_TEXT))
            or _clean_text(trademark_row.get(sc.COL_TM_APPLICATION_NO))
            or _clean_text(trademark_row.get(sc.COL_TM_REGISTRATION_NO))
            or f"Trademark {trademark_id}"
        )
        jurisdiction = _clean_text(trademark_row.get(sc.COL_TM_JURISDICTION)) or _clean_text(
            trademark_row.get(sc.COL_TM_JURISDICTION_OTHER)
        )
        application_no = _clean_text(trademark_row.get(sc.COL_TM_APPLICATION_NO))
        registration_no = _clean_text(trademark_row.get(sc.COL_TM_REGISTRATION_NO))
        matter_ref = _clean_text(trademark_row.get(sc.COL_TM_MATTER_NUMBER))
        assignment_type, matter_id, matter_name, matter_client_name = self._resolve_trademark_deadline_assignment(trademark_row)
        trademark_client_name = _clean_text(trademark_row.get(sc.COL_TM_CLIENT_NAME))
        client_name = matter_client_name or trademark_client_name

        deadline_fields: List[Tuple[str, str, str]] = [
            ("renewalDeadline", "Renewal Deadline", sc.COL_TM_RENEWAL_DEADLINE),
            ("officeActionResponseDeadline", "Office Action Response Deadline", sc.COL_TM_OFFICE_ACTION_RESPONSE_DEADLINE),
            ("oppositionDeadline", "Opposition Deadline", sc.COL_TM_OPPOSITION_DEADLINE),
            ("souDeadline", "Statement of Use Deadline", sc.COL_TM_SOU_DEADLINE),
            ("section8Deadline", "Section 8 Maintenance Deadline", sc.COL_TM_SECTION8_DEADLINE),
            ("section15Deadline", "Section 15 Maintenance Deadline", sc.COL_TM_SECTION15_DEADLINE),
            ("section9Deadline", "Section 9 Renewal Deadline", sc.COL_TM_SECTION9_DEADLINE),
            ("oppositionPeriodEndDate", "Opposition Period End Date", sc.COL_TM_OPPOSITION_PERIOD_END_DATE),
            ("upcomingLocalDeadlineOfficeActionDate", "Upcoming Local Deadline / Office Action Date", sc.COL_TM_UPCOMING_LOCAL_DEADLINE),
        ]
        info_fields: List[Tuple[str, str, str]] = [
            ("filingDate", "Filing Date", sc.COL_TM_FILING_DATE),
            ("registrationDate", "Registration Date", sc.COL_TM_REGISTRATION_DATE),
            ("examinersReportDate", "Examiner's Report Date", sc.COL_TM_EXAMINERS_REPORT_DATE),
            ("approvalDate", "Approval Date", sc.COL_TM_APPROVAL_DATE),
            ("advertisementDate", "Advertisement Date", sc.COL_TM_ADVERTISEMENT_DATE),
            ("allowanceDate", "Allowance Date", sc.COL_TM_ALLOWANCE_DATE),
            ("publicationDate", "Publication Date", sc.COL_TM_PUBLICATION_DATE),
            ("noticeOfAllowanceDate", "Notice of Allowance Date", sc.COL_TM_NOTICE_OF_ALLOWANCE_DATE),
            ("publicationAdvertisementDate", "Publication/Advertisement Date", sc.COL_TM_PUBLICATION_ADVERTISEMENT_DATE),
        ]

        out: List[Dict[str, Any]] = []
        for source_field, label, col_key in deadline_fields:
            due_date = _clean_text(trademark_row.get(col_key))
            if not _is_valid_iso_date(due_date):
                continue
            detail_bits: List[str] = []
            if jurisdiction:
                detail_bits.append(jurisdiction)
            if application_no:
                detail_bits.append(f"App {application_no}")
            elif registration_no:
                detail_bits.append(f"Reg {registration_no}")
            if matter_ref:
                detail_bits.append(f"Matter {matter_ref}")

            description = f"{title} | {label}"
            if detail_bits:
                description += " | " + " | ".join(detail_bits)

            out.append(
                {
                    "date": due_date,
                    "description": description,
                    "escalated": False,
                    "completed": False,
                    "assignmentType": assignment_type,
                    "matterId": matter_id if assignment_type == "Matter" else "",
                    "matterName": matter_name if assignment_type == "Matter" else "",
                    "clientName": client_name,
                    "entryType": "Deadline",
                    "sourceType": "Trademark",
                    "sourceId": trademark_id,
                    "sourceField": source_field,
                    "sourceLabel": label,
                    "autoGenerated": True,
                }
            )
        for source_field, label, col_key in info_fields:
            info_date = _clean_text(trademark_row.get(col_key))
            if not _is_valid_iso_date(info_date):
                continue
            detail_bits: List[str] = []
            if jurisdiction:
                detail_bits.append(jurisdiction)
            if application_no:
                detail_bits.append(f"App {application_no}")
            elif registration_no:
                detail_bits.append(f"Reg {registration_no}")
            if matter_ref:
                detail_bits.append(f"Matter {matter_ref}")
            description = f"{title} | {label}"
            if detail_bits:
                description += " | " + " | ".join(detail_bits)
            out.append(
                {
                    "date": info_date,
                    "description": description,
                    "escalated": False,
                    "completed": False,
                    "assignmentType": assignment_type,
                    "matterId": matter_id if assignment_type == "Matter" else "",
                    "matterName": matter_name if assignment_type == "Matter" else "",
                    "clientName": client_name,
                    "entryType": "Information",
                    "sourceType": "Trademark",
                    "sourceId": trademark_id,
                    "sourceField": source_field,
                    "sourceLabel": label,
                    "autoGenerated": True,
                }
            )
        return out

    def _sync_all_trademark_generated_deadlines(self) -> None:
        try:
            rows = [self._canonicalize_trademark_row(r) for r in self._read_table_rows(TBL_TRADEMARKS)]
        except Exception:
            return
        for row in rows:
            try:
                self._sync_trademark_generated_deadlines(row)
            except Exception:
                continue

    def _sync_trademark_generated_deadlines(self, trademark_row: Dict[str, Any]) -> None:
        trademark_id = _clean_text(trademark_row.get(sc.COL_TM_ID))
        if not trademark_id:
            return

        candidates = self._build_trademark_deadline_candidates(trademark_row)
        candidate_by_field: Dict[str, Dict[str, Any]] = {
            _clean_text(c.get("sourceField")).lower(): c for c in candidates if _clean_text(c.get("sourceField"))
        }

        existing = self._load_deadlines()
        retained: List[Dict[str, Any]] = []
        generated: List[Dict[str, Any]] = []

        for entry in existing:
            entry_source_type = _clean_text(entry.get("sourceType")).lower()
            entry_source_id = _clean_text(entry.get("sourceId")).lower()
            if entry_source_type != "trademark" or entry_source_id != trademark_id.lower():
                retained.append(entry)
                continue

            source_field = _clean_text(entry.get("sourceField")).lower()
            if not source_field:
                # Older generated entries without explicit source-field keys are dropped
                # so current trademark deadlines can be regenerated deterministically.
                continue

            candidate = candidate_by_field.pop(source_field, None)
            if not candidate:
                # Field removed/blanked on trademark record; remove stale generated entry.
                continue

            merged = dict(entry)
            merged.update(candidate)
            merged["id"] = _clean_text(entry.get("id")) or _clean_text(merged.get("id"))
            merged["createdAt"] = _clean_text(entry.get("createdAt")) or _clean_text(merged.get("createdAt"))
            merged["escalated"] = bool(entry.get("escalated", False))
            merged["completed"] = bool(entry.get("completed", False))
            if merged["completed"]:
                merged["completedAt"] = _clean_text(entry.get("completedAt")) or self._now_utc_iso()
            else:
                merged["completedAt"] = ""
            merged["updatedAt"] = self._now_utc_iso()
            generated.append(self._normalize_deadline_entry(merged))

        for candidate in candidate_by_field.values():
            generated.append(self._normalize_deadline_entry(candidate))

        all_entries = self._sort_deadline_entries(retained + generated)
        self._write_deadlines(all_entries)

    def search_global_entities(self, query: str, mode: str = "any", limit: int = 250) -> Dict[str, Any]:
        raw_query = _clean_text(query)
        normalized_query = _normalize_search_text(raw_query)
        normalized_mode = _clean_text(mode).lower()
        search_mode = "boolean" if normalized_mode == "boolean" else "any"
        try:
            max_results = int(limit)
        except Exception:
            max_results = 250
        max_results = max(1, min(max_results, 1000))

        facets: Dict[str, int] = {
            "client": 0,
            "matter": 0,
            "parent": 0,
            "invoice": 0,
            "transaction": 0,
            "account": 0,
            "category": 0,
            "business_unit": 0,
            "payee": 0,
            "tickler": 0,
            "deadline": 0,
            "docket": 0,
            "trademark": 0,
        }
        entity_labels = {
            "client": "Client",
            "matter": "Matter",
            "parent": "Parent",
            "invoice": "Invoice",
            "transaction": "Transaction",
            "account": "Account",
            "category": "Category",
            "business_unit": "Business Unit",
            "payee": "Payee",
            "tickler": "Tickler",
            "deadline": "Deadline",
            "docket": "Docket",
            "trademark": "Trademark",
        }

        empty_payload = {
            "ok": True,
            "query": raw_query,
            "mode": search_mode,
            "results": [],
            "facets": dict(facets),
            "total": 0,
            "returnedCount": 0,
        }
        if not normalized_query:
            return empty_payload

        any_terms = _search_terms(normalized_query)
        boolean_tokens = _boolean_query_tokens(raw_query)
        boolean_postfix = _boolean_to_postfix(boolean_tokens) if search_mode == "boolean" else []
        boolean_terms = [t for t in _search_terms(raw_query) if t not in ("and", "or", "not")]
        if search_mode == "boolean" and not boolean_postfix:
            search_mode = "any"

        try:
            self.ensure_schema()
            client_rows = [self._canonicalize_client_row(r) for r in self._read_table_rows(TBL_CLIENTS)]
            profile_rows = [
                self._canonicalize_client_profile_row(r)
                for r in self._read_table_rows(TBL_CLIENT_PROFILES)
            ]
            parent_rows = [self._canonicalize_parent_row(r) for r in self._read_table_rows(TBL_PARENTS)]
            matter_rows = [self._canonicalize_matter_row(r) for r in self._read_table_rows(TBL_MATTERS)]
            time_rows = [self._canonicalize_time_row(r) for r in self._read_table_rows(TBL_TIME)]
            trademark_rows = [self._canonicalize_trademark_row(r) for r in self._read_table_rows(TBL_TRADEMARKS)]
            txn_rows = [
                self._canonicalize_transaction_row(r)
                for r in self._read_table_rows(TBL_TRANSACTIONS_MASTER)
            ]
            account_rows = [
                self._canonicalize_transaction_account_row(r)
                for r in self._read_table_rows(TBL_TRANSACTION_ACCOUNTS)
            ]
            category_rows = [
                self._canonicalize_transaction_category_row(r)
                for r in self._read_table_rows(TBL_TRANSACTION_CATEGORIES)
            ]
            business_unit_rows = [
                self._canonicalize_transaction_business_unit_row(r)
                for r in self._read_table_rows(TBL_TRANSACTION_BUSINESS_UNITS)
            ]
            payee_rows = [
                self._canonicalize_transaction_payee_row(r)
                for r in self._read_table_rows(TBL_TRANSACTION_PAYEES)
            ]
            receivable_rows = [
                self._canonicalize_row(TBL_RECEIVABLES, r)
                for r in self._read_table_rows(TBL_RECEIVABLES)
            ]
        except Exception as exc:
            return {
                "ok": False,
                "query": raw_query,
                "mode": search_mode,
                "results": [],
                "facets": dict(facets),
                "total": 0,
                "returnedCount": 0,
                "message": str(exc),
            }

        profile_by_client_id: Dict[str, Dict[str, Any]] = {}
        for row in profile_rows:
            client_id = _clean_text(row.get(sc.COL_PROFILE_CLIENT_ID))
            if client_id:
                profile_by_client_id[client_id] = row

        client_name_by_id: Dict[str, str] = {}
        for row in client_rows:
            client_id = _clean_text(row.get(sc.COL_CLIENT_ID))
            client_name = _clean_text(row.get(sc.COL_CLIENT_NAME))
            if client_id and client_name:
                client_name_by_id[client_id] = client_name

        parent_name_by_id: Dict[str, str] = {}
        for row in parent_rows:
            parent_id = _clean_text(row.get(sc.COL_PARENT_ID))
            parent_name = _clean_text(row.get(sc.COL_PARENT_NAME))
            if parent_id and parent_name:
                parent_name_by_id[parent_id] = parent_name

        matter_name_by_id: Dict[str, str] = {}
        for row in matter_rows:
            matter_id = _clean_text(row.get(sc.COL_MATTER_ID))
            matter_name = _clean_text(row.get(sc.COL_MATTER_DISPLAY_NAME)) or _clean_text(
                row.get(sc.COL_MATTER_NAME)
            )
            if matter_id and matter_name:
                matter_name_by_id[matter_id] = matter_name

        results: List[Dict[str, Any]] = []

        def _field_weight(label: str) -> int:
            normalized_label = _normalize_search_text(label)
            if any(token in normalized_label for token in ("notes", "description", "details", "audit")):
                return 4
            if normalized_label in {
                "client name",
                "display name",
                "legal name",
                "principal name",
                "parent name",
                "matter name",
                "account name",
                "category name",
                "business unit",
                "payee",
                "payee name",
                "trademark text",
            }:
                return 34
            if any(
                token in normalized_label
                for token in (
                    "email",
                    "phone",
                    "responsible lawyer",
                    "billing contact",
                    "member",
                    "owner",
                    "applicant",
                )
            ):
                return 20
            if any(token in normalized_label for token in ("id", "number", "code", "ref")):
                return 18
            if any(token in normalized_label for token in ("status", "type", "class", "scope", "tax")):
                return 10
            return 12

        def _match_fields(field_pairs: List[Tuple[str, Any]], title_text: str) -> Optional[Tuple[int, List[str]]]:
            normalized_fields: List[Tuple[str, str]] = []
            for label, raw_value in field_pairs:
                value_text = _clean_text(raw_value)
                if not value_text:
                    continue
                normalized_value = _normalize_search_text(value_text)
                if not normalized_value:
                    continue
                normalized_fields.append((str(label), normalized_value))
            if not normalized_fields:
                return None

            haystack = " | ".join(v for _, v in normalized_fields if v)
            if not haystack:
                return None

            active_terms = any_terms
            if search_mode == "boolean" and boolean_postfix:
                if not _match_boolean_postfix(boolean_postfix, haystack):
                    return None
                active_terms = boolean_terms
            else:
                if not any(term and term in haystack for term in any_terms):
                    return None

            matched_terms = 0
            seen_terms = set()
            for term in active_terms:
                if not term or term in seen_terms:
                    continue
                if term in haystack:
                    matched_terms += 1
                    seen_terms.add(term)

            matched_fields: List[str] = []
            seen_fields = set()
            for label, value_text in normalized_fields:
                for term in active_terms:
                    if term and term in value_text:
                        key = label.lower()
                        if key not in seen_fields:
                            seen_fields.add(key)
                            matched_fields.append(label)
                        break

            title_lc = _normalize_search_text(title_text)
            score = matched_terms * 6
            for term in active_terms:
                if not term:
                    continue
                if title_lc.startswith(term):
                    score += 28
                elif term in title_lc:
                    score += 14

            for label, value_text in normalized_fields:
                weight = _field_weight(label)
                for term in active_terms:
                    if not term or term not in value_text:
                        continue
                    score += weight
                    if value_text == term:
                        score += 16
                    elif value_text.startswith(term):
                        score += 8
                    break

            return score, matched_fields[:4]

        def _append_result(
            *,
            entity_type: str,
            entity_id: str,
            title: str,
            subtitle: str,
            status: str,
            route_tile_index: int,
            route_node_id: str,
            route_node_title: str,
            field_pairs: List[Tuple[str, Any]],
            extra: Optional[Dict[str, Any]] = None,
        ) -> None:
            title_text = _clean_text(title) or _clean_text(entity_id) or "Untitled"
            match_meta = _match_fields(field_pairs, title_text)
            if not match_meta:
                return
            score, matched_fields = match_meta

            row = {
                "entityType": entity_type,
                "entityTypeLabel": entity_labels.get(entity_type, entity_type.title()),
                "entityId": _clean_text(entity_id),
                "title": title_text,
                "subtitle": _clean_text(subtitle),
                "status": _clean_text(status),
                "routeTileIndex": int(route_tile_index),
                "routeNodeId": _clean_text(route_node_id),
                "routeNodeTitle": _clean_text(route_node_title),
                "matchedFields": matched_fields,
                "_score": int(score),
            }
            if extra:
                row.update(extra)
            results.append(row)
            if entity_type in facets:
                facets[entity_type] += 1

        seen_client_ids = set()
        for client_row in client_rows:
            client_id = _clean_text(client_row.get(sc.COL_CLIENT_ID))
            if client_id:
                seen_client_ids.add(client_id.lower())
            profile_row = profile_by_client_id.get(client_id, {})

            client_name = _clean_text(client_row.get(sc.COL_CLIENT_NAME))
            display_name = _clean_text(profile_row.get(sc.COL_PROFILE_DISPLAY_NAME)) or client_name
            legal_name = _clean_text(profile_row.get(sc.COL_PROFILE_LEGAL_NAME))
            status = _clean_text(client_row.get(sc.COL_CLIENT_STATUS)) or "Active"
            primary_email = _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_EMAIL)) or _clean_text(
                client_row.get(sc.COL_CLIENT_EMAIL)
            )
            primary_phone = _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_PHONE)) or _clean_text(
                client_row.get(sc.COL_CLIENT_PHONE)
            )

            subtitle_bits = [x for x in (legal_name, primary_email, status) if x]
            field_pairs: List[Tuple[str, Any]] = [
                ("Client ID", client_id),
                ("Client Name", client_name),
                ("Display Name", display_name),
                ("Legal Name", legal_name),
                ("First Name", profile_row.get(sc.COL_PROFILE_FIRST_NAME)),
                ("Middle Name", profile_row.get(sc.COL_PROFILE_MIDDLE_NAME)),
                ("Last Name", profile_row.get(sc.COL_PROFILE_LAST_NAME)),
                ("Client Status", status),
                ("Entity Type", profile_row.get(sc.COL_PROFILE_ENTITY_TYPE)),
                ("Principal Name", profile_row.get(sc.COL_PROFILE_PRINCIPAL_NAME)),
                ("Principal Position", profile_row.get(sc.COL_PROFILE_PRINCIPAL_POSITION)),
                ("Primary Email", primary_email),
                ("Primary Phone", primary_phone),
                ("Secondary Contact", profile_row.get(sc.COL_PROFILE_SECONDARY_CONTACT)),
                ("Secondary Position", profile_row.get(sc.COL_PROFILE_SECONDARY_POSITION)),
                ("Secondary Email", profile_row.get(sc.COL_PROFILE_SECONDARY_EMAIL)),
                ("Secondary Phone", profile_row.get(sc.COL_PROFILE_SECONDARY_PHONE)),
                ("Address Line 1", profile_row.get(sc.COL_PROFILE_ADDR1)),
                ("Address Line 2", profile_row.get(sc.COL_PROFILE_ADDR2)),
                ("City", profile_row.get(sc.COL_PROFILE_CITY)),
                ("State/Province", profile_row.get(sc.COL_PROFILE_STATE)),
                ("Postal Code", profile_row.get(sc.COL_PROFILE_POSTAL)),
                ("Country", profile_row.get(sc.COL_PROFILE_COUNTRY)),
                ("Full Address", profile_row.get(sc.COL_PROFILE_FULL_ADDRESS)),
                ("Parent Client ID", profile_row.get(sc.COL_PROFILE_PARENT_ID)),
                ("Parent Client Name", profile_row.get(sc.COL_PROFILE_PARENT_NAME)),
                ("Website", profile_row.get(sc.COL_PROFILE_WEBSITE)),
                ("Tax ID", profile_row.get(sc.COL_PROFILE_TAX_ID)),
                ("Industry", profile_row.get(sc.COL_PROFILE_INDUSTRY)),
                ("Billing Email", profile_row.get(sc.COL_PROFILE_BILLING_EMAIL)),
                ("KYC Status", profile_row.get(sc.COL_PROFILE_KYC_STATUS)),
                ("Onboarding Status", profile_row.get(sc.COL_PROFILE_ONBOARDING_STATUS)),
                ("Retainer Required", profile_row.get(sc.COL_PROFILE_RETAINER_REQUIRED)),
                ("Retainer Amount", profile_row.get(sc.COL_PROFILE_RETAINER_AMOUNT)),
                ("Engagement Start Date", profile_row.get(sc.COL_PROFILE_ENGAGEMENT_START)),
                ("Date Client Added", profile_row.get(sc.COL_PROFILE_DATE_CLIENT_ADDED)),
                ("Birthday", profile_row.get(sc.COL_PROFILE_BIRTHDAY)),
                ("Referral From", profile_row.get(sc.COL_PROFILE_REFERRAL_FROM)),
                ("Conflict Notes", profile_row.get(sc.COL_PROFILE_CONFLICT_NOTES)),
                ("Client Notes", profile_row.get(sc.COL_PROFILE_NOTES) or client_row.get(sc.COL_CLIENT_NOTES)),
            ]
            _append_result(
                entity_type="client",
                entity_id=client_id,
                title=display_name or client_name or legal_name or client_id,
                subtitle=" | ".join(subtitle_bits[:3]),
                status=status,
                route_tile_index=0,
                route_node_id="A03",
                route_node_title="Client Profile 360",
                field_pairs=field_pairs,
                extra={
                    "clientId": client_id,
                    "clientName": client_name,
                    "legalName": legal_name,
                    "firstName": _clean_text(profile_row.get(sc.COL_PROFILE_FIRST_NAME)),
                    "middleName": _clean_text(profile_row.get(sc.COL_PROFILE_MIDDLE_NAME)),
                    "lastName": _clean_text(profile_row.get(sc.COL_PROFILE_LAST_NAME)),
                    "principalName": _clean_text(profile_row.get(sc.COL_PROFILE_PRINCIPAL_NAME)),
                    "primaryEmail": primary_email,
                    "primaryPhone": primary_phone,
                },
            )

        for profile_row in profile_rows:
            client_id = _clean_text(profile_row.get(sc.COL_PROFILE_CLIENT_ID))
            if client_id and client_id.lower() in seen_client_ids:
                continue
            display_name = _clean_text(profile_row.get(sc.COL_PROFILE_DISPLAY_NAME))
            legal_name = _clean_text(profile_row.get(sc.COL_PROFILE_LEGAL_NAME))
            status = "Active"
            subtitle_bits = [
                x
                for x in (
                    legal_name,
                    _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_EMAIL)),
                    status,
                )
                if x
            ]
            field_pairs = [
                ("Client ID", client_id),
                ("Display Name", display_name),
                ("Legal Name", legal_name),
                ("First Name", profile_row.get(sc.COL_PROFILE_FIRST_NAME)),
                ("Middle Name", profile_row.get(sc.COL_PROFILE_MIDDLE_NAME)),
                ("Last Name", profile_row.get(sc.COL_PROFILE_LAST_NAME)),
                ("Entity Type", profile_row.get(sc.COL_PROFILE_ENTITY_TYPE)),
                ("Primary Email", profile_row.get(sc.COL_PROFILE_PRIMARY_EMAIL)),
                ("Primary Phone", profile_row.get(sc.COL_PROFILE_PRIMARY_PHONE)),
                ("Secondary Contact", profile_row.get(sc.COL_PROFILE_SECONDARY_CONTACT)),
                ("Secondary Position", profile_row.get(sc.COL_PROFILE_SECONDARY_POSITION)),
                ("Secondary Email", profile_row.get(sc.COL_PROFILE_SECONDARY_EMAIL)),
                ("Secondary Phone", profile_row.get(sc.COL_PROFILE_SECONDARY_PHONE)),
                ("Full Address", profile_row.get(sc.COL_PROFILE_FULL_ADDRESS)),
                ("Parent Client Name", profile_row.get(sc.COL_PROFILE_PARENT_NAME)),
                ("Website", profile_row.get(sc.COL_PROFILE_WEBSITE)),
                ("Tax ID", profile_row.get(sc.COL_PROFILE_TAX_ID)),
                ("Industry", profile_row.get(sc.COL_PROFILE_INDUSTRY)),
                ("Billing Email", profile_row.get(sc.COL_PROFILE_BILLING_EMAIL)),
                ("KYC Status", profile_row.get(sc.COL_PROFILE_KYC_STATUS)),
                ("Onboarding Status", profile_row.get(sc.COL_PROFILE_ONBOARDING_STATUS)),
                ("Date Client Added", profile_row.get(sc.COL_PROFILE_DATE_CLIENT_ADDED)),
                ("Birthday", profile_row.get(sc.COL_PROFILE_BIRTHDAY)),
                ("Referral From", profile_row.get(sc.COL_PROFILE_REFERRAL_FROM)),
                ("Conflict Notes", profile_row.get(sc.COL_PROFILE_CONFLICT_NOTES)),
                ("Client Notes", profile_row.get(sc.COL_PROFILE_NOTES)),
            ]
            _append_result(
                entity_type="client",
                entity_id=client_id,
                title=display_name or legal_name or client_id,
                subtitle=" | ".join(subtitle_bits[:3]),
                status=status,
                route_tile_index=0,
                route_node_id="A03",
                route_node_title="Client Profile 360",
                field_pairs=field_pairs,
                extra={
                    "clientId": client_id,
                    "clientName": display_name or legal_name,
                    "legalName": legal_name,
                    "firstName": _clean_text(profile_row.get(sc.COL_PROFILE_FIRST_NAME)),
                    "middleName": _clean_text(profile_row.get(sc.COL_PROFILE_MIDDLE_NAME)),
                    "lastName": _clean_text(profile_row.get(sc.COL_PROFILE_LAST_NAME)),
                    "principalName": _clean_text(profile_row.get(sc.COL_PROFILE_PRINCIPAL_NAME)),
                    "primaryEmail": _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_EMAIL)),
                    "primaryPhone": _clean_text(profile_row.get(sc.COL_PROFILE_PRIMARY_PHONE)),
                },
            )

        for parent_row in parent_rows:
            parent_id = _clean_text(parent_row.get(sc.COL_PARENT_ID))
            parent_name = _clean_text(parent_row.get(sc.COL_PARENT_NAME))
            parent_active = self._to_bool_int(parent_row.get(sc.COL_PARENT_ACTIVE), default=1)
            status = "Active" if parent_active else "Inactive"
            _append_result(
                entity_type="parent",
                entity_id=parent_id,
                title=parent_name or parent_id,
                subtitle=" | ".join(
                    x
                    for x in (
                        ("Parent ID: " + parent_id) if parent_id else "",
                        status,
                    )
                    if x
                ),
                status=status,
                route_tile_index=0,
                route_node_id="A05",
                route_node_title="Parent-Child Link Manager",
                field_pairs=[
                    ("Parent ID", parent_id),
                    ("Parent Name", parent_name),
                    ("Default Share %", parent_row.get(sc.COL_PARENT_DEF_SHARE)),
                    ("Default Rate", parent_row.get(sc.COL_PARENT_DEF_RATE)),
                    ("Notes", parent_row.get(sc.COL_PARENT_NOTES)),
                    ("Status", status),
                ],
                extra={"parentId": parent_id, "parentName": parent_name},
            )

        for matter_row in matter_rows:
            matter_id = _clean_text(matter_row.get(sc.COL_MATTER_ID))
            matter_number = _clean_text(matter_row.get(sc.COL_MATTER_NUMBER))
            matter_name = _clean_text(matter_row.get(sc.COL_MATTER_NAME))
            display_name = _clean_text(matter_row.get(sc.COL_MATTER_DISPLAY_NAME)) or matter_name
            status = _clean_text(matter_row.get(sc.COL_MATTER_STATUS)) or "Open"
            client_ref = _clean_text(matter_row.get(sc.COL_MATTER_CLIENT_ID))
            parent_ref = _clean_text(matter_row.get(sc.COL_MATTER_PARENT_ID))
            client_name = _clean_text(matter_row.get(sc.COL_MATTER_CLIENT_NAME)) or client_name_by_id.get(
                client_ref, client_ref
            )
            parent_name = _clean_text(matter_row.get(sc.COL_MATTER_PARENT_NAME)) or parent_name_by_id.get(
                parent_ref, parent_ref
            )
            matter_type = _clean_text(matter_row.get(sc.COL_MATTER_TYPE))
            practice_area = _clean_text(matter_row.get(sc.COL_MATTER_PRACTICE_AREA))
            billing_email = _clean_text(matter_row.get(sc.COL_MATTER_BILLING_EMAIL))
            responsible_lawyer = _clean_text(matter_row.get(sc.COL_MATTER_RESPONSIBLE_LAWYER))

            subtitle_bits = [x for x in (client_name, matter_type or practice_area, status) if x]
            _append_result(
                entity_type="matter",
                entity_id=matter_id,
                title=display_name or matter_name or matter_id,
                subtitle=" | ".join(subtitle_bits[:3]),
                status=status,
                route_tile_index=0,
                route_node_id="A11",
                route_node_title="Matter Profile 360",
                field_pairs=[
                    ("Matter ID", matter_id),
                    ("Matter Number", matter_number),
                    ("Matter Name", matter_name),
                    ("Display Name", display_name),
                    ("Client ID/Ref", client_ref),
                    ("Client Name", client_name),
                    ("Parent ID/Ref", parent_ref),
                    ("Parent Name", parent_name),
                    ("Matter Type", matter_type),
                    ("Practice Area", practice_area),
                    ("Responsible Lawyer", responsible_lawyer),
                    ("Billing Arrangement", matter_row.get(sc.COL_MATTER_BILLING_ARRANGEMENT)),
                    ("Billing Contact", matter_row.get(sc.COL_MATTER_BILLING_CONTACT)),
                    ("Billing Email", billing_email),
                    ("Default Rate", matter_row.get(sc.COL_MATTER_DEF_RATE)),
                    ("Default Share %", matter_row.get(sc.COL_MATTER_DEF_SHARE)),
                    ("Date Of Engagement", matter_row.get(sc.COL_MATTER_ENGAGEMENT_DATE)),
                    ("Date Opened", matter_row.get(sc.COL_MATTER_OPEN_DATE)),
                    ("Date Closed", matter_row.get(sc.COL_MATTER_CLOSE_DATE)),
                    ("Court File Number", matter_row.get(sc.COL_MATTER_COURT_FILE_NO)),
                    ("Opposing Party", matter_row.get(sc.COL_MATTER_OPPOSING_PARTY)),
                    ("Referral From", matter_row.get(sc.COL_MATTER_REFERRAL_FROM)),
                    ("Description", matter_row.get(sc.COL_MATTER_DESCRIPTION)),
                    ("Status", status),
                    ("Notes", matter_row.get(sc.COL_MATTER_NOTES)),
                ],
                extra={
                    "matterId": matter_id,
                    "matterNumber": matter_number,
                    "matterName": matter_name,
                    "displayName": display_name,
                    "clientId": client_ref,
                    "clientName": client_name,
                    "parentId": parent_ref,
                    "parentName": parent_name,
                    "matterType": matter_type,
                    "practiceArea": practice_area,
                    "responsibleLawyer": responsible_lawyer,
                    "billingEmail": billing_email,
                },
            )

        for time_row in time_rows:
            entry_id = _clean_text(time_row.get(sc.COL_TIME_ENTRY_ID))
            date_text = _clean_text(time_row.get(sc.COL_TIME_DATE))
            description = _clean_text(time_row.get(sc.COL_TIME_DESC))
            status = _clean_text(time_row.get(sc.COL_TIME_STATUS)) or "WIP"
            lock_audit = _clean_text(time_row.get(sc.COL_TIME_LOCK_AUDIT))
            client_ref = _clean_text(time_row.get(sc.COL_TIME_CLIENT_ID))
            matter_ref = _clean_text(time_row.get(sc.COL_TIME_MATTER_ID))
            parent_ref = _clean_text(time_row.get(sc.COL_TIME_PARENT_ID))

            client_name = client_name_by_id.get(client_ref, client_ref)
            matter_name = matter_name_by_id.get(matter_ref, matter_ref)
            parent_name = parent_name_by_id.get(parent_ref, parent_ref)

            desc_lc = _normalize_search_text(description)
            status_lc = _normalize_search_text(status)
            entity_type = "docket"
            route_tile_index = 1
            route_node_id = "B01"
            route_node_title = "Time Docket Entry"
            if "deadline" in desc_lc or "deadline" in status_lc:
                entity_type = "deadline"
                route_node_id = "B07"
                route_node_title = "Deadline Master Calendar"
            elif "tickler" in desc_lc or "reminder" in desc_lc:
                entity_type = "tickler"
                route_node_id = "B11"
                route_node_title = "Tickler Scheduler"
            elif "invoice" in desc_lc or status_lc.startswith("invoice"):
                entity_type = "invoice"
                route_tile_index = 2
                route_node_id = "C04"
                route_node_title = "Invoice Directory"

            title = " - ".join(x for x in (date_text, description) if x) or entry_id or "Docket Entry"
            subtitle_bits = [x for x in (client_name, matter_name, status) if x]
            _append_result(
                entity_type=entity_type,
                entity_id=entry_id,
                title=title,
                subtitle=" | ".join(subtitle_bits[:3]),
                status=status,
                route_tile_index=route_tile_index,
                route_node_id=route_node_id,
                route_node_title=route_node_title,
                field_pairs=[
                    ("Entry ID", entry_id),
                    ("Date", date_text),
                    ("Description", description),
                    ("Status", status),
                    ("Client ID/Ref", client_ref),
                    ("Client Name", client_name),
                    ("Matter ID/Ref", matter_ref),
                    ("Matter Name", matter_name),
                    ("Parent ID/Ref", parent_ref),
                    ("Parent Name", parent_name),
                    ("Hours", time_row.get(sc.COL_TIME_HOURS)),
                    ("Client Rate", time_row.get(sc.COL_TIME_RATE)),
                    ("Share %", time_row.get(sc.COL_TIME_SHARE_PCT)),
                    ("Gross to Client", time_row.get(sc.COL_TIME_GROSS)),
                    ("Amount to You", time_row.get(sc.COL_TIME_NET)),
                    ("HST", time_row.get(sc.COL_TIME_HST)),
                    ("Total", time_row.get(sc.COL_TIME_TOTAL)),
                    ("Lock Audit", lock_audit),
                    ("Created At", time_row.get(sc.COL_TIME_CREATED)),
                ],
                extra={
                    "entryId": entry_id,
                    "clientId": client_ref,
                    "clientName": client_name,
                    "matterId": matter_ref,
                    "matterName": matter_name,
                    "parentId": parent_ref,
                    "date": date_text,
                    "description": description,
                    "hours": time_row.get(sc.COL_TIME_HOURS) or 0.0,
                    "rate": time_row.get(sc.COL_TIME_RATE) or 0.0,
                    "sharePct": time_row.get(sc.COL_TIME_SHARE_PCT) or 100.0,
                    "status": status,
                },
            )

        for trademark_row in trademark_rows:
            payload = self._trademark_row_to_payload(trademark_row)
            trademark_id = _clean_text(payload.get("trademarkId"))
            title = _clean_text(payload.get("trademarkText")) or _clean_text(payload.get("applicationNumber")) or trademark_id
            if not title:
                continue

            jurisdiction = _clean_text(payload.get("jurisdiction"))
            jurisdiction_other = _clean_text(payload.get("jurisdictionOther"))
            jurisdiction_label = jurisdiction
            if jurisdiction.lower() == "other" and jurisdiction_other:
                jurisdiction_label = f"{jurisdiction} ({jurisdiction_other})"
            status = _clean_text(payload.get("status")) or "Open"

            subtitle_bits = [
                jurisdiction_label,
                _clean_text(payload.get("applicationNumber")) or _clean_text(payload.get("registrationNumber")),
                status,
            ]
            _append_result(
                entity_type="trademark",
                entity_id=trademark_id,
                title=title,
                subtitle=" | ".join(x for x in subtitle_bits if x),
                status=status,
                route_tile_index=1,
                route_node_id="B16",
                route_node_title="Trademark Filing",
                field_pairs=[
                    ("Trademark ID", payload.get("trademarkId")),
                    ("Trademark Text", payload.get("trademarkText")),
                    ("Jurisdiction", payload.get("jurisdiction")),
                    ("Jurisdiction Other", payload.get("jurisdictionOther")),
                    ("Client Name", payload.get("clientName")),
                    ("Matter Number", payload.get("matterNumber")),
                    ("Application Number", payload.get("applicationNumber")),
                    ("Registration Number", payload.get("registrationNumber")),
                    ("Current Status", payload.get("currentStatus")),
                    ("CIPO Status", payload.get("cipoStatus")),
                    ("USPTO Status", payload.get("usptoStatusIndicator")),
                    ("Mark Type", payload.get("markType")),
                    ("Nice Classes", payload.get("niceClasses")),
                    ("Goods Services", payload.get("goodsServices")),
                    ("Applicant/Owner", payload.get("applicantNameAddress"))
                    or payload.get("ownerNameAddress"),
                    ("Registry Link", payload.get("registryLink")),
                    ("Updated At", payload.get("updatedAt")),
                ],
                extra={
                    "trademarkId": payload.get("trademarkId"),
                    "trademarkText": payload.get("trademarkText"),
                    "jurisdiction": payload.get("jurisdiction"),
                    "applicationNumber": payload.get("applicationNumber"),
                    "registrationNumber": payload.get("registrationNumber"),
                    "registryLink": payload.get("registryLink"),
                },
            )

        for txn_row in txn_rows:
            txn_id = _clean_text(txn_row.get(sc.COL_TXN_ID))
            txn_date = _clean_text(txn_row.get(sc.COL_TXN_DATE))
            txn_type = _clean_text(txn_row.get(sc.COL_TXN_TYPE))
            txn_class = _clean_text(txn_row.get(sc.COL_TXN_CLASS))
            business_unit = _clean_text(txn_row.get(sc.COL_TXN_BUSINESS_UNIT))
            payee = _clean_text(txn_row.get(sc.COL_TXN_PAYEE))
            category_name = _clean_text(txn_row.get(sc.COL_TXN_CATEGORY_NAME))
            category_code = _clean_text(txn_row.get(sc.COL_TXN_CATEGORY_CODE))
            amount = _clean_text(txn_row.get(sc.COL_TXN_AMOUNT))
            currency = _clean_text(txn_row.get(sc.COL_TXN_CURRENCY)) or "CAD"
            status = _clean_text(txn_row.get(sc.COL_TXN_STATUS)) or "Pending"
            title_bits = [x for x in (txn_date, payee or category_name or txn_type, amount and f"{amount} {currency}") if x]
            subtitle_bits = [x for x in (business_unit, txn_class, txn_type, status) if x]
            _append_result(
                entity_type="transaction",
                entity_id=txn_id,
                title=" - ".join(title_bits) or txn_id or "Transaction",
                subtitle=" | ".join(subtitle_bits[:4]),
                status=status,
                route_tile_index=2,
                route_node_id="C11",
                route_node_title="Transactions Master",
                field_pairs=[
                    ("Transaction ID", txn_id),
                    ("Date", txn_date),
                    ("Class", txn_class),
                    ("Business Unit", business_unit),
                    ("Type", txn_type),
                    ("From Account", txn_row.get(sc.COL_TXN_FROM_ACCOUNT)),
                    ("To Account", txn_row.get(sc.COL_TXN_TO_ACCOUNT)),
                    ("Payee", payee),
                    ("Parent", txn_row.get(sc.COL_TXN_PARENT)),
                    ("Client Name", txn_row.get(sc.COL_TXN_CLIENT)),
                    ("Matter Name", txn_row.get(sc.COL_TXN_MATTER)),
                    ("Category Code", category_code),
                    ("Category Name", category_name),
                    ("Member", txn_row.get(sc.COL_TXN_MEMBER)),
                    ("Amount", amount),
                    ("Tax Amount", txn_row.get(sc.COL_TXN_TAX_AMOUNT)),
                    ("Tax Flag", txn_row.get(sc.COL_TXN_TAX_FLAG)),
                    ("Invoice Ref", txn_row.get(sc.COL_TXN_INVOICE_REF)),
                    ("Expense Details", txn_row.get(sc.COL_TXN_EXPENSE_DETAILS)),
                    ("Notes", txn_row.get(sc.COL_TXN_NOTES)),
                    ("Status", status),
                    ("Currency", currency),
                ],
                extra={
                    "transactionId": txn_id,
                    "txnDate": txn_date,
                    "businessUnit": business_unit,
                    "type": txn_type,
                    "payee": payee,
                    "categoryCode": category_code,
                    "categoryName": category_name,
                    "amount": amount,
                    "currency": currency,
                },
            )

        for account_row in account_rows:
            account_code = _clean_text(account_row.get(sc.COL_TXN_ACCOUNT_CODE))
            account_name = _clean_text(account_row.get(sc.COL_TXN_ACCOUNT_NAME))
            account_kind = _clean_text(account_row.get(sc.COL_TXN_ACCOUNT_KIND))
            owner = _clean_text(account_row.get(sc.COL_TXN_ACCOUNT_OWNER))
            active = self._to_bool_int(account_row.get(sc.COL_TXN_ACCOUNT_ACTIVE), default=1)
            status = "Active" if active else "Inactive"
            _append_result(
                entity_type="account",
                entity_id=account_code,
                title=account_name or account_code,
                subtitle=" | ".join(x for x in (account_code, account_kind, owner, status) if x),
                status=status,
                route_tile_index=2,
                route_node_id="C11",
                route_node_title="Transactions Master",
                field_pairs=[
                    ("Account Code", account_code),
                    ("Account Name", account_name),
                    ("Account Kind", account_kind),
                    ("Owner", owner),
                    ("Aliases", account_row.get(sc.COL_TXN_ACCOUNT_ALIASES)),
                    ("Status", status),
                ],
                extra={
                    "accountCode": account_code,
                    "accountName": account_name,
                    "accountKind": account_kind,
                    "owner": owner,
                },
            )

        for category_row in category_rows:
            category_code = _clean_text(category_row.get(sc.COL_TXN_CATEGORY_LKP_CODE))
            category_name = _clean_text(category_row.get(sc.COL_TXN_CATEGORY_LKP_NAME))
            category_type = _clean_text(category_row.get(sc.COL_TXN_CATEGORY_LKP_TYPE))
            class_scope = _clean_text(category_row.get(sc.COL_TXN_CATEGORY_LKP_CLASS_SCOPE))
            active = self._to_bool_int(category_row.get(sc.COL_TXN_CATEGORY_LKP_ACTIVE), default=1)
            status = "Active" if active else "Inactive"
            _append_result(
                entity_type="category",
                entity_id=category_code,
                title=category_name or category_code,
                subtitle=" | ".join(x for x in (category_code, category_type, class_scope, status) if x),
                status=status,
                route_tile_index=2,
                route_node_id="C11",
                route_node_title="Transactions Master",
                field_pairs=[
                    ("Category Code", category_code),
                    ("Category Name", category_name),
                    ("Type", category_type),
                    ("Class Scope", class_scope),
                    ("Tax Flag", category_row.get(sc.COL_TXN_CATEGORY_LKP_TAX_FLAG_DEFAULT)),
                    ("Notes", category_row.get(sc.COL_TXN_CATEGORY_LKP_NOTES)),
                    ("Status", status),
                ],
                extra={
                    "categoryCode": category_code,
                    "categoryName": category_name,
                    "type": category_type,
                    "classScope": class_scope,
                },
            )

        for unit_row in business_unit_rows:
            business_unit = _clean_text(unit_row.get(sc.COL_TXN_BUSINESS_UNIT_NAME))
            owner = _clean_text(unit_row.get(sc.COL_TXN_BUSINESS_UNIT_OWNER))
            active = self._to_bool_int(unit_row.get(sc.COL_TXN_BUSINESS_UNIT_ACTIVE), default=1)
            status = "Active" if active else "Inactive"
            _append_result(
                entity_type="business_unit",
                entity_id=business_unit,
                title=business_unit,
                subtitle=" | ".join(x for x in (owner, status) if x),
                status=status,
                route_tile_index=2,
                route_node_id="C11",
                route_node_title="Transactions Master",
                field_pairs=[
                    ("Business Unit", business_unit),
                    ("Owner", owner),
                    ("Status", status),
                ],
                extra={"businessUnit": business_unit, "owner": owner},
            )

        for payee_row in payee_rows:
            payee_name = _clean_text(payee_row.get(sc.COL_TXN_PAYEE_NAME))
            default_category = _clean_text(payee_row.get(sc.COL_TXN_PAYEE_DEFAULT_CATEGORY_CODE))
            active = self._to_bool_int(payee_row.get(sc.COL_TXN_PAYEE_ACTIVE), default=1)
            status = "Active" if active else "Inactive"
            _append_result(
                entity_type="payee",
                entity_id=payee_name,
                title=payee_name,
                subtitle=" | ".join(x for x in (default_category, status) if x),
                status=status,
                route_tile_index=2,
                route_node_id="C11",
                route_node_title="Transactions Master",
                field_pairs=[
                    ("Payee Name", payee_name),
                    ("Default Category Code", default_category),
                    ("Status", status),
                ],
                extra={"payeeName": payee_name, "defaultCategoryCode": default_category},
            )

        _profile_lookup_lc = {k.lower(): v for k, v in profile_by_client_id.items()}
        _client_name_lookup_lc = {k.lower(): v for k, v in client_name_by_id.items()}

        def _resolve_client_name(cid: str) -> str:
            if not cid:
                return ""
            key = cid.lower()
            if key in _profile_lookup_lc:
                prof = _profile_lookup_lc[key]
                name = _clean_text(prof.get(sc.COL_PROFILE_DISPLAY_NAME)) or _clean_text(prof.get(sc.COL_PROFILE_LEGAL_NAME))
                if name:
                    return name
            return _client_name_lookup_lc.get(key, cid)

        for recv_row in receivable_rows:
            invoice_num = _clean_text(recv_row.get(sc.COL_RECV_INVOICE_NUM))
            if not invoice_num:
                continue
                
            raw_client = _clean_text(recv_row.get(sc.COL_RECV_CLIENT))
            raw_work = _clean_text(recv_row.get(sc.COL_RECV_WORK_CLIENT))
            client = _resolve_client_name(raw_client)
            work_client = _resolve_client_name(raw_work)
            
            date = _clean_text(recv_row.get(sc.COL_RECV_DATE))
            total_invoiced = self._money_round(recv_row.get(sc.COL_RECV_TOTAL_INVOICED))
            balance_due = self._money_round(recv_row.get(sc.COL_RECV_BALANCE_DUE))
            status = _clean_text(recv_row.get(sc.COL_RECV_STATUS)) or "Open"

            subtitle_bits = [x for x in (client, f"Total: ${total_invoiced:,.2f}", status) if x]
            _append_result(
                entity_type="invoice",
                entity_id=invoice_num,
                title=f"Invoice #{invoice_num}",
                subtitle=" | ".join(subtitle_bits[:3]),
                status=status,
                route_tile_index=2,
                route_node_id="C04",
                route_node_title="Invoice Explorer",
                field_pairs=[
                    ("Invoice Number", invoice_num),
                    ("Date", date),
                    ("Client", client),
                    ("Work Client", work_client),
                    ("Total Invoiced", f"${total_invoiced:,.2f}"),
                    ("Balance Due", f"${balance_due:,.2f}"),
                    ("Status", status),
                ],
                extra={
                    "invoiceNum": invoice_num,
                    "date": date,
                    "client": client,
                    "workClient": work_client,
                    "totalInvoiced": total_invoiced,
                    "balanceDue": balance_due,
                },
            )

        results.sort(
            key=lambda row: (
                -int(row.get("_score", 0)),
                _normalize_search_text(row.get("entityType", "")),
                _normalize_search_text(row.get("title", "")),
                _normalize_search_text(row.get("entityId", "")),
            )
        )

        total_count = len(results)
        if total_count > max_results:
            results = results[:max_results]

        for row in results:
            if "_score" in row:
                del row["_score"]

        return {
            "ok": True,
            "query": raw_query,
            "mode": search_mode,
            "results": results,
            "facets": facets,
            "total": total_count,
            "returnedCount": len(results),
        }

    def home_dashboard_summary(self) -> Dict[str, Any]:
        def _empty(ok: bool = True) -> Dict[str, Any]:
            return {
                "ok": bool(ok),
                "asOfDate": datetime.now().strftime("%Y-%m-%d"),
                "deadlinesCount": 0,
                "unbilledDraftCount": 0,
                "clientMeetingCount": 0,
                "queueCount": 0,
                "activeClientCount": 0,
                # matter summary added per UX requirement
                "activeMatterCount": 0,
            }

        try:
            client_rows = self._read_table_rows(TBL_CLIENTS)
            matter_rows = self._read_table_rows(TBL_MATTERS)
            time_rows = self._read_table_rows(TBL_TIME)
        except Exception:
            try:
                self.ensure_schema()
                client_rows = self._read_table_rows(TBL_CLIENTS)
                matter_rows = self._read_table_rows(TBL_MATTERS)
                time_rows = self._read_table_rows(TBL_TIME)
            except Exception:
                return _empty(ok=False)

        today_iso = datetime.now().strftime("%Y-%m-%d")

        active_client_count = 0
        for row in client_rows:
            status = _clean_text(row.get(sc.COL_CLIENT_STATUS)).lower()
            active_flag = self._to_bool_int(row.get(sc.COL_CLIENT_ACTIVE), default=1)
            if active_flag == 1 and status not in ("inactive", "closed", "archived"):
                active_client_count += 1

        # matter activity count uses canonical helper
        active_matter_count = 0
        for row in matter_rows:
            if self._is_matter_row_active(row):
                active_matter_count += 1

        try:
            deadlines_list = self._load_deadlines()
            deadlines_count = sum(1 for d in deadlines_list if not d.get("completed", False))
        except Exception:
            deadlines_count = 0

        unbilled_draft_count = 0
        client_meeting_count = 0
        for row in time_rows:
            description = _clean_text(row.get(sc.COL_TIME_DESC)).lower()
            status = _clean_text(row.get(sc.COL_TIME_STATUS)).lower()
            if status in ("", "wip", "draft", "unbilled", "pending", "open", "ready for billing"):
                unbilled_draft_count += 1

            date_text = _clean_text(row.get(sc.COL_TIME_DATE))
            if date_text == today_iso and (
                "meeting" in description
                or "conference" in description
                or "call" in description
            ):
                client_meeting_count += 1

        queue_count = deadlines_count + unbilled_draft_count + client_meeting_count
        return {
            "ok": True,
            "asOfDate": today_iso,
            "deadlinesCount": int(deadlines_count),
            "unbilledDraftCount": int(unbilled_draft_count),
            "clientMeetingCount": int(client_meeting_count),
            "queueCount": int(queue_count),
            "activeClientCount": int(active_client_count),
            "activeMatterCount": int(active_matter_count),
        }

    def _sanitize_practice_briefing_filters(self, payload: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        defaults = {
            "upcomingDeadlineDays": 14,
            "includeOverdueDeadlinesInToday": True,
            "overdueBillGraceDays": 30,
            "readyToBillMode": "wip_and_ready",
            "readyToBillMinEntries": 1,
            "readyToBillMinAgeDays": 28,
            "considerBillingWipThreshold": 5000,
        }
        if not isinstance(payload, dict):
            return dict(defaults)

        def _safe_int(key: str, minimum: int, maximum: int, default: int) -> int:
            try:
                value = int(payload.get(key, default) or default)
            except (TypeError, ValueError):
                value = default
            return max(minimum, min(maximum, value))

        mode = str(payload.get("readyToBillMode", defaults["readyToBillMode"]) or "").strip().lower()
        if mode not in {"wip_and_ready", "ready_only"}:
            mode = defaults["readyToBillMode"]

        min_age_days = _safe_int("readyToBillMinAgeDays", 0, 3650, defaults["readyToBillMinAgeDays"])
        min_age_days = _safe_int("considerBillingMinAgeDays", 0, 3650, min_age_days)
        wip_threshold = _safe_int("considerBillingWipThreshold", 0, 10000000, defaults["considerBillingWipThreshold"])

        return {
            "upcomingDeadlineDays": _safe_int("upcomingDeadlineDays", 1, 365, defaults["upcomingDeadlineDays"]),
            "includeOverdueDeadlinesInToday": bool(
                payload.get("includeOverdueDeadlinesInToday", defaults["includeOverdueDeadlinesInToday"])
            ),
            "overdueBillGraceDays": _safe_int("overdueBillGraceDays", 0, 365, defaults["overdueBillGraceDays"]),
            "readyToBillMode": mode,
            "readyToBillMinEntries": _safe_int("readyToBillMinEntries", 1, 9999, defaults["readyToBillMinEntries"]),
            "readyToBillMinAgeDays": min_age_days,
            "considerBillingMinAgeDays": min_age_days,
            "considerBillingWipThreshold": wip_threshold,
        }


    def get_productivity_dashboard_data(self) -> dict:
        """Calculates comprehensive productivity metrics for the React dashboard."""
        import calendar
        from datetime import datetime, timedelta

        self.ensure_schema()
        try:
            time_rows = [self._canonicalize_time_row(r) for r in self._read_table_rows(TBL_TIME)]
            receivable_rows = self._read_table_rows(TBL_RECEIVABLES)
            ledger_rows = self._read_table_rows(TBL_LEDGER)
            client_rows = self._read_table_rows(TBL_CLIENTS)
        except Exception as e:
            return {"ok": False, "message": f"Failed to read tables: {e}"}

        now = datetime.now()
        current_year = now.year
        current_month = now.month

        # 1. Production Data
        production_data = []
        for month in range(1, 13):
            month_name = calendar.month_abbr[month]
            fees = 0.0
            for r in time_rows:
                date_str = str(r.get(sc.COL_TIME_DATE) or "")
                if date_str.startswith(f"{current_year}-{month:02d}"):
                    fees += float(r.get(sc.COL_TIME_GROSS) or 0.0)

            target = 50000.0 
            projected = fees
            if month == current_month:
                days_in_month = calendar.monthrange(current_year, current_month)[1]
                if now.day > 0:
                    projected = fees * (days_in_month / now.day)
            elif month > current_month:
                projected = 50000.0
                fees = 0.0 

            production_data.append({
                "name": month_name,
                "fees": round(fees, 2),
                "target": target,
                "projected": round(projected, 2)
            })

        # 2. Pipeline Data
        pipeline_data = []
        today_date = now.date()
        for i in range(5, 0, -1):
            start_date = today_date - timedelta(days=i*7)
            end_date = start_date + timedelta(days=6)
            
            wip = 0.0
            for r in time_rows:
                date_str = str(r.get(sc.COL_TIME_DATE) or "")
                if date_str:
                    try:
                        d = datetime.strptime(date_str, "%Y-%m-%d").date()
                        status = str(r.get(sc.COL_TIME_STATUS) or "").lower()
                        if start_date <= d <= end_date and status in ["unbilled", "wip", ""]:
                            wip += float(r.get(sc.COL_TIME_GROSS) or 0.0)
                    except ValueError:
                        pass
            
            billed = 0.0
            ar = 0.0
            for r in receivable_rows:
                date_str = str(r.get(sc.COL_RECV_DATE) or "")
                if date_str:
                    try:
                        d = datetime.strptime(date_str, "%Y-%m-%d").date()
                        if start_date <= d <= end_date:
                            billed += float(r.get(sc.COL_RECV_TOTAL_INVOICED) or 0.0)
                            ar += float(r.get(sc.COL_RECV_BALANCE_DUE) or 0.0)
                    except ValueError:
                        pass
                        
            income = 0.0
            for r in ledger_rows:
                date_str = str(r.get(sc.COL_LEDGER_DATE) or "")
                if date_str:
                    try:
                        d = datetime.strptime(date_str, "%Y-%m-%d").date()
                        if start_date <= d <= end_date:
                            income += float(r.get(sc.COL_LEDGER_COLLECTED) or 0.0)
                    except ValueError:
                        pass

            pipeline_data.append({
                "name": f"Week {6-i}",
                "wip": round(wip, 2),
                "billed": round(billed, 2),
                "ar": round(ar, 2),
                "income": round(income, 2)
            })

        # 3. Top Clients
        client_fees = {}
        for r in time_rows:
            date_str = str(r.get(sc.COL_TIME_DATE) or "")
            if date_str.startswith(str(current_year)):
                client_id = str(r.get(sc.COL_TIME_CLIENT_ID) or "Unknown")
                client_fees[client_id] = client_fees.get(client_id, 0.0) + float(r.get(sc.COL_TIME_GROSS) or 0.0)
        
        sorted_clients = sorted(client_fees.items(), key=lambda item: item[1], reverse=True)
        top_clients_data = []
        client_map = {str(r.get(sc.COL_CLIENT_ID)): str(r.get(sc.COL_CLIENT_NAME)) for r in client_rows}
        
        for idx, (cid, val) in enumerate(sorted_clients):
            if idx >= 4:
                other_val = sum(v for k, v in sorted_clients[4:])
                if other_val > 0:
                    top_clients_data.append({"name": "Other", "value": round(other_val, 2)})
                break
            top_clients_data.append({"name": client_map.get(cid, cid), "value": round(val, 2)})

        if not top_clients_data:
            top_clients_data = [{"name": "No Data", "value": 1}]

        # 4. KPI Data
        gross = sum(v for k,v in client_fees.items())
        hours = sum(float(r.get(sc.COL_TIME_HOURS) or 0.0) for r in time_rows if str(r.get(sc.COL_TIME_DATE) or "").startswith(str(current_year)))
        total_ar = sum(float(r.get(sc.COL_RECV_BALANCE_DUE) or 0.0) for r in receivable_rows)
        effective_rate = gross / hours if hours > 0 else 0.0

        kpi_data = {
            "gross": round(gross, 2),
            "hours": round(hours, 2),
            "effectiveRate": round(effective_rate, 2),
            "arTotal": round(total_ar, 2)
        }

        # 5. Raw Time for Frontend Extrapolation
        raw_time = []
        for r in time_rows:
            date_str = str(r.get(sc.COL_TIME_DATE) or "")
            if date_str:
                raw_time.append({
                    "date": date_str,
                    "gross": float(r.get(sc.COL_TIME_GROSS) or 0.0),
                    "hours": float(r.get(sc.COL_TIME_HOURS) or 0.0),
                    "status": str(r.get(sc.COL_TIME_STATUS) or "")
                })

        return {
            "ok": True,
            "productionData": production_data,
            "pipelineData": pipeline_data,
            "topClientsData": top_clients_data,
            "kpiData": kpi_data,
            "rawTime": raw_time
        }


    def practice_briefing(self, filters: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        briefing_filters = self._sanitize_practice_briefing_filters(filters)
        def _empty_summary() -> Dict[str, Any]:
            return {
                "ok": False,
                "asOfDate": "",
                "deadlinesCount": 0,
                "unbilledDraftCount": 0,
                "clientMeetingCount": 0,
                "queueCount": 0,
                "activeClientCount": 0,
                "activeMatterCount": 0,
            }

        def _empty(ok: bool = True) -> Dict[str, Any]:
            return {
                "ok": bool(ok),
                "asOfDate": datetime.now().strftime("%Y-%m-%d"),
                "todaysTasks": [],
                "upcomingDeadlines": [],
                "recentWork": [],
                "overdueDeadlines": [],
                "overdueBills": [],
                "readyToBillMatters": [],
                "arSummary": {
                    "totalAr": 0.0,
                    "openInvoiceCount": 0,
                    "overdueAr": 0.0,
                    "overdueInvoiceCount": 0,
                    "overdueGraceDays": int(briefing_filters.get("overdueBillGraceDays", 30) or 30),
                },
                "summary": _empty_summary(),
                "productivitySummary": {
                    "today": {"hours": 0.0, "gross": 0.0, "net": 0.0},
                    "wtd": {"hours": 0.0, "gross": 0.0, "net": 0.0},
                    "last7": {"hours": 0.0, "gross": 0.0, "net": 0.0},
                    "last90": {"hours": 0.0, "gross": 0.0, "net": 0.0},
                    "ytd": {"hours": 0.0, "gross": 0.0, "net": 0.0},
                },
            }

        try:
            summary = self.home_dashboard_summary()
            matter_rows = self._read_table_rows(TBL_MATTERS)
            time_rows = self._read_table_rows(TBL_TIME)
        except Exception:
            try:
                self.ensure_schema()
                summary = self.home_dashboard_summary()
                matter_rows = self._read_table_rows(TBL_MATTERS)
                time_rows = self._read_table_rows(TBL_TIME)
            except Exception:
                return _empty(ok=False)

        today_iso = datetime.now().strftime("%Y-%m-%d")
        today_date = datetime.strptime(today_iso, "%Y-%m-%d").date()
        upcoming_horizon_iso = (today_date + timedelta(days=int(briefing_filters["upcomingDeadlineDays"]))).strftime("%Y-%m-%d")
        matter_name_by_id: Dict[str, str] = {}
        matter_client_by_id: Dict[str, str] = {}
        matter_number_by_id: Dict[str, str] = {}
        for row in matter_rows:
            matter_id = _clean_text(row.get(sc.COL_MATTER_ID) or row.get("MatterId"))
            if not matter_id:
                continue
            matter_name_by_id[matter_id] = _clean_text(
                row.get(sc.COL_MATTER_NAME) or row.get("MatterName") or matter_id
            )
            matter_client_by_id[matter_id] = _clean_text(
                row.get(sc.COL_MATTER_CLIENT_NAME) or row.get("ClientName") or ""
            )
            matter_number_by_id[matter_id] = _clean_text(
                row.get(sc.COL_MATTER_NUMBER) or row.get("MatterNumber") or matter_id
            )

        todays_tasks: List[Dict[str, Any]] = []
        upcoming_deadlines: List[Dict[str, Any]] = []
        overdue_deadlines: List[Dict[str, Any]] = []
        try:
            for entry in self._load_deadlines():
                if entry.get("completed"):
                    continue
                date_text = _clean_text(entry.get("date"))
                if not _is_valid_iso_date(date_text):
                    continue
                item = {
                    "id": _clean_text(entry.get("id")),
                    "date": date_text,
                    "description": _clean_text(entry.get("description")) or "Deadline",
                    "kind": "deadline",
                }
                if date_text < today_iso:
                    overdue_deadlines.append(item)
                elif date_text == today_iso:
                    todays_tasks.append(item)
                elif date_text <= upcoming_horizon_iso:
                    upcoming_deadlines.append(item)
        except Exception:
            pass

        upcoming_deadlines.sort(key=lambda row: str(row.get("date") or ""))
        overdue_deadlines.sort(key=lambda row: str(row.get("date") or ""))

        wip_statuses = {"", "wip", "draft", "unbilled", "pending", "open", "ready for billing"}
        ready_statuses = {"ready for billing", "ready_for_billing", "ready"}
        ready_only = str(briefing_filters.get("readyToBillMode") or "").lower() == "ready_only"
        min_entries = max(1, int(briefing_filters.get("readyToBillMinEntries", 1) or 1))
        min_age_days = max(0, int(briefing_filters.get("readyToBillMinAgeDays", 28) or 28))
        wip_threshold = max(0, int(briefing_filters.get("considerBillingWipThreshold", 5000) or 5000))
        overdue_bill_days = max(0, int(briefing_filters.get("overdueBillGraceDays", 30) or 30))
        matter_wip: Dict[str, Dict[str, Any]] = {}
        
        recent_work_start_iso = (today_date - timedelta(days=7)).strftime("%Y-%m-%d")
        wtd_start_iso = (today_date - timedelta(days=(today_date.weekday() + 1) % 7)).strftime("%Y-%m-%d")
        last90_start_iso = (today_date - timedelta(days=90)).strftime("%Y-%m-%d")
        ytd_start_iso = today_date.replace(month=1, day=1).strftime("%Y-%m-%d")

        prod = {
            "today": {"hours": 0.0, "gross": 0.0, "net": 0.0},
            "wtd": {"hours": 0.0, "gross": 0.0, "net": 0.0},
            "last7": {"hours": 0.0, "gross": 0.0, "net": 0.0},
            "last90": {"hours": 0.0, "gross": 0.0, "net": 0.0},
            "ytd": {"hours": 0.0, "gross": 0.0, "net": 0.0},
        }

        recent_work: List[Dict[str, Any]] = []
        recent_work_hours = 0.0
        
        for row in time_rows:
            description = _clean_text(row.get(sc.COL_TIME_DESC))
            status = _clean_text(row.get(sc.COL_TIME_STATUS)).lower()
            date_text = _clean_text(row.get(sc.COL_TIME_DATE))
            matter_id = _clean_text(
                row.get(sc.COL_TIME_MATTER_ID) or row.get(sc.COL_MATTER_ID) or row.get("MatterId")
            )
            
            invoice_status = _clean_text(row.get(sc.COL_TIME_INVOICE_STATUS)).lower()
            invoice_ref = _clean_text(row.get(sc.COL_TIME_INVOICE_REF))
            if status == "billed" or invoice_status == "billed" or (invoice_ref and invoice_ref.lower() != "draft"):
                continue

            
            try:
                hours = float(row.get(sc.COL_TIME_HOURS) or row.get("Hours") or 0.0)
            except (TypeError, ValueError):
                hours = 0.0

            try:
                gross_value = float(row.get(sc.COL_TIME_GROSS) or row.get("GrossToClient") or 0.0)
            except (TypeError, ValueError):
                gross_value = 0.0

            try:
                net_value = float(row.get(sc.COL_TIME_NET) or row.get("AmountToYou") or 0.0)
            except (TypeError, ValueError):
                net_value = 0.0

            if "no charge" in status or "free" in status or "do not bill" in status:
                gross_value = 0.0
                net_value = 0.0
                
            if hours > 0 or gross_value > 0 or net_value > 0:
                if date_text == today_iso:
                    prod["today"]["hours"] += hours
                    prod["today"]["gross"] += gross_value
                    prod["today"]["net"] += net_value
                if date_text >= wtd_start_iso:
                    prod["wtd"]["hours"] += hours
                    prod["wtd"]["gross"] += gross_value
                    prod["wtd"]["net"] += net_value
                if date_text >= recent_work_start_iso:
                    prod["last7"]["hours"] += hours
                    prod["last7"]["gross"] += gross_value
                    prod["last7"]["net"] += net_value
                if date_text >= last90_start_iso:
                    prod["last90"]["hours"] += hours
                    prod["last90"]["gross"] += gross_value
                    prod["last90"]["net"] += net_value
                if date_text >= ytd_start_iso:
                    prod["ytd"]["hours"] += hours
                    prod["ytd"]["gross"] += gross_value
                    prod["ytd"]["net"] += net_value

            if recent_work_start_iso <= date_text <= today_iso and hours > 0:
                recent_work_hours += hours
                recent_work.append({
                    "id": _clean_text(row.get(sc.COL_TIME_ENTRY_ID) or row.get("TimeEntryId")),
                    "date": date_text,
                    "description": description or "Work entry",
                    "kind": "work",
                    "hours": hours,
                    "gross": gross_value,
                    "net": net_value,
                    "productivityDetail": f"{hours:g} hrs | ${net_value:,.2f}",
                    "matterId": matter_id,
                    "matterName": matter_name_by_id.get(matter_id, ""),
                })
            
            if date_text == today_iso and (
                "meeting" in description.lower()
                or "conference" in description.lower()
                or "call" in description.lower()
            ):
                todays_tasks.append(
                    {
                        "id": _clean_text(row.get(sc.COL_TIME_ENTRY_ID) or row.get("TimeEntryId")),
                        "date": date_text,
                        "description": description or "Client meeting",
                        "kind": "meeting",
                        "matterId": matter_id,
                        "matterName": matter_name_by_id.get(matter_id, ""),
                    }
                )
            if ready_only:
                if status not in ready_statuses:
                    continue
            elif status not in wip_statuses and status not in ready_statuses:
                continue
            if not matter_id:
                continue

            bucket = matter_wip.setdefault(
                matter_id,
                {
                    "matterId": matter_id,
                    "matterNumber": matter_number_by_id.get(matter_id, matter_id),
                    "matterName": matter_name_by_id.get(matter_id, matter_id),
                    "clientName": matter_client_by_id.get(matter_id, ""),
                    "entryCount": 0,
                    "readyCount": 0,
                    "wipAmount": 0.0,
                    "latestDate": "",
                },
            )
            bucket["entryCount"] = int(bucket.get("entryCount", 0)) + 1
            bucket["wipAmount"] = float(bucket.get("wipAmount", 0.0)) + gross_value
            if status in ready_statuses:
                bucket["readyCount"] = int(bucket.get("readyCount", 0)) + 1
            if date_text and date_text > str(bucket.get("latestDate") or ""):
                bucket["latestDate"] = date_text

        ready_to_bill: List[Dict[str, Any]] = []

        overdue_bills: List[Dict[str, Any]] = []
        ar_summary = {
            "totalAr": 0.0,
            "openInvoiceCount": 0,
            "overdueAr": 0.0,
            "overdueInvoiceCount": 0,
            "overdueGraceDays": overdue_bill_days,
        }
        try:
            ar_data = self.ar_aging_report({"excludedInvoices": briefing_filters.get("excludedInvoices", [])})
            report_summary = ar_data.get("summary") if isinstance(ar_data.get("summary"), dict) else {}
            for row in ar_data.get("rows", []):
                age_days = int(row.get("ageDays", 0))
                if age_days >= overdue_bill_days:
                    overdue_bills.append({
                        "description": f"Invoice {row.get('invoice')} ({row.get('client')})",
                        "date": str(row.get("date", "")),
                        "client": str(row.get("client", "")),
                        "workClient": str(row.get("workClient", "")),
                        "invoice": str(row.get("invoice", "")),
                        "wipAmount": float(row.get("balance", 0.0)),
                        "daysOld": age_days,
                        "kind": "overdue-bill",
                    })
            overdue_bills.sort(key=lambda r: -r["daysOld"])
            ar_summary = {
                "totalAr": round(float(report_summary.get("totalAr", 0.0) or 0.0), 2),
                "openInvoiceCount": max(0, int(report_summary.get("invoiceCount", 0) or 0)),
                "overdueAr": round(sum(float(item.get("wipAmount", 0.0) or 0.0) for item in overdue_bills), 2),
                "overdueInvoiceCount": len(overdue_bills),
                "overdueGraceDays": overdue_bill_days,
            }
        except Exception:
            pass

        for row in matter_wip.values():
            entry_count = int(row.get("entryCount", 0) or 0)
            if entry_count < min_entries:
                continue

            latest_date_text = _clean_text(row.get("latestDate"))
            latest_date = None
            if _is_valid_iso_date(latest_date_text):
                latest_date = datetime.strptime(latest_date_text, "%Y-%m-%d").date()

            stale_since_last_bill = bool(latest_date and (today_date - latest_date).days >= min_age_days)
            high_wip_amount = float(row.get("wipAmount", 0.0) or 0.0) > float(wip_threshold)
            if stale_since_last_bill or high_wip_amount:
                ready_to_bill.append(row)


        ready_to_bill.sort(
            key=lambda row: (
                -int(row.get("readyCount", 0) or 0),
                -int(row.get("entryCount", 0) or 0),
                str(row.get("latestDate") or ""),
            ),
        )

        combined_today = list(todays_tasks)
        if bool(briefing_filters.get("includeOverdueDeadlinesInToday", True)):
            combined_today = overdue_deadlines + combined_today
            
        recent_work.sort(key=lambda x: str(x.get("date") or ""), reverse=True)
        
        if isinstance(summary, dict):
            summary["recentWorkHours"] = round(recent_work_hours, 2)
            summary["recentWorkCount"] = len(recent_work)

        for k in prod.keys():
            prod[k]["hours"] = round(prod[k]["hours"], 2)
            prod[k]["gross"] = round(prod[k]["gross"], 2)
            prod[k]["net"] = round(prod[k]["net"], 2)

        return {
            "ok": True,
            "asOfDate": today_iso,
            "todaysTasks": combined_today[:20],
            "upcomingDeadlines": upcoming_deadlines[:20],
            "recentWork": recent_work[:30],
            "overdueDeadlines": overdue_deadlines[:20],
            "overdueBills": overdue_bills[:20],
            "readyToBillMatters": ready_to_bill[:20],
            "arSummary": ar_summary,
            "summary": summary if isinstance(summary, dict) else _empty_summary(),
            "productivitySummary": prod,
            "filters": briefing_filters,
        }

    def _cspm_time_bucket_key_for_row(self, row: dict) -> tuple[str, str, str]:
        """Draft aggregation key: timekeeper + work date + matter id."""
        work_date = _clean_text(
            row.get(sc.COL_TIME_DATE)
            or row.get("WorkDate")
            or row.get("Date")
            or row.get("dateText")
        ).lower()

        matter_id = _clean_text(
            row.get(sc.COL_TIME_MATTER_ID)
            or row.get(sc.COL_MATTER_ID)
            or row.get("MatterId")
            or row.get("matterId")
        ).lower()

        timekeeper_id = _clean_text(
            row.get(getattr(sc, "COL_TIME_TIMEKEEPER_ID", "TimekeeperId"))
            or row.get("TimekeeperId")
            or row.get("timekeeperId")
            or row.get("Timekeeper")
            or row.get("timekeeper")
        ).lower()

        return (timekeeper_id, work_date, matter_id)


    def _cspm_combine_descriptions(self, existing: str, incoming: str) -> str:
        existing = _clean_text(existing)
        incoming = _clean_text(incoming)

        if not existing:
            return incoming
        if not incoming:
            return existing
        if incoming.lower() in existing.lower():
            return existing

        return existing.rstrip() + "\n\n---\n" + incoming


    def _cspm_billable_hours_from_raw_seconds(self, raw_seconds: int) -> float:
        raw_seconds = max(0, int(raw_seconds or 0))
        if raw_seconds <= 0:
            return 0.0
        import math
        return math.ceil((raw_seconds / 3600.0) * 10.0 - 1e-9) / 10.0


    def _is_fee_entry(self, row: Dict[str, Any]) -> bool:
        """Return whether a time-table row was created by Fee Docket Entry.

        Fee entries deliberately remain in tblTimeEntries so they flow through
        WIP, draft creation, invoicing, and ledger history without needing a
        second billing pipeline.  The marker prevents the time-docket daily
        aggregation path from merging a fee with timed work for the same
        matter/date.
        """
        marker = _clean_text((row or {}).get(sc.COL_TIME_LOCK_AUDIT)).lower()
        return "entrytype:fee" in marker



    def add_time_entry(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.ensure_schema()

        normalized = self._normalize_time_payload(payload)
        parent_name = normalized["parentName"]
        parent_row: Optional[Dict[str, Any]] = None
        if parent_name:
            parent_row = self._get_or_create_parent(parent_name)

        client_row = self._get_or_create_client(normalized["clientName"])
        client_id = _clean_text(client_row.get(sc.COL_CLIENT_ID))
        parent_id = _clean_text(parent_row.get(sc.COL_PARENT_ID)) if parent_row else ""

        matter_id = ""
        requested_matter_id = _clean_text(normalized.get("matterId"))
        matter_row = None

        if requested_matter_id:
            for candidate in [self._canonicalize_matter_row(r) for r in self._read_table_rows(TBL_MATTERS)]:
                if _clean_text(candidate.get(sc.COL_MATTER_ID)).lower() == requested_matter_id.lower():
                    matter_row = candidate
                    matter_id = _clean_text(candidate.get(sc.COL_MATTER_ID))
                    break

        if matter_row is not None:
            matter_client_id = _clean_text(matter_row.get(sc.COL_MATTER_CLIENT_ID))
            if matter_client_id:
                client_id = matter_client_id
            matter_parent_id = _clean_text(matter_row.get(sc.COL_MATTER_PARENT_ID))
            if matter_parent_id:
                parent_id = matter_parent_id
        elif normalized["matterName"]:
            matter_row = self._get_or_create_matter(
                matter_name=normalized["matterName"],
                client_id=client_id,
                parent_id=parent_id,
                default_rate=normalized["clientRate"],
                default_share_pct=normalized["sharePct"],
            )
            matter_id = _clean_text(matter_row.get(sc.COL_MATTER_ID))
            if not parent_id:
                parent_id = _clean_text(matter_row.get(sc.COL_MATTER_PARENT_ID))
        elif not normalized.get("allowClientOnlyDraft", False):
            raise ValueError("Matter is required for time entry.")

        normalized_status = self._normalize_time_status(normalized.get("status"))

        rows = [self._canonicalize_time_row(r) for r in self._read_table_rows(TBL_TIME)]
        requested_entry_id = _clean_text(normalized.get("entryId"))
        force_duplicate = bool(normalized.get("forceDuplicate")) and not requested_entry_id
        requested_row: Optional[Dict[str, Any]] = None
        if requested_entry_id:
            for row in rows:
                if _clean_text(row.get(sc.COL_TIME_ENTRY_ID)).lower() == requested_entry_id.lower():
                    requested_row = row
                    break

        matching_rows: List[Dict[str, Any]] = []
        if requested_row is not None:
            matching_rows = [requested_row]
        elif not force_duplicate:
            for row in rows:
                if self._is_fee_entry(row):
                    continue
                if _clean_text(row.get(sc.COL_TIME_DATE)) != normalized["date"]:
                    continue
                if _clean_text(row.get(sc.COL_TIME_CLIENT_ID)).lower() != client_id.lower():
                    continue
                row_matter_id = _clean_text(row.get(sc.COL_TIME_MATTER_ID))
                if matter_id:
                    if row_matter_id.lower() != matter_id.lower():
                        continue
                elif row_matter_id:
                    continue
                matching_rows.append(row)

        billed_rows = [
            r for r in matching_rows if self._normalize_time_status(r.get(sc.COL_TIME_STATUS)) == "Billed"
        ]
        if billed_rows and not normalized.get("forceEditBilled"):
            return {
                "ok": False,
                "verifiedExact": False,
                "entryId": _clean_text(billed_rows[0].get(sc.COL_TIME_ENTRY_ID)),
                "savedRow": {},
                "message": "Time entry is already billed and locked.",
            }

        editable_rows = [
            r for r in matching_rows if _clean_text(r.get(sc.COL_TIME_STATUS)).lower() != "merged"
        ]
        editable_rows.sort(
            key=lambda row: (
                _clean_text(row.get(sc.COL_TIME_CREATED)),
                _clean_text(row.get(sc.COL_TIME_ENTRY_ID)),
            )
        )

        primary_row: Optional[Dict[str, Any]] = editable_rows[0] if editable_rows else None
        duplicate_rows: List[Dict[str, Any]] = editable_rows[1:] if len(editable_rows) > 1 else []
        existing_total_seconds = 0
        for row in editable_rows:
            existing_total_seconds += int(self._parse_float(row.get(sc.COL_TIME_SECONDS)) or 0)

        incoming_raw_seconds = int(normalized.get("rawSeconds") or 0)
        incoming_segment_seconds = normalized.get("segmentSeconds")
        mode = _clean_text(normalized.get("rawSecondsMode")).lower()
        if mode not in ("increment", "replace", "auto"):
            mode = "auto"

        if mode == "replace":
            aggregate_raw_seconds = max(0, incoming_raw_seconds)
        elif mode == "increment":
            increment = int(incoming_segment_seconds) if incoming_segment_seconds is not None else incoming_raw_seconds
            aggregate_raw_seconds = max(0, existing_total_seconds + max(0, int(increment)))
        else:
            # Auto mode: accept monotonic absolute totals; otherwise treat as incremental segment.
            if incoming_raw_seconds >= existing_total_seconds:
                aggregate_raw_seconds = max(0, incoming_raw_seconds)
            else:
                increment = int(incoming_segment_seconds) if incoming_segment_seconds is not None else incoming_raw_seconds
                aggregate_raw_seconds = max(0, existing_total_seconds + max(0, int(increment)))

        # When importing legacy data with known exact hours, skip the rawSeconds→ceiling
        # roundup which inflates fractional hours (e.g. 0.25 → 0.3).
        if normalized.get("useExactHours") and normalized.get("hours") is not None:
            aggregate_hours = float(normalized["hours"])
            aggregate_raw_seconds = int(aggregate_hours * 3600)
        else:
            aggregate_hours = (
                math.ceil(((aggregate_raw_seconds / 3600.0) * 10.0) - 1e-9) / 10.0
                if aggregate_raw_seconds > 0
                else 0.0
            )
        amounts = calc_amounts(
            hours=aggregate_hours,
            client_rate=normalized["clientRate"],
            your_share_pct=normalized["sharePct"],
            hst_rate=0.13,
        )

        # Editing an existing docket should replace the user-facing description text
        # (fallback to prior description only when a new one is blank).
        incoming_description = _clean_text(normalized.get("description"))
        existing_description = _clean_text((primary_row or {}).get(sc.COL_TIME_DESC))
        if incoming_description:
            description_text = incoming_description
        elif existing_description:
            description_text = existing_description
        else:
            description_text = "Time entry"

        lock_audit_parts: List[str] = []
        seen_audit = set()
        for candidate in [
            _clean_text((primary_row or {}).get(sc.COL_TIME_LOCK_AUDIT)),
            _clean_text(normalized.get("lockAudit")),
        ]:
            if not candidate:
                continue
            split_parts = [part.strip() for part in candidate.split("||")] if "||" in candidate else [candidate]
            for part in split_parts:
                if not part:
                    continue
                key = part.lower()
                if key in seen_audit:
                    continue
                seen_audit.add(key)
                lock_audit_parts.append(part)
        lock_audit_text = " || ".join(lock_audit_parts)

        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        entry_id = (
            requested_entry_id
            or _clean_text((primary_row or {}).get(sc.COL_TIME_ENTRY_ID))
            or self._new_id("T")
        )
        created_at = _clean_text((primary_row or {}).get(sc.COL_TIME_CREATED)) or now_stamp
        effective_parent_id = parent_id or _clean_text((primary_row or {}).get(sc.COL_TIME_PARENT_ID))

        entry_row = {
            sc.COL_TIME_ENTRY_ID: entry_id,
            sc.COL_TIME_DATE: normalized["date"],
            sc.COL_TIME_CLIENT_ID: client_id,
            sc.COL_TIME_MATTER_ID: matter_id,
            sc.COL_TIME_PARENT_ID: effective_parent_id,
            sc.COL_TIME_DESC: description_text,
            sc.COL_TIME_HOURS: round(float(aggregate_hours), 2),
            sc.COL_TIME_RATE: round(float(normalized["clientRate"]), 2),
            sc.COL_TIME_SHARE_PCT: round(float(normalized["sharePct"]), 2),
            sc.COL_TIME_GROSS: amounts["gross_to_client"],
            sc.COL_TIME_NET: amounts["amount_to_you"],
            sc.COL_TIME_HST: amounts["hst_on_you"],
            sc.COL_TIME_TOTAL: amounts["total_you_incl_hst"],
            sc.COL_TIME_SECONDS: int(aggregate_raw_seconds),
            sc.COL_TIME_STATUS: normalized_status,
            sc.COL_TIME_INVOICE_REF: normalized["invoiceRef"],
            sc.COL_TIME_INVOICE_STATUS: normalized["invoiceStatus"],
            sc.COL_TIME_PAYMENT_STATUS: normalized["paymentStatus"],
            sc.COL_TIME_INVOICE_TOTAL: normalized["invoiceTotal"],
            sc.COL_TIME_INVOICE_AMOUNT_PAID: normalized["invoiceAmountPaid"],
            sc.COL_TIME_INVOICE_BALANCE_DUE: normalized["invoiceBalanceDue"],
            sc.COL_TIME_INVOICE_DATE: normalized["invoiceDate"],
            sc.COL_TIME_LOCK_AUDIT: lock_audit_text,
            sc.COL_TIME_CREATED: created_at,
        }

        self._upsert_row_by_key(TBL_TIME, sc.COL_TIME_ENTRY_ID, entry_id, entry_row)

        # Merge duplicate editable rows into the primary bucket, preserving an audit breadcrumb.
        if duplicate_rows:
            for dup in duplicate_rows:
                dup_id = _clean_text(dup.get(sc.COL_TIME_ENTRY_ID))
                if not dup_id:
                    continue
                dup_row = dict(dup)
                dup_row[sc.COL_TIME_HOURS] = 0.0
                dup_row[sc.COL_TIME_GROSS] = 0.0
                dup_row[sc.COL_TIME_NET] = 0.0
                dup_row[sc.COL_TIME_HST] = 0.0
                dup_row[sc.COL_TIME_TOTAL] = 0.0
                dup_row[sc.COL_TIME_SECONDS] = 0
                dup_row[sc.COL_TIME_STATUS] = "Merged"
                dup_desc = _clean_text(dup_row.get(sc.COL_TIME_DESC))
                dup_row[sc.COL_TIME_DESC] = (
                    f"[MERGED->{entry_id}] {dup_desc}".strip()
                    if dup_desc
                    else f"[MERGED->{entry_id}]"
                )
                self._upsert_row_by_key(TBL_TIME, sc.COL_TIME_ENTRY_ID, dup_id, dup_row)

        persisted = self._find_time_entry(entry_id)
        verified = self._compare_rows_loose(entry_row, persisted)

        return {
            "ok": bool(verified),
            "verifiedExact": bool(verified),
            "entryId": entry_id,
            "savedRow": entry_row,
            "aggregateRawSeconds": int(aggregate_raw_seconds),
            "aggregateHoursRounded": round(float(aggregate_hours), 2),
            "mergedRowCount": len(duplicate_rows),
            "message": "" if verified else "Entry write verification failed.",
        }

    def add_fee_entry(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """Create one matter-linked, invoiceable fee WIP line.

        A fee is stored in the standard time/WIP table with zero hours and a
        zero rate.  Its positive gross/net amount is intentional: the invoice
        renderer treats that shape as a flat-fee service line.  This keeps a
        matter containing only direct fee entries invoiceable without routing
        the user through the invoice builder.
        """
        self.ensure_schema()
        raw = dict(payload or {})

        date_text = self._pick_text(raw, ["date", "dateText", sc.COL_TIME_DATE])
        if not date_text:
            raise ValueError("Date is required.")
        if not _is_valid_iso_date(date_text):
            raise ValueError("Date must be in YYYY-MM-DD format.")

        amount = self._pick_float(raw, ["amount", "feeAmount", "amountText", sc.COL_TIME_GROSS])
        if amount is None or not math.isfinite(float(amount)) or float(amount) <= 0:
            raise ValueError("Fee amount must be greater than zero.")
        amount = round(float(amount), 2)

        requested_matter_id = self._pick_text(
            raw,
            ["matterId", "selectedMatterId", sc.COL_TIME_MATTER_ID, sc.COL_MATTER_ID],
        )
        requested_matter_text = self._pick_text(
            raw,
            ["matterName", "matterText", sc.COL_MATTER_NAME, "Matter"],
        )
        if not requested_matter_id and not requested_matter_text:
            raise ValueError("Matter is required for a fee entry.")

        matter_rows = [self._canonicalize_matter_row(r) for r in self._read_table_rows(TBL_MATTERS)]
        matter_row: Optional[Dict[str, Any]] = None
        if requested_matter_id:
            for candidate in matter_rows:
                if _clean_text(candidate.get(sc.COL_MATTER_ID)).lower() == requested_matter_id.lower():
                    matter_row = candidate
                    break

        if matter_row is None and requested_matter_text:
            lookup = requested_matter_text.lower()
            for candidate in matter_rows:
                matter_id = _clean_text(candidate.get(sc.COL_MATTER_ID)).lower()
                matter_number = _clean_text(candidate.get(sc.COL_MATTER_NUMBER)).lower()
                matter_name = _clean_text(candidate.get(sc.COL_MATTER_NAME)).lower()
                display_name = _clean_text(candidate.get(sc.COL_MATTER_DISPLAY_NAME)).lower()
                combined = (matter_number + " - " + matter_name).strip(" -")
                if lookup in (matter_id, matter_number, matter_name, display_name, combined):
                    matter_row = candidate
                    break

        if matter_row is None:
            raise ValueError("Select an existing matter for the fee entry.")

        matter_id = _clean_text(matter_row.get(sc.COL_MATTER_ID))
        client_id = _clean_text(matter_row.get(sc.COL_MATTER_CLIENT_ID))
        if not matter_id or not client_id:
            raise ValueError("The selected matter is missing its client link.")

        parent_id = _clean_text(matter_row.get(sc.COL_MATTER_PARENT_ID))
        description = self._pick_text(raw, ["description", "descriptionText", sc.COL_TIME_DESC])
        if not description:
            description = "Fee Entry"
        status = self._normalize_time_status(self._pick_text(raw, ["status", sc.COL_TIME_STATUS]))
        tax = round(amount * 0.13, 2)
        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        entry_id = self._new_id("T")
        entry_row = {
            sc.COL_TIME_ENTRY_ID: entry_id,
            sc.COL_TIME_DATE: date_text,
            sc.COL_TIME_CLIENT_ID: client_id,
            sc.COL_TIME_MATTER_ID: matter_id,
            sc.COL_TIME_PARENT_ID: parent_id,
            sc.COL_TIME_DESC: description,
            sc.COL_TIME_HOURS: 0.0,
            sc.COL_TIME_RATE: 0.0,
            sc.COL_TIME_SHARE_PCT: 100.0,
            sc.COL_TIME_GROSS: amount,
            sc.COL_TIME_NET: amount,
            sc.COL_TIME_HST: tax,
            sc.COL_TIME_TOTAL: round(amount + tax, 2),
            sc.COL_TIME_SECONDS: 0,
            sc.COL_TIME_STATUS: status,
            sc.COL_TIME_INVOICE_REF: "",
            sc.COL_TIME_INVOICE_STATUS: "",
            sc.COL_TIME_PAYMENT_STATUS: "",
            sc.COL_TIME_INVOICE_TOTAL: 0.0,
            sc.COL_TIME_INVOICE_AMOUNT_PAID: 0.0,
            sc.COL_TIME_INVOICE_BALANCE_DUE: 0.0,
            sc.COL_TIME_INVOICE_DATE: "",
            sc.COL_TIME_LOCK_AUDIT: "EntryType:Fee",
            sc.COL_TIME_CREATED: now_stamp,
        }
        self._upsert_row_by_key(TBL_TIME, sc.COL_TIME_ENTRY_ID, entry_id, entry_row)
        persisted = self._find_time_entry(entry_id)
        verified = self._compare_rows_loose(entry_row, persisted)
        return {
            "ok": bool(verified),
            "verifiedExact": bool(verified),
            "entryId": entry_id,
            "savedRow": entry_row,
            "message": "" if verified else "Fee entry write verification failed.",
        }

    def _bulk_docket_move_context(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        raw = dict(payload or {})
        source_key = _clean_text(raw.get("sourceMatterId") or raw.get("sourceMatter"))
        target_key = _clean_text(raw.get("targetMatterId") or raw.get("targetMatter"))
        start_date = self._parse_date_value(raw.get("fromDate") or raw.get("startDate"))
        end_date = self._parse_date_value(raw.get("toDate") or raw.get("endDate"))

        if not source_key:
            raise ValueError("Select the matter that currently owns the dockets.")
        if start_date is None or end_date is None:
            raise ValueError("Enter both dates as YYYY-MM-DD.")
        if start_date > end_date:
            raise ValueError("The start date cannot be after the end date.")

        source_result = self.get_matter_profile(source_key)
        if not bool(source_result.get("ok")):
            raise ValueError(source_result.get("message") or "Source matter was not found.")
        source = dict(source_result.get("matter") or {})

        target: Dict[str, Any] = {}
        if target_key:
            target_result = self.get_matter_profile(target_key)
            if not bool(target_result.get("ok")):
                raise ValueError(target_result.get("message") or "Destination matter was not found.")
            target = dict(target_result.get("matter") or {})
            source_id = _clean_text(source.get("matterId"))
            target_id = _clean_text(target.get("matterId"))
            if source_id.lower() == target_id.lower():
                raise ValueError("Choose a different destination matter.")
            if not bool(target.get("active")):
                raise ValueError("The destination matter must be active.")
        return {
            "source": source,
            "target": target,
            "fromDate": start_date,
            "toDate": end_date,
        }

    def preview_bulk_docket_move(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """List movable dockets for a source matter and inclusive date range.

        Only entries that have not been assigned to an invoice are eligible.  This
        preserves finalized and draft-invoice history while still permitting a
        user to correct a whole period of unbilled time or direct fees.
        """
        self.ensure_schema()
        context = self._bulk_docket_move_context(payload)
        source = context["source"]
        source_id = _clean_text(source.get("matterId"))
        start_date = context["fromDate"]
        end_date = context["toDate"]

        candidates: List[Dict[str, Any]] = []
        matched_count = 0
        blocked_count = 0
        blocked_statuses = {"billed", "merged", "locked", "invoiced", "finalized"}

        for row in self._read_table_rows(TBL_TIME):
            if _clean_text(row.get(sc.COL_TIME_MATTER_ID)).lower() != source_id.lower():
                continue
            row_date = self._parse_date_value(row.get(sc.COL_TIME_DATE))
            if row_date is None or row_date < start_date or row_date > end_date:
                continue
            matched_count += 1
            invoice_ref = _clean_text(row.get(sc.COL_TIME_INVOICE_REF))
            invoice_status = _clean_text(row.get(sc.COL_TIME_INVOICE_STATUS)).lower()
            status = _clean_text(row.get(sc.COL_TIME_STATUS)) or "Draft"
            if invoice_ref or invoice_status in blocked_statuses or status.lower() in blocked_statuses:
                blocked_count += 1
                continue

            entry_id = _clean_text(row.get(sc.COL_TIME_ENTRY_ID))
            if not entry_id:
                blocked_count += 1
                continue
            hours = float(self._parse_float(row.get(sc.COL_TIME_HOURS)) or 0.0)
            amount = float(
                self._parse_float(row.get(sc.COL_TIME_GROSS))
                or self._parse_float(row.get(sc.COL_TIME_NET))
                or 0.0
            )
            is_fee = self._is_fee_entry(row)
            candidates.append({
                "entryId": entry_id,
                "date": row_date.isoformat(),
                "description": _clean_text(row.get(sc.COL_TIME_DESC)) or ("Fee entry" if is_fee else "Time docket"),
                "hours": round(hours, 2),
                "amount": round(amount, 2),
                "isFee": bool(is_fee),
                "status": status,
            })

        candidates.sort(key=lambda item: (item["date"], item["entryId"]))
        return {
            "ok": True,
            "message": "",
            "sourceMatter": source,
            "fromDate": start_date.isoformat(),
            "toDate": end_date.isoformat(),
            "matchedCount": matched_count,
            "blockedCount": blocked_count,
            "candidates": candidates,
        }

    @with_db_lock
    def move_dockets_to_matter(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """Atomically move selected, still-unbilled time/fee dockets to one matter."""
        self.ensure_schema()
        context = self._bulk_docket_move_context(payload)
        source = context["source"]
        target = context["target"]
        if not target:
            raise ValueError("Select the destination matter.")

        requested_ids = {
            _clean_text(entry_id)
            for entry_id in list((payload or {}).get("entryIds") or [])
            if _clean_text(entry_id)
        }
        if not requested_ids:
            raise ValueError("Select at least one docket to move.")

        preview = self.preview_bulk_docket_move(payload)
        eligible_ids = {
            _clean_text(row.get("entryId"))
            for row in list(preview.get("candidates") or [])
            if _clean_text(row.get("entryId"))
        }
        unavailable_ids = sorted(requested_ids - eligible_ids)
        if unavailable_ids:
            raise ValueError(
                "One or more selected dockets are no longer eligible to move. Refresh the list and try again."
            )

        source_id = _clean_text(source.get("matterId"))
        source_name = _clean_text(source.get("displayName")) or _clean_text(source.get("matterName")) or source_id
        target_id = _clean_text(target.get("matterId"))
        target_name = _clean_text(target.get("displayName")) or _clean_text(target.get("matterName")) or target_id
        target_client_id = _clean_text(target.get("clientId"))
        target_parent_id = _clean_text(target.get("parentId"))
        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        audit_note = (
            f"[Bulk docket move {now_stamp}] {source_name} -> {target_name}; "
            f"{context['fromDate'].isoformat()} to {context['toDate'].isoformat()}"
        )

        rows = self._read_table_rows(TBL_TIME)
        moved_ids: List[str] = []
        for row in rows:
            entry_id = _clean_text(row.get(sc.COL_TIME_ENTRY_ID))
            if entry_id not in requested_ids:
                continue
            if _clean_text(row.get(sc.COL_TIME_MATTER_ID)).lower() != source_id.lower():
                raise ValueError("A selected docket no longer belongs to the source matter. Refresh and try again.")
            row[sc.COL_TIME_MATTER_ID] = target_id
            if target_client_id:
                row[sc.COL_TIME_CLIENT_ID] = target_client_id
            if target_parent_id:
                row[sc.COL_TIME_PARENT_ID] = target_parent_id
            row[sc.COL_TIME_LOCK_AUDIT] = self._append_note_line(row.get(sc.COL_TIME_LOCK_AUDIT), audit_note)
            moved_ids.append(entry_id)

        if len(moved_ids) != len(requested_ids):
            raise ValueError("Not every selected docket could be located. No changes were saved.")

        self._replace_table_rows(TBL_TIME, rows)
        verified_rows = {
            _clean_text(row.get(sc.COL_TIME_ENTRY_ID)): row
            for row in self._read_table_rows(TBL_TIME)
        }
        verified = all(
            _clean_text(verified_rows.get(entry_id, {}).get(sc.COL_TIME_MATTER_ID)).lower() == target_id.lower()
            for entry_id in moved_ids
        )
        return {
            "ok": bool(verified),
            "verifiedExact": bool(verified),
            "movedCount": len(moved_ids),
            "movedEntryIds": moved_ids,
            "sourceMatter": source,
            "targetMatter": target,
            "message": "" if verified else "Docket move verification failed.",
        }

    def update_time_entry(self, entry_id: str, changes: Dict[str, Any]) -> Dict[str, Any]:
        self.ensure_schema()
        row = self._find_time_entry(entry_id)
        if not row:
            return {"ok": False, "message": f"Time entry {entry_id} not found."}
            
        updated = False
        
        if "description" in changes:
            row[sc.COL_TIME_DESC] = _clean_text(changes["description"])
            updated = True
            
        if "date" in changes:
            row[sc.COL_TIME_DATE] = _clean_text(changes["date"])
            updated = True
            
        hours_changed = "hours" in changes
        amount_changed = "amount" in changes
        
        if hours_changed or amount_changed:
            old_hours = float(row.get(sc.COL_TIME_HOURS) or 0)
            old_net = float(row.get(sc.COL_TIME_NET) or 0)
            
            if hours_changed:
                new_hours = float(changes["hours"])
            else:
                new_hours = old_hours
                
            if amount_changed:
                new_net = float(changes["amount"])
            else:
                if old_hours != 0:
                    rate = old_net / old_hours
                    new_net = new_hours * rate
                else:
                    new_net = 0
                          
            hst = new_net * 0.13
            total = new_net + hst
            
            row[sc.COL_TIME_HOURS] = new_hours
            row[sc.COL_TIME_NET] = round(new_net, 2)
            row[sc.COL_TIME_HST] = round(hst, 2)
            row[sc.COL_TIME_TOTAL] = round(total, 2)
            row[sc.COL_TIME_GROSS] = round(new_net, 2)
            
            # also update seconds
            row[sc.COL_TIME_SECONDS] = int(new_hours * 3600)
            updated = True
            
        if updated:
            self._upsert_row_by_key(TBL_TIME, sc.COL_TIME_ENTRY_ID, entry_id, row)
            return {"ok": True, "message": "Time entry updated.", "savedRow": row}
            
        return {"ok": True, "message": "No changes applied."}

    def delete_time_entry(self, entry_id: str) -> Dict[str, Any]:
        """Deletes a time entry from TBL_TIME."""
        self.ensure_schema()
        deleted = self._delete_row_by_key_hard(TBL_TIME, sc.COL_TIME_ENTRY_ID, entry_id)
        if deleted:
            return {"ok": True, "message": f"Time entry {entry_id} deleted."}
        return {"ok": False, "message": f"Time entry {entry_id} not found."}

    def _cipo_trademark_application_url(self, application_no: Any, registration_no: Any, trademark_text: Any = "") -> str:
        app_no = re.sub(r"\D+", "", _clean_text(application_no))
        reg_no = re.sub(r"\D+", "", _clean_text(registration_no))
        lookup_no = app_no or reg_no
        query = _clean_text(trademark_text).lower() or lookup_no
        if not lookup_no:
            return self._cipo_trademark_search_url(query)
        encoded_payload = self._encoded_cipo_trademark_payload(query)
        return (
            f"https://ised-isde.canada.ca/cipo/trademark-search/{lookup_no}"
            f"?lang=eng&payload={encoded_payload}&pageNum=0&pageLen=100"
        )

    def _encoded_cipo_trademark_payload(self, query: Any) -> str:
        query_text = _clean_text(query)
        if not query_text:
            return ""
        payload = {
            "domIntlFilter": "1",
            "searchfield1": "all",
            "textfield1": query_text,
            "display": "list",
            "maxReturn": "1000",
            "nicetextfield1": None,
            "cipotextfield1": None,
        }
        return quote_plus(json.dumps(payload, separators=(",", ":")))

    def _cipo_trademark_search_url(self, query: Any) -> str:
        encoded_payload = self._encoded_cipo_trademark_payload(query)
        if not encoded_payload:
            return "https://ised-isde.canada.ca/cipo/trademark-search/srch?lang=eng"
        return f"https://ised-isde.canada.ca/cipo/trademark-search/srch?lang=eng&payload={encoded_payload}"

    def _is_cipo_search_results_link(self, url: Any) -> bool:
        raw = _clean_text(url).lower()
        return "ised-isde.canada.ca/cipo/trademark-search/srch" in raw

    def _default_trademark_registry_link(
        self,
        jurisdiction: Any,
        application_no: Any,
        registration_no: Any,
        trademark_text: Any = "",
    ) -> str:
        office = _clean_text(jurisdiction).upper()
        app_no = re.sub(r"\D+", "", _clean_text(application_no))
        reg_no = re.sub(r"\D+", "", _clean_text(registration_no))
        lookup_no = app_no or reg_no
        if office == "USPTO":
            if lookup_no:
                return (
                    f"https://tsdr.uspto.gov/#caseNumber={lookup_no}"
                    "&caseSearchType=US_APPLICATION&caseType=DEFAULT&searchType=statusSearch"
            )
            return "https://tsdr.uspto.gov/"
        if office == "CIPO":
            return self._cipo_trademark_application_url(
                application_no,
                registration_no,
                trademark_text,
            )
        return ""

    def _trademark_row_to_payload(self, row: Dict[str, Any]) -> Dict[str, Any]:
        payload = {
            "trademarkId": _clean_text(row.get(sc.COL_TM_ID)),
            "jurisdiction": _clean_text(row.get(sc.COL_TM_JURISDICTION)),
            "jurisdictionOther": _clean_text(row.get(sc.COL_TM_JURISDICTION_OTHER)),
            "clientName": _clean_text(row.get(sc.COL_TM_CLIENT_NAME)),
            "matterNumber": _clean_text(row.get(sc.COL_TM_MATTER_NUMBER)),
            "internalNotes": _clean_text(row.get(sc.COL_TM_INTERNAL_NOTES)),
            "trademarkText": _clean_text(row.get(sc.COL_TM_TRADEMARK_TEXT)),
            "markType": _clean_text(row.get(sc.COL_TM_MARK_TYPE)),
            "designRepresentation": _clean_text(row.get(sc.COL_TM_DESIGN_REPRESENTATION)),
            "designImagePaste": _clean_text(row.get(sc.COL_TM_DESIGN_IMAGE_PASTE)),
            "colorClaimed": "Yes"
            if int(self._to_bool_int(row.get(sc.COL_TM_COLOR_CLAIMED), default=0)) == 1
            else "No",
            "colorDescription": _clean_text(row.get(sc.COL_TM_COLOR_DESCRIPTION)),
            "niceClasses": _clean_text(row.get(sc.COL_TM_NICE_CLASSES)),
            "goodsServices": _clean_text(row.get(sc.COL_TM_GOODS_SERVICES)),
            "foreignPriorityClaim": _clean_text(row.get(sc.COL_TM_FOREIGN_PRIORITY)),
            "registryLink": _clean_text(row.get(sc.COL_TM_REGISTRY_LINK)),
            "applicationNumber": _clean_text(row.get(sc.COL_TM_APPLICATION_NO)),
            "registrationNumber": _clean_text(row.get(sc.COL_TM_REGISTRATION_NO)),
            "currentStatus": _clean_text(row.get(sc.COL_TM_CURRENT_STATUS)),
            "applicantNameAddress": _clean_text(row.get(sc.COL_TM_APPLICANT_NAME_ADDRESS)),
            "filingDate": _clean_text(row.get(sc.COL_TM_FILING_DATE)),
            "registrationDate": _clean_text(row.get(sc.COL_TM_REGISTRATION_DATE)),
            "renewalDeadline": _clean_text(row.get(sc.COL_TM_RENEWAL_DEADLINE)),
            "cipoStatus": _clean_text(row.get(sc.COL_TM_CIPO_STATUS)),
            "tm5Status": _clean_text(row.get(sc.COL_TM_TM5_STATUS)),
            "examinersReportDate": _clean_text(row.get(sc.COL_TM_EXAMINERS_REPORT_DATE)),
            "officeActionResponseDeadline": _clean_text(row.get(sc.COL_TM_OFFICE_ACTION_RESPONSE_DEADLINE)),
            "approvalDate": _clean_text(row.get(sc.COL_TM_APPROVAL_DATE)),
            "advertisementDate": _clean_text(row.get(sc.COL_TM_ADVERTISEMENT_DATE)),
            "advertisementVolIssue": _clean_text(row.get(sc.COL_TM_ADVERTISEMENT_VOL_ISSUE)),
            "oppositionDeadline": _clean_text(row.get(sc.COL_TM_OPPOSITION_DEADLINE)),
            "allowanceDate": _clean_text(row.get(sc.COL_TM_ALLOWANCE_DATE)),
            "registerType": _clean_text(row.get(sc.COL_TM_REGISTER_TYPE)),
            "usptoStatusIndicator": _clean_text(row.get(sc.COL_TM_USPTO_STATUS_INDICATOR)),
            "ownerNameAddress": _clean_text(row.get(sc.COL_TM_OWNER_NAME_ADDRESS)),
            "attorneyOfRecord": _clean_text(row.get(sc.COL_TM_ATTORNEY_OF_RECORD)),
            "publicationDate": _clean_text(row.get(sc.COL_TM_PUBLICATION_DATE)),
            "noticeOfAllowanceDate": _clean_text(row.get(sc.COL_TM_NOTICE_OF_ALLOWANCE_DATE)),
            "souDeadline": _clean_text(row.get(sc.COL_TM_SOU_DEADLINE)),
            "souExtensionTracking": _clean_text(row.get(sc.COL_TM_SOU_EXTENSION_TRACKING)),
            "section8Deadline": _clean_text(row.get(sc.COL_TM_SECTION8_DEADLINE)),
            "section15Deadline": _clean_text(row.get(sc.COL_TM_SECTION15_DEADLINE)),
            "section9Deadline": _clean_text(row.get(sc.COL_TM_SECTION9_DEADLINE)),
            "localForeignAssociate": _clean_text(row.get(sc.COL_TM_LOCAL_FOREIGN_ASSOCIATE)),
            "applicationReferenceNumber": _clean_text(row.get(sc.COL_TM_APPLICATION_REFERENCE_NO)),
            "publicationAdvertisementDate": _clean_text(row.get(sc.COL_TM_PUBLICATION_ADVERTISEMENT_DATE)),
            "oppositionPeriodEndDate": _clean_text(row.get(sc.COL_TM_OPPOSITION_PERIOD_END_DATE)),
            "upcomingLocalDeadlineOfficeActionDate": _clean_text(row.get(sc.COL_TM_UPCOMING_LOCAL_DEADLINE)),
            "createdAt": _clean_text(row.get(sc.COL_TM_CREATED_AT)),
            "updatedAt": _clean_text(row.get(sc.COL_TM_UPDATED_AT)),
        }
        if not payload["registryLink"] or self._is_cipo_search_results_link(payload["registryLink"]):
            payload["registryLink"] = self._default_trademark_registry_link(
                payload["jurisdiction"],
                payload["applicationNumber"],
                payload["registrationNumber"],
                payload["trademarkText"],
            )
        payload["title"] = payload["trademarkText"] or payload["applicationNumber"] or payload["trademarkId"]
        payload["status"] = payload["currentStatus"] or payload["cipoStatus"] or payload["usptoStatusIndicator"]
        return payload

    def save_trademark_filing(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.ensure_schema()
        raw_payload = dict(payload or {})
        mapped_payload = dict(raw_payload)
        key_map = {
            "trademarkId": sc.COL_TM_ID,
            "jurisdiction": sc.COL_TM_JURISDICTION,
            "jurisdictionOther": sc.COL_TM_JURISDICTION_OTHER,
            "clientName": sc.COL_TM_CLIENT_NAME,
            "matterNumber": sc.COL_TM_MATTER_NUMBER,
            "internalNotes": sc.COL_TM_INTERNAL_NOTES,
            "trademarkText": sc.COL_TM_TRADEMARK_TEXT,
            "markType": sc.COL_TM_MARK_TYPE,
            "designRepresentation": sc.COL_TM_DESIGN_REPRESENTATION,
            "designImagePaste": sc.COL_TM_DESIGN_IMAGE_PASTE,
            "colorClaimed": sc.COL_TM_COLOR_CLAIMED,
            "colorDescription": sc.COL_TM_COLOR_DESCRIPTION,
            "niceClasses": sc.COL_TM_NICE_CLASSES,
            "goodsServices": sc.COL_TM_GOODS_SERVICES,
            "foreignPriorityClaim": sc.COL_TM_FOREIGN_PRIORITY,
            "registryLink": sc.COL_TM_REGISTRY_LINK,
            "applicationNumber": sc.COL_TM_APPLICATION_NO,
            "registrationNumber": sc.COL_TM_REGISTRATION_NO,
            "currentStatus": sc.COL_TM_CURRENT_STATUS,
            "applicantNameAddress": sc.COL_TM_APPLICANT_NAME_ADDRESS,
            "filingDate": sc.COL_TM_FILING_DATE,
            "registrationDate": sc.COL_TM_REGISTRATION_DATE,
            "renewalDeadline": sc.COL_TM_RENEWAL_DEADLINE,
            "cipoStatus": sc.COL_TM_CIPO_STATUS,
            "tm5Status": sc.COL_TM_TM5_STATUS,
            "examinersReportDate": sc.COL_TM_EXAMINERS_REPORT_DATE,
            "officeActionResponseDeadline": sc.COL_TM_OFFICE_ACTION_RESPONSE_DEADLINE,
            "approvalDate": sc.COL_TM_APPROVAL_DATE,
            "advertisementDate": sc.COL_TM_ADVERTISEMENT_DATE,
            "advertisementVolIssue": sc.COL_TM_ADVERTISEMENT_VOL_ISSUE,
            "oppositionDeadline": sc.COL_TM_OPPOSITION_DEADLINE,
            "allowanceDate": sc.COL_TM_ALLOWANCE_DATE,
            "registerType": sc.COL_TM_REGISTER_TYPE,
            "usptoStatusIndicator": sc.COL_TM_USPTO_STATUS_INDICATOR,
            "ownerNameAddress": sc.COL_TM_OWNER_NAME_ADDRESS,
            "attorneyOfRecord": sc.COL_TM_ATTORNEY_OF_RECORD,
            "publicationDate": sc.COL_TM_PUBLICATION_DATE,
            "noticeOfAllowanceDate": sc.COL_TM_NOTICE_OF_ALLOWANCE_DATE,
            "souDeadline": sc.COL_TM_SOU_DEADLINE,
            "souExtensionTracking": sc.COL_TM_SOU_EXTENSION_TRACKING,
            "section8Deadline": sc.COL_TM_SECTION8_DEADLINE,
            "section15Deadline": sc.COL_TM_SECTION15_DEADLINE,
            "section9Deadline": sc.COL_TM_SECTION9_DEADLINE,
            "localForeignAssociate": sc.COL_TM_LOCAL_FOREIGN_ASSOCIATE,
            "applicationReferenceNumber": sc.COL_TM_APPLICATION_REFERENCE_NO,
            "publicationAdvertisementDate": sc.COL_TM_PUBLICATION_ADVERTISEMENT_DATE,
            "oppositionPeriodEndDate": sc.COL_TM_OPPOSITION_PERIOD_END_DATE,
            "upcomingLocalDeadlineOfficeActionDate": sc.COL_TM_UPCOMING_LOCAL_DEADLINE,
            "createdAt": sc.COL_TM_CREATED_AT,
            "updatedAt": sc.COL_TM_UPDATED_AT,
        }
        for source_key, target_key in key_map.items():
            if source_key in raw_payload and target_key not in mapped_payload:
                mapped_payload[target_key] = raw_payload.get(source_key)
        incoming = self._canonicalize_trademark_row(mapped_payload)
        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        trademark_id = _clean_text(incoming.get(sc.COL_TM_ID))
        if not trademark_id:
            trademark_id = self._new_id("TM")
        incoming[sc.COL_TM_ID] = trademark_id
        incoming[sc.COL_TM_UPDATED_AT] = now_stamp
        if not _clean_text(incoming.get(sc.COL_TM_CREATED_AT)):
            incoming[sc.COL_TM_CREATED_AT] = now_stamp
        if not _clean_text(incoming.get(sc.COL_TM_REGISTRY_LINK)) or self._is_cipo_search_results_link(
            incoming.get(sc.COL_TM_REGISTRY_LINK)
        ):
            incoming[sc.COL_TM_REGISTRY_LINK] = self._default_trademark_registry_link(
                incoming.get(sc.COL_TM_JURISDICTION),
                incoming.get(sc.COL_TM_APPLICATION_NO),
                incoming.get(sc.COL_TM_REGISTRATION_NO),
                incoming.get(sc.COL_TM_TRADEMARK_TEXT),
            )

        rows = [self._canonicalize_trademark_row(r) for r in self._read_table_rows(TBL_TRADEMARKS)]
        found = False
        for idx, row in enumerate(rows):
            row_id = _clean_text(row.get(sc.COL_TM_ID))
            if not row_id or row_id.lower() != trademark_id.lower():
                continue
            incoming[sc.COL_TM_CREATED_AT] = _clean_text(row.get(sc.COL_TM_CREATED_AT)) or incoming.get(
                sc.COL_TM_CREATED_AT
            )
            rows[idx] = incoming
            found = True
            break
        if not found:
            rows.append(incoming)

        self._replace_table_rows(TBL_TRADEMARKS, rows)
        self._sync_trademark_generated_deadlines(incoming)

        persisted = {}
        verify_rows = [self._canonicalize_trademark_row(r) for r in self._read_table_rows(TBL_TRADEMARKS)]
        for row in verify_rows:
            if _clean_text(row.get(sc.COL_TM_ID)).lower() == trademark_id.lower():
                persisted = row
                break
        verified = self._compare_rows_loose(incoming, persisted)

        return {
            "ok": bool(verified),
            "verifiedExact": bool(verified),
            "trademarkId": trademark_id,
            "savedRow": self._trademark_row_to_payload(incoming),
            "message": "" if verified else "Trademark write verification failed.",
        }

    def list_trademark_directory(self, query: str = "") -> List[Dict[str, Any]]:
        self.ensure_schema()
        rows = [self._canonicalize_trademark_row(r) for r in self._read_table_rows(TBL_TRADEMARKS)]
        normalized_query = _normalize_search_text(query)
        terms = _search_terms(normalized_query) if normalized_query else []

        out: List[Dict[str, Any]] = []
        for row in rows:
            payload = self._trademark_row_to_payload(row)
            haystack = _normalize_search_text(
                " | ".join(
                    [
                        payload.get("trademarkId", ""),
                        payload.get("trademarkText", ""),
                        payload.get("jurisdiction", ""),
                        payload.get("jurisdictionOther", ""),
                        payload.get("applicationNumber", ""),
                        payload.get("registrationNumber", ""),
                        payload.get("currentStatus", ""),
                        payload.get("cipoStatus", ""),
                        payload.get("usptoStatusIndicator", ""),
                        payload.get("clientName", ""),
                        payload.get("matterNumber", ""),
                        payload.get("niceClasses", ""),
                        payload.get("goodsServices", ""),
                        payload.get("registryLink", ""),
                    ]
                )
            )
            if terms and not all(term in haystack for term in terms):
                continue
            out.append(payload)

        out.sort(
            key=lambda row: (
                _clean_text(row.get("updatedAt")).lower(),
                _clean_text(row.get("trademarkText")).lower(),
                _clean_text(row.get("trademarkId")).lower(),
            ),
            reverse=True,
        )
        return out

    def get_time_docket_aggregate(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.ensure_schema()

        raw = dict(payload or {})
        date_text = self._pick_text(raw, ["date", "dateText", sc.COL_TIME_DATE])
        if not date_text:
            date_text = datetime.now().strftime("%Y-%m-%d")
        if not _is_valid_iso_date(date_text):
            raise ValueError("Date must be in YYYY-MM-DD format.")

        client_name = self._pick_text(raw, ["clientName", "clientText", sc.COL_CLIENT_NAME, "Client"])
        matter_name = self._pick_text(raw, ["matterName", "matterText", sc.COL_MATTER_NAME, "Matter"])
        requested_matter_id = self._pick_text(raw, ["matterId", "selectedMatterId", sc.COL_TIME_MATTER_ID, sc.COL_MATTER_ID])
        allow_client_only = self._to_bool_int(
            self._pick_value(raw, ["allowClientOnlyDraft", "allowClientOnly", "clientOnlyOverride"]),
            default=0,
        ) == 1

        if not client_name:
            return {
                "ok": True,
                "exists": False,
                "entryId": "",
                "aggregateRawSeconds": 0,
                "aggregateHoursRounded": 0.0,
                "status": "Draft",
                "message": "",
            }

        client_rows = [self._canonicalize_client_row(r) for r in self._read_table_rows(TBL_CLIENTS)]
        matter_rows = [self._canonicalize_matter_row(r) for r in self._read_table_rows(TBL_MATTERS)]
        time_rows = [self._canonicalize_time_row(r) for r in self._read_table_rows(TBL_TIME)]

        client_id = ""
        for row in client_rows:
            if _clean_text(row.get(sc.COL_CLIENT_NAME)).lower() == client_name.lower():
                client_id = _clean_text(row.get(sc.COL_CLIENT_ID))
                client_name = _clean_text(row.get(sc.COL_CLIENT_NAME))
                break

        matter_id = ""
        if requested_matter_id:
            for row in matter_rows:
                row_id = _clean_text(row.get(sc.COL_MATTER_ID))
                if row_id.lower() == requested_matter_id.lower():
                    row_name = _clean_text(row.get(sc.COL_MATTER_NAME))
                    row_display = _clean_text(row.get(sc.COL_MATTER_DISPLAY_NAME))
                    row_number = _clean_text(row.get(sc.COL_MATTER_NUMBER))
                    row_client_id = _clean_text(row.get(sc.COL_MATTER_CLIENT_ID))
                    matter_id = row_id
                    matter_name = row_display or row_name or row_number or matter_name
                    if row_client_id:
                        client_id = row_client_id
                    break
        elif matter_name:
            for row in matter_rows:
                row_name = _clean_text(row.get(sc.COL_MATTER_NAME))
                row_display = _clean_text(row.get(sc.COL_MATTER_DISPLAY_NAME))
                row_number = _clean_text(row.get(sc.COL_MATTER_NUMBER))
                row_label = (row_number + " - " + row_name).strip(" -")
                if (
                    row_name.lower() == matter_name.lower()
                    or row_display.lower() == matter_name.lower()
                    or row_number.lower() == matter_name.lower()
                    or row_label.lower() == matter_name.lower()
                ):
                    row_client_id = _clean_text(row.get(sc.COL_MATTER_CLIENT_ID))
                    if client_id and row_client_id and row_client_id.lower() != client_id.lower():
                        continue
                    matter_id = _clean_text(row.get(sc.COL_MATTER_ID))
                    matter_name = row_display or row_name or matter_name
                    if not client_id:
                        client_id = row_client_id
                    break
        elif not allow_client_only:
            return {
                "ok": True,
                "exists": False,
                "entryId": "",
                "aggregateRawSeconds": 0,
                "aggregateHoursRounded": 0.0,
                "status": "Draft",
                "message": "Matter not selected.",
            }

        if not client_id:
            return {
                "ok": True,
                "exists": False,
                "entryId": "",
                "aggregateRawSeconds": 0,
                "aggregateHoursRounded": 0.0,
                "status": "Draft",
                "message": "",
            }

        matches: List[Dict[str, Any]] = []
        for row in time_rows:
            if self._is_fee_entry(row):
                continue
            if _clean_text(row.get(sc.COL_TIME_DATE)) != date_text:
                continue
            if _clean_text(row.get(sc.COL_TIME_CLIENT_ID)).lower() != client_id.lower():
                continue
            row_matter_id = _clean_text(row.get(sc.COL_TIME_MATTER_ID))
            if matter_id:
                if row_matter_id.lower() != matter_id.lower():
                    continue
            elif row_matter_id:
                continue
            if _clean_text(row.get(sc.COL_TIME_STATUS)).lower() == "merged":
                continue
            matches.append(row)

        if not matches:
            return {
                "ok": True,
                "exists": False,
                "entryId": "",
                "date": date_text,
                "clientId": client_id,
                "clientName": client_name,
                "matterId": matter_id,
                "matterName": matter_name,
                "aggregateRawSeconds": 0,
                "aggregateHoursRounded": 0.0,
                "status": "Draft",
                "message": "",
            }

        matches.sort(
            key=lambda row: (
                _clean_text(row.get(sc.COL_TIME_CREATED)),
                _clean_text(row.get(sc.COL_TIME_ENTRY_ID)),
            )
        )

        primary = matches[0]
        aggregate_raw_seconds = sum(int(self._parse_float(r.get(sc.COL_TIME_SECONDS)) or 0) for r in matches)
        aggregate_raw_seconds = max(0, aggregate_raw_seconds)
        aggregate_hours = (
            math.ceil(((aggregate_raw_seconds / 3600.0) * 10.0) - 1e-9) / 10.0
            if aggregate_raw_seconds > 0
            else 0.0
        )

        normalized_statuses = [self._normalize_time_status(r.get(sc.COL_TIME_STATUS)) for r in matches]
        effective_status = "Draft"
        if "Billed" in normalized_statuses:
            effective_status = "Billed"
        elif "Ready for Billing" in normalized_statuses:
            effective_status = "Ready for Billing"

        return {
            "ok": True,
            "exists": True,
            "entryId": _clean_text(primary.get(sc.COL_TIME_ENTRY_ID)),
            "date": date_text,
            "clientId": client_id,
            "clientName": client_name,
            "matterId": matter_id,
            "matterName": matter_name,
            "description": _clean_text(primary.get(sc.COL_TIME_DESC)),
            "lockAudit": _clean_text(primary.get(sc.COL_TIME_LOCK_AUDIT)),
            "rate": round(float(self._parse_float(primary.get(sc.COL_TIME_RATE)) or 0.0), 2),
            "sharePct": round(float(self._parse_float(primary.get(sc.COL_TIME_SHARE_PCT)) or 100.0), 2),
            "aggregateRawSeconds": int(aggregate_raw_seconds),
            "aggregateHoursRounded": round(float(aggregate_hours), 2),
            "status": effective_status,
            "sourceRowCount": len(matches),
            "message": "",
        }




    def get_docket_activity_report(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        from openpyxl import load_workbook
        import logging, time, os, json, csv
        from datetime import datetime
        
        logger = logging.getLogger("DocketReport")
        trace_mode = os.environ.get("CSPM_DOCKET_TRACE") == "1" or "--trace-docket-report" in sys.argv
        if trace_mode: logger.setLevel(logging.DEBUG)
        detail_log = logger.info if trace_mode else logger.debug

        t_start = time.perf_counter()
        
        # [A] File resolution
        path = self.paths.workbook_path()
        detail_log(f"[A] EXCEL SOURCE: path={path.absolute()}")
        detail_log(f"[A] EXCEL SOURCE: exists={path.exists()} size={path.stat().st_size if path.exists() else 0}")
        
        if not path.exists():
            logger.error("[A] CRITICAL: Workbook path does not exist.")
            raise FileNotFoundError("Excel file missing.")

        raw_rows = []
        headers = []
        try:
            # [B] Workbook open details
            wb = load_workbook(path, read_only=True, data_only=True)
            t_open_end = time.perf_counter()
            detail_log(f"[B] WORKBOOK OPEN: {(t_open_end - t_start)*1000:.1f}ms. Engine: openpyxl (read_only)")
            
            # [C] Sheet selection
            detail_log(f"[C] SHEET SELECTION: sheets_found={wb.sheetnames}")
            target_sheet = sc.SHEET_TIME
            if target_sheet not in wb.sheetnames:
                logger.error(f"[C] CRITICAL: Sheet '{target_sheet}' missing!")
                raise ValueError(f"Sheet {target_sheet} missing.")
            
            detail_log(f"[C] SHEET SELECTION: sheet_selected={target_sheet}")
            ws = wb[target_sheet]
            
            # [D] Header detection
            header_row_index = 1
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
                str_row = [str(c).strip() if c else "" for c in row]
                if sc.COL_TIME_ENTRY_ID in str_row:
                    headers = str_row
                    header_row_index = r_idx
                    break
            
            detail_log(f"[D] HEADER DETECTION: row_index={header_row_index}, headers={headers}")
            
            mapping = {
                "date_col": headers.index(sc.COL_TIME_DATE) if sc.COL_TIME_DATE in headers else "MISSING",
                "client_col": headers.index(sc.COL_TIME_CLIENT_ID) if sc.COL_TIME_CLIENT_ID in headers else "MISSING",
                "matter_col": headers.index(sc.COL_TIME_MATTER_ID) if sc.COL_TIME_MATTER_ID in headers else "MISSING",
                "status_col": headers.index(sc.COL_TIME_STATUS) if sc.COL_TIME_STATUS in headers else "MISSING",
                "hours_col": headers.index(sc.COL_TIME_HOURS) if sc.COL_TIME_HOURS in headers else "MISSING",
                "gross_col": headers.index(sc.COL_TIME_GROSS) if sc.COL_TIME_GROSS in headers else "MISSING"
            }
            detail_log(f"[D] MAPPING: {mapping}")

            if "MISSING" in mapping.values():
                logger.error(f"[D] CRITICAL: Missing required headers in {target_sheet}!")
                raise ValueError("Missing required headers in Excel sheet.")

            # [E] Range scan
            max_r = ws.max_row
            detail_log(f"[E] RANGE SCAN: {header_row_index + 1}..{max_r}")
            for row in ws.iter_rows(min_row=header_row_index + 1, max_row=max_r, values_only=True):
                if any(row):
                    raw_dict = {headers[i]: val for i, val in enumerate(row) if i < len(headers) and headers[i]}
                    raw_rows.append(raw_dict)
                    
            detail_log(f"[E] RANGE SCAN: rows_scanned_total={len(raw_rows)}")
            
            if trace_mode and raw_rows:
                logger.debug(f"[F] RAW SAMPLE (first 3): {raw_rows[:3]}")
                
        finally:
            if 'wb' in locals(): wb.close()

        # [G] Parse records & Lookups
        try:
            client_rows = [self._canonicalize_client_row(r) for r in self._read_table_rows(TBL_CLIENTS)]
            matter_rows = [self._canonicalize_matter_row(r) for r in self._read_table_rows(TBL_MATTERS)]
            parent_rows = [self._canonicalize_parent_row(r) for r in self._read_table_rows(TBL_PARENTS)]
        except Exception as e:
            logger.error(f"Lookup table read failure: {e}", exc_info=True)
            client_rows, matter_rows, parent_rows = [], [], []

        client_name_by_id = {
            _clean_text(r.get(sc.COL_CLIENT_ID)): _clean_text(r.get(sc.COL_CLIENT_NAME))
            for r in client_rows
            if _clean_text(r.get(sc.COL_CLIENT_ID))
        }
        matter_name_by_id = {
            _clean_text(r.get(sc.COL_MATTER_ID)): (
                _clean_text(r.get(sc.COL_MATTER_DISPLAY_NAME))
                or _clean_text(r.get(sc.COL_MATTER_NAME))
                or _clean_text(r.get(sc.COL_MATTER_NUMBER))
            )
            for r in matter_rows
            if _clean_text(r.get(sc.COL_MATTER_ID))
        }
        parent_name_by_id = {
            _clean_text(r.get(sc.COL_PARENT_ID)): _clean_text(r.get(sc.COL_PARENT_NAME))
            for r in parent_rows
            if _clean_text(r.get(sc.COL_PARENT_ID))
        }

        parsed_rows = []
        for r in raw_rows:
            parsed_rows.append(self._canonicalize_time_row(r))
            
        detail_log(f"[G] PARSE RESULTS: rows_parsed_ok={len(parsed_rows)}")
        
        valid_dates = [str(r.get(sc.COL_TIME_DATE)) for r in parsed_rows if r.get(sc.COL_TIME_DATE) and _is_valid_iso_date(str(r.get(sc.COL_TIME_DATE)))]
        if valid_dates:
            detail_log(f"[E] DATASET SANITY: min_date={min(valid_dates)}, max_date={max(valid_dates)}")
        else:
            logger.warning("[E] DATASET SANITY: No valid dates parsed!")
            
        if trace_mode and parsed_rows:
            logger.debug(f"[H] PARSED SAMPLE (first 3): {parsed_rows[:3]}")

        # [I] Filters
        raw = dict(payload or {})
        start_text = self._pick_text(raw, ["fromDate", "dateFrom", "startDate", "from"]) or datetime.now().strftime("%Y-%m-%d")
        end_text = self._pick_text(raw, ["toDate", "dateTo", "endDate", "to"]) or start_text
        status_mode = _normalize_search_text(self._pick_text(raw, ["statusMode", "statusFilter", "status"])) or "all_except_merged"
        client_filter = _normalize_search_text(self._pick_text(raw, ["clientFilter", "clientName"]))
        matter_filter = _normalize_search_text(self._pick_text(raw, ["matterFilter", "matterName"]))
        matter_id_filter = _clean_text(self._pick_text(raw, ["matterId", "selectedMatterId", sc.COL_TIME_MATTER_ID, sc.COL_MATTER_ID])).lower()
        text_query = _normalize_search_text(self._pick_text(raw, ["query", "search"]))

        detail_log(f"[I] FILTER LOGGING: filters={{from:{start_text}, to:{end_text}, status:{status_mode}, client:'{client_filter}', matter:'{matter_filter}', search:'{text_query}'}}")

        start_date = datetime.strptime(start_text, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_text, "%Y-%m-%d").date()
        if end_date < start_date:
            start_date, end_date = end_date, start_date

        def _pair_matches(filter_token: str, value_id: str, value_name: str) -> bool:
            if not filter_token:
                return True
            return filter_token in f"{value_id} {value_name}".lower()

        relation_pairs = []
        client_candidates = set()
        matter_candidates = set()
        saw_no_matter = False

        for row in client_rows:
            cname = _clean_text(row.get(sc.COL_CLIENT_NAME))
            if cname:
                client_candidates.add(cname)

        for row in matter_rows:
            cid = _clean_text(row.get(sc.COL_MATTER_CLIENT_ID))
            cname = _clean_text(row.get(sc.COL_MATTER_CLIENT_NAME)) or client_name_by_id.get(cid, "")
            mid = _clean_text(row.get(sc.COL_MATTER_ID))
            mname = (
                _clean_text(row.get(sc.COL_MATTER_DISPLAY_NAME))
                or _clean_text(row.get(sc.COL_MATTER_NAME))
                or _clean_text(row.get(sc.COL_MATTER_NUMBER))
                or mid
            )
            if cname:
                client_candidates.add(cname)
            if mname:
                matter_candidates.add(mname)
            if cid or cname or mid or mname:
                relation_pairs.append(
                    {
                        "clientId": cid,
                        "clientName": cname,
                        "matterId": mid,
                        "matterName": mname,
                    }
                )

        for row in parsed_rows:
            cid = _clean_text(row.get(sc.COL_TIME_CLIENT_ID))
            cname = client_name_by_id.get(cid, cid)
            mid = _clean_text(row.get(sc.COL_TIME_MATTER_ID))
            if mid:
                mname = matter_name_by_id.get(mid, mid)
                if mname:
                    matter_candidates.add(mname)
            else:
                mname = "No Matter"
                saw_no_matter = True

            if cname:
                client_candidates.add(cname)
            relation_pairs.append(
                {
                    "clientId": cid,
                    "clientName": cname,
                    "matterId": mid,
                    "matterName": mname,
                }
            )

        if saw_no_matter:
            matter_candidates.add("No Matter")

        option_clients_set = set()
        if matter_filter:
            for pair in relation_pairs:
                if _pair_matches(matter_filter, _clean_text(pair.get("matterId")), _clean_text(pair.get("matterName"))):
                    cname = _clean_text(pair.get("clientName"))
                    if cname:
                        option_clients_set.add(cname)
        else:
            option_clients_set.update(client_candidates)

        option_matters_set = set()
        if client_filter:
            for pair in relation_pairs:
                if _pair_matches(client_filter, _clean_text(pair.get("clientId")), _clean_text(pair.get("clientName"))):
                    mname = _clean_text(pair.get("matterName"))
                    if mname:
                        option_matters_set.add(mname)
        else:
            option_matters_set.update(matter_candidates)

        if not option_clients_set:
            option_clients_set.update(client_candidates)
        if not option_matters_set:
            option_matters_set.update(matter_candidates)

        raw_client_selection = _clean_text(self._pick_text(raw, ["clientFilter", "clientName"]))
        raw_matter_selection = _clean_text(self._pick_text(raw, ["matterFilter", "matterName"]))
        if raw_client_selection:
            option_clients_set.add(raw_client_selection)
        if raw_matter_selection:
            option_matters_set.add(raw_matter_selection)

        option_clients = ["All Clients"]
        option_clients.extend(
            sorted(
                [
                    label
                    for label in option_clients_set
                    if _normalize_search_text(label) and _normalize_search_text(label) != "all clients"
                ],
                key=lambda label: _normalize_search_text(label),
            )
        )
        option_matters = ["All Matters"]
        option_matters.extend(
            sorted(
                [
                    label
                    for label in option_matters_set
                    if _normalize_search_text(label) and _normalize_search_text(label) != "all matters"
                ],
                key=lambda label: _normalize_search_text(label),
            )
        )

        detail_rows = []
        rejects = {"invalid_date": 0, "out_of_range": 0, "status_excluded": 0, "client_mismatch": 0, "matter_mismatch": 0, "search_miss": 0}
        counts = {"post_date": 0, "post_status": 0, "post_client": 0, "post_matter": 0, "post_search": 0}
        
        for row in parsed_rows:
            date_text = _clean_text(row.get(sc.COL_TIME_DATE))
            if not _is_valid_iso_date(date_text):
                rejects["invalid_date"] += 1
                continue
            row_date = datetime.strptime(date_text, "%Y-%m-%d").date()
            if row_date < start_date or row_date > end_date:
                rejects["out_of_range"] += 1
                continue
            counts["post_date"] += 1
                
            raw_status = _clean_text(row.get(sc.COL_TIME_STATUS)).lower()
            norm_status = "Merged" if raw_status == "merged" else self._normalize_time_status(raw_status)
            
            inc_status = True
            if status_mode == "all": inc_status = True
            elif status_mode == "all_except_merged": inc_status = raw_status != "merged"
            elif status_mode in ("unbilled_wip", "unbilled_work"):
                inc_status = norm_status in {"Draft", "Ready for Billing"}
            elif status_mode in ("draft", "draft_only", "unbilled", "open"): inc_status = norm_status == "Draft"
            elif status_mode in ("ready", "ready_for_billing"): inc_status = norm_status == "Ready for Billing"
            elif status_mode in ("billed", "billed_only"): inc_status = norm_status == "Billed"
            elif status_mode in ("merged", "merged_only"): inc_status = raw_status == "merged"
            else: inc_status = raw_status != "merged"
            
            if not inc_status:
                rejects["status_excluded"] += 1
                continue
            counts["post_status"] += 1

            c_id = _clean_text(row.get(sc.COL_TIME_CLIENT_ID))
            c_name = client_name_by_id.get(c_id, c_id)
            if client_filter and client_filter not in f"{c_id} {c_name}".lower():
                rejects["client_mismatch"] += 1
                continue
            counts["post_client"] += 1

            m_id = _clean_text(row.get(sc.COL_TIME_MATTER_ID))
            m_name = matter_name_by_id.get(m_id, m_id) or "No Matter"
            if matter_id_filter and m_id.lower() != matter_id_filter:
                rejects["matter_mismatch"] += 1
                continue
            if matter_filter and matter_filter not in f"{m_id} {m_name}".lower():
                rejects["matter_mismatch"] += 1
                continue
            counts["post_matter"] += 1

            desc = _clean_text(row.get(sc.COL_TIME_DESC))
            entry_id = _clean_text(row.get(sc.COL_TIME_ENTRY_ID))
            p_id = _clean_text(row.get(sc.COL_TIME_PARENT_ID))
            p_name = parent_name_by_id.get(p_id, p_id)
            
            if text_query:
                haystack = f"{entry_id} {desc} {c_id} {c_name} {m_id} {m_name} {p_id} {p_name} {norm_status}".lower()
                if text_query not in haystack:
                    rejects["search_miss"] += 1
                    continue
            counts["post_search"] += 1

            row["clientName"] = c_name
            row["matterName"] = m_name
            row["parentName"] = p_name
            row["status"] = norm_status
            
            row["hours"] = float(self._parse_float(row.get(sc.COL_TIME_HOURS)) or 0.0)
            row["grossToClient"] = float(self._parse_float(row.get(sc.COL_TIME_GROSS)) or 0.0)
            
            detail_rows.append(row)

        detail_log(f"[I] FILTER LOGGING: counts={counts}")
        detail_log(f"[G] REJECT REASONS: {rejects}")

        detail_rows.sort(key=lambda r: (_clean_text(r.get("date")), _normalize_search_text(r.get("clientName")), _normalize_search_text(r.get("matterName")), _normalize_search_text(r.get("entryId"))))

        summary_map = {}
        total_hours = 0.0
        total_gross = 0.0
        for r in detail_rows:
            total_hours += r["hours"]
            total_gross += r["grossToClient"]
            m_id = _clean_text(r.get("matterId"))
            c_key = _clean_text(r.get("clientId")) or _clean_text(r.get("clientName")).lower()
            s_key = f"matter::{m_id.lower()}" if m_id else f"no_matter::{c_key.lower()}"
            
            if s_key not in summary_map:
                summary_map[s_key] = {
                    "matterId": m_id,
                    "matterName": r["matterName"],
                    "clientId": r.get("clientId", ""),
                    "clientName": r["clientName"],
                    "entryCount": 0,
                    "totalHours": 0.0,
                    "totalGrossToClient": 0.0
                }
            summary_map[s_key]["entryCount"] += 1
            summary_map[s_key]["totalHours"] += r["hours"]
            summary_map[s_key]["totalGrossToClient"] += r["grossToClient"]

        summary_rows = list(summary_map.values())
        summary_rows.sort(key=lambda r: (1 if _normalize_search_text(r.get("matterName")) == "no matter" else 0, _normalize_search_text(r.get("matterName")), _normalize_search_text(r.get("clientName"))))
        for r in summary_rows:
            r["totalHours"] = round(r["totalHours"], 2)
            r["totalGrossToClient"] = round(r["totalGrossToClient"], 2)

        # [J] Publish to UI
        detail_log(f"[J] PUBLISH: publishing_rows={len(detail_rows)}")
        detail_log(f"[J] PUBLISH: summary_by_matter_rows={len(summary_rows)}")
        detail_log(f"[J] PUBLISH: aggregate_entries={len(detail_rows)} hours={total_hours:.2f} gross={total_gross:.2f}")
        detail_log(f"[Perf] [DocketReport] total_time={time.perf_counter()-t_start:.3f}s")
        
        if trace_mode:
            try:
                out_dir = self.paths.root / "outputs"
                out_dir.mkdir(exist_ok=True)
                with open(out_dir / "docket_raw_preview.csv", "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=headers)
                    writer.writeheader()
                    writer.writerows(raw_rows[:200])
                with open(out_dir / "docket_parsed_preview.json", "w", encoding="utf-8") as f:
                    json.dump(parsed_rows[:200], f, indent=2, default=str)
                with open(out_dir / "docket_reject_summary.json", "w", encoding="utf-8") as f:
                    json.dump(rejects, f, indent=2)
                logger.info(f"[DocketReport] Dumped TRACE artifacts to {out_dir}")
            except Exception as e:
                logger.error(f"[DocketReport] Failed to write TRACE artifacts: {e}")

        return {
            "ok": True,
            "filters": {
                "fromDate": start_date.strftime("%Y-%m-%d"),
                "toDate": end_date.strftime("%Y-%m-%d"),
                "statusMode": status_mode,
                "clientFilter": client_filter,
                "matterFilter": matter_filter,
                "query": text_query
            },
            "rows": detail_rows,
            "summaryRows": summary_rows,
            "totalRows": len(detail_rows),
            "summaryCount": len(summary_rows),
            "optionClients": option_clients,
            "optionMatters": option_matters,
            "totals": {
                "totalHours": round(total_hours, 2),
                "totalGrossToClient": round(total_gross, 2)
            },
            "debug": {
                "scanned": len(raw_rows),
                "accepted": len(detail_rows),
                "rejected": sum(rejects.values()),
                "rejects": rejects
            },
            "message": ""
        }

    def export_docket_activity_csv(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        raw_payload = dict(payload or {})
        report_filters = raw_payload.get("filters") if isinstance(raw_payload.get("filters"), dict) else raw_payload
        report = self.get_docket_activity_report(dict(report_filters or {}))
        if not bool(report.get("ok")):
            return {
                "ok": False,
                "path": "",
                "filename": "",
                "rowCount": 0,
                "message": _clean_text(report.get("message")) or "Could not build report.",
            }

        # ensure we have the expected types so the static analyzer stops complaining
        # about union values. initialize first and then overwrite under guard.
        rows: List[Dict[str, Any]] = []
        if isinstance(report.get("rows"), list):
            # mypy/Pylance still think `.get` could return None, so silence with ignore
            rows = report.get("rows")  # type: ignore[assignment]
        summary_rows: List[Dict[str, Any]] = []
        if isinstance(report.get("summaryRows"), list):
            summary_rows = report.get("summaryRows")  # type: ignore[assignment]
        totals: Dict[str, Any] = {}
        if isinstance(report.get("totals"), dict):
            totals = report.get("totals")  # type: ignore[assignment]

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        candidate_dirs = [
            self.paths.exports_dir(),
            self.paths.data_dir() / "exports",
            self.paths.root / "outputs",
        ]
        output_path: Optional[Path] = None
        last_error: Optional[Exception] = None

        for export_dir in candidate_dirs:
            try:
                export_dir.mkdir(parents=True, exist_ok=True)
            except Exception as exc:
                last_error = exc
                continue
            candidate_path = export_dir / f"docket_activity_{stamp}.csv"
            try:
                with candidate_path.open("w", encoding="utf-8-sig", newline="") as handle:
                    writer = csv.writer(handle)
                    writer.writerow(
                        [
                            "Date",
                            "Entry ID",
                            "Client",
                            "Matter",
                            "Description",
                            "Hours",
                            "GrossToClient",
                            "Status",
                            "Rate",
                            "SharePct",
                            "RawSeconds",
                        ]
                    )
                    for row in rows:
                        writer.writerow(
                            [
                                _clean_text(row.get("date")),
                                _clean_text(row.get("entryId")),
                                _clean_text(row.get("clientName")),
                                _clean_text(row.get("matterName")),
                                _clean_text(row.get("description")),
                                round(float(self._parse_float(row.get("hours")) or 0.0), 2),
                                round(float(self._parse_float(row.get("grossToClient")) or 0.0), 2),
                                _clean_text(row.get("status")),
                                round(float(self._parse_float(row.get("rate")) or 0.0), 2),
                                round(float(self._parse_float(row.get("sharePct")) or 0.0), 2),
                                int(self._parse_float(row.get("rawSeconds")) or 0),
                            ]
                        )

                    writer.writerow([])
                    writer.writerow(["Summary by Matter"])
                    writer.writerow(["Matter", "Client", "Entries", "Hours", "GrossToClient"])
                    for row in summary_rows:
                        writer.writerow(
                            [
                                _clean_text(row.get("matterName")),
                                _clean_text(row.get("clientName")),
                                int(self._parse_float(row.get("entryCount")) or 0),
                                round(float(self._parse_float(row.get("totalHours")) or 0.0), 2),
                                round(float(self._parse_float(row.get("totalGrossToClient")) or 0.0), 2),
                            ]
                        )
                    writer.writerow([])
                    writer.writerow(
                        [
                            "Totals",
                            "",
                            "",
                            round(float(self._parse_float(totals.get("totalHours")) or 0.0), 2),
                            round(float(self._parse_float(totals.get("totalGrossToClient")) or 0.0), 2),
                        ]
                    )
                output_path = candidate_path
                break
            except PermissionError as exc:
                last_error = exc
                continue
            except Exception as exc:
                last_error = exc
                continue

        if output_path is None:
            raise PermissionError(f"Could not write CSV export to any configured directory: {last_error}")

        return {
            "ok": True,
            "path": str(output_path),
            "filename": output_path.name,
            "rowCount": len(rows),
            "message": f"CSV exported: {output_path.name} | Saved to: {output_path.parent}",
        }

    def save_client_profile(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.ensure_schema()
        normalized = self._normalize_client_profile_payload(payload)
        force_duplicate = bool(normalized.get("forceDuplicate"))

        parent_row: Optional[Dict[str, Any]] = None
        if normalized["parentClientName"]:
            parent_row = self._get_or_create_parent(normalized["parentClientName"])

        requested_client_id = _clean_text(normalized.get("clientId"))
        client_row: Optional[Dict[str, Any]] = None
        if requested_client_id:
            client_rows = [self._canonicalize_client_row(r) for r in self._read_table_rows(TBL_CLIENTS)]
            for row in client_rows:
                if _clean_text(row.get(sc.COL_CLIENT_ID)).lower() == requested_client_id.lower():
                    client_row = row
                    break
        if client_row is None and not (force_duplicate and not requested_client_id):
            client_row = self._get_or_create_client(normalized["clientName"])

        client_id = requested_client_id or _clean_text((client_row or {}).get(sc.COL_CLIENT_ID))
        if not client_id:
            client_id = self._new_id("C")
        parent_id = _clean_text(parent_row.get(sc.COL_PARENT_ID)) if parent_row else normalized["parentClientId"]

        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        existing_profile = self._find_client_profile(client_id)
        created_at = _clean_text(existing_profile.get(sc.COL_PROFILE_CREATED)) if existing_profile else now_stamp
        if not created_at:
            created_at = now_stamp

        client_master_row = {
            sc.COL_CLIENT_ID: client_id,
            sc.COL_CLIENT_NAME: normalized["clientName"],
            sc.COL_CLIENT_EMAIL: normalized["primaryEmail"],
            sc.COL_CLIENT_PHONE: normalized["primaryPhone"],
            sc.COL_CLIENT_STATUS: normalized["status"],
            sc.COL_CLIENT_ACTIVE: normalized["active"],
            sc.COL_CLIENT_NOTES: normalized["notes"],
        }
        self._upsert_row_by_key(TBL_CLIENTS, sc.COL_CLIENT_ID, client_id, client_master_row)

        full_address = normalized["fullAddress"] or self._format_full_address(
            line1=normalized["addressLine1"],
            line2=normalized["addressLine2"],
            city=normalized["city"],
            state_province=normalized["stateProvince"],
            postal_code=normalized["postalCode"],
            country=normalized["country"],
        )

        profile_row = {
            sc.COL_PROFILE_CLIENT_ID: client_id,
            sc.COL_PROFILE_LEGAL_NAME: normalized["legalName"],
            sc.COL_PROFILE_DISPLAY_NAME: normalized["displayName"] or normalized["clientName"],
            sc.COL_PROFILE_FIRST_NAME: normalized["firstName"],
            sc.COL_PROFILE_MIDDLE_NAME: normalized["middleName"],
            sc.COL_PROFILE_LAST_NAME: normalized["lastName"],
            sc.COL_PROFILE_ENTITY_TYPE: normalized["entityType"],
            sc.COL_PROFILE_PRINCIPAL_NAME: normalized["principalName"],
            sc.COL_PROFILE_PRINCIPAL_POSITION: normalized["principalPosition"],
            sc.COL_PROFILE_PRIMARY_EMAIL: normalized["primaryEmail"],
            sc.COL_PROFILE_PRIMARY_PHONE: normalized["primaryPhone"],
            sc.COL_PROFILE_SECONDARY_CONTACT: normalized["secondaryContactName"],
            sc.COL_PROFILE_SECONDARY_POSITION: normalized["secondaryContactPosition"],
            sc.COL_PROFILE_SECONDARY_EMAIL: normalized["secondaryContactEmail"],
            sc.COL_PROFILE_SECONDARY_PHONE: normalized["secondaryContactPhone"],
            sc.COL_PROFILE_ADDR1: normalized["addressLine1"],
            sc.COL_PROFILE_ADDR2: normalized["addressLine2"],
            sc.COL_PROFILE_CITY: normalized["city"],
            sc.COL_PROFILE_STATE: normalized["stateProvince"],
            sc.COL_PROFILE_POSTAL: normalized["postalCode"],
            sc.COL_PROFILE_COUNTRY: normalized["country"],
            sc.COL_PROFILE_FULL_ADDRESS: full_address,
            sc.COL_PROFILE_PARENT_ID: parent_id,
            sc.COL_PROFILE_PARENT_NAME: normalized["parentClientName"],
            sc.COL_PROFILE_WEBSITE: normalized["website"],
            sc.COL_PROFILE_TAX_ID: normalized["taxId"],
            sc.COL_PROFILE_INDUSTRY: normalized["industry"],
            sc.COL_PROFILE_BILLING_EMAIL: normalized["billingEmail"],
            sc.COL_PROFILE_KYC_STATUS: normalized["kycStatus"],
            sc.COL_PROFILE_ONBOARDING_STATUS: normalized["onboardingStatus"],
            sc.COL_PROFILE_RETAINER_REQUIRED: normalized["retainerRequired"],
            sc.COL_PROFILE_RETAINER_AMOUNT: round(float(normalized["retainerAmount"]), 2),
            sc.COL_PROFILE_ENGAGEMENT_START: normalized["engagementStartDate"],
            sc.COL_PROFILE_DATE_CLIENT_ADDED: normalized["dateClientAdded"],
            sc.COL_PROFILE_BIRTHDAY: normalized["birthday"],
            sc.COL_PROFILE_REFERRAL_FROM: normalized["referralFrom"],
            sc.COL_PROFILE_CONFLICT_NOTES: normalized["conflictNotes"],
            sc.COL_PROFILE_NOTES: normalized["notes"],
            sc.COL_PROFILE_CREATED: created_at,
            sc.COL_PROFILE_UPDATED: now_stamp,
        }

        self._upsert_row_by_key(TBL_CLIENT_PROFILES, sc.COL_PROFILE_CLIENT_ID, client_id, profile_row)
        persisted = self._find_client_profile(client_id)
        verified = self._compare_client_profile_rows_loose(profile_row, persisted)

        return {
            "ok": bool(verified),
            "verifiedExact": bool(verified),
            "clientId": client_id,
            "savedRow": profile_row,
            "message": "" if verified else "Client profile verification failed.",
        }

    def save_matter_profile(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.ensure_schema()
        normalized = self._normalize_matter_profile_payload(payload)
        # Joint matter and party rows are a single logical record.  Use the
        # repository batch path so a failed party validation never leaves a
        # newly-created matter pointing at only its legacy file anchor.
        if normalized["parties"] and not self._import_batch_active:
            with self.import_batch():
                return self.save_matter_profile(payload)
        force_duplicate = bool(normalized.get("forceDuplicate"))

        client_id = normalized["clientId"]
        client_name = normalized["clientName"]
        resolved_parties: List[Dict[str, Any]] = []
        if normalized["parties"]:
            if normalized["parentId"] or normalized["parentName"]:
                raise ValueError(
                    "A joint retainer cannot use a Billing Client relationship. Select invoice recipients from the matter parties."
                )
            client_rows_for_parties = [
                self._canonicalize_client_row(row) for row in self._read_table_rows(TBL_CLIENTS)
            ]
            client_by_id = {
                _clean_text(row.get(sc.COL_CLIENT_ID)).casefold(): row
                for row in client_rows_for_parties
                if _clean_text(row.get(sc.COL_CLIENT_ID))
            }
            client_by_name = {
                _clean_text(row.get(sc.COL_CLIENT_NAME)).casefold(): row
                for row in client_rows_for_parties
                if _clean_text(row.get(sc.COL_CLIENT_NAME))
            }
            seen_party_ids: set[str] = set()
            for index, raw_party in enumerate(normalized["parties"], start=1):
                requested_id = _clean_text(raw_party.get("clientId") or raw_party.get(sc.COL_MATTER_PARTY_CLIENT_ID))
                requested_name = _clean_text(raw_party.get("clientName") or raw_party.get(sc.COL_MATTER_PARTY_CLIENT_NAME))
                source_row = client_by_id.get(requested_id.casefold()) if requested_id else None
                if source_row is None and requested_name:
                    source_row = client_by_name.get(requested_name.casefold())
                if source_row is None:
                    raise ValueError(
                        f"Matter party {index} must be an existing client. Create the client record first."
                    )
                resolved_id = _clean_text(source_row.get(sc.COL_CLIENT_ID))
                resolved_name = _clean_text(source_row.get(sc.COL_CLIENT_NAME))
                if resolved_id.casefold() in seen_party_ids:
                    raise ValueError("A client can be listed only once on a matter.")
                seen_party_ids.add(resolved_id.casefold())
                resolved_parties.append({
                    "clientId": resolved_id,
                    "clientName": resolved_name,
                    "role": _clean_text(raw_party.get("role") or raw_party.get(sc.COL_MATTER_PARTY_ROLE)) or "Client",
                    "isFileAnchor": self._to_bool_int(
                        raw_party.get("isFileAnchor", raw_party.get(sc.COL_MATTER_PARTY_IS_FILE_ANCHOR)),
                        default=0,
                    ),
                    "isBillingRecipient": self._to_bool_int(
                        raw_party.get(
                            "isBillingRecipient",
                            raw_party.get(sc.COL_MATTER_PARTY_IS_BILLING_RECIPIENT),
                        ),
                        default=1,
                    ),
                    "sortOrder": index,
                    "notes": _clean_text(raw_party.get("notes") or raw_party.get(sc.COL_MATTER_PARTY_NOTES)),
                })
            anchors = [party for party in resolved_parties if party["isFileAnchor"] == 1]
            if not anchors:
                resolved_parties[0]["isFileAnchor"] = 1
                anchors = [resolved_parties[0]]
            if len(anchors) != 1:
                raise ValueError("Select exactly one file anchor for the matter's legacy timekeeping link.")
            client_id = anchors[0]["clientId"]
            client_name = anchors[0]["clientName"]
        if client_id:
            client_rows = [self._canonicalize_client_row(r) for r in self._read_table_rows(TBL_CLIENTS)]
            for row in client_rows:
                if _clean_text(row.get(sc.COL_CLIENT_ID)).lower() == client_id.lower():
                    client_name = _clean_text(row.get(sc.COL_CLIENT_NAME))
                    break
            if not client_name:
                client_id = ""
        if not client_id and client_name:
            client_row = self._get_or_create_client(client_name)
            client_id = _clean_text(client_row.get(sc.COL_CLIENT_ID))
            client_name = _clean_text(client_row.get(sc.COL_CLIENT_NAME))

        if not client_id:
            raise ValueError("Client is required.")

        parent_id = normalized["parentId"]
        parent_name = normalized["parentName"]
        if parent_name:
            parent_row = self._get_or_create_parent(parent_name)
            parent_id = _clean_text(parent_row.get(sc.COL_PARENT_ID))
            parent_name = _clean_text(parent_row.get(sc.COL_PARENT_NAME))
        elif parent_id:
            parent_rows = [self._canonicalize_parent_row(r) for r in self._read_table_rows(TBL_PARENTS)]
            for row in parent_rows:
                if _clean_text(row.get(sc.COL_PARENT_ID)).lower() == parent_id.lower():
                    parent_name = _clean_text(row.get(sc.COL_PARENT_NAME))
                    break

        existing_row: Optional[Dict[str, Any]] = None
        matter_id = normalized["matterId"]
        if matter_id:
            existing_row = self._find_matter_row(matter_id)
        if existing_row is None and not (force_duplicate and not matter_id):
            matter_rows = [self._canonicalize_matter_row(r) for r in self._read_table_rows(TBL_MATTERS)]
            target_name_lc = normalized["matterName"].lower()
            for row in matter_rows:
                if _clean_text(row.get(sc.COL_MATTER_NAME)).lower() != target_name_lc:
                    continue
                row_client_id = _clean_text(row.get(sc.COL_MATTER_CLIENT_ID))
                if client_id and row_client_id and row_client_id.lower() != client_id.lower():
                    continue
                existing_row = row
                break
            if existing_row is None and normalized.get("matterNumber"):
                target_num_lc = normalized["matterNumber"].lower()
                for row in matter_rows:
                    if _clean_text(row.get(sc.COL_MATTER_NUMBER)).lower() == target_num_lc:
                        existing_row = row
                        break
        if existing_row is not None:
            matter_id = _clean_text(existing_row.get(sc.COL_MATTER_ID))
        if not matter_id:
            matter_id = self._new_id("M")

        # A matter with open financial work must remain operational.  This is
        # deliberately enforced here (rather than only in QML), so imports or
        # future screens cannot archive/close/on-hold a file which still needs
        # billing or collection work.
        if existing_row is not None:
            prior_status = _clean_text(existing_row.get(sc.COL_MATTER_STATUS)).casefold()
            requested_status = _clean_text(normalized.get("status")).casefold()
            protected_statuses = {"on hold", "closed", "archived"}
            if requested_status in protected_statuses and requested_status != prior_status:
                financial = self.get_matter_financial_summary(matter_id)
                if financial.get("ok") and financial.get("hasFinancialBlockers"):
                    return {
                        "ok": False,
                        "matterId": matter_id,
                        "financialSummary": financial,
                        "message": self._matter_financial_blocker_message(
                            financial,
                            action=f"set the matter status to {normalized.get('status')}",
                        ),
                    }
                if not financial.get("ok"):
                    return {
                        "ok": False,
                        "matterId": matter_id,
                        "message": financial.get(
                            "message",
                            "Could not verify unbilled WIP and unpaid invoices. The status was not changed for safety.",
                        ),
                    }

        # Resolve entity_type from the client profile so individual-vs-company
        # naming rules work correctly.
        entity_type = ""
        if client_id:
            profile_rows = self._read_table_rows(TBL_CLIENT_PROFILES)
            for pr in profile_rows:
                if _clean_text(pr.get(sc.COL_PROFILE_CLIENT_ID, "")).lower() == client_id.lower():
                    entity_type = _clean_text(pr.get(sc.COL_PROFILE_ENTITY_TYPE, ""))
                    break

        matter_number = _clean_text(normalized.get("matterNumber"))
        if not matter_number and existing_row is not None:
            matter_number = _clean_text(existing_row.get(sc.COL_MATTER_NUMBER))
        if not matter_number:
            matter_number = self._build_matter_number(
                client_name=client_name,
                matter_type=normalized["matterType"],
                date_opened=normalized["dateOpened"],
                existing_matter_id=matter_id,
                parent_name=parent_name,
                entity_type=entity_type,
            )
        matter_number = _clean_text(matter_number).upper()
        if self._matter_number_in_use(matter_number, ignore_matter_id=matter_id):
            if _clean_text(normalized.get("matterNumber")):
                raise ValueError(f"Matter number already exists: {matter_number}")
            matter_number = self._build_matter_number(
                client_name=client_name,
                matter_type=normalized["matterType"],
                date_opened=normalized["dateOpened"],
                existing_matter_id=matter_id,
                parent_name=parent_name,
                entity_type=entity_type,
            )

        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        created_at = _clean_text((existing_row or {}).get(sc.COL_MATTER_CREATED)) or now_stamp

        matter_row = {
            sc.COL_MATTER_ID: matter_id,
            sc.COL_MATTER_NUMBER: matter_number,
            sc.COL_MATTER_NAME: normalized["matterName"],
            sc.COL_MATTER_DISPLAY_NAME: normalized["displayName"] or normalized["matterName"],
            sc.COL_MATTER_CLIENT_ID: client_id,
            sc.COL_MATTER_CLIENT_NAME: client_name,
            sc.COL_MATTER_PARENT_ID: parent_id,
            sc.COL_MATTER_PARENT_NAME: parent_name,
            sc.COL_MATTER_TYPE: normalized["matterType"],
            sc.COL_MATTER_PRACTICE_AREA: normalized["practiceArea"],
            sc.COL_MATTER_STATUS: normalized["status"],
            sc.COL_MATTER_RESPONSIBLE_LAWYER: normalized["responsibleLawyer"],
            sc.COL_MATTER_BILLING_ARRANGEMENT: normalized["billingArrangement"],
            sc.COL_MATTER_BILLING_CONTACT: normalized["billingContact"],
            sc.COL_MATTER_BILLING_EMAIL: normalized["billingEmail"],
            sc.COL_MATTER_DEF_RATE: round(float(normalized["defaultRate"]), 2),
            sc.COL_MATTER_DEF_SHARE: round(float(normalized["defaultSharePct"]), 2),
            sc.COL_MATTER_ENGAGEMENT_DATE: normalized["dateOfEngagement"],
            sc.COL_MATTER_OPEN_DATE: normalized["dateOpened"],
            sc.COL_MATTER_CLOSE_DATE: normalized["dateClosed"],
            sc.COL_MATTER_COURT_FILE_NO: normalized["courtFileNumber"],
            sc.COL_MATTER_OPPOSING_PARTY: normalized["opposingParty"],
            sc.COL_MATTER_REFERRAL_FROM: normalized["referralFrom"],
            sc.COL_MATTER_DESCRIPTION: normalized["description"],
            sc.COL_MATTER_NOTES: normalized["notes"],
            sc.COL_MATTER_REPRESENTATION_MODE: normalized["representationMode"],
            sc.COL_MATTER_JOINT_NO_CONFIDENTIALITY_CONFIRMED: normalized["jointNoConfidentialityConfirmed"],
            sc.COL_MATTER_JOINT_INSTRUCTIONS_REQUIRE_ALL: normalized["jointInstructionsRequireAll"],
            sc.COL_MATTER_JOINT_ENGAGEMENT_DOCUMENT: normalized["jointEngagementDocument"],
            sc.COL_MATTER_CREATED: created_at,
            sc.COL_MATTER_UPDATED: now_stamp,
        }

        self._upsert_row_by_key(TBL_MATTERS, sc.COL_MATTER_ID, matter_id, matter_row)
        if resolved_parties:
            existing_parties = [
                self._canonicalize_matter_party_row(row)
                for row in self._read_table_rows(TBL_MATTER_PARTIES)
            ]
            existing_by_client = {
                _clean_text(row.get(sc.COL_MATTER_PARTY_CLIENT_ID)).casefold(): row
                for row in existing_parties
                if _clean_text(row.get(sc.COL_MATTER_PARTY_MATTER_ID)).casefold() == matter_id.casefold()
            }
            retained_parties = [
                row
                for row in existing_parties
                if _clean_text(row.get(sc.COL_MATTER_PARTY_MATTER_ID)).casefold() != matter_id.casefold()
            ]
            for party in resolved_parties:
                existing_party = existing_by_client.get(party["clientId"].casefold(), {})
                retained_parties.append({
                    sc.COL_MATTER_PARTY_ID: _clean_text(existing_party.get(sc.COL_MATTER_PARTY_ID)) or self._new_id("MP"),
                    sc.COL_MATTER_PARTY_MATTER_ID: matter_id,
                    sc.COL_MATTER_PARTY_CLIENT_ID: party["clientId"],
                    sc.COL_MATTER_PARTY_CLIENT_NAME: party["clientName"],
                    sc.COL_MATTER_PARTY_ROLE: party["role"],
                    sc.COL_MATTER_PARTY_IS_FILE_ANCHOR: party["isFileAnchor"],
                    sc.COL_MATTER_PARTY_IS_BILLING_RECIPIENT: party["isBillingRecipient"],
                    sc.COL_MATTER_PARTY_SORT_ORDER: party["sortOrder"],
                    sc.COL_MATTER_PARTY_NOTES: party["notes"],
                    sc.COL_MATTER_PARTY_CREATED: _clean_text(existing_party.get(sc.COL_MATTER_PARTY_CREATED)) or now_stamp,
                    sc.COL_MATTER_PARTY_UPDATED: now_stamp,
                })
            self._replace_table_rows(TBL_MATTER_PARTIES, retained_parties)

        if self._import_batch_active:
            verified = True
        else:
            persisted = self._find_matter_row(matter_id)
            verified = self._compare_matter_profile_rows_loose(matter_row, persisted)

        return {
            "ok": bool(verified),
            "verifiedExact": bool(verified),
            "matterId": matter_id,
            "matterNumber": matter_number,
            "savedRow": matter_row,
            "partyCount": len(resolved_parties),
            "message": "" if verified else "Matter profile verification failed.",
        }

    def run_conflict_check(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        data = dict(payload or {})
        client_key = _clean_text(data.get("clientKey") or data.get("client") or data.get("clientName"))
        matter_key = _clean_text(data.get("matterKey") or data.get("matter") or data.get("matterName"))
        subject_text = _clean_text(data.get("subjectText") or data.get("query"))
        extra_text = _clean_text(data.get("extraTerms"))

        client_profile: Dict[str, Any] = {}
        matter_profile: Dict[str, Any] = {}
        if client_key:
            loaded = self.get_client_profile(client_key)
            if bool(loaded.get("ok")):
                client_profile = dict(loaded.get("client") or {})
        if matter_key:
            loaded = self.get_matter_profile(matter_key)
            if bool(loaded.get("ok")):
                matter_profile = dict(loaded.get("matter") or {})

        term_candidates = [
            subject_text,
            client_key,
            matter_key,
            _clean_text(client_profile.get("displayName")),
            _clean_text(client_profile.get("legalName")),
            _clean_text(client_profile.get("primaryEmail")),
            _clean_text(client_profile.get("primaryPhone")),
            _clean_text(client_profile.get("secondaryContactEmail")),
            _clean_text(client_profile.get("secondaryContactPhone")),
            _clean_text(client_profile.get("taxId")),
            _clean_text(client_profile.get("parentClientName")),
            _clean_text(matter_profile.get("displayName")),
            _clean_text(matter_profile.get("matterName")),
            _clean_text(matter_profile.get("matterNumber")),
            _clean_text(matter_profile.get("opposingParty")),
            _clean_text(matter_profile.get("clientName")),
        ]
        if extra_text:
            term_candidates.extend(_search_terms(extra_text))

        terms: List[str] = []
        seen_terms = set()
        for raw in term_candidates:
            text = _clean_text(raw)
            if not text:
                continue
            key = text.lower()
            if key in seen_terms:
                continue
            seen_terms.add(key)
            terms.append(text)
            if len(terms) >= 12:
                break

        if not terms:
            return {
                "ok": False,
                "message": "Enter at least one client, matter, or subject term.",
                "riskLevel": "none",
                "termsUsed": [],
                "matches": [],
                "totalMatches": 0,
                "summary": {"client": 0, "matter": 0, "parent": 0},
                "checkedAtUtc": datetime.now(UTC).replace(tzinfo=None).isoformat() + "Z",
            }

        selected_client_id = _clean_text(client_profile.get("clientId")).lower()
        selected_matter_id = _clean_text(matter_profile.get("matterId")).lower()
        aggregate: Dict[str, Dict[str, Any]] = {}
        summary = {"client": 0, "matter": 0, "parent": 0}

        for term in terms:
            search = self.search_global_entities(term, mode="any", limit=250)
            for row in search.get("results", []):
                entity_type = _clean_text(row.get("entityType")).lower()
                if entity_type not in ("client", "matter", "parent"):
                    continue
                entity_id = _clean_text(row.get("entityId"))
                title = _clean_text(row.get("title")) or entity_id
                if not title:
                    continue
                key = f"{entity_type}|{entity_id.lower()}|{title.lower()}"
                existing = aggregate.get(key)
                if existing is None:
                    existing = {
                        "entityType": entity_type,
                        "entityId": entity_id,
                        "title": title,
                        "subtitle": _clean_text(row.get("subtitle")),
                        "status": _clean_text(row.get("status")),
                        "routeTileIndex": int(row.get("routeTileIndex", 0)),
                        "routeNodeId": _clean_text(row.get("routeNodeId")),
                        "routeNodeTitle": _clean_text(row.get("routeNodeTitle")),
                        "matchedFields": list(row.get("matchedFields") or []),
                        "_hitTerms": [],
                        "_hits": 0,
                    }
                    aggregate[key] = existing

                term_lc = term.lower()
                if term_lc not in existing["_hitTerms"]:
                    existing["_hitTerms"].append(term_lc)
                    existing["_hits"] += 1

        matches: List[Dict[str, Any]] = []
        for row in aggregate.values():
            entity_type = _clean_text(row.get("entityType")).lower()
            entity_id_lc = _clean_text(row.get("entityId")).lower()
            if entity_type == "client" and selected_client_id and entity_id_lc == selected_client_id:
                continue
            if entity_type == "matter" and selected_matter_id and entity_id_lc == selected_matter_id:
                continue

            score = int(row.get("_hits", 0)) * 10
            if entity_type == "client":
                score += 4
            elif entity_type == "matter":
                score += 2
            else:
                score += 1

            clean_row = {
                "entityType": entity_type,
                "entityId": _clean_text(row.get("entityId")),
                "title": _clean_text(row.get("title")),
                "subtitle": _clean_text(row.get("subtitle")),
                "status": _clean_text(row.get("status")),
                "routeTileIndex": int(row.get("routeTileIndex", 0)),
                "routeNodeId": _clean_text(row.get("routeNodeId")),
                "routeNodeTitle": _clean_text(row.get("routeNodeTitle")),
                "matchedFields": list(row.get("matchedFields") or []),
                "matchedTerms": list(row.get("_hitTerms") or []),
                "score": score,
            }
            matches.append(clean_row)
            if entity_type in summary:
                summary[entity_type] += 1

        matches.sort(
            key=lambda row: (
                -int(row.get("score", 0)),
                _normalize_search_text(row.get("entityType")),
                _normalize_search_text(row.get("title")),
                _normalize_search_text(row.get("entityId")),
            )
        )
        for row in matches:
            row.pop("score", None)

        top_hits = max((len(row.get("matchedTerms") or []) for row in matches), default=0)
        total = len(matches)
        if total <= 0:
            risk_level = "none"
        elif top_hits >= 3 or total >= 8:
            risk_level = "high"
        elif top_hits >= 2 or total >= 3:
            risk_level = "medium"
        else:
            risk_level = "low"

        return {
            "ok": True,
            "message": f"{total} potential conflict match(es) found.",
            "riskLevel": risk_level,
            "termsUsed": terms,
            "matches": matches[:120],
            "totalMatches": total,
            "summary": summary,
            "checkedAtUtc": datetime.now(UTC).replace(tzinfo=None).isoformat() + "Z",
        }

    def reassign_matter(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.ensure_schema()
        data = dict(payload or {})
        matter_key = _clean_text(data.get("matterKey") or data.get("matterId") or data.get("matterName"))
        from_lawyer = _clean_text(data.get("fromLawyer"))
        to_lawyer = _clean_text(
            data.get("toLawyer")
            or data.get("newResponsibleLawyer")
            or data.get("responsibleLawyer")
        )
        reason = _clean_text(data.get("reason"))
        actor = _clean_text(data.get("actor"))

        if not matter_key:
            return {"ok": False, "message": "Matter key is required.", "changed": False}
        if not to_lawyer:
            return {"ok": False, "message": "New responsible lawyer is required.", "changed": False}

        matter_row = self._find_matter_row(matter_key)
        if not matter_row:
            return {"ok": False, "message": f"Matter not found: {matter_key}", "changed": False}

        matter_id = _clean_text(matter_row.get(sc.COL_MATTER_ID))
        matter_name = _clean_text(matter_row.get(sc.COL_MATTER_DISPLAY_NAME)) or _clean_text(
            matter_row.get(sc.COL_MATTER_NAME)
        )
        current_lawyer = _clean_text(matter_row.get(sc.COL_MATTER_RESPONSIBLE_LAWYER))

        if from_lawyer and current_lawyer and from_lawyer.lower() != current_lawyer.lower():
            return {
                "ok": False,
                "message": f"Current lawyer mismatch (expected '{from_lawyer}', found '{current_lawyer}').",
                "changed": False,
                "matterId": matter_id,
            }

        if current_lawyer.lower() == to_lawyer.lower():
            return {
                "ok": True,
                "changed": False,
                "matterId": matter_id,
                "matterName": matter_name,
                "fromLawyer": current_lawyer,
                "toLawyer": to_lawyer,
                "message": "Matter already assigned to that lawyer.",
                "updatedAtUtc": datetime.now(UTC).replace(tzinfo=None).isoformat() + "Z",
            }

        rows = [self._canonicalize_matter_row(r) for r in self._read_table_rows(TBL_MATTERS)]
        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        audit_note = (
            f"[Reassigned {now_stamp}] {current_lawyer or '[unassigned]'} -> {to_lawyer}"
            + (f"; Reason: {reason}" if reason else "")
            + (f"; Actor: {actor}" if actor else "")
        )

        changed_rows = 0
        for row in rows:
            row_id = _clean_text(row.get(sc.COL_MATTER_ID))
            if row_id.lower() != matter_id.lower():
                continue
            row[sc.COL_MATTER_RESPONSIBLE_LAWYER] = to_lawyer
            row[sc.COL_MATTER_UPDATED] = now_stamp
            row[sc.COL_MATTER_NOTES] = self._append_note_line(row.get(sc.COL_MATTER_NOTES), audit_note)
            changed_rows += 1
            break

        if changed_rows <= 0:
            return {"ok": False, "message": "Matter reassignment failed to apply.", "changed": False}

        self._replace_table_rows(TBL_MATTERS, rows)
        persisted = self._find_matter_row(matter_id)
        persisted_lawyer = _clean_text((persisted or {}).get(sc.COL_MATTER_RESPONSIBLE_LAWYER))
        verified = persisted_lawyer.lower() == to_lawyer.lower()

        return {
            "ok": bool(verified),
            "changed": True,
            "verifiedExact": bool(verified),
            "matterId": matter_id,
            "matterName": matter_name,
            "fromLawyer": current_lawyer,
            "toLawyer": to_lawyer,
            "message": "" if verified else "Matter reassignment verification failed.",
            "updatedAtUtc": datetime.now(UTC).replace(tzinfo=None).isoformat() + "Z",
            "auditNote": audit_note,
        }

    def merge_duplicate_entities(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        data = dict(payload or {})
        merge_type = _clean_text(data.get("mergeType") or "client").lower()
        if merge_type in ("matter", "matters"):
            return self._merge_duplicate_matters(data)
        return self._merge_duplicate_clients(data)

    def _merge_duplicate_clients(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.ensure_schema()
        source_key = _clean_text(payload.get("sourceKey") or payload.get("sourceClientKey") or payload.get("source"))
        target_key = _clean_text(payload.get("targetKey") or payload.get("targetClientKey") or payload.get("target"))
        reason = _clean_text(payload.get("reason"))
        actor = _clean_text(payload.get("actor"))

        if not source_key or not target_key:
            return {"ok": False, "message": "Source and target clients are required.", "changed": False}

        source_profile = self.get_client_profile(source_key)
        target_profile = self.get_client_profile(target_key)
        if not bool(source_profile.get("ok")):
            return {"ok": False, "message": f"Source client not found: {source_key}", "changed": False}
        if not bool(target_profile.get("ok")):
            return {"ok": False, "message": f"Target client not found: {target_key}", "changed": False}

        source_client = dict(source_profile.get("client") or {})
        target_client = dict(target_profile.get("client") or {})
        source_id = _clean_text(source_client.get("clientId"))
        target_id = _clean_text(target_client.get("clientId"))
        source_name = _clean_text(source_client.get("displayName")) or _clean_text(source_client.get("clientName"))
        target_name = _clean_text(target_client.get("displayName")) or _clean_text(target_client.get("clientName"))

        if not source_id or not target_id:
            return {"ok": False, "message": "Both source and target client IDs are required.", "changed": False}
        if source_id.lower() == target_id.lower():
            return {"ok": False, "message": "Source and target clients must be different.", "changed": False}

        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        audit_note = (
            f"[Merged {now_stamp}] Source client '{source_name or source_id}' -> '{target_name or target_id}'"
            + (f"; Reason: {reason}" if reason else "")
            + (f"; Actor: {actor}" if actor else "")
        )

        clients_rows = [self._canonicalize_client_row(r) for r in self._read_table_rows(TBL_CLIENTS)]
        profile_rows = [self._canonicalize_client_profile_row(r) for r in self._read_table_rows(TBL_CLIENT_PROFILES)]
        matter_rows = [self._canonicalize_matter_row(r) for r in self._read_table_rows(TBL_MATTERS)]
        time_rows = [self._canonicalize_time_row(r) for r in self._read_table_rows(TBL_TIME)]
        txn_rows = [self._canonicalize_transaction_row(r) for r in self._read_table_rows(TBL_TRANSACTIONS_MASTER)]
        try:
            matter_party_rows = [
                self._canonicalize_matter_party_row(row)
                for row in self._read_table_rows(TBL_MATTER_PARTIES)
            ]
        except Exception:
            # Compatibility for pre-joint-retainer workbooks and lightweight
            # repository adapters used by older integrations.
            matter_party_rows = []

        counts = {"clients": 0, "profiles": 0, "matters": 0, "timeEntries": 0, "transactions": 0}

        for row in clients_rows:
            row_id = _clean_text(row.get(sc.COL_CLIENT_ID))
            if row_id.lower() != source_id.lower():
                continue
            row[sc.COL_CLIENT_STATUS] = "Archived"
            row[sc.COL_CLIENT_ACTIVE] = 0
            row[sc.COL_CLIENT_NOTES] = self._append_note_line(row.get(sc.COL_CLIENT_NOTES), audit_note)
            counts["clients"] += 1

        source_aliases = {source_id.lower()}
        if source_name:
            source_aliases.add(source_name.lower())

        for row in profile_rows:
            row_client_id = _clean_text(row.get(sc.COL_PROFILE_CLIENT_ID))
            parent_id = _clean_text(row.get(sc.COL_PROFILE_PARENT_ID))
            parent_name = _clean_text(row.get(sc.COL_PROFILE_PARENT_NAME))

            if row_client_id.lower() == source_id.lower():
                row[sc.COL_PROFILE_NOTES] = self._append_note_line(row.get(sc.COL_PROFILE_NOTES), audit_note)
                row[sc.COL_PROFILE_ONBOARDING_STATUS] = "Archived"
                row[sc.COL_PROFILE_UPDATED] = now_stamp
                counts["profiles"] += 1
                continue

            if parent_id.lower() == source_id.lower() or (parent_name and parent_name.lower() in source_aliases):
                row[sc.COL_PROFILE_PARENT_ID] = target_id
                row[sc.COL_PROFILE_PARENT_NAME] = target_name
                row[sc.COL_PROFILE_UPDATED] = now_stamp
                counts["profiles"] += 1

        for row in matter_rows:
            row_client_id = _clean_text(row.get(sc.COL_MATTER_CLIENT_ID))
            row_client_name = _clean_text(row.get(sc.COL_MATTER_CLIENT_NAME))
            if row_client_id.lower() != source_id.lower() and row_client_name.lower() not in source_aliases:
                continue
            row[sc.COL_MATTER_CLIENT_ID] = target_id
            row[sc.COL_MATTER_CLIENT_NAME] = target_name
            row[sc.COL_MATTER_UPDATED] = now_stamp
            row[sc.COL_MATTER_NOTES] = self._append_note_line(row.get(sc.COL_MATTER_NOTES), audit_note)
            counts["matters"] += 1

        for row in time_rows:
            row_client_id = _clean_text(row.get(sc.COL_TIME_CLIENT_ID))
            if row_client_id.lower() != source_id.lower():
                continue
            row[sc.COL_TIME_CLIENT_ID] = target_id
            row[sc.COL_TIME_LOCK_AUDIT] = self._append_note_line(row.get(sc.COL_TIME_LOCK_AUDIT), audit_note)
            counts["timeEntries"] += 1

        for row in txn_rows:
            row_client = _clean_text(row.get(sc.COL_TXN_CLIENT))
            if row_client.lower() not in source_aliases:
                continue
            row[sc.COL_TXN_CLIENT] = target_name or target_id
            row[sc.COL_TXN_UPDATED_AT] = now_stamp
            row[sc.COL_TXN_NOTES] = self._append_note_line(row.get(sc.COL_TXN_NOTES), audit_note)
            counts["transactions"] += 1

        matter_party_changed = False
        for row in matter_party_rows:
            if _clean_text(row.get(sc.COL_MATTER_PARTY_CLIENT_ID)).lower() != source_id.lower():
                continue
            row[sc.COL_MATTER_PARTY_CLIENT_ID] = target_id
            row[sc.COL_MATTER_PARTY_CLIENT_NAME] = target_name
            row[sc.COL_MATTER_PARTY_NOTES] = self._append_note_line(
                row.get(sc.COL_MATTER_PARTY_NOTES), audit_note
            )
            row[sc.COL_MATTER_PARTY_UPDATED] = now_stamp
            matter_party_changed = True

        self._replace_table_rows(TBL_CLIENTS, clients_rows)
        self._replace_table_rows(TBL_CLIENT_PROFILES, profile_rows)
        self._replace_table_rows(TBL_MATTERS, matter_rows)
        self._replace_table_rows(TBL_TIME, time_rows)
        self._replace_table_rows(TBL_TRANSACTIONS_MASTER, txn_rows)
        if matter_party_changed:
            self._replace_table_rows(TBL_MATTER_PARTIES, matter_party_rows)

        source_after = self.get_client_profile(source_id)
        target_after = self.get_client_profile(target_id)
        source_archived = False
        if bool(source_after.get("ok")):
            src = dict(source_after.get("client") or {})
            src_status = _clean_text(src.get("status")).lower()
            src_active = self._to_bool_int(src.get("active"), default=1)
            source_archived = src_status in ("archived", "inactive", "closed") or src_active == 0
        verified = bool(target_after.get("ok")) and source_archived

        return {
            "ok": bool(verified),
            "changed": True,
            "verifiedExact": bool(verified),
            "mergeType": "client",
            "sourceId": source_id,
            "targetId": target_id,
            "sourceName": source_name,
            "targetName": target_name,
            "counts": counts,
            "message": "" if verified else "Client merge verification incomplete.",
            "updatedAtUtc": datetime.now(UTC).replace(tzinfo=None).isoformat() + "Z",
            "auditNote": audit_note,
        }

    def _merge_duplicate_matters(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.ensure_schema()
        source_key = _clean_text(payload.get("sourceKey") or payload.get("sourceMatterKey") or payload.get("source"))
        target_key = _clean_text(payload.get("targetKey") or payload.get("targetMatterKey") or payload.get("target"))
        reason = _clean_text(payload.get("reason"))
        actor = _clean_text(payload.get("actor"))

        if not source_key or not target_key:
            return {"ok": False, "message": "Source and target matters are required.", "changed": False}

        source_profile = self.get_matter_profile(source_key)
        target_profile = self.get_matter_profile(target_key)
        if not bool(source_profile.get("ok")):
            return {"ok": False, "message": f"Source matter not found: {source_key}", "changed": False}
        if not bool(target_profile.get("ok")):
            return {"ok": False, "message": f"Target matter not found: {target_key}", "changed": False}

        source_matter = dict(source_profile.get("matter") or {})
        target_matter = dict(target_profile.get("matter") or {})
        source_id = _clean_text(source_matter.get("matterId"))
        target_id = _clean_text(target_matter.get("matterId"))
        source_name = _clean_text(source_matter.get("displayName")) or _clean_text(source_matter.get("matterName"))
        target_name = _clean_text(target_matter.get("displayName")) or _clean_text(target_matter.get("matterName"))
        source_number = _clean_text(source_matter.get("matterNumber"))
        target_number = _clean_text(target_matter.get("matterNumber"))

        if not source_id or not target_id:
            return {"ok": False, "message": "Both source and target matter IDs are required.", "changed": False}
        if source_id.lower() == target_id.lower():
            return {"ok": False, "message": "Source and target matters must be different.", "changed": False}

        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        audit_note = (
            f"[Merged {now_stamp}] Source matter '{source_name or source_id}' -> '{target_name or target_id}'"
            + (f"; Reason: {reason}" if reason else "")
            + (f"; Actor: {actor}" if actor else "")
        )

        matter_rows = [self._canonicalize_matter_row(r) for r in self._read_table_rows(TBL_MATTERS)]
        time_rows = [self._canonicalize_time_row(r) for r in self._read_table_rows(TBL_TIME)]
        disb_rows = [self._canonicalize_row(TBL_DISBURSEMENTS, r) for r in self._read_table_rows(TBL_DISBURSEMENTS)]
        txn_rows = [self._canonicalize_transaction_row(r) for r in self._read_table_rows(TBL_TRANSACTIONS_MASTER)]
        try:
            matter_party_rows = [
                self._canonicalize_matter_party_row(row)
                for row in self._read_table_rows(TBL_MATTER_PARTIES)
            ]
        except Exception:
            matter_party_rows = []
        counts = {"matters": 0, "timeEntries": 0, "disbursements": 0, "transactions": 0}

        for row in matter_rows:
            row_id = _clean_text(row.get(sc.COL_MATTER_ID))
            if row_id.lower() != source_id.lower():
                continue
            row[sc.COL_MATTER_STATUS] = "Archived"
            row[sc.COL_MATTER_UPDATED] = now_stamp
            row[sc.COL_MATTER_NOTES] = self._append_note_line(row.get(sc.COL_MATTER_NOTES), audit_note)
            counts["matters"] += 1

        source_aliases = {source_id.lower()}
        if source_name:
            source_aliases.add(source_name.lower())
        if source_number:
            source_aliases.add(source_number.lower())

        for row in time_rows:
            row_matter_id = _clean_text(row.get(sc.COL_TIME_MATTER_ID))
            if row_matter_id.lower() != source_id.lower():
                continue
            row[sc.COL_TIME_MATTER_ID] = target_id
            row[sc.COL_TIME_LOCK_AUDIT] = self._append_note_line(row.get(sc.COL_TIME_LOCK_AUDIT), audit_note)
            counts["timeEntries"] += 1

        for row in disb_rows:
            row_matter_id = _clean_text(row.get(sc.COL_DISB_MATTER_ID))
            if row_matter_id.lower() != source_id.lower():
                continue
            row[sc.COL_DISB_MATTER_ID] = target_id
            counts["disbursements"] += 1

        for row in txn_rows:
            row_matter = _clean_text(row.get(sc.COL_TXN_MATTER))
            if row_matter.lower() not in source_aliases:
                continue
            row[sc.COL_TXN_MATTER] = target_number or target_name or target_id
            row[sc.COL_TXN_UPDATED_AT] = now_stamp
            row[sc.COL_TXN_NOTES] = self._append_note_line(row.get(sc.COL_TXN_NOTES), audit_note)
            counts["transactions"] += 1

        target_party_client_ids = {
            _clean_text(row.get(sc.COL_MATTER_PARTY_CLIENT_ID)).casefold()
            for row in matter_party_rows
            if _clean_text(row.get(sc.COL_MATTER_PARTY_MATTER_ID)).casefold() == target_id.casefold()
        }
        retained_party_rows: List[Dict[str, Any]] = []
        matter_party_changed = False
        for row in matter_party_rows:
            if _clean_text(row.get(sc.COL_MATTER_PARTY_MATTER_ID)).casefold() != source_id.casefold():
                retained_party_rows.append(row)
                continue
            matter_party_changed = True
            client_key = _clean_text(row.get(sc.COL_MATTER_PARTY_CLIENT_ID)).casefold()
            if client_key in target_party_client_ids:
                continue
            row[sc.COL_MATTER_PARTY_MATTER_ID] = target_id
            row[sc.COL_MATTER_PARTY_NOTES] = self._append_note_line(
                row.get(sc.COL_MATTER_PARTY_NOTES), audit_note
            )
            row[sc.COL_MATTER_PARTY_UPDATED] = now_stamp
            retained_party_rows.append(row)
            target_party_client_ids.add(client_key)

        self._replace_table_rows(TBL_MATTERS, matter_rows)
        self._replace_table_rows(TBL_TIME, time_rows)
        self._replace_table_rows(TBL_DISBURSEMENTS, disb_rows)
        self._replace_table_rows(TBL_TRANSACTIONS_MASTER, txn_rows)
        if matter_party_changed:
            self._replace_table_rows(TBL_MATTER_PARTIES, retained_party_rows)

        source_after = self.get_matter_profile(source_id)
        target_after = self.get_matter_profile(target_id)
        source_archived = False
        if bool(source_after.get("ok")):
            src = dict(source_after.get("matter") or {})
            src_status = _clean_text(src.get("status")).lower()
            source_archived = src_status in ("archived", "inactive", "closed")
        verified = bool(target_after.get("ok")) and source_archived

        return {
            "ok": bool(verified),
            "changed": True,
            "verifiedExact": bool(verified),
            "mergeType": "matter",
            "sourceId": source_id,
            "targetId": target_id,
            "sourceName": source_name,
            "targetName": target_name,
            "counts": counts,
            "message": "" if verified else "Matter merge verification incomplete.",
            "updatedAtUtc": datetime.now(UTC).replace(tzinfo=None).isoformat() + "Z",
            "auditNote": audit_note,
        }

    def _ensure_table_schema(self, wb, tref: TableRef) -> Tuple[bool, str]:
        headers = TABLE_COLUMNS[tref.table]
        if tref.sheet in wb.sheetnames:
            ws = wb[tref.sheet]
        else:
            ws = wb.create_sheet(tref.sheet)

        existing = ws.tables[tref.table] if tref.table in ws.tables else None
        if existing is not None:
            existing_headers, existing_rows = self._rows_from_table(ws, existing)
            if existing_headers == headers:
                return False, "unchanged"

            migrated_rows = [self._canonicalize_row(tref, row) for row in existing_rows]
            self._write_table(
                ws=ws,
                existing_table=existing,
                table_name=tref.table,
                headers=headers,
                rows=migrated_rows,
                style=existing.tableStyleInfo,
            )
            return True, "migrated"

        # If this table already exists on another sheet (legacy layout), relocate it.
        source_ws, source_table = self._find_table_in_workbook(wb, tref.table)
        if source_ws is not None and source_table is not None:
            source_headers, source_rows = self._rows_from_table(source_ws, source_table)
            migrated_rows = [self._canonicalize_row(tref, row) for row in source_rows]
            source_style = source_table.tableStyleInfo
            if source_ws.title != ws.title:
                self._delete_table(source_ws, tref.table)
                self._remove_empty_legacy_sheet(wb, source_ws)
            self._write_table(
                ws=ws,
                existing_table=None,
                table_name=tref.table,
                headers=headers,
                rows=migrated_rows,
                style=source_style,
            )
            if source_headers == headers:
                return True, "relocated"
            return True, "relocated+migrated"

        self._write_table(ws, None, tref.table, headers, [], None)
        return True, "created"

    def _seed_transaction_lookup_tables_if_empty(self, wb) -> List[Dict[str, Any]]:
        seed_specs = [
            {
                "tref": TBL_TRANSACTION_ACCOUNTS,
                "seedFile": "transactions_master.accounts.seed.csv",
                "keyColumn": sc.COL_TXN_ACCOUNT_CODE,
            },
            {
                "tref": TBL_TRANSACTION_CATEGORIES,
                "seedFile": "transactions_master.categories.seed.csv",
                "keyColumn": sc.COL_TXN_CATEGORY_LKP_CODE,
            },
            {
                "tref": TBL_TRANSACTION_BUSINESS_UNITS,
                "seedFile": "transactions_master.business_units.seed.csv",
                "keyColumn": sc.COL_TXN_BUSINESS_UNIT_NAME,
            },
            {
                "tref": TBL_TRANSACTION_PAYEES,
                "seedFile": "transactions_master.payees.seed.csv",
                "keyColumn": sc.COL_TXN_PAYEE_NAME,
            },
        ]

        seeded: List[Dict[str, Any]] = []
        for spec in seed_specs:
            tref = spec["tref"]
            if tref.sheet not in wb.sheetnames:
                continue
            ws = wb[tref.sheet]
            if tref.table not in ws.tables:
                continue

            table = ws.tables[tref.table]
            _headers, existing_rows = self._rows_from_table(ws, table)
            if existing_rows:
                continue

            seed_file = _clean_text(spec.get("seedFile"))
            seed_rows = self._load_seed_csv_rows(seed_file)
            if not seed_rows:
                continue

            key_column = _clean_text(spec.get("keyColumn"))
            canonical_rows: List[Dict[str, Any]] = []
            seen_keys = set()
            for seed_row in seed_rows:
                canonical_row = self._canonicalize_row(tref, seed_row)
                dedupe_key = _clean_text(canonical_row.get(key_column)).lower() if key_column else ""
                if dedupe_key and dedupe_key in seen_keys:
                    continue
                if dedupe_key:
                    seen_keys.add(dedupe_key)
                canonical_rows.append(canonical_row)

            if not canonical_rows:
                continue

            self._write_table(
                ws=ws,
                existing_table=table,
                table_name=tref.table,
                headers=TABLE_COLUMNS[tref.table],
                rows=canonical_rows,
                style=table.tableStyleInfo,
            )
            seeded.append(
                {
                    "table": tref.table,
                    "rows": len(canonical_rows),
                    "seedFile": seed_file,
                }
            )

        return seeded

    def _find_table_in_workbook(self, wb, table_name: str) -> Tuple[Optional[Worksheet], Optional[Table]]:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if table_name in ws.tables:
                return ws, ws.tables[table_name]
        return None, None

    def _delete_table(self, ws: Worksheet, table_name: str) -> None:
        if table_name not in ws.tables:
            return
        table = ws.tables[table_name]
        min_col, min_row, max_col, max_row = _safe_range_boundaries(table.ref)
        del ws.tables[table_name]
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                _set_cell_value(ws, row_idx, col_idx, None)

    def _remove_empty_legacy_sheet(self, wb, ws: Worksheet) -> None:
        if ws.title != sc.SHEET_TRANSACTION_LOOKUPS:
            return
        if len(ws.tables) > 0:
            return
        if len(wb.sheetnames) <= 1:
            return

        has_content = False
        if ws.max_row > 0 and ws.max_column > 0:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if _clean_text(cell.value):
                        has_content = True
                        break
                if has_content:
                    break
        if not has_content:
            wb.remove(ws)

    def _write_table(
        self,
        ws: Worksheet,
        existing_table: Optional[Table],
        table_name: str,
        headers: List[str],
        rows: List[Dict[str, Any]],
        style: Optional[TableStyleInfo],
    ) -> None:
        if existing_table is not None:
            min_col, min_row, old_max_col, old_max_row = _safe_range_boundaries(existing_table.ref)
        else:
            min_col = 1
            min_row = self._next_table_start_row(ws)
            old_max_col = len(headers)
            old_max_row = min_row + 1

        write_rows = rows if rows else [{}]
        new_max_col = min_col + len(headers) - 1
        new_max_row = min_row + len(write_rows)

        clear_max_col = max(old_max_col, new_max_col)
        clear_max_row = max(old_max_row, new_max_row)
        for row_idx in range(min_row, clear_max_row + 1):
            for col_idx in range(min_col, clear_max_col + 1):
                _set_cell_value(ws, row_idx, col_idx, None)

        for col_idx, header in enumerate(headers, start=min_col):
            _set_cell_value(ws, min_row, col_idx, header)

        for offset, row_data in enumerate(write_rows, start=1):
            row_idx = min_row + offset
            for col_idx, header in enumerate(headers, start=min_col):
                _set_cell_value(ws, row_idx, col_idx, row_data.get(header, ""))

        ref = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(new_max_col)}{new_max_row}"
        )

        if existing_table is not None:
            if len(existing_table.tableColumns) == len(headers):
                existing_table.ref = ref
                if style is not None:
                    existing_table.tableStyleInfo = style
                return
            else:
                # Remove the old table to avoid XML schema corruption when column counts mismatch
                try:
                    if existing_table.name in ws.tables:
                        del ws.tables[existing_table.name]
                except Exception:
                    pass

        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = style or TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    def _next_table_start_row(self, ws: Worksheet) -> int:
        max_row = 0
        for existing in ws.tables.values():
            _, _, _, table_max_row = _safe_range_boundaries(existing.ref)
            if table_max_row > max_row:
                max_row = table_max_row
        return max_row + 2 if max_row > 0 else 1

    def _rows_from_table(self, ws: Worksheet, table: Table) -> Tuple[List[str], List[Dict[str, Any]]]:
        min_col, min_row, max_col, max_row = _safe_range_boundaries(table.ref)
        headers = []
        for col_idx in range(min_col, max_col + 1):
            headers.append(_clean_text(ws.cell(row=min_row, column=col_idx).value))

        rows: List[Dict[str, Any]] = []
        for row_idx in range(min_row + 1, max_row + 1):
            row: Dict[str, Any] = {}
            has_data = False
            for offset, col_idx in enumerate(range(min_col, max_col + 1)):
                key = headers[offset]
                value = ws.cell(row=row_idx, column=col_idx).value
                if value not in (None, ""):
                    has_data = True
                row[key] = value
            if has_data:
                rows.append(row)
        return headers, rows

    def _read_raw_excel_table_rows(
        self,
        sheet_name: str,
        table_name: str = "",
        workbook_path: Optional[Path] = None,
    ) -> List[Dict[str, Any]]:
        """Read a non-schema worksheet/table without repairing or canonicalizing it."""
        _lazy_load_heavy_libs()
        path = workbook_path or self.paths.workbook_path()
        if not path.exists():
            return []

        wb = load_workbook(path, keep_vba=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return []
            ws = wb[sheet_name]
            if table_name:
                try:
                    if hasattr(ws, "tables") and ws.tables and table_name in ws.tables:
                        _headers, rows = self._rows_from_table(ws, ws.tables[table_name])
                        return rows
                except Exception:
                    pass

            header_row = 1
            headers = [_clean_text(cell.value) for cell in ws[header_row]]
            if not any(headers):
                return []

            rows: List[Dict[str, Any]] = []
            for row_idx in range(header_row + 1, int(ws.max_row or header_row) + 1):
                row: Dict[str, Any] = {}
                has_data = False
                for col_idx, header in enumerate(headers, start=1):
                    if not header:
                        continue
                    value = ws.cell(row=row_idx, column=col_idx).value
                    row[header] = value
                    if value not in (None, ""):
                        has_data = True
                if has_data:
                    rows.append(row)
            return rows
        except Exception:
            return []
        finally:
            self._close_workbook(wb)

    def _read_legacy_docket_rows(self) -> List[Dict[str, Any]]:
        """Return cached legacy docket rows used only to preserve invoice context.

        Older invoices may have been finalized before their docket rows were
        copied into ``tblTimeEntries``.  Their client/matter identity remains
        in the companion Dockets workbook and is needed for a recipient-facing
        statement; this method never writes to either workbook.
        """
        path = self.paths.dockets_workbook_path()
        signature = self._workbook_signature(path)
        if signature and signature == self._legacy_docket_rows_signature:
            return self._legacy_docket_rows_cache
        rows = self._read_raw_excel_table_rows("Dockets", workbook_path=path)
        self._legacy_docket_rows_cache = rows
        self._legacy_docket_rows_signature = signature
        return rows

    def _canonicalize_row(self, tref: TableRef, row: Dict[str, Any]) -> Dict[str, Any]:
        table_name = tref.table
        if table_name == TBL_PARENTS.table:
            return self._canonicalize_parent_row(row)
        if table_name == TBL_CLIENTS.table:
            return self._canonicalize_client_row(row)
        if table_name == TBL_CLIENT_PROFILES.table:
            return self._canonicalize_client_profile_row(row)
        if table_name == TBL_MATTERS.table:
            return self._canonicalize_matter_row(row)
        if table_name == TBL_MATTER_PARTIES.table:
            return self._canonicalize_matter_party_row(row)
        if table_name == TBL_TIME.table:
            return self._canonicalize_time_row(row)
        if table_name == TBL_TRADEMARKS.table:
            return self._canonicalize_trademark_row(row)
        if table_name in {
            TBL_DISBURSEMENTS.table,
            TBL_LEDGER.table,
            TBL_RECEIVABLES.table,
            TBL_INVOICE_LOG.table,
        }:
            return self._canonicalize_simple_table_row(tref, row)
        if table_name == TBL_TRANSACTIONS_MASTER.table:
            return self._canonicalize_transaction_row(row)
        if table_name == TBL_TRANSACTION_ACCOUNTS.table:
            return self._canonicalize_transaction_account_row(row)
        if table_name == TBL_TRANSACTION_CATEGORIES.table:
            return self._canonicalize_transaction_category_row(row)
        if table_name == TBL_TRANSACTION_BUSINESS_UNITS.table:
            return self._canonicalize_transaction_business_unit_row(row)
        if table_name == TBL_TRANSACTION_PAYEES.table:
            return self._canonicalize_transaction_payee_row(row)
        return self._canonicalize_simple_table_row(tref, row)

    def _canonicalize_simple_table_row(self, tref: TableRef, row: Dict[str, Any]) -> Dict[str, Any]:
        headers = TABLE_COLUMNS.get(tref.table)
        if not headers:
            return dict(row or {})
        normalized: Dict[str, Any] = {}
        for header in headers:
            value = self._value_with_alias(tref.table, row, header)
            if isinstance(value, datetime):
                value = value.date().isoformat()
            elif isinstance(value, date):
                value = value.isoformat()
            normalized[header] = "" if value is None else value
        return normalized

    def _canonicalize_parent_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        parent_id = _clean_text(self._value_with_alias(TBL_PARENTS.table, row, sc.COL_PARENT_ID))
        if not parent_id:
            parent_id = self._new_id("P")

        parent_name = _clean_text(self._value_with_alias(TBL_PARENTS.table, row, sc.COL_PARENT_NAME))
        share_pct = normalize_pct(self._value_with_alias(TBL_PARENTS.table, row, sc.COL_PARENT_DEF_SHARE), default_pct=100.0)
        default_rate = float(self._parse_float(self._value_with_alias(TBL_PARENTS.table, row, sc.COL_PARENT_DEF_RATE)) or 0.0)
        active = self._to_bool_int(self._value_with_alias(TBL_PARENTS.table, row, sc.COL_PARENT_ACTIVE), default=1)
        notes = _clean_text(self._value_with_alias(TBL_PARENTS.table, row, sc.COL_PARENT_NOTES))

        return {
            sc.COL_PARENT_ID: parent_id,
            sc.COL_PARENT_NAME: parent_name,
            sc.COL_PARENT_DEF_SHARE: round(share_pct, 2),
            sc.COL_PARENT_DEF_RATE: round(default_rate, 2),
            sc.COL_PARENT_ACTIVE: active,
            sc.COL_PARENT_NOTES: notes,
        }

    def _canonicalize_client_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        client_id = _clean_text(self._value_with_alias(TBL_CLIENTS.table, row, sc.COL_CLIENT_ID))
        if not client_id:
            client_id = self._new_id("C")

        status = _clean_text(self._value_with_alias(TBL_CLIENTS.table, row, sc.COL_CLIENT_STATUS))
        if not status:
            status = "Active"

        return {
            sc.COL_CLIENT_ID: client_id,
            sc.COL_CLIENT_NAME: _clean_text(self._value_with_alias(TBL_CLIENTS.table, row, sc.COL_CLIENT_NAME)),
            sc.COL_CLIENT_EMAIL: _clean_text(self._value_with_alias(TBL_CLIENTS.table, row, sc.COL_CLIENT_EMAIL)),
            sc.COL_CLIENT_PHONE: _clean_text(self._value_with_alias(TBL_CLIENTS.table, row, sc.COL_CLIENT_PHONE)),
            sc.COL_CLIENT_STATUS: status,
            sc.COL_CLIENT_ACTIVE: self._to_bool_int(self._value_with_alias(TBL_CLIENTS.table, row, sc.COL_CLIENT_ACTIVE), default=1),
            sc.COL_CLIENT_NOTES: _clean_text(self._value_with_alias(TBL_CLIENTS.table, row, sc.COL_CLIENT_NOTES)),
        }

    def _canonicalize_client_profile_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        client_id = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_CLIENT_ID)
        )
        if not client_id:
            client_id = self._new_id("C")

        display_name = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_DISPLAY_NAME)
        )
        legal_name = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_LEGAL_NAME)
        )
        if not display_name:
            display_name = legal_name
        entity_type = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_ENTITY_TYPE)
        )
        first_name = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_FIRST_NAME)
        )
        middle_name = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_MIDDLE_NAME)
        )
        last_name = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_LAST_NAME)
        )
        if _is_individual_entity_type(entity_type) and not (first_name and last_name):
            split_first, split_middle, split_last = _split_person_name_parts(display_name or legal_name)
            if not first_name:
                first_name = split_first
            if not middle_name:
                middle_name = split_middle
            if not last_name:
                last_name = split_last

        address_line1 = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_ADDR1)
        )
        address_line2 = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_ADDR2)
        )
        city = _clean_text(self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_CITY))
        state_province = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_STATE)
        )
        postal_code = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_POSTAL)
        )
        country = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_COUNTRY)
        )
        full_address = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_FULL_ADDRESS)
        )
        if not full_address:
            full_address = self._format_full_address(
                line1=address_line1,
                line2=address_line2,
                city=city,
                state_province=state_province,
                postal_code=postal_code,
                country=country,
            )

        onboarding_status = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_ONBOARDING_STATUS)
        )
        if not onboarding_status:
            onboarding_status = "Prospect"

        kyc_status = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_KYC_STATUS)
        )
        if not kyc_status:
            kyc_status = "Pending"

        created_at = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_CREATED)
        )
        if not created_at:
            created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        updated_at = _clean_text(
            self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_UPDATED)
        )
        if not updated_at:
            updated_at = created_at

        return {
            sc.COL_PROFILE_CLIENT_ID: client_id,
            sc.COL_PROFILE_LEGAL_NAME: legal_name,
            sc.COL_PROFILE_DISPLAY_NAME: display_name,
            sc.COL_PROFILE_FIRST_NAME: first_name,
            sc.COL_PROFILE_MIDDLE_NAME: middle_name,
            sc.COL_PROFILE_LAST_NAME: last_name,
            sc.COL_PROFILE_ENTITY_TYPE: entity_type,
            sc.COL_PROFILE_PRINCIPAL_NAME: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_PRINCIPAL_NAME)
            ),
            sc.COL_PROFILE_PRINCIPAL_POSITION: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_PRINCIPAL_POSITION)
            ),
            sc.COL_PROFILE_PRIMARY_EMAIL: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_PRIMARY_EMAIL)
            ),
            sc.COL_PROFILE_PRIMARY_PHONE: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_PRIMARY_PHONE)
            ),
            sc.COL_PROFILE_SECONDARY_CONTACT: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_SECONDARY_CONTACT)
            ),
            sc.COL_PROFILE_SECONDARY_POSITION: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_SECONDARY_POSITION)
            ),
            sc.COL_PROFILE_SECONDARY_EMAIL: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_SECONDARY_EMAIL)
            ),
            sc.COL_PROFILE_SECONDARY_PHONE: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_SECONDARY_PHONE)
            ),
            sc.COL_PROFILE_ADDR1: address_line1,
            sc.COL_PROFILE_ADDR2: address_line2,
            sc.COL_PROFILE_CITY: city,
            sc.COL_PROFILE_STATE: state_province,
            sc.COL_PROFILE_POSTAL: postal_code,
            sc.COL_PROFILE_COUNTRY: country,
            sc.COL_PROFILE_FULL_ADDRESS: full_address,
            sc.COL_PROFILE_PARENT_ID: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_PARENT_ID)
            ),
            sc.COL_PROFILE_PARENT_NAME: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_PARENT_NAME)
            ),
            sc.COL_PROFILE_WEBSITE: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_WEBSITE)
            ),
            sc.COL_PROFILE_TAX_ID: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_TAX_ID)
            ),
            sc.COL_PROFILE_INDUSTRY: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_INDUSTRY)
            ),
            sc.COL_PROFILE_BILLING_EMAIL: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_BILLING_EMAIL)
            ),
            sc.COL_PROFILE_KYC_STATUS: kyc_status,
            sc.COL_PROFILE_ONBOARDING_STATUS: onboarding_status,
            sc.COL_PROFILE_RETAINER_REQUIRED: self._to_bool_int(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_RETAINER_REQUIRED),
                default=0,
            ),
            sc.COL_PROFILE_RETAINER_AMOUNT: round(
                float(
                    self._parse_float(
                        self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_RETAINER_AMOUNT)
                    )
                    or 0.0
                ),
                2,
            ),
            sc.COL_PROFILE_ENGAGEMENT_START: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_ENGAGEMENT_START)
            ),
            sc.COL_PROFILE_DATE_CLIENT_ADDED: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_DATE_CLIENT_ADDED)
            ),
            sc.COL_PROFILE_BIRTHDAY: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_BIRTHDAY)
            ),
            sc.COL_PROFILE_REFERRAL_FROM: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_REFERRAL_FROM)
            ),
            sc.COL_PROFILE_CONFLICT_NOTES: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_CONFLICT_NOTES)
            ),
            sc.COL_PROFILE_NOTES: _clean_text(
                self._value_with_alias(TBL_CLIENT_PROFILES.table, row, sc.COL_PROFILE_NOTES)
            ),
            sc.COL_PROFILE_CREATED: created_at,
            sc.COL_PROFILE_UPDATED: updated_at,
        }

    def _canonicalize_matter_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        matter_id = _clean_text(self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_ID))
        if not matter_id:
            matter_id = self._new_id("M")

        matter_number = _clean_text(
            self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_NUMBER)
        ).upper()
        if not matter_number:
            matter_number = matter_id

        matter_name = _clean_text(self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_NAME))
        display_name = _clean_text(self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_DISPLAY_NAME))
        if not display_name:
            display_name = matter_name

        status = _clean_text(self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_STATUS))
        if not status:
            status = "Open"

        matter_type = _clean_text(self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_TYPE))
        if not matter_type:
            matter_type = "General"

        billing_arrangement = _clean_text(
            self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_BILLING_ARRANGEMENT)
        )
        if not billing_arrangement:
            billing_arrangement = "Hourly"

        created_at = _clean_text(self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_CREATED))
        if not created_at:
            created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        updated_at = _clean_text(self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_UPDATED))
        if not updated_at:
            updated_at = created_at

        return {
            sc.COL_MATTER_ID: matter_id,
            sc.COL_MATTER_NUMBER: matter_number,
            sc.COL_MATTER_NAME: matter_name,
            sc.COL_MATTER_DISPLAY_NAME: display_name,
            sc.COL_MATTER_CLIENT_ID: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_CLIENT_ID)
            ),
            sc.COL_MATTER_CLIENT_NAME: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_CLIENT_NAME)
            ),
            sc.COL_MATTER_PARENT_ID: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_PARENT_ID)
            ),
            sc.COL_MATTER_PARENT_NAME: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_PARENT_NAME)
            ),
            sc.COL_MATTER_TYPE: matter_type,
            sc.COL_MATTER_PRACTICE_AREA: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_PRACTICE_AREA)
            ),
            sc.COL_MATTER_STATUS: status,
            sc.COL_MATTER_RESPONSIBLE_LAWYER: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_RESPONSIBLE_LAWYER)
            ),
            sc.COL_MATTER_BILLING_ARRANGEMENT: billing_arrangement,
            sc.COL_MATTER_BILLING_CONTACT: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_BILLING_CONTACT)
            ),
            sc.COL_MATTER_BILLING_EMAIL: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_BILLING_EMAIL)
            ),
            sc.COL_MATTER_DEF_RATE: round(
                float(self._parse_float(self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_DEF_RATE)) or 0.0),
                2,
            ),
            sc.COL_MATTER_DEF_SHARE: round(
                normalize_pct(self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_DEF_SHARE), default_pct=100.0),
                2,
            ),
            sc.COL_MATTER_ENGAGEMENT_DATE: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_ENGAGEMENT_DATE)
            ),
            sc.COL_MATTER_OPEN_DATE: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_OPEN_DATE)
            ),
            sc.COL_MATTER_CLOSE_DATE: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_CLOSE_DATE)
            ),
            sc.COL_MATTER_COURT_FILE_NO: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_COURT_FILE_NO)
            ),
            sc.COL_MATTER_OPPOSING_PARTY: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_OPPOSING_PARTY)
            ),
            sc.COL_MATTER_REFERRAL_FROM: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_REFERRAL_FROM)
            ),
            sc.COL_MATTER_DESCRIPTION: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_DESCRIPTION)
            ),
            sc.COL_MATTER_NOTES: _clean_text(self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_NOTES)),
            sc.COL_MATTER_REPRESENTATION_MODE: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_REPRESENTATION_MODE)
            ) or "Single Client",
            sc.COL_MATTER_JOINT_NO_CONFIDENTIALITY_CONFIRMED: self._to_bool_int(
                self._value_with_alias(
                    TBL_MATTERS.table,
                    row,
                    sc.COL_MATTER_JOINT_NO_CONFIDENTIALITY_CONFIRMED,
                ),
                default=0,
            ),
            sc.COL_MATTER_JOINT_INSTRUCTIONS_REQUIRE_ALL: self._to_bool_int(
                self._value_with_alias(
                    TBL_MATTERS.table,
                    row,
                    sc.COL_MATTER_JOINT_INSTRUCTIONS_REQUIRE_ALL,
                ),
                default=0,
            ),
            sc.COL_MATTER_JOINT_ENGAGEMENT_DOCUMENT: _clean_text(
                self._value_with_alias(TBL_MATTERS.table, row, sc.COL_MATTER_JOINT_ENGAGEMENT_DOCUMENT)
            ),
            sc.COL_MATTER_CREATED: created_at,
            sc.COL_MATTER_UPDATED: updated_at,
        }

    def _canonicalize_matter_party_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        party_id = _clean_text(
            self._value_with_alias(TBL_MATTER_PARTIES.table, row, sc.COL_MATTER_PARTY_ID)
        )
        if not party_id:
            party_id = self._new_id("MP")
        created_at = _clean_text(
            self._value_with_alias(TBL_MATTER_PARTIES.table, row, sc.COL_MATTER_PARTY_CREATED)
        ) or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        updated_at = _clean_text(
            self._value_with_alias(TBL_MATTER_PARTIES.table, row, sc.COL_MATTER_PARTY_UPDATED)
        ) or created_at
        return {
            sc.COL_MATTER_PARTY_ID: party_id,
            sc.COL_MATTER_PARTY_MATTER_ID: _clean_text(
                self._value_with_alias(TBL_MATTER_PARTIES.table, row, sc.COL_MATTER_PARTY_MATTER_ID)
            ),
            sc.COL_MATTER_PARTY_CLIENT_ID: _clean_text(
                self._value_with_alias(TBL_MATTER_PARTIES.table, row, sc.COL_MATTER_PARTY_CLIENT_ID)
            ),
            sc.COL_MATTER_PARTY_CLIENT_NAME: _clean_text(
                self._value_with_alias(TBL_MATTER_PARTIES.table, row, sc.COL_MATTER_PARTY_CLIENT_NAME)
            ),
            sc.COL_MATTER_PARTY_ROLE: _clean_text(
                self._value_with_alias(TBL_MATTER_PARTIES.table, row, sc.COL_MATTER_PARTY_ROLE)
            ) or "Client",
            sc.COL_MATTER_PARTY_IS_FILE_ANCHOR: self._to_bool_int(
                self._value_with_alias(TBL_MATTER_PARTIES.table, row, sc.COL_MATTER_PARTY_IS_FILE_ANCHOR),
                default=0,
            ),
            sc.COL_MATTER_PARTY_IS_BILLING_RECIPIENT: self._to_bool_int(
                self._value_with_alias(
                    TBL_MATTER_PARTIES.table,
                    row,
                    sc.COL_MATTER_PARTY_IS_BILLING_RECIPIENT,
                ),
                default=1,
            ),
            sc.COL_MATTER_PARTY_SORT_ORDER: int(
                self._parse_int(
                    self._value_with_alias(TBL_MATTER_PARTIES.table, row, sc.COL_MATTER_PARTY_SORT_ORDER)
                )
                or 0
            ),
            sc.COL_MATTER_PARTY_NOTES: _clean_text(
                self._value_with_alias(TBL_MATTER_PARTIES.table, row, sc.COL_MATTER_PARTY_NOTES)
            ),
            sc.COL_MATTER_PARTY_CREATED: created_at,
            sc.COL_MATTER_PARTY_UPDATED: updated_at,
        }

    def _canonicalize_time_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        entry_id = _clean_text(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_ENTRY_ID))
        if not entry_id:
            entry_id = self._new_id("T")

        hours = float(self._parse_float(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_HOURS)) or 0.0)
        rate = float(self._parse_float(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_RATE)) or 0.0)
        share_pct = normalize_pct(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_SHARE_PCT), default_pct=100.0)

        raw_seconds = self._parse_int(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_SECONDS))
        if raw_seconds is None:
            raw_seconds = int(round(hours * 3600.0))
        raw_seconds = int(max(0, raw_seconds))

        if hours <= 0.0 and raw_seconds > 0:
            hours = math.ceil(((raw_seconds / 3600.0) * 10.0) - 1e-9) / 10.0

        amounts = calc_amounts(hours=hours, client_rate=rate, your_share_pct=share_pct, hst_rate=0.13)

        gross = self._parse_float(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_GROSS))
        amount_to_you = self._parse_float(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_NET))
        hst = self._parse_float(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_HST))
        total = self._parse_float(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_TOTAL))

        if gross is None:
            gross = amounts["gross_to_client"]
        if amount_to_you is None:
            amount_to_you = amounts["amount_to_you"]
        if hst is None:
            hst = amounts["hst_on_you"]
        if total is None:
            total = amounts["total_you_incl_hst"]

        status_raw = _clean_text(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_STATUS))
        if status_raw.lower() == "merged":
            status = "Merged"
        else:
            status = self._normalize_time_status(status_raw)

        created_at = _clean_text(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_CREATED))
        if not created_at:
            created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        return {
            sc.COL_TIME_ENTRY_ID: entry_id,
            sc.COL_TIME_DATE: _clean_text(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_DATE)),
            sc.COL_TIME_CLIENT_ID: _clean_text(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_CLIENT_ID)),
            sc.COL_TIME_MATTER_ID: _clean_text(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_MATTER_ID)),
            sc.COL_TIME_PARENT_ID: _clean_text(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_PARENT_ID)),
            sc.COL_TIME_DESC: _clean_text(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_DESC)),
            sc.COL_TIME_HOURS: round(hours, 2),
            sc.COL_TIME_RATE: round(rate, 2),
            sc.COL_TIME_SHARE_PCT: round(share_pct, 2),
            sc.COL_TIME_GROSS: round(float(gross), 2),
            sc.COL_TIME_NET: round(float(amount_to_you), 2),
            sc.COL_TIME_HST: round(float(hst), 2),
            sc.COL_TIME_TOTAL: round(float(total), 2),
            sc.COL_TIME_SECONDS: int(raw_seconds),
            sc.COL_TIME_STATUS: status,
            sc.COL_TIME_INVOICE_REF: _clean_text(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_INVOICE_REF)),
            sc.COL_TIME_INVOICE_STATUS: _clean_text(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_INVOICE_STATUS)),
            sc.COL_TIME_PAYMENT_STATUS: _clean_text(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_PAYMENT_STATUS)),
            sc.COL_TIME_INVOICE_TOTAL: round(float(self._parse_float(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_INVOICE_TOTAL)) or 0.0), 2),
            sc.COL_TIME_INVOICE_AMOUNT_PAID: round(float(self._parse_float(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_INVOICE_AMOUNT_PAID)) or 0.0), 2),
            sc.COL_TIME_INVOICE_BALANCE_DUE: round(float(self._parse_float(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_INVOICE_BALANCE_DUE)) or 0.0), 2),
            sc.COL_TIME_INVOICE_DATE: _clean_text(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_INVOICE_DATE)),
            sc.COL_TIME_LOCK_AUDIT: _clean_text(self._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_LOCK_AUDIT)),
            sc.COL_TIME_CREATED: created_at,
        }

    def _canonicalize_trademark_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        trademark_id = _clean_text(self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_ID))
        if not trademark_id:
            trademark_id = self._new_id("TM")

        jurisdiction = _normalize_choice(
            self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_JURISDICTION),
            self.TM_JURISDICTION_OPTIONS,
            "CIPO",
        )
        if not jurisdiction:
            jurisdiction = "CIPO"

        mark_type = _normalize_choice(
            self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_MARK_TYPE),
            self.TM_MARK_TYPE_OPTIONS,
            "Standard Character",
        )
        if not mark_type:
            mark_type = "Standard Character"

        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        created_at = _clean_text(self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_CREATED_AT))
        if not created_at:
            created_at = now_stamp
        updated_at = _clean_text(self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_UPDATED_AT))
        if not updated_at:
            updated_at = created_at

        return {
            sc.COL_TM_ID: trademark_id,
            sc.COL_TM_JURISDICTION: jurisdiction,
            sc.COL_TM_JURISDICTION_OTHER: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_JURISDICTION_OTHER)
            ),
            sc.COL_TM_CLIENT_NAME: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_CLIENT_NAME)
            ),
            sc.COL_TM_MATTER_NUMBER: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_MATTER_NUMBER)
            ),
            sc.COL_TM_INTERNAL_NOTES: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_INTERNAL_NOTES)
            ),
            sc.COL_TM_TRADEMARK_TEXT: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_TRADEMARK_TEXT)
            ),
            sc.COL_TM_MARK_TYPE: mark_type,
            sc.COL_TM_DESIGN_REPRESENTATION: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_DESIGN_REPRESENTATION)
            ),
            sc.COL_TM_DESIGN_IMAGE_PASTE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_DESIGN_IMAGE_PASTE)
            ),
            sc.COL_TM_COLOR_CLAIMED: self._to_bool_int(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_COLOR_CLAIMED),
                default=0,
            ),
            sc.COL_TM_COLOR_DESCRIPTION: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_COLOR_DESCRIPTION)
            ),
            sc.COL_TM_NICE_CLASSES: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_NICE_CLASSES)
            ),
            sc.COL_TM_GOODS_SERVICES: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_GOODS_SERVICES)
            ),
            sc.COL_TM_FOREIGN_PRIORITY: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_FOREIGN_PRIORITY)
            ),
            sc.COL_TM_REGISTRY_LINK: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_REGISTRY_LINK)
            ),
            sc.COL_TM_APPLICATION_NO: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_APPLICATION_NO)
            ),
            sc.COL_TM_REGISTRATION_NO: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_REGISTRATION_NO)
            ),
            sc.COL_TM_CURRENT_STATUS: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_CURRENT_STATUS)
            ),
            sc.COL_TM_APPLICANT_NAME_ADDRESS: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_APPLICANT_NAME_ADDRESS)
            ),
            sc.COL_TM_FILING_DATE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_FILING_DATE)
            ),
            sc.COL_TM_REGISTRATION_DATE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_REGISTRATION_DATE)
            ),
            sc.COL_TM_RENEWAL_DEADLINE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_RENEWAL_DEADLINE)
            ),
            sc.COL_TM_CIPO_STATUS: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_CIPO_STATUS)
            ),
            sc.COL_TM_TM5_STATUS: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_TM5_STATUS)
            ),
            sc.COL_TM_EXAMINERS_REPORT_DATE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_EXAMINERS_REPORT_DATE)
            ),
            sc.COL_TM_OFFICE_ACTION_RESPONSE_DEADLINE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_OFFICE_ACTION_RESPONSE_DEADLINE)
            ),
            sc.COL_TM_APPROVAL_DATE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_APPROVAL_DATE)
            ),
            sc.COL_TM_ADVERTISEMENT_DATE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_ADVERTISEMENT_DATE)
            ),
            sc.COL_TM_ADVERTISEMENT_VOL_ISSUE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_ADVERTISEMENT_VOL_ISSUE)
            ),
            sc.COL_TM_OPPOSITION_DEADLINE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_OPPOSITION_DEADLINE)
            ),
            sc.COL_TM_ALLOWANCE_DATE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_ALLOWANCE_DATE)
            ),
            sc.COL_TM_REGISTER_TYPE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_REGISTER_TYPE)
            ),
            sc.COL_TM_USPTO_STATUS_INDICATOR: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_USPTO_STATUS_INDICATOR)
            ),
            sc.COL_TM_OWNER_NAME_ADDRESS: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_OWNER_NAME_ADDRESS)
            ),
            sc.COL_TM_ATTORNEY_OF_RECORD: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_ATTORNEY_OF_RECORD)
            ),
            sc.COL_TM_PUBLICATION_DATE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_PUBLICATION_DATE)
            ),
            sc.COL_TM_NOTICE_OF_ALLOWANCE_DATE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_NOTICE_OF_ALLOWANCE_DATE)
            ),
            sc.COL_TM_SOU_DEADLINE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_SOU_DEADLINE)
            ),
            sc.COL_TM_SOU_EXTENSION_TRACKING: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_SOU_EXTENSION_TRACKING)
            ),
            sc.COL_TM_SECTION8_DEADLINE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_SECTION8_DEADLINE)
            ),
            sc.COL_TM_SECTION15_DEADLINE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_SECTION15_DEADLINE)
            ),
            sc.COL_TM_SECTION9_DEADLINE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_SECTION9_DEADLINE)
            ),
            sc.COL_TM_LOCAL_FOREIGN_ASSOCIATE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_LOCAL_FOREIGN_ASSOCIATE)
            ),
            sc.COL_TM_APPLICATION_REFERENCE_NO: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_APPLICATION_REFERENCE_NO)
            ),
            sc.COL_TM_PUBLICATION_ADVERTISEMENT_DATE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_PUBLICATION_ADVERTISEMENT_DATE)
            ),
            sc.COL_TM_OPPOSITION_PERIOD_END_DATE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_OPPOSITION_PERIOD_END_DATE)
            ),
            sc.COL_TM_UPCOMING_LOCAL_DEADLINE: _clean_text(
                self._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_UPCOMING_LOCAL_DEADLINE)
            ),
            sc.COL_TM_CREATED_AT: created_at,
            sc.COL_TM_UPDATED_AT: updated_at,
        }

    def _canonicalize_transaction_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        txn_id = _clean_text(self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_ID))
        if not txn_id:
            txn_id = self._new_id("TXN")

        txn_date = _clean_text(self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_DATE))
        if not _is_valid_iso_date(txn_date):
            txn_date = datetime.now().strftime("%Y-%m-%d")

        txn_class = _normalize_choice(
            self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_CLASS),
            self.TXN_CLASS_OPTIONS,
            "Family",
        )
        if not txn_class:
            txn_class = "Family"

        txn_type = _normalize_choice(
            self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_TYPE),
            self.TXN_TYPE_OPTIONS,
            "Expense",
        )
        if not txn_type:
            txn_type = "Expense"

        member = _normalize_choice(
            self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_MEMBER),
            self.TXN_MEMBER_OPTIONS,
            "Joint",
        )
        if not member:
            member = "Joint"

        tax_flag = _normalize_choice(
            self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_TAX_FLAG),
            self.TXN_TAX_FLAG_OPTIONS,
            "None",
        )
        if not tax_flag:
            tax_flag = "None"

        status = _normalize_choice(
            self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_STATUS),
            self.TXN_STATUS_OPTIONS,
            "Pending",
        )
        if not status:
            status = "Pending"

        currency = _normalize_choice(
            self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_CURRENCY),
            self.TXN_CURRENCY_OPTIONS,
            "CAD",
        )
        if not currency:
            currency = "CAD"

        amount = float(
            self._parse_float(self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_AMOUNT))
            or 0.0
        )
        tax_amount = float(
            self._parse_float(self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_TAX_AMOUNT))
            or 0.0
        )
        bill_claim_pct = float(
            self._parse_float(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_BILL_CLAIM_PCT)
            )
            or 0.0
        )
        if bill_claim_pct < 0:
            bill_claim_pct = 0.0
        if bill_claim_pct > 100:
            bill_claim_pct = 100.0

        hst_exempt = self._to_bool_int(
            self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_HST_EXEMPT),
            default=0,
        )
        if hst_exempt == 1:
            tax_amount = 0.0

        total_claim_existing = self._parse_float(
            self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_TOTAL_CLAIM_AMOUNT)
        )
        total_claim_amount = (
            float(total_claim_existing)
            if total_claim_existing is not None
            else round((amount + tax_amount) * (bill_claim_pct / 100.0), 2)
        )

        created_at = _clean_text(
            self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_CREATED_AT)
        )
        if not created_at:
            created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        updated_at = _clean_text(
            self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_UPDATED_AT)
        )
        if not updated_at:
            updated_at = created_at

        return {
            sc.COL_TXN_ID: txn_id,
            sc.COL_TXN_DATE: txn_date,
            sc.COL_TXN_CLASS: txn_class,
            sc.COL_TXN_BUSINESS_UNIT: _clean_text(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_BUSINESS_UNIT)
            ),
            sc.COL_TXN_TYPE: txn_type,
            sc.COL_TXN_FROM_ACCOUNT: _clean_text(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_FROM_ACCOUNT)
            ),
            sc.COL_TXN_TO_ACCOUNT: _clean_text(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_TO_ACCOUNT)
            ),
            sc.COL_TXN_PAYEE: _clean_text(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_PAYEE)
            ),
            sc.COL_TXN_PARENT: _clean_text(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_PARENT)
            ),
            sc.COL_TXN_CLIENT: _clean_text(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_CLIENT)
            ),
            sc.COL_TXN_MATTER: _clean_text(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_MATTER)
            ),
            sc.COL_TXN_CATEGORY_CODE: _clean_text(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_CATEGORY_CODE)
            ),
            sc.COL_TXN_CATEGORY_NAME: _clean_text(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_CATEGORY_NAME)
            ),
            sc.COL_TXN_MEMBER: member,
            sc.COL_TXN_AMOUNT: round(amount, 2),
            sc.COL_TXN_TAX_AMOUNT: round(tax_amount, 2),
            sc.COL_TXN_TAX_FLAG: tax_flag,
            sc.COL_TXN_HST_EXEMPT: hst_exempt,
            sc.COL_TXN_GENERAL_OFFICE_EXPENSE: self._to_bool_int(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_GENERAL_OFFICE_EXPENSE),
                default=0,
            ),
            sc.COL_TXN_SHADOW: self._to_bool_int(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_SHADOW),
                default=0,
            ),
            sc.COL_TXN_INVOICE_REF: _clean_text(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_INVOICE_REF)
            ),
            sc.COL_TXN_BILL_CLAIM_PCT: round(bill_claim_pct, 2),
            sc.COL_TXN_TOTAL_CLAIM_AMOUNT: round(float(total_claim_amount), 2),
            sc.COL_TXN_EXPENSE_DETAILS: _clean_text(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_EXPENSE_DETAILS)
            ),
            sc.COL_TXN_NOTES: _clean_text(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_NOTES)
            ),
            sc.COL_TXN_STATUS: status,
            sc.COL_TXN_CURRENCY: currency,
            sc.COL_TXN_VOID_REASON: _clean_text(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_VOID_REASON)
            ),
            sc.COL_TXN_CLEARED_AT: _clean_text(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_CLEARED_AT)
            ),
            sc.COL_TXN_RECONCILED_AT: _clean_text(
                self._value_with_alias(TBL_TRANSACTIONS_MASTER.table, row, sc.COL_TXN_RECONCILED_AT)
            ),
            sc.COL_TXN_CREATED_AT: created_at,
            sc.COL_TXN_UPDATED_AT: updated_at,
        }

    def _canonicalize_transaction_account_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        account_code = _clean_text(
            self._value_with_alias(TBL_TRANSACTION_ACCOUNTS.table, row, sc.COL_TXN_ACCOUNT_CODE)
        )
        account_name = _clean_text(
            self._value_with_alias(TBL_TRANSACTION_ACCOUNTS.table, row, sc.COL_TXN_ACCOUNT_NAME)
        )
        if not account_code:
            account_code = _to_code_token(account_name, fallback="ACCOUNT")
        if not account_name:
            account_name = account_code

        return {
            sc.COL_TXN_ACCOUNT_CODE: account_code,
            sc.COL_TXN_ACCOUNT_NAME: account_name,
            sc.COL_TXN_ACCOUNT_KIND: _clean_text(
                self._value_with_alias(TBL_TRANSACTION_ACCOUNTS.table, row, sc.COL_TXN_ACCOUNT_KIND)
            ).lower(),
            sc.COL_TXN_ACCOUNT_OWNER: _clean_text(
                self._value_with_alias(TBL_TRANSACTION_ACCOUNTS.table, row, sc.COL_TXN_ACCOUNT_OWNER)
            ),
            sc.COL_TXN_ACCOUNT_ACTIVE: self._to_bool_int(
                self._value_with_alias(TBL_TRANSACTION_ACCOUNTS.table, row, sc.COL_TXN_ACCOUNT_ACTIVE),
                default=1,
            ),
            sc.COL_TXN_ACCOUNT_ALIASES: _clean_text(
                self._value_with_alias(TBL_TRANSACTION_ACCOUNTS.table, row, sc.COL_TXN_ACCOUNT_ALIASES)
            ),
        }

    def _canonicalize_transaction_category_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        code = _clean_text(
            self._value_with_alias(TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_CODE)
        )
        name = _clean_text(
            self._value_with_alias(TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_NAME)
        )
        if not code:
            code = _to_code_token(name, fallback="CATEGORY")
        if not name:
            name = code

        sort_order = self._parse_int(
            self._value_with_alias(TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_SORT_ORDER)
        )
        if sort_order is None:
            sort_order = 0

        return {
            sc.COL_TXN_CATEGORY_LKP_CODE: code,
            sc.COL_TXN_CATEGORY_LKP_NAME: name,
            sc.COL_TXN_CATEGORY_LKP_TYPE: _clean_text(
                self._value_with_alias(TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_TYPE)
            ),
            sc.COL_TXN_CATEGORY_LKP_CLASS_SCOPE: _clean_text(
                self._value_with_alias(TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_CLASS_SCOPE)
            ),
            sc.COL_TXN_CATEGORY_LKP_TAX_FLAG_DEFAULT: _clean_text(
                self._value_with_alias(
                    TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_TAX_FLAG_DEFAULT
                )
            ),
            sc.COL_TXN_CATEGORY_LKP_BILLABLE_ALLOWED: self._to_bool_int(
                self._value_with_alias(
                    TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_BILLABLE_ALLOWED
                ),
                default=0,
            ),
            sc.COL_TXN_CATEGORY_LKP_MEDICAL_ELIGIBLE: self._to_bool_int(
                self._value_with_alias(
                    TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_MEDICAL_ELIGIBLE
                ),
                default=0,
            ),
            sc.COL_TXN_CATEGORY_LKP_DEDUCTIBLE_ELIGIBLE: self._to_bool_int(
                self._value_with_alias(
                    TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_DEDUCTIBLE_ELIGIBLE
                ),
                default=0,
            ),
            sc.COL_TXN_CATEGORY_LKP_BUSINESS_DEDUCTIBLE_ELIGIBLE: self._to_bool_int(
                self._value_with_alias(
                    TBL_TRANSACTION_CATEGORIES.table,
                    row,
                    sc.COL_TXN_CATEGORY_LKP_BUSINESS_DEDUCTIBLE_ELIGIBLE,
                ),
                default=0,
            ),
            sc.COL_TXN_CATEGORY_LKP_ACTIVE: self._to_bool_int(
                self._value_with_alias(TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_ACTIVE),
                default=1,
            ),
            sc.COL_TXN_CATEGORY_LKP_SORT_ORDER: int(sort_order),
            sc.COL_TXN_CATEGORY_LKP_NOTES: _clean_text(
                self._value_with_alias(TBL_TRANSACTION_CATEGORIES.table, row, sc.COL_TXN_CATEGORY_LKP_NOTES)
            ),
        }

    def _canonicalize_transaction_business_unit_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        unit_name = _clean_text(
            self._value_with_alias(TBL_TRANSACTION_BUSINESS_UNITS.table, row, sc.COL_TXN_BUSINESS_UNIT_NAME)
        )
        if not unit_name:
            unit_name = "Unassigned"
        return {
            sc.COL_TXN_BUSINESS_UNIT_NAME: unit_name,
            sc.COL_TXN_BUSINESS_UNIT_OWNER: _clean_text(
                self._value_with_alias(TBL_TRANSACTION_BUSINESS_UNITS.table, row, sc.COL_TXN_BUSINESS_UNIT_OWNER)
            ),
            sc.COL_TXN_BUSINESS_UNIT_ACTIVE: self._to_bool_int(
                self._value_with_alias(TBL_TRANSACTION_BUSINESS_UNITS.table, row, sc.COL_TXN_BUSINESS_UNIT_ACTIVE),
                default=1,
            ),
        }

    def _canonicalize_transaction_payee_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        payee_name = _clean_text(self._value_with_alias(TBL_TRANSACTION_PAYEES.table, row, sc.COL_TXN_PAYEE_NAME))
        if not payee_name:
            payee_name = "Unknown Payee"
        return {
            sc.COL_TXN_PAYEE_NAME: payee_name,
            sc.COL_TXN_PAYEE_DEFAULT_CATEGORY_CODE: _clean_text(
                self._value_with_alias(
                    TBL_TRANSACTION_PAYEES.table, row, sc.COL_TXN_PAYEE_DEFAULT_CATEGORY_CODE
                )
            ),
            sc.COL_TXN_PAYEE_ACTIVE: self._to_bool_int(
                self._value_with_alias(TBL_TRANSACTION_PAYEES.table, row, sc.COL_TXN_PAYEE_ACTIVE),
                default=1,
            ),
        }

    def _txn_scope_match(self, source_scope: str, expected: str) -> bool:
        expected_norm = _normalize_search_text(expected)
        source_norm = _normalize_search_text(source_scope)
        if not expected_norm:
            return True
        if not source_norm:
            return True
        if source_norm in ("all", "any", "both", "*"):
            return True
        tokens = [tok.strip() for tok in re.split(r"[|,;/]", source_norm) if tok.strip()]
        if not tokens:
            tokens = [source_norm]
        return expected_norm in tokens

    def _value_with_alias(self, table_name: str, row: Dict[str, Any], canonical_key: str) -> Any:
        if canonical_key in row and row[canonical_key] not in (None, ""):
            return row[canonical_key]
        alias_list = TABLE_ALIASES.get(table_name, {}).get(canonical_key, [])
        for alias in alias_list:
            if alias in row and row[alias] not in (None, ""):
                return row[alias]
        return None

    def _ensure_sheet_schema(self, wb, tref: TableRef) -> Tuple[bool, str]:
        created_sheet = False
        if tref.sheet not in wb.sheetnames:
            wb.create_sheet(tref.sheet)
            created_sheet = True
        ws = wb[tref.sheet]
        header_row, header_map = self._locate_sheet_header_map(ws, tref, create_missing=True)
        headers_changed = self._ensure_required_headers(ws, header_row, header_map, TABLE_COLUMNS[tref.table])
        table_changed = self._ensure_excel_table_object(ws, tref, header_row, header_map, TABLE_COLUMNS[tref.table])
        if created_sheet:
            return True, "created_sheet"
        if table_changed:
            return True, "table_updated"
        if headers_changed:
            return True, "header_updated"
        return False, "ok"

    def _locate_sheet_header_map(
        self,
        ws: Worksheet,
        tref: TableRef,
        create_missing: bool,
    ) -> Tuple[int, Dict[str, int]]:
        expected_headers = TABLE_COLUMNS[tref.table]
        if not expected_headers:
            return 0, {}
        anchor_header = expected_headers[0]

        header_row: int = 0
        header_map: Dict[str, int] = {}

        # Fast path: when table metadata exists, header row is the table's top row.
        table_obj = None
        try:
            if hasattr(ws, "tables") and ws.tables and tref.table in ws.tables:
                table_obj = ws.tables[tref.table]
        except Exception:
            table_obj = None
        if table_obj is not None:
            try:
                min_col, min_row, max_col, _ = _safe_range_boundaries(table_obj.ref)
                min_col = int(min_col or 0)
                min_row = int(min_row or 0)
                max_col = int(max_col or 0)
                if min_row <= 0 or min_col <= 0 or max_col < min_col:
                    raise ValueError("invalid table header range")
                row_map: Dict[str, int] = {}
                for col_idx in range(min_col, max_col + 1):
                    label = _clean_text(ws.cell(row=min_row, column=col_idx).value)
                    if not label or label in row_map:
                        continue
                    row_map[label] = col_idx
                if anchor_header in row_map or create_missing:
                    header_row = min_row
                    header_map = row_map
            except Exception:
                header_row = 0
                header_map = {}

        if header_row <= 0:
            max_scan_row = max(20, min(200, int(ws.max_row or 200) if ws.max_row else 200))
            # Keep fallback scan bounded so we never walk sparse/styled full-sheet column extents.
            max_scan_col = min(256, max(64, len(expected_headers) + 24))
            for row in ws.iter_rows(min_row=1, max_row=max_scan_row, min_col=1, max_col=max_scan_col):
                if not row:
                    continue
                row_idx = int(row[0].row or 0)
                if row_idx <= 0:
                    continue
                row_map = {}
                for offset, cell in enumerate(row, start=1):
                    if cell.value is None:
                        continue
                    label = _clean_text(cell.value)
                    if not label or label in row_map:
                        continue
                    raw_col_idx = getattr(cell, "col_idx", 0) or getattr(cell, "column", 0) or offset
                    try:
                        col_idx = int(raw_col_idx)
                    except Exception:
                        col_idx = offset
                    if col_idx <= 0:
                        continue
                    row_map[label] = col_idx
                if anchor_header in row_map:
                    header_row = row_idx
                    header_map = row_map
                    break

        if header_row <= 0:
            if not create_missing:
                return 0, {}
            header_row = 1
            header_map = {}

        if create_missing:
            self._ensure_required_headers(ws, header_row, header_map, expected_headers)
        return header_row, header_map

    def _ensure_required_headers(
        self,
        ws: Worksheet,
        header_row: int,
        header_map: Dict[str, int],
        required_headers: List[str],
    ) -> bool:
        changed = False
        next_col = max(header_map.values(), default=0) + 1
        for header in required_headers:
            if header in header_map:
                continue
            _set_cell_value(ws, header_row, next_col, header)
            header_map[header] = next_col
            next_col += 1
            changed = True
        return changed

    def _ensure_excel_table_object(
        self,
        ws: Worksheet,
        tref: TableRef,
        header_row: int,
        header_map: Dict[str, int],
        required_headers: List[str],
    ) -> bool:
        _lazy_load_heavy_libs()
        if header_row <= 0 or not required_headers:
            return False
        col_indices = [int(header_map[h]) for h in required_headers if h in header_map and int(header_map[h]) > 0]
        if not col_indices:
            return False
        min_col = min(col_indices)
        max_col = max(col_indices)
        last_data_row = self._last_data_row_for_columns(ws, header_row, col_indices)
        if last_data_row <= header_row:
            last_data_row = header_row + 1
        ref = (
            f"{get_column_letter(min_col)}{header_row}:"
            f"{get_column_letter(max_col)}{last_data_row}"
        )

        existing = None
        try:
            if hasattr(ws, "tables") and ws.tables and tref.table in ws.tables:
                existing = ws.tables[tref.table]
        except Exception:
            existing = None

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        if existing is not None:
            changed = False
            if existing.ref != ref:
                existing.ref = ref
                changed = True
            if existing.tableStyleInfo is None:
                existing.tableStyleInfo = style
                changed = True
            return changed

        table = Table(displayName=tref.table, ref=ref)
        table.tableStyleInfo = style
        ws.add_table(table)
        return True

    def _last_data_row_for_columns(
        self,
        ws: Worksheet,
        header_row: int,
        column_indices: List[int],
        tref: Optional[TableRef] = None,
    ) -> int:
        cols = [int(c) for c in column_indices if int(c) > 0]
        if not cols:
            return int(header_row)
        max_row = int(ws.max_row or 0)
        if tref is not None:
            try:
                if hasattr(ws, "tables") and ws.tables and tref.table in ws.tables:
                    _, _, _, table_max_row = _safe_range_boundaries(ws.tables[tref.table].ref)
                    if table_max_row > 0:
                        max_row = table_max_row if max_row <= 0 else min(max_row, table_max_row)
            except Exception:
                pass
        if max_row <= header_row:
            return int(header_row)
        for row_idx in range(max_row, header_row, -1):
            for col_idx in cols:
                if ws.cell(row=row_idx, column=col_idx).value not in (None, ""):
                    return row_idx
        return int(header_row)

    @with_db_lock
    def _write_table_rows(self, tref: Union[TableRef, str], rows: List[Dict[str, Any]]) -> None:
        """Overwrite an entire table with the given dictionary rows."""
        if isinstance(tref, str):
            tref = next((t for t in TABLES_IN_ORDER if t.table == tref), tref)
        self.ensure_schema()
        _lazy_load_heavy_libs()
        wb = load_workbook(self.paths.workbook_path(), keep_vba=True)
        try:
            ws = wb[tref.sheet]
            existing_table = None
            headers = TABLE_COLUMNS[tref.table]
            if hasattr(ws, "tables") and tref.table in ws.tables:
                existing_table = ws.tables[tref.table]
                
                # Note: We purposely DO NOT override `headers` with `existing_table.tableColumns`.
                # The canonical schema `TABLE_COLUMNS` is the source of truth. If the Excel table
                # is out of sync, `_write_table` will correct its `ref` boundary based on `headers`.
            
            self._write_table(ws, existing_table, tref.table, headers, rows, None)
            
            # Use _safe_save to prevent OneDrive/SharePoint corruption
            self._safe_save(wb, self.paths.workbook_path())
        finally:
            self._close_workbook(wb)

    def _read_table_rows_read_only(
        self,
        table_refs: List[TableRef],
        workbook_signature: str,
    ) -> Optional[Dict[str, List[Dict[str, Any]]]]:
        """Read table rows from one streaming workbook snapshot.

        ``openpyxl`` cannot expose Excel ``Table`` objects in read-only mode,
        but CSPM already records each table's headers and bounds in its
        signature-validated metadata cache.  Reading through that information
        avoids parsing the complete VBA workbook object graph for read-only
        screen transitions.  If a workbook is unfamiliar or malformed, the
        caller falls back to the existing full-fidelity reader.
        """
        if not table_refs:
            return {}

        _lazy_load_heavy_libs()
        path = self.paths.workbook_path()
        workbook = None
        try:
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=False,
                keep_vba=False,
            )
            results: Dict[str, List[Dict[str, Any]]] = {}

            for tref in table_refs:
                if tref.sheet not in workbook.sheetnames:
                    return None

                worksheet = workbook[tref.sheet]
                expected_headers = TABLE_COLUMNS[tref.table]
                cached_meta = self._get_cached_table_meta(tref, workbook_signature) or {}
                try:
                    header_row = int(cached_meta.get("headerRow", 0) or 0)
                except (TypeError, ValueError):
                    header_row = 0

                # A missing cache is expected after an upgrade or for a newly
                # added table.  Locate a likely header in the first rows rather
                # than giving up the fast path for every other table.
                if header_row <= 0:
                    scan_limit = min(max(int(worksheet.max_row or 1), 1), 40)
                    for row_index, values in enumerate(
                        worksheet.iter_rows(
                            min_row=1,
                            max_row=scan_limit,
                            values_only=True,
                        ),
                        start=1,
                    ):
                        labels = {_clean_text(value) for value in values if _clean_text(value)}
                        matches = 0
                        for expected in expected_headers:
                            aliases = TABLE_ALIASES.get(tref.table, {}).get(expected, [])
                            if expected in labels or any(alias in labels for alias in aliases):
                                matches += 1
                        if matches >= min(2, len(expected_headers)):
                            header_row = row_index
                            break
                if header_row <= 0:
                    return None

                try:
                    last_data_row = int(cached_meta.get("lastDataRow", 0) or 0)
                except (TypeError, ValueError):
                    last_data_row = 0
                if last_data_row <= header_row:
                    last_data_row = int(worksheet.max_row or header_row)
                if last_data_row < header_row:
                    last_data_row = header_row

                values_iter = worksheet.iter_rows(
                    min_row=header_row,
                    max_row=last_data_row,
                    values_only=True,
                )
                source_headers = next(values_iter, None)
                if source_headers is None:
                    return None
                source_index = {
                    _clean_text(header): index
                    for index, header in enumerate(source_headers)
                    if _clean_text(header)
                }
                column_for_header: Dict[str, int] = {}
                for expected in expected_headers:
                    candidates = [expected, *TABLE_ALIASES.get(tref.table, {}).get(expected, [])]
                    index = next((source_index[item] for item in candidates if item in source_index), None)
                    if index is not None:
                        column_for_header[expected] = int(index)

                # A table with no recognizable columns needs the legacy reader
                # to diagnose or repair it rather than silently returning an
                # empty result.
                if not column_for_header:
                    return None

                rows: List[Dict[str, Any]] = []
                for source_row in values_iter:
                    row_data: Dict[str, Any] = {}
                    has_data = False
                    for header in expected_headers:
                        index = column_for_header.get(header)
                        value = source_row[index] if index is not None and index < len(source_row) else None
                        row_data[header] = value
                        if value not in (None, ""):
                            has_data = True
                    if has_data:
                        rows.append(row_data)

                self._row_cache[self._table_cache_key(tref)] = rows
                self._set_cached_table_meta(
                    tref,
                    workbook_signature,
                    {
                        "headerRow": int(header_row),
                        "lastDataRow": int(last_data_row),
                        "columnForHeader": {
                            header: int(index + 1)
                            for header, index in column_for_header.items()
                        },
                    },
                )
                results[tref.table] = rows

            return results
        except Exception:
            logger.debug("Read-only table snapshot unavailable; using full workbook reader.", exc_info=True)
            return None
        finally:
            self._close_workbook(workbook)

    @with_db_lock
    def _read_table_rows(self, tref: Union[TableRef, str]) -> List[Dict[str, Any]]:
        if isinstance(tref, str):
            tref = next((t for t in TABLES_IN_ORDER if t.table == tref), tref)
        if self._import_batch_active:
            return [dict(row) for row in self._batch_table_rows(tref)]

        _lazy_load_heavy_libs()
        import logging
        import time
        logger = logging.getLogger("ExcelRepo")
        t_start = time.perf_counter()
        
        path = self.paths.workbook_path()
        if not path.exists():
            logger.warning(f"[ExcelRepo] Workbook missing at {path}; creating schema workbook.")
            self.ensure_schema()

        cache_key = self._table_cache_key(tref)
        workbook_signature = ""
        try:
            stat_info = os.stat(str(path))
            current_mtime = float(stat_info.st_mtime)
            workbook_signature = f"{int(stat_info.st_size)}:{int(stat_info.st_mtime_ns)}"
            if self._row_cache_mtime == current_mtime:
                if cache_key in self._row_cache:
                    return self._row_cache[cache_key]
            else:
                self._row_cache.clear()
                self._row_cache_mtime = current_mtime
        except Exception:
            current_mtime = 0.0

        fast_rows = self._read_table_rows_read_only([tref], workbook_signature)
        if fast_rows is not None and tref.table in fast_rows:
            self._persist_table_meta_cache()
            logger.debug(
                "[ExcelRepo] Streaming read recovered %d rows from %s in %.3fs",
                len(fast_rows[tref.table]),
                tref.sheet,
                time.perf_counter() - t_start,
            )
            return fast_rows[tref.table]

        cached_meta = self._get_cached_table_meta(tref, workbook_signature)
        wb = load_workbook(path, keep_vba=True)
        try:
            if tref.sheet not in wb.sheetnames:
                available = list(wb.sheetnames)
                self._close_workbook(wb)
                logger.warning(
                    f"[ExcelRepo] Expected sheet '{tref.sheet}' missing. Available: {available}; repairing schema."
                )
                self.ensure_schema()
                wb = load_workbook(path, keep_vba=True)
                if tref.sheet not in wb.sheetnames:
                    logger.error(
                        f"[ExcelRepo] Expected sheet '{tref.sheet}' still missing after repair. Available: {wb.sheetnames}"
                    )
                    return []

            ws = wb[tref.sheet]
            expected_headers = TABLE_COLUMNS[tref.table]

            table_obj = None
            try:
                if hasattr(ws, "tables") and ws.tables and tref.table in ws.tables:
                    table_obj = ws.tables[tref.table]
            except Exception:
                table_obj = None

            if table_obj is not None:
                min_col, min_row, max_col, max_row = _safe_range_boundaries(table_obj.ref)
                _headers, raw_rows = self._rows_from_table(ws, table_obj)
                rows: List[Dict[str, Any]] = []
                header_to_col: Dict[str, int] = {}
                for idx, header in enumerate(_headers):
                    if not header:
                        continue
                    header_to_col[header] = min_col + idx
                column_for_header: Dict[str, int] = {}
                for raw_row in raw_rows:
                    row_data: Dict[str, Any] = {}
                    has_data = False
                    for header in expected_headers:
                        if header not in column_for_header:
                            if header in header_to_col:
                                column_for_header[header] = int(header_to_col[header])
                            else:
                                alias_list = TABLE_ALIASES.get(tref.table, {}).get(header, [])
                                for alias in alias_list:
                                    if alias in header_to_col:
                                        column_for_header[header] = int(header_to_col[alias])
                                        break
                        value = self._value_with_alias(tref.table, raw_row, header)
                        row_data[header] = value
                        if value not in (None, ""):
                            has_data = True
                    if has_data:
                        rows.append(row_data)
                logger.debug(
                    f"[ExcelRepo] Table-range read successful. Recovered {len(rows)} rows from {tref.sheet} in {time.perf_counter() - t_start:.3f}s"
                )
                self._row_cache[cache_key] = rows
                self._set_cached_table_meta(
                    tref,
                    workbook_signature,
                    {
                        "headerRow": int(min_row),
                        "lastDataRow": int(max_row),
                        "tableRef": str(table_obj.ref),
                        "columnForHeader": {k: int(v) for k, v in column_for_header.items()},
                    },
                )
                self._persist_table_meta_cache()
                return rows

            header_row_idx = 0
            column_for_header: Dict[str, int] = {}
            last_data_row = 0
            if cached_meta:
                cached_header_row = int(cached_meta.get("headerRow", 0) or 0)
                cached_last_row = int(cached_meta.get("lastDataRow", 0) or 0)
                cached_columns = cached_meta.get("columnForHeader")
                if isinstance(cached_columns, dict) and cached_header_row > 0:
                    for canonical in expected_headers:
                        raw_idx = cached_columns.get(canonical, 0)
                        try:
                            idx = int(raw_idx)
                        except Exception:
                            idx = 0
                        if idx > 0:
                            column_for_header[canonical] = idx
                    if column_for_header:
                        anchor_col = int(column_for_header.get(expected_headers[0], 0) or 0)
                        anchor_label = _clean_text(ws.cell(row=cached_header_row, column=anchor_col).value) if anchor_col > 0 else ""
                        if anchor_label:
                            header_row_idx = cached_header_row
                            if cached_last_row > header_row_idx:
                                max_row = int(ws.max_row or cached_last_row or 0)
                                if max_row > 0:
                                    last_data_row = min(max_row, cached_last_row)

            if header_row_idx <= 0:
                header_row_idx, header_map = self._locate_sheet_header_map(ws, tref, create_missing=False)
                if header_row_idx <= 0:
                    logger.error(
                        f"[ExcelRepo] Range scan failed: Could not locate headers for table '{tref.table}' on sheet '{tref.sheet}'."
                    )
                    return []

                column_for_header = {}
                for canonical in expected_headers:
                    if canonical in header_map:
                        column_for_header[canonical] = header_map[canonical]
                        continue
                    alias_list = TABLE_ALIASES.get(tref.table, {}).get(canonical, [])
                    for alias in alias_list:
                        if alias in header_map:
                            column_for_header[canonical] = header_map[alias]
                            break

            if last_data_row <= header_row_idx:
                last_data_row = self._last_data_row_for_columns(
                    ws,
                    header_row_idx,
                    [column_for_header.get(h, 0) for h in expected_headers],
                    tref=tref,
                )

            rows: List[Dict[str, Any]] = []
            for row_idx in range(header_row_idx + 1, last_data_row + 1):
                row_data: Dict[str, Any] = {}
                has_data = False
                for header in expected_headers:
                    col_idx = column_for_header.get(header, 0)
                    value = ws.cell(row=row_idx, column=col_idx).value if col_idx > 0 else None
                    row_data[header] = value
                    if value not in (None, ""):
                        has_data = True
                if has_data:
                    rows.append(row_data)

            logger.debug(
                f"[ExcelRepo] Range read successful. Recovered {len(rows)} rows from {tref.sheet} in {time.perf_counter() - t_start:.3f}s"
            )
            self._row_cache[cache_key] = rows
            self._set_cached_table_meta(
                tref,
                workbook_signature,
                {
                    "headerRow": int(header_row_idx),
                    "lastDataRow": int(last_data_row),
                    "columnForHeader": {
                        k: int(v) for k, v in column_for_header.items() if int(v) > 0
                    },
                },
            )
            self._persist_table_meta_cache()
            return rows
        except Exception as e:
            logger.error(f"[ExcelRepo] Crash while reading {tref.table}: {e}")
            return []
        finally:
            self._close_workbook(wb)

    @with_db_lock
    def _read_table_rows_bulk(
        self,
        table_refs: List[Union[TableRef, str]],
    ) -> Dict[str, List[Dict[str, Any]]]:
        """Read a related group of tables from a single workbook opening.

        Cold ``_read_table_rows`` calls each parse the full macro workbook.
        Screens that need a small related group of tables (such as WIP-to-Bill)
        should use this read-only helper instead.  The normal row cache remains
        authoritative, and no workbook content or formatting is modified.
        """

        normalized: List[TableRef] = []
        for raw_ref in table_refs:
            tref = raw_ref
            if isinstance(tref, str):
                tref = next((item for item in TABLES_IN_ORDER if item.table == tref), tref)
            if not isinstance(tref, TableRef):
                raise ValueError(f"Unknown workbook table reference: {raw_ref}")
            if tref not in normalized:
                normalized.append(tref)

        if not normalized:
            return {}
        if self._import_batch_active:
            return {
                tref.table: [dict(row) for row in self._batch_table_rows(tref)]
                for tref in normalized
            }

        _lazy_load_heavy_libs()
        import time

        path = self.paths.workbook_path()
        if not path.exists():
            logger.warning("[ExcelRepo] Workbook missing at %s; creating schema workbook.", path)
            self.ensure_schema()

        workbook_signature = ""
        try:
            stat_info = os.stat(str(path))
            current_mtime = float(stat_info.st_mtime)
            workbook_signature = f"{int(stat_info.st_size)}:{int(stat_info.st_mtime_ns)}"
            if self._row_cache_mtime != current_mtime:
                self._row_cache.clear()
                self._row_cache_mtime = current_mtime
        except Exception:
            current_mtime = 0.0

        results: Dict[str, List[Dict[str, Any]]] = {}
        pending: List[TableRef] = []
        for tref in normalized:
            cached = self._row_cache.get(self._table_cache_key(tref))
            if cached is not None and self._row_cache_mtime == current_mtime:
                results[tref.table] = cached
            else:
                pending.append(tref)
        if not pending:
            return results

        started = time.perf_counter()
        fast_rows = self._read_table_rows_read_only(pending, workbook_signature)
        if fast_rows is not None:
            results.update(fast_rows)
            self._persist_table_meta_cache()
            logger.info(
                "[ExcelRepo] Streaming bulk table read: %d table(s), %d cold table(s), %.3fs",
                len(normalized),
                len(pending),
                time.perf_counter() - started,
            )
            return results

        fallback: List[TableRef] = []
        wb = load_workbook(path, keep_vba=True)
        try:
            for tref in pending:
                if tref.sheet not in wb.sheetnames:
                    fallback.append(tref)
                    continue

                ws = wb[tref.sheet]
                try:
                    table_obj = ws.tables[tref.table] if hasattr(ws, "tables") and tref.table in ws.tables else None
                except Exception:
                    table_obj = None
                if table_obj is None:
                    fallback.append(tref)
                    continue

                min_col, min_row, _max_col, max_row = _safe_range_boundaries(table_obj.ref)
                headers, raw_rows = self._rows_from_table(ws, table_obj)
                header_to_col = {
                    header: min_col + index
                    for index, header in enumerate(headers)
                    if header
                }
                expected_headers = TABLE_COLUMNS[tref.table]
                column_for_header: Dict[str, int] = {}
                rows: List[Dict[str, Any]] = []
                for raw_row in raw_rows:
                    row_data: Dict[str, Any] = {}
                    has_data = False
                    for header in expected_headers:
                        if header not in column_for_header:
                            if header in header_to_col:
                                column_for_header[header] = int(header_to_col[header])
                            else:
                                for alias in TABLE_ALIASES.get(tref.table, {}).get(header, []):
                                    if alias in header_to_col:
                                        column_for_header[header] = int(header_to_col[alias])
                                        break
                        value = self._value_with_alias(tref.table, raw_row, header)
                        row_data[header] = value
                        if value not in (None, ""):
                            has_data = True
                    if has_data:
                        rows.append(row_data)

                self._row_cache[self._table_cache_key(tref)] = rows
                self._set_cached_table_meta(
                    tref,
                    workbook_signature,
                    {
                        "headerRow": int(min_row),
                        "lastDataRow": int(max_row),
                        "tableRef": str(table_obj.ref),
                        "columnForHeader": {
                            key: int(value)
                            for key, value in column_for_header.items()
                        },
                    },
                )
                results[tref.table] = rows
        finally:
            self._close_workbook(wb)

        self._persist_table_meta_cache()
        for tref in fallback:
            results[tref.table] = self._read_table_rows(tref)

        logger.info(
            "[ExcelRepo] Bulk table read: %d table(s), %d cold table(s), %.3fs",
            len(normalized),
            len(pending),
            time.perf_counter() - started,
        )
        return results

    @with_db_lock
    def _write_table_rows_bulk(
        self,
        table_rows: Dict[Union[TableRef, str], List[Dict[str, Any]]],
    ) -> None:
        """Replace several workbook tables in one durable save.

        Financial commands often alter several related tables at the same time.
        Saving each table separately reparses and rewrites the full macro
        workbook for every table, which is both slow and exposes short-lived
        partially-updated states.  This helper applies all supplied table
        snapshots to one workbook object and performs the existing atomic
        save protocol once.
        """
        if not table_rows:
            return

        normalized: List[Tuple[TableRef, List[Dict[str, Any]]]] = []
        seen_tables = set()
        for raw_ref, rows in table_rows.items():
            tref = raw_ref
            if isinstance(tref, str):
                tref = next((item for item in TABLES_IN_ORDER if item.table == tref), tref)
            if not isinstance(tref, TableRef):
                raise ValueError(f"Unknown workbook table reference: {raw_ref}")
            if tref.table in seen_tables:
                raise ValueError(f"Table {tref.table} was supplied more than once.")
            seen_tables.add(tref.table)
            normalized.append((tref, list(rows or [])))

        # Preserve the canonical workbook table order.  It makes a financial
        # save deterministic and keeps related source sheets stable for users
        # who inspect the workbook directly in Excel.
        table_order = {item.table: index for index, item in enumerate(TABLES_IN_ORDER)}
        normalized.sort(key=lambda item: table_order.get(item[0].table, len(table_order)))

        self.ensure_schema()
        _lazy_load_heavy_libs()
        wb = load_workbook(self.paths.workbook_path(), keep_vba=True)
        try:
            for tref, rows in normalized:
                if tref.sheet not in wb.sheetnames:
                    raise ValueError(f"Worksheet for table {tref.table} was not found: {tref.sheet}")
                ws = wb[tref.sheet]
                existing_table = None
                if hasattr(ws, "tables") and tref.table in ws.tables:
                    existing_table = ws.tables[tref.table]
                self._write_table(
                    ws,
                    existing_table,
                    tref.table,
                    TABLE_COLUMNS[tref.table],
                    rows,
                    None,
                )

            # _safe_save atomically replaces the workbook and invalidates all
            # row/meta caches only after every table is in the new snapshot.
            self._safe_save(wb, self.paths.workbook_path())
        finally:
            self._close_workbook(wb)

        self._persist_table_meta_cache()

    def _append_row_to_table(self, tref: TableRef, row: Dict[str, Any]) -> None:
        if self._import_batch_active:
            self._batch_append_row(tref, row)
            return

        _lazy_load_heavy_libs()
        self.ensure_schema()
        wb = load_workbook(self.paths.workbook_path(), keep_vba=True)
        try:
            ws = wb[tref.sheet]
            header_row, header_map = self._locate_sheet_header_map(ws, tref, create_missing=True)
            headers = TABLE_COLUMNS[tref.table]
            self._ensure_required_headers(ws, header_row, header_map, headers)
            insert_row = self._last_data_row_for_columns(
                ws,
                header_row,
                [header_map[h] for h in headers if h in header_map],
                tref=tref,
            ) + 1
            for header in headers:
                col_idx = header_map.get(header)
                if col_idx:
                    _set_cell_value(ws, insert_row, col_idx, row.get(header, ""))
            self._safe_save(wb, self.paths.workbook_path())
        finally:
            self._close_workbook(wb)

    def _delete_row_by_key_hard(
        self,
        tref: TableRef,
        key_column: str,
        key_value: Any,
    ) -> bool:
        if self._import_batch_active:
            raise ValueError("Hard deletes are not supported during batch import.")

        _lazy_load_heavy_libs()
        self.ensure_schema()
        wb = load_workbook(self.paths.workbook_path(), keep_vba=True)
        try:
            ws = wb[tref.sheet]
            header_row, header_map = self._locate_sheet_header_map(ws, tref, create_missing=True)
            if key_column not in header_map:
                return False

            key_col_idx = header_map[key_column]
            target_key = _clean_text(key_value).lower()
            target_row_idx = -1
            
            headers = TABLE_COLUMNS[tref.table]
            last_data_row = self._last_data_row_for_columns(
                ws,
                header_row,
                [header_map[h] for h in headers if h in header_map],
                tref=tref,
            )

            for row_idx in range(header_row + 1, last_data_row + 1):
                current_key = _clean_text(ws.cell(row=row_idx, column=key_col_idx).value).lower()
                if current_key and current_key == target_key:
                    target_row_idx = row_idx
                    break

            if target_row_idx > 0:
                ws.delete_rows(target_row_idx, 1)
                
                if hasattr(ws, "tables") and tref.table in ws.tables:
                    table_obj = ws.tables[tref.table]
                    min_col, min_row, max_col, max_row = _safe_range_boundaries(table_obj.ref)
                    if target_row_idx <= max_row:
                        max_row -= 1
                        if max_row < min_row + 1:
                            max_row = min_row + 1
                        from openpyxl.utils import get_column_letter
                        table_obj.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                
                self._safe_save(wb, self.paths.workbook_path())
                return True
            return False
        finally:
            wb.close()

    def _find_row_by_key(self, tref: TableRef, key_column: str, key_value: Any) -> Optional[Dict[str, Any]]:
        rows = self._read_table_rows(tref)
        target = _clean_text(key_value).lower()
        for r in rows:
            if _clean_text(r.get(key_column)).lower() == target:
                return dict(r)
        return None

    def _upsert_row_by_key(
        self,
        tref: TableRef,
        key_column: str,
        key_value: Any,
        row: Dict[str, Any],
    ) -> None:
        if self._import_batch_active:
            self._batch_upsert_row(tref, key_column, key_value, row)
            return

        _lazy_load_heavy_libs()
        self.ensure_schema()
        wb = load_workbook(self.paths.workbook_path(), keep_vba=True)
        try:
            ws = wb[tref.sheet]
            header_row, header_map = self._locate_sheet_header_map(ws, tref, create_missing=True)
            headers = TABLE_COLUMNS[tref.table]
            self._ensure_required_headers(ws, header_row, header_map, headers)
            if key_column not in header_map:
                raise ValueError(f"Key column not found in table {tref.table}: {key_column}")

            key_col_idx = header_map[key_column]
            target_key = _clean_text(key_value).lower()
            target_row_idx = -1
            last_data_row = self._last_data_row_for_columns(
                ws,
                header_row,
                [header_map[h] for h in headers if h in header_map],
                tref=tref,
            )

            for row_idx in range(header_row + 1, last_data_row + 1):
                current_key = _clean_text(ws.cell(row=row_idx, column=key_col_idx).value).lower()
                if current_key and current_key == target_key:
                    target_row_idx = row_idx
                    break

            if target_row_idx <= 0:
                target_row_idx = last_data_row + 1

            for header in headers:
                col_idx = header_map.get(header)
                if col_idx:
                    _set_cell_value(ws, target_row_idx, col_idx, row.get(header, ""))
            
            # Dynamically push table bounds down to include the new row
            if hasattr(ws, "tables") and tref.table in ws.tables:
                table_obj = ws.tables[tref.table]
                min_col, min_row, max_col, max_row = _safe_range_boundaries(table_obj.ref)
                if target_row_idx > max_row:
                    from openpyxl.utils.cell import get_column_letter
                    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{target_row_idx}"
                    table_obj.ref = new_ref

            self._safe_save(wb, self.paths.workbook_path())
        finally:
            self._close_workbook(wb)

    def _replace_table_rows(self, tref: TableRef, rows: List[Dict[str, Any]]) -> None:
        if self._import_batch_active:
            self._batch_replace_rows(tref, rows)
            return

        _lazy_load_heavy_libs()
        self.ensure_schema()
        wb = load_workbook(self.paths.workbook_path(), keep_vba=True)
        try:
            ws = wb[tref.sheet]
            headers = TABLE_COLUMNS[tref.table]
            header_row, header_map = self._locate_sheet_header_map(ws, tref, create_missing=True)
            self._ensure_required_headers(ws, header_row, header_map, headers)
            last_data_row = self._last_data_row_for_columns(
                ws,
                header_row,
                [header_map[h] for h in headers if h in header_map],
                tref=tref,
            )
            for row_idx in range(header_row + 1, last_data_row + 1):
                for header in headers:
                    col_idx = header_map.get(header)
                    if col_idx:
                        _set_cell_value(ws, row_idx, col_idx, None)

            canonical_rows = [self._canonicalize_row(tref, dict(row or {})) for row in list(rows or [])]
            write_row_idx = header_row + 1
            for row_data in canonical_rows:
                for header in headers:
                    col_idx = header_map.get(header)
                    if col_idx:
                        _set_cell_value(ws, write_row_idx, col_idx, row_data.get(header, ""))
                write_row_idx += 1
            
            if hasattr(ws, "tables") and tref.table in ws.tables:
                table_obj = ws.tables[tref.table]
                min_col, min_row, max_col, max_row = _safe_range_boundaries(table_obj.ref)
                if write_row_idx - 1 > max_row:
                    from openpyxl.utils.cell import get_column_letter
                    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{write_row_idx - 1}"
                    table_obj.ref = new_ref

            self._safe_save(wb, self.paths.workbook_path())
        finally:
            self._close_workbook(wb)

    def _append_note_line(self, existing: Any, line: Any) -> str:
        existing_text = _clean_text(existing)
        next_line = _clean_text(line)
        if not next_line:
            return existing_text
        if not existing_text:
            return next_line

        lines = [chunk.strip() for chunk in re.split(r"\r?\n", existing_text) if _clean_text(chunk)]
        line_lc = next_line.lower()
        for chunk in lines:
            if _clean_text(chunk).lower() == line_lc:
                return "\n".join(lines)

        lines.append(next_line)
        if len(lines) > 80:
            lines = lines[-80:]
        return "\n".join(lines)

    def _row_is_blank(self, ws: Worksheet, min_col: int, max_col: int, row_idx: int) -> bool:
        for col_idx in range(min_col, max_col + 1):
            if ws.cell(row=row_idx, column=col_idx).value not in (None, ""):
                return False
        return True

    def _format_full_address(
        self,
        line1: str,
        line2: str,
        city: str,
        state_province: str,
        postal_code: str,
        country: str,
    ) -> str:
        lines: List[str] = []
        if _clean_text(line1):
            lines.append(_clean_text(line1))
        if _clean_text(line2):
            lines.append(_clean_text(line2))

        local_parts = [p for p in [_clean_text(city), _clean_text(state_province)] if p]
        locality = ", ".join(local_parts)
        if _clean_text(postal_code):
            locality = (locality + " " + _clean_text(postal_code)).strip() if locality else _clean_text(postal_code)
        if locality:
            lines.append(locality)
        if _clean_text(country):
            lines.append(_clean_text(country))
        return "\n".join(lines)

    def _is_client_row_active(self, row: Dict[str, Any]) -> bool:
        status = _clean_text(row.get(sc.COL_CLIENT_STATUS)).lower()
        active_flag = self._to_bool_int(row.get(sc.COL_CLIENT_ACTIVE), default=1)
        return active_flag == 1 and status not in ("inactive", "closed", "archived")

    def _is_matter_row_active(self, row: Dict[str, Any]) -> bool:
        status = _clean_text(row.get(sc.COL_MATTER_STATUS)).lower()
        return status not in ("inactive", "closed", "archived")

    def _matter_number_in_use(self, matter_number: str, ignore_matter_id: str = "") -> bool:
        target = _clean_text(matter_number).upper()
        if not target:
            return False
        ignore_lc = _clean_text(ignore_matter_id).lower()
        rows = [self._canonicalize_matter_row(r) for r in self._read_table_rows(TBL_MATTERS)]
        for row in rows:
            row_id = _clean_text(row.get(sc.COL_MATTER_ID)).lower()
            if ignore_lc and row_id == ignore_lc:
                continue
            row_number = _clean_text(row.get(sc.COL_MATTER_NUMBER)).upper()
            if row_number and row_number == target:
                return True
        return False

    def _all_client_codes_in_use(self) -> set:
        """
        Collect every unique 4-char client code already embedded in any
        existing matter number.  Used to detect/avoid collisions when
        generating a new client code.
        """
        codes: set = set()
        rows = self._read_table_rows(TBL_MATTERS)
        for row in rows:
            num = _clean_text(row.get(sc.COL_MATTER_NUMBER, "")).upper()
            # Accept any reasonable legacy format – extract the leading token
            m = re.match(r"^([A-Z0-9]{2,6})-", num)
            if m:
                codes.add(m.group(1)[:4])
        return codes

    def _legacy_client_code(
        self,
        client_name: str,
        entity_type: str = "",
        existing_matter_id: str = "",
    ) -> str:
        """
        Return a collision-free 4-char client code for *client_name*.

        If the matter already has a number whose leading 4 chars match the
        raw code derived from *client_name*, that existing code is reused
        (idempotent on edit).  Otherwise, we scan all existing matter numbers
        and increment an integer suffix until we find a free slot.
        """
        raw = _legacy_raw_client_code(client_name, entity_type)

        # If this matter already exists and its number starts with the raw
        # code, preserve it (don't bump the counter on edits).
        if _clean_text(existing_matter_id):
            existing_row = self._find_matter_row(existing_matter_id)
            if existing_row:
                existing_num = _clean_text(existing_row.get(sc.COL_MATTER_NUMBER, "")).upper()
                if existing_num.startswith(raw + "-"):
                    return raw

        in_use = self._all_client_codes_in_use()
        # Also include UUID-style client IDs so we never clash with those
        client_rows = self._read_table_rows(TBL_CLIENTS)
        for r in client_rows:
            cid = _clean_text(r.get(sc.COL_CLIENT_ID, "")).upper()
            if cid:
                in_use.add(cid[:4])

        candidate = raw
        counter = 1
        while candidate in in_use:
            suffix = str(counter)
            candidate = raw[: 4 - len(suffix)] + suffix
            counter += 1

        return candidate

    def _next_matter_sequence_global(
        self,
        year_two_digits: str,
        existing_matter_id: str = "",
    ) -> int:
        """
        Return the next globally-unique sequence number for *year_two_digits*.

        We scan every matter number in the database looking for anything that
        ends with ``-{year}-{4-digit-seq}`` and return  max_found + 1.
        """
        ignore_lc = _clean_text(existing_matter_id).lower()
        # Pattern: anything ending in -26-0017  (2-digit year then 4-digit seq)
        pattern = re.compile(r"-" + re.escape(year_two_digits) + r"-(\d{4})$")
        max_seq = 0
        rows = self._read_table_rows(TBL_MATTERS)
        for row in rows:
            row_id = _clean_text(row.get(sc.COL_MATTER_ID, "")).lower()
            if ignore_lc and row_id == ignore_lc:
                continue
            num = _clean_text(row.get(sc.COL_MATTER_NUMBER, "")).upper()
            m = pattern.search(num)
            if m:
                seq = int(m.group(1))
                if seq > max_seq:
                    max_seq = seq
        return max_seq + 1

    def delete_matter_profile(self, matter_id: str) -> Dict[str, Any]:
        """Permanently delete only a matter with no remaining dependencies.

        The old editor exposed this destructive action directly.  A fresh
        server-side check keeps it safe even if the confirmation dialog is
        bypassed or a future UI calls this method directly.
        """
        dependencies = self.check_matter_dependencies(matter_id)
        if not dependencies.get("ok"):
            return {
                "ok": False,
                "message": dependencies.get("message", "Could not verify linked matter records."),
                "dependencies": dependencies,
            }
        if not dependencies.get("canDelete"):
            return {
                "ok": False,
                "message": self._matter_delete_blocker_message(dependencies),
                "dependencies": dependencies,
            }
        deleted = ExcelRepo._delete_matter_with_parties(self, matter_id)
        return {
            "ok": deleted,
            "message": "Matter permanently deleted." if deleted else "Matter not found.",
            "dependencies": dependencies,
        }

    def delete_archived_matter_profile(self, matter_id: str) -> Dict[str, Any]:
        """Permanently delete an archived matter only after a fresh dependency check."""
        matter = self._find_matter_row(matter_id)
        if matter is None:
            return {"ok": False, "message": "Matter not found."}

        status = _clean_text(matter.get(sc.COL_MATTER_STATUS)).lower()
        if status != "archived":
            return {
                "ok": False,
                "message": "Only archived matters can be permanently deleted from the directory.",
            }

        dependencies = self.check_matter_dependencies(matter_id)
        if not dependencies.get("ok"):
            return {
                "ok": False,
                "message": dependencies.get("message", "Could not verify linked records."),
            }
        if not dependencies.get("canDelete"):
            return {
                "ok": False,
                "message": self._matter_delete_blocker_message(dependencies),
                "dependencies": dependencies,
            }

        deleted = ExcelRepo._delete_matter_with_parties(self, matter_id)
        return {
            "ok": deleted,
            "message": "Archived matter permanently deleted." if deleted else "Matter not found.",
            "dependencies": dependencies,
        }

    def _delete_matter_with_parties(self, matter_id: str) -> bool:
        """Remove an otherwise-safe matter and its internal party rows together."""
        target = _clean_text(matter_id).casefold()
        try:
            matter_rows = [self._canonicalize_matter_row(row) for row in self._read_table_rows(TBL_MATTERS)]
            if not any(_clean_text(row.get(sc.COL_MATTER_ID)).casefold() == target for row in matter_rows):
                return False
            party_rows = [
                self._canonicalize_matter_party_row(row)
                for row in self._read_table_rows(TBL_MATTER_PARTIES)
            ]
        except Exception:
            return self._delete_row_by_key_hard(TBL_MATTERS, sc.COL_MATTER_ID, matter_id)

        related_party_rows = [
            row
            for row in party_rows
            if _clean_text(row.get(sc.COL_MATTER_PARTY_MATTER_ID)).casefold() == target
        ]
        if not related_party_rows:
            return self._delete_row_by_key_hard(TBL_MATTERS, sc.COL_MATTER_ID, matter_id)

        remaining_matters = [
            row for row in matter_rows if _clean_text(row.get(sc.COL_MATTER_ID)).casefold() != target
        ]
        remaining_parties = [
            row
            for row in party_rows
            if _clean_text(row.get(sc.COL_MATTER_PARTY_MATTER_ID)).casefold() != target
        ]
        self._write_table_rows_bulk({
            TBL_MATTERS: remaining_matters,
            TBL_MATTER_PARTIES: remaining_parties,
        })
        return True

    def check_matter_dependencies(self, matter_id: str) -> Dict[str, Any]:
        """
        Checks whether a matter can be permanently deleted and returns the
        financial work that blocks archive/close/on-hold.  It reads the related
        tables together so a single user action does not repeatedly open the
        macro workbook.
        """
        self.ensure_schema()
        matter_id_lower = _clean_text(matter_id).lower()
        if not matter_id_lower:
            return {"ok": False, "message": "Invalid matter ID", "canDelete": False}

        tables = self._read_table_rows_bulk([
            TBL_MATTERS,
            TBL_TIME,
            TBL_DISBURSEMENTS,
            TBL_RECEIVABLES,
            TBL_TRANSACTIONS_MASTER,
            TBL_TRADEMARKS,
        ])
        matter_rows = [self._canonicalize_matter_row(row) for row in tables.get(TBL_MATTERS.table, [])]
        matter = next(
            (
                row for row in matter_rows
                if _clean_text(row.get(sc.COL_MATTER_ID)).casefold() == matter_id_lower
            ),
            None,
        )
        if matter is None:
            return {"ok": False, "message": "Matter not found.", "canDelete": False}

        matter_aliases = {
            matter_id_lower,
            _clean_text(matter.get(sc.COL_MATTER_NAME)).casefold(),
            _clean_text(matter.get(sc.COL_MATTER_NUMBER)).casefold(),
            _clean_text(matter.get(sc.COL_MATTER_DISPLAY_NAME)).casefold(),
        }
        matter_aliases.discard("")

        time_rows = tables.get(TBL_TIME.table, [])
        disbursement_rows = tables.get(TBL_DISBURSEMENTS.table, [])
        financial = self._matter_financial_summary_from_rows(
            matter_id,
            time_rows,
            disbursement_rows,
            tables.get(TBL_RECEIVABLES.table, []),
        )
        time_count = sum(
            1
            for row in time_rows
            if _clean_text(row.get(sc.COL_TIME_MATTER_ID)).casefold() == matter_id_lower
        )
        disb_count = sum(
            1
            for row in disbursement_rows
            if _clean_text(row.get(sc.COL_DISB_MATTER_ID)).casefold() == matter_id_lower
        )
        txn_count = sum(
            1
            for row in tables.get(TBL_TRANSACTIONS_MASTER.table, [])
            if _clean_text(row.get(sc.COL_TXN_MATTER)).casefold() in matter_aliases
        )
        tm_count = sum(
            1
            for row in tables.get(TBL_TRADEMARKS.table, [])
            if _clean_text(row.get(sc.COL_TM_MATTER_NUMBER)).casefold() in matter_aliases
        )
        total_dependencies = time_count + disb_count + txn_count + tm_count

        return {
            "ok": True,
            "canDelete": total_dependencies == 0,
            "timeCount": time_count,
            "disbursementCount": disb_count,
            "transactionCount": txn_count,
            "invoiceCount": int(financial.get("unpaidInvoiceCount", 0) or 0),
            "trademarkCount": tm_count,
            "totalDependencies": total_dependencies,
            "financialSummary": financial,
            "hasFinancialBlockers": bool(financial.get("hasFinancialBlockers")),
        }

    def get_matter_financial_summary(self, matter_id: str) -> Dict[str, Any]:
        """Return the live WIP and unpaid-A/R picture for one matter.

        Receivables do not store MatterID themselves.  The authoritative link
        is the docket/disbursement's invoice reference, so this derives the
        invoice list from those underlying financial records rather than
        accidentally treating the billing client as the work matter.
        """
        self.ensure_schema()
        normalized_id = _clean_text(matter_id)
        if not normalized_id:
            return {"ok": False, "message": "Invalid matter ID."}

        tables = self._read_table_rows_bulk([
            TBL_MATTERS,
            TBL_TIME,
            TBL_DISBURSEMENTS,
            TBL_RECEIVABLES,
        ])
        matter_exists = any(
            _clean_text(row.get(sc.COL_MATTER_ID)).casefold() == normalized_id.casefold()
            for row in tables.get(TBL_MATTERS.table, [])
        )
        if not matter_exists:
            return {"ok": False, "message": "Matter not found.", "matterId": normalized_id}

        return self._matter_financial_summary_from_rows(
            normalized_id,
            tables.get(TBL_TIME.table, []),
            tables.get(TBL_DISBURSEMENTS.table, []),
            tables.get(TBL_RECEIVABLES.table, []),
        )

    def _matter_financial_summary_from_rows(
        self,
        matter_id: str,
        time_rows: List[Dict[str, Any]],
        disbursement_rows: List[Dict[str, Any]],
        receivable_rows: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """Compute a matter-level financial summary from already loaded rows."""
        matter_id_lower = _clean_text(matter_id).casefold()
        billed_statuses = {
            "billed", "posted", "invoiced", "finalized", "locked", "merged",
            "void", "voided", "reversed", "cancelled", "canceled", "superseded",
        }
        closed_receivable_statuses = {
            "paid", "void", "voided", "reversed", "cancelled", "canceled", "superseded",
        }

        def amount(value: Any) -> float:
            return self._money_round(self._parse_float(value) or 0.0)

        def docket_is_unbilled(row: Dict[str, Any]) -> bool:
            status = _clean_text(row.get(sc.COL_TIME_STATUS)).casefold()
            invoice_status = _clean_text(row.get(sc.COL_TIME_INVOICE_STATUS)).casefold()
            invoice_ref = _clean_text(row.get(sc.COL_TIME_INVOICE_REF)).casefold()
            if status in billed_statuses or invoice_status in billed_statuses:
                return False
            # A real invoice reference means that the docket has left WIP.
            # Generic legacy markers ("Draft"/"WIP") still represent active,
            # unbilled work and therefore keep the matter operational.
            return invoice_ref in {"", "draft", "wip", "unbilled"}

        unbilled_dockets: List[Dict[str, Any]] = []
        unbilled_disbursements: List[Dict[str, Any]] = []
        linked_invoices: set[str] = set()

        for row in time_rows:
            if _clean_text(row.get(sc.COL_TIME_MATTER_ID)).casefold() != matter_id_lower:
                continue
            invoice_ref = _clean_text(row.get(sc.COL_TIME_INVOICE_REF))
            if docket_is_unbilled(row):
                unbilled_dockets.append({
                    "entryId": _clean_text(row.get(sc.COL_TIME_ENTRY_ID)),
                    "date": _clean_text(row.get(sc.COL_TIME_DATE)),
                    "description": _clean_text(row.get(sc.COL_TIME_DESC)),
                    "status": _clean_text(row.get(sc.COL_TIME_STATUS)) or "Draft",
                    "amount": amount(row.get(sc.COL_TIME_TOTAL)) or amount(row.get(sc.COL_TIME_GROSS)),
                })
            elif invoice_ref:
                linked_invoices.add(invoice_ref.casefold())

        for row in disbursement_rows:
            if _clean_text(row.get(sc.COL_DISB_MATTER_ID)).casefold() != matter_id_lower:
                continue
            invoice_ref = _clean_text(row.get(sc.COL_DISB_INVOICE_REF))
            if invoice_ref.casefold() in {"", "draft", "wip", "unbilled"}:
                unbilled_disbursements.append({
                    "entryId": _clean_text(row.get(sc.COL_DISB_ID)),
                    "date": _clean_text(row.get(sc.COL_DISB_DATE)),
                    "description": _clean_text(row.get(sc.COL_DISB_DESCRIPTION)) or "Disbursement",
                    "status": "Unbilled",
                    "amount": amount(row.get(sc.COL_DISB_AMOUNT)),
                })
            else:
                linked_invoices.add(invoice_ref.casefold())

        unpaid_invoices: List[Dict[str, Any]] = []
        for row in receivable_rows:
            invoice_num = _clean_text(row.get(sc.COL_RECV_INVOICE_NUM))
            if not invoice_num or invoice_num.casefold() not in linked_invoices:
                continue
            balance = amount(row.get(sc.COL_RECV_BALANCE_DUE))
            status = _clean_text(row.get(sc.COL_RECV_STATUS))
            if balance <= 0.004 or status.casefold() in closed_receivable_statuses:
                continue
            unpaid_invoices.append({
                "invoiceNum": invoice_num,
                "date": _clean_text(row.get(sc.COL_RECV_DATE)),
                "status": status or "Open",
                "client": _clean_text(row.get(sc.COL_RECV_CLIENT)),
                "workClient": _clean_text(row.get(sc.COL_RECV_WORK_CLIENT)),
                "invoiceTotal": amount(row.get(sc.COL_RECV_TOTAL_INVOICED)),
                "amountDue": balance,
            })

        unbilled_dockets.sort(key=lambda item: (item["date"], item["entryId"]))
        unbilled_disbursements.sort(key=lambda item: (item["date"], item["entryId"]))
        unpaid_invoices.sort(key=lambda item: (item["date"], item["invoiceNum"]))
        docket_total = self._money_round(sum(item["amount"] for item in unbilled_dockets))
        disbursement_total = self._money_round(sum(item["amount"] for item in unbilled_disbursements))
        wip_total = self._money_round(docket_total + disbursement_total)
        unpaid_total = self._money_round(sum(item["amountDue"] for item in unpaid_invoices))
        unbilled_count = len(unbilled_dockets) + len(unbilled_disbursements)

        return {
            "ok": True,
            "matterId": _clean_text(matter_id),
            "unbilledDocketCount": len(unbilled_dockets),
            "unbilledDocketAmount": docket_total,
            "unbilledDisbursementCount": len(unbilled_disbursements),
            "unbilledDisbursementAmount": disbursement_total,
            "unbilledWipCount": unbilled_count,
            "unbilledWipAmount": wip_total,
            "unbilledDockets": unbilled_dockets,
            "unbilledDisbursements": unbilled_disbursements,
            "unpaidInvoiceCount": len(unpaid_invoices),
            "unpaidInvoiceAmount": unpaid_total,
            "unpaidInvoices": unpaid_invoices,
            "hasUnbilledWip": unbilled_count > 0,
            "hasUnpaidInvoices": len(unpaid_invoices) > 0,
            "hasFinancialBlockers": unbilled_count > 0 or len(unpaid_invoices) > 0,
        }

    def _matter_financial_blocker_message(self, summary: Dict[str, Any], *, action: str) -> str:
        parts: List[str] = []
        wip_count = int(summary.get("unbilledWipCount", 0) or 0)
        if wip_count:
            parts.append(
                f"{wip_count} unbilled WIP item(s) (${float(summary.get('unbilledWipAmount', 0.0) or 0.0):,.2f})"
            )
        invoice_count = int(summary.get("unpaidInvoiceCount", 0) or 0)
        if invoice_count:
            parts.append(
                f"{invoice_count} unpaid invoice(s) (${float(summary.get('unpaidInvoiceAmount', 0.0) or 0.0):,.2f})"
            )
        detail = " and ".join(parts) if parts else "open financial records"
        return (
            f"Cannot {action} because this matter has {detail}. "
            "Bill, reassign, or settle the work first."
        )

    def _matter_delete_blocker_message(self, dependencies: Dict[str, Any]) -> str:
        financial = dict(dependencies.get("financialSummary") or {})
        if financial.get("hasFinancialBlockers"):
            return self._matter_financial_blocker_message(financial, action="permanently delete this matter")
        return (
            "This matter still has linked records and cannot be permanently deleted. "
            "Archive it instead, or remove/reassign the linked records first."
        )

    def _build_matter_number(
        self,
        client_name: str,
        matter_type: str,
        date_opened: str,
        existing_matter_id: str = "",
        parent_name: str = "",
        entity_type: str = "",
    ) -> str:
        """
        Build a Legacy-format matter number.

        Format (no parent):   CLNT-TYP-YY-NNNN
        Format (with parent): CLNT-PRNT-TYP-YY-NNNN

        where:
          CLNT  = 4-char client code
          PRNT  = 4-char parent-entity code (only when parent is set)
          TYP   = 3-char matter-type code  (TMK / COM / CRP / EST / TAX / AUD / GEN)
          YY    = 2-digit year
          NNNN  = 4-digit globally-incrementing sequence within the year
        """
        year_code   = _matter_year_two_digits(date_opened)
        client_code = self._legacy_client_code(client_name, entity_type, existing_matter_id)
        type_code   = _legacy_matter_type_code(matter_type)

        # Parent code – only inject when a parent name is present
        parent_code = ""
        if _clean_text(parent_name):
            parent_code = _legacy_raw_client_code(parent_name, entity_type="")

        # Idempotency: if the existing matter already carries a valid number,
        # keep it as long as the key segments still match.
        if _clean_text(existing_matter_id):
            existing_row = self._find_matter_row(existing_matter_id)
            if existing_row:
                existing_num = _clean_text(existing_row.get(sc.COL_MATTER_NUMBER, "")).upper()
                if parent_code:
                    expected_prefix = f"{client_code}-{parent_code}-{type_code}-{year_code}-"
                else:
                    expected_prefix = f"{client_code}-{type_code}-{year_code}-"
                if existing_num.startswith(expected_prefix):
                    return existing_num

        next_seq = self._next_matter_sequence_global(
            year_two_digits=year_code,
            existing_matter_id=existing_matter_id,
        )

        if parent_code:
            return f"{client_code}-{parent_code}-{type_code}-{year_code}-{next_seq:04d}"
        return f"{client_code}-{type_code}-{year_code}-{next_seq:04d}"

    def _normalize_client_profile_payload(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        display_name = self._pick_text(payload, ["displayName", sc.COL_PROFILE_DISPLAY_NAME, "DisplayName"])
        legal_name = self._pick_text(payload, ["legalName", sc.COL_PROFILE_LEGAL_NAME, "LegalName"])
        first_name = self._pick_text(
            payload,
            ["firstName", "givenName", sc.COL_PROFILE_FIRST_NAME, "First Name", "Given Name"],
        )
        middle_name = self._pick_text(
            payload,
            ["middleName", "middleInitial", sc.COL_PROFILE_MIDDLE_NAME, "Middle Name", "Middle Initial"],
        )
        last_name = self._pick_text(
            payload,
            ["lastName", "surname", "familyName", sc.COL_PROFILE_LAST_NAME, "Last Name", "Family Name"],
        )

        client_name = self._pick_text(payload, ["clientName", "clientText", sc.COL_CLIENT_NAME, "Client"])
        if not client_name:
            client_name = display_name or legal_name or _join_person_name_parts(first_name, middle_name, last_name)
        if not client_name:
            raise ValueError("Client name is required.")

        if not legal_name:
            legal_name = display_name or client_name

        status = self._pick_text(payload, ["status", "clientStatus", sc.COL_CLIENT_STATUS])
        if not status:
            status = "Active"
        active_raw = self._pick_value(payload, ["active", sc.COL_CLIENT_ACTIVE])
        if active_raw is None:
            active = 0 if _clean_text(status).lower() in ("inactive", "closed", "archived") else 1
        else:
            active = self._to_bool_int(active_raw, default=1)

        retainer_required = self._to_bool_int(
            self._pick_value(payload, ["retainerRequired", sc.COL_PROFILE_RETAINER_REQUIRED]),
            default=0,
        )
        retainer_amount = self._pick_float(payload, ["retainerAmount", sc.COL_PROFILE_RETAINER_AMOUNT])
        if retainer_amount is None:
            retainer_amount = 0.0
        if retainer_amount < 0:
            raise ValueError("Retainer amount must be >= 0.")

        address_line1 = self._pick_text(payload, ["addressLine1", sc.COL_PROFILE_ADDR1, "Address1"])
        address_line2 = self._pick_text(payload, ["addressLine2", sc.COL_PROFILE_ADDR2, "Address2"])
        city = self._pick_text(payload, ["city", sc.COL_PROFILE_CITY])
        state_province = self._pick_text(payload, ["stateProvince", "state", "province", sc.COL_PROFILE_STATE])
        postal_code = self._pick_text(payload, ["postalCode", "zip", "zipCode", sc.COL_PROFILE_POSTAL])
        country = self._pick_text(payload, ["country", sc.COL_PROFILE_COUNTRY])
        full_address = self._pick_text(payload, ["fullAddress", "formattedAddress", sc.COL_PROFILE_FULL_ADDRESS])
        if not full_address:
            full_address = self._format_full_address(
                line1=address_line1,
                line2=address_line2,
                city=city,
                state_province=state_province,
                postal_code=postal_code,
                country=country,
            )

        entity_type = self._pick_text(payload, ["entityType", "clientType", sc.COL_PROFILE_ENTITY_TYPE])
        if not entity_type:
            entity_type = "Corporation"
        if _is_individual_entity_type(entity_type) and not (first_name and last_name):
            split_first, split_middle, split_last = _split_person_name_parts(
                display_name or legal_name or client_name
            )
            if not first_name:
                first_name = split_first
            if not middle_name:
                middle_name = split_middle
            if not last_name:
                last_name = split_last

        onboarding_status = self._pick_text(
            payload,
            ["onboardingStatus", "onboardingStage", sc.COL_PROFILE_ONBOARDING_STATUS],
        )
        if not onboarding_status:
            onboarding_status = "Prospect"

        kyc_status = self._pick_text(payload, ["kycStatus", "KYC", sc.COL_PROFILE_KYC_STATUS])
        if not kyc_status:
            kyc_status = "Pending"

        date_client_added = self._pick_text(
            payload,
            ["dateClientAdded", "clientSinceDate", "dateAdded", sc.COL_PROFILE_DATE_CLIENT_ADDED],
        )
        if not date_client_added:
            date_client_added = datetime.now().strftime("%Y-%m-%d")

        birthday = self._pick_text(
            payload,
            ["birthday", "dob", "birthDate", "dateOfBirth", sc.COL_PROFILE_BIRTHDAY],
        )

        return {
            "clientId": self._pick_text(payload, ["clientId", sc.COL_CLIENT_ID, sc.COL_PROFILE_CLIENT_ID]),
            "forceDuplicate": self._to_bool_int(
                self._pick_value(payload, ["forceDuplicate", "addDuplicate"]),
                default=0,
            ) == 1,
            "clientName": client_name,
            "displayName": display_name,
            "legalName": legal_name,
            "firstName": first_name,
            "middleName": middle_name,
            "lastName": last_name,
            "entityType": entity_type,
            "principalName": self._pick_text(payload, ["principalName", sc.COL_PROFILE_PRINCIPAL_NAME]),
            "principalPosition": self._pick_text(payload, ["principalPosition", "principalTitle", sc.COL_PROFILE_PRINCIPAL_POSITION]),
            "primaryEmail": self._pick_text(payload, ["primaryEmail", "email", sc.COL_PROFILE_PRIMARY_EMAIL, sc.COL_CLIENT_EMAIL]),
            "primaryPhone": _normalize_us_phone(
                self._pick_text(payload, ["primaryPhone", "phone", sc.COL_PROFILE_PRIMARY_PHONE, sc.COL_CLIENT_PHONE])
            ),
            "secondaryContactName": self._pick_text(payload, ["secondaryContactName", "secondaryName", sc.COL_PROFILE_SECONDARY_CONTACT]),
            "secondaryContactPosition": self._pick_text(payload, ["secondaryContactPosition", "secondaryPosition", sc.COL_PROFILE_SECONDARY_POSITION]),
            "secondaryContactEmail": self._pick_text(payload, ["secondaryContactEmail", "secondaryEmail", sc.COL_PROFILE_SECONDARY_EMAIL]),
            "secondaryContactPhone": _normalize_us_phone(
                self._pick_text(payload, ["secondaryContactPhone", "secondaryPhone", sc.COL_PROFILE_SECONDARY_PHONE])
            ),
            "addressLine1": address_line1,
            "addressLine2": address_line2,
            "city": city,
            "stateProvince": state_province,
            "postalCode": postal_code,
            "country": country,
            "fullAddress": full_address,
            "parentClientId": self._pick_text(payload, ["parentClientId", "parentId", sc.COL_PROFILE_PARENT_ID]),
            "parentClientName": self._pick_text(payload, ["parentClientName", "parentClient", "parentName", sc.COL_PROFILE_PARENT_NAME]),
            "website": self._pick_text(payload, ["website", sc.COL_PROFILE_WEBSITE]),
            "taxId": self._pick_text(payload, ["taxId", "taxID", sc.COL_PROFILE_TAX_ID]),
            "industry": self._pick_text(payload, ["industry", sc.COL_PROFILE_INDUSTRY]),
            "billingEmail": self._pick_text(payload, ["billingEmail", sc.COL_PROFILE_BILLING_EMAIL]),
            "kycStatus": kyc_status,
            "onboardingStatus": onboarding_status,
            "retainerRequired": retainer_required,
            "retainerAmount": float(retainer_amount),
            "engagementStartDate": self._pick_text(payload, ["engagementStartDate", "startDate", sc.COL_PROFILE_ENGAGEMENT_START]),
            "dateClientAdded": date_client_added,
            "birthday": birthday,
            "referralFrom": self._pick_text(payload, ["referralFrom", "referredBy", sc.COL_PROFILE_REFERRAL_FROM]),
            "conflictNotes": self._pick_text(payload, ["conflictNotes", "conflictCheckNotes", sc.COL_PROFILE_CONFLICT_NOTES]),
            "notes": self._pick_text(payload, ["notes", "notesText", sc.COL_PROFILE_NOTES, sc.COL_CLIENT_NOTES]),
            "status": status,
            "active": active,
        }

    def _normalize_matter_profile_payload(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        matter_name = self._pick_text(
            payload,
            ["matterName", "matterText", "matterTitle", sc.COL_MATTER_NAME, "Matter"],
        )
        if not matter_name:
            raise ValueError("Matter name is required.")

        client_name = self._pick_text(
            payload,
            ["clientName", "clientText", sc.COL_MATTER_CLIENT_NAME, sc.COL_CLIENT_NAME, "Client"],
        )
        client_id = self._pick_text(payload, ["clientId", sc.COL_MATTER_CLIENT_ID])
        if not client_name and not client_id:
            raise ValueError("Client is required.")

        default_rate = self._pick_float(payload, ["defaultRate", "rate", sc.COL_MATTER_DEF_RATE])
        if default_rate is None:
            default_rate = 0.0
        if default_rate < 0:
            raise ValueError("Default rate must be >= 0.")

        default_share_pct = normalize_pct(
            self._pick_value(payload, ["defaultSharePct", "sharePct", sc.COL_MATTER_DEF_SHARE]),
            default_pct=100.0,
        )

        import json
        rate_history_raw = self._pick_text(payload, ["rateHistory", sc.COL_MATTER_RATE_HISTORY])
        rate_history_json = "[]"
        if rate_history_raw:
            if isinstance(rate_history_raw, str) and rate_history_raw.strip().startswith("["):
                rate_history_json = rate_history_raw.strip()
            elif isinstance(payload.get("rateHistory"), list):
                rate_history_json = json.dumps(payload.get("rateHistory"))

        status = self._pick_text(payload, ["status", "matterStatus", sc.COL_MATTER_STATUS])
        if not status:
            status = "Open"

        billing_email = self._pick_text(payload, ["billingEmail", sc.COL_MATTER_BILLING_EMAIL])
        if billing_email and not _looks_like_email_address(billing_email):
            raise ValueError("Billing Email is not in a valid email format.")

        date_of_engagement = self._pick_text(
            payload,
            ["dateOfEngagement", "engagementDate", sc.COL_MATTER_ENGAGEMENT_DATE],
        )
        if date_of_engagement and not _is_valid_iso_date(date_of_engagement):
            raise ValueError("Date Of Engagement must be in YYYY-MM-DD format.")

        date_opened = self._pick_text(payload, ["dateOpened", "openDate", sc.COL_MATTER_OPEN_DATE])
        if not date_opened:
            date_opened = datetime.now().strftime("%Y-%m-%d")
        if not _is_valid_iso_date(date_opened):
            raise ValueError("Date Opened must be in YYYY-MM-DD format.")

        date_closed = self._pick_text(payload, ["dateClosed", "closeDate", sc.COL_MATTER_CLOSE_DATE])
        if date_closed and not _is_valid_iso_date(date_closed):
            raise ValueError("Date Closed must be in YYYY-MM-DD format.")
        if date_closed and date_opened and date_closed < date_opened:
            raise ValueError("Date Closed cannot be earlier than Date Opened.")

        representation_mode = self._pick_text(
            payload,
            ["representationMode", "retainerMode", sc.COL_MATTER_REPRESENTATION_MODE],
        ) or "Single Client"
        raw_parties = payload.get("parties", payload.get("matterParties", []))
        if isinstance(raw_parties, str):
            try:
                raw_parties = json.loads(raw_parties)
            except (TypeError, ValueError):
                raise ValueError("Matter parties must be a list of client records.")
        if raw_parties is None:
            raw_parties = []
        if not isinstance(raw_parties, list):
            raise ValueError("Matter parties must be a list of client records.")
        normalized_parties = [dict(item or {}) for item in raw_parties if isinstance(item, dict)]
        if len(normalized_parties) != len(raw_parties):
            raise ValueError("Each matter party must identify an existing client.")
        is_joint_retainer = representation_mode.casefold() == "joint retainer" or len(normalized_parties) > 1
        if is_joint_retainer and len(normalized_parties) < 2:
            raise ValueError("A joint retainer requires at least two independent client records.")
        if is_joint_retainer:
            representation_mode = "Joint Retainer"

        return {
            "matterId": self._pick_text(payload, ["matterId", sc.COL_MATTER_ID]),
            "forceDuplicate": self._to_bool_int(
                self._pick_value(payload, ["forceDuplicate", "addDuplicate"]),
                default=0,
            ) == 1,
            "matterNumber": self._pick_text(payload, ["matterNumber", sc.COL_MATTER_NUMBER]),
            "matterName": matter_name,
            "displayName": self._pick_text(payload, ["displayName", sc.COL_MATTER_DISPLAY_NAME]),
            "clientId": client_id,
            "clientName": client_name,
            "parentId": self._pick_text(payload, ["parentId", sc.COL_MATTER_PARENT_ID]),
            "parentName": self._pick_text(payload, ["parentName", sc.COL_MATTER_PARENT_NAME, "Parent"]),
            "matterType": self._pick_text(payload, ["matterType", "type", sc.COL_MATTER_TYPE]) or "General",
            "practiceArea": self._pick_text(payload, ["practiceArea", "practice", sc.COL_MATTER_PRACTICE_AREA]),
            "status": status,
            "responsibleLawyer": self._pick_text(
                payload,
                ["responsibleLawyer", "responsibleAttorney", "owner", sc.COL_MATTER_RESPONSIBLE_LAWYER],
            ),
            "billingArrangement": self._pick_text(
                payload,
                ["billingArrangement", "billingType", sc.COL_MATTER_BILLING_ARRANGEMENT],
            )
            or "Hourly",
            "billingContact": self._pick_text(payload, ["billingContact", sc.COL_MATTER_BILLING_CONTACT]),
            "billingEmail": billing_email,
            "defaultRate": float(default_rate),
            "defaultSharePct": round(float(default_share_pct), 2),
            "rateHistory": rate_history_json,
            "dateOfEngagement": date_of_engagement,
            "dateOpened": date_opened,
            "dateClosed": date_closed,
            "courtFileNumber": self._pick_text(payload, ["courtFileNumber", "fileNumber", sc.COL_MATTER_COURT_FILE_NO]),
            "opposingParty": self._pick_text(payload, ["opposingParty", "opposingCounsel", sc.COL_MATTER_OPPOSING_PARTY]),
            "referralFrom": self._pick_text(payload, ["referralFrom", "referralSource", sc.COL_MATTER_REFERRAL_FROM]),
            "description": self._pick_text(payload, ["description", "summary", sc.COL_MATTER_DESCRIPTION]),
            "notes": self._pick_text(payload, ["notes", "notesText", sc.COL_MATTER_NOTES]),
            "representationMode": representation_mode,
            "jointNoConfidentialityConfirmed": self._to_bool_int(
                self._pick_value(
                    payload,
                    ["jointNoConfidentialityConfirmed", sc.COL_MATTER_JOINT_NO_CONFIDENTIALITY_CONFIRMED],
                ),
                default=0,
            ),
            "jointInstructionsRequireAll": self._to_bool_int(
                self._pick_value(
                    payload,
                    ["jointInstructionsRequireAll", sc.COL_MATTER_JOINT_INSTRUCTIONS_REQUIRE_ALL],
                ),
                default=0,
            ),
            "jointEngagementDocument": self._pick_text(
                payload,
                ["jointEngagementDocument", "engagementDocument", sc.COL_MATTER_JOINT_ENGAGEMENT_DOCUMENT],
            ),
            "parties": normalized_parties,
        }

    def _normalize_time_payload(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        date_text = self._pick_text(payload, ["date", "dateText", sc.COL_TIME_DATE])
        if not date_text:
            raise ValueError("Date is required.")
        if not _is_valid_iso_date(date_text):
            raise ValueError("Date must be in YYYY-MM-DD format.")

        client_name = self._pick_text(payload, ["clientName", "clientText", sc.COL_CLIENT_NAME, "Client"])
        if not client_name:
            raise ValueError("Client is required.")

        allow_client_only = self._to_bool_int(
            self._pick_value(payload, ["allowClientOnlyDraft", "allowClientOnly", "clientOnlyOverride"]),
            default=0,
        ) == 1
        matter_name = self._pick_text(payload, ["matterName", "matterText", sc.COL_MATTER_NAME, "Matter"])
        if not matter_name and not allow_client_only:
            raise ValueError("Matter is required. Create/select a matter or use client-only draft override.")

        description = self._pick_text(payload, ["description", "descriptionText", sc.COL_TIME_DESC])
        if not description:
            task_text = self._pick_text(payload, ["taskText", "taskName", "Task"])
            description = task_text or "Time entry"

        hours = self._pick_float(payload, ["hours", sc.COL_TIME_HOURS, "timeHours"])
        if hours is None:
            hours = self._pick_float(payload, ["timeText", "Time (in hrs) or Units"])
        if hours is None:
            raise ValueError("Hours must be numeric.")
        if hours < 0:
            raise ValueError("Hours must be >= 0.")

        rate = self._pick_float(payload, ["rate", "rateText", "clientRate", "ClientRate"])
        if rate is None:
            raise ValueError("Rate must be numeric.")

        share_source = self._pick_value(payload, ["sharePct", sc.COL_TIME_SHARE_PCT, "billText", "BillPct", "yourSharePct", "CutPct"])
        share_pct = normalize_pct(share_source, default_pct=100.0)

        raw_seconds = self._pick_int(payload, ["rawSeconds", sc.COL_TIME_SECONDS, "elapsedSeconds"])
        if raw_seconds is None:
            raw_seconds = int(round(hours * 3600.0))
        if raw_seconds < 0:
            raise ValueError("RawSeconds must be >= 0.")

        segment_seconds = self._pick_int(payload, ["segmentSeconds", "elapsedSegmentSeconds"])
        if segment_seconds is not None and segment_seconds < 0:
            segment_seconds = 0

        status = self._normalize_time_status(self._pick_text(payload, ["status", sc.COL_TIME_STATUS]))
        invoice_ref = self._pick_text(payload, ["invoiceRef", "invoice", sc.COL_TIME_INVOICE_REF])
        invoice_status = self._pick_text(
            payload,
            ["invoiceStatus", "invoiceState", sc.COL_TIME_INVOICE_STATUS],
        )
        payment_status = self._pick_text(
            payload,
            ["paymentStatus", "paymentState", sc.COL_TIME_PAYMENT_STATUS],
        )
        invoice_total = self._pick_float(payload, ["invoiceTotal", "totalInvoiced", sc.COL_TIME_INVOICE_TOTAL])
        invoice_amount_paid = self._pick_float(
            payload,
            ["invoiceAmountPaid", "amountPaid", sc.COL_TIME_INVOICE_AMOUNT_PAID],
        )
        invoice_balance_due = self._pick_float(
            payload,
            ["invoiceBalanceDue", "balanceDue", sc.COL_TIME_INVOICE_BALANCE_DUE],
        )
        invoice_date = self._pick_text(payload, ["invoiceDate", sc.COL_TIME_INVOICE_DATE])

        raw_seconds_mode = self._pick_text(payload, ["rawSecondsMode", "secondsMode"]).lower()
        if raw_seconds_mode not in ("", "increment", "replace", "auto"):
            raw_seconds_mode = ""
        if not raw_seconds_mode:
            raw_seconds_mode = "auto"

        parent_name = self._pick_text(payload, ["parentName", "parentText", sc.COL_PARENT_NAME, "Parent"])

        return {
            "entryId": self._pick_text(payload, ["entryId", sc.COL_TIME_ENTRY_ID]),
            "forceDuplicate": self._to_bool_int(
                self._pick_value(payload, ["forceDuplicate", "addDuplicate"]),
                default=0,
            ) == 1,
            "date": date_text,
            "clientName": client_name,
            "matterName": matter_name,
            "allowClientOnlyDraft": allow_client_only,
            "parentName": parent_name,
            "description": description,
            "lockAudit": self._pick_text(payload, ["lockAudit", sc.COL_TIME_LOCK_AUDIT, "timerAudit"]),
            "hours": float(hours),
            "clientRate": float(rate),
            "sharePct": float(share_pct),
            "rawSeconds": int(raw_seconds),
            "segmentSeconds": int(segment_seconds) if segment_seconds is not None else None,
            "rawSecondsMode": raw_seconds_mode,
            "status": status,
            "invoiceRef": invoice_ref,
            "invoiceStatus": invoice_status,
            "paymentStatus": payment_status,
            "invoiceTotal": round(float(invoice_total or 0.0), 2),
            "invoiceAmountPaid": round(float(invoice_amount_paid or 0.0), 2),
            "invoiceBalanceDue": round(float(invoice_balance_due or 0.0), 2),
            "invoiceDate": invoice_date,
            "useExactHours": bool(payload.get("useExactHours")),
        }

    def _normalize_time_status(self, raw_status: Any) -> str:
        value = _clean_text(raw_status).lower()
        if value in ("", "wip", "unbilled", "open", "pending"):
            return "Draft"
        if value in ("ready", "ready for billing", "ready_for_billing", "review"):
            return "Ready for Billing"
        if value in ("billed", "posted", "invoiced", "finalized", "locked"):
            return "Billed"
        return "Draft"

    def _normalize_transaction_payload(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        warnings: List[str] = []
        today_iso = datetime.now().strftime("%Y-%m-%d")

        txn_id = self._pick_text(payload, ["transactionId", "txnId", sc.COL_TXN_ID, "TransactionID"])
        txn_date = self._pick_text(payload, ["txnDate", "date", "dateText", sc.COL_TXN_DATE])
        if not txn_date:
            txn_date = today_iso
        if not _is_valid_iso_date(txn_date):
            raise ValueError("TxnDate must be in YYYY-MM-DD format.")

        txn_class = _normalize_choice(
            self._pick_text(payload, ["class", "txnClass", sc.COL_TXN_CLASS]),
            self.TXN_CLASS_OPTIONS,
            "Family",
        )
        if not txn_class:
            raise ValueError(f"Class must be one of: {', '.join(self.TXN_CLASS_OPTIONS)}")

        txn_type = _normalize_choice(
            self._pick_text(payload, ["type", "txnType", sc.COL_TXN_TYPE]),
            self.TXN_TYPE_OPTIONS,
            "Expense",
        )
        if not txn_type:
            raise ValueError(f"Type must be one of: {', '.join(self.TXN_TYPE_OPTIONS)}")

        business_unit = self._pick_text(payload, ["businessUnit", "bu", sc.COL_TXN_BUSINESS_UNIT])
        if txn_class == "Business" and not business_unit:
            raise ValueError("BusinessUnit is required when Class is Business.")

        from_account = self._pick_text(payload, ["fromAccount", "sourceAccount", sc.COL_TXN_FROM_ACCOUNT])
        to_account = self._pick_text(payload, ["toAccount", "destinationAccount", sc.COL_TXN_TO_ACCOUNT])
        if not from_account:
            raise ValueError("FromAccount is required.")

        txn_type_lc = txn_type.lower()
        movement_type = txn_type_lc in self.TXN_MOVEMENT_TYPES
        if movement_type and not to_account:
            raise ValueError("ToAccount is required for Transfer and Debt Repayment.")
        if movement_type and to_account and from_account.lower() == to_account.lower():
            raise ValueError("FromAccount and ToAccount must differ for Transfer and Debt Repayment.")

        payee = self._pick_text(payload, ["payee", "vendor", sc.COL_TXN_PAYEE])
        if txn_type_lc != "transfer" and not payee:
            raise ValueError("Payee is required for non-transfer transactions.")

        member = _normalize_choice(
            self._pick_text(payload, ["member", sc.COL_TXN_MEMBER]),
            self.TXN_MEMBER_OPTIONS,
            "Joint",
        )
        if not member:
            raise ValueError(f"Member must be one of: {', '.join(self.TXN_MEMBER_OPTIONS)}")

        tax_flag = _normalize_choice(
            self._pick_text(payload, ["taxFlag", "taxCategory", sc.COL_TXN_TAX_FLAG]),
            self.TXN_TAX_FLAG_OPTIONS,
            "None",
        )
        if not tax_flag:
            raise ValueError(f"TaxFlag must be one of: {', '.join(self.TXN_TAX_FLAG_OPTIONS)}")

        status = _normalize_choice(
            self._pick_text(payload, ["status", sc.COL_TXN_STATUS]),
            self.TXN_STATUS_OPTIONS,
            "Pending",
        )
        if not status:
            raise ValueError(f"Status must be one of: {', '.join(self.TXN_STATUS_OPTIONS)}")

        currency = _normalize_choice(
            self._pick_text(payload, ["currency", sc.COL_TXN_CURRENCY]),
            self.TXN_CURRENCY_OPTIONS,
            "CAD",
        )
        if not currency:
            raise ValueError(f"Currency must be one of: {', '.join(self.TXN_CURRENCY_OPTIONS)}")

        amount = self._pick_float(payload, ["amount", sc.COL_TXN_AMOUNT])
        if amount is None:
            raise ValueError("Amount is required.")
        amount = float(amount)
        if amount <= 0:
            raise ValueError("Amount must be greater than 0.")

        tax_amount = self._pick_float(payload, ["taxAmount", "tax", "hstAmount", sc.COL_TXN_TAX_AMOUNT])
        if tax_amount is None:
            tax_amount = 0.0
        tax_amount = float(tax_amount)
        if tax_amount < 0:
            raise ValueError("TaxAmount must be >= 0.")

        hst_exempt = self._to_bool_int(self._pick_value(payload, ["hstExempt", sc.COL_TXN_HST_EXEMPT]), default=0)
        if hst_exempt == 1 and abs(tax_amount) > 0.0001:
            tax_amount = 0.0
            warnings.append("HSTExempt=1 forced TaxAmount to 0.00")

        general_office_expense = self._to_bool_int(
            self._pick_value(payload, ["generalOfficeExpense", "officeExpense", sc.COL_TXN_GENERAL_OFFICE_EXPENSE]),
            default=0,
        )
        shadow = self._to_bool_int(self._pick_value(payload, ["shadow", sc.COL_TXN_SHADOW]), default=0)

        parent = self._pick_text(payload, ["parent", "parentName", sc.COL_TXN_PARENT])
        client = self._pick_text(payload, ["client", "clientName", sc.COL_TXN_CLIENT])
        matter = self._pick_text(payload, ["matter", "matterName", sc.COL_TXN_MATTER])

        bill_claim_pct = self._pick_float(payload, ["billClaimPct", "claimPct", sc.COL_TXN_BILL_CLAIM_PCT])
        if bill_claim_pct is None:
            bill_claim_pct = 0.0
        bill_claim_pct = float(bill_claim_pct)
        if bill_claim_pct < 0 or bill_claim_pct > 100:
            raise ValueError("BillClaimPct must be between 0 and 100.")

        if general_office_expense == 1:
            if parent or client or matter:
                raise ValueError("GeneralOfficeExpense cannot include Parent/Client/Matter context.")
            if abs(bill_claim_pct) > 0.0001:
                raise ValueError("GeneralOfficeExpense cannot have BillClaimPct > 0.")
            bill_claim_pct = 0.0

        if txn_type_lc == "expense" and bill_claim_pct > 0:
            if not (parent and client and matter):
                raise ValueError("Billable expense requires Parent, Client, and Matter when BillClaimPct > 0.")

        total_claim_amount = round((amount + tax_amount) * (bill_claim_pct / 100.0), 2)

        category_code = self._pick_text(payload, ["categoryCode", "catCode", sc.COL_TXN_CATEGORY_CODE])
        category_name = self._pick_text(payload, ["categoryName", "category", sc.COL_TXN_CATEGORY_NAME])
        category_by_code, category_by_name = self._transaction_category_lookup_maps()
        if category_code and not category_name:
            found = category_by_code.get(category_code.lower())
            if found:
                category_name = _clean_text(found.get("categoryName"))
        if category_name and not category_code:
            found = category_by_name.get(category_name.lower())
            if found:
                category_code = _clean_text(found.get("categoryCode"))
        if category_code and category_by_code.get(category_code.lower()):
            canonical_name = _clean_text(category_by_code[category_code.lower()].get("categoryName"))
            if canonical_name:
                category_name = canonical_name
        if not category_code:
            raise ValueError("CategoryCode is required.")
        if not category_name:
            raise ValueError("CategoryName is required.")

        if txn_type_lc == "debt repayment":
            account_kind_map = self._transaction_account_kind_lookup()
            if account_kind_map:
                to_kind = _clean_text(account_kind_map.get(to_account.lower(), "")).lower()
                if not to_kind:
                    raise ValueError("Debt Repayment ToAccount must exist in transaction account lookup.")
                if to_kind not in self.TXN_DEBT_DEST_ACCOUNT_KINDS:
                    allowed = ", ".join(sorted(self.TXN_DEBT_DEST_ACCOUNT_KINDS))
                    raise ValueError(f"Debt Repayment ToAccount must be a debt account kind ({allowed}).")

        invoice_ref = self._pick_text(payload, ["invoiceRef", "invoice", sc.COL_TXN_INVOICE_REF])
        expense_details = self._pick_text(payload, ["expenseDetails", "details", sc.COL_TXN_EXPENSE_DETAILS])
        notes = self._pick_text(payload, ["notes", "notesText", sc.COL_TXN_NOTES])

        void_reason = self._pick_text(payload, ["voidReason", "voidNotes", sc.COL_TXN_VOID_REASON])
        cleared_at = self._pick_text(payload, ["clearedAt", "clearedDate", sc.COL_TXN_CLEARED_AT])
        reconciled_at = self._pick_text(payload, ["reconciledAt", "reconciledDate", sc.COL_TXN_RECONCILED_AT])

        if status == "Void" and not void_reason:
            raise ValueError("VoidReason is required when Status is Void.")

        if status in ("Cleared", "Reconciled") and not cleared_at:
            cleared_at = today_iso
        if status == "Reconciled" and not reconciled_at:
            reconciled_at = today_iso

        if cleared_at and not _is_valid_iso_date(cleared_at):
            raise ValueError("ClearedAt must be in YYYY-MM-DD format.")
        if reconciled_at and not _is_valid_iso_date(reconciled_at):
            raise ValueError("ReconciledAt must be in YYYY-MM-DD format.")

        if status == "Reconciled" and (not cleared_at or not reconciled_at):
            raise ValueError("Reconciled status requires ClearedAt and ReconciledAt.")

        return {
            "transactionId": txn_id,
            "txnDate": txn_date,
            "class": txn_class,
            "businessUnit": business_unit,
            "type": txn_type,
            "fromAccount": from_account,
            "toAccount": to_account,
            "payee": payee,
            "parent": parent,
            "client": client,
            "matter": matter,
            "categoryCode": category_code,
            "categoryName": category_name,
            "member": member,
            "amount": amount,
            "taxAmount": tax_amount,
            "taxFlag": tax_flag,
            "hstExempt": hst_exempt,
            "generalOfficeExpense": general_office_expense,
            "shadow": shadow,
            "invoiceRef": invoice_ref,
            "billClaimPct": bill_claim_pct,
            "totalClaimAmount": total_claim_amount,
            "expenseDetails": expense_details,
            "notes": notes,
            "status": status,
            "currency": currency,
            "voidReason": void_reason,
            "clearedAt": cleared_at,
            "reconciledAt": reconciled_at,
            "createdAt": self._pick_text(payload, ["createdAt", sc.COL_TXN_CREATED_AT]),
            "warnings": warnings,
        }

    def _validate_transaction_status_transition(
        self,
        old_status: str,
        new_status: str,
        has_existing: bool,
    ) -> None:
        if not has_existing:
            return

        previous = _clean_text(old_status)
        target = _clean_text(new_status)
        if not previous or not target or previous == target:
            return

        allowed_transitions = {
            "Pending": {"Pending", "Cleared", "Void"},
            "Cleared": {"Cleared", "Reconciled", "Void"},
            "Reconciled": {"Reconciled", "Void"},
            "Void": {"Void"},
        }
        if previous not in allowed_transitions:
            return

        allowed = allowed_transitions[previous]
        if target not in allowed:
            raise ValueError(
                f"Invalid Status transition: {previous} -> {target}. "
                f"Allowed next statuses: {', '.join(sorted(allowed))}."
            )

    def _transaction_account_kind_lookup(self) -> Dict[str, str]:
        out: Dict[str, str] = {}
        rows = self.list_transaction_accounts(include_inactive=True)
        for row in rows:
            kind = _clean_text(row.get("accountKind")).lower()
            code = _clean_text(row.get("accountCode")).lower()
            name = _clean_text(row.get("accountName")).lower()
            aliases = row.get("aliases") if isinstance(row, dict) else []
            if kind:
                if code:
                    out[code] = kind
                if name:
                    out[name] = kind
                for alias in aliases if isinstance(aliases, list) else []:
                    alias = _clean_text(alias).lower()
                    if not alias:
                        continue
                    out[alias] = kind
        return out

    def _transaction_category_lookup_maps(self) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
        by_code: Dict[str, Dict[str, Any]] = {}
        by_name: Dict[str, Dict[str, Any]] = {}
        rows = self.list_transaction_categories(txn_type="", txn_class="", include_inactive=True)
        for row in rows:
            code = _clean_text(row.get("categoryCode")).lower()
            name = _clean_text(row.get("categoryName")).lower()
            if code:
                by_code[code] = row
            if name:
                by_name[name] = row
        return by_code, by_name

    def _load_seed_csv_rows(self, filename: str) -> List[Dict[str, Any]]:
        seed_path = self.paths.schema_dir() / filename
        if not seed_path.exists():
            return []
        out: List[Dict[str, Any]] = []
        try:
            with seed_path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                for row in reader:
                    cleaned: Dict[str, Any] = {}
                    for key, value in (row or {}).items():
                        cleaned[_clean_text(key)] = _clean_text(value)
                    if any(_clean_text(v) for v in cleaned.values()):
                        out.append(cleaned)
        except Exception:
            return []
        return out

    def _get_or_create_parent(self, parent_name: str) -> Dict[str, Any]:
        name = _clean_text(parent_name)
        rows = self._read_table_rows(TBL_PARENTS)
        for row in rows:
            if _clean_text(row.get(sc.COL_PARENT_NAME)).lower() == name.lower():
                return self._canonicalize_parent_row(row)

        new_row = {
            sc.COL_PARENT_ID: self._new_id("P"),
            sc.COL_PARENT_NAME: name,
            sc.COL_PARENT_DEF_SHARE: 100.0,
            sc.COL_PARENT_DEF_RATE: 0.0,
            sc.COL_PARENT_ACTIVE: 1,
            sc.COL_PARENT_NOTES: "",
        }
        self._append_row_to_table(TBL_PARENTS, new_row)
        return new_row

    def _get_or_create_client(self, client_name: str) -> Dict[str, Any]:
        name = _clean_text(client_name)
        rows = self._read_table_rows(TBL_CLIENTS)
        for row in rows:
            if _clean_text(row.get(sc.COL_CLIENT_NAME)).lower() == name.lower():
                return self._canonicalize_client_row(row)

        new_row = {
            sc.COL_CLIENT_ID: self._new_id("C"),
            sc.COL_CLIENT_NAME: name,
            sc.COL_CLIENT_EMAIL: "",
            sc.COL_CLIENT_PHONE: "",
            sc.COL_CLIENT_STATUS: "Active",
            sc.COL_CLIENT_ACTIVE: 1,
            sc.COL_CLIENT_NOTES: "",
        }
        self._append_row_to_table(TBL_CLIENTS, new_row)
        return new_row

    def _get_or_create_matter(
        self,
        matter_name: str,
        client_id: str,
        parent_id: str,
        default_rate: float,
        default_share_pct: float,
    ) -> Dict[str, Any]:
        name = _clean_text(matter_name)
        rows = self._read_table_rows(TBL_MATTERS)
        for row in rows:
            if _clean_text(row.get(sc.COL_MATTER_NAME)).lower() == name.lower():
                canonical = self._canonicalize_matter_row(row)
                row_client_id = _clean_text(canonical.get(sc.COL_MATTER_CLIENT_ID))
                if client_id and row_client_id and row_client_id.lower() != _clean_text(client_id).lower():
                    continue
                updated = False
                if not canonical.get(sc.COL_MATTER_CLIENT_ID) and client_id:
                    canonical[sc.COL_MATTER_CLIENT_ID] = client_id
                    updated = True
                if not canonical.get(sc.COL_MATTER_CLIENT_NAME) and client_id:
                    client_rows = [self._canonicalize_client_row(r) for r in self._read_table_rows(TBL_CLIENTS)]
                    for client_row in client_rows:
                        if _clean_text(client_row.get(sc.COL_CLIENT_ID)).lower() == _clean_text(client_id).lower():
                            canonical[sc.COL_MATTER_CLIENT_NAME] = _clean_text(
                                client_row.get(sc.COL_CLIENT_NAME)
                            )
                            updated = True
                            break
                if not canonical.get(sc.COL_MATTER_PARENT_ID) and parent_id:
                    canonical[sc.COL_MATTER_PARENT_ID] = parent_id
                    updated = True
                if not canonical.get(sc.COL_MATTER_OPEN_DATE):
                    canonical[sc.COL_MATTER_OPEN_DATE] = datetime.now().strftime("%Y-%m-%d")
                    updated = True
                if not canonical.get(sc.COL_MATTER_DISPLAY_NAME):
                    canonical[sc.COL_MATTER_DISPLAY_NAME] = canonical.get(sc.COL_MATTER_NAME, "")
                    updated = True
                if not canonical.get(sc.COL_MATTER_NUMBER):
                    canonical[sc.COL_MATTER_NUMBER] = self._build_matter_number(
                        client_name=canonical.get(sc.COL_MATTER_CLIENT_NAME, ""),
                        matter_type=canonical.get(sc.COL_MATTER_TYPE, "General"),
                        date_opened=canonical.get(sc.COL_MATTER_OPEN_DATE, ""),
                        existing_matter_id=canonical.get(sc.COL_MATTER_ID, ""),
                        parent_name=canonical.get(sc.COL_MATTER_PARENT_NAME, ""),
                        entity_type="",
                    )
                    updated = True
                if updated:
                    canonical[sc.COL_MATTER_UPDATED] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    self._upsert_row_by_key(
                        TBL_MATTERS,
                        sc.COL_MATTER_ID,
                        canonical.get(sc.COL_MATTER_ID, ""),
                        canonical,
                    )
                return canonical

        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        opened_date = datetime.now().strftime("%Y-%m-%d")
        new_row = {
            sc.COL_MATTER_ID: self._new_id("M"),
            sc.COL_MATTER_NUMBER: "",
            sc.COL_MATTER_NAME: name,
            sc.COL_MATTER_DISPLAY_NAME: name,
            sc.COL_MATTER_CLIENT_ID: client_id,
            sc.COL_MATTER_CLIENT_NAME: "",
            sc.COL_MATTER_PARENT_ID: parent_id,
            sc.COL_MATTER_PARENT_NAME: "",
            sc.COL_MATTER_TYPE: "General",
            sc.COL_MATTER_PRACTICE_AREA: "",
            sc.COL_MATTER_STATUS: "Open",
            sc.COL_MATTER_RESPONSIBLE_LAWYER: "",
            sc.COL_MATTER_BILLING_ARRANGEMENT: "Hourly",
            sc.COL_MATTER_BILLING_CONTACT: "",
            sc.COL_MATTER_BILLING_EMAIL: "",
            sc.COL_MATTER_DEF_RATE: round(float(default_rate), 2),
            sc.COL_MATTER_DEF_SHARE: round(float(default_share_pct), 2),
            sc.COL_MATTER_ENGAGEMENT_DATE: "",
            sc.COL_MATTER_OPEN_DATE: opened_date,
            sc.COL_MATTER_CLOSE_DATE: "",
            sc.COL_MATTER_COURT_FILE_NO: "",
            sc.COL_MATTER_OPPOSING_PARTY: "",
            sc.COL_MATTER_REFERRAL_FROM: "",
            sc.COL_MATTER_DESCRIPTION: "",
            sc.COL_MATTER_NOTES: "",
            sc.COL_MATTER_CREATED: created_at,
            sc.COL_MATTER_UPDATED: created_at,
        }
        if client_id:
            client_rows = [self._canonicalize_client_row(r) for r in self._read_table_rows(TBL_CLIENTS)]
            for client_row in client_rows:
                if _clean_text(client_row.get(sc.COL_CLIENT_ID)).lower() == _clean_text(client_id).lower():
                    new_row[sc.COL_MATTER_CLIENT_NAME] = _clean_text(client_row.get(sc.COL_CLIENT_NAME))
                    break
        if parent_id:
            parent_rows = [self._canonicalize_parent_row(r) for r in self._read_table_rows(TBL_PARENTS)]
            for parent_row in parent_rows:
                if _clean_text(parent_row.get(sc.COL_PARENT_ID)).lower() == _clean_text(parent_id).lower():
                    new_row[sc.COL_MATTER_PARENT_NAME] = _clean_text(parent_row.get(sc.COL_PARENT_NAME))
                    break
        new_row[sc.COL_MATTER_NUMBER] = self._build_matter_number(
            client_name=new_row.get(sc.COL_MATTER_CLIENT_NAME, ""),
            matter_type=new_row.get(sc.COL_MATTER_TYPE, "General"),
            date_opened=opened_date,
            existing_matter_id="",
            parent_name=new_row.get(sc.COL_MATTER_PARENT_NAME, ""),
            entity_type="",
        )
        self._append_row_to_table(TBL_MATTERS, new_row)
        return new_row

    def _find_transaction_row(self, transaction_key: str) -> Optional[Dict[str, Any]]:
        target = _clean_text(transaction_key)
        if not target:
            return None
        target_lc = target.lower()
        rows = [
            self._canonicalize_transaction_row(r)
            for r in self._read_table_rows(TBL_TRANSACTIONS_MASTER)
        ]
        for row in rows:
            if _clean_text(row.get(sc.COL_TXN_ID)).lower() == target_lc:
                return row
        return None

    def _find_time_entry(self, entry_id: str) -> Optional[Dict[str, Any]]:
        rows = self._read_table_rows(TBL_TIME)
        target = _clean_text(entry_id)
        for row in rows:
            if _clean_text(row.get(sc.COL_TIME_ENTRY_ID)) == target:
                return self._canonicalize_time_row(row)
        return None

    def _find_client_profile(self, client_id: str) -> Optional[Dict[str, Any]]:
        rows = self._read_table_rows(TBL_CLIENT_PROFILES)
        target = _clean_text(client_id)
        for row in rows:
            if _clean_text(row.get(sc.COL_PROFILE_CLIENT_ID)) == target:
                return self._canonicalize_client_profile_row(row)
        return None

    def _find_matter_row(self, matter_key: str) -> Optional[Dict[str, Any]]:
        target = _clean_text(matter_key)
        if not target:
            return None
        target_lc = target.lower()
        rows = [self._canonicalize_matter_row(r) for r in self._read_table_rows(TBL_MATTERS)]
        for row in rows:
            if _clean_text(row.get(sc.COL_MATTER_ID)).lower() == target_lc:
                return row
        for row in rows:
            if _clean_text(row.get(sc.COL_MATTER_NAME)).lower() == target_lc:
                return row
        for row in rows:
            if _clean_text(row.get(sc.COL_MATTER_DISPLAY_NAME)).lower() == target_lc:
                return row
        for row in rows:
            if _clean_text(row.get(sc.COL_MATTER_NUMBER)).lower() == target_lc:
                return row
        return None

    def _compare_client_profile_rows_loose(
        self,
        expected: Dict[str, Any],
        actual: Optional[Dict[str, Any]],
    ) -> bool:
        if not actual:
            return False

        numeric_keys = {sc.COL_PROFILE_RETAINER_AMOUNT}
        bool_int_keys = {sc.COL_PROFILE_RETAINER_REQUIRED}

        for key, expected_value in expected.items():
            actual_value = actual.get(key)
            if key in numeric_keys:
                ev = self._parse_float(expected_value)
                av = self._parse_float(actual_value)
                if ev is None and av is None:
                    continue
                if ev is None or av is None:
                    return False
                if abs(float(ev) - float(av)) > 0.001:
                    return False
                continue

            if key in bool_int_keys:
                ev = self._to_bool_int(expected_value, default=0)
                av = self._to_bool_int(actual_value, default=0)
                if ev != av:
                    return False
                continue

            if _clean_text(expected_value) != _clean_text(actual_value):
                return False
        return True

    def _compare_matter_profile_rows_loose(
        self,
        expected: Dict[str, Any],
        actual: Optional[Dict[str, Any]],
    ) -> bool:
        if not actual:
            return False

        numeric_keys = {sc.COL_MATTER_DEF_RATE, sc.COL_MATTER_DEF_SHARE}
        for key, expected_value in expected.items():
            actual_value = actual.get(key)
            if key in numeric_keys:
                ev = self._parse_float(expected_value)
                av = self._parse_float(actual_value)
                if ev is None and av is None:
                    continue
                if ev is None or av is None:
                    return False
                if abs(float(ev) - float(av)) > 0.001:
                    return False
                continue

            if _clean_text(expected_value) != _clean_text(actual_value):
                return False
        return True

    def _compare_rows_loose(self, expected: Dict[str, Any], actual: Optional[Dict[str, Any]]) -> bool:
        if not actual:
            return False
        numeric_keys = {
            sc.COL_TIME_HOURS,
            sc.COL_TIME_RATE,
            sc.COL_TIME_SHARE_PCT,
            sc.COL_TIME_GROSS,
            sc.COL_TIME_NET,
            sc.COL_TIME_HST,
            sc.COL_TIME_TOTAL,
            sc.COL_TIME_SECONDS,
        }
        for key, expected_value in expected.items():
            actual_value = actual.get(key)
            if key in numeric_keys:
                ev = self._parse_float(expected_value)
                av = self._parse_float(actual_value)
                if ev is None and av is None:
                    continue
                if ev is None or av is None:
                    return False
                if abs(float(ev) - float(av)) > 0.001:
                    return False
            else:
                if _clean_text(expected_value) != _clean_text(actual_value):
                    return False
        return True

    def _compare_transaction_rows_loose(
        self,
        expected: Dict[str, Any],
        actual: Optional[Dict[str, Any]],
    ) -> bool:
        if not actual:
            return False

        numeric_keys = {
            sc.COL_TXN_AMOUNT,
            sc.COL_TXN_TAX_AMOUNT,
            sc.COL_TXN_BILL_CLAIM_PCT,
            sc.COL_TXN_TOTAL_CLAIM_AMOUNT,
        }
        bool_int_keys = {
            sc.COL_TXN_HST_EXEMPT,
            sc.COL_TXN_GENERAL_OFFICE_EXPENSE,
            sc.COL_TXN_SHADOW,
        }

        for key, expected_value in expected.items():
            actual_value = actual.get(key)
            if key in numeric_keys:
                ev = self._parse_float(expected_value)
                av = self._parse_float(actual_value)
                if ev is None and av is None:
                    continue
                if ev is None or av is None:
                    return False
                if abs(float(ev) - float(av)) > 0.001:
                    return False
                continue

            if key in bool_int_keys:
                ev = self._to_bool_int(expected_value, default=0)
                av = self._to_bool_int(actual_value, default=0)
                if ev != av:
                    return False
                continue

            if _clean_text(expected_value) != _clean_text(actual_value):
                return False
        return True

    def _pick_value(self, payload: Dict[str, Any], keys: Iterable[str]) -> Any:
        for key in keys:
            if key in payload and payload[key] not in (None, ""):
                return payload[key]
        return None

    def _pick_text(self, payload: Dict[str, Any], keys: Iterable[str]) -> str:
        value = self._pick_value(payload, keys)
        return _clean_text(value)

    def _pick_float(self, payload: Dict[str, Any], keys: Iterable[str]) -> Optional[float]:
        value = self._pick_value(payload, keys)
        return self._parse_float(value)

    def _pick_int(self, payload: Dict[str, Any], keys: Iterable[str]) -> Optional[int]:
        value = self._pick_value(payload, keys)
        return self._parse_int(value)

    def _parse_float(self, value: Any) -> Optional[float]:
        s = _clean_text(value)
        if s == "":
            return None
        s = s.replace(",", "").replace("$", "").replace("%", "")
        try:
            return float(s)
        except Exception:
            return None

    def _parse_int(self, value: Any) -> Optional[int]:
        number = self._parse_float(value)
        if number is None:
            return None
        return int(round(number))

    def _parse_date_value(self, value: Any) -> Optional[date]:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = _clean_text(value)
        if not text:
            return None
        normalized = text.replace("T", " ").split(" ")[0]
        for candidate in (normalized, text):
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(candidate, fmt).date()
                except Exception:
                    pass
        return None

    def _date_iso(self, value: Any) -> str:
        parsed = self._parse_date_value(value)
        return parsed.isoformat() if parsed else _clean_text(value)

    def _to_bool_int(self, value: Any, default: int = 1) -> int:
        if value in (None, ""):
            return 1 if default else 0
        if isinstance(value, bool):
            return 1 if value else 0
        s = _clean_text(value).lower()
        if s in ("1", "true", "yes", "y", "active", "open"):
            return 1
        if s in ("0", "false", "no", "n", "inactive", "closed"):
            return 0
        n = self._parse_float(value)
        if n is None:
            return 1 if default else 0
        return 0 if abs(n) < 0.0001 else 1

    def _new_id(self, prefix: str) -> str:
        return f"{prefix}_{uuid4().hex[:10]}"

    def _client_parent_lookup(
        self,
        profile_rows: List[Dict[str, Any]],
        client_rows: List[Dict[str, Any]],
        parent_rows: List[Dict[str, Any]],
    ) -> Tuple[Dict[str, str], Dict[str, Dict[str, str]]]:
        parent_names: Dict[str, str] = {}
        for row in parent_rows:
            parent_id = _clean_text(row.get(sc.COL_PARENT_ID))
            parent_name = _clean_text(row.get(sc.COL_PARENT_NAME))
            if parent_id:
                parent_names[parent_id] = parent_name or parent_id

        client_names_by_id: Dict[str, str] = {}
        for row in client_rows:
            client_id = _clean_text(row.get(sc.COL_CLIENT_ID))
            client_name = _clean_text(row.get(sc.COL_CLIENT_NAME))
            if client_id and client_name:
                client_names_by_id[client_id] = client_name

        lookup: Dict[str, Dict[str, str]] = {}

        def register(value: Any, parent_id: str, parent_name: str) -> None:
            if not _clean_text(value) or not (parent_id or parent_name):
                return
            info = {"id": parent_id, "name": parent_name or parent_names.get(parent_id, "")}
            for key in _report_name_keys(value):
                lookup.setdefault(key, info)

        for row in profile_rows:
            client_id = _clean_text(row.get(sc.COL_PROFILE_CLIENT_ID))
            display_name = (
                _clean_text(row.get(sc.COL_PROFILE_DISPLAY_NAME))
                or _clean_text(row.get(sc.COL_PROFILE_LEGAL_NAME))
                or client_id
            )
            legal_name = _clean_text(row.get(sc.COL_PROFILE_LEGAL_NAME))
            parent_id = _clean_text(row.get(sc.COL_PROFILE_PARENT_ID))
            parent_name = _clean_text(row.get(sc.COL_PROFILE_PARENT_NAME)) or parent_names.get(parent_id, "")
            if not (parent_id or parent_name):
                continue
            register(client_id, parent_id, parent_name)
            register(display_name, parent_id, parent_name)
            register(legal_name, parent_id, parent_name)
            register(client_names_by_id.get(client_id, ""), parent_id, parent_name)

        return parent_names, lookup
    def delete_ledger_entry(self, ledger_id: str) -> Dict[str, Any]:
        self.ensure_schema()
        # The id may be in TBL_LEDGER or TBL_TRANSACTIONS_MASTER
        deleted_ledger = self._delete_row_by_key_hard(TBL_LEDGER, sc.COL_LEDGER_ID, ledger_id)
        deleted_txn = self._delete_row_by_key_hard(TBL_TRANSACTIONS_MASTER, sc.COL_TXN_ID, ledger_id)
        if deleted_ledger or deleted_txn:
            return {"ok": True, "message": f"Ledger/Transaction entry deleted."}
        return {"ok": False, "message": f"Ledger entry not found."}

    def get_client_ledger_report(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        with getattr(self, "_lock", None) or getattr(self, "_FILE_LOCK", None) or __import__("contextlib").nullcontext():
            payload = dict(payload or {})
            selected_client = _clean_text(payload.get("clientId") or payload.get("client"))
            selected_billing = _clean_text(payload.get("billingClientId") or payload.get("billingClient"))
            selected_matter = _clean_text(payload.get("matterId") or payload.get("matter"))
            start_date = self._parse_date_value(payload.get("startDate") or payload.get("fromDate"))
            end_date = self._parse_date_value(payload.get("endDate") or payload.get("toDate"))
            search_text = _clean_text(payload.get("searchText") or payload.get("query")).lower()

            show_time = bool(payload.get("showTime", True))
            show_fees = bool(payload.get("showFees", True))
            show_disb = bool(payload.get("showDisbursements", True))
            show_inv = bool(payload.get("showInvoices", True))
            show_pay = bool(payload.get("showPayments", True))
            show_cred = bool(payload.get("showCredits", True))

            self.ensure_schema()
            parents_rows = self._read_table_rows(TBL_PARENTS)
            clients_rows = self._read_table_rows(TBL_CLIENTS)
            profile_rows = self._read_table_rows(TBL_CLIENT_PROFILES)
            matters_rows = self._read_table_rows(TBL_MATTERS)
            time_rows = self._read_table_rows(TBL_TIME)
            disb_rows = self._read_table_rows(TBL_DISBURSEMENTS)
            invoice_rows = self._read_table_rows(TBL_INVOICE_LOG)
            receivable_rows = self._read_table_rows(TBL_RECEIVABLES)
            
            voided_invoice_refs = set()
            for row in receivable_rows:
                invoice_ref = _clean_text(row.get(sc.COL_RECV_INVOICE_NUM))
                status = _clean_text(row.get(sc.COL_RECV_STATUS)).casefold()
                if invoice_ref and status in {
                    "void",
                    "voided",
                    "reversed",
                    "cancelled",
                    "canceled",
                    "superseded",
                }:
                    voided_invoice_refs.add(invoice_ref.upper())
            for row in invoice_rows:
                invoice_ref = _clean_text(row.get(sc.COL_INV_INVOICE_NUM))
                if invoice_ref.upper().endswith("-V"):
                    voided_invoice_refs.add(invoice_ref[:-2].upper())
                elif invoice_ref.upper().endswith("-SUPERSEDED"):
                    voided_invoice_refs.add(invoice_ref.upper())

            all_invoice_refs = {
                _clean_text(r.get(sc.COL_INV_INVOICE_NUM)).upper()
                for r in invoice_rows
                if _clean_text(r.get(sc.COL_INV_INVOICE_NUM))
                and _clean_text(r.get(sc.COL_INV_INVOICE_NUM)).upper() not in voided_invoice_refs
                and not _clean_text(r.get(sc.COL_INV_INVOICE_NUM)).upper().endswith("-V")
            }

            client_names: Dict[str, str] = {}
            client_parent: Dict[str, str] = {}
            parent_names, client_parent_lookup = self._client_parent_lookup(profile_rows, clients_rows, parents_rows)
            matter_names: Dict[str, str] = {}
            matter_client: Dict[str, str] = {}
            matter_parent: Dict[str, str] = {}

            def put_option(target: Dict[str, Dict[str, str]], item_id: Any, label: Any) -> None:
                opt_id = _clean_text(item_id) or _clean_text(label)
                opt_label = _clean_text(label) or opt_id
                if opt_id and opt_label:
                    target[opt_id.lower()] = {"id": opt_id, "label": opt_label}

            billing_options: Dict[str, Dict[str, str]] = {}
            client_options: Dict[str, Dict[str, str]] = {}
            matter_options: Dict[str, Dict[str, str]] = {}

            for row in parents_rows:
                parent_id = _clean_text(row.get(sc.COL_PARENT_ID))
                parent_name = _clean_text(row.get(sc.COL_PARENT_NAME))
                put_option(billing_options, parent_id, parent_name)

            for row in clients_rows:
                client_id = _clean_text(row.get(sc.COL_CLIENT_ID))
                client_name = _clean_text(row.get(sc.COL_CLIENT_NAME))
                if client_id:
                    client_names[client_id] = client_name or client_id
                put_option(client_options, client_id, client_name)

            for row in profile_rows:
                client_id = _clean_text(row.get(sc.COL_PROFILE_CLIENT_ID))
                display_name = (
                    _clean_text(row.get(sc.COL_PROFILE_DISPLAY_NAME))
                    or _clean_text(row.get(sc.COL_PROFILE_LEGAL_NAME))
                    or client_id
                )
                parent_id = _clean_text(row.get(sc.COL_PROFILE_PARENT_ID))
                parent_name = _clean_text(row.get(sc.COL_PROFILE_PARENT_NAME))
                if client_id:
                    client_names[client_id] = display_name
                    client_parent[client_id] = parent_id
                put_option(client_options, client_id, display_name)
                put_option(billing_options, parent_id or parent_name, parent_name)

            for row in matters_rows:
                matter_id = _clean_text(row.get(sc.COL_MATTER_ID))
                matter_name = _clean_text(row.get(sc.COL_MATTER_NAME)) or _clean_text(row.get(sc.COL_MATTER_DISPLAY_NAME))
                if matter_id:
                    matter_names[matter_id] = matter_name or matter_id
                    matter_client[matter_id] = _clean_text(row.get(sc.COL_MATTER_CLIENT_ID))
                    matter_parent[matter_id] = _clean_text(row.get(sc.COL_MATTER_PARENT_ID))
                put_option(matter_options, matter_id, matter_name)

            for row in receivable_rows:
                billing_name = _clean_text(row.get(sc.COL_RECV_CLIENT))
                work_name = _clean_text(row.get(sc.COL_RECV_WORK_CLIENT))
                put_option(billing_options, billing_name, billing_name)
                put_option(client_options, work_name or billing_name, work_name or billing_name)

            def options_from(mapping: Dict[str, Dict[str, str]]) -> List[Dict[str, str]]:
                return sorted(mapping.values(), key=lambda item: item.get("label", "").lower())

            def selected_all(value: str) -> bool:
                return not value or value.upper() == "ALL"

            def text_matches(selected: str, *values: Any) -> bool:
                if selected_all(selected):
                    return True
                wanted = selected.lower()
                return any(_clean_text(value).lower() == wanted for value in values if _clean_text(value))

            def parent_info_for_client(*values: Any) -> Dict[str, str]:
                for value in values:
                    for key in _report_name_keys(value):
                        info = client_parent_lookup.get(key)
                        if info:
                            return info
                return {"id": "", "name": ""}

            def date_matches(value: Any) -> bool:
                parsed = self._parse_date_value(value)
                if start_date and (parsed is None or parsed < start_date):
                    return False
                if end_date and (parsed is None or parsed > end_date):
                    return False
                return True

            def amount(value: Any) -> float:
                return float(self._parse_float(value) or 0.0)

            def add_entry(entry: Dict[str, Any]) -> None:
                haystack = " ".join(
                    _clean_text(entry.get(key))
                    for key in ("date", "type", "clientName", "matterName", "description", "status", "invoiceRef")
                ).lower()
                if search_text and search_text not in haystack:
                    return
                entry["date"] = self._date_iso(entry.get("date"))
                entry["debit"] = round(float(entry.get("debit") or 0.0), 2)
                entry["credit"] = round(float(entry.get("credit") or 0.0), 2)
                entry["hours"] = round(float(entry.get("hours") or 0.0), 2)
                entry.setdefault("tax", 0.0)
                entry.setdefault("matter", entry.get("matterName", ""))
                entries.append(entry)

            entries: List[Dict[str, Any]] = []
            
            wip_relieved_map: Dict[str, float] = collections.defaultdict(float)
            for row in time_rows:
                inv_ref = _clean_text(row.get(sc.COL_TIME_INVOICE_REF))
                if inv_ref and inv_ref.upper() not in voided_invoice_refs:
                    wip_relieved_map[inv_ref.upper()] += amount(row.get(sc.COL_TIME_TOTAL)) or amount(row.get(sc.COL_TIME_GROSS))
            for row in disb_rows:
                inv_ref = _clean_text(row.get(sc.COL_DISB_INVOICE_REF))
                if inv_ref and inv_ref.upper() not in voided_invoice_refs:
                    wip_relieved_map[inv_ref.upper()] += amount(row.get(sc.COL_DISB_AMOUNT))

            if show_time or show_fees:
                for row in time_rows:
                    client_id = _clean_text(row.get(sc.COL_TIME_CLIENT_ID))
                    client_name = client_names.get(client_id, client_id)
                    matter_id = _clean_text(row.get(sc.COL_TIME_MATTER_ID))
                    matter_name = matter_names.get(matter_id, matter_id)
                    parent_info = parent_info_for_client(client_id, client_name)
                    parent_id = (
                        _clean_text(row.get(sc.COL_TIME_PARENT_ID))
                        or matter_parent.get(matter_id)
                        or client_parent.get(client_id, "")
                        or parent_info.get("id", "")
                    )
                    parent_name = parent_names.get(parent_id, "") or parent_info.get("name", "")
                    if not text_matches(selected_client, client_id, client_name):
                        continue
                    if not text_matches(selected_billing, parent_id, parent_name):
                        continue
                    if not text_matches(selected_matter, matter_id, matter_name):
                        continue
                    if not date_matches(row.get(sc.COL_TIME_DATE)):
                        continue
                    hours = amount(row.get(sc.COL_TIME_HOURS))
                    is_time = hours > 0
                    if is_time and not show_time:
                        continue
                    if not is_time and not show_fees:
                        continue
                    status_raw = _clean_text(row.get(sc.COL_TIME_STATUS))
                    status = "WIP" if status_raw.lower() not in {"billed", "merged"} else status_raw
                    invoice_ref = _clean_text(row.get(sc.COL_TIME_INVOICE_REF))
                    debit = amount(row.get(sc.COL_TIME_TOTAL)) or amount(row.get(sc.COL_TIME_GROSS))
                    add_entry({
                        "date": row.get(sc.COL_TIME_DATE),
                        "type": "Time" if is_time else "Fee",
                        "clientId": client_id,
                        "clientName": client_name,
                        "matterId": matter_id,
                        "matterName": matter_name,
                        "matter": matter_name,
                        "description": _clean_text(row.get(sc.COL_TIME_DESC)),
                        "hours": hours,
                        "debit": debit,
                        "credit": 0.0,
                        "status": status,
                        "invoiceRef": invoice_ref,
                        "billingParentId": parent_id,
                        "billingParentName": parent_name,
                    })
                    
                    if (
                        status.lower() in {"billed", "merged"}
                        and invoice_ref.upper() not in voided_invoice_refs
                        and (not invoice_ref or invoice_ref.upper() not in all_invoice_refs)
                    ):
                        add_entry({
                            "date": row.get(sc.COL_TIME_DATE),
                            "type": "Invoice",
                            "invoiceSubType": "wipTransfer",
                            "clientId": client_id,
                            "clientName": client_name,
                            "matterId": matter_id,
                            "matterName": matter_name,
                            "matter": matter_name,
                            "description": "WIP Relieved (Legacy Orphan)",
                            "hours": 0.0,
                            "debit": 0.0,
                            "credit": debit,
                            "status": status,
                            "invoiceRef": invoice_ref or "LEGACY",
                            "billingParentId": parent_id,
                            "billingParentName": parent_name,
                            "tax": 0.0,
                        })

            if show_disb:
                for row in disb_rows:
                    client_id = _clean_text(row.get(sc.COL_DISB_CLIENT_ID))
                    client_name = (
                        client_names.get(client_id)
                        or _clean_text(row.get(sc.COL_DISB_SUB_CLIENT))
                        or _clean_text(row.get(sc.COL_DISB_CLIENT_NAME))
                        or client_id
                    )
                    matter_id = _clean_text(row.get(sc.COL_DISB_MATTER_ID))
                    matter_name = matter_names.get(matter_id, matter_id)
                    parent_info = parent_info_for_client(client_id, client_name)
                    parent_id = (
                        _clean_text(row.get(sc.COL_DISB_PARENT_ID))
                        or matter_parent.get(matter_id)
                        or client_parent.get(client_id, "")
                        or parent_info.get("id", "")
                    )
                    parent_name = parent_names.get(parent_id, "") or parent_info.get("name", "")
                    invoice_ref = _clean_text(row.get(sc.COL_DISB_INVOICE_REF))
                    if not text_matches(selected_client, client_id, client_name):
                        continue
                    if not text_matches(selected_billing, parent_id, parent_name, _clean_text(row.get(sc.COL_DISB_CLIENT_NAME))):
                        continue
                    if not text_matches(selected_matter, matter_id, matter_name):
                        continue
                    if not date_matches(row.get(sc.COL_DISB_DATE)):
                        continue
                    add_entry({
                        "date": row.get(sc.COL_DISB_DATE),
                        "type": "Disbursement",
                        "clientId": client_id,
                        "clientName": client_name,
                        "matterId": matter_id,
                        "matterName": matter_name,
                        "matter": matter_name,
                        "description": _clean_text(row.get(sc.COL_DISB_DESCRIPTION)),
                        "hours": 0.0,
                        "debit": amount(row.get(sc.COL_DISB_AMOUNT)),
                        "credit": 0.0,
                        "status": "Billed" if invoice_ref else "WIP",
                        "invoiceRef": invoice_ref,
                        "billingParentId": parent_id,
                        "billingParentName": parent_name,
                    })

            if show_inv:
                for row in invoice_rows:
                    billing_name = _clean_text(row.get(sc.COL_INV_BILL_TO_CLIENT)) or _clean_text(row.get(sc.COL_INV_CLIENT_NAME))
                    client_name = _clean_text(row.get(sc.COL_INV_SUB_CLIENT)) or _clean_text(row.get(sc.COL_INV_CLIENT_NAME)) or billing_name
                    parent_info = parent_info_for_client(client_name, billing_name)
                    parent_id = parent_info.get("id", "")
                    parent_name = parent_info.get("name", "")
                    if not text_matches(selected_client, client_name, billing_name):
                        continue
                    if not text_matches(selected_billing, billing_name, parent_id, parent_name):
                        continue
                    if not selected_all(selected_matter):
                        continue
                    if not date_matches(row.get(sc.COL_INV_INVOICE_DATE)):
                        continue
                    invoice_num = _clean_text(row.get(sc.COL_INV_INVOICE_NUM))
                    if invoice_num.upper() in voided_invoice_refs or invoice_num.upper().endswith("-V"):
                        continue
                    inv_total_fees = amount(row.get(sc.COL_INV_TOTAL_FEES))
                    inv_total_disb = amount(row.get(sc.COL_INV_TOTAL_DISBURSEMENTS))
                    inv_total_tax = amount(row.get(sc.COL_INV_TOTAL_TAX))
                    inv_aggregate = amount(row.get(sc.COL_INV_AGGREGATE_BILLED))
                    inv_fees_portion = inv_total_fees + inv_total_disb
                    if inv_fees_portion <= 0 and inv_aggregate > 0:
                        inv_fees_portion = inv_aggregate - inv_total_tax
                    inv_status = _clean_text(row.get(sc.COL_RECV_STATUS))
                    inv_common = {
                        "clientName": client_name,
                        "matterName": "",
                        "matter": "",
                        "hours": 0.0,
                        "status": inv_status,
                        "invoiceRef": invoice_num,
                        "billingParentId": parent_id,
                        "billingParentName": parent_name,
                        "invoiceEditable": True,
                        "invoiceFees": round(inv_total_fees, 2),
                        "invoiceDisb": round(inv_total_disb, 2),
                        "invoiceTax": round(inv_total_tax, 2),
                        "invoiceAggregate": round(inv_aggregate, 2),
                        "invoiceClientName": _clean_text(row.get(sc.COL_INV_CLIENT_NAME)),
                        "invoiceSubClient": _clean_text(row.get(sc.COL_INV_SUB_CLIENT)),
                        "invoiceBillTo": _clean_text(row.get(sc.COL_INV_BILL_TO_CLIENT)),
                    }
                    # Line 1: Billed reversal — credit out the fees that were already docketed as WIP
                    true_wip_credit = wip_relieved_map.get(invoice_num.upper(), 0.0)
                    wip_credit_to_use = true_wip_credit # Fix: Do not inflate WIP relieved beyond actual docketed time
                    if wip_credit_to_use > 0:
                        add_entry({
                            **inv_common,
                            "date": row.get(sc.COL_INV_INVOICE_DATE),
                            "type": "Invoice",
                            "invoiceSubType": "wipTransfer",
                            "description": f"WIP Relieved \u2014 Invoice {invoice_num}",
                            "debit": 0.0,
                            "credit": round(wip_credit_to_use, 2),
                            "tax": 0.0,
                        })
                    # Line 2: Fees receivable — the fee portion now owed as A/R
                    add_entry({
                        **inv_common,
                        "date": row.get(sc.COL_INV_INVOICE_DATE),
                        "type": "Invoice",
                        "invoiceSubType": "fees",
                        "description": f"Fees on Invoice {invoice_num}",
                        "debit": round(inv_fees_portion, 2),
                        "credit": 0.0,
                        "tax": 0.0,
                    })
                    # Line 3: Tax receivable — only the HST is genuinely new
                    add_entry({
                        **inv_common,
                        "date": row.get(sc.COL_INV_INVOICE_DATE),
                        "type": "Invoice",
                        "invoiceSubType": "tax",
                        "description": f"Tax on Invoice {invoice_num}",
                        "debit": round(inv_total_tax, 2),
                        "credit": 0.0,
                        "tax": round(inv_total_tax, 2),
                    })

            for row in receivable_rows:
                billing_name = _clean_text(row.get(sc.COL_RECV_CLIENT))
                work_name = _clean_text(row.get(sc.COL_RECV_WORK_CLIENT))
                client_name = work_name or billing_name
                parent_info = parent_info_for_client(client_name, billing_name)
                parent_id = parent_info.get("id", "")
                parent_name = parent_info.get("name", "")
                if not text_matches(selected_client, client_name, billing_name):
                    continue
                if not text_matches(selected_billing, billing_name, parent_id, parent_name):
                    continue
                if not selected_all(selected_matter):
                    continue
                if not date_matches(row.get(sc.COL_RECV_DATE)):
                    continue
                invoice_num = _clean_text(row.get(sc.COL_RECV_INVOICE_NUM))
                if invoice_num.upper() in voided_invoice_refs:
                    continue
                status = _clean_text(row.get(sc.COL_RECV_STATUS))
                paid = amount(row.get(sc.COL_RECV_AMOUNT_PAID))
                credits = amount(row.get(sc.COL_RECV_CREDITS_ADJ))
                if show_pay and paid > 0:
                    add_entry({
                        "date": row.get(sc.COL_RECV_DATE),
                        "type": "Payment",
                        "clientName": client_name,
                        "matterName": "",
                        "matter": "",
                        "description": f"Payment applied to invoice {invoice_num}",
                        "hours": 0.0,
                        "debit": 0.0,
                        "credit": paid,
                        "status": status,
                        "invoiceRef": invoice_num,
                        "billingParentId": parent_id,
                        "billingParentName": parent_name,
                    })
                if show_cred and credits > 0:
                    add_entry({
                        "date": row.get(sc.COL_RECV_DATE),
                        "type": "Credit/Adj",
                        "clientName": client_name,
                        "matterName": "",
                        "matter": "",
                        "description": f"Credit/adjustment on invoice {invoice_num}",
                        "hours": 0.0,
                        "debit": 0.0,
                        "credit": credits,
                        "status": status,
                        "invoiceRef": invoice_num,
                        "billingParentId": parent_id,
                        "billingParentName": parent_name,
                    })

            type_order = {"Time": 0, "Fee": 1, "Disbursement": 2, "Invoice": 3, "Payment": 4, "Credit/Adj": 5}
            inv_sub_order = {"wipTransfer": 0, "fees": 1, "tax": 2}
            entries.sort(
                key=lambda entry: (
                    self._parse_date_value(entry.get("date")) or date.min,
                    type_order.get(_clean_text(entry.get("type")), 99),
                    inv_sub_order.get(entry.get("invoiceSubType", ""), 99),
                    _clean_text(entry.get("description")).lower(),
                )
            )
            running_balance = 0.0
            for entry in entries:
                running_balance += float(entry.get("debit") or 0.0) - float(entry.get("credit") or 0.0)
                entry["balance"] = round(running_balance, 2)

            selected_label = "All Clients"
            if not selected_all(selected_client):
                selected_label = client_names.get(selected_client, selected_client)

            client_info = {"clientId": selected_client or "ALL", "clientName": selected_label}
            return {
                "ok": True,
                "entries": entries,
                "optionClients": options_from(client_options),
                "optionBillingClients": options_from(billing_options),
                "optionMatters": options_from(matter_options),
                "client": client_info,
                "clientInfo": client_info,
                "message": "" if entries else "No ledger entries found for the selected filters.",
            }

    def ar_aging_report(self, filters: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """A/R detail is generated from open actual invoice rows in tblReceivables."""
        payload = dict(filters or {})
        as_of = self._parse_date_value(payload.get("asOfDate")) or date.today()
        query = _clean_text(payload.get("query") or payload.get("searchText")).lower()
        group_by_raw = _clean_text(payload.get("groupBy") or payload.get("group_by") or payload.get("groupMode")).lower()
        group_by = (
            "billingClient"
            if group_by_raw in {"billing", "billingclient", "billing_client", "billing-client", "billing client"}
            else "client"
        )
        actual_invoice_pattern = re.compile(r"^\d{2}-\d{4}(?:-[A-Z])?$")

        excluded_invoices: set[str] = set()
        for item in payload.get("excludedInvoices", []) or []:
            if isinstance(item, dict):
                invoice_value = (
                    item.get("invoice")
                    or item.get("invoiceNumber")
                    or item.get("reference")
                    or item.get("id")
                )
            else:
                invoice_value = item
            invoice_text = _clean_text(invoice_value).upper()
            if invoice_text:
                excluded_invoices.add(invoice_text)

        def money(value: Any) -> str:
            number = float(value or 0.0)
            sign = "-" if number < 0 else ""
            return f"{sign}${abs(number):,.2f}"

        def amount(value: Any) -> float:
            return float(self._parse_float(value) or 0.0)

        def money_round(value: Any) -> float:
            try:
                return float(Decimal(str(value or 0.0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
            except Exception:
                return 0.0

        def normalize_ledger_client(value: Any) -> str:
            text = _clean_text(value)
            text = re.sub(r"\s*\{.*?\}\s*$", "", text).strip()
            return re.sub(r"\s+", " ", text)

        def ref_year(ref: str) -> Optional[int]:
            if actual_invoice_pattern.match(ref):
                try:
                    return 2000 + int(ref[:2])
                except Exception:
                    return None
            return None

        def row_year(value: Any) -> Optional[int]:
            parsed = self._parse_date_value(value)
            return parsed.year if parsed else None

        def net_of_hst(gross_value: Any) -> float:
            return money_round(Decimal(str(gross_value or 0.0)) / Decimal("1.13"))

        def bucket_key(age_days: int) -> Tuple[str, str]:
            if age_days <= 30:
                return "current", "0-30"
            if age_days <= 60:
                return "days31To60", "31-60"
            if age_days <= 90:
                return "days61To90", "61-90"
            return "days90Plus", "90+"

        self.ensure_schema()
        receivable_rows = self._read_table_rows(TBL_RECEIVABLES)
        ledger_rows = self._read_table_rows(TBL_LEDGER)
        parent_rows = self._read_table_rows(TBL_PARENTS)
        client_rows = self._read_table_rows(TBL_CLIENTS)
        profile_rows = self._read_table_rows(TBL_CLIENT_PROFILES)
        _parent_names, client_parent_lookup = self._client_parent_lookup(profile_rows, client_rows, parent_rows)

        def parent_info_for_client(*values: Any) -> Dict[str, str]:
            for value in values:
                for key in _report_name_keys(value):
                    info = client_parent_lookup.get(key)
                    if info:
                        return info
            return {"id": "", "name": ""}

        open_rows: List[Dict[str, Any]] = []
        open_by_invoice: Dict[str, Dict[str, Any]] = {}
        closed_void_rows: List[Dict[str, Any]] = []
        summary_map: Dict[str, Dict[str, Any]] = {}
        bucket_map: Dict[str, Dict[str, Any]] = {
            "current": {"title": "0-30 days", "invoiceCount": 0, "amount": 0.0},
            "days31To60": {"title": "31-60 days", "invoiceCount": 0, "amount": 0.0},
            "days61To90": {"title": "61-90 days", "invoiceCount": 0, "amount": 0.0},
            "days90Plus": {"title": "90+ days", "invoiceCount": 0, "amount": 0.0},
        }
        stale_closed_balance = 0.0
        closed_statuses = {"void", "cancelled", "canceled", "paid", "closed"}

        for raw in receivable_rows:
            invoice = _clean_text(raw.get(sc.COL_RECV_INVOICE_NUM))
            if not invoice or not actual_invoice_pattern.match(invoice) or invoice.upper() in excluded_invoices:
                continue
            balance = round(amount(raw.get(sc.COL_RECV_BALANCE_DUE)), 2)
            if abs(balance) <= 0.01:
                continue
            status = _clean_text(raw.get(sc.COL_RECV_STATUS))
            status_key = status.lower()
            billing_client = _clean_text(raw.get(sc.COL_RECV_CLIENT))
            work_client = _clean_text(raw.get(sc.COL_RECV_WORK_CLIENT))
            display_client = work_client or billing_client
            parent_info = parent_info_for_client(work_client, billing_client)
            billing_parent_id = parent_info.get("id", "")
            billing_parent_name = parent_info.get("name", "")
            resolved_billing_client = billing_parent_name or billing_client or display_client

            if status_key in closed_statuses:
                stale_closed_balance += balance
                invoice_date = self._parse_date_value(raw.get(sc.COL_RECV_DATE))
                invoice_total = money_round(amount(raw.get(sc.COL_RECV_TOTAL_INVOICED)))
                paid = money_round(amount(raw.get(sc.COL_RECV_AMOUNT_PAID)) + amount(raw.get(sc.COL_RECV_CREDITS_ADJ)))
                closed_void_rows.append(
                    {
                        "invoice": invoice,
                        "date": invoice_date.isoformat() if invoice_date else self._date_iso(raw.get(sc.COL_RECV_DATE)),
                        "client": display_client,
                        "billingClient": resolved_billing_client,
                        "sourceBillingClient": billing_client,
                        "workClient": work_client,
                        "billingParentId": billing_parent_id,
                        "billingParentName": billing_parent_name,
                        "status": status,
                        "invoiceTotal": invoice_total,
                        "invoiceTotalDisplay": money(invoice_total),
                        "paid": paid,
                        "paidDisplay": money(paid),
                        "balance": balance,
                        "balanceDisplay": money(balance),
                        "reason": "Status is closed/paid/void, so this balance is excluded from collectible A/R.",
                    }
                )
                continue

            if query:
                haystack = " ".join([invoice, billing_client, resolved_billing_client, work_client, status]).lower()
                if query not in haystack:
                    continue

            invoice_date = self._parse_date_value(raw.get(sc.COL_RECV_DATE))
            age_days = max(0, (as_of - invoice_date).days) if invoice_date else 0
            bucket_id, bucket_label = bucket_key(age_days)
            invoice_total = round(amount(raw.get(sc.COL_RECV_TOTAL_INVOICED)), 2)
            paid = round(amount(raw.get(sc.COL_RECV_AMOUNT_PAID)) + amount(raw.get(sc.COL_RECV_CREDITS_ADJ)), 2)

            row = {
                "invoice": invoice,
                "date": invoice_date.isoformat() if invoice_date else self._date_iso(raw.get(sc.COL_RECV_DATE)),
                "client": display_client,
                "billingClient": resolved_billing_client,
                "sourceBillingClient": billing_client,
                "workClient": work_client,
                "billingParentId": billing_parent_id,
                "billingParentName": billing_parent_name,
                "matter": "",
                "status": status,
                "ageDays": age_days,
                "bucket": bucket_id,
                "bucketLabel": bucket_label,
                "invoiceTotal": invoice_total,
                "invoiceTotalDisplay": money(invoice_total),
                "total": invoice_total,
                "paid": paid,
                "paidDisplay": money(paid),
                "credits": round(amount(raw.get(sc.COL_RECV_CREDITS_ADJ)), 2),
                "balance": balance,
                "balanceDisplay": money(balance),
                "balanceNet": net_of_hst(balance),
                "balanceNetDisplay": money(net_of_hst(balance)),
            }
            open_rows.append(row)
            open_by_invoice[invoice] = row

            group_key = (
                resolved_billing_client
                if group_by == "billingClient"
                else (display_client or resolved_billing_client)
            ) or "(Unassigned)"
            summary = summary_map.setdefault(
                group_key,
                {
                    "client": group_key,
                    "billingClient": resolved_billing_client,
                    "_childClients": set(),
                    "invoiceCount": 0,
                    "oldestAgeDays": 0,
                    "current": 0.0,
                    "days31To60": 0.0,
                    "days61To90": 0.0,
                    "days90Plus": 0.0,
                    "balance": 0.0,
                },
            )
            child_client = display_client or billing_client
            if child_client:
                summary["_childClients"].add(child_client)
            summary["invoiceCount"] += 1
            summary["oldestAgeDays"] = max(int(summary["oldestAgeDays"]), age_days)
            summary[bucket_id] = round(float(summary[bucket_id]) + balance, 2)
            summary["balance"] = round(float(summary["balance"]) + balance, 2)

            bucket_map[bucket_id]["invoiceCount"] += 1
            bucket_map[bucket_id]["amount"] = round(float(bucket_map[bucket_id]["amount"]) + balance, 2)

        if group_by == "billingClient":
            open_rows.sort(
                key=lambda row: (
                    _clean_text(row.get("billingClient")).lower(),
                    _clean_text(row.get("client")).lower(),
                    -int(row.get("ageDays") or 0),
                    row.get("invoice"),
                )
            )
        else:
            open_rows.sort(key=lambda row: (_clean_text(row.get("client")).lower(), -int(row.get("ageDays") or 0), row.get("invoice")))

        summary_rows = list(summary_map.values())
        for row in summary_rows:
            child_clients = sorted(row.get("_childClients") or [], key=lambda value: _clean_text(value).lower())
            row["childClientCount"] = len(child_clients)
            row["childClients"] = ", ".join(child_clients)
            row.pop("_childClients", None)
            for key in ("current", "days31To60", "days61To90", "days90Plus", "balance"):
                row[key] = round(float(row.get(key) or 0.0), 2)
            row["currentDisplay"] = money(row["current"])
            row["days31To60Display"] = money(row["days31To60"])
            row["days61To90Display"] = money(row["days61To90"])
            row["days90PlusDisplay"] = money(row["days90Plus"])
            row["balanceDisplay"] = money(row["balance"])
        summary_rows.sort(key=lambda row: (-float(row.get("balance") or 0.0), _clean_text(row.get("client")).lower()))

        bucket_rows = []
        for key in ("current", "days31To60", "days61To90", "days90Plus"):
            bucket = dict(bucket_map[key])
            bucket["amount"] = round(float(bucket.get("amount") or 0.0), 2)
            bucket["amountDisplay"] = money(bucket["amount"])
            bucket_rows.append(bucket)

        total_ar = round(sum(float(row.get("balance") or 0.0) for row in open_rows), 2)
        total_net = net_of_hst(total_ar) if total_ar else 0.0
        hst_component = money_round(total_ar - total_net)
        closed_void_rows.sort(key=lambda row: (-float(row.get("balance") or 0.0), _clean_text(row.get("client")).lower(), row.get("invoice")))

        def dashboard_ledger_positive_by_ref(selected_year: int) -> Tuple[Dict[str, Dict[str, Any]], float]:
            def in_scope(row: Dict[str, Any]) -> bool:
                ref = _clean_text(row.get(sc.COL_LEDGER_REFERENCE))
                if not ref:
                    return False
                year_from_ref = ref_year(ref)
                if year_from_ref is not None:
                    return year_from_ref in {selected_year - 1, selected_year}
                parsed_year = row_year(row.get(sc.COL_LEDGER_DATE))
                return parsed_year in {selected_year - 1, selected_year}

            primary_client_by_ref: Dict[str, str] = {}
            for row in ledger_rows:
                if not in_scope(row):
                    continue
                ref = _clean_text(row.get(sc.COL_LEDGER_REFERENCE))
                if amount(row.get(sc.COL_LEDGER_BILLINGS_EXCL_HST)) > 0.0001:
                    primary_client_by_ref.setdefault(
                        ref,
                        normalize_ledger_client(row.get(sc.COL_LEDGER_CLIENT_VENDOR)) or "(Unassigned)",
                    )

            balances: Dict[Tuple[str, str], float] = {}
            for row in ledger_rows:
                if not in_scope(row):
                    continue
                ref = _clean_text(row.get(sc.COL_LEDGER_REFERENCE))
                client = primary_client_by_ref.get(ref) or normalize_ledger_client(row.get(sc.COL_LEDGER_CLIENT_VENDOR)) or "(Unassigned)"
                key = (ref, client)
                balances[key] = balances.get(key, 0.0) + amount(row.get(sc.COL_LEDGER_RECEIVABLE))

            by_ref: Dict[str, Dict[str, Any]] = {}
            for (ref, client), balance in balances.items():
                rounded_balance = money_round(balance)
                if rounded_balance < 0.01:
                    continue
                existing = by_ref.setdefault(ref, {"reference": ref, "client": client, "amount": 0.0})
                existing["amount"] = money_round(float(existing["amount"]) + rounded_balance)
            total = money_round(sum(float(row.get("amount") or 0.0) for row in by_ref.values()))
            return by_ref, total

        issue_rows: List[Dict[str, Any]] = []
        legacy_reconciliation_rows: List[Dict[str, Any]] = []
        legacy_ledger_by_ref, legacy_ledger_ar = dashboard_ledger_positive_by_ref(as_of.year)
        non_invoice_ledger_ar = 0.0
        if not query:
            for reference, ledger_row in sorted(legacy_ledger_by_ref.items()):
                ledger_amount = money_round(ledger_row.get("amount"))
                if not actual_invoice_pattern.match(reference):
                    non_invoice_ledger_ar = money_round(non_invoice_ledger_ar + ledger_amount)
                    legacy_reconciliation_rows.append(
                        {
                            "type": "Legacy ledger-only reference",
                            "reference": reference,
                            "client": ledger_row.get("client", ""),
                            "status": "Accounted",
                            "amount": ledger_amount,
                            "amountDisplay": money(ledger_amount),
                            "note": "Provider/vendor reference is retained for legacy audit, but collectible A/R is governed by client invoice rows in Receivables.",
                        }
                    )
                    continue
                if reference not in open_by_invoice:
                    if ledger_amount <= 0.01:
                        continue
                    legacy_reconciliation_rows.append(
                        {
                            "type": "Legacy ledger invoice not open",
                            "reference": reference,
                            "client": ledger_row.get("client", ""),
                            "status": "Accounted",
                            "amount": ledger_amount,
                            "amountDisplay": money(ledger_amount),
                            "note": "Excluded from collectible A/R because Receivables does not show this invoice as open; payments, discounts, corrections, and closed/void status reduce the payable balance.",
                        }
                    )

            for reference, receivable_row in sorted(open_by_invoice.items()):
                receivable_amount = money_round(receivable_row.get("balance"))
                ledger_row = legacy_ledger_by_ref.get(reference)
                if ledger_row is None:
                    legacy_reconciliation_rows.append(
                        {
                            "type": "Open receivable governed by Receivables",
                            "reference": reference,
                            "client": receivable_row.get("client", ""),
                            "status": "Accounted",
                            "amount": receivable_amount,
                            "amountDisplay": money(receivable_amount),
                            "note": "Included in collectible A/R because Receivables shows the client invoice as open, even if the legacy ledger reference is missing or zero.",
                        }
                    )
                    continue
                ledger_amount = money_round(ledger_row.get("amount"))
                delta = money_round(ledger_amount - receivable_amount)
                if abs(delta) > 0.01:
                    legacy_reconciliation_rows.append(
                        {
                            "type": "Legacy ledger/Receivables amount mismatch",
                            "reference": reference,
                            "client": receivable_row.get("client", ""),
                            "status": "Accounted",
                            "amount": delta,
                            "amountDisplay": money(delta),
                            "note": f"Legacy ledger A/R {money(ledger_amount)} versus open Receivables A/R {money(receivable_amount)}; Receivables governs collectible A/R.",
                        }
                    )
        else:
            legacy_ledger_ar = total_ar

        legacy_ledger_difference = money_round(legacy_ledger_ar - total_ar)
        summary = {
            "totalAr": total_ar,
            "totalArNet": total_net,
            "totalHstComponent": hst_component,
            "invoiceCount": len(open_rows),
            "clientCount": len(summary_rows),
            "groupBy": group_by,
            "ledgerAr": total_ar,
            "dashboardLedgerAr": total_ar,
            "legacyLedgerAr": legacy_ledger_ar,
            "rawLegacyLedgerAr": legacy_ledger_ar,
            "openReceivablesBalance": total_ar,
            "ledgerDifference": 0.0,
            "dashboardLedgerDifference": 0.0,
            "legacyLedgerDifference": legacy_ledger_difference,
            "rawLegacyLedgerDifference": legacy_ledger_difference,
            "nonInvoiceLedgerAr": money_round(non_invoice_ledger_ar),
            "staleClosedBalance": round(stale_closed_balance, 2),
            "closedVoidCount": len(closed_void_rows),
            "issueCount": len(issue_rows),
            "legacyReconciliationCount": len(legacy_reconciliation_rows),
        }
        cards = [
            {"label": "Total Gross A/R", "value": total_ar, "displayValue": money(total_ar), "tone": "primary"},
            {"label": "Total Net A/R", "value": total_net, "displayValue": money(total_net), "tone": "success"},
            {"label": "Open Invoices", "value": len(open_rows), "displayValue": str(len(open_rows)), "tone": "info"},
            {
                "label": "Open Billing Clients" if group_by == "billingClient" else "Open Clients",
                "value": len(summary_rows),
                "displayValue": str(len(summary_rows)),
                "tone": "success",
            },
            {
                "label": "Excluded Closed/Void",
                "value": round(stale_closed_balance, 2),
                "displayValue": money(round(stale_closed_balance, 2)),
                "tone": "warning",
                "actionTable": "closedVoid",
            },
        ]

        return {
            "ok": True,
            "reportId": "ar_aging",
            "title": "A/R Aging & Detail",
            "asOfDate": as_of.isoformat(),
            "summary": summary,
            "cards": cards,
            "rows": open_rows,
            "groupBy": group_by,
            "closedVoidRows": closed_void_rows,
            "summaryRows": summary_rows,
            "bucketRows": bucket_rows,
            "issueRows": issue_rows,
            "legacyReconciliationRows": legacy_reconciliation_rows,
            "notes": [
                "A/R detail is generated from open actual invoice rows in tblReceivables.",
                "Total Net A/R removes estimated 13% HST from the open gross Receivables balance.",
                "Provider/vendor disbursement references, discounts, corrections, and closed/void balances are retained for audit but excluded from headline collectible A/R unless Receivables shows an open client invoice balance.",
            ],
            "message": "",
        }
    def export_ar_aging_csv(self, payload: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        raw_payload = dict(payload or {})
        filters = raw_payload.get("filters") if isinstance(raw_payload.get("filters"), dict) else raw_payload
        report = self.ar_aging_report(dict(filters or {}))
        if not report.get("ok"):
            return {
                "ok": False,
                "path": "",
                "filename": "",
                "rowCount": 0,
                "message": _clean_text(report.get("message")) or "Could not build A/R aging report.",
            }

        rows = report.get("rows") if isinstance(report.get("rows"), list) else []
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        candidate_dirs = [
            self.paths.exports_dir(),
            self.paths.data_dir() / "exports",
            self.paths.root / "outputs",
        ]
        output_path: Optional[Path] = None
        last_error: Optional[Exception] = None
        for export_dir in candidate_dirs:
            try:
                export_dir.mkdir(parents=True, exist_ok=True)
                candidate = export_dir / f"ar_aging_{stamp}.csv"
                with candidate.open("w", encoding="utf-8-sig", newline="") as handle:
                    writer = csv.writer(handle)
                    writer.writerow([
                        "Invoice",
                        "Date",
                        "Billing Client",
                        "Client",
                        "Status",
                        "Age",
                        "Bucket",
                        "Invoice Total",
                        "Paid",
                        "Balance",
                    ])
                    for row in rows:
                        writer.writerow([
                            row.get("invoice", ""),
                            row.get("date", ""),
                            row.get("billingClient", ""),
                            row.get("client", ""),
                            row.get("status", ""),
                            row.get("ageDays", 0),
                            row.get("bucketLabel", ""),
                            row.get("invoiceTotal", 0.0),
                            row.get("paid", 0.0),
                            row.get("balance", 0.0),
                        ])
                output_path = candidate
                break
            except Exception as exc:
                last_error = exc
                continue

        if output_path is None:
            return {
                "ok": False,
                "path": "",
                "filename": "",
                "rowCount": 0,
                "message": f"Could not export CSV: {last_error}",
            }

        return {
            "ok": True,
            "path": str(output_path),
            "filename": output_path.name,
            "rowCount": len(rows),
            "message": f"CSV exported: {output_path.name}",
        }

    def financial_dashboard_report(self, year: Optional[int] = None) -> Dict[str, Any]:
        selected_year = int(year or datetime.now().year)
        self.ensure_schema()
        time_rows = self._read_table_rows(TBL_TIME)
        invoice_rows = self._read_table_rows(TBL_INVOICE_LOG)
        transaction_rows = self._read_table_rows(TBL_TRANSACTIONS_MASTER)
        legacy_docket_rows = self._read_raw_excel_table_rows("Dockets", "tblDockets")
        actual_invoice_pattern = re.compile(r"^\d{2}-\d{4}(?:-[A-Z])?$")

        def amount(value: Any) -> float:
            return float(self._parse_float(value) or 0.0)

        def money_round(value: Any) -> float:
            try:
                return float(Decimal(str(value or 0.0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
            except Exception:
                return 0.0

        def row_year(value: Any) -> Optional[int]:
            parsed = self._parse_date_value(value)
            return parsed.year if parsed else None

        def quarter_label(value: Any) -> Optional[str]:
            parsed = self._parse_date_value(value)
            if not parsed or parsed.year != selected_year:
                return None
            return f"Q{((parsed.month - 1) // 3) + 1}"

        def normalize_ledger_client(value: Any) -> str:
            text = _clean_text(value)
            text = re.sub(r"\s*\{.*?\}\s*$", "", text).strip()
            return re.sub(r"\s+", " ", text)

        def invoice_marker_bucket(value: Any) -> str:
            marker = _clean_text(value).upper()
            if not marker:
                return "blank"
            if actual_invoice_pattern.match(marker):
                return "invoice"
            if "HOLD" in marker:
                return "hold"
            if "NO" in marker or "FREE" in marker:
                return "no_bill"
            return "other"

        def docket_client(row: Dict[str, Any]) -> str:
            return _clean_text(row.get("Client")) or _clean_text(row.get("Parent")) or "(Unassigned)"

        def docket_amount_to_cs(row: Dict[str, Any]) -> float:
            """Return 'Amount to CS', computing from components if the cached formula value is None."""
            cached = amount(row.get("Amount to CS"))
            if cached != 0.0:
                return cached
            # Formula: Time × Rate × Percentage (where Percentage is 0-100 scale, e.g. 88.5)
            hrs = amount(row.get("Time (in hrs) or Units") or row.get("Time (in hrs)"))
            rate = amount(row.get("Hourly Rate/Flat Rate") or row.get("Hourly Rate/Flat Fee"))
            pct = amount(row.get("Percentage"))
            if hrs > 0 and rate > 0 and pct > 0:
                return round(hrs * rate * (pct / 100.0), 2)
            return 0.0

        quarters: Dict[str, Dict[str, Any]] = {
            f"Q{i}": {"quarter": f"Q{i}", "revenue": 0.0, "expenses": 0.0, "hstCollected": 0.0, "hstPaid": 0.0, "netHst": 0.0}
            for i in range(1, 5)
        }

        docketed_amount = 0.0
        wip_hours = 0.0
        wip_amount = 0.0
        work_totals: Dict[str, float] = {}
        billing_totals: Dict[str, float] = {}
        billed_time_amount = 0.0
        ledger_billings = 0.0
        expenses_total = 0.0
        hst_collected_total = 0.0
        hst_paid_total = 0.0
        banked_total = 0.0

        # Pass 1 & 2: Pre-calculate Realization Rates per Invoice
        ledger_billings_by_invoice: Dict[str, float] = {}
        for row in transaction_rows:
            ref = _clean_text(row.get(sc.COL_TXN_INVOICE_REF)).upper()
            if actual_invoice_pattern.match(ref):
                row_type = row.get(sc.COL_TXN_TYPE)
                if row_type == "Income":
                    ledger_billings_by_invoice[ref] = ledger_billings_by_invoice.get(ref, 0.0) + amount(row.get(sc.COL_TXN_AMOUNT))

        docket_fees_by_invoice: Dict[str, float] = {}
        if legacy_docket_rows:
            for row in legacy_docket_rows:
                ref = _clean_text(row.get("Invoice #")).upper()
                if actual_invoice_pattern.match(ref):
                    docket_fees_by_invoice[ref] = docket_fees_by_invoice.get(ref, 0.0) + docket_amount_to_cs(row)
        else:
            for row in time_rows:
                ref = _clean_text(row.get(sc.COL_TIME_INVOICE_REF)).upper()
                if actual_invoice_pattern.match(ref):
                    gross = amount(row.get(sc.COL_TIME_NET))
                    docket_fees_by_invoice[ref] = docket_fees_by_invoice.get(ref, 0.0) + gross

        for row in transaction_rows:
            label = quarter_label(row.get(sc.COL_TXN_DATE))
            if not label:
                continue
            
            cls = row.get(sc.COL_TXN_CLASS)
            typ = row.get(sc.COL_TXN_TYPE)
            amt = amount(row.get(sc.COL_TXN_AMOUNT))
            tax = amount(row.get(sc.COL_TXN_TAX_AMOUNT))
            
            bucket = quarters[label]
            if typ == "Income":
                ledger_billings += amt
                hst_collected_total += tax
                bucket["revenue"] += amt
                bucket["hstCollected"] += tax
            elif typ == "Expense":
                expenses_total += amt
                hst_paid_total += tax
                bucket["expenses"] += amt
                bucket["hstPaid"] += tax
            elif typ == "Transfer":
                # A/R set-offs clear a receivable and A/P together without
                # cash entering a bank account.  They remain fully visible in
                # the ledger and A/R audit trail but must not inflate the
                # Executive Dashboard's cash/banked card.
                from_account = _clean_text(row.get(sc.COL_TXN_FROM_ACCOUNT)).upper()
                to_account = _clean_text(row.get(sc.COL_TXN_TO_ACCOUNT)).upper()
                if from_account != "AR_SET_OFF" and to_account != "AR_SET_OFF":
                    banked_total += amt
            elif typ == "Adjustment":
                ledger_billings -= amt
                bucket["revenue"] -= amt
            
            # Note: client mapping could be enhanced if needed
            client_key = normalize_ledger_client(row.get(sc.COL_TXN_CLIENT))
            if client_key and typ == "Income":
                billing_totals[client_key] = billing_totals.get(client_key, 0.0) + amt

        if legacy_docket_rows:
            for row in legacy_docket_rows:
                if row_year(row.get("Date")) != selected_year:
                    continue
                client_key = docket_client(row)
                ref = _clean_text(row.get("Invoice #")).upper()
                bucket = invoice_marker_bucket(ref)
                raw_production = docket_amount_to_cs(row)
                hours = amount(row.get("Time (in hrs) or Units"))
                
                if bucket == "no_bill":
                    production = 0.0
                elif bucket == "invoice" and actual_invoice_pattern.match(ref):
                    invoice_billings = ledger_billings_by_invoice.get(ref, 0.0)
                    invoice_dockets = docket_fees_by_invoice.get(ref, 0.0)
                    if invoice_dockets > 0.0001:
                        realization_rate = invoice_billings / invoice_dockets
                        production = raw_production * realization_rate
                    else:
                        production = 0.0
                else:
                    production = raw_production

                docketed_amount += production
                work_totals[client_key] = work_totals.get(client_key, 0.0) + production
                if bucket in ("blank", "hold"):
                    wip_amount += production
                    wip_hours += hours
                    billing_totals[client_key] = billing_totals.get(client_key, 0.0) + production
                    label = quarter_label(row.get("Date"))
                    if label:
                        quarters[label]["revenue"] += production
        else:
            for row in time_rows:
                if row_year(row.get(sc.COL_TIME_DATE)) != selected_year:
                    continue
                gross = amount(row.get(sc.COL_TIME_NET))
                hours = amount(row.get(sc.COL_TIME_HOURS))
                status = _clean_text(row.get(sc.COL_TIME_STATUS)).lower()
                
                ref = _clean_text(row.get(sc.COL_TIME_INVOICE_REF)).upper()
                if status in {"billed", "merged"} and actual_invoice_pattern.match(ref):
                    invoice_billings = ledger_billings_by_invoice.get(ref, 0.0)
                    invoice_dockets = docket_fees_by_invoice.get(ref, 0.0)
                    if invoice_dockets > 0.0001:
                        realization_rate = invoice_billings / invoice_dockets
                        gross = gross * realization_rate
                    else:
                        gross = 0.0
                
                docketed_amount += gross
                if status not in {"billed", "merged"}:
                    wip_hours += hours
                    wip_amount += gross
                    label = quarter_label(row.get(sc.COL_TIME_DATE))
                    if label:
                        quarters[label]["revenue"] += gross
                client_key = _clean_text(row.get(sc.COL_TIME_CLIENT_ID)) or "(Unassigned)"
                work_totals[client_key] = work_totals.get(client_key, 0.0) + gross

            for row in invoice_rows:
                if row_year(row.get(sc.COL_INV_INVOICE_DATE)) != selected_year:
                    continue
                billed = amount(row.get(sc.COL_INV_AGGREGATE_BILLED))
                billed_time_amount += billed
                billing_client = _clean_text(row.get(sc.COL_INV_BILL_TO_CLIENT)) or _clean_text(row.get(sc.COL_INV_CLIENT_NAME)) or "(Unassigned)"
                work_client = _clean_text(row.get(sc.COL_INV_SUB_CLIENT)) or _clean_text(row.get(sc.COL_INV_CLIENT_NAME)) or billing_client
                billing_totals[billing_client] = billing_totals.get(billing_client, 0.0) + billed
                work_totals[work_client] = work_totals.get(work_client, 0.0) + billed

        if legacy_docket_rows:
            billed_time_amount = ledger_billings
        revenue_including_wip = money_round(ledger_billings + wip_amount)
        net_income_accrual = money_round(revenue_including_wip - expenses_total)
        banked_net_est = money_round(banked_total / 1.13) if banked_total else 0.0
        cash_net_income_est = money_round(banked_net_est - expenses_total)

        for bucket in quarters.values():
            for key in ("revenue", "expenses", "hstCollected", "hstPaid"):
                bucket[key] = money_round(bucket.get(key) or 0.0)
            bucket["netHst"] = money_round(float(bucket["hstCollected"]) - float(bucket["hstPaid"]))

        def dashboard_ar_from_ledger() -> Tuple[List[Dict[str, Any]], float, float, float]:
            selected_prefix = f"{selected_year % 100:02d}-"
            prior_prefix = f"{(selected_year - 1) % 100:02d}-"

            def ref_year(ref: str) -> Optional[int]:
                if actual_invoice_pattern.match(ref):
                    try:
                        return 2000 + int(ref[:2])
                    except Exception:
                        return None
                return None

            def in_scope(row: Dict[str, Any]) -> bool:
                ref = _clean_text(row.get(sc.COL_TXN_INVOICE_REF))
                if not ref:
                    return False
                year_from_ref = ref_year(ref)
                if year_from_ref is not None:
                    return year_from_ref in {selected_year - 1, selected_year}
                parsed = self._parse_date_value(row.get(sc.COL_TXN_DATE))
                return bool(parsed and parsed.year in {selected_year - 1, selected_year})

            primary_client_by_ref: Dict[str, str] = {}
            for row in transaction_rows:
                if not in_scope(row):
                    continue
                ref = _clean_text(row.get(sc.COL_TXN_INVOICE_REF))
                row_type = row.get(sc.COL_TXN_TYPE)
                if row_type == "Income" and amount(row.get(sc.COL_TXN_AMOUNT)) > 0.0001:
                    primary_client_by_ref.setdefault(
                        ref,
                        normalize_ledger_client(row.get(sc.COL_TXN_CLIENT)) or "(Unassigned)",
                    )

            balances_by_ref_client: Dict[Tuple[str, str], float] = {}
            for row in transaction_rows:
                if not in_scope(row):
                    continue
                ref = _clean_text(row.get(sc.COL_TXN_INVOICE_REF))
                client = primary_client_by_ref.get(ref) or normalize_ledger_client(row.get(sc.COL_TXN_CLIENT)) or "(Unassigned)"
                key = (ref, client)
                
                row_type = row.get(sc.COL_TXN_TYPE)
                amt = amount(row.get(sc.COL_TXN_AMOUNT))
                tax = amount(row.get(sc.COL_TXN_TAX_AMOUNT))
                
                receivable_delta = 0.0
                if row_type == "Income":
                    receivable_delta = amt + tax
                elif row_type == "Transfer":
                    receivable_delta = -amt
                elif row_type == "Adjustment":
                    # Write-offs reduce the receivable. amt is negative (e.g. -500), so we just add it
                    receivable_delta = amt
                
                balances_by_ref_client[key] = balances_by_ref_client.get(key, 0.0) + receivable_delta

            totals: Dict[str, float] = {}
            pre_year_total = 0.0
            selected_year_total = 0.0
            for (ref, client), balance in balances_by_ref_client.items():
                rounded_balance = money_round(balance)
                if rounded_balance < 0.01:
                    continue
                totals[client] = totals.get(client, 0.0) + rounded_balance
                if ref.startswith(selected_prefix):
                    selected_year_total += rounded_balance
                elif ref.startswith(prior_prefix):
                    pre_year_total += rounded_balance
                else:
                    selected_year_total += rounded_balance

            rows = [
                {"client": client, "amount": money_round(total)}
                for client, total in sorted(totals.items(), key=lambda item: float(item[1] or 0.0), reverse=True)
                if money_round(total) >= 0.01
            ]
            total = money_round(sum(row["amount"] for row in rows))
            return rows, money_round(pre_year_total), money_round(selected_year_total), total

        ar_report = self.ar_aging_report({"asOfDate": datetime.now().strftime("%Y-%m-%d")})
        legacy_ar_details, legacy_pre_year_ar, legacy_selected_year_ar, legacy_dashboard_total_ar = dashboard_ar_from_ledger()
        ar_rows = ar_report.get("rows") if isinstance(ar_report.get("rows"), list) else []
        ar_summary_rows = ar_report.get("summaryRows") if isinstance(ar_report.get("summaryRows"), list) else []
        dashboard_ar_details = [
            {"client": _clean_text(row.get("client")), "amount": money_round(row.get("balance") or 0.0)}
            for row in ar_summary_rows
            if money_round(row.get("balance") or 0.0) >= 0.01
        ]

        def ar_invoice_year(row: Dict[str, Any]) -> Optional[int]:
            invoice_ref = _clean_text(row.get("invoice"))
            if actual_invoice_pattern.match(invoice_ref):
                try:
                    return 2000 + int(invoice_ref[:2])
                except Exception:
                    return None
            return row_year(row.get("date"))

        pre_year_ar = 0.0
        selected_year_ar = 0.0
        for row in ar_rows:
            row_amount = money_round(row.get("balance") or 0.0)
            if row_amount < 0.01:
                continue
            invoice_year = ar_invoice_year(row)
            if invoice_year is not None and invoice_year < selected_year:
                pre_year_ar += row_amount
            else:
                selected_year_ar += row_amount
        pre_year_ar = money_round(pre_year_ar)
        selected_year_ar = money_round(selected_year_ar)
        dashboard_total_ar = money_round(pre_year_ar + selected_year_ar)

        def top_rows(source: Dict[str, float], limit: int = 10) -> List[Dict[str, Any]]:
            total = sum(max(0.0, float(value or 0.0)) for value in source.values())
            rows = []
            for client, value in sorted(source.items(), key=lambda item: float(item[1] or 0.0), reverse=True)[:limit]:
                value_float = money_round(value or 0.0)
                rows.append({
                    "client": client,
                    "amount": value_float,
                    "sharePct": round((value_float / total * 100.0), 1) if total else 0.0,
                })
            return rows

        collectible_ar = money_round((ar_report.get("summary") or {}).get("totalAr") or 0.0)
        summary = {
            "revenueIncludingWip": revenue_including_wip,
            "expenses": money_round(expenses_total),
            "netIncomeAccrual": net_income_accrual,
            "wipAmount": money_round(wip_amount),
            "bankedAmount": money_round(banked_total),
            "bankedNetEstimate": banked_net_est,
            "cashNetIncomeEstimate": cash_net_income_est,
            "preYearAr": pre_year_ar,
            "selectedYearAr": selected_year_ar,
            "dashboardAr": dashboard_total_ar,
            "collectibleAr": collectible_ar,
            "legacyPreYearAr": legacy_pre_year_ar,
            "legacySelectedYearAr": legacy_selected_year_ar,
            "legacyDashboardAr": legacy_dashboard_total_ar,
            "legacyDashboardArDifference": money_round(legacy_dashboard_total_ar - dashboard_total_ar),
            "docketedAmount": money_round(docketed_amount),
            "billedTimeAmount": money_round(billed_time_amount),
            "wipHours": round(wip_hours, 1),
            "totalAr": dashboard_total_ar,
        }
        cards = [
            {"label": "Revenue (Inc. WIP)", "value": revenue_including_wip, "subvalue": str(selected_year), "tone": "success"},
            {"label": f"Expenses ({selected_year})", "value": money_round(expenses_total), "subvalue": "Ledger expenses", "tone": "warning"},
            {"label": "Net Income (Accrual)", "value": net_income_accrual, "subvalue": "Revenue less expenses", "tone": "success"},
            {"label": f"WIP ({selected_year})", "value": money_round(wip_amount), "subvalue": f"{round(wip_hours, 1)} hours", "tone": "primary"},
            {"label": f"Banked ({selected_year})", "value": money_round(banked_total), "subvalue": f"Est. net {banked_net_est:,.0f}", "tone": "info"},
            {"label": "Cash Net Inc (Est)", "value": cash_net_income_est, "subvalue": "Banked net less expenses", "tone": "success"},
            {"label": f"Pre-{selected_year} A/R", "value": pre_year_ar, "subvalue": "Open prior invoices", "tone": "warning"},
            {"label": f"{selected_year} A/R", "value": selected_year_ar, "subvalue": "Open current invoices", "tone": "warning"},
            {"label": f"Total A/R (End {selected_year})", "value": dashboard_total_ar, "subvalue": "Gross collectible A/R", "tone": "warning"},
        ]

        return {
            "ok": True,
            "year": selected_year,
            "asOfDate": datetime.now().strftime("%Y-%m-%d"),
            "summary": summary,
            "cards": cards,
            "quarters": [quarters[f"Q{i}"] for i in range(1, 5)],
            "arDetails": dashboard_ar_details[:20],
            "topBillingClients": top_rows(billing_totals),
            "topWorkClients": top_rows(work_totals),
            "notes": [
                "Revenue is summarized from tblLedger billings plus blank-invoice WIP rows from the imported legacy Dockets table.",
                "Financial dashboard A/R is governed by open client invoice balances in tblReceivables, matching the A/R aging report.",
                "Legacy positive-reference ledger A/R is retained in summary metadata for audit, not used as headline collectible A/R.",
            ],
            "message": "",
        }


    def _canonical_ar_ledger(self, payload: dict) -> dict:
        """
        Canonical account-balance engine.
        Processes TBL_RECEIVABLES (debits) and TBL_LEDGER (credits) to build a chronological AR stream.
        """
        payload = dict(payload or {})
        requested_client = _clean_text(payload.get("client") or payload.get("clientId"))
        requested_matter = _clean_text(payload.get("matter") or payload.get("matterId"))
        client_level = _clean_text(payload.get("client_level") or "work").lower()
        if client_level not in {"billing", "work"}:
            client_level = "work"
            
        start_date = self._parse_date_value(payload.get("startDate") or payload.get("fromDate"))
        end_date = self._parse_date_value(payload.get("endDate") or payload.get("toDate") or payload.get("asOfDate") or payload.get("as_of_date"))
        if not end_date:
            end_date = date.today()
            
        open_items_only = str(payload.get("openItemsOnly", "")).lower() == "true"

        try:
            self.ensure_schema()
            # A statement needs all of these related tables to build its
            # canonical A/R stream.  Reading them one by one re-opened and
            # parsed the entire macro workbook for every table on a cold
            # cache, which freezes the QML UI while a report tab is opening.
            # The bulk reader preserves the exact same row-cache semantics,
            # but obtains every missing table from one read-only workbook
            # opening instead.
            statement_table_refs = [
                TBL_RECEIVABLES,
                TBL_LEDGER,
                TBL_CLIENT_PROFILES,
                TBL_CLIENTS,
                TBL_PARENTS,
                TBL_MATTERS,
                TBL_TIME,
                TBL_DISBURSEMENTS,
                TBL_INVOICE_LOG,
            ]
            bulk_reader = getattr(self, "_read_table_rows_bulk", None)
            if callable(bulk_reader):
                statement_tables = bulk_reader(statement_table_refs)
            else:
                # Keep lightweight repository adapters and narrow unit-test
                # doubles compatible with the original single-table contract.
                statement_tables = {
                    tref.table: self._read_table_rows(tref)
                    for tref in statement_table_refs
                }
            receivable_rows = statement_tables.get(TBL_RECEIVABLES.table, [])
            ledger_rows = statement_tables.get(TBL_LEDGER.table, [])
            profile_rows = statement_tables.get(TBL_CLIENT_PROFILES.table, [])
            client_rows = statement_tables.get(TBL_CLIENTS.table, [])
            parent_rows = statement_tables.get(TBL_PARENTS.table, [])
            matter_rows = statement_tables.get(TBL_MATTERS.table, [])
            time_rows = statement_tables.get(TBL_TIME.table, [])
            disbursement_rows = statement_tables.get(TBL_DISBURSEMENTS.table, [])
            invoice_rows = statement_tables.get(TBL_INVOICE_LOG.table, [])
        except Exception as e:
            return {"ok": False, "message": f"Data load failed: {e}"}

        def _name_keys(value: Any) -> set[str]:
            text = _clean_text(value).lower()
            if not text:
                return set()
            def compact(candidate: str) -> str:
                return re.sub(r"[^a-z0-9]+", " ", candidate).strip()
            keys = {compact(text)}
            if "," in text:
                last, first = [part.strip() for part in text.split(",", 1)]
                if first and last:
                    keys.add(compact(f"{first} {last}"))
            return {key for key in keys if key}

        requested_keys = _name_keys(requested_client)
        _parent_names, client_parent_lookup = self._client_parent_lookup(profile_rows, client_rows, parent_rows)

        def _parent_info_for_client(*values: Any) -> Dict[str, str]:
            for value in values:
                for key in _report_name_keys(value):
                    info = client_parent_lookup.get(key)
                    if info:
                        return info
            return {"id": "", "name": ""}

        def _matches_client(value: Any) -> bool:
            return bool(requested_keys.intersection(_name_keys(value)))

        def _matches_parent(info: Dict[str, str]) -> bool:
            return _matches_client(info.get("id", "")) or _matches_client(info.get("name", ""))

        def _client_scope(billing_client: Any, work_client: Any) -> str:
            if not requested_client:
                return "any"
            billing_text = _clean_text(billing_client)
            work_text = _clean_text(work_client)
            billing_match = _matches_client(billing_text)
            work_match = _matches_client(work_text)
            parent_match = _matches_parent(_parent_info_for_client(work_text, billing_text))
            if client_level == "billing":
                return "billing" if billing_match or parent_match else ""
            if work_match:
                return "work"
            if billing_match or parent_match:
                return "billing"
            return ""

        def _row_matches_scope(row_billing_client: Any, row_work_client: Any) -> bool:
            if not requested_client:
                return True
            return bool(_client_scope(row_billing_client, row_work_client))

        client_name_by_id: Dict[str, str] = {}
        for row in client_rows:
            client_id = _clean_text(row.get(sc.COL_CLIENT_ID))
            client_name = _clean_text(row.get(sc.COL_CLIENT_NAME))
            if client_id and client_name:
                client_name_by_id[client_id.casefold()] = client_name

        invoice_date_by_key: Dict[str, Optional[date]] = {}
        for row in receivable_rows:
            invoice_key = _clean_text(row.get(sc.COL_RECV_INVOICE_NUM)).casefold()
            if invoice_key:
                invoice_date_by_key[invoice_key] = self._parse_date_value(row.get(sc.COL_RECV_DATE))

        matter_context_by_id: Dict[str, Dict[str, str]] = {}
        matter_contexts_by_client: Dict[str, List[Dict[str, Any]]] = collections.defaultdict(list)
        for row in matter_rows:
            matter_id = _clean_text(row.get(sc.COL_MATTER_ID))
            if not matter_id:
                continue
            matter_client_id = _clean_text(row.get(sc.COL_MATTER_CLIENT_ID))
            matter_client = _clean_text(row.get(sc.COL_MATTER_CLIENT_NAME))
            if not matter_client and matter_client_id:
                matter_client = client_name_by_id.get(matter_client_id.casefold(), matter_client_id)
            # The Description is the user-facing, plain-English matter name.
            # Legacy codes and internal names are intentionally last resorts.
            matter_label = (
                _clean_text(row.get(sc.COL_MATTER_DESCRIPTION))
                or _clean_text(row.get(sc.COL_MATTER_DISPLAY_NAME))
                or _clean_text(row.get(sc.COL_MATTER_NAME))
                or _clean_text(row.get(sc.COL_MATTER_NUMBER))
                or matter_id
            )
            context = {
                "client": matter_client,
                "matter": matter_label,
                # Preserve the stable key as well as the plain-English label.
                # Statement rows use this to open the exact Matter Profile 360
                # record rather than attempting a name-based lookup.
                "matterId": matter_id,
            }
            # Use the same deliberately conservative historic fallback as the
            # Invoice Directory: only one non-legacy matter for the actual
            # work client that was already open on the invoice date may fill a
            # completely missing matter link.
            if matter_client and "legacy" not in matter_label.casefold() and "unassigned" not in matter_label.casefold():
                client_contexts = matter_contexts_by_client[matter_client.casefold()]
                if not any(_clean_text(item.get("matter")) == matter_label for item in client_contexts):
                    client_contexts.append(
                        {
                            "matter": matter_label,
                            "matterId": matter_id,
                            "opened": self._parse_date_value(row.get(sc.COL_MATTER_OPEN_DATE)),
                        }
                    )
            # A legacy docket stores the human-facing MatterNumber, whereas
            # current time rows store the UUID MatterID.  Index both (plus the
            # older display/name forms) so a statement can resolve either
            # generation of records to the same client and description.
            for lookup_value in (
                matter_id,
                _clean_text(row.get(sc.COL_MATTER_NUMBER)),
                _clean_text(row.get(sc.COL_MATTER_DISPLAY_NAME)),
                _clean_text(row.get(sc.COL_MATTER_NAME)),
            ):
                lookup_key = _clean_text(lookup_value).casefold()
                if lookup_key:
                    matter_context_by_id.setdefault(lookup_key, context)

        invoice_contexts: Dict[str, List[Dict[str, str]]] = {}

        def _append_invoice_context(
            invoice_ref: Any,
            client_ref: Any = "",
            matter_ref: Any = "",
            matter_label: Any = "",
        ) -> None:
            invoice_key = _clean_text(invoice_ref).casefold()
            if not invoice_key:
                return
            source_client = _clean_text(client_ref)
            source_matter = _clean_text(matter_label)
            resolved = matter_context_by_id.get(_clean_text(matter_ref).casefold(), {})
            client_name = _clean_text(resolved.get("client")) or client_name_by_id.get(
                source_client.casefold(), source_client
            )
            matter_name = _clean_text(resolved.get("matter")) or source_matter
            if not client_name and not matter_name:
                return
            context = {
                "client": client_name,
                "matter": matter_name,
                "matterId": _clean_text(resolved.get("matterId")),
            }
            contexts = invoice_contexts.setdefault(invoice_key, [])
            for index, existing in enumerate(contexts):
                if _clean_text(existing.get("client")).casefold() != client_name.casefold():
                    continue
                existing_matter = _clean_text(existing.get("matter"))
                if existing_matter == matter_name:
                    # Prefer a current structured lookup when the same
                    # legacy/invoice-log label lacked its durable Matter_ID.
                    if (not _clean_text(existing.get("matterId"))
                            and _clean_text(context.get("matterId"))):
                        contexts[index] = context
                    return
                if not existing_matter and matter_name:
                    contexts[index] = context
                    return
                if existing_matter and not matter_name:
                    return
            if context not in contexts:
                contexts.append(context)

        # Prefer current structured sources.  These identify both the actual
        # work client and its matter, whereas tblReceivables only identifies
        # the bill-to party.
        for row in time_rows:
            _append_invoice_context(
                row.get(sc.COL_TIME_INVOICE_REF),
                row.get(sc.COL_TIME_CLIENT_ID),
                row.get(sc.COL_TIME_MATTER_ID),
            )
        for row in disbursement_rows:
            _append_invoice_context(
                row.get(sc.COL_DISB_INVOICE_REF),
                row.get(sc.COL_DISB_CLIENT_ID) or row.get(sc.COL_DISB_SUB_CLIENT) or row.get(sc.COL_DISB_CLIENT_NAME),
                row.get(sc.COL_DISB_MATTER_ID),
            )
        # Invoice-log sub-client is a reliable secondary source when no work
        # row was retained.  It is deliberately not replaced with the bill-to
        # client below.
        for row in invoice_rows:
            _append_invoice_context(
                row.get(sc.COL_INV_INVOICE_NUM),
                row.get(sc.COL_INV_SUB_CLIENT) or row.get(sc.COL_INV_CLIENT_NAME),
            )
        # Preserve historical context stored only in the companion legacy
        # docket workbook, but only when the requested account actually has
        # an invoice whose modern records cannot identify a matter.  Opening
        # that OneDrive workbook can take many seconds; doing so for every
        # Statement or A/R tab made normal navigation appear frozen even when
        # current CSPM records already supplied every label.
        invoice_pattern = re.compile(r"^\d{2}-\d{4}(?:-[A-Z])?$")
        requires_legacy_context = False
        for row in receivable_rows:
            invoice_key = _clean_text(row.get(sc.COL_RECV_INVOICE_NUM)).casefold()
            if not invoice_key or not invoice_pattern.match(invoice_key.upper()):
                continue
            if not _row_matches_scope(
                row.get(sc.COL_RECV_CLIENT),
                row.get(sc.COL_RECV_WORK_CLIENT),
            ):
                continue
            known_contexts = invoice_contexts.get(invoice_key, [])
            if not any(_clean_text(context.get("matter")) for context in known_contexts):
                requires_legacy_context = True
                break

        if requires_legacy_context:
            for row in self._read_legacy_docket_rows():
                legacy_invoice_ref = (
                    row.get("Invoice") or row.get("InvoiceRef") or row.get("Invoice #")
                )
                legacy_invoice_key = _clean_text(legacy_invoice_ref).casefold()
                # A legacy Dockets row is fallback-only.  It may describe an
                # invoice that was subsequently reversed and reissued against
                # a different current matter.  Never append that historical
                # context when a live time/disbursement row has already
                # resolved a matter for the same invoice.
                existing_contexts = invoice_contexts.get(legacy_invoice_key, [])
                if any(_clean_text(context.get("matter")) for context in existing_contexts):
                    continue
                _append_invoice_context(
                    legacy_invoice_ref,
                    row.get("Sub-Client") or row.get("SubClient") or row.get("Client"),
                    row.get("Matter_ID") or row.get("MatterID") or row.get("Matter"),
                )

        for invoice_key, contexts in invoice_contexts.items():
            invoice_date = invoice_date_by_key.get(invoice_key)
            if not invoice_date:
                continue
            for index, context in enumerate(contexts):
                if _clean_text(context.get("matter")):
                    continue
                client_name = _clean_text(context.get("client"))
                if not client_name:
                    continue
                candidates = [
                    item
                    for item in matter_contexts_by_client.get(client_name.casefold(), [])
                    if not item.get("opened") or item["opened"] <= invoice_date
                ]
                if len(candidates) == 1:
                    contexts[index] = {
                        "client": client_name,
                        "matter": _clean_text(candidates[0].get("matter")),
                        "matterId": _clean_text(candidates[0].get("matterId")),
                    }

        client_by_invoice = {}
        for r in receivable_rows:
            inv = _clean_text(r.get(sc.COL_RECV_INVOICE_NUM))
            if inv:
                client_by_invoice[inv] = (
                    r.get(sc.COL_RECV_CLIENT),
                    r.get(sc.COL_RECV_WORK_CLIENT),
                    invoice_contexts.get(inv.casefold(), []),
                )
        
        # Build chronological events
        events = []
        
        # 1. Debits (Invoices)
        actual_invoice_pattern = re.compile(r"^\d{2}-\d{4}(?:-[A-Z])?$")
        for row in receivable_rows:
            invoice = _clean_text(row.get(sc.COL_RECV_INVOICE_NUM))
            if not invoice or not actual_invoice_pattern.match(invoice):
                continue
            
            billing_c = _clean_text(row.get(sc.COL_RECV_CLIENT))
            work_c = _clean_text(row.get(sc.COL_RECV_WORK_CLIENT))
            if requested_client and not _row_matches_scope(billing_c, work_c):
                continue
            
            # NOTE: Matter filtering would happen here if requested_matter was used.
            # Currently Receivables table does not contain matter ID. We assume client-level AR for now.
            
            invoice_date = self._parse_date_value(row.get(sc.COL_RECV_DATE))
            if invoice_date and invoice_date > end_date:
                continue
                
            amount = self._money_round(row.get(sc.COL_RECV_TOTAL_INVOICED))
            if amount <= 0:
                continue

            # Receivables is the authoritative source for the client-facing
            # invoice balance.  ``CreditsAdj`` can contain a private legacy
            # rounding/write-off adjustment (for example, a one-cent tidy-up),
            # not a payment or a credit the billing client received.  Keep the
            # gross amount in the internal AR event, but carry a separate
            # statement view that absorbs that adjustment into the displayed
            # invoice amount instead of exposing it in "Paid / Credits".
            statement_values_available = all(
                key in row
                for key in (
                    sc.COL_RECV_TOTAL_INVOICED,
                    sc.COL_RECV_AMOUNT_PAID,
                    sc.COL_RECV_CREDITS_ADJ,
                    sc.COL_RECV_BALANCE_DUE,
                )
            )
            statement_invoice_total = amount
            statement_payment_amount = 0.0
            statement_balance_due = amount
            if statement_values_available:
                adjustment = self._money_round(row.get(sc.COL_RECV_CREDITS_ADJ))
                statement_invoice_total = self._money_round(max(0.0, amount - adjustment))
                statement_payment_amount = self._money_round(row.get(sc.COL_RECV_AMOUNT_PAID))
                statement_balance_due = self._money_round(row.get(sc.COL_RECV_BALANCE_DUE))
                
            events.append({
                "date": invoice_date or date.min,
                "date_iso": invoice_date.isoformat() if invoice_date else "0000-00-00",
                "reference": invoice,
                "invoice": invoice,
                "type": "Invoice",
                "description": f"Invoice {invoice}",
                "debit": amount,
                "credit": 0.0,
                "status": _clean_text(row.get(sc.COL_RECV_STATUS)),
                "billingClient": billing_c,
                "workClient": work_c,
                "workContexts": invoice_contexts.get(invoice.casefold(), []),
                "statementValuesAvailable": statement_values_available,
                "statementInvoiceTotal": statement_invoice_total,
                "statementPaymentAmount": statement_payment_amount,
                "statementBalanceDue": statement_balance_due,
            })
            
        # 2. Credits (Payments & Adjustments)
        for row in ledger_rows:
            invoice = _clean_text(row.get(sc.COL_LEDGER_REFERENCE))
            if not invoice or not actual_invoice_pattern.match(invoice):
                continue
                
            client_tuple = client_by_invoice.get(invoice)
            if not client_tuple:
                continue
            
            if requested_client and not _row_matches_scope(client_tuple[0], client_tuple[1]):
                continue
                
            ledger_date = self._parse_date_value(row.get(sc.COL_LEDGER_DATE))
            if ledger_date and ledger_date > end_date:
                continue
                
            collected = self._money_round(row.get(sc.COL_LEDGER_COLLECTED))
            write_off = self._money_round(row.get(sc.COL_LEDGER_WRITE_OFF))
            total_credit = collected + write_off
            if total_credit <= 0:
                continue
                
            desc = _clean_text(row.get(sc.COL_LEDGER_DESCRIPTION)) or "Payment"
            method = _clean_text(row.get(sc.COL_LEDGER_CATEGORY))
            if method and method not in desc:
                desc = f"{desc} ({method})"
                
            events.append({
                "date": ledger_date or date.min,
                "date_iso": ledger_date.isoformat() if ledger_date else "0000-00-00",
                "reference": invoice,
                "invoice": invoice,
                "type": "Payment" if collected > 0 else "Credit/Adj",
                "description": desc,
                "debit": 0.0,
                "credit": total_credit,
                "status": "Applied",
                "billingClient": client_tuple[0],
                "workClient": client_tuple[1],
                "workContexts": client_tuple[2],
            })
            
        # 3. Sort chronologically
        events.sort(key=lambda e: (e["date_iso"], e["reference"], 0 if e["type"] == "Invoice" else 1))
        
        # 4. Process Running Balance and Opening Balance
        running_balance = 0.0
        opening_balance = 0.0
        final_events = []
        
        for e in events:
            if start_date and e["date"] < start_date:
                opening_balance += (e["debit"] - e["credit"])
                running_balance = opening_balance
            else:
                running_balance += (e["debit"] - e["credit"])
                e_copy = dict(e)
                e_copy["runningBalance"] = round(running_balance, 2)
                e_copy["debit"] = round(e_copy["debit"], 2)
                e_copy["credit"] = round(e_copy["credit"], 2)
                
                # Format money
                e_copy["debitFormatted"] = f"${e_copy['debit']:,.2f}" if e_copy["debit"] > 0 else ""
                e_copy["creditFormatted"] = f"${e_copy['credit']:,.2f}" if e_copy["credit"] > 0 else ""
                e_copy["balanceFormatted"] = f"${e_copy['runningBalance']:,.2f}"
                
                # Cleanup internal date object
                e_copy.pop("date", None)
                e_copy["date"] = e_copy.pop("date_iso")
                final_events.append(e_copy)
                
        # Handle Open Items Only filter
        if open_items_only:
            # Re-calculate balances grouped by invoice to hide fully paid ones
            invoice_balances = {}
            for e in events:
                inv = e["invoice"]
                invoice_balances[inv] = invoice_balances.get(inv, 0.0) + (e["debit"] - e["credit"])
                
            final_events = [e for e in final_events if round(invoice_balances.get(e["invoice"], 0.0), 2) > 0.00]
            
            # Recalculate running balances for the filtered view
            display_balance = 0.0
            for e in final_events:
                display_balance += (e["debit"] - e["credit"])
                e["runningBalance"] = round(display_balance, 2)
                e["balanceFormatted"] = f"${display_balance:,.2f}"
            
            # Since open items only recalculates display balance from 0 for open items,
            # opening balance conceptually becomes 0 for the displayed list.
            opening_balance = 0.0

        return {
            "ok": True,
            "openingBalance": round(opening_balance, 2),
            "openingBalanceFormatted": f"${opening_balance:,.2f}",
            "closingBalance": round(running_balance, 2),
            "closingBalanceFormatted": f"${running_balance:,.2f}",
            "events": final_events,
            "client": requested_client
        }

    @with_db_lock
    def repair_rogue_vendor_receivables(
        self,
        vendor_name: str,
        *,
        apply: bool = False,
    ) -> Dict[str, Any]:
        """Remove imported expense artifacts that were incorrectly staged as A/R.

        An imported expense can occasionally create matching Receivables and
        Invoice Log rows under its vendor.  Those rows make a vendor appear as
        a statement recipient.  This repair is intentionally conservative:
        a candidate must be both a vendor-named receivable and linked to an
        expense transaction with the same reference.  It never removes the
        expense transaction, disbursement, or normal vendor ledger expense.

        ``apply=False`` is a read-only plan.  ``apply=True`` writes only the
        affected table snapshots in one atomic workbook replacement.
        """
        vendor = _clean_text(vendor_name)
        if not vendor:
            return {"ok": False, "message": "A vendor name is required."}
        vendor_key = vendor.casefold()

        if apply:
            self.ensure_schema()

        tables = self._read_table_rows_bulk(
            [
                TBL_RECEIVABLES,
                TBL_INVOICE_LOG,
                TBL_TRANSACTIONS_MASTER,
                TBL_LEDGER,
            ]
        )
        receivables = [dict(row) for row in tables.get(TBL_RECEIVABLES.table, [])]
        invoice_log = [dict(row) for row in tables.get(TBL_INVOICE_LOG.table, [])]
        transactions = [dict(row) for row in tables.get(TBL_TRANSACTIONS_MASTER.table, [])]
        ledger = [dict(row) for row in tables.get(TBL_LEDGER.table, [])]

        def _matches_vendor(value: Any) -> bool:
            return _clean_text(value).casefold() == vendor_key

        expense_refs = {
            _clean_text(row.get(sc.COL_TXN_INVOICE_REF)).casefold()
            for row in transactions
            if _clean_text(row.get(sc.COL_TXN_TYPE)).casefold() == "expense"
            and _matches_vendor(row.get(sc.COL_TXN_CLIENT))
            and _clean_text(row.get(sc.COL_TXN_INVOICE_REF))
        }

        rogue_ref_keys = {
            _clean_text(row.get(sc.COL_RECV_INVOICE_NUM)).casefold()
            for row in receivables
            if _clean_text(row.get(sc.COL_RECV_INVOICE_NUM)).casefold() in expense_refs
            and (
                _matches_vendor(row.get(sc.COL_RECV_CLIENT))
                or _matches_vendor(row.get(sc.COL_RECV_WORK_CLIENT))
            )
        }
        rogue_refs = sorted(
            {
                _clean_text(row.get(sc.COL_RECV_INVOICE_NUM))
                for row in receivables
                if _clean_text(row.get(sc.COL_RECV_INVOICE_NUM)).casefold() in rogue_ref_keys
            },
            key=str.casefold,
        )

        def _is_rogue_receivable(row: Dict[str, Any]) -> bool:
            return (
                _clean_text(row.get(sc.COL_RECV_INVOICE_NUM)).casefold() in rogue_ref_keys
                and (
                    _matches_vendor(row.get(sc.COL_RECV_CLIENT))
                    or _matches_vendor(row.get(sc.COL_RECV_WORK_CLIENT))
                )
            )

        def _is_rogue_invoice_log(row: Dict[str, Any]) -> bool:
            return (
                _clean_text(row.get(sc.COL_INV_INVOICE_NUM)).casefold() in rogue_ref_keys
                and any(
                    _matches_vendor(value)
                    for value in (
                        row.get(sc.COL_INV_CLIENT_NAME),
                        row.get(sc.COL_INV_SUB_CLIENT),
                        row.get(sc.COL_INV_BILL_TO_CLIENT),
                    )
                )
            )

        def _is_cleanup_ledger_artifact(row: Dict[str, Any]) -> bool:
            if _clean_text(row.get(sc.COL_LEDGER_REFERENCE)).casefold() not in rogue_ref_keys:
                return False
            if not (
                _matches_vendor(row.get(sc.COL_LEDGER_CLIENT_VENDOR))
                or _matches_vendor(row.get(sc.COL_LEDGER_WORK_CLIENT))
            ):
                return False
            return "rogue vendor row" in _clean_text(row.get(sc.COL_LEDGER_DESCRIPTION)).casefold()

        remaining_receivables = [row for row in receivables if not _is_rogue_receivable(row)]
        remaining_invoice_log = [row for row in invoice_log if not _is_rogue_invoice_log(row)]
        remaining_ledger = [row for row in ledger if not _is_cleanup_ledger_artifact(row)]
        removed = {
            "receivables": len(receivables) - len(remaining_receivables),
            "invoiceLog": len(invoice_log) - len(remaining_invoice_log),
            "ledgerCleanupArtifacts": len(ledger) - len(remaining_ledger),
            "transactions": 0,
            "disbursements": 0,
        }
        changed = any(value > 0 for key, value in removed.items() if key not in {"transactions", "disbursements"})

        if apply and changed:
            self._write_table_rows_bulk(
                {
                    TBL_LEDGER: remaining_ledger,
                    TBL_RECEIVABLES: remaining_receivables,
                    TBL_INVOICE_LOG: remaining_invoice_log,
                }
            )

        return {
            "ok": True,
            "applied": bool(apply and changed),
            "vendor": vendor,
            "rogueInvoiceRefs": rogue_refs,
            "removed": removed,
            "message": (
                f"Removed {removed['receivables']} rogue {vendor} receivable row(s)."
                if apply and changed
                else f"Identified {removed['receivables']} rogue {vendor} receivable row(s)."
            ),
        }

    def _statement_open_invoice_candidates(self, payload: dict) -> dict:
        """Return one current open-invoice row per invoice for a billing client.

        Statements must be based on the same as-of AR stream as the ledger, but
        the recipient-facing document is invoice-oriented rather than an event
        ledger.  Payments and adjustments are therefore rolled into their
        invoice, producing the exact outstanding balance at the statement date.
        """
        request = dict(payload or {})
        billing_client = _clean_text(
            request.get("billingClient") or request.get("client") or request.get("clientId")
        )
        if not billing_client:
            return {"ok": False, "message": "Choose a billing client before generating a statement."}

        ledger_request = dict(request)
        ledger_request["client"] = billing_client
        ledger_request["client_level"] = "billing"
        # We need all activity up to the as-of date so each invoice's remaining
        # balance can be calculated before excluding paid invoices.
        ledger_request["openItemsOnly"] = False
        ledger_res = self._canonical_ar_ledger(ledger_request)
        if not ledger_res.get("ok"):
            return ledger_res

        def _client_matter_label(context: Dict[str, Any]) -> str:
            client_name = _clean_text(context.get("client") or context.get("workClient"))
            matter_name = _clean_text(
                context.get("matter")
                or context.get("matterDescription")
                or context.get("matterName")
            )
            if client_name and matter_name:
                return f"{client_name} — {matter_name}"
            if client_name:
                return f"{client_name} — Matter not recorded"
            return matter_name

        grouped: Dict[str, Dict[str, Any]] = {}
        closed_statuses = {"paid", "closed", "void", "voided", "cancelled", "canceled", "reversed"}
        for event in ledger_res.get("events", []):
            if not isinstance(event, dict):
                continue
            invoice = _clean_text(event.get("invoice") or event.get("reference"))
            if not invoice:
                continue
            key = invoice.casefold()
            group = grouped.setdefault(
                key,
                {
                    "invoice": invoice,
                    "date": _clean_text(event.get("date")),
                    "invoiceTotal": 0.0,
                    "paidCredits": 0.0,
                    "statementInvoiceTotal": 0.0,
                    "statementPaymentAmount": 0.0,
                    "statementBalanceDue": 0.0,
                    "hasStatementValues": False,
                    "status": _clean_text(event.get("status")),
                    "workClients": [],
                    "clientMatters": [],
                    "matterLinks": [],
                },
            )
            if not group["date"] and _clean_text(event.get("date")):
                group["date"] = _clean_text(event.get("date"))
            if _clean_text(event.get("type")).lower() == "invoice":
                group["invoiceTotal"] += self._money_round(event.get("debit"))
                group["status"] = _clean_text(event.get("status")) or group["status"]
                if event.get("statementValuesAvailable"):
                    group["statementInvoiceTotal"] += self._money_round(event.get("statementInvoiceTotal"))
                    group["statementPaymentAmount"] += self._money_round(event.get("statementPaymentAmount"))
                    group["statementBalanceDue"] += self._money_round(event.get("statementBalanceDue"))
                    group["hasStatementValues"] = True
            else:
                group["paidCredits"] += self._money_round(event.get("credit"))
            work_client = _clean_text(event.get("workClient"))
            if work_client and work_client not in group["workClients"]:
                group["workClients"].append(work_client)
            raw_contexts = event.get("workContexts") or []
            if isinstance(raw_contexts, dict):
                raw_contexts = [raw_contexts]
            for context in raw_contexts:
                if not isinstance(context, dict):
                    continue
                context_client = _clean_text(context.get("client") or context.get("workClient"))
                context_matter = _clean_text(
                    context.get("matter")
                    or context.get("matterDescription")
                    or context.get("matterName")
                )
                if context_client.casefold() == billing_client.casefold() and not context_matter:
                    # An invoice-log row that only repeats the bill-to party
                    # adds no work-context information.
                    continue
                label = _client_matter_label(context)
                if label and label not in group["clientMatters"]:
                    group["clientMatters"].append(label)
                if context_matter:
                    detail = {
                        "client": context_client,
                        "matter": context_matter,
                        "matterId": _clean_text(context.get("matterId")),
                    }
                    if not any(
                        _clean_text(existing.get("client")).casefold() == context_client.casefold()
                        and _clean_text(existing.get("matter")) == context_matter
                        for existing in group["matterLinks"]
                    ):
                        group["matterLinks"].append(detail)
            # Test fixtures and older callers may provide a work matter
            # directly on the event.  Preserve it without requiring a new
            # ledger contract everywhere at once.
            if any(_clean_text(event.get(key)) for key in ("matter", "matterDescription", "matterName")):
                direct_label = _client_matter_label(event)
                if direct_label and direct_label not in group["clientMatters"]:
                    group["clientMatters"].append(direct_label)
                direct_client = _clean_text(event.get("client") or event.get("workClient"))
                direct_matter = _clean_text(
                    event.get("matter") or event.get("matterDescription") or event.get("matterName")
                )
                if direct_matter and not any(
                    _clean_text(existing.get("client")).casefold() == direct_client.casefold()
                    and _clean_text(existing.get("matter")) == direct_matter
                    for existing in group["matterLinks"]
                ):
                    group["matterLinks"].append(
                        {
                            "client": direct_client,
                            "matter": direct_matter,
                            "matterId": _clean_text(event.get("matterId")),
                        }
                    )

        candidates: List[Dict[str, Any]] = []
        for group in grouped.values():
            if group["hasStatementValues"]:
                invoice_total = round(float(group["statementInvoiceTotal"]), 2)
                paid_credits = round(float(group["statementPaymentAmount"]), 2)
                balance_due = round(float(group["statementBalanceDue"]), 2)
            else:
                invoice_total = round(float(group["invoiceTotal"]), 2)
                paid_credits = round(float(group["paidCredits"]), 2)
                balance_due = round(invoice_total - paid_credits, 2)
            if invoice_total <= 0 or balance_due <= 0.009:
                continue
            status_key = _clean_text(group["status"]).casefold()
            if status_key in closed_statuses:
                continue
            client_matters = group["clientMatters"]
            # A later migration can leave both a literal "Legacy Matter …"
            # marker and the resolved contemporary matter for one client.
            # Show the resolved matter only; retain all genuinely distinct
            # matters on a multi-matter invoice.
            resolved_clients = {
                label.split(" — ", 1)[0].casefold()
                for label in client_matters
                if "legacy matter" not in label.casefold()
            }
            client_matters = [
                label
                for label in client_matters
                if "legacy matter" not in label.casefold()
                or label.split(" — ", 1)[0].casefold() not in resolved_clients
            ]
            matter_links = [
                link
                for link in group["matterLinks"]
                if "legacy matter" not in _clean_text(link.get("matter")).casefold()
                or _clean_text(link.get("client")).casefold() not in resolved_clients
            ]
            if not client_matters:
                # A bill-to party is not evidence of the matter receiving the
                # work.  Never label a statement row as the billing client
                # merely because historical context was unavailable.
                work_clients = [
                    name
                    for name in group["workClients"]
                    if _clean_text(name).casefold() != billing_client.casefold()
                ]
                client_matters = work_clients
            service_for = ", ".join(client_matters) if client_matters else "Client matter not recorded"
            candidates.append(
                {
                    "invoice": group["invoice"],
                    "reference": group["invoice"],
                    "date": group["date"],
                    "description": f"Legal services for {service_for}",
                    "serviceFor": service_for,
                    # An unresolved historic label remains plain text in QML.
                    # Only a stable Matter_ID is allowed to open an editable
                    # Matter 360 record, preventing a name-based misroute.
                    "matterLinks": [
                        {
                            "clientName": _clean_text(link.get("client")),
                            "matterName": _clean_text(link.get("matter")),
                            "matterId": _clean_text(link.get("matterId")),
                        }
                        for link in matter_links
                    ],
                    "status": "Partially Paid" if paid_credits > 0.009 else "Unpaid",
                    "invoiceTotal": invoice_total,
                    "invoiceTotalFormatted": f"${invoice_total:,.2f}",
                    "paidCredits": paid_credits,
                    "paidCreditsFormatted": f"${paid_credits:,.2f}" if paid_credits > 0.009 else "—",
                    "balanceDue": balance_due,
                    "balanceDueFormatted": f"${balance_due:,.2f}",
                }
            )
        candidates.sort(key=lambda row: (row.get("date") or "", row.get("invoice") or ""))

        as_of_date = self._parse_date_value(
            request.get("asOfDate") or request.get("endDate") or request.get("as_of_date")
        ) or date.today()
        return {
            "ok": True,
            "billingClient": billing_client,
            "asOfDate": as_of_date.isoformat(),
            "invoices": candidates,
        }

    def list_statement_billing_clients(self) -> List[Dict[str, Any]]:
        """List bill-to parties with current open receivables for statement selection."""
        try:
            self.ensure_schema()
            # Keep the first visible Statement screen responsive: all four
            # lists live in CSPM.xlsm, so loading them independently needlessly
            # reparses the same macro workbook four times.
            statement_table_refs = [
                TBL_RECEIVABLES,
                TBL_CLIENT_PROFILES,
                TBL_CLIENTS,
                TBL_PARENTS,
            ]
            bulk_reader = getattr(self, "_read_table_rows_bulk", None)
            if callable(bulk_reader):
                statement_tables = bulk_reader(statement_table_refs)
            else:
                statement_tables = {
                    tref.table: self._read_table_rows(tref)
                    for tref in statement_table_refs
                }
            receivables = statement_tables.get(TBL_RECEIVABLES.table, [])
            profiles = statement_tables.get(TBL_CLIENT_PROFILES.table, [])
            clients = statement_tables.get(TBL_CLIENTS.table, [])
            parents = statement_tables.get(TBL_PARENTS.table, [])
        except Exception:
            return []

        _parent_names, parent_lookup = self._client_parent_lookup(profiles, clients, parents)
        actual_invoice_pattern = re.compile(r"^\d{2}-\d{4}(?:-[A-Z])?$")
        closed_statuses = {"paid", "closed", "void", "voided", "cancelled", "canceled", "reversed"}
        choices: Dict[str, Dict[str, Any]] = {}
        for row in receivables:
            invoice = _clean_text(row.get(sc.COL_RECV_INVOICE_NUM))
            if not invoice or not actual_invoice_pattern.match(invoice):
                continue
            balance = self._money_round(row.get(sc.COL_RECV_BALANCE_DUE))
            status = _clean_text(row.get(sc.COL_RECV_STATUS)).casefold()
            if balance <= 0.009 or status in closed_statuses:
                continue
            billing_client = _clean_text(row.get(sc.COL_RECV_CLIENT))
            work_client = _clean_text(row.get(sc.COL_RECV_WORK_CLIENT))
            parent = None
            for value in (work_client, billing_client):
                for key in _report_name_keys(value):
                    parent = parent_lookup.get(key)
                    if parent:
                        break
                if parent:
                    break
            name = _clean_text((parent or {}).get("name")) or billing_client or work_client
            if not name:
                continue
            key = name.casefold()
            choice = choices.setdefault(
                key,
                {"name": name, "openInvoiceCount": 0, "openBalance": 0.0},
            )
            choice["openInvoiceCount"] += 1
            choice["openBalance"] = round(choice["openBalance"] + balance, 2)

        rows = list(choices.values())
        for row in rows:
            row["openBalanceFormatted"] = f"${row['openBalance']:,.2f}"
        return sorted(rows, key=lambda row: row["name"].casefold())

    def statement_of_account_report(self, payload: dict) -> dict:
        """Create a selectable, billing-client Statement of Account document."""
        request = dict(payload or {})
        candidate_res = self._statement_open_invoice_candidates(request)
        if not candidate_res.get("ok"):
            return candidate_res

        candidates = candidate_res.get("invoices", [])
        selection_was_supplied = "selectedInvoiceNumbers" in request
        raw_selection = request.get("selectedInvoiceNumbers") or []
        if isinstance(raw_selection, (str, bytes)):
            raw_selection = [raw_selection]
        selected_keys = {
            _clean_text(value).casefold()
            for value in raw_selection
            if _clean_text(value)
        }
        selected_rows = [
            dict(row)
            for row in candidates
            if not selection_was_supplied or row["invoice"].casefold() in selected_keys
        ]
        selected_invoice_numbers = [row["invoice"] for row in selected_rows]
        selected_total = round(sum(float(row["balanceDue"]) for row in selected_rows), 2)
        total_open_balance = round(sum(float(row["balanceDue"]) for row in candidates), 2)
        billing_client = candidate_res["billingClient"]
        as_of = candidate_res["asOfDate"]

        cards = [
            {
                "label": "Selected Invoices",
                "value": str(len(selected_rows)),
                "displayValue": f"{len(selected_rows)} of {len(candidates)}",
                "tone": "secondary",
            },
            {
                "label": "Amount Due",
                "value": f"${selected_total:,.2f}",
                "displayValue": f"${selected_total:,.2f}",
                "tone": "warning",
            },
        ]
        summary_rows = [
            {"label": "Billing Client", "value": billing_client},
            {"label": "Statement Date", "value": as_of},
            {"label": "Invoices Included", "value": str(len(selected_rows))},
            {"label": "Amount Due", "value": f"${selected_total:,.2f}"},
        ]
        summary_columns = [
            {"key": "label", "label": "Item", "width": 180, "minWidth": 120},
            {"key": "value", "label": "Details", "width": 420, "minWidth": 200},
        ]
        columns = [
            {"key": "date", "label": "Invoice Date", "width": 90, "minWidth": 80},
            {"key": "reference", "label": "Invoice", "width": 90, "minWidth": 80},
            {"key": "serviceFor", "label": "Client & Matter", "width": 230, "minWidth": 140},
            {"key": "invoiceTotalFormatted", "label": "Invoice Total", "width": 105, "minWidth": 90, "align": "right"},
            {"key": "paidCreditsFormatted", "label": "Paid / Credits", "width": 105, "minWidth": 90, "align": "right"},
            {"key": "balanceDueFormatted", "label": "Amount Due", "width": 105, "minWidth": 90, "align": "right"},
        ]
        sections = [
            {
                "sectionId": "header",
                "title": "Statement Summary",
                "columns": summary_columns,
                "rows": summary_rows,
                "defaultExpanded": True,
            },
            {
                "sectionId": "detail",
                "title": "Outstanding Invoices",
                "columns": columns,
                "rows": selected_rows,
                "defaultExpanded": True,
            },
        ]

        return {
            "ok": True,
            "reportId": "statement_of_account",
            "title": "Statement of Account",
            "client": billing_client,
            "billingClient": billing_client,
            "clientLevel": "billing",
            "asOfDate": as_of,
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "filterSummary": f"Billing Client: {billing_client} | Open invoices only | As of: {as_of}",
            "cards": cards,
            "rows": selected_rows,
            "availableInvoices": candidates,
            "selectedInvoiceNumbers": selected_invoice_numbers,
            "selectionProvided": selection_was_supplied,
            "summaryRows": summary_rows,
            "summary": {
                "invoiceCount": len(selected_rows),
                "availableInvoiceCount": len(candidates),
                "amountDue": selected_total,
                "totalOpenBalance": total_open_balance,
            },
            "exportPayload": {
                "statement": {
                    "billingClient": billing_client,
                    "asOfDate": as_of,
                    "invoiceCount": len(selected_rows),
                    "amountDue": selected_total,
                    "amountDueFormatted": f"${selected_total:,.2f}",
                },
            },
            "sections": sections,
            "message": "" if candidates else "No unpaid invoices found for this billing client.",
        }

    def export_statement_of_account_csv(self, payload: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        raw_payload = dict(payload or {})
        report = raw_payload
        if "sections" not in report and raw_payload.get("client"):
            report = self.statement_of_account_report(raw_payload)
        if not report.get("ok", True):
            return {
                "ok": False,
                "path": "",
                "filename": "",
                "rowCount": 0,
                "message": _clean_text(report.get("message")) or "Could not build Statement of Account.",
            }

        rows: List[Dict[str, Any]] = []
        sections = report.get("sections") if isinstance(report.get("sections"), list) else []
        for section in sections:
            if isinstance(section, dict) and _clean_text(section.get("sectionId")).lower() == "detail":
                rows = section.get("rows") if isinstance(section.get("rows"), list) else []
                break
        if not rows and isinstance(report.get("rows"), list):
            rows = report.get("rows") or []

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        candidate_dirs = [
            self.paths.exports_dir(),
            self.paths.data_dir() / "exports",
            self.paths.root / "outputs",
        ]
        output_path: Optional[Path] = None
        last_error: Optional[Exception] = None
        for export_dir in candidate_dirs:
            try:
                export_dir.mkdir(parents=True, exist_ok=True)
                candidate = export_dir / f"statement_of_account_{stamp}.csv"
                with candidate.open("w", encoding="utf-8-sig", newline="") as handle:
                    writer = csv.writer(handle)
                    writer.writerow(["Date", "Invoice", "Description", "Invoice Total", "Paid/Credits", "Balance Due"])
                    for row in rows:
                        writer.writerow([
                            row.get("date", ""),
                            row.get("reference", ""),
                            row.get("description", ""),
                            row.get("invoiceTotalAmount", row.get("chargeAmount", row.get("invoiceTotal", ""))),
                            row.get("paymentAmount", row.get("payment", row.get("paidCredits", ""))),
                            row.get("balanceAmount", row.get("balance", row.get("balanceDue", ""))),
                        ])
                output_path = candidate
                break
            except Exception as exc:
                last_error = exc
                continue

        if output_path is None:
            return {
                "ok": False,
                "path": "",
                "filename": "",
                "rowCount": 0,
                "message": f"Could not export CSV: {last_error}",
            }

        return {
            "ok": True,
            "path": str(output_path),
            "filename": output_path.name,
            "rowCount": len(rows),
            "message": f"CSV exported: {output_path.name}",
        }

    # =========================================================================
    # CORPORATE MODULE
    # =========================================================================

    def list_corporate_entities(self) -> List[Dict[str, Any]]:
        return self._read_table_rows(TBL_CORP_ENTITIES)

    def save_corporate_entity(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        with _DB_LOCK:
            wb = self._get_cached_wb()
            sheet = wb.sheets[TBL_CORP_ENTITIES.sheet]
            table = sheet.tables[TBL_CORP_ENTITIES.table]

            entity_id = _clean_text(payload.get(sc.COL_CORP_ENTITY_ID))
            if not entity_id:
                entity_id = self._generate_id("CE")
                payload[sc.COL_CORP_ENTITY_ID] = entity_id

            if sc.COL_CORP_CREATED not in payload:
                payload[sc.COL_CORP_CREATED] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            row_map, row_idx = self._find_row_by_id(table, sc.COL_CORP_ENTITY_ID, entity_id)
            if row_map is None:
                self._append_row(table, payload)
            else:
                merged = dict(row_map)
                merged.update(payload)
                self._update_row(table, row_idx, merged)

            self.save()
            return {"ok": True, "entityId": entity_id, "savedRow": payload}

    def list_corporate_relationships(self) -> List[Dict[str, Any]]:
        return self._read_table_rows(TBL_CORP_RELATIONSHIPS)

    def save_corporate_relationship(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        with _DB_LOCK:
            wb = self._get_cached_wb()
            sheet = wb.sheets[TBL_CORP_RELATIONSHIPS.sheet]
            table = sheet.tables[TBL_CORP_RELATIONSHIPS.table]

            rel_id = _clean_text(payload.get(sc.COL_CREL_ID))
            if not rel_id:
                rel_id = self._generate_id("CR")
                payload[sc.COL_CREL_ID] = rel_id

            row_map, row_idx = self._find_row_by_id(table, sc.COL_CREL_ID, rel_id)
            if row_map is None:
                self._append_row(table, payload)
            else:
                merged = dict(row_map)
                merged.update(payload)
                self._update_row(table, row_idx, merged)

            self.save()
            return {"ok": True, "relationshipId": rel_id, "savedRow": payload}

    def list_corporate_transactions(self) -> List[Dict[str, Any]]:
        return self._read_table_rows(TBL_CORP_TRANSACTIONS)

    def save_corporate_transaction(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        with _DB_LOCK:
            wb = self._get_cached_wb()
            sheet = wb.sheets[TBL_CORP_TRANSACTIONS.sheet]
            table = sheet.tables[TBL_CORP_TRANSACTIONS.table]

            txn_id = _clean_text(payload.get(sc.COL_CTX_ID))
            if not txn_id:
                txn_id = self._generate_id("CTX")
                payload[sc.COL_CTX_ID] = txn_id

            if sc.COL_CTX_CREATED not in payload:
                payload[sc.COL_CTX_CREATED] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            row_map, row_idx = self._find_row_by_id(table, sc.COL_CTX_ID, txn_id)
            if row_map is None:
                self._append_row(table, payload)
            else:
                merged = dict(row_map)
                merged.update(payload)
                self._update_row(table, row_idx, merged)

            self.save()
            return {"ok": True, "transactionId": txn_id, "savedRow": payload}
