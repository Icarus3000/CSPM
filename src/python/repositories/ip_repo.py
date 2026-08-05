from __future__ import annotations

import sys
import csv
import math
import json
import os
from dataclasses import dataclass
from datetime import datetime, UTC
from pathlib import Path
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple

import typing
if typing.TYPE_CHECKING:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils.cell import get_column_letter, range_boundaries
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.worksheet import Worksheet
else:
    Workbook = None
    
    MergedCell = None
    get_column_letter = None
    range_boundaries = None
    Table = None
    TableStyleInfo = None
    Worksheet = None

def _lazy_load_heavy_libs():
    global Workbook, load_workbook, MergedCell, get_column_letter, range_boundaries, Table, TableStyleInfo, Worksheet
    if Workbook is None:
        from openpyxl import Workbook, load_workbook
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils.cell import get_column_letter, range_boundaries
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.worksheet.worksheet import Worksheet


from uuid import uuid4

# Support direct execution via:
# python src/python/repositories/excel_repo.py
if __package__ in (None, ""):
    _THIS_FILE = Path(__file__).resolve()
    _PYTHON_ROOT = _THIS_FILE.parents[1]
    _PYTHON_ROOT_STR = str(_PYTHON_ROOT)
    if _PYTHON_ROOT_STR not in sys.path:
        sys.path.insert(0, _PYTHON_ROOT_STR)


from domain.money import calc_amounts, normalize_pct
from domain import schema_constants as sc
from services.paths import AppPaths
from repositories.excel_repo import ExcelRepo, TableRef

import logging
logger = logging.getLogger(__name__)

TBL_PARENTS = TableRef(sc.SHEET_PARENTS, sc.TBL_PARENTS)
TBL_CLIENTS = TableRef(sc.SHEET_CLIENTS, sc.TBL_CLIENTS)
TBL_CLIENT_PROFILES = TableRef(sc.SHEET_CLIENT_PROFILES, sc.TBL_CLIENT_PROFILES)
TBL_MATTERS = TableRef(sc.SHEET_MATTERS, sc.TBL_MATTERS)
TBL_TIME = TableRef(sc.SHEET_TIME, sc.TBL_TIME)
TBL_TRADEMARKS = TableRef(sc.SHEET_TRADEMARKS, sc.TBL_TRADEMARKS)
TBL_TRANSACTIONS_MASTER = TableRef(sc.SHEET_TRANSACTIONS, sc.TBL_TRANSACTIONS_MASTER)
TBL_TRANSACTION_ACCOUNTS = TableRef(sc.SHEET_TRANSACTION_ACCOUNTS, sc.TBL_TRANSACTION_ACCOUNTS)
TBL_TRANSACTION_CATEGORIES = TableRef(sc.SHEET_TRANSACTION_CATEGORIES, sc.TBL_TRANSACTION_CATEGORIES)
TBL_TRANSACTION_BUSINESS_UNITS = TableRef(sc.SHEET_TRANSACTION_BUSINESS_UNITS, sc.TBL_TRANSACTION_BUSINESS_UNITS)
TBL_TRANSACTION_PAYEES = TableRef(sc.SHEET_TRANSACTION_PAYEES, sc.TBL_TRANSACTION_PAYEES)

TABLES_IN_ORDER = [
    TBL_PARENTS,
    TBL_CLIENTS,
    TBL_CLIENT_PROFILES,
    TBL_MATTERS,
    TBL_TIME,
    TBL_TRADEMARKS,
    TBL_TRANSACTIONS_MASTER,
    TBL_TRANSACTION_ACCOUNTS,
    TBL_TRANSACTION_CATEGORIES,
    TBL_TRANSACTION_BUSINESS_UNITS,
    TBL_TRANSACTION_PAYEES,
]
TABLE_COLUMNS = sc.TABLE_COLUMNS
TABLE_ALIASES = sc.TABLE_ALIASES
TABLE_META_CACHE_SCHEMA_VERSION = 1


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_us_phone(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    digits = re.sub(r"\D+", "", text)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) != 10:
        return text
    return f"{digits[0:3]}-{digits[3:6]}-{digits[6:10]}"


def _looks_like_email_address(value: Any) -> bool:
    text = _clean_text(value)
    if not text:
        return True
    return bool(re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", text))


def _normalize_search_text(value: Any) -> str:
    text = _clean_text(value).lower()
    if not text:
        return ""
    return re.sub(r"\s+", " ", text)


def _matter_token_fragment(value: Any, width: int = 3, fallback: str = "XXX") -> str:
    """Legacy compat shim – kept for any call-sites not yet migrated."""
    token = re.sub(r"[^A-Za-z0-9]+", "", _clean_text(value).upper())
    if not token:
        token = _clean_text(fallback).upper() or "X"
    if len(token) >= width:
        return token[:width]
    return (token + ("X" * width))[:width]


# ---------------------------------------------------------------------------
# Legacy matter-number helpers
# ---------------------------------------------------------------------------

# Hardcoded 3-char matter-type codes.  Keys are normalised lowercase.
_MATTER_TYPE_CODES: Dict[str, str] = {
    "trademark": "TMK",
    "trademarks": "TMK",
    "tmk": "TMK",
    "commercial": "COM",
    "com": "COM",
    "corporate": "CRP",
    "crp": "CRP",
    "corporation": "CRP",
    "estate": "EST",
    "estates": "EST",
    "est": "EST",
    "tax": "TAX",
    "taxation": "TAX",
    "taxes": "TAX",
    "audit": "AUD",
    "audits": "AUD",
    "aud": "AUD",
    "income": "INC",
    "inc": "INC",
    "general": "GEN",
    "gen": "GEN",
}

# Words that should be stripped from the *end* of a company name before coding
_COMPANY_NOISE_SUFFIXES = {
    "inc", "ltd", "llc", "llp", "lp", "corp", "co", "plc", "pte", "pty",
    "sa", "srl", "sas", "gmbh", "ag", "nv", "bv", "ab",
}


def _legacy_matter_type_code(matter_type: Any) -> str:
    """Return the canonical 3-char matter-type code for the given string."""
    key = re.sub(r"[^a-z]", "", _clean_text(matter_type).lower())
    return _MATTER_TYPE_CODES.get(key, "GEN")


def _legacy_raw_client_code(name: Any, entity_type: Any = "") -> str:
    """
    Derive an *uncollision-checked* 4-character client code from *name*.

    Rules
    -----
    Individual  (EntityType == "Individual"):
        last[:4].  If len(last) < 4, pad with first[:needed].
    Company / anything else:
        Strip trailing noise suffixes, then take first-word[:4].
    In both cases, keep alphanumerics (including digits), pad with "X",
    return exactly 4 uppercase characters.
    """
    cleaned = re.sub(r"[^A-Za-z0-9 ]+", "", _clean_text(name)).strip()
    parts = cleaned.split()
    if not parts:
        return "XXXX"

    is_individual = _clean_text(entity_type).lower() == "individual"

    if is_individual:
        last = parts[-1].upper() if len(parts) >= 2 else parts[0].upper()
        raw = last[:4]
        if len(raw) < 4 and len(parts) >= 2:
            first = parts[0].upper()
            raw = raw + first[: 4 - len(raw)]
    else:
        # company: try the first meaningful word
        meaningful_parts = [p for p in parts if p.lower() not in _COMPANY_NOISE_SUFFIXES]
        main = (meaningful_parts[0] if meaningful_parts else parts[0]).upper()
        raw = main[:4]

    return (raw + "XXXX")[:4]


def _matter_year_two_digits(date_text: Any) -> str:
    text = _clean_text(date_text)
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", text)
    if match:
        return match.group(1)[2:]
    return datetime.now().strftime("%y")


def _normalize_choice(value: Any, allowed: List[str], default_value: str = "") -> str:
    normalized_allowed: Dict[str, str] = {
        _clean_text(choice).lower(): _clean_text(choice)
        for choice in allowed
        if _clean_text(choice)
    }
    if not normalized_allowed:
        return _clean_text(value) or _clean_text(default_value)
    raw = _clean_text(value)
    if not raw:
        return _clean_text(default_value) if _clean_text(default_value) else next(iter(normalized_allowed.values()))
    return normalized_allowed.get(raw.lower(), "")


def _is_valid_iso_date(text: Any) -> bool:
    value = _clean_text(text)
    if not value:
        return False
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", value):
        return False
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return True
    except Exception:
        return False


def _to_code_token(value: Any, fallback: str = "X") -> str:
    token = re.sub(r"[^A-Za-z0-9]+", "_", _clean_text(value).strip()).strip("_").upper()
    if token:
        return token
    return _clean_text(fallback).upper() or "X"


def _search_terms(value: Any) -> List[str]:
    normalized = _normalize_search_text(value)
    if not normalized:
        return []

    terms: List[str] = []
    seen = set()
    for raw_token in re.findall(r"[a-z0-9@._:/#\-]+", normalized, flags=re.IGNORECASE):
        token = raw_token.strip(" .,_-")
        if not token:
            continue
        if token in seen:
            continue
        seen.add(token)
        terms.append(token)

    if " " in normalized and normalized not in seen:
        terms.insert(0, normalized)
    return terms


def _boolean_query_tokens(value: Any) -> List[str]:
    raw = _clean_text(value)
    if not raw:
        return []
    return [tok for tok in re.findall(r"\(|\)|AND|OR|NOT|[^\s()]+", raw, flags=re.IGNORECASE) if tok]


def _boolean_to_postfix(tokens: List[str]) -> List[str]:
    precedence = {"OR": 1, "AND": 2, "NOT": 3}
    output: List[str] = []
    stack: List[str] = []

    for raw_tok in tokens:
        tok = str(raw_tok or "").strip()
        if not tok:
            continue
        upper = tok.upper()
        if upper in ("AND", "OR", "NOT"):
            if upper == "NOT":
                while stack and stack[-1] in precedence and precedence[stack[-1]] > precedence[upper]:
                    output.append(stack.pop())
            else:
                while stack and stack[-1] in precedence and precedence[stack[-1]] >= precedence[upper]:
                    output.append(stack.pop())
            stack.append(upper)
            continue
        if tok == "(":
            stack.append(tok)
            continue
        if tok == ")":
            while stack and stack[-1] != "(":
                output.append(stack.pop())
            if stack and stack[-1] == "(":
                stack.pop()
            continue
        output.append(_normalize_search_text(tok))

    while stack:
        top = stack.pop()
        if top in ("(", ")"):
            continue
        output.append(top)
    return output


def _match_boolean_postfix(postfix: List[str], haystack: str) -> bool:
    if not postfix:
        return False
    stack: List[bool] = []

    for tok in postfix:
        upper = tok.upper()
        if upper == "NOT":
            value = stack.pop() if stack else False
            stack.append(not value)
            continue
        if upper in ("AND", "OR"):
            right = stack.pop() if stack else False
            left = stack.pop() if stack else False
            stack.append(left and right if upper == "AND" else left or right)
            continue
        needle = _normalize_search_text(tok)
        stack.append(bool(needle) and needle in haystack)

    return stack[-1] if stack else False


def _normalize_index(value: Any, fallback: int) -> int:
    try:
        parsed = int(value)
        if parsed > 0:
            return parsed
    except Exception:
        pass
    return fallback


def _safe_range_boundaries(ref: str) -> Tuple[int, int, int, int]:
    _lazy_load_heavy_libs()
    min_col_raw, min_row_raw, max_col_raw, max_row_raw = range_boundaries(ref)
    min_col = _normalize_index(min_col_raw, 1)
    min_row = _normalize_index(min_row_raw, 1)
    max_col = _normalize_index(max_col_raw, min_col)
    max_row = _normalize_index(max_row_raw, min_row)
    if max_col < min_col:
        max_col = min_col
    if max_row < min_row:
        max_row = min_row
    return min_col, min_row, max_col, max_row


def _set_cell_value(ws: Worksheet, row_idx: int, col_idx: int, value: Any) -> None:
    _lazy_load_heavy_libs()
    cell_obj = ws.cell(row=row_idx, column=col_idx)
    if isinstance(cell_obj, MergedCell):
        return
    cell_obj.value = value

class IPRepo:
    def __init__(self, db: ExcelRepo):
        self.db = db

    def save_trademark_filing(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.db.ensure_schema()
        raw_payload = dict(payload or {})
        mapped_payload = dict(raw_payload)
        key_map = {
            "trademarkId": sc.COL_TM_ID,
            "jurisdiction": sc.COL_TM_JURISDICTION,
            "jurisdictionOther": sc.COL_TM_JURISDICTION_OTHER,
            "clientName": sc.COL_TM_CLIENT_NAME,
            "matterNumber": sc.COL_TM_MATTER_NUMBER,
            "internalNotes": sc.COL_TM_INTERNAL_NOTES,
            "trademarkText": sc.COL_TM_TRADEMARK_TEXT,
            "markType": sc.COL_TM_MARK_TYPE,
            "designRepresentation": sc.COL_TM_DESIGN_REPRESENTATION,
            "designImagePaste": sc.COL_TM_DESIGN_IMAGE_PASTE,
            "colorClaimed": sc.COL_TM_COLOR_CLAIMED,
            "colorDescription": sc.COL_TM_COLOR_DESCRIPTION,
            "niceClasses": sc.COL_TM_NICE_CLASSES,
            "goodsServices": sc.COL_TM_GOODS_SERVICES,
            "foreignPriorityClaim": sc.COL_TM_FOREIGN_PRIORITY,
            "registryLink": sc.COL_TM_REGISTRY_LINK,
            "applicationNumber": sc.COL_TM_APPLICATION_NO,
            "registrationNumber": sc.COL_TM_REGISTRATION_NO,
            "currentStatus": sc.COL_TM_CURRENT_STATUS,
            "applicantNameAddress": sc.COL_TM_APPLICANT_NAME_ADDRESS,
            "filingDate": sc.COL_TM_FILING_DATE,
            "registrationDate": sc.COL_TM_REGISTRATION_DATE,
            "renewalDeadline": sc.COL_TM_RENEWAL_DEADLINE,
            "cipoStatus": sc.COL_TM_CIPO_STATUS,
            "tm5Status": sc.COL_TM_TM5_STATUS,
            "examinersReportDate": sc.COL_TM_EXAMINERS_REPORT_DATE,
            "officeActionResponseDeadline": sc.COL_TM_OFFICE_ACTION_RESPONSE_DEADLINE,
            "approvalDate": sc.COL_TM_APPROVAL_DATE,
            "advertisementDate": sc.COL_TM_ADVERTISEMENT_DATE,
            "advertisementVolIssue": sc.COL_TM_ADVERTISEMENT_VOL_ISSUE,
            "oppositionDeadline": sc.COL_TM_OPPOSITION_DEADLINE,
            "allowanceDate": sc.COL_TM_ALLOWANCE_DATE,
            "registerType": sc.COL_TM_REGISTER_TYPE,
            "usptoStatusIndicator": sc.COL_TM_USPTO_STATUS_INDICATOR,
            "ownerNameAddress": sc.COL_TM_OWNER_NAME_ADDRESS,
            "attorneyOfRecord": sc.COL_TM_ATTORNEY_OF_RECORD,
            "publicationDate": sc.COL_TM_PUBLICATION_DATE,
            "noticeOfAllowanceDate": sc.COL_TM_NOTICE_OF_ALLOWANCE_DATE,
            "souDeadline": sc.COL_TM_SOU_DEADLINE,
            "souExtensionTracking": sc.COL_TM_SOU_EXTENSION_TRACKING,
            "section8Deadline": sc.COL_TM_SECTION8_DEADLINE,
            "section15Deadline": sc.COL_TM_SECTION15_DEADLINE,
            "section9Deadline": sc.COL_TM_SECTION9_DEADLINE,
            "localForeignAssociate": sc.COL_TM_LOCAL_FOREIGN_ASSOCIATE,
            "applicationReferenceNumber": sc.COL_TM_APPLICATION_REFERENCE_NO,
            "publicationAdvertisementDate": sc.COL_TM_PUBLICATION_ADVERTISEMENT_DATE,
            "oppositionPeriodEndDate": sc.COL_TM_OPPOSITION_PERIOD_END_DATE,
            "upcomingLocalDeadlineOfficeActionDate": sc.COL_TM_UPCOMING_LOCAL_DEADLINE,
            "createdAt": sc.COL_TM_CREATED_AT,
            "updatedAt": sc.COL_TM_UPDATED_AT,
        }
        for source_key, target_key in key_map.items():
            if source_key in raw_payload and target_key not in mapped_payload:
                mapped_payload[target_key] = raw_payload.get(source_key)
        incoming = self.db._canonicalize_trademark_row(mapped_payload)
        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        trademark_id = _clean_text(incoming.get(sc.COL_TM_ID))
        if not trademark_id:
            trademark_id = self.db._new_id("TM")
        incoming[sc.COL_TM_ID] = trademark_id
        incoming[sc.COL_TM_UPDATED_AT] = now_stamp
        if not _clean_text(incoming.get(sc.COL_TM_CREATED_AT)):
            incoming[sc.COL_TM_CREATED_AT] = now_stamp
        if not _clean_text(incoming.get(sc.COL_TM_REGISTRY_LINK)):
            incoming[sc.COL_TM_REGISTRY_LINK] = self.db._default_trademark_registry_link(
                incoming.get(sc.COL_TM_JURISDICTION),
                incoming.get(sc.COL_TM_APPLICATION_NO),
                incoming.get(sc.COL_TM_REGISTRATION_NO),
                incoming.get(sc.COL_TM_TRADEMARK_TEXT),
            )

        rows = [self.db._canonicalize_trademark_row(r) for r in self.db._read_table_rows(TBL_TRADEMARKS)]
        found = False
        for idx, row in enumerate(rows):
            row_id = _clean_text(row.get(sc.COL_TM_ID))
            if not row_id or row_id.lower() != trademark_id.lower():
                continue
            incoming[sc.COL_TM_CREATED_AT] = _clean_text(row.get(sc.COL_TM_CREATED_AT)) or incoming.get(
                sc.COL_TM_CREATED_AT
            )
            rows[idx] = incoming
            found = True
            break
        if not found:
            rows.append(incoming)

        self.db._replace_table_rows(TBL_TRADEMARKS, rows)
        self.db._sync_trademark_generated_deadlines(incoming)

        persisted = {}
        verify_rows = [self.db._canonicalize_trademark_row(r) for r in self.db._read_table_rows(TBL_TRADEMARKS)]
        for row in verify_rows:
            if _clean_text(row.get(sc.COL_TM_ID)).lower() == trademark_id.lower():
                persisted = row
                break
        verified = self.db._compare_rows_loose(incoming, persisted)

        return {
            "ok": bool(verified),
            "verifiedExact": bool(verified),
            "trademarkId": trademark_id,
            "savedRow": self.db._trademark_row_to_payload(incoming),
            "message": "" if verified else "Trademark write verification failed.",
        }

    def list_trademark_directory(self, query: str = "") -> List[Dict[str, Any]]:
        self.db.ensure_schema()
        rows = [self.db._canonicalize_trademark_row(r) for r in self.db._read_table_rows(TBL_TRADEMARKS)]
        normalized_query = _normalize_search_text(query)
        terms = _search_terms(normalized_query) if normalized_query else []

        out: List[Dict[str, Any]] = []
        for row in rows:
            payload = self.db._trademark_row_to_payload(row)
            haystack = _normalize_search_text(
                " | ".join(
                    [
                        payload.get("trademarkId", ""),
                        payload.get("trademarkText", ""),
                        payload.get("jurisdiction", ""),
                        payload.get("jurisdictionOther", ""),
                        payload.get("applicationNumber", ""),
                        payload.get("registrationNumber", ""),
                        payload.get("currentStatus", ""),
                        payload.get("cipoStatus", ""),
                        payload.get("usptoStatusIndicator", ""),
                        payload.get("clientName", ""),
                        payload.get("matterNumber", ""),
                        payload.get("niceClasses", ""),
                        payload.get("goodsServices", ""),
                        payload.get("registryLink", ""),
                    ]
                )
            )
            if terms and not all(term in haystack for term in terms):
                continue
            out.append(payload)

        out.sort(
            key=lambda row: (
                _clean_text(row.get("updatedAt")).lower(),
                _clean_text(row.get("trademarkText")).lower(),
                _clean_text(row.get("trademarkId")).lower(),
            ),
            reverse=True,
        )
        return out

    def _canonicalize_trademark_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        trademark_id = _clean_text(self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_ID))
        if not trademark_id:
            trademark_id = self.db._new_id("TM")

        jurisdiction = _normalize_choice(
            self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_JURISDICTION),
            self.db.TM_JURISDICTION_OPTIONS,
            "CIPO",
        )
        if not jurisdiction:
            jurisdiction = "CIPO"

        mark_type = _normalize_choice(
            self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_MARK_TYPE),
            self.db.TM_MARK_TYPE_OPTIONS,
            "Standard Character",
        )
        if not mark_type:
            mark_type = "Standard Character"

        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        created_at = _clean_text(self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_CREATED_AT))
        if not created_at:
            created_at = now_stamp
        updated_at = _clean_text(self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_UPDATED_AT))
        if not updated_at:
            updated_at = created_at

        return {
            sc.COL_TM_ID: trademark_id,
            sc.COL_TM_JURISDICTION: jurisdiction,
            sc.COL_TM_JURISDICTION_OTHER: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_JURISDICTION_OTHER)
            ),
            sc.COL_TM_CLIENT_NAME: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_CLIENT_NAME)
            ),
            sc.COL_TM_MATTER_NUMBER: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_MATTER_NUMBER)
            ),
            sc.COL_TM_INTERNAL_NOTES: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_INTERNAL_NOTES)
            ),
            sc.COL_TM_TRADEMARK_TEXT: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_TRADEMARK_TEXT)
            ),
            sc.COL_TM_MARK_TYPE: mark_type,
            sc.COL_TM_DESIGN_REPRESENTATION: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_DESIGN_REPRESENTATION)
            ),
            sc.COL_TM_DESIGN_IMAGE_PASTE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_DESIGN_IMAGE_PASTE)
            ),
            sc.COL_TM_COLOR_CLAIMED: self.db._to_bool_int(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_COLOR_CLAIMED),
                default=0,
            ),
            sc.COL_TM_COLOR_DESCRIPTION: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_COLOR_DESCRIPTION)
            ),
            sc.COL_TM_NICE_CLASSES: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_NICE_CLASSES)
            ),
            sc.COL_TM_GOODS_SERVICES: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_GOODS_SERVICES)
            ),
            sc.COL_TM_FOREIGN_PRIORITY: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_FOREIGN_PRIORITY)
            ),
            sc.COL_TM_REGISTRY_LINK: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_REGISTRY_LINK)
            ),
            sc.COL_TM_APPLICATION_NO: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_APPLICATION_NO)
            ),
            sc.COL_TM_REGISTRATION_NO: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_REGISTRATION_NO)
            ),
            sc.COL_TM_CURRENT_STATUS: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_CURRENT_STATUS)
            ),
            sc.COL_TM_APPLICANT_NAME_ADDRESS: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_APPLICANT_NAME_ADDRESS)
            ),
            sc.COL_TM_FILING_DATE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_FILING_DATE)
            ),
            sc.COL_TM_REGISTRATION_DATE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_REGISTRATION_DATE)
            ),
            sc.COL_TM_RENEWAL_DEADLINE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_RENEWAL_DEADLINE)
            ),
            sc.COL_TM_CIPO_STATUS: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_CIPO_STATUS)
            ),
            sc.COL_TM_TM5_STATUS: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_TM5_STATUS)
            ),
            sc.COL_TM_EXAMINERS_REPORT_DATE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_EXAMINERS_REPORT_DATE)
            ),
            sc.COL_TM_OFFICE_ACTION_RESPONSE_DEADLINE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_OFFICE_ACTION_RESPONSE_DEADLINE)
            ),
            sc.COL_TM_APPROVAL_DATE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_APPROVAL_DATE)
            ),
            sc.COL_TM_ADVERTISEMENT_DATE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_ADVERTISEMENT_DATE)
            ),
            sc.COL_TM_ADVERTISEMENT_VOL_ISSUE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_ADVERTISEMENT_VOL_ISSUE)
            ),
            sc.COL_TM_OPPOSITION_DEADLINE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_OPPOSITION_DEADLINE)
            ),
            sc.COL_TM_ALLOWANCE_DATE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_ALLOWANCE_DATE)
            ),
            sc.COL_TM_REGISTER_TYPE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_REGISTER_TYPE)
            ),
            sc.COL_TM_USPTO_STATUS_INDICATOR: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_USPTO_STATUS_INDICATOR)
            ),
            sc.COL_TM_OWNER_NAME_ADDRESS: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_OWNER_NAME_ADDRESS)
            ),
            sc.COL_TM_ATTORNEY_OF_RECORD: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_ATTORNEY_OF_RECORD)
            ),
            sc.COL_TM_PUBLICATION_DATE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_PUBLICATION_DATE)
            ),
            sc.COL_TM_NOTICE_OF_ALLOWANCE_DATE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_NOTICE_OF_ALLOWANCE_DATE)
            ),
            sc.COL_TM_SOU_DEADLINE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_SOU_DEADLINE)
            ),
            sc.COL_TM_SOU_EXTENSION_TRACKING: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_SOU_EXTENSION_TRACKING)
            ),
            sc.COL_TM_SECTION8_DEADLINE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_SECTION8_DEADLINE)
            ),
            sc.COL_TM_SECTION15_DEADLINE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_SECTION15_DEADLINE)
            ),
            sc.COL_TM_SECTION9_DEADLINE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_SECTION9_DEADLINE)
            ),
            sc.COL_TM_LOCAL_FOREIGN_ASSOCIATE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_LOCAL_FOREIGN_ASSOCIATE)
            ),
            sc.COL_TM_APPLICATION_REFERENCE_NO: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_APPLICATION_REFERENCE_NO)
            ),
            sc.COL_TM_PUBLICATION_ADVERTISEMENT_DATE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_PUBLICATION_ADVERTISEMENT_DATE)
            ),
            sc.COL_TM_OPPOSITION_PERIOD_END_DATE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_OPPOSITION_PERIOD_END_DATE)
            ),
            sc.COL_TM_UPCOMING_LOCAL_DEADLINE: _clean_text(
                self.db._value_with_alias(TBL_TRADEMARKS.table, row, sc.COL_TM_UPCOMING_LOCAL_DEADLINE)
            ),
            sc.COL_TM_CREATED_AT: created_at,
            sc.COL_TM_UPDATED_AT: updated_at,
        }
