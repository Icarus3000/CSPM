from __future__ import annotations

import sys
import csv
import math
import json
import os
from dataclasses import dataclass
from datetime import datetime, UTC
from pathlib import Path
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple

import typing
if typing.TYPE_CHECKING:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils.cell import get_column_letter, range_boundaries
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.worksheet import Worksheet
else:
    Workbook = None
    
    MergedCell = None
    get_column_letter = None
    range_boundaries = None
    Table = None
    TableStyleInfo = None
    Worksheet = None

def _lazy_load_heavy_libs():
    global Workbook, load_workbook, MergedCell, get_column_letter, range_boundaries, Table, TableStyleInfo, Worksheet
    if Workbook is None:
        from openpyxl import Workbook, load_workbook
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils.cell import get_column_letter, range_boundaries
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.worksheet.worksheet import Worksheet


from uuid import uuid4

# Support direct execution via:
# python src/python/repositories/excel_repo.py
if __package__ in (None, ""):
    _THIS_FILE = Path(__file__).resolve()
    _PYTHON_ROOT = _THIS_FILE.parents[1]
    _PYTHON_ROOT_STR = str(_PYTHON_ROOT)
    if _PYTHON_ROOT_STR not in sys.path:
        sys.path.insert(0, _PYTHON_ROOT_STR)


from domain.money import calc_amounts, normalize_pct
from domain import schema_constants as sc
from services.paths import AppPaths
from repositories.excel_repo import ExcelRepo, TableRef

import logging
logger = logging.getLogger(__name__)

TBL_PARENTS = TableRef(sc.SHEET_PARENTS, sc.TBL_PARENTS)
TBL_CLIENTS = TableRef(sc.SHEET_CLIENTS, sc.TBL_CLIENTS)
TBL_CLIENT_PROFILES = TableRef(sc.SHEET_CLIENT_PROFILES, sc.TBL_CLIENT_PROFILES)
TBL_MATTERS = TableRef(sc.SHEET_MATTERS, sc.TBL_MATTERS)
TBL_TIME = TableRef(sc.SHEET_TIME, sc.TBL_TIME)
TBL_TRADEMARKS = TableRef(sc.SHEET_TRADEMARKS, sc.TBL_TRADEMARKS)
TBL_TRANSACTIONS_MASTER = TableRef(sc.SHEET_TRANSACTIONS, sc.TBL_TRANSACTIONS_MASTER)
TBL_TRANSACTION_ACCOUNTS = TableRef(sc.SHEET_TRANSACTION_ACCOUNTS, sc.TBL_TRANSACTION_ACCOUNTS)
TBL_TRANSACTION_CATEGORIES = TableRef(sc.SHEET_TRANSACTION_CATEGORIES, sc.TBL_TRANSACTION_CATEGORIES)
TBL_TRANSACTION_BUSINESS_UNITS = TableRef(sc.SHEET_TRANSACTION_BUSINESS_UNITS, sc.TBL_TRANSACTION_BUSINESS_UNITS)
TBL_TRANSACTION_PAYEES = TableRef(sc.SHEET_TRANSACTION_PAYEES, sc.TBL_TRANSACTION_PAYEES)

TABLES_IN_ORDER = [
    TBL_PARENTS,
    TBL_CLIENTS,
    TBL_CLIENT_PROFILES,
    TBL_MATTERS,
    TBL_TIME,
    TBL_TRADEMARKS,
    TBL_TRANSACTIONS_MASTER,
    TBL_TRANSACTION_ACCOUNTS,
    TBL_TRANSACTION_CATEGORIES,
    TBL_TRANSACTION_BUSINESS_UNITS,
    TBL_TRANSACTION_PAYEES,
]
TABLE_COLUMNS = sc.TABLE_COLUMNS
TABLE_ALIASES = sc.TABLE_ALIASES
TABLE_META_CACHE_SCHEMA_VERSION = 1


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_us_phone(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    digits = re.sub(r"\D+", "", text)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) != 10:
        return text
    return f"{digits[0:3]}-{digits[3:6]}-{digits[6:10]}"


def _looks_like_email_address(value: Any) -> bool:
    text = _clean_text(value)
    if not text:
        return True
    return bool(re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", text))


def _normalize_search_text(value: Any) -> str:
    text = _clean_text(value).lower()
    if not text:
        return ""
    return re.sub(r"\s+", " ", text)


def _matter_token_fragment(value: Any, width: int = 3, fallback: str = "XXX") -> str:
    """Legacy compat shim – kept for any call-sites not yet migrated."""
    token = re.sub(r"[^A-Za-z0-9]+", "", _clean_text(value).upper())
    if not token:
        token = _clean_text(fallback).upper() or "X"
    if len(token) >= width:
        return token[:width]
    return (token + ("X" * width))[:width]


# ---------------------------------------------------------------------------
# Legacy matter-number helpers
# ---------------------------------------------------------------------------

# Hardcoded 3-char matter-type codes.  Keys are normalised lowercase.
_MATTER_TYPE_CODES: Dict[str, str] = {
    "trademark": "TMK",
    "trademarks": "TMK",
    "tmk": "TMK",
    "commercial": "COM",
    "com": "COM",
    "corporate": "CRP",
    "crp": "CRP",
    "corporation": "CRP",
    "estate": "EST",
    "estates": "EST",
    "est": "EST",
    "tax": "TAX",
    "taxation": "TAX",
    "taxes": "TAX",
    "audit": "AUD",
    "audits": "AUD",
    "aud": "AUD",
    "income": "INC",
    "inc": "INC",
    "general": "GEN",
    "gen": "GEN",
}

# Words that should be stripped from the *end* of a company name before coding
_COMPANY_NOISE_SUFFIXES = {
    "inc", "ltd", "llc", "llp", "lp", "corp", "co", "plc", "pte", "pty",
    "sa", "srl", "sas", "gmbh", "ag", "nv", "bv", "ab",
}


def _legacy_matter_type_code(matter_type: Any) -> str:
    """Return the canonical 3-char matter-type code for the given string."""
    key = re.sub(r"[^a-z]", "", _clean_text(matter_type).lower())
    return _MATTER_TYPE_CODES.get(key, "GEN")


def _legacy_raw_client_code(name: Any, entity_type: Any = "") -> str:
    """
    Derive an *uncollision-checked* 4-character client code from *name*.

    Rules
    -----
    Individual  (EntityType == "Individual"):
        last[:4].  If len(last) < 4, pad with first[:needed].
    Company / anything else:
        Strip trailing noise suffixes, then take first-word[:4].
    In both cases, keep alphanumerics (including digits), pad with "X",
    return exactly 4 uppercase characters.
    """
    cleaned = re.sub(r"[^A-Za-z0-9 ]+", "", _clean_text(name)).strip()
    parts = cleaned.split()
    if not parts:
        return "XXXX"

    is_individual = _clean_text(entity_type).lower() == "individual"

    if is_individual:
        last = parts[-1].upper() if len(parts) >= 2 else parts[0].upper()
        raw = last[:4]
        if len(raw) < 4 and len(parts) >= 2:
            first = parts[0].upper()
            raw = raw + first[: 4 - len(raw)]
    else:
        # company: try the first meaningful word
        meaningful_parts = [p for p in parts if p.lower() not in _COMPANY_NOISE_SUFFIXES]
        main = (meaningful_parts[0] if meaningful_parts else parts[0]).upper()
        raw = main[:4]

    return (raw + "XXXX")[:4]


def _matter_year_two_digits(date_text: Any) -> str:
    text = _clean_text(date_text)
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", text)
    if match:
        return match.group(1)[2:]
    return datetime.now().strftime("%y")


def _normalize_choice(value: Any, allowed: List[str], default_value: str = "") -> str:
    normalized_allowed: Dict[str, str] = {
        _clean_text(choice).lower(): _clean_text(choice)
        for choice in allowed
        if _clean_text(choice)
    }
    if not normalized_allowed:
        return _clean_text(value) or _clean_text(default_value)
    raw = _clean_text(value)
    if not raw:
        return _clean_text(default_value) if _clean_text(default_value) else next(iter(normalized_allowed.values()))
    return normalized_allowed.get(raw.lower(), "")


def _is_valid_iso_date(text: Any) -> bool:
    value = _clean_text(text)
    if not value:
        return False
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", value):
        return False
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return True
    except Exception:
        return False


def _to_code_token(value: Any, fallback: str = "X") -> str:
    token = re.sub(r"[^A-Za-z0-9]+", "_", _clean_text(value).strip()).strip("_").upper()
    if token:
        return token
    return _clean_text(fallback).upper() or "X"


def _search_terms(value: Any) -> List[str]:
    normalized = _normalize_search_text(value)
    if not normalized:
        return []

    terms: List[str] = []
    seen = set()
    for raw_token in re.findall(r"[a-z0-9@._:/#\-]+", normalized, flags=re.IGNORECASE):
        token = raw_token.strip(" .,_-")
        if not token:
            continue
        if token in seen:
            continue
        seen.add(token)
        terms.append(token)

    if " " in normalized and normalized not in seen:
        terms.insert(0, normalized)
    return terms


def _boolean_query_tokens(value: Any) -> List[str]:
    raw = _clean_text(value)
    if not raw:
        return []
    return [tok for tok in re.findall(r"\(|\)|AND|OR|NOT|[^\s()]+", raw, flags=re.IGNORECASE) if tok]


def _boolean_to_postfix(tokens: List[str]) -> List[str]:
    precedence = {"OR": 1, "AND": 2, "NOT": 3}
    output: List[str] = []
    stack: List[str] = []

    for raw_tok in tokens:
        tok = str(raw_tok or "").strip()
        if not tok:
            continue
        upper = tok.upper()
        if upper in ("AND", "OR", "NOT"):
            if upper == "NOT":
                while stack and stack[-1] in precedence and precedence[stack[-1]] > precedence[upper]:
                    output.append(stack.pop())
            else:
                while stack and stack[-1] in precedence and precedence[stack[-1]] >= precedence[upper]:
                    output.append(stack.pop())
            stack.append(upper)
            continue
        if tok == "(":
            stack.append(tok)
            continue
        if tok == ")":
            while stack and stack[-1] != "(":
                output.append(stack.pop())
            if stack and stack[-1] == "(":
                stack.pop()
            continue
        output.append(_normalize_search_text(tok))

    while stack:
        top = stack.pop()
        if top in ("(", ")"):
            continue
        output.append(top)
    return output


def _match_boolean_postfix(postfix: List[str], haystack: str) -> bool:
    if not postfix:
        return False
    stack: List[bool] = []

    for tok in postfix:
        upper = tok.upper()
        if upper == "NOT":
            value = stack.pop() if stack else False
            stack.append(not value)
            continue
        if upper in ("AND", "OR"):
            right = stack.pop() if stack else False
            left = stack.pop() if stack else False
            stack.append(left and right if upper == "AND" else left or right)
            continue
        needle = _normalize_search_text(tok)
        stack.append(bool(needle) and needle in haystack)

    return stack[-1] if stack else False


def _normalize_index(value: Any, fallback: int) -> int:
    try:
        parsed = int(value)
        if parsed > 0:
            return parsed
    except Exception:
        pass
    return fallback


def _safe_range_boundaries(ref: str) -> Tuple[int, int, int, int]:
    _lazy_load_heavy_libs()
    min_col_raw, min_row_raw, max_col_raw, max_row_raw = range_boundaries(ref)
    min_col = _normalize_index(min_col_raw, 1)
    min_row = _normalize_index(min_row_raw, 1)
    max_col = _normalize_index(max_col_raw, min_col)
    max_row = _normalize_index(max_row_raw, min_row)
    if max_col < min_col:
        max_col = min_col
    if max_row < min_row:
        max_row = min_row
    return min_col, min_row, max_col, max_row


def _set_cell_value(ws: Worksheet, row_idx: int, col_idx: int, value: Any) -> None:
    _lazy_load_heavy_libs()
    cell_obj = ws.cell(row=row_idx, column=col_idx)
    if isinstance(cell_obj, MergedCell):
        return
    cell_obj.value = value

class DocketRepo:
    def __init__(self, db: ExcelRepo):
        self.db = db

    def list_deadline_entries(self) -> list:
        """Return all deadline records."""
        return self.db._sort_deadline_entries(self.db._load_deadlines())

    def create_deadline_entry(self, entry: Dict[str, Any]) -> Dict[str, Any]:
        """Add a new deadline and return the saved object (with id)."""
        entries = self.db._load_deadlines()
        new = self.db._normalize_deadline_entry(entry)
        now_iso = self.db._now_utc_iso()
        new["createdAt"] = now_iso
        new["updatedAt"] = now_iso
        entries.append(new)
        entries = self.db._sort_deadline_entries(entries)
        self.db._write_deadlines(entries)
        return dict(new)

    def update_deadline_entry(self, entry_id: str, changes: Dict[str, Any]) -> Dict[str, Any] | None:
        entries = self.db._load_deadlines()
        target_id = _clean_text(entry_id)
        for idx, existing in enumerate(entries):
            if _clean_text(existing.get("id")) == target_id:
                merged = dict(existing)
                merged.update(dict(changes or {}))
                normalized = self.db._normalize_deadline_entry(merged)
                normalized["id"] = target_id or normalized["id"]
                normalized["createdAt"] = _clean_text(existing.get("createdAt")) or normalized.get("createdAt")
                normalized["updatedAt"] = self.db._now_utc_iso()
                entries[idx] = normalized
                entries = self.db._sort_deadline_entries(entries)
                self.db._write_deadlines(entries)
                return dict(normalized)
        return None

    def delete_deadline_entry(self, entry_id: str) -> bool:
        entries = self.db._load_deadlines()
        target_id = _clean_text(entry_id)
        new_list = [e for e in entries if _clean_text(e.get("id")) != target_id]
        if len(new_list) != len(entries):
            self.db._write_deadlines(new_list)
            return True
        return False

    def add_time_entry(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.db.ensure_schema()

        normalized = self.db._normalize_time_payload(payload)
        parent_name = normalized["parentName"]
        parent_row: Optional[Dict[str, Any]] = None
        if parent_name:
            parent_row = self.db._get_or_create_parent(parent_name)

        client_row = self.db._get_or_create_client(normalized["clientName"])
        client_id = _clean_text(client_row.get(sc.COL_CLIENT_ID))
        parent_id = _clean_text(parent_row.get(sc.COL_PARENT_ID)) if parent_row else ""

        matter_id = ""
        if normalized["matterName"]:
            matter_row = self.db._get_or_create_matter(
                matter_name=normalized["matterName"],
                client_id=client_id,
                parent_id=parent_id,
                default_rate=normalized["clientRate"],
                default_share_pct=normalized["sharePct"],
            )
            matter_id = _clean_text(matter_row.get(sc.COL_MATTER_ID))
            if not parent_id:
                parent_id = _clean_text(matter_row.get(sc.COL_MATTER_PARENT_ID))
        elif not normalized.get("allowClientOnlyDraft", False):
            raise ValueError("Matter is required for time entry.")

        normalized_status = self.db._normalize_time_status(normalized.get("status"))

        incoming_entry_id = _clean_text(normalized.get("entryId"))
        rows = [self.db._canonicalize_time_row(r) for r in self.db._read_table_rows(TBL_TIME)]
        matching_rows: List[Dict[str, Any]] = []
        
        if incoming_entry_id:
            for row in rows:
                if _clean_text(row.get(sc.COL_TIME_ENTRY_ID)) == incoming_entry_id:
                    matching_rows.append(row)
                    break
        else:
            for row in rows:
                if _clean_text(row.get(sc.COL_TIME_DATE)) != normalized["date"]:
                    continue
                if _clean_text(row.get(sc.COL_TIME_CLIENT_ID)).lower() != client_id.lower():
                    continue
                row_matter_id = _clean_text(row.get(sc.COL_TIME_MATTER_ID))
                if matter_id:
                    if row_matter_id.lower() != matter_id.lower():
                        continue
                elif row_matter_id:
                    continue
                matching_rows.append(row)

        billed_rows = [
            r for r in matching_rows if self.db._normalize_time_status(r.get(sc.COL_TIME_STATUS)) == "Billed"
        ]
        if billed_rows and not normalized.get("forceEditBilled"):
            return {
                "ok": False,
                "verifiedExact": False,
                "entryId": _clean_text(billed_rows[0].get(sc.COL_TIME_ENTRY_ID)),
                "savedRow": {},
                "message": "Time entry is already billed and locked.",
            }

        editable_rows = [
            r for r in matching_rows if _clean_text(r.get(sc.COL_TIME_STATUS)).lower() != "merged"
        ]
        editable_rows.sort(
            key=lambda row: (
                _clean_text(row.get(sc.COL_TIME_CREATED)),
                _clean_text(row.get(sc.COL_TIME_ENTRY_ID)),
            )
        )

        primary_row: Optional[Dict[str, Any]] = editable_rows[0] if editable_rows else None
        duplicate_rows: List[Dict[str, Any]] = editable_rows[1:] if len(editable_rows) > 1 else []
        existing_total_seconds = 0
        for row in editable_rows:
            existing_total_seconds += int(self.db._parse_float(row.get(sc.COL_TIME_SECONDS)) or 0)

        incoming_raw_seconds = int(normalized.get("rawSeconds") or 0)
        incoming_segment_seconds = normalized.get("segmentSeconds")
        mode = _clean_text(normalized.get("rawSecondsMode")).lower()
        if mode not in ("increment", "replace", "auto"):
            mode = "auto"

        if mode == "replace":
            aggregate_raw_seconds = max(0, incoming_raw_seconds)
        elif mode == "increment":
            increment = int(incoming_segment_seconds) if incoming_segment_seconds is not None else incoming_raw_seconds
            aggregate_raw_seconds = max(0, existing_total_seconds + max(0, int(increment)))
        else:
            # Auto mode: accept monotonic absolute totals; otherwise treat as incremental segment.
            if incoming_raw_seconds >= existing_total_seconds:
                aggregate_raw_seconds = max(0, incoming_raw_seconds)
            else:
                increment = int(incoming_segment_seconds) if incoming_segment_seconds is not None else incoming_raw_seconds
                aggregate_raw_seconds = max(0, existing_total_seconds + max(0, int(increment)))

        aggregate_hours = (
            math.ceil(((aggregate_raw_seconds / 3600.0) * 10.0) - 1e-9) / 10.0
            if aggregate_raw_seconds > 0
            else 0.0
        )
        amounts = calc_amounts(
            hours=aggregate_hours,
            client_rate=normalized["clientRate"],
            your_share_pct=normalized["sharePct"],
            hst_rate=0.13,
        )

        # When the same client+matter+date docket is saved again with a new
        # description, append the new segment with "; " so the daily aggregate
        # reads like a journal: "Drafting email; Setting task reminder; ...".
        # Duplicate segments are skipped (timer auto-saves resend the same text).
        incoming_description = _clean_text(normalized.get("description"))
        existing_description = _clean_text((primary_row or {}).get(sc.COL_TIME_DESC))
        
        if incoming_entry_id:
            description_text = incoming_description or existing_description or "Time entry"
        elif incoming_description and existing_description:
            # Split existing into its already-recorded segments
            existing_segments = [
                seg.strip() for seg in existing_description.split(";") if seg.strip()
            ]
            # Only append if the incoming text isn't already present as a segment
            incoming_lower = incoming_description.lower()
            already_present = any(
                seg.lower() == incoming_lower for seg in existing_segments
            )
            if already_present:
                description_text = existing_description
            else:
                description_text = existing_description + "; " + incoming_description
        elif incoming_description:
            description_text = incoming_description
        elif existing_description:
            description_text = existing_description
        else:
            description_text = "Time entry"

        lock_audit_parts: List[str] = []
        seen_audit = set()
        for candidate in [
            _clean_text((primary_row or {}).get(sc.COL_TIME_LOCK_AUDIT)),
            _clean_text(normalized.get("lockAudit")),
        ]:
            if not candidate:
                continue
            split_parts = [part.strip() for part in candidate.split("||")] if "||" in candidate else [candidate]
            for part in split_parts:
                if not part:
                    continue
                key = part.lower()
                if key in seen_audit:
                    continue
                seen_audit.add(key)
                lock_audit_parts.append(part)
        lock_audit_text = " || ".join(lock_audit_parts)

        now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        entry_id = (
            _clean_text((primary_row or {}).get(sc.COL_TIME_ENTRY_ID))
            or self.db._new_id("T")
        )
        created_at = _clean_text((primary_row or {}).get(sc.COL_TIME_CREATED)) or now_stamp
        effective_parent_id = parent_id or _clean_text((primary_row or {}).get(sc.COL_TIME_PARENT_ID))

        entry_row = {
            sc.COL_TIME_ENTRY_ID: entry_id,
            sc.COL_TIME_DATE: normalized["date"],
            sc.COL_TIME_CLIENT_ID: client_id,
            sc.COL_TIME_MATTER_ID: matter_id,
            sc.COL_TIME_PARENT_ID: effective_parent_id,
            sc.COL_TIME_DESC: description_text,
            sc.COL_TIME_HOURS: round(float(aggregate_hours), 2),
            sc.COL_TIME_RATE: round(float(normalized["clientRate"]), 2),
            sc.COL_TIME_SHARE_PCT: round(float(normalized["sharePct"]), 2),
            sc.COL_TIME_GROSS: amounts["gross_to_client"],
            sc.COL_TIME_NET: amounts["amount_to_you"],
            sc.COL_TIME_HST: amounts["hst_on_you"],
            sc.COL_TIME_TOTAL: amounts["total_you_incl_hst"],
            sc.COL_TIME_SECONDS: int(aggregate_raw_seconds),
            sc.COL_TIME_STATUS: normalized_status,
            sc.COL_TIME_LOCK_AUDIT: lock_audit_text,
            sc.COL_TIME_CREATED: created_at,
        }

        self.db._upsert_row_by_key(TBL_TIME, sc.COL_TIME_ENTRY_ID, entry_id, entry_row)

        # Merge duplicate editable rows into the primary bucket, preserving an audit breadcrumb.
        if duplicate_rows:
            for dup in duplicate_rows:
                dup_id = _clean_text(dup.get(sc.COL_TIME_ENTRY_ID))
                if not dup_id:
                    continue
                dup_row = dict(dup)
                dup_row[sc.COL_TIME_HOURS] = 0.0
                dup_row[sc.COL_TIME_GROSS] = 0.0
                dup_row[sc.COL_TIME_NET] = 0.0
                dup_row[sc.COL_TIME_HST] = 0.0
                dup_row[sc.COL_TIME_TOTAL] = 0.0
                dup_row[sc.COL_TIME_SECONDS] = 0
                dup_row[sc.COL_TIME_STATUS] = "Merged"
                dup_desc = _clean_text(dup_row.get(sc.COL_TIME_DESC))
                dup_row[sc.COL_TIME_DESC] = (
                    f"[MERGED->{entry_id}] {dup_desc}".strip()
                    if dup_desc
                    else f"[MERGED->{entry_id}]"
                )
                self.db._upsert_row_by_key(TBL_TIME, sc.COL_TIME_ENTRY_ID, dup_id, dup_row)

        persisted = self.db._find_time_entry(entry_id)
        verified = self.db._compare_rows_loose(entry_row, persisted)

        return {
            "ok": bool(verified),
            "verifiedExact": bool(verified),
            "entryId": entry_id,
            "savedRow": entry_row,
            "aggregateRawSeconds": int(aggregate_raw_seconds),
            "aggregateHoursRounded": round(float(aggregate_hours), 2),
            "mergedRowCount": len(duplicate_rows),
            "message": "" if verified else "Entry write verification failed.",
        }

    def get_time_docket_aggregate(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        self.db.ensure_schema()

        raw = dict(payload or {})
        date_text = self.db._pick_text(raw, ["date", "dateText", sc.COL_TIME_DATE])
        if not date_text:
            date_text = datetime.now().strftime("%Y-%m-%d")
        if not _is_valid_iso_date(date_text):
            raise ValueError("Date must be in YYYY-MM-DD format.")

        client_name = self.db._pick_text(raw, ["clientName", "clientText", sc.COL_CLIENT_NAME, "Client"])
        matter_name = self.db._pick_text(raw, ["matterName", "matterText", sc.COL_MATTER_NAME, "Matter"])
        allow_client_only = self.db._to_bool_int(
            self.db._pick_value(raw, ["allowClientOnlyDraft", "allowClientOnly", "clientOnlyOverride"]),
            default=0,
        ) == 1

        if not client_name:
            return {
                "ok": True,
                "exists": False,
                "entryId": "",
                "aggregateRawSeconds": 0,
                "aggregateHoursRounded": 0.0,
                "status": "Draft",
                "message": "",
            }

        client_rows = [self.db._canonicalize_client_row(r) for r in self.db._read_table_rows(TBL_CLIENTS)]
        matter_rows = [self.db._canonicalize_matter_row(r) for r in self.db._read_table_rows(TBL_MATTERS)]
        time_rows = [self.db._canonicalize_time_row(r) for r in self.db._read_table_rows(TBL_TIME)]

        client_id = ""
        client_row: Optional[Dict[str, Any]] = None
        for row in client_rows:
            if _clean_text(row.get(sc.COL_CLIENT_NAME)).lower() == client_name.lower():
                client_id = _clean_text(row.get(sc.COL_CLIENT_ID))
                client_name = _clean_text(row.get(sc.COL_CLIENT_NAME))
                client_row = row
                break

        matter_id = ""
        matter_row: Optional[Dict[str, Any]] = None
        if matter_name:
            for row in matter_rows:
                row_name = _clean_text(row.get(sc.COL_MATTER_NAME))
                row_display = _clean_text(row.get(sc.COL_MATTER_DISPLAY_NAME))
                row_number = _clean_text(row.get(sc.COL_MATTER_NUMBER))
                if (
                    row_name.lower() == matter_name.lower()
                    or row_display.lower() == matter_name.lower()
                    or row_number.lower() == matter_name.lower()
                ):
                    row_client_id = _clean_text(row.get(sc.COL_MATTER_CLIENT_ID))
                    if client_id and row_client_id and row_client_id.lower() != client_id.lower():
                        continue
                    matter_id = _clean_text(row.get(sc.COL_MATTER_ID))
                    matter_name = row_display or row_name or matter_name
                    matter_row = row
                    if not client_id:
                        client_id = row_client_id
                    break
        elif not allow_client_only:
            return {
                "ok": True,
                "exists": False,
                "entryId": "",
                "aggregateRawSeconds": 0,
                "aggregateHoursRounded": 0.0,
                "status": "Draft",
                "message": "Matter not selected.",
            }

        if not client_id:
            return {
                "ok": True,
                "exists": False,
                "entryId": "",
                "aggregateRawSeconds": 0,
                "aggregateHoursRounded": 0.0,
                "status": "Draft",
                "message": "",
            }

        matches: List[Dict[str, Any]] = []
        for row in time_rows:
            if _clean_text(row.get(sc.COL_TIME_DATE)) != date_text:
                continue
            if _clean_text(row.get(sc.COL_TIME_CLIENT_ID)).lower() != client_id.lower():
                continue
            row_matter_id = _clean_text(row.get(sc.COL_TIME_MATTER_ID))
            if matter_id:
                if row_matter_id.lower() != matter_id.lower():
                    continue
            elif row_matter_id:
                continue
            if _clean_text(row.get(sc.COL_TIME_STATUS)).lower() == "merged":
                continue
            matches.append(row)

        # Resolve Default Rate from Matter (or Client)
        def _resolve_rate() -> Tuple[float, float]:
            target_row = matter_row or client_row
            if not target_row:
                return 0.0, 100.0

            default_rate = float(self.db._parse_float(target_row.get(sc.COL_MATTER_DEF_RATE, target_row.get("defaultRate"))) or 0.0)
            default_share = float(self.db._parse_float(target_row.get(sc.COL_MATTER_DEF_SHARE, target_row.get("defaultSharePct"))) or 100.0)
            
            rate_history_str = _clean_text(target_row.get(sc.COL_MATTER_RATE_HISTORY, target_row.get("rateHistory")))
            if not rate_history_str or rate_history_str == "[]":
                return default_rate, default_share

            try:
                history = json.loads(rate_history_str)
                if not isinstance(history, list) or not history:
                    return default_rate, default_share
                
                # Sort by effective date ascending
                history.sort(key=lambda x: _clean_text(x.get("effectiveDate", "")))
                
                applicable_epoch = history[0]
                for epoch in history:
                    epoch_date = _clean_text(epoch.get("effectiveDate", ""))
                    if epoch_date <= date_text:
                        applicable_epoch = epoch
                    else:
                        break

                r = float(self.db._parse_float(applicable_epoch.get("rate")) or default_rate)
                s = float(self.db._parse_float(applicable_epoch.get("sharePct")) or default_share)
                return r, s
            except Exception:
                return default_rate, default_share

        resolved_rate, resolved_share = _resolve_rate()

        if not matches:
            return {
                "ok": True,
                "exists": False,
                "entryId": "",
                "date": date_text,
                "clientId": client_id,
                "clientName": client_name,
                "matterId": matter_id,
                "matterName": matter_name,
                "rate": round(resolved_rate, 2),
                "sharePct": round(resolved_share, 2),
                "aggregateRawSeconds": 0,
                "aggregateHoursRounded": 0.0,
                "status": "Draft",
                "message": "",
            }

        matches.sort(
            key=lambda row: (
                _clean_text(row.get(sc.COL_TIME_CREATED)),
                _clean_text(row.get(sc.COL_TIME_ENTRY_ID)),
            )
        )

        primary = matches[0]
        aggregate_raw_seconds = sum(int(self.db._parse_float(r.get(sc.COL_TIME_SECONDS)) or 0) for r in matches)
        aggregate_raw_seconds = max(0, aggregate_raw_seconds)
        aggregate_hours = (
            math.ceil(((aggregate_raw_seconds / 3600.0) * 10.0) - 1e-9) / 10.0
            if aggregate_raw_seconds > 0
            else 0.0
        )

        normalized_statuses = [self.db._normalize_time_status(r.get(sc.COL_TIME_STATUS)) for r in matches]
        effective_status = "Draft"
        if "Billed" in normalized_statuses:
            effective_status = "Billed"
        elif "Ready for Billing" in normalized_statuses:
            effective_status = "Ready for Billing"

        return {
            "ok": True,
            "exists": True,
            "entryId": _clean_text(primary.get(sc.COL_TIME_ENTRY_ID)),
            "date": date_text,
            "clientId": client_id,
            "clientName": client_name,
            "matterId": matter_id,
            "matterName": matter_name,
            "description": _clean_text(primary.get(sc.COL_TIME_DESC)),
            "lockAudit": _clean_text(primary.get(sc.COL_TIME_LOCK_AUDIT)),
            "rate": round(float(self.db._parse_float(primary.get(sc.COL_TIME_RATE)) or 0.0), 2),
            "sharePct": round(float(self.db._parse_float(primary.get(sc.COL_TIME_SHARE_PCT)) or 100.0), 2),
            "aggregateRawSeconds": int(aggregate_raw_seconds),
            "aggregateHoursRounded": round(float(aggregate_hours), 2),
            "status": effective_status,
            "sourceRowCount": len(matches),
            "message": "",
        }




    def get_docket_activity_report(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        from openpyxl import load_workbook
        import logging, time, os, json, csv
        from datetime import datetime
        
        logger = logging.getLogger("DocketReport")
        trace_mode = os.environ.get("CSPM_DOCKET_TRACE") == "1" or "--trace-docket-report" in sys.argv
        if trace_mode: logger.setLevel(logging.DEBUG)
        detail_log = logger.info if trace_mode else logger.debug

        t_start = time.perf_counter()
        
        # [A] File resolution
        path = self.db.paths.workbook_path()
        detail_log(f"[A] EXCEL SOURCE: path={path.absolute()}")
        detail_log(f"[A] EXCEL SOURCE: exists={path.exists()} size={path.stat().st_size if path.exists() else 0}")
        
        if not path.exists():
            logger.error("[A] CRITICAL: Workbook path does not exist.")
            raise FileNotFoundError("Excel file missing.")

        raw_rows = []
        headers = []
        try:
            # [B] Workbook open details
            wb = load_workbook(path, read_only=True, data_only=True)
            t_open_end = time.perf_counter()
            detail_log(f"[B] WORKBOOK OPEN: {(t_open_end - t_start)*1000:.1f}ms. Engine: openpyxl (read_only)")
            
            # [C] Sheet selection
            detail_log(f"[C] SHEET SELECTION: sheets_found={wb.sheetnames}")
            target_sheet = sc.SHEET_TIME
            if target_sheet not in wb.sheetnames:
                logger.error(f"[C] CRITICAL: Sheet '{target_sheet}' missing!")
                raise ValueError(f"Sheet {target_sheet} missing.")
            
            detail_log(f"[C] SHEET SELECTION: sheet_selected={target_sheet}")
            ws = wb[target_sheet]
            
            # [D] Header detection
            header_row_index = 1
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
                str_row = [str(c).strip() if c else "" for c in row]
                if sc.COL_TIME_ENTRY_ID in str_row:
                    headers = str_row
                    header_row_index = r_idx
                    break
            
            detail_log(f"[D] HEADER DETECTION: row_index={header_row_index}, headers={headers}")
            
            mapping = {
                "date_col": headers.index(sc.COL_TIME_DATE) if sc.COL_TIME_DATE in headers else "MISSING",
                "client_col": headers.index(sc.COL_TIME_CLIENT_ID) if sc.COL_TIME_CLIENT_ID in headers else "MISSING",
                "matter_col": headers.index(sc.COL_TIME_MATTER_ID) if sc.COL_TIME_MATTER_ID in headers else "MISSING",
                "status_col": headers.index(sc.COL_TIME_STATUS) if sc.COL_TIME_STATUS in headers else "MISSING",
                "hours_col": headers.index(sc.COL_TIME_HOURS) if sc.COL_TIME_HOURS in headers else "MISSING",
                "gross_col": headers.index(sc.COL_TIME_GROSS) if sc.COL_TIME_GROSS in headers else "MISSING"
            }
            detail_log(f"[D] MAPPING: {mapping}")

            if "MISSING" in mapping.values():
                logger.error(f"[D] CRITICAL: Missing required headers in {target_sheet}!")
                raise ValueError("Missing required headers in Excel sheet.")

            # [E] Range scan
            max_r = ws.max_row
            detail_log(f"[E] RANGE SCAN: {header_row_index + 1}..{max_r}")
            for row in ws.iter_rows(min_row=header_row_index + 1, max_row=max_r, values_only=True):
                if any(row):
                    raw_dict = {headers[i]: val for i, val in enumerate(row) if i < len(headers) and headers[i]}
                    raw_rows.append(raw_dict)
                    
            detail_log(f"[E] RANGE SCAN: rows_scanned_total={len(raw_rows)}")
            
            if trace_mode and raw_rows:
                logger.debug(f"[F] RAW SAMPLE (first 3): {raw_rows[:3]}")
                
        finally:
            if 'wb' in locals(): wb.close()

        # [G] Parse records & Lookups
        try:
            client_rows = [self.db._canonicalize_client_row(r) for r in self.db._read_table_rows(TBL_CLIENTS)]
            matter_rows = [self.db._canonicalize_matter_row(r) for r in self.db._read_table_rows(TBL_MATTERS)]
            parent_rows = [self.db._canonicalize_parent_row(r) for r in self.db._read_table_rows(TBL_PARENTS)]
        except Exception as e:
            logger.error(f"Lookup table read failure: {e}", exc_info=True)
            client_rows, matter_rows, parent_rows = [], [], []

        client_name_by_id = {
            _clean_text(r.get(sc.COL_CLIENT_ID)): _clean_text(r.get(sc.COL_CLIENT_NAME))
            for r in client_rows
            if _clean_text(r.get(sc.COL_CLIENT_ID))
        }
        matter_name_by_id = {
            _clean_text(r.get(sc.COL_MATTER_ID)): (
                _clean_text(r.get(sc.COL_MATTER_DISPLAY_NAME))
                or _clean_text(r.get(sc.COL_MATTER_NAME))
                or _clean_text(r.get(sc.COL_MATTER_NUMBER))
            )
            for r in matter_rows
            if _clean_text(r.get(sc.COL_MATTER_ID))
        }
        parent_name_by_id = {
            _clean_text(r.get(sc.COL_PARENT_ID)): _clean_text(r.get(sc.COL_PARENT_NAME))
            for r in parent_rows
            if _clean_text(r.get(sc.COL_PARENT_ID))
        }

        parsed_rows = []
        for r in raw_rows:
            parsed_rows.append(self.db._canonicalize_time_row(r))
            
        detail_log(f"[G] PARSE RESULTS: rows_parsed_ok={len(parsed_rows)}")
        
        valid_dates = [str(r.get(sc.COL_TIME_DATE)) for r in parsed_rows if r.get(sc.COL_TIME_DATE) and _is_valid_iso_date(str(r.get(sc.COL_TIME_DATE)))]
        if valid_dates:
            detail_log(f"[E] DATASET SANITY: min_date={min(valid_dates)}, max_date={max(valid_dates)}")
        else:
            logger.warning("[E] DATASET SANITY: No valid dates parsed!")
            
        if trace_mode and parsed_rows:
            logger.debug(f"[H] PARSED SAMPLE (first 3): {parsed_rows[:3]}")

        # [I] Filters
        raw = dict(payload or {})
        start_text = self.db._pick_text(raw, ["fromDate", "dateFrom", "startDate", "from"]) or datetime.now().strftime("%Y-%m-%d")
        end_text = self.db._pick_text(raw, ["toDate", "dateTo", "endDate", "to"]) or start_text
        status_mode = _normalize_search_text(self.db._pick_text(raw, ["statusMode", "statusFilter", "status"])) or "all_except_merged"
        client_filter = _normalize_search_text(self.db._pick_text(raw, ["clientFilter", "clientName"]))
        matter_filter = _normalize_search_text(self.db._pick_text(raw, ["matterFilter", "matterName"]))
        text_query = _normalize_search_text(self.db._pick_text(raw, ["query", "search"]))

        detail_log(f"[I] FILTER LOGGING: filters={{from:{start_text}, to:{end_text}, status:{status_mode}, client:'{client_filter}', matter:'{matter_filter}', search:'{text_query}'}}")

        start_date = datetime.strptime(start_text, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_text, "%Y-%m-%d").date()
        if end_date < start_date:
            start_date, end_date = end_date, start_date

        def _pair_matches(filter_token: str, value_id: str, value_name: str) -> bool:
            if not filter_token:
                return True
            return filter_token in f"{value_id} {value_name}".lower()

        relation_pairs = []
        client_candidates = set()
        matter_candidates = set()
        saw_no_matter = False

        for row in client_rows:
            cname = _clean_text(row.get(sc.COL_CLIENT_NAME))
            if cname:
                client_candidates.add(cname)

        for row in matter_rows:
            cid = _clean_text(row.get(sc.COL_MATTER_CLIENT_ID))
            cname = _clean_text(row.get(sc.COL_MATTER_CLIENT_NAME)) or client_name_by_id.get(cid, "")
            mid = _clean_text(row.get(sc.COL_MATTER_ID))
            mname = (
                _clean_text(row.get(sc.COL_MATTER_DISPLAY_NAME))
                or _clean_text(row.get(sc.COL_MATTER_NAME))
                or _clean_text(row.get(sc.COL_MATTER_NUMBER))
                or mid
            )
            if cname:
                client_candidates.add(cname)
            if mname:
                matter_candidates.add(mname)
            if cid or cname or mid or mname:
                relation_pairs.append(
                    {
                        "clientId": cid,
                        "clientName": cname,
                        "matterId": mid,
                        "matterName": mname,
                    }
                )

        for row in parsed_rows:
            cid = _clean_text(row.get(sc.COL_TIME_CLIENT_ID))
            cname = client_name_by_id.get(cid, cid)
            mid = _clean_text(row.get(sc.COL_TIME_MATTER_ID))
            if mid:
                mname = matter_name_by_id.get(mid, mid)
                if mname:
                    matter_candidates.add(mname)
            else:
                mname = "No Matter"
                saw_no_matter = True

            if cname:
                client_candidates.add(cname)
            relation_pairs.append(
                {
                    "clientId": cid,
                    "clientName": cname,
                    "matterId": mid,
                    "matterName": mname,
                }
            )

        if saw_no_matter:
            matter_candidates.add("No Matter")

        option_clients_set = set()
        if matter_filter:
            for pair in relation_pairs:
                if _pair_matches(matter_filter, _clean_text(pair.get("matterId")), _clean_text(pair.get("matterName"))):
                    cname = _clean_text(pair.get("clientName"))
                    if cname:
                        option_clients_set.add(cname)
        else:
            option_clients_set.update(client_candidates)

        option_matters_set = set()
        if client_filter:
            for pair in relation_pairs:
                if _pair_matches(client_filter, _clean_text(pair.get("clientId")), _clean_text(pair.get("clientName"))):
                    mname = _clean_text(pair.get("matterName"))
                    if mname:
                        option_matters_set.add(mname)
        else:
            option_matters_set.update(matter_candidates)

        if not option_clients_set:
            option_clients_set.update(client_candidates)
        if not option_matters_set:
            option_matters_set.update(matter_candidates)

        raw_client_selection = _clean_text(self.db._pick_text(raw, ["clientFilter", "clientName"]))
        raw_matter_selection = _clean_text(self.db._pick_text(raw, ["matterFilter", "matterName"]))
        if raw_client_selection:
            option_clients_set.add(raw_client_selection)
        if raw_matter_selection:
            option_matters_set.add(raw_matter_selection)

        option_clients = ["All Clients"]
        option_clients.extend(
            sorted(
                [
                    label
                    for label in option_clients_set
                    if _normalize_search_text(label) and _normalize_search_text(label) != "all clients"
                ],
                key=lambda label: _normalize_search_text(label),
            )
        )
        option_matters = ["All Matters"]
        option_matters.extend(
            sorted(
                [
                    label
                    for label in option_matters_set
                    if _normalize_search_text(label) and _normalize_search_text(label) != "all matters"
                ],
                key=lambda label: _normalize_search_text(label),
            )
        )

        detail_rows = []
        rejects = {"invalid_date": 0, "out_of_range": 0, "status_excluded": 0, "client_mismatch": 0, "matter_mismatch": 0, "search_miss": 0}
        counts = {"post_date": 0, "post_status": 0, "post_client": 0, "post_matter": 0, "post_search": 0}
        
        for row in parsed_rows:
            date_text = _clean_text(row.get(sc.COL_TIME_DATE))
            if not _is_valid_iso_date(date_text):
                rejects["invalid_date"] += 1
                continue
            row_date = datetime.strptime(date_text, "%Y-%m-%d").date()
            if row_date < start_date or row_date > end_date:
                rejects["out_of_range"] += 1
                continue
            counts["post_date"] += 1
                
            raw_status = _clean_text(row.get(sc.COL_TIME_STATUS)).lower()
            norm_status = "Merged" if raw_status == "merged" else self.db._normalize_time_status(raw_status)
            
            inc_status = True
            if status_mode == "all": inc_status = True
            elif status_mode == "all_except_merged": inc_status = raw_status != "merged"
            elif status_mode in ("draft", "draft_only", "unbilled", "open"): inc_status = norm_status == "Draft"
            elif status_mode in ("ready", "ready_for_billing"): inc_status = norm_status == "Ready for Billing"
            elif status_mode in ("billed", "billed_only"): inc_status = norm_status == "Billed"
            elif status_mode in ("merged", "merged_only"): inc_status = raw_status == "merged"
            else: inc_status = raw_status != "merged"
            
            if not inc_status:
                rejects["status_excluded"] += 1
                continue
            counts["post_status"] += 1

            c_id = _clean_text(row.get(sc.COL_TIME_CLIENT_ID))
            c_name = client_name_by_id.get(c_id, c_id)
            if client_filter and client_filter not in f"{c_id} {c_name}".lower():
                rejects["client_mismatch"] += 1
                continue
            counts["post_client"] += 1

            m_id = _clean_text(row.get(sc.COL_TIME_MATTER_ID))
            m_name = matter_name_by_id.get(m_id, m_id) or "No Matter"
            if matter_filter and matter_filter not in f"{m_id} {m_name}".lower():
                rejects["matter_mismatch"] += 1
                continue
            counts["post_matter"] += 1

            desc = _clean_text(row.get(sc.COL_TIME_DESC))
            entry_id = _clean_text(row.get(sc.COL_TIME_ENTRY_ID))
            p_id = _clean_text(row.get(sc.COL_TIME_PARENT_ID))
            p_name = parent_name_by_id.get(p_id, p_id)
            
            if text_query:
                haystack = f"{entry_id} {desc} {c_id} {c_name} {m_id} {m_name} {p_id} {p_name} {norm_status}".lower()
                if text_query not in haystack:
                    rejects["search_miss"] += 1
                    continue
            counts["post_search"] += 1

            row["clientName"] = c_name
            row["matterName"] = m_name
            row["parentName"] = p_name
            row["status"] = norm_status
            
            row["hours"] = float(self.db._parse_float(row.get(sc.COL_TIME_HOURS)) or 0.0)
            row["grossToClient"] = float(self.db._parse_float(row.get(sc.COL_TIME_GROSS)) or 0.0)
            
            detail_rows.append(row)

        detail_log(f"[I] FILTER LOGGING: counts={counts}")
        detail_log(f"[G] REJECT REASONS: {rejects}")

        detail_rows.sort(key=lambda r: (_clean_text(r.get("date")), _normalize_search_text(r.get("clientName")), _normalize_search_text(r.get("matterName")), _normalize_search_text(r.get("entryId"))))

        summary_map = {}
        total_hours = 0.0
        total_gross = 0.0
        for r in detail_rows:
            total_hours += r["hours"]
            total_gross += r["grossToClient"]
            m_id = _clean_text(r.get("matterId"))
            c_key = _clean_text(r.get("clientId")) or _clean_text(r.get("clientName")).lower()
            s_key = f"matter::{m_id.lower()}" if m_id else f"no_matter::{c_key.lower()}"
            
            if s_key not in summary_map:
                summary_map[s_key] = {
                    "matterId": m_id,
                    "matterName": r["matterName"],
                    "clientId": r.get("clientId", ""),
                    "clientName": r["clientName"],
                    "entryCount": 0,
                    "totalHours": 0.0,
                    "totalGrossToClient": 0.0
                }
            summary_map[s_key]["entryCount"] += 1
            summary_map[s_key]["totalHours"] += r["hours"]
            summary_map[s_key]["totalGrossToClient"] += r["grossToClient"]

        summary_rows = list(summary_map.values())
        summary_rows.sort(key=lambda r: (1 if _normalize_search_text(r.get("matterName")) == "no matter" else 0, _normalize_search_text(r.get("matterName")), _normalize_search_text(r.get("clientName"))))
        for r in summary_rows:
            r["totalHours"] = round(r["totalHours"], 2)
            r["totalGrossToClient"] = round(r["totalGrossToClient"], 2)

        # [J] Publish to UI
        detail_log(f"[J] PUBLISH: publishing_rows={len(detail_rows)}")
        detail_log(f"[J] PUBLISH: summary_by_matter_rows={len(summary_rows)}")
        detail_log(f"[J] PUBLISH: aggregate_entries={len(detail_rows)} hours={total_hours:.2f} gross={total_gross:.2f}")
        detail_log(f"[Perf] [DocketReport] total_time={time.perf_counter()-t_start:.3f}s")
        
        if trace_mode:
            try:
                out_dir = self.db.paths.root / "outputs"
                out_dir.mkdir(exist_ok=True)
                with open(out_dir / "docket_raw_preview.csv", "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=headers)
                    writer.writeheader()
                    writer.writerows(raw_rows[:200])
                with open(out_dir / "docket_parsed_preview.json", "w", encoding="utf-8") as f:
                    json.dump(parsed_rows[:200], f, indent=2, default=str)
                with open(out_dir / "docket_reject_summary.json", "w", encoding="utf-8") as f:
                    json.dump(rejects, f, indent=2)
                logger.info(f"[DocketReport] Dumped TRACE artifacts to {out_dir}")
            except Exception as e:
                logger.error(f"[DocketReport] Failed to write TRACE artifacts: {e}")

        return {
            "ok": True,
            "filters": {
                "fromDate": start_date.strftime("%Y-%m-%d"),
                "toDate": end_date.strftime("%Y-%m-%d"),
                "statusMode": status_mode,
                "clientFilter": client_filter,
                "matterFilter": matter_filter,
                "query": text_query
            },
            "rows": detail_rows,
            "summaryRows": summary_rows,
            "totalRows": len(detail_rows),
            "summaryCount": len(summary_rows),
            "optionClients": option_clients,
            "optionMatters": option_matters,
            "totals": {
                "totalHours": round(total_hours, 2),
                "totalGrossToClient": round(total_gross, 2)
            },
            "debug": {
                "scanned": len(raw_rows),
                "accepted": len(detail_rows),
                "rejected": sum(rejects.values()),
                "rejects": rejects
            },
            "message": ""
        }

    def export_docket_activity_csv(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        raw_payload = dict(payload or {})
        report_filters = raw_payload.get("filters") if isinstance(raw_payload.get("filters"), dict) else raw_payload
        report = self.get_docket_activity_report(dict(report_filters or {}))
        if not bool(report.get("ok")):
            return {
                "ok": False,
                "path": "",
                "filename": "",
                "rowCount": 0,
                "message": _clean_text(report.get("message")) or "Could not build report.",
            }

        # ensure we have the expected types so the static analyzer stops complaining
        # about union values. initialize first and then overwrite under guard.
        rows: List[Dict[str, Any]] = []
        if isinstance(report.get("rows"), list):
            # mypy/Pylance still think `.get` could return None, so silence with ignore
            rows = report.get("rows")  # type: ignore[assignment]
        summary_rows: List[Dict[str, Any]] = []
        if isinstance(report.get("summaryRows"), list):
            summary_rows = report.get("summaryRows")  # type: ignore[assignment]
        totals: Dict[str, Any] = {}
        if isinstance(report.get("totals"), dict):
            totals = report.get("totals")  # type: ignore[assignment]

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        candidate_dirs = [
            self.db.paths.exports_dir(),
            self.db.paths.data_dir() / "exports",
            self.db.paths.root / "outputs",
        ]
        output_path: Optional[Path] = None
        last_error: Optional[Exception] = None

        for export_dir in candidate_dirs:
            try:
                export_dir.mkdir(parents=True, exist_ok=True)
            except Exception as exc:
                last_error = exc
                continue
            candidate_path = export_dir / f"docket_activity_{stamp}.csv"
            try:
                with candidate_path.open("w", encoding="utf-8-sig", newline="") as handle:
                    writer = csv.writer(handle)
                    writer.writerow(
                        [
                            "Date",
                            "Entry ID",
                            "Client",
                            "Matter",
                            "Description",
                            "Hours",
                            "GrossToClient",
                            "Status",
                            "Rate",
                            "SharePct",
                            "RawSeconds",
                        ]
                    )
                    for row in rows:
                        writer.writerow(
                            [
                                _clean_text(row.get("date")),
                                _clean_text(row.get("entryId")),
                                _clean_text(row.get("clientName")),
                                _clean_text(row.get("matterName")),
                                _clean_text(row.get("description")),
                                round(float(self.db._parse_float(row.get("hours")) or 0.0), 2),
                                round(float(self.db._parse_float(row.get("grossToClient")) or 0.0), 2),
                                _clean_text(row.get("status")),
                                round(float(self.db._parse_float(row.get("rate")) or 0.0), 2),
                                round(float(self.db._parse_float(row.get("sharePct")) or 0.0), 2),
                                int(self.db._parse_float(row.get("rawSeconds")) or 0),
                            ]
                        )

                    writer.writerow([])
                    writer.writerow(["Summary by Matter"])
                    writer.writerow(["Matter", "Client", "Entries", "Hours", "GrossToClient"])
                    for row in summary_rows:
                        writer.writerow(
                            [
                                _clean_text(row.get("matterName")),
                                _clean_text(row.get("clientName")),
                                int(self.db._parse_float(row.get("entryCount")) or 0),
                                round(float(self.db._parse_float(row.get("totalHours")) or 0.0), 2),
                                round(float(self.db._parse_float(row.get("totalGrossToClient")) or 0.0), 2),
                            ]
                        )
                    writer.writerow([])
                    writer.writerow(
                        [
                            "Totals",
                            "",
                            "",
                            round(float(self.db._parse_float(totals.get("totalHours")) or 0.0), 2),
                            round(float(self.db._parse_float(totals.get("totalGrossToClient")) or 0.0), 2),
                        ]
                    )
                output_path = candidate_path
                break
            except PermissionError as exc:
                last_error = exc
                continue
            except Exception as exc:
                last_error = exc
                continue

        if output_path is None:
            raise PermissionError(f"Could not write CSV export to any configured directory: {last_error}")

        return {
            "ok": True,
            "path": str(output_path),
            "filename": output_path.name,
            "rowCount": len(rows),
            "message": f"CSV exported: {output_path.name} | Saved to: {output_path.parent}",
        }

    def _canonicalize_time_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        entry_id = _clean_text(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_ENTRY_ID))
        if not entry_id:
            entry_id = self.db._new_id("T")

        hours = float(self.db._parse_float(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_HOURS)) or 0.0)
        rate = float(self.db._parse_float(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_RATE)) or 0.0)
        share_pct = normalize_pct(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_SHARE_PCT), default_pct=100.0)

        raw_seconds = self.db._parse_int(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_SECONDS))
        if raw_seconds is None:
            raw_seconds = int(round(hours * 3600.0))
        raw_seconds = int(max(0, raw_seconds))

        if hours <= 0.0 and raw_seconds > 0:
            hours = math.ceil(((raw_seconds / 3600.0) * 10.0) - 1e-9) / 10.0

        amounts = calc_amounts(hours=hours, client_rate=rate, your_share_pct=share_pct, hst_rate=0.13)

        gross = self.db._parse_float(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_GROSS))
        amount_to_you = self.db._parse_float(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_NET))
        hst = self.db._parse_float(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_HST))
        total = self.db._parse_float(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_TOTAL))

        if gross is None:
            gross = amounts["gross_to_client"]
        if amount_to_you is None:
            amount_to_you = amounts["amount_to_you"]
        if hst is None:
            hst = amounts["hst_on_you"]
        if total is None:
            total = amounts["total_you_incl_hst"]

        status_raw = _clean_text(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_STATUS))
        if status_raw.lower() == "merged":
            status = "Merged"
        else:
            status = self.db._normalize_time_status(status_raw)

        created_at = _clean_text(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_CREATED))
        if not created_at:
            created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        return {
            sc.COL_TIME_ENTRY_ID: entry_id,
            sc.COL_TIME_DATE: _clean_text(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_DATE)),
            sc.COL_TIME_CLIENT_ID: _clean_text(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_CLIENT_ID)),
            sc.COL_TIME_MATTER_ID: _clean_text(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_MATTER_ID)),
            sc.COL_TIME_PARENT_ID: _clean_text(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_PARENT_ID)),
            sc.COL_TIME_DESC: _clean_text(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_DESC)),
            sc.COL_TIME_HOURS: round(hours, 2),
            sc.COL_TIME_RATE: round(rate, 2),
            sc.COL_TIME_SHARE_PCT: round(share_pct, 2),
            sc.COL_TIME_GROSS: round(float(gross), 2),
            sc.COL_TIME_NET: round(float(amount_to_you), 2),
            sc.COL_TIME_HST: round(float(hst), 2),
            sc.COL_TIME_TOTAL: round(float(total), 2),
            sc.COL_TIME_SECONDS: int(raw_seconds),
            sc.COL_TIME_STATUS: status,
            sc.COL_TIME_LOCK_AUDIT: _clean_text(self.db._value_with_alias(TBL_TIME.table, row, sc.COL_TIME_LOCK_AUDIT)),
            sc.COL_TIME_CREATED: created_at,
        }

