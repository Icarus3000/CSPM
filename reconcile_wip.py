import sys, os, re
from pathlib import Path
from datetime import datetime

root = Path(__file__).resolve().parent
sys.path.insert(0, str(root / "src" / "python"))

import openpyxl
from services.paths import AppPaths
from repositories.excel_repo import ExcelRepo
from domain import schema_constants as sc
from services.dockets_import_service import DocketsImportService

cspm_path = root / "data" / "CSPM.xlsm"
dockets_path = root / "data" / "Dockets.xlsm"

paths = AppPaths(root)
repo = ExcelRepo(paths)

def _clean_text(val):
    if val is None:
        return ""
    return str(val).strip()

def amount(val) -> float:
    if val is None:
        return 0.0
    try:
        if isinstance(val, str):
            cleaned = re.sub(r'[^\d.\-]', '', val)
            return float(cleaned) if cleaned else 0.0
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def row_year(date_val) -> int:
    if isinstance(date_val, datetime):
        return date_val.year
    if isinstance(date_val, str) and len(date_val) >= 4:
        try:
            return int(date_val[:4])
        except Exception:
            pass
    return 0

def docket_amount_to_cs(row):
    return amount(row.get("Amount to CS"))

def invoice_marker_bucket(ref: str) -> str:
    r = _clean_text(ref).lower()
    if not r:
        return "blank"
    if r == "hold":
        return "hold"
    if r in {"no bill", "nobill", "n/a", "no", "w/o"}:
        return "no_bill"
    if r in {"legacy billed", "legacy", "billed legacy"}:
        return "legacy_billed"
    return "invoice"

def check_wip_reconciliation():
    # Load Legacy Dockets
    print("Loading Legacy Dockets...")
    legacy_wb = openpyxl.load_workbook(str(dockets_path), data_only=True)
    legacy_ws = legacy_wb["Dockets"]
    legacy_headers = [str(c.value).strip() if c.value else "" for c in legacy_ws[1]]
    
    legacy_wip_rows = []
    legacy_wip_total = 0.0
    for r in legacy_ws.iter_rows(min_row=2, values_only=True):
        if not any(r): continue
        row = dict(zip(legacy_headers, r))
        
        date_val = row.get("Date")
        if row_year(date_val) != 2026:
            continue
            
        ref = _clean_text(row.get("Invoice")).upper()
        bucket = invoice_marker_bucket(ref)
        prod = docket_amount_to_cs(row)

        
        if bucket in ("blank", "hold"):
            legacy_wip_total += prod
            desc = str(row.get("Description"))[:30]
            legacy_wip_rows.append({"desc": desc, "amount": prod, "ref": ref, "bucket": bucket})
            
    legacy_wb.close()
    
    # Load CSPM TimeEntries
    print("Loading CSPM TimeEntries...")
    cspm_wb = openpyxl.load_workbook(str(cspm_path), data_only=True)
    time_ws = cspm_wb["TimeEntries"]
    time_headers = [str(c.value).strip() if c.value else "" for c in time_ws[1]]
    
    python_wip_rows = []
    python_wip_total = 0.0
    for r in time_ws.iter_rows(min_row=2, values_only=True):
        if not any(r): continue
        row = dict(zip(time_headers, r))
        
        date_val = row.get(sc.COL_TIME_DATE)
        if row_year(date_val) != 2026:
            continue
            
        status = _clean_text(row.get(sc.COL_TIME_STATUS)).lower()
        net = amount(row.get(sc.COL_TIME_NET))
        
        if status not in {"billed", "merged"}:
            python_wip_total += net
            desc = str(row.get(sc.COL_TIME_DESC))[:30]
            ref = _clean_text(row.get(sc.COL_TIME_INVOICE_REF)).upper()
            python_wip_rows.append({"desc": desc, "amount": net, "status": status, "ref": ref})
            
    cspm_wb.close()
    
    print(f"\nLegacy WIP Total: ${legacy_wip_total:,.2f} ({len(legacy_wip_rows)} rows)")
    print(f"Python WIP Total: ${python_wip_total:,.2f} ({len(python_wip_rows)} rows)")
    print(f"Difference: ${python_wip_total - legacy_wip_total:,.2f}")
    
    # Find the specific row(s) causing the difference
    legacy_amounts = {}
    for r in legacy_wip_rows:
        amt = round(r["amount"], 2)
        legacy_amounts[amt] = legacy_amounts.get(amt, 0) + 1
        
    python_amounts = {}
    for r in python_wip_rows:
        amt = round(r["amount"], 2)
        python_amounts[amt] = python_amounts.get(amt, 0) + 1
        
    diffs = []
    for r in python_wip_rows:
        amt = round(r["amount"], 2)
        if amt in legacy_amounts and legacy_amounts[amt] > 0:
            legacy_amounts[amt] -= 1
        else:
            diffs.append(r)
            
    print("\nWIP items in Python that were excluded from Legacy:")
    for d in diffs:
        print(f"  Amount: ${d['amount']:,.2f} | Status: {d['status']} | Ref: {d['ref']} | Desc: {d['desc']}")
        
    legacy_amounts_again = {}
    for r in legacy_wip_rows:
        amt = round(r["amount"], 2)
        legacy_amounts_again[amt] = legacy_amounts_again.get(amt, 0) + 1
        
    python_amounts_again = {}
    for r in python_wip_rows:
        amt = round(r["amount"], 2)
        python_amounts_again[amt] = python_amounts_again.get(amt, 0) + 1
        
    missing = []
    for r in legacy_wip_rows:
        amt = round(r["amount"], 2)
        if amt in python_amounts_again and python_amounts_again[amt] > 0:
            python_amounts_again[amt] -= 1
        else:
            missing.append(r)
            
    print("\nWIP items in Legacy that were excluded from Python:")
    for m in missing:
        print(f"  Amount: ${m['amount']:,.2f} | Bucket: {m['bucket']} | Ref: {m['ref']} | Desc: {m['desc']}")

if __name__ == "__main__":
    check_wip_reconciliation()
